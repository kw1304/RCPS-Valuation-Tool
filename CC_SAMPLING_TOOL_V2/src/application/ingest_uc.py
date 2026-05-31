"""IngestUC — 파일 → Population[AR/AP] persist orchestration.

설계서 §6.1 [2].
"""
from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional
import openpyxl

from src.domain.entities import Account, Kind
from src.domain.fx import convert_to_base, FxRateMissing
from src.infrastructure.db.repository import ProjectRepo, AccountRepo
from src.infrastructure.ingest.excel_loader import (
    detect_sheet_kind, load_account_sheet,
)
from src.infrastructure.ingest.rp_parser import parse_related_parties
from src.infrastructure.ingest.allowance_parser import parse_allowance
from src.infrastructure.ingest.fs_parser import parse_fs_totals


@dataclass
class IngestResult:
    project_id: int
    ar_count: int
    ap_count: int
    ar_total_krw: float
    ap_total_krw: float
    confidence_ar: float
    confidence_ap: float
    fs_totals: dict
    needs_mapping_confirmation: bool
    # BS vs 명세서 cross-check (표준절차 1~2단계)
    cross_check: dict = field(default_factory=dict)
    # 시트별 매핑 상세 — 매핑확인 UI용
    sheets_meta: list = field(default_factory=list)


class IngestUC:
    def __init__(self, session, fx_client):
        self.s = session
        self.fx = fx_client
        self.proj = ProjectRepo(session)
        self.acc = AccountRepo(session)

    def ingest(
        self,
        project_id: int,
        ledger_path: Path,
        fs_path: Optional[Path],
        rp_path: Optional[Path],
        allowance_path: Optional[Path],
        explicit_mapping: Optional[dict] = None,
    ) -> IngestResult:
        """원장 적재.

        explicit_mapping: 마법사 확정 매핑 {sheet_name: {kind, header_row, column_map}}.
                          주어지면 시트·헤더·컬럼 자동감지 전부 skip. None이면 자동경로.
        """
        project = self.proj.get(project_id)

        # 1) 시트 종류 결정 + 시트별 매핑 spec.
        # spec = None(자동) | {header_row, column_map, kind}(명시).
        # MIXED 시트는 AR/AP 양쪽 등록 후 load_account_sheet kind_filter로 분리.
        from src.infrastructure.ingest.excel_loader import classify_sheet_by_content
        # assignment entry = (sheet_name, kind_filter, spec)
        sheet_assignments: dict[str, list[tuple]] = {"AR": [], "AP": []}
        if explicit_mapping:
            for sn, m in explicit_mapping.items():
                kind = (m.get("kind") or "IGNORE").upper()
                spec = {
                    "header_row": int(m.get("header_row", 0)),
                    "column_map": {k: int(v) for k, v in (m.get("column_map") or {}).items()},
                    "kind": kind,
                }
                if kind in ("AR", "AP"):
                    sheet_assignments[kind].append((sn, None, spec))
                elif kind == "MIXED":
                    sheet_assignments["AR"].append((sn, "AR", {**spec, "kind": "AR"}))
                    sheet_assignments["AP"].append((sn, "AP", {**spec, "kind": "AP"}))
                # FS/RP/ALLOWANCE/IGNORE in-ledger 시트는 별도 업로드로 처리 → skip
        else:
            wb = openpyxl.load_workbook(ledger_path, read_only=True)
            for sn in wb.sheetnames:
                kind = detect_sheet_kind(sn)
                if kind is None:
                    # 시트명 alias 매칭 실패 → 내용으로 분류
                    kind = classify_sheet_by_content(ledger_path, sn)
                if kind in ("AR", "AP"):
                    sheet_assignments[kind].append((sn, None, None))
                elif kind == "MIXED":
                    sheet_assignments["AR"].append((sn, "AR", None))
                    sheet_assignments["AP"].append((sn, "AP", None))
            wb.close()

        # 2) RP/충당금 사전로드
        rp_names: set[str] = set()
        synonym_groups: list[list[str]] = []
        rp_source = "none"
        if rp_path is not None:
            rp_sheet = self._auto_sheet(rp_path, "RP")
            if rp_sheet:
                rp_names = parse_related_parties(rp_path, rp_sheet)
                from src.infrastructure.ingest.rp_parser import parse_rp_synonym_groups
                synonym_groups = parse_rp_synonym_groups(rp_path, rp_sheet)
                rp_source = "manual_upload"
        # RP 파일 없으면 DART 자동 fetch (DART_API_KEY 있을 때).
        # 실패해도 ingest는 계속하되, 실패 단계를 rp_source에 기록 → 화면 노출(조용한 0건 방지).
        if not rp_names:
            try:
                from src.infrastructure.dart import DartClient
                from src.infrastructure.dart.rp_extractor import (
                    extract_rp_from_document,
                )
                dart = DartClient()
                if not dart.enabled:
                    rp_source = "dart_error:키 미설정(.env DART_API_KEY)"
                else:
                    corp_code = dart.find_corp_code(project.client)
                    if not corp_code:
                        rp_source = f"dart_error:'{project.client}' 회사 못 찾음"
                    else:
                        end_str = project.period_end.strftime("%Y%m%d")
                        rep = dart.latest_audit_report(corp_code, end_date=end_str)
                        if not rep:
                            rp_source = "dart_error:조회기간 내 보고서 없음"
                        else:
                            doc = dart.fetch_document(rep["rcept_no"])
                            if not doc:
                                rp_source = "dart_error:보고서 원문 fetch 실패"
                            else:
                                names = extract_rp_from_document(doc)
                                if names:
                                    rp_names = set(names)
                                    rp_source = (
                                        f"dart:{rep['report_nm']}@{rep['rcept_dt']}"
                                    )
                                else:
                                    rp_source = (
                                        f"dart_error:{rep['report_nm']} 특수관계자 주석 추출 0건"
                                    )
            except Exception as e:
                import logging
                logging.getLogger("cc_sampling_v2.ingest").warning(
                    "DART RP fetch failed: %s", e
                )
                rp_source = f"dart_error:{e}"

        from src.domain.party_normalize import (
            build_synonym_groups, load_default_synonyms, merge_synonym_maps,
            normalize_party_name,
        )
        # default(한·영·중) + RP 시트 synonym → 머지
        synonym_map = merge_synonym_maps(
            load_default_synonyms(),
            build_synonym_groups(synonym_groups),
        )
        # RP canonical 집합 — 루프 밖 1회 계산 (행마다 재계산 방지, O(N·M)→O(N+M)).
        rp_canons: set[str] = set()
        for _rp in rp_names:
            _n = normalize_party_name(_rp)
            if _n:
                rp_canons.add(synonym_map.get(_n, _n))

        allow_map: dict[str, dict] = {}
        if allowance_path is not None:
            allow_sheet = self._auto_sheet(allowance_path, "ALLOWANCE")
            if allow_sheet:
                allow_map = parse_allowance(allowance_path, allow_sheet)

        # 3) FS totals (cross-check 정보)
        fs_totals: dict[str, float] = {}
        if fs_path is not None:
            fs_sheet = self._auto_sheet(fs_path, "FS")
            if fs_sheet:
                fs_totals = parse_fs_totals(fs_path, fs_sheet)

        # 4) AR/AP 로드 + 플래그 + 환산 + persist
        from collections import defaultdict as _dd
        counts = {"AR": 0, "AP": 0}
        totals = {"AR": 0.0, "AP": 0.0}
        ledger_buckets = {"AR": _dd(float), "AP": _dd(float)}  # 계정과목(버킷)별 합
        confidences = {"AR": 0.0, "AP": 0.0}
        integrity_summary: dict = {"AR": {"checked": 0, "anomalies": 0},
                                     "AP": {"checked": 0, "anomalies": 0}}
        all_sheets_meta: list[dict] = []
        for kind_str in ("AR", "AP"):
            sns = sheet_assignments.get(kind_str, [])
            if not sns:
                continue

            all_enriched: list[Account] = []
            confidences_per_sheet: list[float] = []
            sheets_meta_local: list[dict] = []

            for sn_entry in sns:
                sn, kind_filter, spec = sn_entry
                if spec is not None:
                    accs, meta = load_account_sheet(
                        ledger_path, sn, kind_filter=kind_filter,
                        header_row_idx=spec["header_row"],
                        explicit_mapping=spec["column_map"],
                        kind_override=spec["kind"],
                    )
                else:
                    accs, meta = load_account_sheet(ledger_path, sn, kind_filter=kind_filter)
                confidences_per_sheet.append(meta["confidence"])
                integrity_summary[kind_str]["checked"] += meta.get("integrity_checked", 0)
                integrity_summary[kind_str]["anomalies"] += meta.get("integrity_anomalies", 0)
                # 매핑 누락 필드 식별 (gl_account·ccy는 optional)
                mapping = meta.get("mapping") or {}
                required_fields = ["party_id", "name", "balance"]
                missing = [f for f in required_fields if f not in mapping]
                sheets_meta_local.append({
                    "kind": kind_str, "sheet": sn,
                    "confidence": meta["confidence"],
                    "rows": len(accs),
                    "mapped_fields": list(mapping.keys()),
                    "missing_fields": missing,
                })

                for a in accs:
                    rate = a.fx_rate
                    if a.ccy.upper() != project.base_ccy.upper():
                        try:
                            rate = self.fx.lookup(a.ccy, project.period_end)
                        except Exception:
                            rate = a.fx_rate
                    try:
                        balance_krw = convert_to_base(
                            a.balance_orig, a.ccy, project.base_ccy, rate
                        )
                    except FxRateMissing:
                        balance_krw = a.balance_orig

                    allow = allow_map.get(a.party_id, {})
                    # RP 매칭 — 정규화(corp prefix·공백·구분자 제거) 후:
                    #  1) exact: `(주)코스맥스바이오`·`코스맥스바이오㈜` → `코스맥스바이오`
                    #  2) prefix: `코스맥스바이오 인천공장`·`코스맥스바이오지점`
                    #     → `코스맥스바이오인천공장`.startswith(`코스맥스바이오`) → RP O
                    #     (특관자 지점·공장·사업장 누락 방지; 길이 ≥3 RP명만 prefix 적용해
                    #      짧은 토큰 과매칭 방지).
                    # `삼성웰스토리코스맥스바이오` → startswith 불일치 → RP X (suffix 포함은 무시).
                    is_rp = False
                    if rp_canons:
                        norm_acc = normalize_party_name(a.name)
                        acc_canon = synonym_map.get(norm_acc, norm_acc) if norm_acc else ""
                        if acc_canon:
                            is_rp = acc_canon in rp_canons or any(
                                len(rp) >= 3 and acc_canon.startswith(rp)
                                for rp in rp_canons
                            )
                    all_enriched.append(Account(
                        party_id=a.party_id, name=a.name,
                        gl_account=a.gl_account,
                        balance_orig=a.balance_orig,
                        ccy=a.ccy.upper(), fx_rate=rate,
                        balance_krw=balance_krw,
                        is_related_party=is_rp,
                        is_bad_debt=bool(allow.get("is_bad_debt", False)),
                        allowance_amt=float(allow.get("allowance_amt",
                                                      a.allowance_amt)),
                        aging_bucket=a.aging_bucket,
                        src_sheet=a.src_sheet, src_row=a.src_row,
                        debit_amt=a.debit_amt, credit_amt=a.credit_amt,
                        business_number=a.business_number,
                        # excel_loader가 만든 breakdowns 보존 (gl_account 우선,
                        # 없으면 시트명). KRW 환산값으로 재계산.
                        account_breakdowns={
                            k: balance_krw for k in (a.account_breakdowns or {}).keys()
                        } if a.account_breakdowns else (
                            {a.src_sheet: balance_krw} if a.src_sheet else {}
                        ),
                    ))

            # 같은 거래처 집계 — 사업자번호/정규화 이름/synonym map 기준
            aggregated = _aggregate_by_party(all_enriched, synonym_map=synonym_map)

            self.acc.replace_all(project_id, Kind(kind_str), aggregated)
            counts[kind_str] = len(aggregated)
            totals[kind_str] = sum(abs(a.balance_krw) for a in aggregated)
            # 계정과목(BS 버킷)별 명세서 합 — 집계 전 원장 행(단일 gl_account)으로
            # 누적. account_breakdowns는 키별 값 오염 이력 있어 사용 안 함.
            from src.domain.account_classify import bs_bucket
            for a in all_enriched:
                label, _ = bs_bucket(a.gl_account, kind_str)
                ledger_buckets[kind_str][label] += a.balance_krw
            # 평균 confidence (시트별)
            confidences[kind_str] = (
                sum(confidences_per_sheet) / len(confidences_per_sheet)
                if confidences_per_sheet else 0.0
            )
            all_sheets_meta.extend(sheets_meta_local)

        needs_confirm = (
            (confidences["AR"] > 0 and confidences["AR"] < 0.95)
            or (confidences["AP"] > 0 and confidences["AP"] < 0.95)
        )

        # Cross-check (표준절차 1~2단계) — BS vs 명세서 비교
        def _cc(fs_v, ledger_v):
            fs_v = float(fs_v or 0)
            ledger_v = float(ledger_v or 0)
            diff = ledger_v - fs_v
            pct = (abs(diff) / max(abs(fs_v), 1) * 100) if fs_v else 0.0
            return {
                "fs_bs": fs_v, "ledger_total": ledger_v,
                "diff": diff, "diff_pct": pct,
                "match": pct < 1.0 and abs(diff) < 1000,  # ±1% 또는 ±1000원
            }

        # 계정과목(BS 버킷)별 대사 — 명세서합 vs BS금액. 범위가 같은 계정끼리 비교.
        fs_by_acct = (fs_totals.get("by_account") or {}) if isinstance(fs_totals, dict) else {}

        def _by_account(kind_str):
            # 모집단은 원장 → 원장에 있는 계정만 대사 표시. BS 전용 계정(예수금·
            # 미지급비용 등 조회대상 외)은 제외.
            rows = []
            for b, led in ledger_buckets[kind_str].items():
                bs = float(fs_by_acct.get(b, 0.0) or 0.0)
                diff = led - bs
                pct = (abs(diff) / max(abs(bs), 1) * 100) if bs else (
                    0.0 if abs(led) < 1e-9 else 100.0)
                rows.append({
                    "account": b, "ledger": led, "bs": bs,
                    "diff": diff, "diff_pct": pct,
                    "bs_found": abs(bs) > 1e-9,
                    "match": bool(bs) and pct < 1.0 and abs(diff) < 1000,
                })
            rows.sort(key=lambda r: -abs(r["ledger"]))
            return rows

        cross_check = {
            "AR": _cc(fs_totals.get("AR") if isinstance(fs_totals, dict) else 0, totals["AR"]),
            "AP": _cc(fs_totals.get("AP") if isinstance(fs_totals, dict) else 0, totals["AP"]),
            "by_account": {"AR": _by_account("AR"), "AP": _by_account("AP")},
            "integrity": integrity_summary,
        }

        cross_check["rp_source"] = rp_source
        cross_check["rp_count"] = len(rp_names)
        return IngestResult(
            project_id=project_id,
            ar_count=counts["AR"], ap_count=counts["AP"],
            ar_total_krw=totals["AR"], ap_total_krw=totals["AP"],
            confidence_ar=confidences["AR"],
            confidence_ap=confidences["AP"],
            fs_totals=fs_totals,
            needs_mapping_confirmation=needs_confirm,
            cross_check=cross_check,
            sheets_meta=all_sheets_meta,
        )

    @staticmethod
    def _auto_sheet(path: Path, target_kind: str) -> Optional[str]:
        wb = openpyxl.load_workbook(path, read_only=True)
        for sn in wb.sheetnames:
            if detect_sheet_kind(sn) == target_kind:
                return sn
        return wb.sheetnames[0] if wb.sheetnames else None


def _aggregate_by_party(
    accounts: list[Account],
    synonym_map: dict[str, str] = None,
) -> list[Account]:
    """거래처 fuzzy 합산.

    매칭 우선순위:
    1. 사업자번호 동일 → 합산
    2. 정규화 이름 동일 → 합산
    3. synonym_map 같은 그룹 → 합산

    합산 시 account_breakdowns dict는 시트별로 누적.
    RP/BAD/allowance는 OR로 합침. name은 한글 우선 + 가장 긴 것 채택.
    src_sheet는 여러 시트 join ("외상매출금+미수금" 형태).
    """
    from collections import defaultdict
    from src.domain.party_normalize import canonical_party_key

    synonym_map = synonym_map or {}
    groups: dict[str, list[Account]] = defaultdict(list)

    for a in accounts:
        key = canonical_party_key(a.name, a.business_number, synonym_map)
        if not key:
            key = a.party_id or f"_{id(a)}"  # fallback
        groups[key].append(a)

    out: list[Account] = []
    for key, items in groups.items():
        if len(items) == 1:
            out.append(items[0])
            continue

        first = items[0]
        balance_orig = sum(x.balance_orig for x in items)
        balance_krw = sum(x.balance_krw for x in items)
        debit_amt = sum(x.debit_amt for x in items)
        credit_amt = sum(x.credit_amt for x in items)
        allowance_amt = sum(x.allowance_amt for x in items)
        is_rp = any(x.is_related_party for x in items)
        is_bad = any(x.is_bad_debt for x in items)
        # 한글 포함 + 가장 긴 name 우선
        def _name_score(n):
            has_kr = any('가' <= ch <= '힯' for ch in n) if n else False
            return (1 if has_kr else 0, len(n or ""))
        name = max((x.name for x in items), key=_name_score)
        # business_number 첫 유효값
        biz = next((x.business_number for x in items if x.business_number), None)
        # party_id 첫 유효값
        pid = next((x.party_id for x in items if x.party_id), first.party_id)
        # src_sheet — 모든 시트 join
        sheets = sorted({x.src_sheet for x in items if x.src_sheet})
        src_sheet = "+".join(sheets) if sheets else first.src_sheet
        # account_breakdowns — dict merge (시트별 잔액)
        breakdowns: dict = {}
        for x in items:
            for sheet, amt in (x.account_breakdowns or {}).items():
                breakdowns[sheet] = breakdowns.get(sheet, 0.0) + amt

        out.append(Account(
            party_id=pid, name=name,
            gl_account=first.gl_account,
            balance_orig=balance_orig,
            ccy=first.ccy, fx_rate=first.fx_rate,
            balance_krw=balance_krw,
            is_related_party=is_rp,
            is_bad_debt=is_bad,
            allowance_amt=allowance_amt,
            aging_bucket=first.aging_bucket,
            src_sheet=src_sheet, src_row=first.src_row,
            debit_amt=debit_amt, credit_amt=credit_amt,
            business_number=biz,
            account_breakdowns=breakdowns,
        ))
    return out
