"""통합 조서(C100/AA100) Excel 빌더.

YAML 양식 로드 → state dict → openpyxl Workbook.
"""
from __future__ import annotations
import io
from pathlib import Path
from typing import Any
import yaml
import openpyxl
from openpyxl.utils import get_column_letter
from src.infrastructure.excel_writer.styles import (
    HEADER_FILL, HEADER_FONT, HEADER_ALIGN,
    BODY_FONT, NUM_ALIGN, TEXT_ALIGN, CELL_BORDER,
    TITLE_FONT, SUBTITLE_FONT, META_FONT, SIGN_FONT, TICKMARK_FONT,
)


_TEMPLATES_DIR = Path(__file__).resolve().parent.parent.parent.parent / "configs" / "templates"


def load_template(name: str) -> dict:
    p = _TEMPLATES_DIR / f"{name}.yaml"
    with open(p, encoding="utf-8") as f:
        return yaml.safe_load(f)


def _get_nested(d: dict, dotpath: str) -> Any:
    cur: Any = d
    for k in dotpath.split("."):
        if cur is None:
            return None
        cur = cur.get(k) if isinstance(cur, dict) else getattr(cur, k, None)
    return cur


def _write_header(ws, tpl: dict, state: dict) -> int:
    ws.cell(row=1, column=1, value=tpl["title"]).font = TITLE_FONT
    row = 3
    for h in tpl.get("header", []):
        ws.cell(row=row, column=1, value=h["label"]).font = META_FONT
        val = _get_nested(state, h["field"])
        if h.get("format") == "currency" and isinstance(val, (int, float)):
            cell = ws.cell(row=row, column=2, value=val)
            cell.number_format = "#,##0"
        else:
            ws.cell(row=row, column=2, value=val)
        ws.cell(row=row, column=2).font = META_FONT
        row += 1
    return row + 1


def _write_signature(ws, tpl: dict, start_row: int) -> int:
    sig = tpl.get("signature", {})
    ws.cell(row=start_row, column=1, value="-" * 40).font = SIGN_FONT
    r = start_row + 1
    for key in ("prepared_by", "reviewed_by", "date_field"):
        label = sig.get(key, key)
        ws.cell(row=r, column=1, value=label + ":").font = SIGN_FONT
        ws.cell(row=r, column=2, value="(   )").font = SIGN_FONT
        r += 1
    return r + 1


def _write_tickmark_legend(ws, tpl: dict, start_row: int) -> int:
    tm = tpl.get("tickmarks", {})
    if not tm:
        return start_row
    ws.cell(row=start_row, column=1, value="tickmark 범례").font = SUBTITLE_FONT
    r = start_row + 1
    for mark, desc in tm.items():
        ws.cell(row=r, column=1, value=mark).font = TICKMARK_FONT
        ws.cell(row=r, column=2, value=desc).font = META_FONT
        r += 1
    return r + 1


def build_workpaper(template_name: str, state: dict) -> bytes:
    tpl = load_template(template_name)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    kind = tpl["kind"]

    for sheet_spec in tpl["sheets"]:
        ws = wb.create_sheet(sheet_spec["name"])
        row = _write_header(ws, tpl, state)
        ws.cell(row=row, column=1, value=sheet_spec["title"]).font = SUBTITLE_FONT
        row += 2

        section = sheet_spec["section"]
        if kind == "COMBINED":
            if section == "design_summary":
                row = _write_summary_7620(ws, state, row)
            elif section == "control_sheet":
                row = _write_c100_control(ws, state, row)
            elif section == "size_determination":
                row = _write_c100_1_size(ws, state, row)
            elif section == "key_item":
                row = _write_c100_2_keyitem(ws, state, row)
            elif section == "mus_sample":
                row = _write_c100_3_mus(ws, state, row)
            elif section == "recovery_management":
                row = _write_recovery_management(ws, state, row)
            elif section == "alternative":
                row = _write_alternative_combined(ws, state, row)
        else:
            if section == "design_summary":
                row = _write_design_summary(ws, state, kind, row)
            elif section == "sendlist":
                row = _write_sendlist(ws, state, kind, row)
            elif section == "matching":
                row = _write_matching(ws, state, kind, row)
            elif section == "alternative":
                row = _write_alternative(ws, state, kind, row)
            elif section == "projection":
                row = _write_projection(ws, state, kind, row)

        row += 2
        row = _write_signature(ws, tpl, row)
        if section == "design_summary":
            row = _write_tickmark_legend(ws, tpl, row)

        for col_idx in range(1, 8):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_design_summary(ws, state, kind, row):
    pop = state.get("populations", {}).get(kind, {})
    samp = state.get("samples", {}).get(kind, {})
    items = [
        ("모집단 건수", pop.get("count", 0)),
        ("모집단 잔액 (KRW)", pop.get("total_krw", 0)),
        ("표본 건수", samp.get("count", 0)),
        ("표본 잔액 (KRW)", samp.get("total_krw", 0)),
        ("커버리지", (samp.get("total_krw", 0) / pop["total_krw"])
                       if pop.get("total_krw") else 0),
    ]
    for label, val in items:
        ws.cell(row=row, column=1, value=label).font = BODY_FONT
        c = ws.cell(row=row, column=2, value=val)
        c.font = BODY_FONT
        c.alignment = NUM_ALIGN
        c.number_format = "#,##0" if label != "커버리지" else "0.0%"
        row += 1
    return row


def _write_sendlist(ws, state, kind, row):
    headers = ["거래처코드", "거래처명", "계정과목", "잔액(KRW)", "통화", "선정사유"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
    row += 1
    for it in state.get("samples", {}).get(kind, {}).get("items", []):
        cells = [it["party_id"], it["name"], it["gl_account"],
                 it["balance_krw"], it["ccy"], it["selection_reason"]]
        for c_idx, v in enumerate(cells, start=1):
            c = ws.cell(row=row, column=c_idx, value=v)
            c.font = BODY_FONT
            c.border = CELL_BORDER
            if c_idx == 4:
                c.alignment = NUM_ALIGN
                c.number_format = "#,##0"
            else:
                c.alignment = TEXT_ALIGN
        row += 1
    return row


def _write_matching(ws, state, kind, row):
    headers = ["거래처", "장부잔액", "회신금액", "차이", "차이사유", "판정", "PDF경로"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
    row += 1
    for cf in state.get("confirmations", {}).get(kind, []):
        cells = [
            f"{cf['name']} ({cf['party_id']})",
            cf["expected"], cf["confirmed"], cf["diff"],
            cf.get("diff_reason"), cf.get("verdict"), cf.get("pdf_path"),
        ]
        for c_idx, v in enumerate(cells, start=1):
            c = ws.cell(row=row, column=c_idx, value=v)
            c.font = BODY_FONT
            c.border = CELL_BORDER
            if c_idx in (2, 3, 4):
                c.alignment = NUM_ALIGN
                c.number_format = "#,##0"
            else:
                c.alignment = TEXT_ALIGN
        row += 1
    return row


def _write_alternative(ws, state, kind, row):
    headers = ["거래처", "절차유형", "증빙금액(KRW)", "비고"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
    row += 1
    for ap_item in state.get("alternatives", {}).get(kind, []):
        cells = [
            f"{ap_item.get('name', '')} ({ap_item['party_id']})",
            ap_item["procedure_type"], ap_item["evidence_sum"],
            ap_item.get("note", ""),
        ]
        for c_idx, v in enumerate(cells, start=1):
            c = ws.cell(row=row, column=c_idx, value=v)
            c.font = BODY_FONT
            c.border = CELL_BORDER
            if c_idx == 3:
                c.alignment = NUM_ALIGN
                c.number_format = "#,##0"
            else:
                c.alignment = TEXT_ALIGN
        row += 1
    return row


def _write_projection(ws, state, kind, row):
    p = state.get("projection", {}).get(kind)
    if not p:
        ws.cell(row=row, column=1, value="(Projection 미산출)").font = META_FONT
        return row + 1
    items = [
        ("신뢰수준", p["confidence"]),
        ("Sampling interval (KRW)", p["sampling_interval"]),
        ("Projected misstatement", p["projected_misstatement"]),
        ("Basic precision", p["basic_precision"]),
        ("Incremental allowance", p["incremental_allowance"]),
        ("Upper limit", p["upper_limit"]),
        ("Tolerable", p["tolerable"]),
        ("판정", p["verdict"]),
    ]
    for label, val in items:
        ws.cell(row=row, column=1, value=label).font = BODY_FONT
        c = ws.cell(row=row, column=2, value=val)
        c.font = BODY_FONT
        c.alignment = NUM_ALIGN if isinstance(val, (int, float)) else TEXT_ALIGN
        if isinstance(val, (int, float)) and label != "신뢰수준":
            c.number_format = "#,##0"
        elif label == "신뢰수준":
            c.number_format = "0.0%"
        row += 1
    return row


def _write_design_summary_combined(ws, state, row):
    """채권·채무 모집단·표본 요약 — 종류 컬럼 + 합계 행."""
    headers = ["종류", "모집단 건수", "모집단 잔액(KRW)", "표본 건수",
               "표본 잔액(KRW)", "커버리지"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
    row += 1

    totals = {"pop_count": 0, "pop_total": 0.0,
              "samp_count": 0, "samp_total": 0.0}
    for kind_code, label in (("AR", "채권"), ("AP", "채무")):
        pop = state.get("populations", {}).get(kind_code, {})
        samp = state.get("samples", {}).get(kind_code, {})
        pop_c = pop.get("count", 0)
        pop_t = pop.get("total_krw", 0)
        samp_c = samp.get("count", 0)
        samp_t = samp.get("total_krw", 0)
        cov = samp_t / pop_t if pop_t else 0

        cells = [label, pop_c, pop_t, samp_c, samp_t, cov]
        for c_idx, v in enumerate(cells, start=1):
            c = ws.cell(row=row, column=c_idx, value=v)
            c.font = BODY_FONT
            c.border = CELL_BORDER
            if c_idx in (2, 4):
                c.alignment = NUM_ALIGN
                c.number_format = "#,##0"
            elif c_idx in (3, 5):
                c.alignment = NUM_ALIGN
                c.number_format = "#,##0"
            elif c_idx == 6:
                c.alignment = NUM_ALIGN
                c.number_format = "0.0%"
            else:
                c.alignment = TEXT_ALIGN
        totals["pop_count"] += pop_c
        totals["pop_total"] += pop_t
        totals["samp_count"] += samp_c
        totals["samp_total"] += samp_t
        row += 1

    # 합계
    cov_total = (totals["samp_total"] / totals["pop_total"]
                 if totals["pop_total"] else 0)
    cells = ["합계", totals["pop_count"], totals["pop_total"],
             totals["samp_count"], totals["samp_total"], cov_total]
    for c_idx, v in enumerate(cells, start=1):
        c = ws.cell(row=row, column=c_idx, value=v)
        c.font = SUBTITLE_FONT
        c.border = CELL_BORDER
        if c_idx in (2, 4):
            c.alignment = NUM_ALIGN
            c.number_format = "#,##0"
        elif c_idx in (3, 5):
            c.alignment = NUM_ALIGN
            c.number_format = "#,##0"
        elif c_idx == 6:
            c.alignment = NUM_ALIGN
            c.number_format = "0.0%"
        else:
            c.alignment = TEXT_ALIGN
    row += 1
    return row


def _write_sendlist_combined(ws, state, row):
    headers = ["종류", "거래처코드", "거래처명", "계정과목",
               "잔액(KRW)", "통화", "선정사유"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
    row += 1

    for kind_code, label in (("AR", "채권"), ("AP", "채무")):
        for it in state.get("samples", {}).get(kind_code, {}).get("items", []):
            cells = [label, it["party_id"], it["name"], it["gl_account"],
                     it["balance_krw"], it["ccy"], it["selection_reason"]]
            for c_idx, v in enumerate(cells, start=1):
                c = ws.cell(row=row, column=c_idx, value=v)
                c.font = BODY_FONT
                c.border = CELL_BORDER
                if c_idx == 5:
                    c.alignment = NUM_ALIGN
                    c.number_format = "#,##0"
                else:
                    c.alignment = TEXT_ALIGN
            row += 1
    return row


def _write_matching_combined(ws, state, row):
    headers = ["종류", "거래처", "장부잔액", "회신금액", "차이",
               "차이사유", "판정", "PDF경로"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
    row += 1

    for kind_code, label in (("AR", "채권"), ("AP", "채무")):
        for cf in state.get("confirmations", {}).get(kind_code, []):
            cells = [
                label,
                f"{cf['name']} ({cf['party_id']})",
                cf["expected"], cf["confirmed"], cf["diff"],
                cf.get("diff_reason"), cf.get("verdict"), cf.get("pdf_path"),
            ]
            for c_idx, v in enumerate(cells, start=1):
                c = ws.cell(row=row, column=c_idx, value=v)
                c.font = BODY_FONT
                c.border = CELL_BORDER
                if c_idx in (3, 4, 5):
                    c.alignment = NUM_ALIGN
                    c.number_format = "#,##0"
                else:
                    c.alignment = TEXT_ALIGN
            row += 1
    return row


def _write_alternative_combined(ws, state, row):
    headers = ["종류", "거래처", "절차유형", "증빙금액(KRW)", "비고"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
    row += 1

    for kind_code, label in (("AR", "채권"), ("AP", "채무")):
        for ap_item in state.get("alternatives", {}).get(kind_code, []):
            cells = [
                label,
                f"{ap_item.get('name', '')} ({ap_item['party_id']})",
                ap_item["procedure_type"], ap_item["evidence_sum"],
                ap_item.get("note", ""),
            ]
            for c_idx, v in enumerate(cells, start=1):
                c = ws.cell(row=row, column=c_idx, value=v)
                c.font = BODY_FONT
                c.border = CELL_BORDER
                if c_idx == 4:
                    c.alignment = NUM_ALIGN
                    c.number_format = "#,##0"
                else:
                    c.alignment = TEXT_ALIGN
            row += 1
    return row


def _write_projection_combined(ws, state, row):
    """채권·채무 별로 projection 표시 + 마지막에 합산 verdict."""
    # 컬럼: 항목 | 채권 | 채무 | 합산
    headers = ["항목", "채권", "채무", "합산"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
    row += 1

    ar = state.get("projection", {}).get("AR") or {}
    ap = state.get("projection", {}).get("AP") or {}

    def _val(d, key, default=0):
        return d.get(key, default) if d else default

    rows_data = [
        ("신뢰수준", _val(ar, "confidence"), _val(ap, "confidence"), None),
        ("Sampling interval", _val(ar, "sampling_interval"),
         _val(ap, "sampling_interval"), None),
        ("Projected misstatement",
         _val(ar, "projected_misstatement"),
         _val(ap, "projected_misstatement"),
         _val(ar, "projected_misstatement") + _val(ap, "projected_misstatement")),
        ("Basic precision",
         _val(ar, "basic_precision"), _val(ap, "basic_precision"),
         _val(ar, "basic_precision") + _val(ap, "basic_precision")),
        ("Incremental allowance",
         _val(ar, "incremental_allowance"), _val(ap, "incremental_allowance"),
         _val(ar, "incremental_allowance") + _val(ap, "incremental_allowance")),
        ("Upper limit",
         _val(ar, "upper_limit"), _val(ap, "upper_limit"),
         _val(ar, "upper_limit") + _val(ap, "upper_limit")),
        ("Tolerable",
         _val(ar, "tolerable"), _val(ap, "tolerable"),
         _val(ar, "tolerable") + _val(ap, "tolerable")),
    ]

    for label, ar_v, ap_v, sum_v in rows_data:
        cells = [label, ar_v, ap_v, sum_v if sum_v is not None else ""]
        for c_idx, v in enumerate(cells, start=1):
            c = ws.cell(row=row, column=c_idx, value=v)
            c.font = BODY_FONT
            c.border = CELL_BORDER
            if c_idx == 1:
                c.alignment = TEXT_ALIGN
            elif label == "신뢰수준":
                c.alignment = NUM_ALIGN
                if isinstance(v, (int, float)):
                    c.number_format = "0.0%"
            else:
                c.alignment = NUM_ALIGN
                if isinstance(v, (int, float)):
                    c.number_format = "#,##0"
        row += 1

    # 판정 행
    ar_verdict = ar.get("verdict") if ar else "—"
    ap_verdict = ap.get("verdict") if ap else "—"
    sum_upper = _val(ar, "upper_limit") + _val(ap, "upper_limit")
    sum_tol = _val(ar, "tolerable") + _val(ap, "tolerable")
    sum_verdict = "WITHIN_TOLERABLE" if sum_upper <= sum_tol else "EXCEED"

    cells = ["판정", ar_verdict, ap_verdict, sum_verdict]
    for c_idx, v in enumerate(cells, start=1):
        c = ws.cell(row=row, column=c_idx, value=v)
        c.font = SUBTITLE_FONT
        c.border = CELL_BORDER
        c.alignment = TEXT_ALIGN
    row += 1
    return row


def _write_summary_7620(ws, state, row):
    """샘플링 요약 — overview 행 + 강제포함/제외 내역."""
    headers = ["종류", "모집단 건수", "모집단 잔액(KRW)", "표본 건수",
               "강제포함(RP)", "강제포함(KEY)", "대표(REP)",
               "표본 잔액(KRW)", "커버리지"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
    row += 1

    totals = [0] * 6  # pop_c, pop_t, samp_c, samp_t, rp, key, rep
    for kind_code, label in (("AR", "채권"), ("AP", "채무")):
        pop = state.get("populations", {}).get(kind_code, {})
        samp = state.get("samples", {}).get(kind_code, {})
        items = samp.get("items", [])
        n_rp = sum(1 for it in items if it.get("selection_reason") == "FORCED_RP")
        n_key = sum(1 for it in items if it.get("selection_reason") == "FORCED_KEY")
        n_rep = sum(1 for it in items if it.get("selection_reason") == "REP")
        pop_c = pop.get("count", 0)
        pop_t = pop.get("total_krw", 0)
        samp_c = samp.get("count", 0)
        samp_t = samp.get("total_krw", 0)
        cov = samp_t / pop_t if pop_t else 0
        cells = [label, pop_c, pop_t, samp_c, n_rp, n_key, n_rep, samp_t, cov]
        for c_idx, v in enumerate(cells, start=1):
            c = ws.cell(row=row, column=c_idx, value=v)
            c.font = BODY_FONT
            c.border = CELL_BORDER
            if c_idx == 1:
                c.alignment = TEXT_ALIGN
            elif c_idx == 9:
                c.alignment = NUM_ALIGN
                c.number_format = "0.0%"
            else:
                c.alignment = NUM_ALIGN
                c.number_format = "#,##0"
        totals[0] += pop_c
        totals[1] += pop_t
        totals[2] += samp_c
        totals[3] += samp_t
        totals[4] += n_rp
        totals[5] += n_key
        row += 1

    # 합계
    cov_total = totals[3] / totals[1] if totals[1] else 0
    cells = ["합계", totals[0], totals[1], totals[2],
             totals[4], totals[5], totals[2] - totals[4] - totals[5],
             totals[3], cov_total]
    for c_idx, v in enumerate(cells, start=1):
        c = ws.cell(row=row, column=c_idx, value=v)
        c.font = SUBTITLE_FONT
        c.border = CELL_BORDER
        if c_idx == 1:
            c.alignment = TEXT_ALIGN
        elif c_idx == 9:
            c.alignment = NUM_ALIGN
            c.number_format = "0.0%"
        else:
            c.alignment = NUM_ALIGN
            c.number_format = "#,##0"
    row += 1
    return row


def _write_c100_control(ws, state, row):
    """C100 — 거래처별 계정과목별 컬럼 (동적 — 시트명 자동 수집)."""
    # default_aliases.yaml의 AR/AP sheet aliases 로드
    aliases_path = Path(__file__).resolve().parent.parent.parent.parent / \
        "configs" / "schema_mapping" / "default_aliases.yaml"
    with open(aliases_path, encoding="utf-8") as f:
        aliases = yaml.safe_load(f)
    ar_alias = set(aliases["sheets"].get("AR", []))
    ap_alias = set(aliases["sheets"].get("AP", []))

    # state items에서 모든 account_breakdowns 키 수집
    ar_items = state.get("samples", {}).get("AR", {}).get("items", [])
    ap_items = state.get("samples", {}).get("AP", {}).get("items", [])

    ar_sheets_used: set[str] = set()
    ap_sheets_used: set[str] = set()
    for it in ar_items + ap_items:
        bd = it.get("account_breakdowns", {}) or {}
        for sn in bd.keys():
            if sn in ar_alias:
                ar_sheets_used.add(sn)
            elif sn in ap_alias:
                ap_sheets_used.add(sn)
            else:
                # alias에 없는 시트 — kind 기준으로 분류
                if it in ar_items:
                    ar_sheets_used.add(sn)
                else:
                    ap_sheets_used.add(sn)

    # 컬럼 순서: alias yaml의 순서 우선, 그 외는 알파벳 순
    def _ordered(used, alias_list):
        ordered_from_alias = [s for s in alias_list if s in used]
        rest = sorted(used - set(ordered_from_alias))
        return ordered_from_alias + rest

    AR_ACCOUNTS = _ordered(ar_sheets_used, aliases["sheets"].get("AR", []))
    AP_ACCOUNTS = _ordered(ap_sheets_used, aliases["sheets"].get("AP", []))

    headers = (["No", "거래처코드", "거래처명", "사업자번호"]
               + AR_ACCOUNTS + ["채권 계"]
               + AP_ACCOUNTS + ["채무 계"]
               + ["채권+채무 합계", "선정사유"])
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
    row += 1

    # 거래처별 합산 — party_id가 있으면 그걸로, 없으면 정규화 name으로
    from src.domain.party_normalize import normalize_party_name
    by_party: dict[str, dict] = {}
    for it in ar_items + ap_items:
        pid = it.get("party_id") or ""
        name = it.get("name") or ""
        # canonical key: party_id 우선, 그 외 정규화 이름
        key = pid if pid else normalize_party_name(name)
        is_ar = it in ar_items
        ent = by_party.setdefault(key, {
            "party_id": pid,
            "name": name,
            "business_number": it.get("business_number"),
            "ar_acc": {a: 0 for a in AR_ACCOUNTS},
            "ap_acc": {a: 0 for a in AP_ACCOUNTS},
            "reasons": set(),
        })
        # name 갱신 — 더 긴 한글 이름 선호
        def _name_score(n):
            has_kr = any('가' <= ch <= '힯' for ch in n)
            return (1 if has_kr else 0, len(n))
        if _name_score(name) > _name_score(ent["name"]):
            ent["name"] = name
        # business_number 보강
        if not ent["business_number"] and it.get("business_number"):
            ent["business_number"] = it.get("business_number")
        # party_id 보강
        if not ent["party_id"] and pid:
            ent["party_id"] = pid

        bd = it.get("account_breakdowns", {}) or {}
        for sheet, amt in bd.items():
            if sheet in AR_ACCOUNTS:
                ent["ar_acc"][sheet] = ent["ar_acc"].get(sheet, 0) + abs(amt)
            elif sheet in AP_ACCOUNTS:
                ent["ap_acc"][sheet] = ent["ap_acc"].get(sheet, 0) + abs(amt)
        ent["reasons"].add(it.get("selection_reason", ""))

    def _total(e):
        return sum(e["ar_acc"].values()) + sum(e["ap_acc"].values())
    rows_sorted = sorted(by_party.values(), key=lambda e: -_total(e))

    for i, ent in enumerate(rows_sorted, start=1):
        ar_sum = sum(ent["ar_acc"].values())
        ap_sum = sum(ent["ap_acc"].values())
        total = ar_sum + ap_sum
        cells = [i, ent.get("party_id"), ent.get("name"), ent.get("business_number")]
        for acc in AR_ACCOUNTS:
            v = ent["ar_acc"].get(acc, 0)
            cells.append(v if v else "")
        cells.append(ar_sum if ar_sum else "")
        for acc in AP_ACCOUNTS:
            v = ent["ap_acc"].get(acc, 0)
            cells.append(v if v else "")
        cells.append(ap_sum if ap_sum else "")
        cells.append(total)
        cells.append(",".join(sorted(ent["reasons"])))

        for c_idx, v in enumerate(cells, start=1):
            c = ws.cell(row=row, column=c_idx, value=v)
            c.font = BODY_FONT
            c.border = CELL_BORDER
            if c_idx in (2, 3, 4) or c_idx == len(cells):
                c.alignment = TEXT_ALIGN
            else:
                c.alignment = NUM_ALIGN
                if isinstance(v, (int, float)) and v:
                    c.number_format = "#,##0"
        row += 1
    return row


def _write_c100_1_size(ws, state, row):
    """C100-1 표본규모 결정 — key-value 표."""
    proj = state.get("project", {})
    materiality = proj.get("materiality", 0)
    tolerable = proj.get("tolerable", 0)

    items = [
        ("감사목적", "보고기간말 현재 채권채무의 실재성·완전성 검토를 위한 표본 규모 결정"),
        ("", ""),
        ("[모집단 정보]", ""),
    ]
    for kind_code, label in (("AR", "채권 (AR) 모집단"), ("AP", "채무 (AP) 모집단")):
        pop = state.get("populations", {}).get(kind_code, {})
        items.append((f"  {label} 건수", pop.get("count", 0)))
        items.append((f"  {label} 잔액(KRW)", pop.get("total_krw", 0)))

    items.extend([
        ("", ""),
        ("[표본규모 결정요소]", ""),
        ("수행중요성 (materiality)", materiality),
        ("허용왜곡표시 (tolerable)", tolerable),
        ("Key item 기준금액", tolerable * 0.5),
        ("신뢰수준", "95% (RF 3.0)"),
        ("", ""),
        ("[샘플링 결과]", ""),
    ])

    for kind_code, label in (("AR", "AR 표본"), ("AP", "AP 표본")):
        samp = state.get("samples", {}).get(kind_code, {})
        s_items = samp.get("items", [])
        n_rp = sum(1 for it in s_items if it.get("selection_reason") == "FORCED_RP")
        n_key = sum(1 for it in s_items if it.get("selection_reason") == "FORCED_KEY")
        n_rep = sum(1 for it in s_items if it.get("selection_reason") == "REP")
        items.append((f"  {label} 총 건수", samp.get("count", 0)))
        items.append((f"  {label} 특관자(RP)", n_rp))
        items.append((f"  {label} Key item", n_key))
        items.append((f"  {label} Representative", n_rep))

    for label, val in items:
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = SUBTITLE_FONT if label.startswith("[") else BODY_FONT
        c1.alignment = TEXT_ALIGN
        if val != "":
            c2 = ws.cell(row=row, column=2, value=val)
            c2.font = BODY_FONT
            if isinstance(val, (int, float)):
                c2.alignment = NUM_ALIGN
                c2.number_format = "#,##0"
            else:
                c2.alignment = TEXT_ALIGN
        row += 1
    return row


def _write_c100_2_keyitem(ws, state, row):
    """C100-2 Key item 추출 — 강제포함 거래처 list (RP + KEY)."""
    headers = ["No", "종류", "거래처코드", "거래처명",
               "잔액(KRW)", "선정사유", "사유 상세"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
    row += 1

    forced_items = []
    for kind_code, label in (("AR", "채권"), ("AP", "채무")):
        for it in state.get("samples", {}).get(kind_code, {}).get("items", []):
            reason = it.get("selection_reason", "")
            if reason in ("FORCED_RP", "FORCED_KEY"):
                forced_items.append({**it, "kind_label": label, "kind": kind_code})

    forced_items.sort(key=lambda x: -abs(x.get("balance_krw", 0)))

    reason_desc = {
        "FORCED_RP": "특수관계자 (ISA 550 강제포함)",
        "FORCED_KEY": "Key item — 잔액 ≥ 기준금액 (ISA 530)",
    }

    for i, it in enumerate(forced_items, start=1):
        cells = [i, it["kind_label"], it.get("party_id"), it.get("name"),
                 abs(it.get("balance_krw", 0)),
                 it.get("selection_reason"),
                 reason_desc.get(it.get("selection_reason"), "")]
        for c_idx, v in enumerate(cells, start=1):
            c = ws.cell(row=row, column=c_idx, value=v)
            c.font = BODY_FONT
            c.border = CELL_BORDER
            if c_idx in (1, 5):
                c.alignment = NUM_ALIGN
                if c_idx == 5:
                    c.number_format = "#,##0"
            else:
                c.alignment = TEXT_ALIGN
        row += 1
    return row


def _write_c100_3_mus(ws, state, row):
    """C100-3 Representative MUS 표본 — 누적합산 + 선정 표시."""
    ws.cell(row=row, column=1, value="Representative sample (MUS 방법)").font = SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value="모집단 중 강제포함 외 REP로 선정된 거래처 목록").font = META_FONT
    row += 2

    headers = ["No", "종류", "거래처코드", "거래처명",
               "잔액(KRW)", "선정사유"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
    row += 1

    rep_items = []
    for kind_code, label in (("AR", "채권"), ("AP", "채무")):
        for it in state.get("samples", {}).get(kind_code, {}).get("items", []):
            if it.get("selection_reason") == "REP":
                rep_items.append({**it, "kind_label": label, "kind": kind_code})

    rep_items.sort(key=lambda x: -abs(x.get("balance_krw", 0)))

    for i, it in enumerate(rep_items, start=1):
        cells = [i, it["kind_label"], it.get("party_id"), it.get("name"),
                 abs(it.get("balance_krw", 0)), "REP (PPS 추출)"]
        for c_idx, v in enumerate(cells, start=1):
            c = ws.cell(row=row, column=c_idx, value=v)
            c.font = BODY_FONT
            c.border = CELL_BORDER
            if c_idx in (1, 5):
                c.alignment = NUM_ALIGN
                if c_idx == 5:
                    c.number_format = "#,##0"
            else:
                c.alignment = TEXT_ALIGN
        row += 1
    return row


def _write_recovery_management(ws, state, row):
    """C100-4 조회서 회수 관리 — 표본 전체 회신 추적."""
    # 요약 줄
    ar_items = state.get("samples", {}).get("AR", {}).get("items", [])
    ap_items = state.get("samples", {}).get("AP", {}).get("items", [])
    ar_confs = state.get("confirmations", {}).get("AR", [])
    ap_confs = state.get("confirmations", {}).get("AP", [])

    # party_id → confirmation 매핑
    conf_map_ar = {c.get("party_id"): c for c in ar_confs}
    conf_map_ap = {c.get("party_id"): c for c in ap_confs}

    # 요약
    n_sent = len(ar_items) + len(ap_items)
    n_recv_ar = sum(1 for c in ar_confs if c.get("verdict") is not None)
    n_recv_ap = sum(1 for c in ap_confs if c.get("verdict") is not None)
    n_recv = n_recv_ar + n_recv_ap
    n_match = sum(1 for c in ar_confs + ap_confs if c.get("verdict") == "MATCH")
    n_disc = sum(1 for c in ar_confs + ap_confs if c.get("verdict") == "DISCREPANCY")
    n_recon = sum(1 for c in ar_confs + ap_confs if c.get("verdict") == "RECONCILED")
    recv_rate = n_recv / n_sent if n_sent else 0

    ws.cell(row=row, column=1, value="회수 현황 요약").font = SUBTITLE_FONT
    row += 1
    summary = [
        ("발송 표본 수", n_sent),
        ("회신 수", n_recv),
        ("회신율", f"{recv_rate*100:.1f}%"),
        ("MATCH (일치)", n_match),
        ("DISCREPANCY (차이)", n_disc),
        ("RECONCILED (시점차이 등)", n_recon),
        ("미회신", n_sent - n_recv),
    ]
    for label, val in summary:
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = BODY_FONT
        c1.border = CELL_BORDER
        c2 = ws.cell(row=row, column=2, value=val)
        c2.font = BODY_FONT
        c2.alignment = NUM_ALIGN if isinstance(val, (int, float)) else TEXT_ALIGN
        if isinstance(val, (int, float)):
            c2.number_format = "#,##0"
        c2.border = CELL_BORDER
        row += 1
    row += 2

    # 본 표
    headers = ["No", "종류", "거래처코드", "거래처명",
               "장부잔액(KRW)", "발송일", "상태", "회신금액",
               "차이", "차이사유", "판정", "PDF경로"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
    row += 1

    # 모든 표본 거래처 (회신 여부 무관)
    all_samples = []
    for it in ar_items:
        conf = conf_map_ar.get(it.get("party_id"), {})
        all_samples.append({**it, "kind_label": "채권", "kind": "AR", "conf": conf})
    for it in ap_items:
        conf = conf_map_ap.get(it.get("party_id"), {})
        all_samples.append({**it, "kind_label": "채무", "kind": "AP", "conf": conf})

    # 잔액 큰 순
    all_samples.sort(key=lambda x: -abs(x.get("balance_krw", 0)))

    for i, it in enumerate(all_samples, start=1):
        conf = it.get("conf") or {}
        status = conf.get("status") or "PENDING"
        verdict = conf.get("verdict") or "—"
        if not conf:
            status = "발송됨"
            verdict = "미회신"
        cells = [
            i, it["kind_label"],
            it.get("party_id"), it.get("name"),
            it.get("balance_krw", 0),
            "",  # 발송일 (sent_at은 state.confirmations 별도 미노출 — 추후 보강)
            status,
            conf.get("confirmed"),
            conf.get("diff"),
            conf.get("diff_reason"),
            verdict,
            conf.get("pdf_path"),
        ]
        for c_idx, v in enumerate(cells, start=1):
            c = ws.cell(row=row, column=c_idx, value=v)
            c.font = BODY_FONT
            c.border = CELL_BORDER
            if c_idx in (1, 5, 8, 9):
                c.alignment = NUM_ALIGN
                if isinstance(v, (int, float)) and v is not None:
                    c.number_format = "#,##0"
            else:
                c.alignment = TEXT_ALIGN
        row += 1
    return row
