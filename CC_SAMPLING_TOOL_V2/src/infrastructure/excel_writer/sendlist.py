"""발송명단 Excel 생성."""
from __future__ import annotations
import io
from typing import Mapping
import openpyxl
from src.domain.entities import Account, Kind, SelectionReason
from src.infrastructure.excel_writer.styles import (
    HEADER_FILL, HEADER_FONT, HEADER_ALIGN,
    BODY_FONT, NUM_ALIGN, TEXT_ALIGN, CELL_BORDER,
)


COLUMNS = [
    ("종류", 8, "text"),
    ("거래처코드", 16, "text"),
    ("거래처명", 30, "text"),
    ("계정과목", 12, "text"),
    ("기말잔액(KRW)", 18, "num"),
    ("통화", 8, "text"),
    ("선정사유", 14, "text"),
]


def build_sendlist(
    client_name: str,
    period_end: str,
    samples: Mapping[Kind, list[tuple[Account, SelectionReason]]],
) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("발송명단")

    # 메타 행
    ws.append([f"회사명: {client_name}", f"평가기준일: {period_end}"])
    ws.append([])
    header_row_idx = 3

    # 헤더
    headers = [c[0] for c in COLUMNS]
    ws.append(headers)
    for col, (_, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
        ws.column_dimensions[cell.column_letter].width = width

    # 바디
    for kind in (Kind.AR, Kind.AP):
        for acc, reason in samples.get(kind, []):
            row = [
                kind.value, acc.party_id, acc.name, acc.gl_account,
                acc.balance_krw, acc.ccy, reason.value,
            ]
            ws.append(row)
            r_idx = ws.max_row
            for c_idx, (_, _, kind_t) in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = BODY_FONT
                cell.alignment = NUM_ALIGN if kind_t == "num" else TEXT_ALIGN
                cell.border = CELL_BORDER
                if kind_t == "num":
                    cell.number_format = "#,##0"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
