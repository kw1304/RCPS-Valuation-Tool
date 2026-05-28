"""거래처별 대손충당금 + 부실 플래그."""
from __future__ import annotations
from pathlib import Path
import openpyxl


def parse_allowance(path: Path, sheet_name: str
                    ) -> dict[str, dict[str, float | bool]]:
    """party_id → {allowance_amt, is_bad_debt}."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {}

    headers = [str(h or "").strip().lower() for h in rows[0]]

    def idx(*candidates):
        for i, h in enumerate(headers):
            if any(c in h for c in candidates):
                return i
        return -1

    i_party = idx("거래처코드", "거래처번호", "code", "party")
    i_allow = idx("충당금", "allowance")
    i_bad = idx("부실", "bad")

    out: dict[str, dict] = {}
    for row in rows[1:]:
        if i_party < 0 or i_party >= len(row) or row[i_party] is None:
            continue
        pid = str(row[i_party]).strip()
        allowance_amt = 0.0
        if 0 <= i_allow < len(row) and row[i_allow] is not None:
            try:
                allowance_amt = float(row[i_allow])
            except (TypeError, ValueError):
                allowance_amt = 0.0
        is_bad = False
        if 0 <= i_bad < len(row) and row[i_bad] is not None:
            val = str(row[i_bad]).strip().upper()
            is_bad = val in {"Y", "YES", "TRUE", "1", "부실"}
        out[pid] = {"allowance_amt": allowance_amt, "is_bad_debt": is_bad}
    return out
