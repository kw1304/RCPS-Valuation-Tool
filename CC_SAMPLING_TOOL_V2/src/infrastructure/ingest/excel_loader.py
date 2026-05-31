"""Excel 원장·시트 자동감지·로드.

설계서 §6.1 [2]. confidence < 0.95이면 UI 매핑확인 차단 (호출자 책임).
"""
from __future__ import annotations
from pathlib import Path
from typing import Optional, Literal
import yaml
import openpyxl
from src.domain.entities import Account
from src.infrastructure.ingest.header_detect import to_number


MappingConfidence = float
_CFG_PATH = Path(__file__).resolve().parent.parent.parent.parent / \
    "configs" / "schema_mapping" / "default_aliases.yaml"


def _load_aliases() -> dict:
    with open(_CFG_PATH, encoding="utf-8") as f:
        return yaml.safe_load(f)


_ALIASES = _load_aliases()


def detect_sheet_kind(sheet_name: str) -> Optional[Literal["AR","AP","FS","RP","ALLOWANCE","MIXED"]]:
    """시트명에서 종류 추정. 공백 제거 후 alias 매칭."""
    name = _normalize_header(sheet_name)
    for kind, aliases in _ALIASES["sheets"].items():
        for a in aliases:
            a_norm = _normalize_header(a)
            if a_norm and (a_norm == name or a_norm in name):
                return kind
    return None


# AR/AP 계정과목 keyword — 시트 내용으로 분류 fallback
_AR_GL_KEYWORDS = (
    "외상매출", "받을어음", "미수금", "미수수익", "선급",
    "단기대여", "장기대여", "임대보증", "임차보증",
    "매출채권", "receivable",
)
_AP_GL_KEYWORDS = (
    "외상매입", "지급어음", "미지급", "선수",
    "단기차입", "장기차입", "매입채무", "사채",
    "임대보증", "payable",
)


def classify_sheet_by_content(path, sheet_name: str) -> Optional[str]:
    """시트명 alias 매칭 실패 시 — 시트 안 계정과목 데이터로 AR/AP 판별.

    `계정과목명` 컬럼 값에 AR/AP keyword 빈도 비교.
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        ar_hits = 0
        ap_hits = 0
        # 첫 200행만 sampling
        for r in ws.iter_rows(min_row=1, max_row=200, values_only=True):
            for cell in r:
                if cell is None:
                    continue
                s = str(cell)
                for kw in _AR_GL_KEYWORDS:
                    if kw in s:
                        ar_hits += 1
                        break
                for kw in _AP_GL_KEYWORDS:
                    if kw in s:
                        ap_hits += 1
                        break
        if ar_hits == 0 and ap_hits == 0:
            return None
        if ar_hits > ap_hits * 1.5:
            return "AR"
        if ap_hits > ar_hits * 1.5:
            return "AP"
        if ar_hits > 0 and ap_hits > 0:
            return "MIXED"
        return "AR" if ar_hits > 0 else "AP"
    except Exception:
        return None


import re as _re_hdr
_HDR_WHITESPACE_RE = _re_hdr.compile(r"\s+")


def _normalize_header(h) -> str:
    """헤더 정규화: 모든 공백 제거 + 소문자.

    `거 래 처 명` `증    가` 같은 공백 포함 헤더도 매칭됨.
    """
    if h is None:
        return ""
    s = str(h).strip()
    s = _HDR_WHITESPACE_RE.sub("", s)
    return s.lower()


def detect_columns(headers: list[Optional[str]]) -> tuple[dict[str, int], MappingConfidence]:
    """헤더 행에서 컬럼명 → index 매핑.

    헤더·alias 모두 공백 제거 후 비교 (`거 래 처 명` → `거래처명`).
    """
    # ccy·gl_account는 optional. ccy 없으면 KRW. gl_account 없으면 시트명 fallback.
    required = ["party_id", "name", "balance"]
    mapping: dict[str, int] = {}
    used_idx: set[int] = set()
    norm_headers = [_normalize_header(h) for h in headers]

    # 1차: exact match — alias·header 모두 공백 제거 후 비교
    for field, aliases in _ALIASES["columns"].items():
        for alias in aliases:
            a_norm = _normalize_header(alias)
            matched = False
            for idx, h in enumerate(norm_headers):
                if idx in used_idx or not h:
                    continue
                if a_norm == h:
                    mapping[field] = idx
                    used_idx.add(idx)
                    matched = True
                    break
            if matched:
                break

    # 2차: partial match (substring) — 1차 누락 보완.
    for field, aliases in _ALIASES["columns"].items():
        if field in mapping:
            continue
        sorted_aliases = sorted(aliases, key=lambda a: -len(a))
        for alias in sorted_aliases:
            a_norm = _normalize_header(alias)
            matched = False
            for idx, h in enumerate(norm_headers):
                if idx in used_idx or not h:
                    continue
                if a_norm in h:
                    mapping[field] = idx
                    used_idx.add(idx)
                    matched = True
                    break
            if matched:
                break

    found_required = sum(1 for f in required if f in mapping)
    confidence = found_required / len(required)
    return mapping, confidence


def load_account_sheet(
    path: Path,
    sheet_name: str,
    kind_filter: Optional[str] = None,
    header_row_idx: Optional[int] = None,
    explicit_mapping: Optional[dict] = None,
    kind_override: Optional[str] = None,
) -> tuple[list[Account], dict]:
    """엑셀 시트에서 Account 목록 + meta 반환.

    Args:
        kind_filter: MIXED 시트일 때 `구분` 컬럼 값과 매칭되는 행만 반환 ("AR"/"AP").
                      None이면 모든 행. MIXED 양식(AR/AP 한 시트)에서 사용.
        header_row_idx: 명시 헤더 행(0-based). 주어지면 자동탐지 skip.
        explicit_mapping: 명시 컬럼매핑 {field: col_idx}. 주어지면 alias 자동매핑 skip.
        kind_override: 명시 시트 종류. 주어지면 시트명 추정 skip.

    헤더가 1행이 아닌 경우 (R1·R2에 제목, 실제 헤더 R3 등)도 자동 감지.

    Returns:
        (accounts, meta). meta = {sheet_kind, confidence, mapping}.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"sheet {sheet_name!r} not found in {path}")
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], {"sheet_kind": None, "confidence": 0.0, "mapping": {}}

    if explicit_mapping is not None:
        # 명시 매핑 주입 — 자동감지 전부 skip.
        mapping = {k: int(v) for k, v in explicit_mapping.items()}
        header_row_idx = int(header_row_idx or 0)
        confidence = 1.0
    else:
        # 헤더 행 자동탐지 — 첫 15행 중 confidence 가장 높은 행을 헤더로.
        # 한국 회사 ledger는 제목·요약·필터행이 위에 있고 헤더가 R3~R10에 흔함.
        best_idx = 0
        best_mapping: dict[str, int] = {}
        best_conf = -1.0
        for i in range(min(15, len(rows))):
            m, c = detect_columns(list(rows[i]))
            if c > best_conf:
                best_conf = c
                best_idx = i
                best_mapping = m
        mapping = best_mapping
        confidence = best_conf if best_conf > 0 else 0.0
        header_row_idx = best_idx
    sheet_kind = kind_override or detect_sheet_kind(sheet_name)

    _AR_VALUES = {"채권", "AR", "ar", "받을", "Receivable", "receivable"}
    _AP_VALUES = {"채무", "AP", "ap", "줄", "Payable", "payable"}
    kind_col_idx = mapping.get("kind_col")
    # MIXED(채권채무 혼재) 시트인데 '구분' 컬럼이 없으면 AR/AP 분리 불가.
    # 가드 없으면 필터가 통째로 skip돼 같은 행이 AR·AP 양쪽 모집단에 이중 적재됨.
    if kind_filter and kind_col_idx is None:
        raise ValueError(
            f"'{sheet_name}' 채권·채무 혼재 시트지만 '구분(AR/AP)' 컬럼 매핑이 없습니다. "
            f"구분 컬럼을 지정하거나 시트 종류를 채권/채무 중 하나로 바꾸세요."
        )

    accounts: list[Account] = []
    # 정합성 검증 카운터 (기초+차변-대변 ≠ 기말 행 수).
    # opening·debit·credit·balance 4개 모두 매핑됐을 때만 — credit 누락 시
    # credit_amt=0으로 계산돼 회수(수금) 있는 정상 거래처가 전부 anomaly로 오탐됨.
    integrity_check_enabled = (
        "opening" in mapping and "balance" in mapping
        and "debit" in mapping and "credit" in mapping
    )
    integrity_anomalies = 0
    integrity_checked = 0
    for r_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        if "balance" not in mapping:
            break
        # name 또는 party_id 둘 중 하나는 있어야 매칭 가능
        if "party_id" not in mapping and "name" not in mapping:
            break

        def cell(field, default=None):
            i = mapping.get(field)
            if i is None or i >= len(row):
                return default
            v = row[i]
            return default if v is None else v

        # MIXED 시트 행별 AR/AP 필터링
        if kind_filter and kind_col_idx is not None:
            kv = row[kind_col_idx] if kind_col_idx < len(row) else None
            kv_s = str(kv).strip() if kv is not None else ""
            row_kind = None
            if kv_s in _AR_VALUES:
                row_kind = "AR"
            elif kv_s in _AP_VALUES:
                row_kind = "AP"
            if row_kind != kind_filter:
                continue

        party_id = str(cell("party_id", "")).strip()
        name = str(cell("name", "")).strip()
        # party_id 없으면 name으로 식별 (fuzzy 집계 단계에서 통합 매칭됨)
        if not party_id and not name:
            continue
        # skip summary/subtotal rows — 내부 공백 무시 ("합 계"·"소 계"·"총 계"도 차단)
        import re as _re_sum
        _SUMMARY = {"합계", "소계", "계", "총계", "누계", "total", "subtotal", "sum"}
        pid_ns = _re_sum.sub(r"\s", "", party_id).lower()
        name_ns = _re_sum.sub(r"\s", "", name).lower()
        if pid_ns in _SUMMARY or name_ns in _SUMMARY:
            continue
        gl_account = str(cell("gl_account", "")).strip()
        # BTI 양식 subtotal: party_id만 있고 name·계정과목(원본 셀) 둘 다 빈값
        # (거래처 블록 끝 합계 행 — 잔액 컬럼에 거래처 합계 들어감 → 이중계상).
        # 반드시 시트명 backfill 前에 판정해야 함 (backfill 후엔 gl_account가 항상
        # 채워져 가드가 죽고 합계행이 통과됨).
        if "gl_account" in mapping and not gl_account and not name:
            continue
        # gl_account 컬럼 없거나 빈값이면 시트명에서 추출.
        # "1_외상매출금(10800)" → "외상매출금" / "외상매출금" → "외상매출금"
        if not gl_account:
            import re as _re
            sn_clean = _re.sub(r"^[\d_\s\-]+", "", sheet_name)  # 앞 숫자·_·공백 제거
            sn_clean = _re.sub(r"\([^)]*\)", "", sn_clean).strip()  # 괄호 제거
            sn_clean = sn_clean.strip("_- ")
            if sn_clean:
                gl_account = sn_clean
        # 숫자 파싱은 공용 to_number 사용 — 회계 괄호음수 (1,234)·통화기호·콤마 처리.
        # (로컬 _to_float는 괄호음수 미지원 → 충당금/FS 시트와 값 불일치·침묵손실 유발했음)
        balance_orig = to_number(cell("balance", 0))
        business_number = str(cell("business_number", "") or "").strip() or None
        ccy = str(cell("ccy", "KRW")).strip() or "KRW"
        fx_rate = to_number(cell("fx_rate", 1.0), 1.0)
        balance_krw = balance_orig * fx_rate
        debit_amt = to_number(cell("debit", 0)) * fx_rate
        credit_amt = to_number(cell("credit", 0)) * fx_rate
        # AP 정합성 검증용 — 기초+증감(=차변-대변) = 기말 비교
        opening_amt = to_number(cell("opening", 0)) * fx_rate
        # KRW 시트만 정합성 검증 (외화 시트는 기초 환율이 달라 false positive)
        if integrity_check_enabled and ccy.upper() == "KRW":
            expected_end = opening_amt + debit_amt - credit_amt
            integrity_checked += 1
            # ±1원 또는 잔액의 ±0.1% 이내 OK
            tol = max(1.0, abs(balance_krw) * 0.001)
            if abs(expected_end - balance_krw) > tol:
                integrity_anomalies += 1

        # account_breakdowns key — gl_account 컬럼 우선, 없으면 시트명.
        # BTI 양식: 한 시트(`채권`)에 여러 계정과목 행 — gl_account별 분리 필요.
        # 네오 양식: 시트명 자체가 계정과목 (`외상매출금` 등) — fallback.
        breakdown_key = gl_account if gl_account else sheet_name
        accounts.append(Account(
            party_id=party_id, name=name, gl_account=gl_account,
            balance_orig=balance_orig, ccy=ccy, fx_rate=fx_rate,
            balance_krw=balance_krw,
            aging_bucket=str(cell("aging", "")).strip() or None,
            allowance_amt=float(cell("allowance", 0) or 0),
            src_sheet=sheet_name, src_row=r_idx,
            debit_amt=debit_amt, credit_amt=credit_amt,
            business_number=business_number,
            account_breakdowns={breakdown_key: balance_krw},
        ))

    return accounts, {
        "sheet_kind": sheet_kind,
        "confidence": confidence,
        "mapping": mapping,
        "integrity_checked": integrity_checked,
        "integrity_anomalies": integrity_anomalies,
    }
