"""특수관계자 거래처명 set 반환.

거래처명 컬럼 자동감지 (헤더에 '거래처명'/'상대거래선'/'name' 등 매칭).
헤더 자동감지 실패 시 첫 컬럼 fallback.
"""
from __future__ import annotations
from pathlib import Path
import openpyxl


_NAME_ALIASES = ["거래처명", "거래처", "상대거래선", "상호", "회사명",
                  "name", "company", "party"]


def parse_related_parties(path: Path, sheet_name: str) -> set[str]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return set()
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return set()

    headers = [str(h or "").strip().lower() for h in rows[0]]
    name_idx = -1
    for i, h in enumerate(headers):
        if any(a.lower() in h for a in _NAME_ALIASES):
            name_idx = i
            break
    if name_idx < 0:
        name_idx = 0  # fallback first column

    out: set[str] = set()
    for r in rows[1:]:
        if not r or len(r) <= name_idx:
            continue
        v = r[name_idx]
        if v is None:
            continue
        s = str(v).strip()
        if s:
            out.add(s)
    return out


def parse_rp_synonym_groups(path: Path, sheet_name: str) -> list[list[str]]:
    """RP 시트의 각 행 = 동일 회사 이름들 그룹.

    7620 같은 양식: 한 행에 ['상대거래선명*', '상대거래선명(영어)*']
    각 행을 그룹 list (한·영 모두 동일 회사).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    out: list[list[str]] = []
    for r in rows[1:]:
        if not r:
            continue
        # 각 셀이 다른 표기 — 같은 회사로 묶음
        group: list[str] = []
        for v in r:
            if v is None:
                continue
            s = str(v).strip()
            if s and not s.replace(".", "").isdigit():  # NO 같은 숫자 skip
                group.append(s)
        if group:
            out.append(group)
    return out
