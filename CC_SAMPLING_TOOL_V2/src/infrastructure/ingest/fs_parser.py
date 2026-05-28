"""재무제표 — AR/AP 합계 추출 (모집단 cross-check용)."""
from __future__ import annotations
from pathlib import Path
import openpyxl


_AR_LABELS = {"매출채권", "외상매출금", "trade receivables"}
_AP_LABELS = {"매입채무", "외상매입금", "trade payables"}


def parse_fs_totals(path: Path, sheet_name: str) -> dict[str, float]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    out: dict[str, float] = {}
    for row in ws.iter_rows(values_only=True):
        if len(row) < 2:
            continue
        label = str(row[0] or "").strip().lower()
        try:
            amount = float(row[1] or 0)
        except (TypeError, ValueError):
            continue
        if any(lab in label for lab in _AR_LABELS):
            out["AR"] = out.get("AR", 0) + amount
        if any(lab in label for lab in _AP_LABELS):
            out["AP"] = out.get("AP", 0) + amount
    return out
