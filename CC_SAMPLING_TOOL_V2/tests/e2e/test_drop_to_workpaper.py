"""E2E — Phase 4 final: 전 라이프사이클 + 워크페이퍼 다운로드."""
import pytest
import io
from pathlib import Path
from unittest.mock import MagicMock
import openpyxl
from api.app import create_app
from src.infrastructure.db.session import make_engine, make_session
from src.infrastructure.db.models import Base


FIXTURES = Path(__file__).parent / "fixtures"


def _dynamic_pdf(name: str, amount: float) -> bytes:
    try:
        from reportlab.pdfgen import canvas
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        from reportlab.pdfbase import pdfmetrics
        pdfmetrics.registerFont(UnicodeCIDFont("HYSMyeongJo-Medium"))
    except ImportError:
        pytest.skip("reportlab not installed")
    buf = io.BytesIO()
    c = canvas.Canvas(buf)
    c.setFont("HYSMyeongJo-Medium", 11)
    c.drawString(50, 800, "회신서")
    c.drawString(50, 780, f"조회처: {name}")
    c.drawString(50, 760, f"잔액: {int(amount):,}원")
    c.save()
    return buf.getvalue()


@pytest.fixture(scope="module")
def fixtures_ready():
    for n in ("dummy_ledger", "dummy_fs", "dummy_rp", "dummy_allowance"):
        if not (FIXTURES / f"{n}.xlsx").exists():
            pytest.skip(f"fixture {n}.xlsx missing")


@pytest.fixture
def client(fixtures_ready):
    engine = make_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    SF = make_session(engine)
    fx_mock = MagicMock()
    fx_mock.lookup.return_value = 1300.0
    app = create_app(testing=True, session_factory=SF)
    app.config["FX_CLIENT"] = fx_mock
    return app.test_client()


def _file(name):
    return (open(FIXTURES / f"{name}.xlsx", "rb"), f"{name}.xlsx")


def test_e2e_drop_to_workpaper(client):
    r = client.post("/api/projects", json={
        "client": "DUMMY_CLIENT", "period_end": "2025-12-31",
        "base_ccy": "KRW", "materiality": 50_000_000, "tolerable": 25_000_000,
    })
    pid = r.get_json()["id"]
    client.post(f"/api/projects/{pid}/ingest", data={
        "ledger": _file("dummy_ledger"), "fs": _file("dummy_fs"),
        "rp": _file("dummy_rp"), "allowance": _file("dummy_allowance"),
    }, content_type="multipart/form-data")

    for kind in ("AR", "AP"):
        client.post(f"/api/projects/{pid}/sampling/design", json={
            "kind": kind, "confidence": 0.95, "expected_ms_pct": 0.0,
            "key_threshold": 5_000_000, "n_strata": 4, "seed": 42,
        })

    r = client.get(f"/api/projects/{pid}/sendlist")
    assert r.status_code == 200

    state = client.get(f"/api/projects/{pid}/state").get_json()
    ar_items = state["samples"]["AR"]["items"]
    if ar_items:
        target = ar_items[0]
        pdf = _dynamic_pdf(target["name"], target["balance_krw"])
        client.post(f"/api/projects/{pid}/confirmations/upload",
                    data={"kind": "AR",
                          "pdf": (io.BytesIO(pdf), "x.pdf")},
                    content_type="multipart/form-data")
    if len(ar_items) > 1:
        client.post(f"/api/projects/{pid}/confirmations/correct", json={
            "kind": "AR", "party_id": ar_items[1]["party_id"],
            "confirmed": ar_items[1]["balance_krw"] * 0.95,
        })

    for kind in ("AR", "AP"):
        client.post(f"/api/projects/{pid}/projection",
                    json={"kind": kind, "confidence": 0.95})

    r = client.get(f"/api/projects/{pid}/workpaper/c100")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert "C100_summary" in wb.sheetnames
    assert "C101_sendlist" in wb.sheetnames
    assert "C102_matching" in wb.sheetnames
    assert "C103_alternative" in wb.sheetnames
    assert "C104_projection" in wb.sheetnames

    r = client.get(f"/api/projects/{pid}/workpaper/aa100")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert "AA100_summary" in wb.sheetnames

    r = client.delete(f"/api/projects/{pid}")
    assert r.status_code == 200
    r = client.get(f"/api/projects/{pid}")
    assert r.status_code == 404
