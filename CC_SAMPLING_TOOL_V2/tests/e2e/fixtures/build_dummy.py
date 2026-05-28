"""더미 데이터 4종 생성 스크립트.

실행: python tests/e2e/fixtures/build_dummy.py
출력: tests/e2e/fixtures/{ledger, fs, rp, allowance}.xlsx
"""
from __future__ import annotations
from pathlib import Path
import random
import openpyxl


OUT = Path(__file__).parent


def build_ledger():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ar = wb.create_sheet("매출채권")
    ar.append(["거래처코드", "거래처명", "계정과목", "기말잔액", "통화", "환율",
               "차변", "대변"])
    rng = random.Random(7)
    # AR050-054는 충당금명세와 정합되도록 고정 잔액
    fixed_bal = {
        "AR050": 500_000,
        "AR051": 800_000,
        "AR052": 200_000,
        "AR053": 600_000,
        "AR054": 400_000,
    }
    for i in range(120):
        pid = f"AR{i:03d}"
        bal_choice = rng.choice([10_000, 50_000, 100_000, 500_000, 2_000_000, 10_000_000])
        bal_jitter = rng.uniform(0.5, 2.0)
        if pid in fixed_bal:
            bal = float(fixed_bal[pid])
        else:
            bal = bal_choice * bal_jitter
        ccy = "USD" if i < 5 else "KRW"
        fx = 1300.0 if ccy == "USD" else 1.0
        debit = bal * rng.uniform(2.0, 6.0)
        credit = bal * rng.uniform(1.5, 5.0)
        ar.append([pid, f"고객사{i:03d}", "11200",
                   round(bal, 0), ccy, fx,
                   round(debit, 0), round(credit, 0)])

    ap = wb.create_sheet("매입채무")
    ap.append(["거래처코드", "거래처명", "계정과목", "기말잔액", "통화", "환율",
               "차변", "대변"])
    for i in range(80):
        bal = rng.choice([5_000, 30_000, 200_000, 1_000_000]) * rng.uniform(0.5, 2.0)
        ccy = "USD" if i < 3 else "KRW"
        fx = 1300.0 if ccy == "USD" else 1.0
        debit = bal * rng.uniform(1.5, 5.0)
        credit = bal * rng.uniform(2.0, 6.0)
        ap.append([f"AP{i:03d}", f"공급사{i:03d}", "21100",
                   round(bal, 0), ccy, fx,
                   round(debit, 0), round(credit, 0)])

    wb.save(OUT / "dummy_ledger.xlsx")


def build_fs():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("재무제표")
    ws.append(["계정", "기말금액"])
    ws.append(["매출채권", 250_000_000])
    ws.append(["매입채무", 120_000_000])
    wb.save(OUT / "dummy_fs.xlsx")


def build_rp():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("특관자")
    ws.append(["거래처명"])
    for i in range(5):
        ws.append([f"고객사{i:03d}"])  # AR 첫 5건 RP 매칭
    wb.save(OUT / "dummy_rp.xlsx")


def build_allowance():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("충당금명세")
    ws.append(["거래처코드", "거래처명", "잔액", "충당금", "부실여부"])
    # 3건 부실 + 2건 부분충당
    ws.append(["AR050", "고객사050", 500_000, 500_000, "Y"])
    ws.append(["AR051", "고객사051", 800_000, 800_000, "Y"])
    ws.append(["AR052", "고객사052", 200_000, 200_000, "Y"])
    ws.append(["AR053", "고객사053", 600_000, 300_000, "N"])
    ws.append(["AR054", "고객사054", 400_000, 100_000, "N"])
    wb.save(OUT / "dummy_allowance.xlsx")


if __name__ == "__main__":
    build_ledger()
    build_fs()
    build_rp()
    build_allowance()
    print("dummy fixtures built at:", OUT)
