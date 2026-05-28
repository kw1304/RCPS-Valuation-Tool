import pytest
import io
import openpyxl
from src.infrastructure.excel_writer.workpaper import (
    build_workpaper, load_template,
)


def _sample_state():
    return {
        "project": {
            "client": "ACME", "period_end": "2025-12-31",
            "base_ccy": "KRW", "materiality": 500_000_000,
            "tolerable": 250_000_000,
        },
        "populations": {"AR": {"count": 100, "total_krw": 5_000_000_000},
                         "AP": {"count": 0, "total_krw": 0}},
        "samples": {"AR": {"count": 10, "total_krw": 1_500_000_000,
                            "items": []}, "AP": {"count": 0, "items": []}},
        "confirmations": {"AR": [], "AP": []},
        "alternatives": {"AR": [], "AP": []},
        "projection": {"AR": None, "AP": None},
    }


def test_load_template_c100():
    tpl = load_template("c100")
    assert tpl["workpaper_code"] == "C100"
    assert tpl["kind"] == "AR"
    assert len(tpl["sheets"]) == 5


def test_build_c100_minimal():
    state = _sample_state()
    blob = build_workpaper("c100", state)
    assert isinstance(blob, bytes) and len(blob) > 0
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    assert "C100_summary" in wb.sheetnames
    assert "C101_sendlist" in wb.sheetnames
    assert "C104_projection" in wb.sheetnames


def test_workpaper_header_contains_client():
    state = _sample_state()
    blob = build_workpaper("c100", state)
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    ws = wb["C100_summary"]
    flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert any("ACME" in v for v in flat)


def test_workpaper_signature_block():
    state = _sample_state()
    blob = build_workpaper("c100", state)
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    ws = wb["C100_summary"]
    flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert any("작성자" in v for v in flat)


def test_c100_full_state():
    state = {
        "project": {
            "client": "ACME", "period_end": "2025-12-31",
            "base_ccy": "KRW", "materiality": 500_000_000,
            "tolerable": 250_000_000,
        },
        "populations": {
            "AR": {"count": 120, "total_krw": 250_000_000},
            "AP": {"count": 0, "total_krw": 0},
        },
        "samples": {
            "AR": {"count": 5, "total_krw": 50_000_000, "items": [
                {"party_id": "AR000", "name": "고객사000", "gl_account": "11200",
                 "balance_krw": 10_000_000, "ccy": "KRW",
                 "selection_reason": "FORCED_RP",
                 "is_related_party": True, "is_bad_debt": False},
                {"party_id": "AR050", "name": "고객사050", "gl_account": "11200",
                 "balance_krw": 8_000_000, "ccy": "KRW",
                 "selection_reason": "REP",
                 "is_related_party": False, "is_bad_debt": False},
            ]},
            "AP": {"count": 0, "items": []},
        },
        "confirmations": {
            "AR": [
                {"party_id": "AR000", "name": "고객사000",
                 "expected": 10_000_000, "confirmed": 10_000_000,
                 "diff": 0, "diff_reason": None, "verdict": "MATCH",
                 "status": "RECEIVED", "pdf_path": "/tmp/c1.pdf"},
            ],
            "AP": [],
        },
        "alternatives": {
            "AR": [{"party_id": "AR050", "name": "고객사050",
                    "procedure_type": "후속회수",
                    "evidence_sum": 8_000_000, "note": "회수증빙"}],
            "AP": [],
        },
        "projection": {
            "AR": {"confidence": 0.95, "sampling_interval": 50_000_000,
                   "tolerable": 250_000_000,
                   "projected_misstatement": 0,
                   "basic_precision": 150_000_000,
                   "incremental_allowance": 0,
                   "upper_limit": 150_000_000,
                   "verdict": "WITHIN_TOLERABLE",
                   "strata_snapshot": []},
            "AP": None,
        },
    }
    blob = build_workpaper("c100", state)
    wb = openpyxl.load_workbook(io.BytesIO(blob))

    ws = wb["C101_sendlist"]
    flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert "AR000" in flat
    assert "AR050" in flat

    ws = wb["C102_matching"]
    flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert "MATCH" in flat

    ws = wb["C103_alternative"]
    flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert "후속회수" in flat

    ws = wb["C104_projection"]
    flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert "WITHIN_TOLERABLE" in flat


def test_aa100_kind_ap():
    state = {
        "project": {
            "client": "ACME", "period_end": "2025-12-31",
            "base_ccy": "KRW", "materiality": 100_000_000,
            "tolerable": 50_000_000,
        },
        "populations": {
            "AR": {"count": 0, "total_krw": 0},
            "AP": {"count": 80, "total_krw": 120_000_000},
        },
        "samples": {
            "AR": {"count": 0, "items": []},
            "AP": {"count": 3, "total_krw": 20_000_000, "items": [
                {"party_id": "AP000", "name": "공급사000", "gl_account": "21100",
                 "balance_krw": 5_000_000, "ccy": "KRW",
                 "selection_reason": "FORCED_KEY",
                 "is_related_party": False, "is_bad_debt": False},
            ]},
        },
        "confirmations": {"AR": [], "AP": []},
        "alternatives": {"AR": [], "AP": []},
        "projection": {
            "AR": None,
            "AP": {"confidence": 0.95, "sampling_interval": 40_000_000,
                   "tolerable": 50_000_000,
                   "projected_misstatement": 0,
                   "basic_precision": 120_000_000,
                   "incremental_allowance": 0,
                   "upper_limit": 120_000_000,
                   "verdict": "EXCEED",
                   "strata_snapshot": []},
        },
    }
    blob = build_workpaper("aa100", state)
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    assert "AA100_summary" in wb.sheetnames
    assert "AA101_sendlist" in wb.sheetnames

    ws = wb["AA101_sendlist"]
    flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert "AP000" in flat
    assert not any("AR000" in v for v in flat)

    ws = wb["AA104_projection"]
    flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert "EXCEED" in flat
