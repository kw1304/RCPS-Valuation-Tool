import pytest
from pathlib import Path
import openpyxl
from src.infrastructure.ingest.excel_loader import (
    detect_sheet_kind, detect_columns, load_account_sheet, MappingConfidence,
)


def _make_xlsx(tmp_path, sheets: dict[str, list[list]]) -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        for r in rows:
            ws.append(r)
    p = tmp_path / "test.xlsx"
    wb.save(p)
    return p


def test_detect_sheet_kind_exact_match():
    assert detect_sheet_kind("매출채권") == "AR"
    assert detect_sheet_kind("매입채무") == "AP"
    assert detect_sheet_kind("재무제표") == "FS"


def test_detect_sheet_kind_partial_match():
    assert detect_sheet_kind("매출채권 원장") == "AR"
    assert detect_sheet_kind("Trade Receivables 2024") == "AR"


def test_detect_sheet_kind_unknown():
    assert detect_sheet_kind("기타시트") is None


def test_detect_columns_korean(tmp_path):
    p = _make_xlsx(tmp_path, {
        "매출채권": [
            ["거래처코드", "거래처명", "계정과목", "기말잔액", "통화"],
            ["P1", "갑", "11200", 1_000_000, "KRW"],
        ],
    })
    wb = openpyxl.load_workbook(p)
    ws = wb["매출채권"]
    headers = [c.value for c in ws[1]]
    mapping, confidence = detect_columns(headers)
    assert mapping["party_id"] == 0
    assert mapping["name"] == 1
    assert mapping["gl_account"] == 2
    assert mapping["balance"] == 3
    assert mapping["ccy"] == 4
    assert confidence >= 0.95


def test_detect_columns_arbitrary_order(tmp_path):
    p = _make_xlsx(tmp_path, {
        "매출채권": [
            ["기말잔액", "거래처명", "통화", "계정", "거래처코드"],
        ],
    })
    wb = openpyxl.load_workbook(p)
    headers = [c.value for c in wb["매출채권"][1]]
    mapping, _ = detect_columns(headers)
    assert mapping["balance"] == 0
    assert mapping["party_id"] == 4


def test_detect_columns_low_confidence(tmp_path):
    p = _make_xlsx(tmp_path, {
        "매출채권": [
            ["col1", "col2", "col3"],
        ],
    })
    wb = openpyxl.load_workbook(p)
    headers = [c.value for c in wb["매출채권"][1]]
    _, confidence = detect_columns(headers)
    assert confidence < 0.95


def test_load_account_sheet_full(tmp_path):
    p = _make_xlsx(tmp_path, {
        "매출채권": [
            ["거래처코드", "거래처명", "계정과목", "기말잔액", "통화", "환율"],
            ["P1", "갑", "11200", 1_000_000, "KRW", 1.0],
            ["P2", "을", "11200", 5_000, "USD", 1300.0],
        ],
    })
    accs, meta = load_account_sheet(p, sheet_name="매출채권")
    assert len(accs) == 2
    assert accs[0].party_id == "P1"
    assert accs[1].ccy == "USD"
    assert accs[1].balance_orig == 5_000
    assert meta["sheet_kind"] == "AR"
    assert meta["confidence"] >= 0.95


def test_load_account_sheet_skips_blank_rows(tmp_path):
    p = _make_xlsx(tmp_path, {
        "매출채권": [
            ["거래처코드", "거래처명", "계정", "기말잔액"],
            ["P1", "갑", "11200", 1000],
            [None, None, None, None],
            ["P2", "을", "11200", 2000],
        ],
    })
    accs, _ = load_account_sheet(p, sheet_name="매출채권")
    assert len(accs) == 2
