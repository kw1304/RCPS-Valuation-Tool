import pytest
import io
from datetime import date
import openpyxl
from api.app import create_app
from src.infrastructure.db.session import make_engine, make_session
from src.infrastructure.db.models import Base
from src.infrastructure.db.repository import ProjectRepo


@pytest.fixture
def client():
    e = make_engine("sqlite:///:memory:")
    Base.metadata.create_all(e)
    SF = make_session(e)
    app = create_app(testing=True, session_factory=SF)
    s = SF()
    pid = ProjectRepo(s).create(
        client="X", period_end=date(2025, 12, 31), base_ccy="KRW",
        materiality=1, tolerable=1)
    s.close()
    return app.test_client(), pid


def test_download_c100(client):
    c, pid = client
    r = c.get(f"/api/projects/{pid}/workpaper/c100")
    assert r.status_code == 200
    assert r.content_type.startswith("application/vnd.openxmlformats")
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert "C100_summary" in wb.sheetnames


def test_download_aa100(client):
    c, pid = client
    r = c.get(f"/api/projects/{pid}/workpaper/aa100")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert "AA100_summary" in wb.sheetnames


def test_download_invalid_template_404(client):
    c, pid = client
    r = c.get(f"/api/projects/{pid}/workpaper/zzz")
    assert r.status_code == 404
