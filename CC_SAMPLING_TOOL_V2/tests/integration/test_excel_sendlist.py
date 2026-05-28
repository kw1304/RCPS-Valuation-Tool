import pytest
import io
import openpyxl
from src.domain.entities import Account, Kind, SelectionReason
from src.infrastructure.excel_writer.sendlist import build_sendlist


def _acc(pid, name, balance):
    return Account(party_id=pid, name=name, gl_account="11200",
                   balance_orig=balance, ccy="KRW", fx_rate=1.0,
                   balance_krw=balance)


def test_sendlist_builds_xlsx_bytes():
    selections = [
        (_acc("AR001", "고객사001", 1_000_000), SelectionReason.FORCED_RP),
        (_acc("AR002", "고객사002", 5_000_000), SelectionReason.FORCED_KEY),
    ]
    samples = {Kind.AR: selections, Kind.AP: []}
    blob = build_sendlist(client_name="ACME", period_end="2025-12-31",
                          samples=samples)
    assert isinstance(blob, (bytes, bytearray))
    assert len(blob) > 0
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    assert "발송명단" in wb.sheetnames


def test_sendlist_rows_present():
    selections = [
        (_acc("AR001", "고객사001", 1_000_000), SelectionReason.FORCED_RP),
    ]
    blob = build_sendlist(client_name="X", period_end="2025-12-31",
                          samples={Kind.AR: selections, Kind.AP: []})
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    ws = wb["발송명단"]
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) >= 2
    headers = rows[0]
    # 헤더는 메타 정보 또는 컬럼명 — 어느 위치에 있든 "거래처코드" 발견되어야
    # 메타가 위에 있을 수 있으니 모든 행 검색
    flat = [v for row in rows for v in row if v is not None]
    assert "거래처코드" in flat
    assert "거래처명" in flat


def test_sendlist_merges_ar_ap():
    ar = [(_acc("AR1", "ar", 100), SelectionReason.REP)]
    ap = [(_acc("AP1", "ap", 200), SelectionReason.REP)]
    blob = build_sendlist(client_name="X", period_end="2025-12-31",
                          samples={Kind.AR: ar, Kind.AP: ap})
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    ws = wb["발송명단"]
    rows = list(ws.iter_rows(values_only=True))
    flat = [v for row in rows for v in row if v is not None]
    assert "AR1" in flat
    assert "AP1" in flat
