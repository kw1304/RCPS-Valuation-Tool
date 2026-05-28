import pytest
import openpyxl
from datetime import date
from pathlib import Path
from unittest.mock import MagicMock
from src.application.ingest_uc import IngestUC
from src.domain.entities import Kind
from src.infrastructure.db.session import make_engine, make_session
from src.infrastructure.db.models import Base
from src.infrastructure.db.repository import ProjectRepo, AccountRepo


def _make_multi_sheet_ledger(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # 외상매출금 시트
    ar1 = wb.create_sheet("외상매출금")
    ar1.append(["거래처코드", "거래처명", "계정과목", "기말잔액", "통화"])
    ar1.append(["P1", "고객A", "11200", 1_000_000, "KRW"])
    ar1.append(["P2", "고객B", "11200", 2_000_000, "KRW"])
    # 미수금 시트 — AR로 합쳐져야
    ar2 = wb.create_sheet("미수금")
    ar2.append(["거래처코드", "거래처명", "계정과목", "기말잔액", "통화"])
    ar2.append(["P3", "고객C", "12100", 500_000, "KRW"])
    # 받을어음 시트 — AR로 합쳐져야
    ar3 = wb.create_sheet("받을어음")
    ar3.append(["거래처코드", "거래처명", "계정과목", "기말잔액", "통화"])
    ar3.append(["P4", "고객D", "11300", 300_000, "KRW"])
    # 외상매입금 시트 — AP
    ap1 = wb.create_sheet("외상매입금")
    ap1.append(["거래처코드", "거래처명", "계정과목", "기말잔액", "통화"])
    ap1.append(["P5", "공급A", "21100", 800_000, "KRW"])
    # 미지급금 시트 — AP로 합쳐져야
    ap2 = wb.create_sheet("미지급금")
    ap2.append(["거래처코드", "거래처명", "계정과목", "기말잔액", "통화"])
    ap2.append(["P6", "공급B", "22100", 200_000, "KRW"])
    p = tmp_path / "multi.xlsx"
    wb.save(p)
    return p


@pytest.fixture
def session():
    e = make_engine("sqlite:///:memory:")
    Base.metadata.create_all(e)
    S = make_session(e)
    s = S()
    yield s
    s.close()


def test_ingest_combines_ar_sheets(session, tmp_path):
    pid = ProjectRepo(session).create(
        client="X", period_end=date(2025, 12, 31), base_ccy="KRW",
        materiality=1, tolerable=1)
    ledger = _make_multi_sheet_ledger(tmp_path)
    fx = MagicMock(lookup=MagicMock(return_value=1.0))
    uc = IngestUC(session, fx_client=fx)
    result = uc.ingest(pid, ledger, None, None, None)
    # 외상매출금 2 + 미수금 1 + 받을어음 1 = AR 4건
    assert result.ar_count == 4
    # 외상매입금 1 + 미지급금 1 = AP 2건
    assert result.ap_count == 2

    ar = AccountRepo(session).list_by_project_kind(pid, Kind.AR)
    ar_ids = {a.party_id for a in ar}
    assert ar_ids == {"P1", "P2", "P3", "P4"}

    ap = AccountRepo(session).list_by_project_kind(pid, Kind.AP)
    ap_ids = {a.party_id for a in ap}
    assert ap_ids == {"P5", "P6"}


def test_ingest_src_sheet_recorded(session, tmp_path):
    pid = ProjectRepo(session).create(
        client="X", period_end=date(2025, 12, 31), base_ccy="KRW",
        materiality=1, tolerable=1)
    ledger = _make_multi_sheet_ledger(tmp_path)
    fx = MagicMock(lookup=MagicMock(return_value=1.0))
    IngestUC(session, fx_client=fx).ingest(pid, ledger, None, None, None)
    ar = AccountRepo(session).list_by_project_kind(pid, Kind.AR)
    sheets = {a.party_id: a.src_sheet for a in ar}
    assert sheets["P1"] == "외상매출금"
    assert sheets["P3"] == "미수금"
    assert sheets["P4"] == "받을어음"
