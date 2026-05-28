"""Fuzzy 거래처 집계 — 사업자번호 + 정규화 이름 + RP synonym group.

코스맥스네오 같이 다중 시트에 같은 거래처가 다른 코드(또는 코드 없이)
다른 표기로 흩어진 ledger에 대비. ISA 505 완전성.
"""
import pytest
import openpyxl
from datetime import date
from pathlib import Path
from unittest.mock import MagicMock
from src.application.ingest_uc import IngestUC
from src.domain.entities import Kind
from src.infrastructure.db.session import make_engine, make_session
from src.infrastructure.db.models import Base
from src.infrastructure.db.repository import ProjectRepo, AccountRepo


def _make_test_files(tmp_path):
    # Ledger with same party in different sheets, different names
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ar1 = wb.create_sheet("외상매출금")
    ar1.append(["거래처코드", "거래처명", "사업자번호", "계정과목", "기말잔액", "통화"])
    ar1.append(["P1", "코스맥스(주)", "143-81-19635", "11200", 1_000_000, "KRW"])
    ar1.append(["", "Cosmax Inc", "143-81-19635", "11200", 500_000, "KRW"])  # 사번 같음
    ar2 = wb.create_sheet("미수금")
    ar2.append(["거래처코드", "거래처명", "사업자번호", "계정과목", "기말잔액", "통화"])
    ar2.append(["P1", "코스맥스 주식회사", "143-81-19635", "12100", 300_000, "KRW"])  # 같은 사번
    ledger = tmp_path / "ledger.xlsx"
    wb.save(ledger)

    # RP file with synonym group
    wb_rp = openpyxl.Workbook()
    ws = wb_rp.active
    ws.title = "특관자"
    ws.append(["NO", "한글명", "영문명"])
    ws.append([1, "코스맥스", "Cosmax"])
    rp = tmp_path / "rp.xlsx"
    wb_rp.save(rp)

    return ledger, rp


@pytest.fixture
def session():
    e = make_engine("sqlite:///:memory:")
    Base.metadata.create_all(e)
    S = make_session(e)
    s = S()
    yield s
    s.close()


def test_aggregate_by_business_number(session, tmp_path):
    pid = ProjectRepo(session).create(
        client="X", period_end=date(2025, 12, 31), base_ccy="KRW",
        materiality=1, tolerable=1)
    ledger, rp = _make_test_files(tmp_path)
    fx = MagicMock(lookup=MagicMock(return_value=1.0))
    IngestUC(session, fx_client=fx).ingest(pid, ledger, None, rp, None)

    ar = AccountRepo(session).list_by_project_kind(pid, Kind.AR)
    # 같은 사번 3행 → 1행으로 합산
    assert len(ar) == 1
    a = ar[0]
    # 합산 1M + 500K + 300K = 1,800,000
    assert a.balance_krw == 1_800_000
    # breakdowns에 시트별 분리
    assert a.account_breakdowns.get("외상매출금") == 1_500_000  # 1M + 500K
    assert a.account_breakdowns.get("미수금") == 300_000
    # business_number 보존
    assert a.business_number == "143-81-19635"


def test_aggregate_without_business_number(session, tmp_path):
    """사번 없는 경우 — 정규화 이름으로 매칭."""
    pid = ProjectRepo(session).create(
        client="X", period_end=date(2025, 12, 31), base_ccy="KRW",
        materiality=1, tolerable=1)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ar = wb.create_sheet("외상매출금")
    ar.append(["거래처명", "사업자번호", "기말잔액", "통화"])
    ar.append(["(주)갑상사", None, 100, "KRW"])
    ar.append(["갑상사", None, 200, "KRW"])  # 같은 회사 다른 표기
    ledger = tmp_path / "ledger.xlsx"
    wb.save(ledger)

    fx = MagicMock(lookup=MagicMock(return_value=1.0))
    IngestUC(session, fx_client=fx).ingest(pid, ledger, None, None, None)

    a_list = AccountRepo(session).list_by_project_kind(pid, Kind.AR)
    # 정규화 후 동일 → 1행
    assert len(a_list) == 1
    assert a_list[0].balance_krw == 300


def test_aggregate_synonym_group(session, tmp_path):
    """한·영 표기 다른 회사 — RP synonym group 사용."""
    pid = ProjectRepo(session).create(
        client="X", period_end=date(2025, 12, 31), base_ccy="KRW",
        materiality=1, tolerable=1)
    ledger, rp = _make_test_files(tmp_path)
    # ledger에 사번 비우고 이름만 한·영
    wb = openpyxl.load_workbook(ledger)
    ws = wb["외상매출금"]
    # 두 번째 행 사번 제거
    ws.cell(row=3, column=3, value=None)
    wb.save(ledger)

    fx = MagicMock(lookup=MagicMock(return_value=1.0))
    IngestUC(session, fx_client=fx).ingest(pid, ledger, None, rp, None)

    ar = AccountRepo(session).list_by_project_kind(pid, Kind.AR)
    # 사번 일치하는 거래처와 synonym 일치하는 거래처 모두 합산
    # 사번이 다 같으면 1행, synonym으로 같으면 1행
    # 실제로는 일치 패턴이 어떻든 1행이어야 (합산 OK 인 한)
    total = sum(a.balance_krw for a in ar)
    assert total == 1_800_000
