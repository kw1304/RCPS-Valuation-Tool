import pytest
from datetime import date
from pathlib import Path
from unittest.mock import MagicMock
import openpyxl

from src.application.ingest_uc import IngestUC, IngestResult
from src.infrastructure.db.session import make_engine, make_session
from src.infrastructure.db.models import Base
from src.infrastructure.db.repository import ProjectRepo, AccountRepo
from src.domain.entities import Kind


@pytest.fixture
def session():
    engine = make_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    Session = make_session(engine)
    s = Session()
    yield s
    s.close()


@pytest.fixture
def project_id(session):
    return ProjectRepo(session).create(
        client="X", period_end=date(2025, 12, 31), base_ccy="KRW",
        materiality=1_000_000, tolerable=500_000,
    )


def _make_ledger(tmp_path) -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ar = wb.create_sheet("매출채권")
    ar.append(["거래처코드", "거래처명", "계정과목", "기말잔액", "통화"])
    ar.append(["P1", "갑", "11200", 1_000_000, "KRW"])
    ar.append(["P2", "을", "11200", 100, "USD"])
    ap = wb.create_sheet("매입채무")
    ap.append(["거래처코드", "거래처명", "계정과목", "기말잔액", "통화"])
    ap.append(["P3", "병", "21100", 500_000, "KRW"])
    p = tmp_path / "ledger.xlsx"
    wb.save(p)
    return p


def test_ingest_creates_ar_ap(session, project_id, tmp_path):
    ledger = _make_ledger(tmp_path)
    fx_client = MagicMock()
    fx_client.lookup.return_value = 1300.0
    uc = IngestUC(session, fx_client=fx_client)

    result = uc.ingest(project_id=project_id, ledger_path=ledger,
                       fs_path=None, rp_path=None, allowance_path=None)

    assert result.ar_count == 2
    assert result.ap_count == 1

    acc_repo = AccountRepo(session)
    ar = acc_repo.list_by_project_kind(project_id, Kind.AR)
    ap = acc_repo.list_by_project_kind(project_id, Kind.AP)
    assert len(ar) == 2
    assert len(ap) == 1
    usd = next(a for a in ar if a.ccy == "USD")
    assert usd.balance_krw == 100 * 1300.0


def test_ingest_with_rp_flags(session, project_id, tmp_path):
    ledger = _make_ledger(tmp_path)
    wb_rp = openpyxl.Workbook()
    ws = wb_rp.active
    ws.title = "특관자"
    ws.append(["거래처명"])
    ws.append(["갑"])
    rp_path = tmp_path / "rp.xlsx"
    wb_rp.save(rp_path)

    fx = MagicMock(lookup=MagicMock(return_value=1300.0))
    uc = IngestUC(session, fx_client=fx)
    uc.ingest(project_id=project_id, ledger_path=ledger,
              fs_path=None, rp_path=rp_path, allowance_path=None)

    ar = AccountRepo(session).list_by_project_kind(project_id, Kind.AR)
    gap = next(a for a in ar if a.name == "갑")
    assert gap.is_related_party is True


def test_ingest_with_allowance(session, project_id, tmp_path):
    ledger = _make_ledger(tmp_path)
    wb_allow = openpyxl.Workbook()
    ws = wb_allow.active
    ws.title = "충당금명세"
    ws.append(["거래처코드", "잔액", "충당금", "부실여부"])
    ws.append(["P1", 1_000_000, 500_000, "N"])
    ws.append(["P2", 130_000, 130_000, "Y"])
    allow_path = tmp_path / "allow.xlsx"
    wb_allow.save(allow_path)

    fx = MagicMock(lookup=MagicMock(return_value=1300.0))
    uc = IngestUC(session, fx_client=fx)
    uc.ingest(project_id=project_id, ledger_path=ledger,
              fs_path=None, rp_path=None, allowance_path=allow_path)

    ar = AccountRepo(session).list_by_project_kind(project_id, Kind.AR)
    p1 = next(a for a in ar if a.party_id == "P1")
    p2 = next(a for a in ar if a.party_id == "P2")
    assert p1.allowance_amt == 500_000
    assert p1.is_bad_debt is False
    assert p2.is_bad_debt is True


def test_ingest_replaces_existing(session, project_id, tmp_path):
    ledger = _make_ledger(tmp_path)
    fx = MagicMock(lookup=MagicMock(return_value=1300.0))
    uc = IngestUC(session, fx_client=fx)
    uc.ingest(project_id, ledger, None, None, None)
    uc.ingest(project_id, ledger, None, None, None)
    ar = AccountRepo(session).list_by_project_kind(project_id, Kind.AR)
    assert len(ar) == 2
