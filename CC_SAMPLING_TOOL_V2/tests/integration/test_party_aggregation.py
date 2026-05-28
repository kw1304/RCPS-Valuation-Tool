"""party_id별 거래처 집계 검증.

다중 시트(외상매출금 + 미수금)에 같은 거래처가 있을 때
ingest 시 1행으로 통합되어야. 잔액·차변·대변 모두 합산.
"""
import pytest
import openpyxl
from datetime import date
from pathlib import Path
from unittest.mock import MagicMock
from src.application.ingest_uc import IngestUC
from src.domain.entities import Kind
from src.infrastructure.db.session import make_engine, make_session
from src.infrastructure.db.models import Base
from src.infrastructure.db.repository import ProjectRepo, AccountRepo


def _make_dup_ledger(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # 외상매출금
    ar1 = wb.create_sheet("외상매출금")
    ar1.append(["거래처코드", "거래처명", "계정과목", "기말잔액", "통화", "차변", "대변"])
    ar1.append(["P1", "갑상사", "11200", 1_000_000, "KRW", 5_000_000, 4_000_000])
    ar1.append(["P2", "을상사", "11200", 500_000, "KRW", 2_000_000, 1_500_000])
    # 미수금 — P1 중복
    ar2 = wb.create_sheet("미수금")
    ar2.append(["거래처코드", "거래처명", "계정과목", "기말잔액", "통화", "차변", "대변"])
    ar2.append(["P1", "갑상사", "12100", 300_000, "KRW", 600_000, 300_000])
    p = tmp_path / "dup.xlsx"
    wb.save(p)
    return p


@pytest.fixture
def session():
    e = make_engine("sqlite:///:memory:")
    Base.metadata.create_all(e)
    S = make_session(e)
    s = S()
    yield s
    s.close()


def test_party_aggregated_balance(session, tmp_path):
    pid = ProjectRepo(session).create(
        client="X", period_end=date(2025, 12, 31), base_ccy="KRW",
        materiality=1, tolerable=1)
    ledger = _make_dup_ledger(tmp_path)
    fx = MagicMock(lookup=MagicMock(return_value=1.0))
    IngestUC(session, fx_client=fx).ingest(pid, ledger, None, None, None)

    ar = AccountRepo(session).list_by_project_kind(pid, Kind.AR)
    # 집계 후 2 unique 거래처
    assert len(ar) == 2
    p1 = next(a for a in ar if a.party_id == "P1")
    # P1: 1,000,000 + 300,000 = 1,300,000
    assert p1.balance_krw == 1_300_000
    # P1 차변: 5M + 600K = 5,600,000
    assert p1.debit_amt == 5_600_000
    # P1 대변: 4M + 300K = 4,300,000
    assert p1.credit_amt == 4_300_000
    # src_sheet에 두 시트 모두 표시
    assert "외상매출금" in p1.src_sheet
    assert "미수금" in p1.src_sheet


def test_single_party_no_aggregation(session, tmp_path):
    """단일 시트 단일 거래처는 그대로 통과."""
    pid = ProjectRepo(session).create(
        client="X", period_end=date(2025, 12, 31), base_ccy="KRW",
        materiality=1, tolerable=1)
    ledger = _make_dup_ledger(tmp_path)
    fx = MagicMock(lookup=MagicMock(return_value=1.0))
    IngestUC(session, fx_client=fx).ingest(pid, ledger, None, None, None)

    ar = AccountRepo(session).list_by_project_kind(pid, Kind.AR)
    p2 = next(a for a in ar if a.party_id == "P2")
    assert p2.balance_krw == 500_000
    assert p2.debit_amt == 2_000_000
