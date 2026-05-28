import pytest
import io
import openpyxl
from src.infrastructure.excel_writer.workpaper import build_workpaper


def _full_state():
    return {
        "project": {
            "client": "ACME", "period_end": "2025-12-31",
            "base_ccy": "KRW", "materiality": 500_000_000,
            "tolerable": 250_000_000,
        },
        "populations": {
            "AR": {"count": 120, "total_krw": 250_000_000},
            "AP": {"count": 80, "total_krw": 120_000_000},
        },
        "samples": {
            "AR": {"count": 2, "total_krw": 18_000_000, "items": [
                {"party_id": "AR000", "name": "고객사000", "gl_account": "11200",
                 "balance_krw": 10_000_000, "ccy": "KRW",
                 "selection_reason": "FORCED_RP",
                 "is_related_party": True, "is_bad_debt": False},
                {"party_id": "AR050", "name": "고객사050", "gl_account": "11200",
                 "balance_krw": 8_000_000, "ccy": "KRW",
                 "selection_reason": "REP",
                 "is_related_party": False, "is_bad_debt": False},
            ]},
            "AP": {"count": 1, "total_krw": 5_000_000, "items": [
                {"party_id": "AP000", "name": "공급사000", "gl_account": "21100",
                 "balance_krw": 5_000_000, "ccy": "KRW",
                 "selection_reason": "FORCED_KEY",
                 "is_related_party": False, "is_bad_debt": False},
            ]},
        },
        "confirmations": {
            "AR": [{"party_id": "AR000", "name": "고객사000",
                    "expected": 10_000_000, "confirmed": 10_000_000,
                    "diff": 0, "diff_reason": None, "verdict": "MATCH",
                    "status": "RECEIVED", "pdf_path": None}],
            "AP": [{"party_id": "AP000", "name": "공급사000",
                    "expected": 5_000_000, "confirmed": 4_900_000,
                    "diff": -100_000, "diff_reason": None,
                    "verdict": "DISCREPANCY", "status": "RECEIVED",
                    "pdf_path": None}],
        },
        "alternatives": {
            "AR": [{"party_id": "AR050", "name": "고객사050",
                    "procedure_type": "후속회수",
                    "evidence_sum": 8_000_000, "note": "회수증빙"}],
            "AP": [],
        },
        "projection": {"AR": None, "AP": None},
    }


def test_combined_7_sheets():
    blob = build_workpaper("cc_combined", _full_state())
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    assert "샘플링 요약" in wb.sheetnames
    assert "C100 조회서 control sheet" in wb.sheetnames
    assert "C100-1 표본규모 결정" in wb.sheetnames
    assert "C100-2 Key item 추출" in wb.sheetnames
    assert "C100-3 표본 추출 MUS" in wb.sheetnames
    assert "C100-4 조회서 회수 관리" in wb.sheetnames
    assert "대체적 절차" in wb.sheetnames


def test_recovery_management_shows_status():
    """회수 관리 시트가 회신·미회신 모두 표시."""
    blob = build_workpaper("cc_combined", _full_state())
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    ws = wb["C100-4 조회서 회수 관리"]
    flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    # 요약 라인
    assert any("회수 현황 요약" in v for v in flat)
    assert any("회신율" in v for v in flat)
    # MATCH 회신 (state에 AR000 MATCH)
    assert any("MATCH" in v for v in flat)
    # DISCREPANCY (state에 AP000 DISCREPANCY)
    assert any("DISCREPANCY" in v for v in flat)
    # 미회신 (state samples에 있지만 confirmations에 없는 거래처: AR050)
    assert any("AR050" in v for v in flat)
    assert any("미회신" in v for v in flat) or any("발송됨" in v for v in flat)


def test_summary_contains_both_kinds():
    blob = build_workpaper("cc_combined", _full_state())
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    ws = wb["샘플링 요약"]
    flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert any("채권" in v for v in flat)
    assert any("채무" in v for v in flat)
    assert any("합계" in v for v in flat)


def test_control_sheet_has_parties():
    blob = build_workpaper("cc_combined", _full_state())
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    ws = wb["C100 조회서 control sheet"]
    flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert any("AR000" in v for v in flat)
    assert any("AP000" in v for v in flat)


def test_key_item_lists_forced():
    blob = build_workpaper("cc_combined", _full_state())
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    ws = wb["C100-2 Key item 추출"]
    flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    # FORCED_RP, FORCED_KEY는 있어야
    assert any("FORCED_RP" in v for v in flat)
    assert any("FORCED_KEY" in v for v in flat)
    # REP는 KEY 시트에 없어야 (REP는 C100-3)
    assert not any(v == "REP" for v in flat)


def test_mus_sample_lists_rep_only():
    blob = build_workpaper("cc_combined", _full_state())
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    ws = wb["C100-3 표본 추출 MUS"]
    flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    # REP 거래처 (AR050) 있어야
    assert any("AR050" in v for v in flat)
    # FORCED 거래처는 KEY 시트에 있고 여기엔 없어야
    assert not any("AR000" in v for v in flat)


def test_alternative_kind_column():
    blob = build_workpaper("cc_combined", _full_state())
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    ws = wb["대체적 절차"]
    flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    # 후속회수 (state alternatives.AR에 있음)
    assert any("후속회수" in v for v in flat)
