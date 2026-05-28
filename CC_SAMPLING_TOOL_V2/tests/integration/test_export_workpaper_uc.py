import pytest
import io
from datetime import date
import openpyxl
from src.application.export_workpaper_uc import ExportWorkpaperUC
from src.domain.entities import Account, Kind, SelectionReason
from src.infrastructure.db.session import make_engine, make_session
from src.infrastructure.db.models import Base
from src.infrastructure.db.repository import ProjectRepo, AccountRepo, SampleRepo


@pytest.fixture
def session():
    e = make_engine("sqlite:///:memory:")
    Base.metadata.create_all(e)
    S = make_session(e)
    s = S()
    yield s
    s.close()


def test_export_c100(session):
    pid = ProjectRepo(session).create(
        client="ACME", period_end=date(2025, 12, 31), base_ccy="KRW",
        materiality=500_000_000, tolerable=250_000_000)
    acc = Account(party_id="AR1", name="고객", gl_account="11200",
                  balance_orig=1_000_000, ccy="KRW", fx_rate=1.0,
                  balance_krw=1_000_000)
    AccountRepo(session).bulk_insert(pid, Kind.AR, [acc])
    accs = AccountRepo(session).list_by_project_kind(pid, Kind.AR)
    SampleRepo(session).persist(pid, Kind.AR,
                                 [(accs[0], SelectionReason.FORCED_RP)])

    uc = ExportWorkpaperUC(session)
    blob = uc.build(pid, "c100")
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    assert "C100_summary" in wb.sheetnames
    assert "C101_sendlist" in wb.sheetnames


def test_export_aa100(session):
    pid = ProjectRepo(session).create(
        client="X", period_end=date(2025, 12, 31), base_ccy="KRW",
        materiality=1, tolerable=1)
    uc = ExportWorkpaperUC(session)
    blob = uc.build(pid, "aa100")
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    assert "AA100_summary" in wb.sheetnames


def test_export_invalid_template_raises(session):
    pid = ProjectRepo(session).create(
        client="X", period_end=date(2025, 12, 31), base_ccy="KRW",
        materiality=1, tolerable=1)
    uc = ExportWorkpaperUC(session)
    with pytest.raises(FileNotFoundError):
        uc.build(pid, "nonexistent_template")
