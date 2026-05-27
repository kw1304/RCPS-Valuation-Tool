from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def generate_workpaper(params, initial_result, sensitivity_result, output_path,
                       tf_tree=None, gs_tree=None, eval_result=None, bdt_cross=None):
    wb = Workbook()

    ws0 = wb.active
    ws0.title = "0.평가의견"
    _write_executive_summary(ws0, params, initial_result, eval_result, sensitivity_result)

    ws1 = wb.create_sheet("1.발행조건")
    _write_cover(ws1, params, initial_result)

    ws2 = wb.create_sheet("2.평가결과")
    _write_valuation(ws2, initial_result)

    # 모형 비교 (TF/GS/MC) — eval_result 있을 때
    if eval_result:
        ws_cmp = wb.create_sheet("3.모형비교")
        _write_model_comparison(ws_cmp, eval_result, params)

    # BDT 교차검증 — bdt_cross 있을 때
    if bdt_cross:
        ws_bdt = wb.create_sheet("4.BDT교차검증")
        _write_bdt_cross(ws_bdt, bdt_cross)

    # 이항트리 — TF/GS 각각 (collect_tree 결과 있을 때만)
    if tf_tree and not tf_tree.get("error"):
        ws_tf = wb.create_sheet("5.TF이항트리")
        _write_tree(ws_tf, "TF (Tsiveriotis-Fernandes)", tf_tree, params)
    if gs_tree and not gs_tree.get("error"):
        ws_gs = wb.create_sheet("6.GS이항트리")
        _write_tree(ws_gs, "GS (Goldman Sachs)", gs_tree, params)

    ws_sens = wb.create_sheet("7.민감도분석")
    _write_sensitivity(ws_sens, sensitivity_result)

    wb.save(output_path)
    return output_path


def _write_model_comparison(ws, eval_result, params):
    """TF·GS·MC 모형별 공정가치 비교."""
    ws.row_dimensions[1].height = 32
    t = ws.cell(row=1, column=1, value="모형별 공정가치 비교 — TF / GS / MC")
    t.font = Font(name="맑은 고딕", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill(fill_type="solid", fgColor="1A365D")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:E1")

    for col, w in zip("ABCDE", [28, 20, 20, 20, 18]):
        ws.column_dimensions[col].width = w

    for i, h in enumerate(["구성요소", "TF (주채택)", "GS (Cox-Ross)", "MC (Monte Carlo)", "비고"], 1):
        _hdr(ws, 3, i, h)

    tf = eval_result.get("tf") or {}
    gs = eval_result.get("gs") or {}
    mc = eval_result.get("mc") or {}
    def _v(d, k):
        v = d.get(k) if isinstance(d, dict) else None
        return v if v not in (None, "", 0) or k.startswith("fair") else None

    # 한국 평가실무 표준: 채권·풋은 모형 무관 공통값(TF값) — GS·MC 모두 동일 사용
    _bond = tf.get("bond_value")
    _put = tf.get("put_option_value")
    _putb = tf.get("put_bond_value")
    _mc_fv = mc.get("fair_value")
    _mc_conv = (_mc_fv - _putb) if (_mc_fv is not None and _putb is not None) else None

    rows = [
        ("순채권가치 (Bond)", _bond, _bond, _bond,
         "모형 무관 공통값 (이자율·CF 기반)"),
        ("풋옵션가치 (Put)", _put, _put, _put,
         "모형 무관 공통값 (보장수익률 IRR)"),
        ("풋채권가치 (Putable Bond)", _putb, _putb, _putb,
         "순채권 + 풋"),
        ("전환권가치 (Conversion)", tf.get("conversion_value"), gs.get("conversion_value"), _mc_conv,
         "공정가치 − 풋채권"),
        ("공정가치 (FV)", tf.get("fair_value"), gs.get("fair_value"), _mc_fv,
         "TF=주채택 / GS=비교 / MC=참고"),
    ]
    for ri, (label, vtf, vgs, vmc, note) in enumerate(rows, 4):
        bold = ri == 8  # 공정가치 행
        bg = "EBF8FF" if bold else ("F7FAFC" if ri % 2 == 0 else "FFFFFF")
        _cell(ws, ri, 1, label, bold=bold, bg=bg)
        _cell(ws, ri, 2, vtf if vtf is not None else "-", bold=bold, bg=bg, num_fmt="#,##0", align="right")
        _cell(ws, ri, 3, vgs if vgs is not None else "-", bg=bg, num_fmt="#,##0", align="right")
        _cell(ws, ri, 4, vmc if vmc is not None else "-", bg=bg, num_fmt="#,##0", align="right")
        _cell(ws, ri, 5, note, bg=bg, color="718096")

    # TF 본래 2-component 분해 — Tsiveriotis-Fernandes(1998) 학술 분해
    tfE = tf.get("equity_component") if tf else None
    tfB = tf.get("bond_component") if tf else None
    if tfE is not None and tfB is not None:
        r2c = 10
        ws.cell(row=r2c, column=1, value="■ TF 본래 2-component 분해 (Tsiveriotis-Fernandes 1998)").font = \
            Font(name="맑은 고딕", bold=True, size=10, color="0369A1")
        ws.merge_cells(start_row=r2c, start_column=1, end_row=r2c, end_column=5)
        for i, (k, v, note) in enumerate([
            ("VE (지분 성분, E[0])", tfE, "전환 가능성 반영"),
            ("VD (채권 성분, B[0])", tfB, "만기상환·풋 보장"),
            ("합계", tfE + tfB, "= TF 공정가치"),
        ], r2c + 1):
            bg = "F0F9FF" if i == r2c + 3 else "FFFFFF"
            _cell(ws, i, 1, k, bg=bg, bold=(i == r2c + 3))
            _cell(ws, i, 2, v, bg=bg, num_fmt="#,##0", align="right", bold=(i == r2c + 3))
            _cell(ws, i, 5, note, bg=bg, color="718096")
        ws.cell(row=r2c + 4, column=1,
                value="※ Tsiveriotis-Fernandes(1998) 본래 분해. 위 3-way(채권/풋옵션/전환권)은 K-IFRS 1109.B4.3.5 발행자 무조건 의무 관점.")\
            .font = Font(name="맑은 고딕", size=9, color="718096", italic=True)
        ws.merge_cells(start_row=r2c + 4, start_column=1, end_row=r2c + 4, end_column=5)

    # 모형 메타정보
    r0 = 16  # 2-comp 블록 다음 (이전 10에서 16으로 밀림)
    ws.cell(row=r0, column=1, value="■ 모형 가정·메타").font = Font(name="맑은 고딕", bold=True, size=11)
    metas = [
        ("이항 트리 단계", str(eval_result.get("steps", "-"))),
        ("u (상승배수)", f"{eval_result.get('u', '-')}"),
        ("d (하락배수)", f"{eval_result.get('d', '-')}"),
        ("위험중립확률 p", f"{eval_result.get('risk_neutral_prob', '-')}"),
        ("선도이자율 곡선 적용", "예" if eval_result.get("term_structure_applied") else "아니오"),
        ("MC 경로수", str(mc.get("n_paths", "-"))),
    ]
    for ri, (k, v) in enumerate(metas, r0 + 1):
        bg = "F7FAFC" if ri % 2 == 0 else "FFFFFF"
        _cell(ws, ri, 1, k, bg=bg)
        _cell(ws, ri, 2, v, bg=bg, align="right")

    # ── 자동 해석 (감사 대응성) — 일반인이 읽고 이해 가능한 한 줄 해설
    interp_row = r0 + len(metas) + 2
    ws.cell(row=interp_row, column=1, value="■ 모형 간 차이 해석").font = Font(name="맑은 고딕", bold=True, size=11)
    ws.merge_cells(start_row=interp_row, start_column=1, end_row=interp_row, end_column=5)
    notes = _interpret_model_diff(tf, gs, mc)
    for i, note in enumerate(notes, interp_row + 1):
        bg = "FFFBEB"
        c = ws.cell(row=i, column=1, value=note)
        c.font = Font(name="맑은 고딕", size=10, color="78350F")
        c.fill = PatternFill(fill_type="solid", fgColor=bg)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 28
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)


def _interpret_model_diff(tf: dict, gs: dict, mc: dict) -> list:
    """TF/GS/MC 결과 차이를 일반인 친화적 한 줄 해설로 자동 생성.

    감사인이 "왜 TF·GS·MC 결과가 다른가?"에 즉답 가능하도록 자동 주석.
    K-IFRS 1113.93(g) 수준 3 측정 불확실성 공시 보조.
    """
    notes = []
    tf_fv = tf.get("fair_value") if tf and not tf.get("error") else None
    gs_fv = gs.get("fair_value") if gs and not gs.get("error") else None
    mc_fv = mc.get("fair_value") if mc and not mc.get("error") else None
    mc_se = mc.get("std_error") if mc else None
    mc_ci_low = mc.get("ci_lower") if mc else None
    mc_ci_high = mc.get("ci_upper") if mc else None

    # TF vs GS 차이 — 임의 구간 임계값 제거 (2026-05-27)
    if tf_fv and gs_fv:
        diff = gs_fv - tf_fv
        diff_pct = diff / tf_fv * 100
        notes.append(
            f"• TF·GS 차이 {diff:+,.0f}원 ({diff_pct:+.2f}%) — "
            f"TF는 채권·주식 분리 할인, GS는 전환확률 가중 블렌딩 할인. "
            f"모형 구조 차이로 결과가 갈림."
        )

    # MC 신뢰구간 표시 (판단 임계값 없이 정보만)
    if mc_fv and tf_fv and mc_ci_low is not None and mc_ci_high is not None:
        in_ci = mc_ci_low <= tf_fv <= mc_ci_high
        notes.append(
            f"• MC 95% 신뢰구간: {mc_ci_low:,.0f}~{mc_ci_high:,.0f}원. "
            f"TF 결과 {tf_fv:,.0f}원 — {'구간 내' if in_ci else '구간 밖'}."
        )

    # SE 표시 (정량 임계값 없이 통계적 의미만)
    if mc_se and mc_fv:
        se_pct = mc_se / mc_fv * 100
        notes.append(
            f"• MC 표본오차(SE) {mc_se:,.0f}원 (공정가치의 {se_pct:.2f}%) — "
            f"경로 수 증가 시 SE는 √n에 반비례 감소."
        )

    if not notes:
        notes.append("• 모형 비교 데이터 부족 — 평가 후 자동 생성됩니다.")
    return notes


def _write_bdt_cross(ws, bdt):
    """BDT 교차검증 — TF vs BDT 풋채권 비교."""
    ws.row_dimensions[1].height = 32
    t = ws.cell(row=1, column=1, value="BDT 교차검증 — TF vs BDT (독립 금리트리)")
    t.font = Font(name="맑은 고딕", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill(fill_type="solid", fgColor="1A365D")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:E1")

    for col, w in zip("ABCDE", [28, 22, 22, 14, 22]):
        ws.column_dimensions[col].width = w

    for i, h in enumerate(["구성요소", "TF (주가트리)", "BDT (금리트리)", "Δ %", "해석"], 1):
        _hdr(ws, 3, i, h)

    def _diff(a, b):
        if a is None or b is None or b == 0:
            return "-"
        return f"{((a - b) / b) * 100:+.2f}%"

    rows = [
        ("① 순채권가치 (Bond)", bdt.get("tf_bond"), bdt.get("bdt_bond"),
         _diff(bdt.get("bdt_bond"), bdt.get("tf_bond")),
         "결정론적 금리 vs 금리트리 — 보통 ±1%"),
        ("② 풋채권가치 (Putable Bond)", bdt.get("tf_put_bond"), bdt.get("bdt_put_bond"),
         _diff(bdt.get("bdt_put_bond"), bdt.get("tf_put_bond")),
         "독립 검증값 — 보통 잘 일치"),
        ("③ 풋옵션가치 (Put)", bdt.get("tf_put"), bdt.get("bdt_put"),
         _diff(bdt.get("bdt_put"), bdt.get("tf_put")),
         "모형 가정 차이로 다소 벌어질 수 있음"),
    ]
    for ri, (label, vtf, vbdt, dpct, note) in enumerate(rows, 4):
        bg = "F7FAFC" if ri % 2 == 0 else "FFFFFF"
        _cell(ws, ri, 1, label, bg=bg)
        _cell(ws, ri, 2, vtf if vtf is not None else "-", bg=bg, num_fmt="#,##0", align="right")
        _cell(ws, ri, 3, vbdt if vbdt is not None else "-", bg=bg, num_fmt="#,##0", align="right")
        _cell(ws, ri, 4, dpct, bg=bg, align="center")
        _cell(ws, ri, 5, note, bg=bg, color="718096")

    r0 = 9
    ws.cell(row=r0, column=1, value="■ BDT 가정").font = Font(name="맑은 고딕", bold=True, size=11)
    metas = [
        ("단기금리 변동성 σ", f"{(bdt.get('rate_vol', 0) or 0) * 100:.2f}%"),
        ("트리 스텝 수", str(bdt.get("n_steps", "-"))),
        ("dt (년)", f"{bdt.get('dt', 0):.4f}"),
        ("풋 행사 시작 step", str(bdt.get("put_step", "-"))),
        ("캘리브레이션 기준", "Rd(신용조정) 곡선 — 시장 zero에 fit"),
    ]
    for ri, (k, v) in enumerate(metas, r0 + 1):
        bg = "F7FAFC" if ri % 2 == 0 else "FFFFFF"
        _cell(ws, ri, 1, k, bg=bg)
        _cell(ws, ri, 2, v, bg=bg, align="right")


def _write_tree(ws, title, tree, params=None):
    """이항트리 — 웹 화면과 동일 레이아웃.

    행 = j (0..N, down moves), 열 = step i (0..N).
    가치 그리드는 RCPS 주식수로 나눠 한주(1RCPS) 기준 표시.
    """
    n = len(tree.get("stock") or [])
    if n == 0:
        return
    N = n - 1
    steps = tree.get("steps", N)
    u = tree.get("u"); d = tree.get("d"); p = tree.get("p")
    rcps_shares = max(int(getattr(params, "rcps_shares", 0) or 0), 1) if params else 1

    # 타이틀
    ws.row_dimensions[1].height = 32
    t = ws.cell(row=1, column=1, value=f"이항트리 — {title}")
    t.font = Font(name="맑은 고딕", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill(fill_type="solid", fgColor="1A365D")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(n + 1, 14))

    p_str = "(커브모드)" if isinstance(p, list) else str(p)
    info = f"steps={steps}  |  u={u}  |  d={d}  |  p={p_str}  |  RCPS 주식수={rcps_shares:,}주  |  ※ 가치 그리드는 1주(1RCPS) 기준"
    c = ws.cell(row=2, column=1, value=info)
    c.font = Font(name="맑은 고딕", size=9, color="718096")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(n + 1, 14))

    # 그리드 정의 — 웹 화면(renderTreeGrids)과 동일 순서/라벨
    is_tf = ("Tsiveriotis" in title) or ("TF" in title.upper())
    if is_tf:
        grids = [
            ("stock",          "① 주가 트리 (원/주)"),
            ("decision",       "의사결정"),
            ("conv_intrinsic", "전환 내재가치 — 즉시 전환 시 (원/주)"),
            ("bond_intrinsic", "채권 내재가치 — 즉시 풋·만기상환 시 (원/주)"),
            ("rcps_value",     "RCPS 가치 — 의사결정 후 채택 (원/주)"),
            ("hold_value",     "보유가치 — 연속보유 (원/주)"),
            ("equity_comp",    "지분가치 — 실현 (원/주)"),
            ("bond_comp",      "채권가치 — 실현 (원/주)"),
            ("equity_hold",    "지분 보유가치 (원/주)"),
            ("bond_hold",      "채권 보유가치 (원/주)"),
        ]
    else:  # GS
        # GS는 단일 V로 후방귀납 (E/B 분리 없음, cp 가중 블렌딩 할인)
        # → 별도 내재가치 그리드 불필요 (decision·rcps_value로 충분)
        grids = [
            ("stock",          "① 주가 트리 (원/주)"),
            ("decision",       "의사결정"),
            ("conv_prob",      "전환확률"),
            ("disc_factor",    "할인계수"),
            ("rcps_value",     "RCPS 가치 — 의사결정 후 채택 (원/주)"),
            ("hold_value",     "보유가치 — 연속보유 (원/주)"),
        ]

    # per-share 변환 대상 (주가/확률/할인계수 제외)
    VALUE_KEYS = {"conv_intrinsic", "bond_intrinsic", "rcps_value", "hold_value",
                  "equity_comp", "bond_comp", "equity_hold", "bond_hold"}
    # 주가는 모형에서 이미 per-share, 확률·할인계수는 단위 없음
    # decision 은 텍스트

    # ── 웹 화면(renderTreeGrids)의 visIdx와 동일: n+1<=20이면 전부, 아니면 앞 10 + ⋯ + 뒤 10 ──
    MAXV, HEAD, TAIL = 20, 10, 10
    if (N + 1) <= MAXV:
        idx_list = list(range(N + 1))
        elided = False
    else:
        idx_list = list(range(HEAD)) + ["…"] + list(range(N - TAIL + 1, N + 1))
        elided = True

    def _grid_block(start_row, key, label, grid):
        r0 = start_row
        head_label = f"■ {label}"
        if elided:
            head_label += "  (중간 노드 생략 · 끝노드까지 표시)"
        c = ws.cell(row=r0, column=1, value=head_label)
        c.font = Font(name="맑은 고딕", bold=True, size=11, color="1A365D")
        # 헤더 행: i (생략 시 …)
        _hdr(ws, r0 + 1, 1, "j \\ i")
        for col_off, i_step in enumerate(idx_list, start=2):
            _hdr(ws, r0 + 1, col_off, "⋯" if i_step == "…" else str(i_step))
        # 데이터 행: j (생략 시 ⋮ 행 한 줄)
        for row_off, j in enumerate(idx_list, start=2):
            row_idx = r0 + row_off
            if j == "…":
                _hdr(ws, row_idx, 1, "⋮")
                for col_off, i_step in enumerate(idx_list, start=2):
                    _cell(ws, row_idx, col_off, "⋯", bg="F7FAFC", align="center", color="A0AEC0")
                continue
            _hdr(ws, row_idx, 1, str(j))
            for col_off, i_step in enumerate(idx_list, start=2):
                if i_step == "…":
                    _cell(ws, row_idx, col_off, "⋯", bg="F7FAFC", align="center", color="A0AEC0")
                    continue
                if j > i_step:
                    _cell(ws, row_idx, col_off, None, bg="F7FAFC")
                    continue
                v = None
                if i_step < len(grid):
                    rr = grid[i_step]
                    if j < len(rr):
                        v = rr[j]
                if v is None or v == "":
                    _cell(ws, row_idx, col_off, None, bg="F7FAFC")
                    continue
                bg = "F7FAFC" if (i_step + j) % 2 == 0 else "FFFFFF"
                if key == "decision":
                    # 웹 화면(renderTreeGrids)과 동일 6단계 색상
                    bg2 = {
                        "전환":     "DCFCE7",   # 녹: 보유자 전환
                        "만기상환": "FED7AA",   # 주: 만기 보장 상환
                        "풋상환":   "FEF3C7",   # 노: 보유자 풋
                        "콜상환":   "FECACA",   # 적: 발행자 콜
                        "강제전환": "A7F3D0",   # 청록: KO
                        "보유":     "F2F4F6",   # 회: 보유
                        # 호환 (기존 4단계 라벨)
                        "상환":     "FEF3C7",
                        "콜":       "FEEBC8",
                    }.get(v, bg)
                    _cell(ws, row_idx, col_off, v, bg=bg2, align="center")
                elif key in ("conv_prob", "disc_factor"):
                    _cell(ws, row_idx, col_off, float(v), bg=bg, num_fmt="0.0000", align="right")
                else:
                    val = float(v)
                    if key in VALUE_KEYS and rcps_shares > 1:
                        val = val / rcps_shares
                    fmt = "#,##0.00" if key == "stock" else "#,##0"
                    _cell(ws, row_idx, col_off, val, bg=bg, num_fmt=fmt, align="right")
        return r0 + 1 + len(idx_list) + 1  # 다음 블록 시작 row

    # 컬럼 폭 (생략 후 실제 표시 열 수에 맞춰)
    ws.column_dimensions["A"].width = 14
    for col_off in range(2, len(idx_list) + 2):
        ws.column_dimensions[get_column_letter(col_off)].width = 13

    next_row = 4
    for key, label in grids:
        g = tree.get(key)
        if not g:
            continue
        next_row = _grid_block(next_row, key, label, g) + 2


def _cell(ws, row, col, value, bold=False, bg=None, num_fmt=None, align="left", color=None):
    # value=None / "" 인 셀에 스타일만 적용 시 openpyxl이 t="n"(빈 numeric) 또는
    # t="inlineStr"(빈 컨텐츠)로 직렬화 → Excel "복구 필요" 오류 유발.
    # 해결: 명시적으로 단일 공백 " " 으로 inline-string 형태 보장
    if value is None or value == "":
        cell = ws.cell(row=row, column=col, value=" ")
    else:
        cell = ws.cell(row=row, column=col, value=value)
    # 문자열이 '=' '+' '-' '@' 로 시작하면 openpyxl이 수식으로 직렬화 → Excel 손상 경고
    # 텍스트로 강제: quotePrefix 활성화 + data_type 명시
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        cell.data_type = "s"
    font_kwargs = dict(name="맑은 고딕", bold=bold, size=10)
    if color:
        font_kwargs["color"] = color
    cell.font = Font(**font_kwargs)
    if bg:
        cell.fill = PatternFill(fill_type="solid", fgColor=bg)
    if num_fmt:
        cell.number_format = num_fmt
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    side = Side(style="thin", color="DDDDDD")
    cell.border = Border(left=side, right=side, top=side, bottom=side)
    return cell


def _hdr(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill(fill_type="solid", fgColor="1A365D")
    c.alignment = Alignment(horizontal="center", vertical="center")
    side = Side(style="thin", color="FFFFFF")
    c.border = Border(left=side, right=side, top=side, bottom=side)
    return c


def _write_executive_summary(ws, params, initial, eval_result, sensitivity):
    """평가의견 표지 — 평가법인 보고서 1페이지 표준.

    구성: 평가목적·기준일·대상 → 결과요약 → 평가절차 → 핵심가정 → 분해 → 이슈사항.
    K-IFRS 1113.91 재현성·1113.93(d)(g) 수준 3 공시 핵심 항목.
    """
    for col, w in zip("ABCDEFG", [4, 22, 18, 18, 18, 18, 4]):
        ws.column_dimensions[col].width = w

    # 제목
    ws.row_dimensions[1].height = 38
    t = ws.cell(row=1, column=2, value="RCPS 공정가치 평가의견")
    t.font = Font(name="맑은 고딕", bold=True, size=16, color="FFFFFF")
    t.fill = PatternFill(fill_type="solid", fgColor="1A365D")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("B1:F1")

    row = 3
    # 1. 평가 개요
    ws.cell(row=row, column=2, value="■ 평가 개요").font = Font(name="맑은 고딕", bold=True, size=12, color="1A365D")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1
    overview = [
        ("평가 대상", "상환전환우선주 (RCPS)"),
        ("평가 기준일", str(params.valuation_date)),
        ("발행일 / 만기일", f"{params.issue_date} ~ {params.maturity_date}"),
        ("잔존 만기", f"{params.T:.3f}년"),
        ("발행금액", f"{int(params.face_value):,}원"),
        ("회계기준", "K-IFRS 1109 · 1113 (당기손익-공정가치 분류 전제)"),
        ("평가목적", "재무제표 작성 목적 공정가치 측정"),
    ]
    for k, v in overview:
        bg = "F7FAFC" if row % 2 == 0 else "FFFFFF"
        _cell(ws, row, 2, k, bold=True, bg=bg)
        _cell(ws, row, 3, v, bg=bg)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        row += 1
    row += 1

    # 2. 평가 결과 (4모형 요약)
    ws.cell(row=row, column=2, value="■ 평가 결과").font = Font(name="맑은 고딕", bold=True, size=12, color="1A365D")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1
    # 모형별 fv
    _hdr(ws, row, 2, "모형")
    _hdr(ws, row, 3, "공정가치")
    _hdr(ws, row, 4, "1주당 가치")
    _hdr(ws, row, 5, "비고")
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    row += 1
    tf = (eval_result or {}).get("tf") or {}
    gs = (eval_result or {}).get("gs") or {}
    mc = (eval_result or {}).get("mc") or {}
    n_rcps = params.rcps_shares or 1
    model_rows = [
        ("TF (Tsiveriotis-Fernandes)", tf.get("fair_value"), "메인 — 채권·주식 분리 할인"),
        ("GS (Goldman Sachs)", gs.get("fair_value"), "비교 — 블렌딩 할인"),
        ("MC (Monte Carlo)", mc.get("fair_value"),
         f"참고 — 경로 {mc.get('n_paths','-')}회"
         + (f", 95% CI {int(mc.get('ci_lower',0)):,}~{int(mc.get('ci_upper',0)):,}원" if mc.get('ci_lower') else "")),
    ]
    for label, fv, note in model_rows:
        bg = "EBF8FF" if "TF" in label else ("F7FAFC" if row % 2 == 0 else "FFFFFF")
        bold = "TF" in label
        _cell(ws, row, 2, label, bold=bold, bg=bg)
        _cell(ws, row, 3, fv if fv else "-", bold=bold, bg=bg, num_fmt="#,##0", align="right")
        per_share = (fv / n_rcps) if fv else None
        _cell(ws, row, 4, per_share if per_share else "-", bold=bold, bg=bg, num_fmt="#,##0", align="right")
        _cell(ws, row, 5, note, bg=bg, color="718096")
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        row += 1
    row += 1

    # 평가의견 한 줄
    main_fv = tf.get("fair_value") or initial.get("fair_value")
    if main_fv:
        opinion = (
            f"본 평가기준일 현재 평가 대상 RCPS의 공정가치는 "
            f"₩{int(main_fv):,} (1주당 ₩{int(main_fv / n_rcps):,}) 으로 산정함."
        )
        c = ws.cell(row=row, column=2, value=opinion)
        c.font = Font(name="맑은 고딕", bold=True, size=11, color="78350F")
        c.fill = PatternFill(fill_type="solid", fgColor="FEF3C7")
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 32
        row += 2

    # 3. 평가 절차
    ws.cell(row=row, column=2, value="■ 평가 절차").font = Font(name="맑은 고딕", bold=True, size=12, color="1A365D")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1
    steps_list = [
        "① 발행조건 검증 — 약정서 기반 발행조건·전환·풋·콜 조항 입력값화",
        "② 시장 데이터 수집 — 무위험이자율(국고채), 신용스프레드(부트스트랩 곡선), 변동성(유사기업 바스켓), 주가",
        "③ 평가 모형 적용 — TF(메인) / GS(비교) / MC(참고) 3개 모형 동시 적용 (동일 단계·동일 곡선)",
        "④ 분해 — 흡수형 분해(채권/풋옵션/전환권) 및 위험 분리 분해(지분/채권) 양쪽 산출",
        "⑤ 민감도 분석 — 변동성·주가·신용스프레드·보장수익률·전환가액 가정 변동의 영향 측정",
        "⑥ 결과 검증 — 모형 간 차이 자동 해석, BDT 풋채권 교차검증",
    ]
    for s in steps_list:
        c = ws.cell(row=row, column=2, value=s)
        c.font = Font(name="맑은 고딕", size=10)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 22
        row += 1
    row += 1

    # 4. 핵심 가정
    ws.cell(row=row, column=2, value="■ 핵심 가정").font = Font(name="맑은 고딕", bold=True, size=12, color="1A365D")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1
    assumptions = [
        ("주가 (S₀)", f"{int(params.stock_price):,}원", "비상장 시 DCF 산출 또는 직전 거래가"),
        ("변동성 (σ)", f"{params.volatility*100:.2f}%", "유사기업 바스켓 역사적 변동성"),
        ("무위험이자율 (Rf)", f"{params.risk_free_rate*100:.2f}%", "국고채 부트스트랩 곡선 (잔여 만기 매칭)"),
        ("신용스프레드", f"{params.credit_spread*100:.2f}%", "신용등급별 회사채 YTM 부트스트랩"),
        ("할인율 합계 (Kd)", f"{params.discount_rate*100:.2f}%", "Rf + 신용스프레드"),
        ("전환가액", f"{int(params.conversion_price):,}원", "약정 기준 (리픽싱 반영 시 동적)"),
        ("보장수익률 (IRR)", f"{params.put_irr*100:.2f}%" if params.put_irr else "—",
         "풋 행사 시 보장 수익률"),
    ]
    for k, v, note in assumptions:
        bg = "F7FAFC" if row % 2 == 0 else "FFFFFF"
        _cell(ws, row, 2, k, bold=True, bg=bg)
        _cell(ws, row, 3, v, bg=bg, align="right")
        _cell(ws, row, 5, note, bg=bg, color="718096")
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        row += 1
    row += 1

    # 5. 분해 (메인 TF 기준)
    if tf and tf.get("fair_value"):
        ws.cell(row=row, column=2, value="■ 공정가치 분해 (TF 메인 기준)").font = Font(name="맑은 고딕", bold=True, size=12, color="1A365D")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1
        _hdr(ws, row, 2, "구성요소")
        _hdr(ws, row, 3, "금액")
        _hdr(ws, row, 4, "비중")
        _hdr(ws, row, 5, "회계적 의미")
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        row += 1
        decomp_rows = [
            ("① 순채권가치", tf.get("bond_value"), "발행자 보장 의무 PV"),
            ("② 풋옵션가치", tf.get("put_option_value"), "조기 상환 청구권 가치"),
            ("③ 전환권가치", tf.get("conversion_value"), "주식 전환권 추가 가치"),
        ]
        for label, val, note in decomp_rows:
            bg = "F7FAFC" if row % 2 == 0 else "FFFFFF"
            _cell(ws, row, 2, label, bg=bg)
            _cell(ws, row, 3, val if val else "-", bg=bg, num_fmt="#,##0", align="right")
            pct_v = (val / main_fv * 100) if val and main_fv else 0
            _cell(ws, row, 4, f"{pct_v:.1f}%", bg=bg, align="right")
            _cell(ws, row, 5, note, bg=bg, color="718096")
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
            row += 1
        # 합계
        bg = "EBF8FF"
        _cell(ws, row, 2, "공정가치 합계", bold=True, bg=bg)
        _cell(ws, row, 3, main_fv, bold=True, bg=bg, num_fmt="#,##0", align="right")
        _cell(ws, row, 4, "100.0%", bold=True, bg=bg, align="right")
        row += 2

    # 6. 측정 불확실성 (K-IFRS 1113.93(g)·(h)(ii))
    ws.cell(row=row, column=2, value="■ 측정 불확실성").font = Font(name="맑은 고딕", bold=True, size=12, color="1A365D")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1
    uncertainty_notes = [
        f"· 본 RCPS는 K-IFRS 1113 수준 3 공정가치 (시장에서 직접 관측 불가한 input 포함).",
        f"· 모형 간 차이: TF와 GS는 할인 방식 차이로 결과가 갈림 — 자세한 내용은 모형비교 시트 참조.",
        f"· 변동성·신용스프레드·전환가액 등 핵심 가정의 변동 시 공정가치 영향은 민감도 분석 시트 참조.",
        f"· 평가일 시점 가정에 의존 — 후속 시점에서는 새 가정으로 재평가 필요.",
    ]
    for note in uncertainty_notes:
        c = ws.cell(row=row, column=2, value=note)
        c.font = Font(name="맑은 고딕", size=10, color="78350F")
        c.fill = PatternFill(fill_type="solid", fgColor="FFFBEB")
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 22
        row += 1


def _write_cover(ws, params, initial):
    for col, w in zip("ABCD", [26, 22, 20, 20]):
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 36

    t = ws.cell(row=1, column=1, value="상환전환우선주(RCPS) 공정가치 평가 조서")
    t.font = Font(name="맑은 고딕", bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill(fill_type="solid", fgColor="1A365D")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:D1")

    ws.cell(row=2, column=1, value=f"작성일: {datetime.today().strftime('%Y-%m-%d')}").font = Font(name="맑은 고딕", size=9, color="718096")
    ws.cell(row=2, column=3, value=f"평가기준일: {params.valuation_date}").font = Font(name="맑은 고딕", size=9, color="718096")

    ws.cell(row=4, column=1, value="■ 발행 조건").font = Font(name="맑은 고딕", bold=True, size=11)
    for i, h in enumerate(["항목", "내용", "항목", "내용"], 1):
        _hdr(ws, 5, i, h)

    issue_rows = [
        ("발행일", str(params.issue_date), "만기일", str(params.maturity_date)),
        ("발행금액", f"{params.face_value:,.0f} 원", "보장수익률(IRR)", f"{params.put_irr*100:.2f}%"),
        ("전환가액", f"{params.conversion_price:,.0f} 원/주", "우선배당률", f"{params.coupon_rate*100:.1f}%"),
        ("전환청구시작", str(params.conversion_start), "리픽싱여부", "있음" if params.refixing else "없음"),
    ]
    if params.refixing:
        issue_rows.append(("리픽싱 하한", f"{params.refixing_floor*100:.0f}%", "리픽싱 트리거", f"{params.refixing_trigger*100:.0f}%"))

    for ri, row_data in enumerate(issue_rows, 6):
        bg = "F7FAFC" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row_data, 1):
            _cell(ws, ri, ci, val, bg=bg)

    r = 6 + len(issue_rows) + 2
    ws.cell(row=r, column=1, value="■ 시장 데이터 (평가기준일)").font = Font(name="맑은 고딕", bold=True, size=11)
    for i, h in enumerate(["항목", "내용", "항목", "내용"], 1):
        _hdr(ws, r + 1, i, h)

    mkt_rows = [
        ("주가", f"{params.stock_price:,.0f} 원", "잔존기간", f"{initial['key_inputs']['time_to_maturity']:.4f}년"),
        ("변동성(연환산)", f"{params.volatility*100:.1f}%", "무위험이자율", f"{params.risk_free_rate*100:.2f}%"),
        ("신용스프레드", f"{params.credit_spread*100:.2f}%", "할인율(합계)", f"{params.discount_rate*100:.2f}%"),
    ]
    for ri, row_data in enumerate(mkt_rows, r + 2):
        bg = "F7FAFC" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row_data, 1):
            _cell(ws, ri, ci, val, bg=bg)


def _write_valuation(ws, initial):
    for col, w in zip("ABC", [30, 24, 14]):
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 32

    t = ws.cell(row=1, column=1, value="평가 결과")
    t.font = Font(name="맑은 고딕", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill(fill_type="solid", fgColor="1A365D")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:C1")

    for i, h in enumerate(["항목", "금액 (원)", "비율"], 1):
        _hdr(ws, 3, i, h)

    fv = initial["fair_value"]
    straight = initial["straight_bond_value"]
    conv = initial["conversion_component"]

    rows = [
        ("공정가치 (FV)", fv, "100.0%"),
        ("  ├ 순채권가치 (Straight Bond)", straight, f"{straight/fv*100:.1f}%"),
        ("  └ 전환권 가치 (Conversion)", conv, f"{conv/fv*100:.1f}%"),
    ]
    for ri, (label, val, pct) in enumerate(rows, 4):
        bold = ri == 4
        bg = "EBF8FF" if ri == 4 else ("F7FAFC" if ri % 2 == 0 else "FFFFFF")
        _cell(ws, ri, 1, label, bold=bold, bg=bg)
        _cell(ws, ri, 2, val, bold=bold, bg=bg, num_fmt="#,##0", align="right")
        _cell(ws, ri, 3, pct, bg=bg, align="center")

    ws.cell(row=9, column=1, value="■ 모델 정보").font = Font(name="맑은 고딕", bold=True, size=11)
    for i, h in enumerate(["항목", "내용"], 1):
        _hdr(ws, 10, i, h)

    model_rows = [
        ("평가모형", initial["model"]),
        ("트리 단계수", str(initial["steps"])),
        ("위험중립확률 (p)", f"{initial['binomial_detail']['risk_neutral_prob']:.4f}"),
        ("채권 할인율", f"{initial['binomial_detail']['discount_rate']*100:.2f}%"),
        ("상승배수 (u)", f"{initial['binomial_detail']['u']:.6f}"),
        ("하락배수 (d)", f"{initial['binomial_detail']['d']:.6f}"),
    ]
    for ri, (k, v) in enumerate(model_rows, 11):
        bg = "F7FAFC" if ri % 2 == 0 else "FFFFFF"
        _cell(ws, ri, 1, k, bg=bg)
        _cell(ws, ri, 2, v, bg=bg)


def _write_sensitivity(ws, sensitivity):
    for col, w in zip("ABCDEFG", [14, 20, 12, 4, 14, 20, 12]):
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 32

    t = ws.cell(row=1, column=1, value="민감도 분석")
    t.font = Font(name="맑은 고딕", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill(fill_type="solid", fgColor="1A365D")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:G1")

    base_fv = sensitivity["base_fair_value"]
    ws.cell(row=2, column=1, value=f"기준 공정가치: {base_fv:,.0f} 원").font = Font(name="맑은 고딕", bold=True, size=10)

    def _sens_block(start_row, start_col, title, data, col_label):
        ws.cell(row=start_row, column=start_col, value=f"■ {title}").font = Font(name="맑은 고딕", bold=True, size=11)
        for i, h in enumerate([col_label, "공정가치(원)", "변동률"], 1):
            _hdr(ws, start_row + 1, start_col + i - 1, h)
        for ri, row in enumerate(data, start_row + 2):
            bg = "EBF8FF" if abs(row["change_pct"]) < 0.5 else ("F7FAFC" if ri % 2 == 0 else "FFFFFF")
            chg_bg = "C6F6D5" if row["change_pct"] > 0 else ("FED7D7" if row["change_pct"] < 0 else bg)
            _cell(ws, ri, start_col, row["label"], bg=bg, align="center")
            _cell(ws, ri, start_col + 1, row["fair_value"], bg=bg, num_fmt="#,##0", align="right")
            _cell(ws, ri, start_col + 2, f"{row['change_pct']:+.2f}%", bg=chg_bg, align="center")

    _sens_block(4, 1, "변동성 민감도", sensitivity["volatility"], "변동성")
    _sens_block(4, 5, "주가 민감도", sensitivity["stock_price"], "주가(원)")

    offset = max(len(sensitivity["volatility"]), len(sensitivity["stock_price"])) + 7
    _sens_block(offset, 1, "신용스프레드 민감도", sensitivity["credit_spread"], "신용스프레드")
    # 보장수익률(IRR) 민감도 — RCPS 핵심 가정 (있을 때만)
    irr_data = sensitivity.get("put_irr", [])
    conv_data = sensitivity.get("conversion_price", [])
    next_offset = offset + len(sensitivity["credit_spread"]) + 4
    if irr_data:
        _sens_block(next_offset, 1, "보장수익률(IRR) 민감도", irr_data, "IRR")
    if conv_data:
        _sens_block(next_offset, 5, "전환가액 민감도", conv_data, "전환가(원)")
