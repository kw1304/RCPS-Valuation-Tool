"""상세 Excel 다운로드 모듈 — DCF, WACC, 부트스트래핑 시트별.

각 함수는 (data dict, output_path) → 파일 저장. 프론트에서 계산된 결과를 그대로 받아 시트화.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# ── 공통 스타일 ────────────────────────────────────────────────
_BORDER = Border(left=Side(style="thin", color="DDDDDD"),
                 right=Side(style="thin", color="DDDDDD"),
                 top=Side(style="thin", color="DDDDDD"),
                 bottom=Side(style="thin", color="DDDDDD"))
_NAVY = "1A365D"
_BLUE = "3182F6"
_LIGHT = "EBF8FF"
_ALT = "F7FAFC"


def _cell(ws, r, c, val, *, bold=False, bg=None, fmt=None, align="left", color=None, size=10):
    cell = ws.cell(row=r, column=c, value=val)
    kw = {"name": "맑은 고딕", "bold": bold, "size": size}
    if color: kw["color"] = color
    cell.font = Font(**kw)
    if bg: cell.fill = PatternFill(fill_type="solid", fgColor=bg)
    if fmt: cell.number_format = fmt
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    cell.border = _BORDER
    return cell


def _hdr(ws, r, c, val):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
    cell.fill = PatternFill(fill_type="solid", fgColor=_NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(left=Side(style="thin", color="FFFFFF"),
                         right=Side(style="thin", color="FFFFFF"),
                         top=Side(style="thin", color="FFFFFF"),
                         bottom=Side(style="thin", color="FFFFFF"))
    return cell


def _title(ws, text, span_cols, height=32):
    ws.row_dimensions[1].height = height
    t = ws.cell(row=1, column=1, value=text)
    t.font = Font(name="맑은 고딕", bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill(fill_type="solid", fgColor=_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    end = chr(ord('A') + span_cols - 1)
    ws.merge_cells(f"A1:{end}1")


def _section(ws, r, text):
    cell = ws.cell(row=r, column=1, value=text)
    cell.font = Font(name="맑은 고딕", bold=True, size=11, color=_NAVY)


# ══════════════════════════════════════════════════════════════
#  DCF 상세 Excel
# ══════════════════════════════════════════════════════════════
def generate_dcf_xlsx(data, output_path):
    """DCF 상세 워크북. data 구조:
        {valuation_date, wacc, g, tax,
         interest_debt, shares, non_op, nci,
         years: [{revenue, ebit, tax_rate, da, capex, dnwc}, ...],
         result: {op_val, pv_explicit, pv_terminal, terminal_value, equity, sp}}
    """
    wb = Workbook()

    # 시트 1: 가정
    ws = wb.active
    ws.title = "1.가정"
    for col, w in zip("ABCD", [28, 24, 22, 22]):
        ws.column_dimensions[col].width = w
    _title(ws, "DCF 평가 가정", 4)
    _cell(ws, 2, 1, f"평가기준일: {data.get('valuation_date','')}", color="718096", size=9)
    _cell(ws, 2, 3, f"작성일: {datetime.today().strftime('%Y-%m-%d')}", color="718096", size=9)
    _section(ws, 4, "■ 밸류에이션 파라미터")
    for i, h in enumerate(["항목", "값", "항목", "값"], 1):
        _hdr(ws, 5, i, h)
    pct = lambda v: f"{(v or 0)*100:.2f}%" if isinstance(v, (int, float)) else (v or "-")
    krw = lambda v: f"{int(v or 0):,} 원"
    rows = [
        ("WACC", pct(data.get("wacc")), "영구성장률 g", pct(data.get("g"))),
        ("기본 세율", pct(data.get("tax")), "터미널 g/WACC", f"{(data.get('g',0)/data.get('wacc',1))*100:.1f}%" if data.get('wacc') else "-"),
        ("이자부부채", krw(data.get("interest_debt")), "총발행주식수", f"{int(data.get('shares',0)):,} 주"),
        ("비영업자산", krw(data.get("non_op")), "비지배지분조정", krw(data.get("nci"))),
    ]
    for ri, row in enumerate(rows, 6):
        bg = _ALT if ri % 2 == 0 else "FFFFFF"
        for ci, v in enumerate(row, 1):
            _cell(ws, ri, ci, v, bg=bg)

    # 시트 2: 연도별 FCFF
    ws2 = wb.create_sheet("2.FCFF추정")
    years = data.get("years", []) or []
    ncol = len(years) + 1
    for c in range(1, ncol + 1):
        ws2.column_dimensions[chr(ord('A') + c - 1)].width = 18
    _title(ws2, "연도별 FCFF 추정", ncol)
    _hdr(ws2, 3, 1, "항목")
    for i, _y in enumerate(years, 2):
        _hdr(ws2, 3, i, f"{i-1}년차")

    lines = [
        ("매출액", lambda y: y.get("revenue", 0), "money", False),
        ("  └ 매출성장률 (YoY)", "GROWTH", "pct", False),
        ("영업이익 (EBIT)", lambda y: y.get("ebit", 0), "money", False),
        ("  └ EBIT 마진", "MARGIN", "pct", False),
        ("세율 (%)", lambda y: y.get("tax", 0), "pct_raw", False),
        ("NOPAT = EBIT × (1−t)", "NOPAT", "money", True),
        ("(+) 감가상각비 D&A", lambda y: y.get("da", 0), "money", False),
        ("(−) 자본지출 CapEx", lambda y: -abs(y.get("capex", 0)), "money", False),
        ("(−) 운전자본 순증가 ΔNWC", lambda y: -abs(y.get("dnwc", 0)), "money", False),
        ("FCFF", "FCFF", "money", True),
        ("  └ 할인계수", "DISC", "decimal", False),
        ("  └ FCFF 현재가치 (PV)", "PV", "money", True),
    ]
    wacc = float(data.get("wacc", 0.12) or 0.12)
    for ri_off, (label, fn, fmt, highlight) in enumerate(lines):
        r = 4 + ri_off
        bg = _LIGHT if highlight else (_ALT if r % 2 == 0 else "FFFFFF")
        _cell(ws2, r, 1, label, bold=highlight, bg=bg)
        for i, y in enumerate(years, 2):
            tax = float(y.get("tax", 25)) / 100
            ebit = float(y.get("ebit", 0))
            nopat = ebit * (1 - tax)
            fcff = nopat + float(y.get("da", 0)) - abs(float(y.get("capex", 0))) - abs(float(y.get("dnwc", 0)))
            disc = (1 / (1 + wacc)) ** (i - 1)
            pv = fcff * disc
            if fn == "GROWTH":
                if i > 2 and float(years[i - 3].get("revenue", 0)):
                    v = (y.get("revenue", 0) - years[i - 3].get("revenue", 0)) / years[i - 3].get("revenue", 0) * 100
                    _cell(ws2, r, i, f"{v:+.1f}%", bg=bg, align="right")
                else:
                    _cell(ws2, r, i, "-", bg=bg, align="right")
            elif fn == "MARGIN":
                v = (y.get("ebit", 0) / y.get("revenue", 1) * 100) if y.get("revenue") else 0
                _cell(ws2, r, i, f"{v:.1f}%", bg=bg, align="right")
            elif fn == "NOPAT":
                _cell(ws2, r, i, nopat, bold=True, bg=bg, fmt="#,##0", align="right")
            elif fn == "FCFF":
                _cell(ws2, r, i, fcff, bold=True, bg=bg, fmt="#,##0", align="right")
            elif fn == "DISC":
                _cell(ws2, r, i, disc, bg=bg, fmt="0.00000", align="right")
            elif fn == "PV":
                _cell(ws2, r, i, pv, bold=True, bg=bg, fmt="#,##0", align="right")
            elif fmt == "money":
                _cell(ws2, r, i, fn(y), bg=bg, fmt="#,##0", align="right")
            elif fmt == "pct_raw":
                _cell(ws2, r, i, f"{fn(y):.1f}%", bg=bg, align="right")

    # 시트 3: 평가 결과
    ws3 = wb.create_sheet("3.평가결과")
    for col, w in zip("ABC", [32, 22, 14]):
        ws3.column_dimensions[col].width = w
    _title(ws3, "DCF 평가 결과 — 자기자본가치 산정", 3)
    _hdr(ws3, 3, 1, "항목"); _hdr(ws3, 3, 2, "금액 (원)"); _hdr(ws3, 3, 3, "비고")

    res = data.get("result", {}) or {}
    pv_ex = float(res.get("pv_explicit", 0))
    pv_tv = float(res.get("pv_terminal", 0))
    tv_nom = float(res.get("terminal_value", 0))
    op_val = pv_ex + pv_tv
    nonop = float(data.get("non_op", 0))
    debt = float(data.get("interest_debt", 0))
    nci = float(data.get("nci", 0))
    equity = float(res.get("equity", op_val + nonop - debt - nci))
    shares = float(data.get("shares", 1))
    sp = float(res.get("sp", equity / max(shares, 1)))

    rows = [
        ("명시기간 FCFF 현재가치 합계", pv_ex, "= Σ PV(FCFF)"),
        ("터미널 밸류 (명목)", tv_nom, "= FCFF_N × (1+g) / (WACC−g)"),
        ("터미널 밸류 현재가치", pv_tv, f"= TV / (1+WACC)^N"),
        ("영업가치 (Operating Value)", op_val, f"TV/Op {(pv_tv/op_val*100):.1f}%" if op_val else "-"),
        ("(+) 비영업자산", nonop, "현금성·금융자산·관계기업투자"),
        ("(−) 이자부부채", -debt, "차입금+리스부채"),
        ("(−) 비지배지분조정", -nci, "연결평가 시"),
        ("자기자본가치", equity, "Equity Value"),
        ("÷ 총발행주식수", -shares, f"{int(shares):,} 주"),
        ("주당 내재가치", sp, "Per share"),
    ]
    for ri, (k, v, note) in enumerate(rows, 4):
        bold = ri in (7, 11, 13)
        bg = _LIGHT if bold else (_ALT if ri % 2 == 0 else "FFFFFF")
        _cell(ws3, ri, 1, k, bold=bold, bg=bg)
        if ri == 12:  # shares 별도 처리
            _cell(ws3, ri, 2, abs(v), bg=bg, fmt="#,##0", align="right")
        else:
            _cell(ws3, ri, 2, v, bold=bold, bg=bg, fmt="#,##0", align="right")
        _cell(ws3, ri, 3, note, bg=bg, color="718096", size=9)

    wb.save(output_path)
    return output_path


# ══════════════════════════════════════════════════════════════
#  WACC 상세 Excel
# ══════════════════════════════════════════════════════════════
def generate_wacc_xlsx(data, output_path):
    """WACC 상세 워크북. data:
        {valuation_date, rf, mrp, country_p, size_p, beta_target, ke,
         tax, kd, kd_after, ev_pct, dv_pct, de_target, wacc,
         peers: [{name, code, beta_l, de, beta_u, beta_r2, beta_period, de_period}, ...],
         market_idx, beta_agg, beta_freq}
    """
    wb = Workbook()
    pct = lambda v: f"{(v or 0)*100:.2f}%" if isinstance(v, (int, float)) else (v or "-")

    # 시트 1: WACC 결과
    ws = wb.active
    ws.title = "1.WACC결과"
    for col, w in zip("ABCD", [28, 22, 28, 22]):
        ws.column_dimensions[col].width = w
    _title(ws, "WACC 산정 결과", 4)
    _cell(ws, 2, 1, f"평가기준일: {data.get('valuation_date','')}", color="718096", size=9)
    _cell(ws, 2, 3, f"작성일: {datetime.today().strftime('%Y-%m-%d')}", color="718096", size=9)

    _section(ws, 4, "■ WACC 공식 분해")
    for i, h in enumerate(["요소", "값", "요소", "값"], 1):
        _hdr(ws, 5, i, h)
    rows = [
        ("E/V (자기자본 비중)", pct((data.get("ev_pct") or 0) / 100), "D/V (부채 비중)", pct((data.get("dv_pct") or 0) / 100)),
        ("Ke (CAPM)", pct(data.get("ke")), "Kd (세전)", pct(data.get("kd"))),
        ("Kd × (1−t) 세후", pct(data.get("kd_after")), "법인세율 t", pct(data.get("tax"))),
        ("D/E_target", f"{(data.get('de_target') or 0):.2f}%", "WACC", pct(data.get("wacc"))),
    ]
    for ri, row in enumerate(rows, 6):
        bg = _LIGHT if ri == 9 else (_ALT if ri % 2 == 0 else "FFFFFF")
        for ci, v in enumerate(row, 1):
            _cell(ws, ri, ci, v, bold=(ri == 9), bg=bg)

    # 시트 2: Ke 분해
    ws2 = wb.create_sheet("2.CAPM_Ke")
    for col, w in zip("ABC", [32, 18, 32]):
        ws2.column_dimensions[col].width = w
    _title(ws2, "자기자본비용 Ke = Rf + β·MRP + Size + CRP", 3)
    _hdr(ws2, 3, 1, "요소"); _hdr(ws2, 3, 2, "값"); _hdr(ws2, 3, 3, "비고")
    ke = data.get("ke", 0) or 0
    rows = [
        ("무위험이자율 Rf", pct(data.get("rf")), "통상 평가기준일 국고채 10년"),
        ("시장위험프리미엄 MRP", pct(data.get("mrp")), "Damodaran 등 연도별 ERP"),
        ("× β_target_L (relevered)", f"{data.get('beta_target', 0):.4f}", "유사기업 β_U 집계 후 D/E_target으로 relever"),
        ("→ β_L × MRP", pct((data.get('beta_target', 0) or 0) * (data.get('mrp', 0) or 0)), ""),
        ("(+) 사이즈 프리미엄", pct(data.get("size_p")), "Duff & Phelps / Kroll"),
        ("(+) 국가위험프리미엄 CRP", pct(data.get("country_p")), "Damodaran 국가별"),
        ("자기자본비용 Ke", pct(ke), "= Rf + β·MRP + Size + CRP"),
    ]
    for ri, (k, v, note) in enumerate(rows, 4):
        bold = ri in (7, 10)
        bg = _LIGHT if bold else (_ALT if ri % 2 == 0 else "FFFFFF")
        _cell(ws2, ri, 1, k, bold=bold, bg=bg)
        _cell(ws2, ri, 2, v, bold=bold, bg=bg, align="right")
        _cell(ws2, ri, 3, note, bg=bg, color="718096", size=9)

    # 시트 3: 유사기업 베타 / D/E
    ws3 = wb.create_sheet("3.유사기업")
    cols = ["A", "B", "C", "D", "E", "F", "G"]
    widths = [28, 14, 14, 14, 14, 24, 24]
    for col, w in zip(cols, widths):
        ws3.column_dimensions[col].width = w
    _title(ws3, "유사기업 베타 회귀 + D/E (Hamada Unlever/Re-lever)", 7)
    _cell(ws3, 2, 1, f"시장지수: {data.get('market_idx', '-')}", color="718096", size=9)
    _cell(ws3, 2, 3, f"수익률 주기: {data.get('beta_freq', 'daily')}", color="718096", size=9)
    _cell(ws3, 2, 5, f"집계: {data.get('beta_agg', 'median')}", color="718096", size=9)
    for i, h in enumerate(["유사기업", "코드", "β_L (raw)", "D/E (%)", "β_U", "β 회귀 기간", "D/E 기준일"], 1):
        _hdr(ws3, 4, i, h)
    tax = float(data.get("tax", 0.25) or 0.25)
    peers = data.get("peers", []) or []
    for ri, p in enumerate(peers, 5):
        bg = _ALT if ri % 2 == 0 else "FFFFFF"
        de = float(p.get("de", 0))
        bl = float(p.get("beta_l", 0))
        bu = bl / (1 + (1 - tax) * de / 100) if (1 + (1 - tax) * de / 100) > 0 else bl
        _cell(ws3, ri, 1, p.get("name", ""), bg=bg)
        _cell(ws3, ri, 2, p.get("code", ""), bg=bg, align="center")
        _cell(ws3, ri, 3, f"{bl:.4f}", bg=bg, align="right")
        _cell(ws3, ri, 4, f"{de:.2f}%", bg=bg, align="right")
        _cell(ws3, ri, 5, f"{bu:.4f}", bg=bg, align="right")
        _cell(ws3, ri, 6, p.get("beta_period", "-"), bg=bg, color="718096", size=9)
        _cell(ws3, ri, 7, p.get("de_period", "-"), bg=bg, color="718096", size=9)

    wb.save(output_path)
    return output_path


# ══════════════════════════════════════════════════════════════
#  부트스트래핑 곡선 Excel
# ══════════════════════════════════════════════════════════════
def generate_bootstrap_xlsx(data, output_path):
    """이자율 부트스트래핑 결과. data:
        {valuation_date, rf_input: [{n, year, ytm}, ...],
         rf_mid: [...], rf_out: [{year, d_spot_y, c_spot_y}, ...],
         rd_input: [...], rd_mid: [...], rd_out: [...],
         rate_data: [{node, date, dt, cum_dt, rf_dspot, rf_cspot, rf_cfwd, rf_cdf, rd_dspot, rd_cspot, rd_cfwd, rd_cdf}, ...]}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "1.Rf곡선"

    def _curve_sheet(ws, label, inp, mid, out):
        for col, w in zip("ABCDEFGHIJKL", [8, 10, 12, 4, 10, 10, 12, 14, 14, 4, 12, 14]):
            ws.column_dimensions[col].width = w
        _title(ws, f"{label} 곡선 — Bootstrapping", 11)
        _cell(ws, 2, 1, f"평가기준일: {data.get('valuation_date','')}", color="718096", size=9)

        # 입력
        _section(ws, 4, f"■ 시장 YTM 입력")
        for i, h in enumerate(["NO", "YEAR", f"{label}-YTM"], 1):
            _hdr(ws, 5, i, h)
        for ri, r in enumerate(inp or [], 6):
            bg = _ALT if ri % 2 == 0 else "FFFFFF"
            _cell(ws, ri, 1, r.get("n"), bg=bg, align="center")
            _cell(ws, ri, 2, r.get("year"), bg=bg, align="center")
            _cell(ws, ri, 3, f"{r.get('ytm',0):.4f}%", bg=bg, align="right")

        # 중간 (기간별 분해)
        r0 = 5 + len(inp or []) + 2
        _section(ws, r0, f"■ 기간별 분해 (반기/분기)")
        period_label = "HALF" if label == "Rf" else "QUARTER"
        for i, h in enumerate(["YEAR", period_label, f"{label}", "D-SPOT", "PV FACTOR"], 5):
            _hdr(ws, r0 + 1, i - 4, h)
        for ri, r in enumerate(mid or [], r0 + 2):
            bg = _ALT if ri % 2 == 0 else "FFFFFF"
            _cell(ws, ri, 1, r.get("year"), bg=bg, align="center")
            _cell(ws, ri, 2, r.get("period"), bg=bg, align="center")
            _cell(ws, ri, 3, f"{r.get('rate',0):.6f}", bg=bg, align="right")
            _cell(ws, ri, 4, f"{r.get('dspot',0):.6f}", bg=bg, align="right")
            _cell(ws, ri, 5, f"{r.get('pvf',0):.6f}", bg=bg, align="right")

        # 출력 (연단위)
        r0 = ri + 2
        _section(ws, r0, f"■ 연단위 곡선 (D-SPOT / C-SPOT)")
        for i, h in enumerate(["YEAR", "D-SPOT (Y)", "C-SPOT (Y)"], 7):
            _hdr(ws, r0 + 1, i - 6, h)
        for ri, r in enumerate(out or [], r0 + 2):
            bg = _ALT if ri % 2 == 0 else "FFFFFF"
            _cell(ws, ri, 1, r.get("year"), bg=bg, align="center")
            _cell(ws, ri, 2, f"{r.get('d_spot_y',0):.6f}", bg=bg, align="right")
            _cell(ws, ri, 3, f"{r.get('c_spot_y',0):.6f}", bg=bg, align="right")

    _curve_sheet(ws, "Rf", data.get("rf_input"), data.get("rf_mid"), data.get("rf_out"))

    ws2 = wb.create_sheet("2.Rd곡선")
    _curve_sheet(ws2, "Rd", data.get("rd_input"), data.get("rd_mid"), data.get("rd_out"))

    # 시트 3: 평가 노드별 선도이자율 데이터
    ws3 = wb.create_sheet("3.이자율DATA")
    cols = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
    for col, w in zip(cols, [6, 12, 8, 10, 12, 12, 12, 12, 12, 12, 12, 12]):
        ws3.column_dimensions[col].width = w
    _title(ws3, "이자율 DATA — 평가 노드별 선도이자율 (BM)", 12)
    headers = ["NODE", "DATE", "dT", "누적dT", "Rf D-SPOT", "Rf C-SPOT", "Rf C-FWD", "Rf C-DF",
               "Rd D-SPOT", "Rd C-SPOT", "Rd C-FWD", "Rd C-DF"]
    for i, h in enumerate(headers, 1):
        _hdr(ws3, 3, i, h)
    for ri, r in enumerate(data.get("rate_data") or [], 4):
        bg = _ALT if ri % 2 == 0 else "FFFFFF"
        _cell(ws3, ri, 1, r.get("node"), bg=bg, align="center")
        _cell(ws3, ri, 2, r.get("date"), bg=bg, align="center")
        _cell(ws3, ri, 3, r.get("dt"), bg=bg, fmt="0.0000", align="right")
        _cell(ws3, ri, 4, r.get("cum_dt"), bg=bg, fmt="0.0000", align="right")
        for ci, key in enumerate(["rf_dspot","rf_cspot","rf_cfwd","rf_cdf","rd_dspot","rd_cspot","rd_cfwd","rd_cdf"], 5):
            v = r.get(key)
            _cell(ws3, ri, ci, v if v is not None else "-", bg=bg, fmt="0.000000", align="right")

    wb.save(output_path)
    return output_path


# ══════════════════════════════════════════════════════════════
#  변동성 평가 상세 Excel
# ══════════════════════════════════════════════════════════════
def generate_volatility_xlsx(data, output_path):
    """변동성 산정 결과 워크북. data:
        {valuation_date, sigma_pct, method, trading_days, start, end,
         outlier_info: {method, k, applied, reason},
         per_ticker: [{ticker, name, sigma, sigma_pct, trading_days_used, n_obs, outlier, reason, marcap}],
         failed: [{ticker, reason}],
         raw: {dates, series: {ticker: [close,...]}}}"""
    from openpyxl.utils import get_column_letter as _gcl
    wb = Workbook()

    # 1. 산정조건
    ws = wb.active
    ws.title = "1.산정조건"
    for col, w in zip("ABCD", [28, 22, 28, 22]):
        ws.column_dimensions[col].width = w
    _title(ws, "변동성(σ) 산정 조건", 4)
    _cell(ws, 2, 1, f"평가기준일: {data.get('valuation_date','')}", color="718096", size=9)
    _cell(ws, 2, 3, f"작성일: {datetime.today().strftime('%Y-%m-%d')}", color="718096", size=9)
    _section(ws, 4, "■ 산정 파라미터")
    for i, h in enumerate(["항목", "값", "항목", "값"], 1):
        _hdr(ws, 5, i, h)
    mlabel = {"median": "중앙값", "mean": "단순평균", "cap_weighted": "시총가중"}.get(data.get("method", ""), data.get("method", ""))
    rows = [
        ("산정 시작일", data.get("start", "-"), "산정 종료일", data.get("end", "-")),
        ("집계 방식", mlabel, "연 거래일 수", str(data.get("trading_days", "auto"))),
    ]
    oi = data.get("outlier_info") or {}
    if oi.get("method") and oi.get("method") != "none":
        ol_label = {"iqr": "IQR (Tukey)", "mad": "MAD"}.get(oi["method"], oi["method"])
        ol_status = f"{ol_label} k={oi.get('k', '-')}× 적용됨" if oi.get("applied") else f"{ol_label} 미적용 ({oi.get('reason', '-')})"
        rows.append(("이상치 필터", ol_status, "", ""))
    for ri, row in enumerate(rows, 6):
        bg = _ALT if ri % 2 == 0 else "FFFFFF"
        for ci, v in enumerate(row, 1):
            _cell(ws, ri, ci, v, bg=bg)
    r0 = 6 + len(rows) + 2
    _section(ws, r0, "■ 산출 결과")
    _hdr(ws, r0 + 1, 1, "항목"); _hdr(ws, r0 + 1, 2, "값")
    sig = data.get("sigma_pct", 0) or 0
    per = data.get("per_ticker") or []
    valid = [p for p in per if p.get("sigma") is not None]
    used = [p for p in valid if not p.get("outlier")]
    failed = data.get("failed") or []
    out_rows = [
        ("집계 변동성 σ (연율)", f"{sig:.2f}%"),
        ("유효 종목 수", f"{len(valid)}개"),
        ("사용 종목 수 (이상치 제외 후)", f"{len(used)}개"),
        ("이상치 제외 종목", f"{len(valid) - len(used)}개"),
        ("조회 실패 종목", f"{len(failed)}개"),
    ]
    for ri, (k, v) in enumerate(out_rows, r0 + 2):
        bg = _LIGHT if ri == r0 + 2 else (_ALT if ri % 2 == 0 else "FFFFFF")
        _cell(ws, ri, 1, k, bold=(ri == r0 + 2), bg=bg)
        _cell(ws, ri, 2, v, bold=(ri == r0 + 2), bg=bg, color=_BLUE if ri == r0 + 2 else None)

    # 2. 유사기업 σ
    ws2 = wb.create_sheet("2.유사기업σ")
    cols = ["A", "B", "C", "D", "E", "F", "G", "H"]
    for col, w in zip(cols, [28, 14, 12, 14, 14, 14, 12, 28]):
        ws2.column_dimensions[col].width = w
    _title(ws2, "유사상장기업별 역사적 변동성 (일별 로그수익률 표준편차 × √연거래일)", 8)
    for i, h in enumerate(["종목명", "코드", "σ (연율)", "σ_decimal", "연거래일", "관측치(일)", "이상치", "비고/제외사유"], 1):
        _hdr(ws2, 3, i, h)
    for ri, p in enumerate(per, 4):
        bg = _ALT if ri % 2 == 0 else "FFFFFF"
        ol_bg = "FEEBC8" if p.get("outlier") else bg
        fail_bg = "FED7D7" if p.get("sigma") is None else bg
        _cell(ws2, ri, 1, p.get("name") or "-", bg=fail_bg)
        _cell(ws2, ri, 2, p.get("ticker") or "-", bg=fail_bg, align="center")
        sg = p.get("sigma_pct")
        if sg is None and p.get("sigma") is not None:
            sg = p["sigma"] * 100
        _cell(ws2, ri, 3, f"{sg:.2f}%" if sg is not None else "-", bg=ol_bg, align="right", bold=True)
        _cell(ws2, ri, 4, f"{p['sigma']:.6f}" if p.get("sigma") is not None else "-", bg=fail_bg, align="right")
        _cell(ws2, ri, 5, p.get("trading_days_used") or "-", bg=bg, align="center")
        _cell(ws2, ri, 6, p.get("n_obs") or "-", bg=bg, align="center")
        _cell(ws2, ri, 7, "이상치" if p.get("outlier") else "", bg=ol_bg, align="center", color="DC2626" if p.get("outlier") else None)
        _cell(ws2, ri, 8, p.get("reason") or "", bg=fail_bg, color="718096", size=9)

    # 3. 실패 종목 (있으면)
    if failed:
        ws3 = wb.create_sheet("3.조회실패")
        for col, w in zip("AB", [20, 60]):
            ws3.column_dimensions[col].width = w
        _title(ws3, "조회 실패 종목 — 사유", 2)
        _hdr(ws3, 3, 1, "코드"); _hdr(ws3, 3, 2, "사유")
        for ri, f in enumerate(failed, 4):
            bg = _ALT if ri % 2 == 0 else "FFFFFF"
            _cell(ws3, ri, 1, f.get("ticker") or "-", bg=bg, align="center")
            _cell(ws3, ri, 2, f.get("reason") or "-", bg=bg, color="DC2626")

    # 4. 원자료 종가 (옵션)
    raw = data.get("raw") or {}
    dates = raw.get("dates") or []
    series = raw.get("series") or {}
    if dates and series:
        ws4 = wb.create_sheet("4.원자료(종가)")
        tickers = list(series.keys())
        ws4.column_dimensions["A"].width = 12
        for ci in range(2, len(tickers) + 2):
            ws4.column_dimensions[_gcl(ci)].width = 14
        _title(ws4, "원자료 — 일별 종가 (감사 증빙)", len(tickers) + 1)
        _hdr(ws4, 3, 1, "날짜")
        for ci, tk in enumerate(tickers, 2):
            _hdr(ws4, 3, ci, tk)
        for ri, d in enumerate(dates, 4):
            bg = _ALT if ri % 2 == 0 else "FFFFFF"
            _cell(ws4, ri, 1, d, bg=bg, align="center")
            for ci, tk in enumerate(tickers, 2):
                v = series[tk][ri - 4] if ri - 4 < len(series[tk]) else None
                _cell(ws4, ri, ci, v if v is not None else "-", bg=bg, fmt="#,##0.##", align="right")

    wb.save(output_path)
    return output_path
