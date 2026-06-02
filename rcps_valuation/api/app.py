import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from flask import Flask, request, jsonify, send_file, send_from_directory
from datetime import date, datetime, timedelta
import tempfile, traceback, json
import urllib.request, urllib.error
import csv, io, re
import openpyxl

import numpy as np

from inputs.deal_params import RCPSParams
from inputs.dcf import DCFParams, dcf_valuation
from models.tsiveriotis_fernandes import tf_rcps
from models.goldman_sachs import gs_rcps
from models.monte_carlo import monte_carlo_rcps
from sensitivity.analysis import sensitivity_analysis
from output.report import generate_workpaper
from output.exports import generate_dcf_xlsx, generate_wacc_xlsx, generate_bootstrap_xlsx, generate_volatility_xlsx

FRONTEND_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "frontend")


def _load_dotenv():
    """경량 .env 로더 (python-dotenv 의존 없이). 이미 설정된 환경변수는 보존."""
    env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), ".env")
    try:
        with open(env_path, encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                k, v = k.strip(), v.strip().strip('"').strip("'")
                if k and k not in os.environ:
                    os.environ[k] = v
    except OSError:
        pass


_load_dotenv()
app = Flask(__name__, static_folder=FRONTEND_DIR)


# ── HTTP Basic Auth (배포용) ─────────────────────────────────────────
# 환경변수 APP_PASSWORD가 설정돼 있으면 모든 요청에 비밀번호 요구.
# 비어 있으면(로컬 개발) auth 비활성화.
from functools import wraps
from flask import Response

_AUTH_USER = os.environ.get("APP_USER", "admin")
_AUTH_PASS = os.environ.get("APP_PASSWORD", "")  # 비워두면 auth 끔

@app.before_request
def _require_auth():
    # 헬스체크 엔드포인트는 인증 면제 (UptimeRobot 등 keep-alive 핑용)
    if request.path == "/healthz":
        return None
    if not _AUTH_PASS:
        return None  # 비밀번호 미설정 시 통과 (로컬)
    auth = request.authorization
    if auth and auth.username == _AUTH_USER and auth.password == _AUTH_PASS:
        return None
    return Response(
        "Authentication required.", 401,
        {"WWW-Authenticate": 'Basic realm="RCPS Valuation Tool"'}
    )


@app.route("/healthz")
def healthz():
    """UptimeRobot 등 외부 모니터링 keep-alive 용. Render sleep 방지."""
    return {"status": "ok"}, 200


def _d(s):
    return date.fromisoformat(s) if s else None

def _f(v, default=None):
    try:
        return float(v) if v not in (None, "", "null") else default
    except Exception:
        return default


from inputs.curves import spot_to_step_forwards as _spot_to_step_forwards
# spot → 스텝별 forward 변환은 inputs/curves.py 단일 출처 사용 (TF·GS·MC·BDT·후속측정 공통)


def parse_params(data: dict) -> RCPSParams:
    return RCPSParams(
        issue_date=_d(data["issue_date"]),
        maturity_date=_d(data["maturity_date"]),
        face_value=float(data["face_value"]),
        coupon_rate=float(data.get("coupon_rate", 0)),
        coupon_frequency=data.get("coupon_frequency", "annual"),
        dividend_cumulative=bool(data.get("dividend_cumulative", True)),
        dividend_first_pay_year=float(data.get("dividend_first_pay_year", 0) or 0),
        conversion_price=float(data.get("conversion_price", 0)),
        conversion_start=_d(data.get("conversion_start")),
        put_start=_d(data.get("put_start")),
        put_irr=float(data.get("put_irr", 0)),
        put_price_mode=data.get("put_price_mode", "irr"),
        put_fixed_price=float(data.get("put_fixed_price", 0) or 0),
        put_coupon_netting=data.get("put_coupon_netting", "accrual"),
        put_contract_ratio=float(data.get("put_contract_ratio", 0) or 0),
        call_start=_d(data.get("call_start")),
        call_irr=float(data.get("call_irr", 0)),
        call_price_mode=data.get("call_price_mode", "irr"),
        call_fixed_price=float(data.get("call_fixed_price", 0) or 0),
        call_contract_ratio=float(data.get("call_contract_ratio", 0) or 0),
        call_coupon_netting=data.get("call_coupon_netting", "accrual"),
        refixing=bool(data.get("refixing", False)),
        refixing_floor=_f(data.get("refixing_floor")),
        refixing_trigger=_f(data.get("refixing_trigger")),
        refixing_frequency=data.get("refixing_frequency", "continuous"),
        refixing_ratchet=bool(data.get("refixing_ratchet", True)),
        refixing_vwap_window=int(data.get("refixing_vwap_window", 0) or 0),
        call_soft_barrier=_f(data.get("call_soft_barrier")),
        call_soft_window=int(data.get("call_soft_window", 1) or 1),
        call_soft_count=int(data.get("call_soft_count", 1) or 1),
        mandatory_conv_barrier=_f(data.get("mandatory_conv_barrier")),
        mandatory_conv_window=int(data.get("mandatory_conv_window", 1) or 1),
        mandatory_conv_count=int(data.get("mandatory_conv_count", 1) or 1),
        put_barrier=_f(data.get("put_barrier")),
        stock_price=float(data.get("stock_price", 0)),
        volatility=float(data.get("volatility", 0)),
        risk_free_rate=float(data.get("risk_free_rate", 0)),
        credit_spread=float(data.get("credit_spread", 0)),
        dividend_yield=float(data.get("dividend_yield", 0)),
        valuation_date=_d(data["valuation_date"]),
        is_unlisted=bool(data.get("is_unlisted", False)),
    )


def serialize(obj):
    if isinstance(obj, date):
        return obj.isoformat()
    if isinstance(obj, dict):
        return {k: serialize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [serialize(i) for i in obj]
    return obj


@app.route("/")
def index():
    return send_from_directory(FRONTEND_DIR, "index.html")


@app.route("/api/bootstrap", methods=["POST"])
def bootstrap_route():
    """이자율 부트스트래핑 — par yield bond bootstrap (서버측 단일 출처).
    프론트가 결과를 직접 표시하거나, 자동 테스트가 호출해 검증 가능.

    Request: {"rows": [{"t", "y"}, ...], "m": 2|4, "max_T": 20}
    Response: {mid_rows, out_rows, dfs, warning, input_max, max_T_used}
    """
    try:
        from inputs.bootstrap import bootstrap_par_yield
        data = request.json or {}
        rows = data.get("rows", [])
        m = int(data.get("m", 2))
        max_T = float(data.get("max_T", 20.0))
        result = bootstrap_par_yield(rows, m=m, max_T=max_T)
        return jsonify({"status": "ok", "result": result})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 400


@app.route("/api/dcf", methods=["POST"])
def dcf():
    """단일 평가 엔진(`inputs.dcf.dcf_valuation`) 호출 — 프론트·Excel 공통 출처.
    두 가지 모드:
      (A) years: [{revenue, ebit, da, capex, dnwc, tax}, ...] — 1차 입력
      (B) fcf_projections: [float, ...] — 사전 계산된 FCFF (legacy)
    """
    try:
        from inputs.dcf import DCFYear
        data = request.json
        years_raw = data.get("years") or []
        years = [DCFYear(
            revenue=float(y.get("revenue", 0)),
            ebit=float(y.get("ebit", 0)),
            da=float(y.get("da", 0)),
            capex=float(y.get("capex", 0)),
            dnwc=float(y.get("dnwc", 0)),
            tax=float(y.get("tax", 25)),
        ) for y in years_raw]
        # 과거 실적(actual) — 예측가정 검증 앵커 (오래된→최근 순)
        hist_raw = data.get("historical") or []
        historical = [DCFYear(
            revenue=float(y.get("revenue", 0)),
            ebit=float(y.get("ebit", 0)),
            da=float(y.get("da", 0)),
            capex=float(y.get("capex", 0)),
            dnwc=float(y.get("dnwc", 0)),
            tax=float(y.get("tax", 25)),
        ) for y in hist_raw]
        params = DCFParams(
            years=years,
            historical=historical,
            fcf_projections=[float(x) for x in (data.get("fcf_projections") or [])],
            wacc=float(data["wacc"]),
            terminal_growth=float(data.get("terminal_growth", 0.02)),
            net_debt=float(data.get("net_debt", 0)),
            non_operating_assets=float(data.get("non_operating_assets", 0)),
            preferred_value=float(data.get("preferred_value", 0)),
            nci_adjustment=float(data.get("nci_adjustment", 0)),
            total_shares=float(data.get("total_shares", 1)),
            mid_year=bool(data.get("mid_year", False)),
            tv_method=str(data.get("tv_method", "gordon")),
            exit_multiple=float(data.get("exit_multiple", 0) or 0),
            tv_weight_gordon=float(data.get("tv_weight_gordon", 0.5)),
        )
        result = dcf_valuation(params)
        return jsonify({"status": "ok", "result": result})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 400


@app.route("/api/dcf/dart_financials", methods=["POST"])
def dcf_dart_financials():
    """DART OPEN API로 대상회사 최근 5개년 실적 자동조회 (예측가정 검증 앵커).

    body: {company_name 또는 corp_code, end_year(선택), max_years(기본 5),
           fs_div('OFS'별도 기본 / 'CFS'연결)}
    반환: {status, fs_div, corp, years[...], notes, holding_warning, alt_available}

    별도(OFS) 기본 — RCPS 비상장 평가는 그 법인 단독 실적 기준. 대상이 지주성격
    (별도 영업수익≈0)이면 holding_warning=True + 연결 재조회 안내.
    상장·사업보고서 제출 비상장은 구조화 FS, 순수 외감 비상장은 감사보고서 원문 파싱.
    """
    try:
        from inputs.dart_financials import DartFinancials, DartFinancialsError
    except Exception as e:
        return jsonify({"status": "error", "message": f"dart 모듈 로드 실패: {e}"}), 200
    data = request.get_json(force=True) or {}
    name = (data.get("company_name") or "").strip()
    corp_code = (data.get("corp_code") or "").strip()
    end_year = int(data.get("end_year") or 0) or (date.today().year - 1)
    max_years = int(data.get("max_years") or 5)
    fs_div = "CFS" if str(data.get("fs_div", "OFS")).upper() == "CFS" else "OFS"

    d = DartFinancials()
    if not d.enabled:
        return jsonify({"status": "error",
                        "message": "DART_API_KEY 미설정 — .env에 키를 추가하세요."}), 200
    corp = None
    try:
        if not corp_code:
            if not name:
                return jsonify({"status": "error",
                                "message": "company_name 또는 corp_code 필요"}), 200
            corp = d.find_corp_code(name)
            if not corp:
                return jsonify({"status": "error",
                                "message": f"'{name}' DART 등록정보 없음 — 회사명 확인"}), 200
            corp_code = corp["corp_code"]
        res = d.fetch_financials(corp_code, end_year=end_year,
                                 max_years=max_years, prefer=fs_div)
        return jsonify({
            "status": "ok",
            "fs_div": res["fs_div"],
            "corp": corp or {"corp_code": corp_code},
            "years": res["years"],
            "notes": [n for n in res["notes"] if n],
            "holding_warning": res.get("holding_warning", False),
            "alt_available": res.get("alt_available", False),
        })
    except DartFinancialsError as e:
        return jsonify({"status": "error", "message": str(e)}), 200
    except Exception as e:
        return jsonify({"status": "error", "message": f"DART 조회 오류: {e}"}), 200


@app.route("/api/evaluate", methods=["POST"])
def evaluate():
    try:
        data = request.json
        params = parse_params(data["params"])
        steps_req = int(data.get("steps", 0)) or None
        # 사용자 지정 단계 우선 — 미입력 시 월별(T*12) 기본값. 안전 cap 520.
        # K-IFRS 13.IE65: 4모형이 동일 단계·동일 곡선이어야 "모형 구조 차이"가 분리됨.
        if steps_req:
            steps_used = max(1, min(int(steps_req), 520))
        else:
            steps_used = max(int(round(params.T * 12)), 12)

        # 이항 파라미터 (u/d/p) — 프론트 참고용
        dt = params.T / steps_used
        u = float(np.exp(params.volatility * np.sqrt(dt)))
        d = 1.0 / u
        p_rn = float((np.exp((params.risk_free_rate - params.dividend_yield) * dt) - d) / (u - d))

        # ── 선도이자율 커브 (term structure, optional)
        rf_curve = None
        kd_curve = None
        curve_warnings = []
        try:
            rf_spot = data.get("rf_spot")
            rd_spot = data.get("rd_spot")
            if rf_spot and rd_spot:
                from inputs.curves import curve_horizon_check
                rf_chk = curve_horizon_check(rf_spot, params.T)
                rd_chk = curve_horizon_check(rd_spot, params.T)
                if rf_chk and rf_chk.get("warning"):
                    curve_warnings.append("무위험이자율 곡선: " + rf_chk["warning"])
                if rd_chk and rd_chk.get("warning"):
                    curve_warnings.append("이자부부채 곡선: " + rd_chk["warning"])
                rf_curve = _spot_to_step_forwards(rf_spot, params.T, steps_used)
                kd_curve = _spot_to_step_forwards(rd_spot, params.T, steps_used)
        except Exception:
            rf_curve = None
            kd_curve = None

        # ── TF 모형 (메인 1)
        try:
            # GS params 에 주식수가 있으면 TF 희석경로도 활성화
            if data.get("gs_params"):
                gp = data["gs_params"]
                cs = _f(gp.get("common_shares"))
                rs = _f(gp.get("rcps_shares"))
                if cs:
                    params.common_shares = cs
                if rs:
                    params.rcps_shares = rs
            tf_kw = {"bond_discrete": False}  # 연속복리: 연속스팟 부트스트랩과 정합 (C-DF=exp(-z·t))
            if rf_curve and kd_curve:
                tf_kw["rf_curve"] = rf_curve
                tf_kw["kd_curve"] = kd_curve
            tf = tf_rcps(params, steps=steps_used, **tf_kw)
            steps_used = tf["steps"]
        except Exception as e:
            tf = {"error": str(e)}

        # ── GS 모형 (메인 2) — TF와 동일 step·동일 곡선 (모형 구조 차이만 분리)
        try:
            gs_kw = {"bond_discrete": False}
            if data.get("gs_params"):
                gp = data["gs_params"]
                for k in ("enterprise_value", "net_debt", "common_shares", "rcps_shares"):
                    v = _f(gp.get(k))
                    if v:
                        gs_kw[k] = v
            if rf_curve and kd_curve:
                gs_kw["rf_curve"] = rf_curve
                gs_kw["kd_curve"] = kd_curve
            gs = gs_rcps(params, steps=steps_used, **gs_kw)
        except Exception as e:
            gs = {"error": str(e)}

        # ── GS 분해 (한국 평가실무 표준): 채권·풋옵션은 모형 무관 공통값(TF 산출),
        #     GS 전환권 = GS 총액 − 풋채권가치(=채권+풋)
        if (isinstance(gs, dict) and not gs.get("error")
                and isinstance(tf, dict) and not tf.get("error")
                and gs.get("fair_value") is not None
                and tf.get("put_bond_value") is not None):
            pbv = tf["put_bond_value"]
            gs["bond_value"] = tf.get("bond_value")
            gs["put_bond_value"] = pbv
            gs["put_option_value"] = tf.get("put_option_value")
            gs["conversion_value"] = round(gs["fair_value"] - pbv)

        # ── 몬테카를로 — TF·GS와 동일 step·동일 곡선 (모형 비교의 분리 원칙)
        try:
            n_paths = int(data.get("mc_paths", 10000))
            mc = monte_carlo_rcps(params, n_paths=n_paths, n_steps=steps_used,
                                  rf_curve=rf_curve, kd_curve=kd_curve)
        except Exception as e:
            mc = {"error": str(e)}

        # ── requires_mc 라우팅 (경로의존 옵션 활성 시): 트리(TF/GS) 값은 구조적으로
        #     과소평가 → 응답에서 마스킹하여 사용자가 잘못 인용하지 않도록 차단.
        #     fair_value만 null 처리하고 진단용 필드는 보존(분해·트리는 참고용으로만).
        if params.requires_mc:
            _mc_required_marker = {
                "mc_required": True,
                "reason": params.mc_only_reason,
                "note": "경로의존 옵션(래칫 리픽싱·VWAP·소프트콜·강제전환·배리어 풋 등) "
                        "활성. recombining 트리(TF/GS)는 구조적 과소평가 → MC 결과 사용 권장."
            }
            if isinstance(tf, dict) and not tf.get("error"):
                tf.update(_mc_required_marker)
                tf["fair_value"] = None
            if isinstance(gs, dict) and not gs.get("error"):
                gs.update(_mc_required_marker)
                gs["fair_value"] = None

        # ── 민감도 (TF 기준) — 본 평가와 동일한 step·동일 곡선 사용 (base fv 정확 일치)
        # K-IFRS 13.93(h)(ii): 민감도 base와 본 평가는 동일 가정 기반이어야 의미
        try:
            sens = sensitivity_analysis(
                params, steps=steps_used,
                rf_curve=rf_curve, kd_curve=kd_curve)
        except Exception as e:
            sens = {"error": str(e)}

        result = {
            "tf": tf,
            "gs": gs,
            "mc": mc,
            "sensitivity": sens,
            "steps": steps_used,
            "time_to_maturity": round(params.T, 4),
            "u": round(u, 6),
            "d": round(d, 6),
            "risk_neutral_prob": round(p_rn, 6),
            "term_structure_applied": bool(rf_curve and kd_curve),
            # 경로의존(래칫 리픽싱 등) 활성 시 트리(TF/GS)는 과소반영 → MC 전용 평가로 라우팅
            "mc_only": bool(params.requires_mc),
            "mc_only_reason": params.mc_only_reason,
        }

        # 곡선 만기 초과 경고 (있을 때만)
        if curve_warnings:
            result["curve_warnings"] = curve_warnings

        return jsonify({
            "status": "ok",
            "result": serialize(result),
        })
    except Exception as e:
        print(f"[/api/evaluate] {e}\n{traceback.format_exc()}", flush=True)
        return jsonify({"status": "error", "message": str(e)}), 400


@app.route("/api/tree", methods=["POST"])
def tree():
    try:
        data = request.json
        params = parse_params(data["params"])

        # propagate share counts from gs_params (same as /api/evaluate)
        if data.get("gs_params"):
            gp = data["gs_params"]
            cs = _f(gp.get("common_shares"))
            rs = _f(gp.get("rcps_shares"))
            if cs:
                params.common_shares = cs
            if rs:
                params.rcps_shares = rs

        # tree step — 메인 평가(/api/evaluate)와 동일 default 사용
        # (caller-supplied 우선, 미입력 시 max(round(T*12), 12) cap 520)
        tree_steps = data.get("steps")
        if tree_steps is not None:
            try:
                tree_steps = max(1, min(int(tree_steps), 520))
            except (TypeError, ValueError):
                tree_steps = None
        if not tree_steps:
            tree_steps = max(int(round(params.T * 12)), 12)
            tree_steps = max(1, min(tree_steps, 520))

        # ── 선도이자율 커브 (term structure, optional)
        tree_rf_curve = None
        tree_kd_curve = None
        try:
            rf_spot = data.get("rf_spot")
            rd_spot = data.get("rd_spot")
            if rf_spot and rd_spot:
                tree_rf_curve = _spot_to_step_forwards(rf_spot, params.T, tree_steps)
                tree_kd_curve = _spot_to_step_forwards(rd_spot, params.T, tree_steps)
        except Exception:
            tree_rf_curve = None
            tree_kd_curve = None

        # TF tree
        try:
            tf_kw = {"bond_discrete": False}
            if tree_rf_curve and tree_kd_curve:
                tf_kw["rf_curve"] = tree_rf_curve
                tf_kw["kd_curve"] = tree_kd_curve
            tf_res = tf_rcps(params, steps=tree_steps, collect_tree=True, **tf_kw)
            tf_tree = tf_res.get("tree")
            tf_fv   = tf_res.get("fair_value")
        except Exception as e:
            tf_tree = {"error": str(e)}
            tf_fv   = None

        # GS tree
        try:
            gs_kw = {"bond_discrete": False}
            if data.get("gs_params"):
                gp = data["gs_params"]
                for k in ("enterprise_value", "net_debt", "common_shares", "rcps_shares"):
                    v = _f(gp.get(k))
                    if v:
                        gs_kw[k] = v
            if tree_rf_curve and tree_kd_curve:
                gs_kw["rf_curve"] = tree_rf_curve
                gs_kw["kd_curve"] = tree_kd_curve
            gs_res = gs_rcps(params, steps=tree_steps, collect_tree=True, **gs_kw)
            gs_tree = gs_res.get("tree")
            gs_fv   = gs_res.get("fair_value")
        except Exception as e:
            gs_tree = {"error": str(e)}
            gs_fv   = None

        return jsonify({
            "status": "ok",
            "steps":  tree_steps,
            "tf":     tf_tree,
            "gs":     gs_tree,
            "tf_fv":  tf_fv,
            "gs_fv":  gs_fv,
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 400


@app.route("/api/download", methods=["POST"])
def download():
    try:
        data = request.json
        params = parse_params(data["params"])

        # propagate share counts from gs_params (same as /api/evaluate, /api/tree)
        # — 트리 1주(1RCPS) 기준 표시·희석경로에 필요
        if data.get("gs_params"):
            gp = data["gs_params"]
            cs = _f(gp.get("common_shares"))
            rs = _f(gp.get("rcps_shares"))
            if cs:
                params.common_shares = cs
            if rs:
                params.rcps_shares = rs

        # ── 메인 평가 step: 화면(/api/evaluate)과 동일 사용자 선택값 (K-IFRS 13.91 재현성)
        # 이전: tf_rcps(params) default 사용 → 화면 240 step 선택했어도 download는 default
        steps_req_dl = int(data.get("steps", 0)) or None
        if steps_req_dl:
            dl_steps_used = max(1, min(int(steps_req_dl), 520))
        else:
            dl_steps_used = max(int(round(params.T * 12)), 12)

        # ── 선도이자율 커브 (term structure, optional)
        dl_rf_curve = None
        dl_kd_curve = None
        try:
            rf_spot = data.get("rf_spot")
            rd_spot = data.get("rd_spot")
            if rf_spot and rd_spot:
                dl_rf_curve = _spot_to_step_forwards(rf_spot, params.T, dl_steps_used)
                dl_kd_curve = _spot_to_step_forwards(rd_spot, params.T, dl_steps_used)
        except Exception:
            dl_rf_curve = None
            dl_kd_curve = None
        dl_kw = {"bond_discrete": False}
        if dl_rf_curve and dl_kd_curve:
            dl_kw["rf_curve"] = dl_rf_curve
            dl_kw["kd_curve"] = dl_kd_curve
        result = tf_rcps(params, steps=dl_steps_used, **dl_kw)
        sens = sensitivity_analysis(params, steps=dl_steps_used,
                                    rf_curve=dl_rf_curve, kd_curve=dl_kd_curve)

        # 이항 파라미터 (u/d/p) — 조서 모델정보용
        steps_r = result["steps"]
        dt = params.T / steps_r if steps_r else params.T
        u = float(np.exp(params.volatility * np.sqrt(dt)))
        d = 1.0 / u
        p_rn = float((np.exp((params.risk_free_rate - params.dividend_yield) * dt) - d) / (u - d))

        initial_adapted = {
            "fair_value": result["fair_value"],
            "straight_bond_value": result["bond_component"],
            "conversion_component": result["equity_component"],
            "model": result["model"],
            "steps": result["steps"],
            "binomial_detail": {
                "risk_neutral_prob": p_rn,
                "discount_rate": params.discount_rate,
                "u": u,
                "d": d,
            },
            "key_inputs": {
                "time_to_maturity": params.T,
                "stock_price": params.stock_price,
                "volatility": params.volatility,
                "risk_free_rate": params.risk_free_rate,
                "credit_spread": params.credit_spread,
            },
        }

        # 이항트리 수집 (TF + GS) — 감사조서에 첨부
        # 웹 화면 /api/tree 와 동일 로직: 월별, 120 cap (수동 지정 시 최대 520)
        ts_raw = data.get("tree_steps")
        try:
            tree_steps_dl = max(1, min(int(ts_raw), 520)) if ts_raw else None
        except (TypeError, ValueError):
            tree_steps_dl = None
        if not tree_steps_dl:
            tree_steps_dl = max(1, min(round(params.T * 12), 120))

        # 트리 전용 커브 — 웹 /api/tree 와 동일하게 tree_steps_dl 길이로 재구성
        tree_dl_kw = {"bond_discrete": False}
        try:
            rf_spot = data.get("rf_spot")
            rd_spot = data.get("rd_spot")
            if rf_spot and rd_spot:
                tree_dl_kw["rf_curve"] = _spot_to_step_forwards(rf_spot, params.T, tree_steps_dl)
                tree_dl_kw["kd_curve"] = _spot_to_step_forwards(rd_spot, params.T, tree_steps_dl)
        except Exception:
            pass
        try:
            tf_tree_res = tf_rcps(params, steps=tree_steps_dl, collect_tree=True, **tree_dl_kw)
            tf_tree_dl = tf_tree_res.get("tree")
        except Exception:
            tf_tree_dl = None
        try:
            gs_kw_dl = dict(tree_dl_kw)
            if data.get("gs_params"):
                gp = data["gs_params"]
                for k in ("enterprise_value", "net_debt", "common_shares", "rcps_shares"):
                    v = _f(gp.get(k))
                    if v:
                        gs_kw_dl[k] = v
            gs_tree_res = gs_rcps(params, steps=tree_steps_dl, collect_tree=True, **gs_kw_dl)
            gs_tree_dl = gs_tree_res.get("tree")
        except Exception:
            gs_tree_dl = None

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        generate_workpaper(params, initial_adapted, sens, tmp.name,
                           tf_tree=tf_tree_dl, gs_tree=gs_tree_dl,
                           eval_result=data.get("eval_result"),
                           bdt_cross=data.get("bdt_cross"))
        return send_file(tmp.name, as_attachment=True,
                         download_name="RCPS_감사조서.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        print(f"[/api/download] {e}\n{traceback.format_exc()}", flush=True)
        return jsonify({"status": "error", "message": str(e)}), 400


# ── ECOS 국채/회사채 수익률 ──────────────────────────────────────────
ECOS_BASE = "https://ecos.bok.or.kr/api"

def _ecos_get(path, timeout=12):
    with urllib.request.urlopen(f"{ECOS_BASE}/{path}", timeout=timeout) as r:
        return json.loads(r.read())

# 817Y002 (시장금리 일별) 항목 — 코드는 ECOS debug로 확인된 값
_ECOS_ITEMS = [
    {"key": "KTB_1Y",    "code": "010190000", "name": "국고채(1년)",           "maturity": 1,  "category": "국고채"},
    {"key": "KTB_2Y",    "code": "010195000", "name": "국고채(2년)",           "maturity": 2,  "category": "국고채"},
    {"key": "KTB_3Y",    "code": "010200000", "name": "국고채(3년)",           "maturity": 3,  "category": "국고채"},
    {"key": "KTB_5Y",    "code": "010200001", "name": "국고채(5년)",           "maturity": 5,  "category": "국고채"},
    {"key": "KTB_10Y",   "code": "010210000", "name": "국고채(10년)",          "maturity": 10, "category": "국고채"},
    {"key": "KTB_20Y",   "code": "010220000", "name": "국고채(20년)",          "maturity": 20, "category": "국고채"},
    {"key": "KTB_30Y",   "code": "010230000", "name": "국고채(30년)",          "maturity": 30, "category": "국고채"},
    {"key": "CB_AA-",    "code": "010300000", "name": "회사채(3년, AA-)",      "maturity": 3,  "category": "회사채 AA-"},
    {"key": "CB_AA-_MP", "code": "010310000", "name": "회사채(3년, AA-, 민평)", "maturity": 3,  "category": "회사채 AA-(민평)"},
    {"key": "CB_BBB-",   "code": "010320000", "name": "회사채(3년, BBB-)",     "maturity": 3,  "category": "회사채 BBB-"},
    {"key": "MSB_1Y",    "code": "010400001", "name": "통안증권(1년)",          "maturity": 1,  "category": "통안증권"},
    {"key": "MSB_2Y",    "code": "010400002", "name": "통안증권(2년)",          "maturity": 2,  "category": "통안증권"},
]

@app.route("/api/rates/ecos/debug")
def ecos_debug():
    """ECOS 테이블/항목 원본 확인용 — 개발 전용"""
    api_key = request.args.get("key", "").strip()
    if not api_key:
        return jsonify({"error": "key 필요"}), 400
    try:
        # 채권 관련 테이블 목록
        td = _ecos_get(f"StatisticTableList/{api_key}/json/kr/1/500/")
        tables = td.get("StatisticTableList", {}).get("row", [])
        bond_tables = [t for t in tables if any(k in t.get("STAT_NAME","") for k in ["채권","금리","수익률"])]

        # 722Y001 항목 원본
        id1 = _ecos_get(f"StatisticItemList/{api_key}/json/kr/1/500/722Y001")
        items_722 = id1.get("StatisticItemList", {}).get("row", [])

        # 817Y002 항목 원본
        id2 = _ecos_get(f"StatisticItemList/{api_key}/json/kr/1/500/817Y002")
        items_817 = id2.get("StatisticItemList", {}).get("row", [])

        return jsonify({
            "bond_tables": bond_tables[:30],
            "items_722Y001_first30": [{"code": r.get("ITEM_CODE"), "name": r.get("ITEM_NAME")} for r in items_722[:30]],
            "items_817Y002_first30": [{"code": r.get("ITEM_CODE"), "name": r.get("ITEM_NAME")} for r in items_817[:30]],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/rates/ecos")
def get_ecos_rates():
    api_key = request.args.get("key", "").strip()
    target_date = request.args.get("date", "").strip()
    if not api_key:
        return jsonify({"status": "error", "message": "ECOS API 키가 필요합니다."}), 400
    try:
        # 최근 45일 범위에서 최신값 조회 (817Y002 = 시장금리 일별)
        if target_date:
            end_d = target_date
            start_d = (datetime.strptime(target_date, "%Y%m%d") - timedelta(days=45)).strftime("%Y%m%d")
        else:
            end_d = datetime.now().strftime("%Y%m%d")
            start_d = (datetime.now() - timedelta(days=45)).strftime("%Y%m%d")

        result = []
        for item in _ECOS_ITEMS:
            try:
                d2 = _ecos_get(
                    f"StatisticSearch/{api_key}/json/kr/1/45/817Y002/D/{start_d}/{end_d}/{item['code']}"
                )
                r2 = d2.get("StatisticSearch", {}).get("row", [])
                valid = [x for x in r2 if x.get("DATA_VALUE", "").strip() not in ("", "-")]
                if not valid:
                    continue
                latest = max(valid, key=lambda x: x["TIME"])
                result.append({
                    "key": item["key"],
                    "name": item["name"],
                    "maturity": item["maturity"],
                    "category": item["category"],
                    "yield": round(float(latest["DATA_VALUE"]) / 100, 6),
                    "yield_pct": round(float(latest["DATA_VALUE"]), 4),
                    "date": latest["TIME"],
                })
            except Exception:
                continue

        if not result:
            return jsonify({"status": "error", "message": "수익률 데이터를 가져오지 못했습니다. API 키 또는 조회 날짜를 확인하세요."}), 400

        return jsonify({"status": "ok", "rates": result})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 400


# ── KIS-Net 민평표 파일 업로드 ────────────────────────────────────────
# 기본 만기 토큰: 3M, 6M, 1Y, 3Y, 3개월, 6개월, 1년, 1.5년 등
_MATURITY_RE = re.compile(
    r'^\s*(\d+(?:\.\d+)?)\s*([MyYmM]|개월|년)\s*$'
)
# 복합 토큰: 1Y6M 형식 (선택적)
_MATURITY_COMPOUND_RE = re.compile(
    r'^\s*(\d+)\s*[Yy년]\s*(\d+)\s*[Mm개월]\s*$'
)


def _parse_mat_token(tok):
    """만기 토큰 → 년 단위 float. 인식 불가이면 None."""
    if tok is None:
        return None
    s = str(tok).strip()
    # 복합형: 1Y6M
    mc = _MATURITY_COMPOUND_RE.match(s)
    if mc:
        return float(mc.group(1)) + float(mc.group(2)) / 12.0
    m = _MATURITY_RE.match(s)
    if not m:
        return None
    val, unit = float(m.group(1)), m.group(2)
    if unit in ('M', 'm', '개월'):
        return val / 12.0
    return val  # Y / y / 년


def _cell_to_float(v):
    """셀 값 → float. 결측·비숫자이면 None."""
    if v is None:
        return None
    sv = str(v).strip().replace(',', '')  # 천단위 콤마 제거
    if sv in ('', '-', 'N/A', 'n/a', 'NA', '—', '–'):
        return None
    try:
        return float(sv)
    except (ValueError, TypeError):
        return None


def _run_row_header_logic(rows):
    """
    행-헤더 방향: 만기 토큰 >=4개인 헤더 행을 찾아
    [{"category","subtype","label","points"}] 반환.
    못 찾으면 빈 리스트.
    """
    if not rows:
        return []

    # 1. 헤더 행 탐색: 만기 토큰 >=4개인 첫 행
    header_idx = None
    mat_cols = {}  # col_index -> years
    for ri, row in enumerate(rows):
        hits = {}
        for ci, cell in enumerate(row):
            yr = _parse_mat_token(cell)
            if yr is not None:
                hits[ci] = yr
        if len(hits) >= 4:
            header_idx = ri
            mat_cols = hits
            break

    if header_idx is None or not mat_cols:
        return []

    first_mat_col = min(mat_cols)
    label_cols = list(range(first_mat_col))  # 0..N-1 레이블 열

    # 2. 레이블 열 forward-fill
    ff = [None] * max(len(label_cols), 1)
    data_rows = []
    for row in rows[header_idx + 1:]:
        padded = list(row) + [None] * max(0, first_mat_col - len(row))
        label_vals = []
        for li, ci in enumerate(label_cols):
            cell_val = padded[ci] if ci < len(padded) else None
            sv = str(cell_val).strip() if cell_val is not None else ''
            if sv and sv not in ('-',):
                ff[li] = sv
            label_vals.append(ff[li])
        data_rows.append((label_vals, row))

    # 3. 각 데이터 행에서 포인트 수집
    series = []
    for label_vals, row in data_rows:
        if not label_vals or label_vals[-1] is None:
            continue
        points = []
        for ci, yr in mat_cols.items():
            fv = _cell_to_float(row[ci] if ci < len(row) else None)
            if fv is not None:
                points.append({"t": yr, "y": fv})
        if not points:
            continue
        points.sort(key=lambda p: p["t"])
        category = label_vals[0] if len(label_vals) > 0 else None
        subtype  = label_vals[1] if len(label_vals) > 1 else None
        label    = label_vals[-1]
        series.append({
            "category": category,
            "subtype":  subtype,
            "label":    label,
            "points":   points,
        })
    return series


def _transpose_grid(rows):
    """그리드 전치: 짧은 행은 None으로 패딩 후 zip."""
    if not rows:
        return []
    max_cols = max(len(r) for r in rows)
    padded = [list(r) + [None] * (max_cols - len(r)) for r in rows]
    return [list(col) for col in zip(*padded)]


def _extract_series_from_grid(grid):
    """
    행-헤더 방향 시도 후 series < 1이면 전치 후 재시도.
    두 방향 모두 실패하면 빈 리스트.
    """
    rows = [list(r) for r in grid]
    series = _run_row_header_logic(rows)
    if len(series) < 1:
        series = _run_row_header_logic(_transpose_grid(rows))
    return series


def _merge_series(existing, new_series):
    """
    (label, subtype) 키로 중복 제거: 포인트 수가 더 많은 쪽 유지.
    """
    index = {}
    for s in existing:
        key = (s.get('label'), s.get('subtype'))
        if key not in index or len(s['points']) > len(index[key]['points']):
            index[key] = s
    for s in new_series:
        key = (s.get('label'), s.get('subtype'))
        if key not in index or len(s['points']) > len(index[key]['points']):
            index[key] = s
    return list(index.values())


def _parse_minpyung_grid(grid):
    """
    2D list(grid) → [{"category","subtype","label","points"}]
    grid는 openpyxl iter_rows(values_only=True) 또는 csv.reader 결과.
    하위 호환: _extract_series_from_grid 에 위임.
    """
    return _extract_series_from_grid(grid)


@app.route("/api/rates/upload", methods=["POST"])
def upload_minpyung():
    try:
        if 'file' not in request.files:
            return jsonify({"status": "error", "message": "파일이 없습니다."}), 400
        f = request.files['file']
        filename = (f.filename or '').lower()
        raw = f.read()

        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
            ws = wb.worksheets[0]
            grid = list(ws.iter_rows(values_only=True))
            series = _extract_series_from_grid(grid)

        elif filename.endswith('.csv'):
            try:
                text = raw.decode('utf-8')
            except UnicodeDecodeError:
                text = raw.decode('cp949', errors='replace')
            grid = list(csv.reader(io.StringIO(text)))
            series = _extract_series_from_grid(grid)

        elif filename.endswith('.pdf'):
            import pdfplumber  # lazy import — PDF 요청에만 비용 발생
            series = []
            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                for page in pdf.pages:
                    page_series = []
                    # 1차: extract_tables() → 각 테이블을 그리드로 처리
                    try:
                        tables = page.extract_tables() or []
                    except Exception:
                        tables = []
                    for tbl in tables:
                        if tbl:
                            found = _extract_series_from_grid(tbl)
                            page_series = _merge_series(page_series, found)
                    # 2차: 테이블에서 못 찾은 경우 텍스트 → 공백 분리 그리드
                    if not page_series:
                        try:
                            txt = page.extract_text() or ''
                        except Exception:
                            txt = ''
                        if txt.strip():
                            text_grid = [
                                re.split(r'\s{2,}', line)
                                for line in txt.splitlines()
                                if line.strip()
                            ]
                            page_series = _extract_series_from_grid(text_grid)
                    series = _merge_series(series, page_series)

        else:
            return jsonify({
                "status": "error",
                "message": "xlsx, xls, csv, pdf 파일만 지원합니다.",
            }), 400

        if not series:
            return jsonify({
                "status": "error",
                "message": "표 형식을 인식하지 못했습니다. 헤더에 3M/1Y/3Y 등 만기 열이 있는지 확인하세요.",
            }), 400

        return jsonify({"status": "ok", "series": series})
    except Exception:
        return jsonify({
            "status": "error",
            "message": "파일 처리 중 오류가 발생했습니다. 파일 형식 및 만기 헤더(3M/1Y/3Y 등)를 확인하세요.",
        }), 400


def _extract_contract_text(filename, raw):
    """계약서 파일에서 전체 텍스트 추출 (PDF: pdfplumber, xlsx: openpyxl)."""
    fn = (filename or '').lower()
    if fn.endswith('.pdf'):
        import pdfplumber
        parts = []
        with pdfplumber.open(io.BytesIO(raw)) as pdf:
            for page in pdf.pages:
                try:
                    t = page.extract_text() or ''
                except Exception:
                    t = ''
                if t.strip():
                    parts.append(t)
        return "\n".join(parts)
    if fn.endswith('.xlsx') or fn.endswith('.xls'):
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        parts = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c not in (None, "")]
                if cells:
                    parts.append("\t".join(cells))
        return "\n".join(parts)
    if fn.endswith('.csv') or fn.endswith('.txt'):
        try:
            return raw.decode('utf-8')
        except UnicodeDecodeError:
            return raw.decode('cp949', errors='replace')
    raise ValueError("지원하지 않는 파일 형식")


_CONTRACT_PROMPT = """당신은 한국 상환전환우선주(RCPS) 신주인수계약서를 분석하는 회계·법무 전문가입니다.
아래 계약서 전문을 읽고, 다음 JSON 스키마에 맞춰 **오직 JSON만** 출력하세요(코드블록·설명 금지).
모든 금액은 숫자(콤마 없이), 날짜는 YYYY-MM-DD.

{
  "section1_발행조건": {
    "종류": "", "우선주의_종류": "", "우선주_의결권": "",
    "발행일": "", "주식수": 0, "주당발행금액": 0, "총발행금액": 0, "액면가액": 0,
    "존속기간": "", "전환비율": "", "전환가액조정": [],
    "전환청구기간": "", "상환청구기간": "", "상환가액": "", "우선배당률": ""
  },
  "valuation_inputs": {
    "issue_date": "", "maturity_date": "", "face_value": 0, "rcps_shares": 0,
    "conversion_price": 0, "coupon_rate": 0.0, "put_irr": 0.0,
    "conversion_start": "", "conversion_ratio": 1.0,
    "refixing": false, "refixing_floor": 0.0, "refixing_trigger": 1.0, "refixing_frequency": "continuous",
    "put_start": "", "put_price_mode": "irr_y",
    "coupon_frequency": "annual",
    "dividend_cumulative": true, "dividend_first_pay_year": 0,
    "call_start": "", "call_irr": 0.0, "call_price_mode": "cp_y",
    "mandatory_conv_barrier": 0.0, "mandatory_conv_window": 1, "mandatory_conv_count": 1
  },
  "analysis_sections": [
    {
      "title": "섹션 제목",
      "type": "table | list | text",
      "columns": ["표일 때 열 제목들"],
      "rows": [["표일 때 행 데이터"]],
      "items": ["목록일 때 항목들"],
      "body": "서술형일 때 본문",
      "note": "보조 설명(선택)"
    }
  ],
  "담보제공자산": ""
}

규칙:
- section1_발행조건 과 valuation_inputs 는 **모든 RCPS 공통 항목**이므로 위 고정 스키마대로 정확히 추출하세요.
  valuation_inputs 매핑(혼동 주의):
    · face_value = 총발행금액(= 주식수 × 주당발행금액), rcps_shares = 우선주 주식수, conversion_price = 전환가액(원/주)
    · coupon_rate = **우선배당률**(소수, 예 2%→0.02)
    · put_irr = **상환/풋옵션 보장수익률(IRR)**(소수, 예 7.5%→0.075)
    · ★ coupon_rate(우선배당률)와 put_irr(보장수익률)은 서로 다른 항목이니 **절대 바꿔 넣지 마세요**.
    · issue_date = 발행일(거래종결일 다음날이면 그 날짜), maturity_date = 만기일, conversion_start = 전환청구 시작일(발행일+6개월 등).
    · refixing = 전환가액 조정(리픽싱) 조항이 있으면 true (저가발행·IPO 하향조정·시가연동 등)
    · refixing_floor = 전환가 조정 하한을 **최초 전환가액 대비 비율**(0~1)로. 하한이 액면가면 액면가÷전환가액, "최초 전환가의 70%"면 0.70
    · refixing_trigger = 주가/발행가가 전환가의 이 비율 미만일 때 조정(0~1). 명시 트리거 없으면 0.90
    · refixing_frequency = 조정 주기 ("continuous"|"quarterly"|"semi-annual"|"annual"). 명시 없으면 "continuous"
    · put_start = 풋 행사 가능 시작일 (예: 발행일+5년 또는 분기말). 명시 없으면 issue_date+coupon_first_pay_year×1년 근사.
    · put_price_mode = 풋 행사가 산정 방식. 기본 "irr_y"(연복리 IRR make-whole). 약정에 "단리"면 "sp", "복리(쿠폰 미차감)"면 "cp_y".
    · coupon_frequency = 우선배당 지급주기 ("annual"|"semi-annual"|"quarterly"|"none"). 명시 없으면 "annual".
    · dividend_cumulative = 누적적 우선배당이면 true (대부분 한국 RCPS). "참가적·누적적" 표현 있으면 true.
    · dividend_first_pay_year = 첫 우선배당 지급일(발행 후 N년). 명시 "발행 5년 후 일괄"이면 5. 명시 없으면 0(매년).
    · call_start, call_irr, call_price_mode = 발행자 콜(상환권). put과 동일 컨벤션. 없으면 0/빈값.
    · mandatory_conv_barrier = 강제전환(Knock-out) 배리어 (전환가 배수, 예: 1.5 = 전환가의 150%). 없으면 0.
    · mandatory_conv_window, mandatory_conv_count = 강제전환 관측창·충족횟수 (Parisian 배리어). 단순 배리어면 둘 다 1.
- **analysis_sections 는 당신이 자율적으로 구성**하세요. RCPS마다 상장 전/후 구조·상환/풋옵션 조건·사전동의·약정·담보 등이
  모두 다르므로, 고정된 틀을 따르지 말고 **이 계약서에 실제로 존재하는 내용**만 골라 가장 적합한 섹션 구분과 표현
  (표/목록/서술)을 직접 결정하세요. 섹션 개수·제목·형식은 계약 내용에 맞춰 자유롭게 정하면 됩니다.
  예시 주제(있는 것만): 상장 전/후 권리구조, 상환사유·방법, 전환가액 조정(리픽싱), 풋/콜옵션 종류별 조건(IRR 등),
  사전동의·통지 사항, 진술·보장, 손해배상, 기한이익상실, 양도제한, 특이조항 등.
- type 이 "table"이면 columns/rows 만, "list"면 items 만, "text"면 body 만 채우고 나머지는 생략 가능.
- ⚠ **"계약서 외 정보 필요"는 analysis_sections 의 의견 작성 시에만** 사용 (지분율·MOU 비교 등 외부 자료가 진짜 필요할 때).
  section1_발행조건 / valuation_inputs 항목은 계약서 본문·별지·부속서류를 끝까지 뒤져서 채우세요.
  찾을 수 없으면 **빈 문자열 ""** 또는 **0**으로 둘 것 ("계약서 외 정보 필요" 절대 금지).
- **별지(別紙)·부속서·Annex 도 계약의 일부**입니다. 별지 1·2·2.1·2.2 등에 적힌 전환가액 조정공식·우선배당률 세부조건 등을
  반드시 읽고 추출하세요. 본문에 "별지 2.2 참조"라고만 적혀 있어도 그 별지를 찾아 실제 값을 가져오세요.

=== 계약서 전문 ===
"""


# 로컬 소형모델(Ollama)용 — 더 직접적·강제적이고 분석 섹션을 고정·단순화
_CONTRACT_PROMPT_LOCAL = """당신은 한국 상환전환우선주(RCPS) 계약서를 분석하는 회계 전문가입니다.
아래 계약서에서 정보를 추출해 **오직 JSON만** 출력하세요. 설명·코드블록 금지.
금액은 콤마 없는 숫자, 비율은 소수(예 7.5%→0.075), 날짜는 YYYY-MM-DD.
**각 항목을 계약서에서 반드시 찾아 채우세요. 숫자 항목을 빈칸/0 으로 방치하지 말고 본문에서 값을 찾으세요.**

{
  "section1_발행조건": {
    "종류": "", "우선주의_종류": "", "우선주_의결권": "",
    "발행일": "", "주식수": 0, "주당발행금액": 0, "총발행금액": 0, "액면가액": 0,
    "존속기간": "", "전환비율": "", "전환가액조정": [],
    "전환청구기간": "", "상환청구기간": "", "상환가액": "", "우선배당률": ""
  },
  "valuation_inputs": {
    "issue_date": "", "maturity_date": "", "face_value": 0, "rcps_shares": 0,
    "conversion_price": 0, "coupon_rate": 0.0, "put_irr": 0.0,
    "conversion_start": "", "conversion_ratio": 1.0,
    "refixing": false, "refixing_floor": 0.0, "refixing_trigger": 1.0, "refixing_frequency": "continuous",
    "put_start": "", "put_price_mode": "irr_y",
    "coupon_frequency": "annual",
    "dividend_cumulative": true, "dividend_first_pay_year": 0,
    "call_start": "", "call_irr": 0.0, "call_price_mode": "cp_y",
    "mandatory_conv_barrier": 0.0, "mandatory_conv_window": 1, "mandatory_conv_count": 1
  },
  "analysis_sections": [
    {"title": "전환·리픽싱 조건", "type": "list", "items": []},
    {"title": "상환·풋옵션·콜옵션 조건", "type": "list", "items": []},
    {"title": "주요 약정·특이사항", "type": "list", "items": []}
  ],
  "담보제공자산": ""
}

추출 지침(중요):
- 총발행금액 = 주식수 × 주당발행금액. face_value = 총발행금액.
- conversion_price = 전환가액(원/주). coupon_rate = 우선배당률(소수). put_irr = 상환/풋 보장수익률 IRR(소수).
- issue_date=발행일, maturity_date=만기일(존속기간 종료일), conversion_start=전환청구 시작일.
- refixing=전환가액 조정(리픽싱) 조항 있으면 true. refixing_floor=조정 하한÷전환가액(액면가 하한이면 액면가÷전환가액). refixing_trigger=명시 없으면 0.90. refixing_frequency=명시 없으면 "continuous".
- analysis_sections 의 items 는 위 3개 섹션 제목 그대로 두고, 계약서에 있는 해당 내용을 짧은 한국어 문장으로 채우세요.
- 위 JSON 키/구조를 그대로 사용하고, 추가 키는 만들지 마세요.

=== 계약서 전문 ===
"""


def _ocr_pdf(raw: bytes, lang: str = "kor+eng", max_pages: int = 80, dpi: int = 250) -> str:
    """스캔 PDF 로컬 OCR: PyMuPDF로 페이지 렌더 → Tesseract로 텍스트 인식.
    Tesseract 바이너리 필요(미설치 시 예외 → 호출부에서 비전 폴백)."""
    import fitz, pytesseract
    from PIL import Image
    tcmd = os.environ.get("TESSERACT_CMD")
    if not tcmd:
        cand = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        if os.path.exists(cand):
            tcmd = cand
    if tcmd:
        pytesseract.pytesseract.tesseract_cmd = tcmd
    doc = fitz.open(stream=raw, filetype="pdf")
    mat = fitz.Matrix(dpi / 72.0, dpi / 72.0)
    parts = []
    try:
        for i, page in enumerate(doc):
            if i >= max_pages:
                break
            pix = page.get_pixmap(matrix=mat)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            # --psm 6(uniform block): 한국어 인식에 기본 psm 3 보다 안정적
            try:
                txt = pytesseract.image_to_string(img, lang=lang, config="--psm 6")
            except pytesseract.TesseractError:
                txt = pytesseract.image_to_string(img, lang="eng", config="--psm 6")  # 한국어 데이터 없으면 영어
            if txt.strip():
                parts.append(txt)
    finally:
        doc.close()
    return "\n".join(parts)


def _parse_gemini_429(detail: str):
    """Gemini 429 응답에서 (quotaId, retryDelay초) 추출."""
    qid, delay = '', 0
    try:
        ej = json.loads(detail)
        for d in ej.get('error', {}).get('details', []):
            t = d.get('@type', '')
            if t.endswith('QuotaFailure'):
                for v in d.get('violations', []):
                    qid = v.get('quotaId') or v.get('quotaMetric') or qid
            if t.endswith('RetryInfo'):
                m = re.match(r'(\d+)', str(d.get('retryDelay', '')))
                if m:
                    delay = int(m.group(1))
    except Exception:
        pass
    return qid, delay


def _llm_summarize(provider: str, api_key: str, model: str, prompt: str,
                   pdf_bytes: bytes = None) -> str:
    """계약서 요약 LLM 호출 — provider: 'gemini'(무료티어) | 'anthropic'.
    pdf_bytes 가 주어지면 비전(멀티모달)으로 PDF 를 직접 첨부 → 스캔본도 모델이 OCR/이해.
    Gemini 는 추가 의존성 없이 REST(urllib) 로 호출."""
    import base64
    if provider == "gemini":
        m = model if model.startswith("gemini") else "gemini-2.0-flash"
        parts = [{"text": prompt}]
        if pdf_bytes:
            parts.append({"inline_data": {
                "mime_type": "application/pdf",
                "data": base64.b64encode(pdf_bytes).decode("ascii")}})
        url = (f"https://generativelanguage.googleapis.com/v1beta/models/"
               f"{m}:generateContent?key={api_key}")
        body = json.dumps({
            "contents": [{"parts": parts}],
            "generationConfig": {"temperature": 0.2, "maxOutputTokens": 8192},
        }).encode("utf-8")
        import time
        resp = None
        for attempt in range(2):
            req = urllib.request.Request(url, data=body,
                                         headers={"Content-Type": "application/json"})
            try:
                with urllib.request.urlopen(req, timeout=180) as r:
                    resp = json.loads(r.read())
                break
            except urllib.error.HTTPError as he:
                detail = he.read().decode("utf-8", "ignore")
                if he.code == 429:
                    qid, delay = _parse_gemini_429(detail)
                    # 분당 한도(짧은 retryDelay): 자동 1회 대기 후 재시도
                    if attempt == 0 and 0 < delay <= 30:
                        time.sleep(delay + 1)
                        continue
                    # quotaId에 PerDay 포함 시 일일 한도, 그 외 delay로 판별
                    is_daily = bool(qid and 'PerDay' in qid)
                    is_minute = bool(qid and 'PerMinute' in qid)
                    if is_daily:
                        kind = "일일 한도 소진 → PT 자정(한국 오후 5시경) 리셋 / 다른 키·모델 사용 권장"
                    elif is_minute or (delay and delay <= 60):
                        kind = "분당 한도(잠시 후 자동 해소)"
                    else:
                        kind = "한도 초과 → 잠시 후 재시도 또는 다른 키·모델"
                    raise RuntimeError(
                        f"무료 한도 초과(429). 한도종류={qid or '미상'}, 권장대기={delay or '?'}초. {kind}.")
                raise RuntimeError(f"Gemini API {he.code}: {detail[:250]}")
        cands = resp.get("candidates") or []
        if not cands:
            raise RuntimeError(f"Gemini 응답 비어있음: {str(resp.get('promptFeedback'))[:200]}")
        parts_out = (cands[0].get("content", {}) or {}).get("parts") or []
        return "".join(p.get("text", "") for p in parts_out).strip()

    if provider == "groq":
        # Groq (무료·빠름, OpenAI 호환). 비전 모델(llama-4-scout/maverick)이면 PDF → 이미지 변환 후 첨부.
        m = model if model else "llama-3.3-70b-versatile"
        is_vision = bool(pdf_bytes) and ("scout" in m.lower() or "maverick" in m.lower() or "vision" in m.lower())
        # 메시지 구성: 비전이면 이미지 첨부, 아니면 텍스트만
        if is_vision:
            # PDF 페이지 → JPEG(메모리 절약) → base64. Render free 512MB 호환을 위해 페이지별 즉시 해제.
            import fitz, gc  # PyMuPDF
            from PIL import Image
            from io import BytesIO
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            content_parts = [{"type": "text", "text": prompt}]
            max_pages = 3   # 512MB 한도 위해 3장으로 축소 (이전 5)
            dpi = 110       # DPI 110 (이전 150) — 텍스트 인식 충분, 메모리 약 1/2
            mat = fitz.Matrix(dpi/72.0, dpi/72.0)
            try:
                for i, page in enumerate(doc):
                    if i >= max_pages:
                        break
                    pix = page.get_pixmap(matrix=mat)
                    # PNG 대신 JPEG (메모리·파일크기 1/4 수준). 텍스트 인식엔 충분
                    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    buf = BytesIO()
                    img.save(buf, format="JPEG", quality=70, optimize=True)
                    b64 = base64.b64encode(buf.getvalue()).decode("ascii")
                    content_parts.append({
                        "type": "image_url",
                        "image_url": {"url": f"data:image/jpeg;base64,{b64}"}
                    })
                    # 페이지별 메모리 즉시 해제
                    del pix, img, buf
                    gc.collect()
            finally:
                doc.close()
                gc.collect()
            messages = [{"role": "user", "content": content_parts}]
        else:
            messages = [{"role": "user", "content": prompt}]
        body = json.dumps({
            "model": m,
            "messages": messages,
            "temperature": 0.2, "max_tokens": 8000,
        }).encode("utf-8")
        req = urllib.request.Request(
            "https://api.groq.com/openai/v1/chat/completions", data=body,
            headers={"Content-Type": "application/json",
                     "Authorization": f"Bearer {api_key}",
                     "User-Agent": "Mozilla/5.0 (compatible; RCPS-Tool/1.0)"})
        try:
            with urllib.request.urlopen(req, timeout=180) as r:
                resp = json.loads(r.read())
        except urllib.error.HTTPError as he:
            detail = he.read().decode("utf-8", "ignore")[:300]
            raise RuntimeError(f"Groq API {he.code}: {detail}")
        return ((resp.get("choices") or [{}])[0].get("message", {}) or {}).get("content", "").strip()

    if provider == "ollama":
        # 로컬 Ollama (무료·무제한·기밀) — 텍스트 전용. api_key 불필요.
        host = (os.environ.get("OLLAMA_HOST") or "http://localhost:11434").rstrip("/")
        m = model or "qwen2.5:7b"
        # 입력 길이에 맞춰 컨텍스트 설정(과도한 RAM 방지 위해 상한). 초과분은 잘림.
        approx_tokens = len(prompt) // 3 + 1200
        num_ctx = max(8192, min(((approx_tokens // 2048) + 1) * 2048, 32768))
        body = json.dumps({
            "model": m,
            "messages": [{"role": "user", "content": prompt}],
            "stream": False,
            "format": "json",   # 유효 JSON 강제 (로컬 소형모델 JSON 안정화)
            "options": {"temperature": 0.2, "num_ctx": num_ctx},
        }).encode("utf-8")
        req = urllib.request.Request(host + "/api/chat", data=body,
                                     headers={"Content-Type": "application/json"})
        try:
            with urllib.request.urlopen(req, timeout=900) as r:  # CPU 추론 느림 → 최대 15분
                resp = json.loads(r.read())
        except urllib.error.HTTPError as he:
            detail = he.read().decode("utf-8", "ignore")[:250]
            raise RuntimeError(f"Ollama {he.code}: {detail} (모델 설치 확인: ollama pull {m})")
        except urllib.error.URLError as ue:
            raise RuntimeError(f"Ollama 연결 실패: {ue.reason}. 'ollama' 실행 여부와 모델 설치를 확인하세요.")
        return ((resp.get("message") or {}).get("content") or "").strip()

    if provider == "claude_cli":
        # 로컬에 설치된 Claude Code CLI 호출(사용자 Max 구독). 텍스트 전용.
        # 프롬프트는 stdin 으로 전달(대용량 대응), --output-format json 엔벨로프에서 result 추출.
        import subprocess, shutil, tempfile
        cmd = os.environ.get("CLAUDE_CLI_CMD") or shutil.which("claude") or "claude"
        # --allowed-tools "" : 도구 차단(파일읽기 등 에이전트 동작 방지)
        # cwd=중립 임시폴더 : 프로젝트 CLAUDE.md/메모리 자동로드로 인한 컨텍스트 오염 방지
        args = [cmd, "-p", "--output-format", "json",
                "--allowed-tools", "", "--no-session-persistence"]
        if model:
            args += ["--model", model]  # sonnet | opus | haiku
        try:
            proc = subprocess.run(args, input=prompt, capture_output=True,
                                  text=True, encoding="utf-8", timeout=600,
                                  cwd=tempfile.gettempdir(),
                                  # Windows: claude.exe 콘솔창 깜빡임 방지(무창)
                                  creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0))
        except FileNotFoundError:
            raise RuntimeError("claude CLI를 찾을 수 없습니다. 독립형 Claude Code 설치+로그인 후 PATH 등록 "
                               "(또는 환경변수 CLAUDE_CLI_CMD 에 실행파일 경로 지정).")
        except subprocess.TimeoutExpired:
            raise RuntimeError("claude CLI 응답 시간 초과(10분).")
        if proc.returncode != 0:
            raise RuntimeError(f"claude CLI 오류: {((proc.stderr or proc.stdout) or '')[:250]}")
        out = (proc.stdout or "").strip()
        try:                       # --output-format json → {"result": "...", ...}
            env = json.loads(out)
            if isinstance(env, dict) and env.get("result") is not None:
                return str(env["result"]).strip()
        except Exception:
            pass
        return out

    # anthropic (유료)
    import anthropic
    client = anthropic.Anthropic(api_key=api_key)
    m = model if model.startswith("claude") else "claude-sonnet-4-6"
    if pdf_bytes:
        content = [
            {"type": "document", "source": {
                "type": "base64", "media_type": "application/pdf",
                "data": base64.b64encode(pdf_bytes).decode("ascii")}},
            {"type": "text", "text": prompt},
        ]
    else:
        content = prompt
    msg = client.messages.create(
        model=m, max_tokens=8000,
        messages=[{"role": "user", "content": content}],
    )
    return "".join(b.text for b in msg.content if getattr(b, "type", "") == "text").strip()


@app.route("/api/contract/parse", methods=["POST"])
def contract_parse():
    try:
        if 'file' not in request.files:
            return jsonify({"status": "error", "message": "파일이 없습니다."}), 200
        f = request.files['file']
        raw = f.read()
        api_key = (request.form.get('api_key') or '').strip()
        provider = (request.form.get('provider') or '').strip().lower()
        model = (request.form.get('model') or '').strip()
        if not provider:
            if api_key.startswith('sk-ant'):
                provider = 'anthropic'
            elif api_key.startswith('gsk_'):
                provider = 'groq'
            else:
                provider = 'gemini'
        if provider in ('ollama', 'claude_cli'):
            api_key = api_key or 'local'  # 로컬 Ollama / Claude CLI(구독)는 키 불필요
        if not api_key:
            _env = {'gemini': ('GEMINI_API_KEY', 'GOOGLE_API_KEY'),
                    'groq': ('GROQ_API_KEY',),
                    'anthropic': ('ANTHROPIC_API_KEY',)}.get(provider, ())
            for k in _env:
                if os.environ.get(k):
                    api_key = os.environ[k].strip(); break
        if not api_key:
            nm = {'gemini': 'Gemini', 'groq': 'Groq', 'anthropic': 'Anthropic'}.get(provider, provider)
            return jsonify({"status": "error",
                            "message": f"{nm} API 키가 필요합니다. 키를 입력하세요."}), 200

        fn = (f.filename or '').lower()
        is_pdf = fn.endswith('.pdf')
        try:
            text = _extract_contract_text(f.filename, raw)
        except Exception:
            text = ''
        # 텍스트가 거의 없으면(스캔 이미지 PDF) → 비전으로 PDF 직접 전달(모델이 OCR)
        scanned = is_pdf and len((text or '').strip()) < 100
        # Ollama(로컬)는 num_ctx 상한이 있어 지시문이 잘리지 않게 본문을 더 짧게 cap
        maxchars = 85000 if provider == 'ollama' else 200000
        # 로컬 소형모델은 강제·단순 프롬프트, 클라우드는 자율 분석 프롬프트
        cprompt = _CONTRACT_PROMPT_LOCAL if provider == 'ollama' else _CONTRACT_PROMPT
        try:
            if scanned:
                # 1순위: 로컬 OCR(Tesseract) → 가벼운 텍스트만 클라우드 전송(429 회피)
                ocr_text = ''
                try:
                    ocr_text = _ocr_pdf(raw)
                except Exception:
                    ocr_text = ''
                # Groq 비전 모델 선택 시 OCR 우회하고 직접 이미지 분석 (무료·빠름)
                groq_vision = provider == 'groq' and any(k in (model or '').lower() for k in ('scout','maverick','vision'))
                if ocr_text and len(ocr_text.strip()) >= 100:
                    out = _llm_summarize(provider, api_key, model,
                                         cprompt + ocr_text[:maxchars])
                elif provider in ('gemini', 'anthropic') or groq_vision:
                    # 2순위(비전 지원 모델): 로컬 OCR 불가/실패 → PDF 직접 전달
                    if len(raw) > 18_000_000:
                        return jsonify({"status": "error",
                                        "message": "스캔 PDF가 큽니다(18MB 초과). 페이지 분할 후 재시도하세요."}), 200
                    out = _llm_summarize(provider, api_key, model,
                                         _CONTRACT_PROMPT + "(첨부된 PDF 계약서 이미지를 직접 읽고 분석하세요. 한국어 OCR 수행.)",
                                         pdf_bytes=raw)
                else:
                    # 텍스트 전용 모델: OCR 실패 시 비전 불가
                    return jsonify({"status": "error",
                                    "message": "스캔 PDF인데 OCR 실패. 비전 지원 모델 선택 권장: Gemini, Claude API, 또는 Groq의 Llama 4 Scout/Maverick."}), 200
            else:
                if not text or not text.strip():
                    return jsonify({"status": "error",
                                    "message": "계약서 텍스트가 비어 있습니다. PDF/Excel 형식을 확인하세요."}), 200
                out = _llm_summarize(provider, api_key, model, cprompt + text[:maxchars])
        except Exception as e:
            return jsonify({"status": "error",
                            "message": f"LLM 호출 오류: {str(e)[:250]}"}), 200

        # JSON 추출 (코드블록/잡텍스트 방어)
        m = re.search(r'\{.*\}', out, re.DOTALL)
        if not m:
            return jsonify({"status": "error",
                            "message": "요약 JSON을 파싱하지 못했습니다."}), 200
        try:
            summary = json.loads(m.group(0))
        except Exception:
            return jsonify({"status": "error",
                            "message": "요약 JSON 형식 오류."}), 200

        return jsonify({"status": "ok", "summary": summary})
    except Exception:
        return jsonify({"status": "error",
                        "message": "계약서 처리 중 오류가 발생했습니다."}), 200


@app.route("/api/contract/loadfile", methods=["GET"])
def contract_loadfile():
    """Claude Code(사용자 Max 구독)가 생성한 요약 JSON 파일을 읽어 반환.
    경로: <project>/contract_summary.json — API 키·비용 없이 내 Claude로 요약 → 툴 연동."""
    path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "contract_summary.json")
    if not os.path.exists(path):
        return jsonify({"status": "error",
                        "message": "contract_summary.json 이 없습니다. Claude Code에 '계약서 요약해서 contract_summary.json 으로 저장해줘'라고 요청하세요."}), 200
    try:
        with open(path, encoding="utf-8") as fp:
            summary = json.load(fp)
    except Exception as e:
        return jsonify({"status": "error",
                        "message": f"JSON 파싱 오류: {str(e)[:150]}"}), 200
    return jsonify({"status": "ok", "summary": summary})


# ══════════════════════════════════════════════════════════════════
#  변동성 평가 — 유사상장기업 바스켓 역사적 변동성 (FinanceDataReader)
# ══════════════════════════════════════════════════════════════════
_STOCK_LISTING = None   # KRX + 해외(NASDAQ/NYSE/AMEX) 종목목록 캐시 (프로세스 1회 로드)


def _stock_listing():
    """상장종목 마스터 캐시 로드 — KRX(코드·시총) + 미국 거래소(티커·종목명).

    KRX는 Code/Name/Market/Marcap, 미국은 Symbol/Name 제공(시총 없음→None).
    개별 거래소 조회 실패는 건너뛰어 일부만으로도 동작하게 한다.
    """
    global _STOCK_LISTING
    if _STOCK_LISTING is None:
        import FinanceDataReader as fdr  # lazy import — 변동성 요청에만 비용
        rows, seen = [], set()
        # 1) KRX — FDR이 Render Singapore에서 가끔 실패 → 정적 CSV 폴백
        krx_ok = False
        try:
            for r in fdr.StockListing('KRX').to_dict('records'):
                code = str(r.get('Code') or '').strip()
                if not code or code in seen:
                    continue
                seen.add(code)
                rows.append({"code": code, "name": str(r.get('Name') or '').strip(),
                             "market": str(r.get('Market') or '').strip(),
                             "marcap": r.get('Marcap')})
            if len(rows) > 100:
                krx_ok = True
                print(f"[listing] KRX via FDR: {len(rows)} stocks", flush=True)
        except Exception as e:
            print(f"[listing] KRX FDR failed: {e}", flush=True)
        # KRX FDR 실패 시 정적 CSV로 폴백
        if not krx_ok:
            try:
                csv_path = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                                        'data', 'krx_listing.csv')
                if os.path.exists(csv_path):
                    import csv as _csv
                    with open(csv_path, 'r', encoding='utf-8') as f:
                        cnt = 0
                        for r in _csv.DictReader(f):
                            code = (r.get('Code') or '').strip()
                            if not code or code in seen:
                                continue
                            seen.add(code)
                            marcap = r.get('Marcap')
                            try: marcap = float(marcap) if marcap else None
                            except: marcap = None
                            rows.append({"code": code, "name": (r.get('Name') or '').strip(),
                                         "market": (r.get('Market') or '').strip(),
                                         "marcap": marcap})
                            cnt += 1
                        print(f"[listing] KRX via static CSV: {cnt} stocks", flush=True)
            except Exception as e:
                print(f"[listing] KRX static CSV failed: {e}", flush=True)
        # 2) 미국 거래소 (NASDAQ/NYSE/AMEX) — FDR 우선, 실패 시 정적 CSV
        data_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data')
        import csv as _csv
        for mk, fn in (('NASDAQ', 'nasdaq_listing.csv'),
                       ('NYSE', 'nyse_listing.csv'),
                       ('AMEX', 'amex_listing.csv')):
            mk_ok = False
            mk_start = len(rows)
            try:
                for r in fdr.StockListing(mk).to_dict('records'):
                    sym = str(r.get('Symbol') or r.get('Code') or '').strip()
                    if not sym or sym in seen:
                        continue
                    seen.add(sym)
                    rows.append({"code": sym, "name": str(r.get('Name') or '').strip(),
                                 "market": mk, "marcap": r.get('Marcap')})
                if (len(rows) - mk_start) > 50:
                    mk_ok = True
                    print(f"[listing] {mk} via FDR: {len(rows)-mk_start} stocks", flush=True)
            except Exception as e:
                print(f"[listing] {mk} FDR failed: {e}", flush=True)
            # FDR 실패 시 정적 CSV 폴백
            if not mk_ok:
                csv_path = os.path.join(data_dir, fn)
                try:
                    if os.path.exists(csv_path):
                        cnt = 0
                        with open(csv_path, 'r', encoding='utf-8') as f:
                            for r in _csv.DictReader(f):
                                sym = (r.get('Symbol') or '').strip()
                                if not sym or sym in seen:
                                    continue
                                seen.add(sym)
                                rows.append({"code": sym, "name": (r.get('Name') or '').strip(),
                                             "market": mk, "marcap": None})
                                cnt += 1
                            print(f"[listing] {mk} via static CSV: {cnt} stocks", flush=True)
                except Exception as e:
                    print(f"[listing] {mk} static CSV failed: {e}", flush=True)
        _STOCK_LISTING = rows
    return _STOCK_LISTING


def _yahoo_search(q, want=8):
    """Yahoo Finance 검색으로 회사명→티커 해석(전 세계 거래소). 실패 시 빈 리스트.

    유럽 등 FDR 종목목록에 없는 종목을 회사명으로 찾기 위함. EQUITY만 반환.
    """
    try:
        url = ("https://query1.finance.yahoo.com/v1/finance/search?q="
               + urllib.parse.quote(q) + f"&quotesCount={want}&newsCount=0")
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=8) as r:
            d = json.load(r)
        # 거래데이터가 빈약한 보조상장(Cboe Europe DXE, IOB, OTC 등)은 뒤로 — 본상장 우선
        SECONDARY = {"XD", "IL", "PNK", "OTC", "DXE", "IOB"}
        res = []
        for x in d.get("quotes", []):
            if x.get("quoteType") != "EQUITY":
                continue
            sym = str(x.get("symbol") or "").strip()
            if not sym:
                continue
            suffix = sym.rsplit(".", 1)[1].upper() if "." in sym else ""
            exch = str(x.get("exchange") or "").upper()
            secondary = suffix in SECONDARY or exch in SECONDARY
            res.append({"code": sym,
                        "name": str(x.get("shortname") or x.get("longname") or sym).strip(),
                        "market": str(x.get("exchDisp") or x.get("exchange") or "").strip(),
                        "marcap": None, "_secondary": secondary})
        res.sort(key=lambda r: r.pop("_secondary"))   # 본상장(False) 먼저
        return res[:want]
    except Exception:
        return []


@app.route("/api/volatility/search", methods=["GET"])
def volatility_search():
    """종목명/코드(티커) 자동완성. KRX+미국 로컬 목록 우선, 부족하면 Yahoo로 전 세계 보강."""
    q = (request.args.get('q') or '').strip()
    if not q:
        return jsonify({"status": "ok", "results": []})
    try:
        listing = _stock_listing()
    except Exception as e:
        return jsonify({"status": "error",
                        "message": f"종목목록 조회 실패: {str(e)[:150]}"}), 200
    ql = q.lower()
    out, seen = [], set()
    for r in listing:
        if ql in r["name"].lower() or ql in r["code"].lower():
            out.append(r); seen.add(r["code"])
            if len(out) >= 30:
                break
    # 로컬(KRX+미국)에 충분히 없으면 Yahoo 검색으로 전 세계 티커 보강(유럽 등)
    if len(out) < 5 and len(q) >= 2:
        for r in _yahoo_search(q):
            if r["code"] not in seen:
                out.append(r); seen.add(r["code"])
    # 통화 일관성·관련성 정렬:
    # ① 이름이 query 로 시작하면 강한 가산점 (Toyota Motor Corp이 Toyota Tsusho 앞으로)
    # ② ADR/예탁증서/DR/OTC는 후순위 — 본상장이 BS·시총 동일 통화 보장
    #    예: "Toyota Motor Corp ADR"(TM, NYSE, USD) vs "TOYOTA MOTOR CORP"(7203.T, 도쿄, JPY)
    # 주력 사업 descriptor — 글로벌 대형사가 흔히 갖는 corp suffix 직전 단어
    _PRIMARY = {"motor","motors","industries","industrial","electronics","electric",
                "financial","group","holdings","holding","bank","banking","energy",
                "oil","gas","pharma","pharmaceutical","pharmaceuticals","semiconductor",
                "technology","tech","chemical","chemicals","steel","life","health",
                "healthcare","insurance","aerospace","aviation","airlines","automotive",
                "communications","networks","systems","software","media","entertainment",
                "telecom","mobile","retail","consumer","beverages","foods","food"}
    def _rank(r):
        nm = (r.get("name") or "").lower().strip()
        market = (r.get("market") or "").upper()
        # 감점 (큰 숫자 = 후순위)
        penalty = 0
        if "adr" in nm or " ads " in f" {nm} " or "depositary" in nm or "dep recpt" in nm or "_dr " in nm or nm.endswith("_dr") or " dr " in f" {nm} ":
            penalty += 100
        if market in {"OTC MARKETS", "PINK SHEETS", "PNK", "OTC"}:
            penalty += 50
        # 관련성 가산점 (작은 숫자 = 우선)
        rel = 50
        if nm.startswith(ql):           # 정확 시작 매칭
            rel = 0
        elif (" " + ql) in (" " + nm):  # 단어 경계 매칭
            rel = 10
        # query 바로 뒤가 주력 사업 descriptor면 대형 본사일 확률↑ (Toyota Motor vs Toyota Tsusho)
        if rel == 0:
            rest = nm[len(ql):].strip()
            first_word = rest.split()[0].rstrip(",.;") if rest else ""
            if first_word in _PRIMARY:
                rel -= 5
        return penalty + rel
    out.sort(key=_rank)
    return jsonify({"status": "ok", "results": out})


import threading
_BS_LOCK = threading.Lock()   # baostock 전역 세션 보호(동시요청 대비)


def _to_baostock(code):
    """중국 본토 A주 코드 → baostock 심볼(sh./sz.). A주가 아니면 None."""
    c = (code or "").strip()
    if c.lower().startswith(("sh.", "sz.")):
        return c.lower()
    u = c.upper()
    if u.endswith(".SS"):
        return "sh." + c[:-3]
    if u.endswith(".SZ"):
        return "sz." + c[:-3]
    return None


def _fetch_china_ashare(bs_symbol, start, end):
    """baostock로 A주 일별 종가·거래량 조회(전복권). (dates, closes, volumes) 반환."""
    import baostock as bs
    with _BS_LOCK:
        lg = bs.login()
        try:
            if lg.error_code != '0':
                raise ValueError(f"baostock 로그인 실패: {lg.error_msg}")
            rs = bs.query_history_k_data_plus(
                bs_symbol, "date,close,volume",
                start_date=(start or "2015-01-01"), end_date=(end or ""),
                frequency="d", adjustflag="2")   # 2=전복권(qfq)
            if rs.error_code != '0':
                raise ValueError(f"baostock 조회 실패: {rs.error_msg}")
            dates, closes, vols = [], [], []
            while rs.next():
                d, c, v = rs.get_row_data()
                if c in (None, ''):
                    continue
                dates.append(d); closes.append(float(c))
                vols.append(float(v) if v not in (None, '') else None)
            return dates, closes, vols
        finally:
            bs.logout()


def _fetch_price_series(code, start, end):
    """코드 → (dates, closes, volumes). 중국 본토 A주는 baostock, 그 외 FDR(Yahoo)."""
    bsym = _to_baostock(code)
    if bsym:
        return _fetch_china_ashare(bsym, start, end)
    import FinanceDataReader as fdr
    df = fdr.DataReader(code, start, end)
    if df is None or df.empty or 'Close' not in df.columns:
        raise ValueError("조회 결과가 비어 있습니다.")
    closes = [float(x) for x in df['Close'].tolist()]
    vols = [float(x) for x in df['Volume'].tolist()] if 'Volume' in df.columns else None
    dates = [d.strftime('%Y-%m-%d') for d in df.index]
    return dates, closes, vols


@app.route("/api/stock_price", methods=["GET"])
def stock_price_on_date():
    """평가대상이 상장사인 경우, 평가기준일(이하 마지막 거래일) 종가 조회 → S₀ 자동입력용."""
    code = (request.args.get('ticker') or '').strip()
    date = (request.args.get('date') or '').strip()
    if not code or not date:
        return jsonify({"status": "error", "message": "종목코드와 평가기준일이 필요합니다."}), 200
    try:
        d = _d(date)
    except Exception:
        return jsonify({"status": "error", "message": "날짜 형식 오류(YYYY-MM-DD)."}), 200
    start = (d - timedelta(days=14)).isoformat()
    end = (d + timedelta(days=1)).isoformat()   # 기준일 포함되도록 +1
    try:
        dates, closes, _ = _fetch_price_series(code, start, end)
    except Exception as e:
        return jsonify({"status": "error", "message": f"조회 실패: {str(e)[:150]}"}), 200
    pick = None
    for dt, c in zip(dates, closes):
        if dt <= date:
            pick = (dt, c)
    if not pick:
        return jsonify({"status": "error",
                        "message": "기준일 이전 거래 데이터가 없습니다(상장 전이거나 휴장)."}), 200
    return jsonify({"status": "ok", "ticker": code, "date": pick[0], "close": pick[1]})


@app.route("/api/volatility", methods=["POST"])
def volatility_eval():
    """유사기업 바스켓 역사적 변동성 산출.

    body: {tickers:[code...], start, end, trading_days, method, log}
    각 종목을 FDR로 조회 → 종목별 σ·시총 → 집계 σ. 조회 실패 종목은 스킵.
    원자료(날짜·종가)도 함께 반환(감사 재현성·CSV 내보내기용).
    """
    from models.volatility import basket_volatility
    data = request.get_json(force=True) or {}
    tickers = [str(t).strip() for t in (data.get("tickers") or []) if str(t).strip()]
    if not tickers:
        return jsonify({"status": "error", "message": "종목이 비어 있습니다."}), 200
    start = data.get("start") or None
    end = data.get("end") or None
    td_raw = data.get("trading_days", 252)
    trading_days = "auto" if isinstance(td_raw, str) and td_raw.lower() == "auto" else int(td_raw or 252)
    method = data.get("method") or "median"
    log = data.get("log", True)
    outlier_method = (data.get("outlier_method") or "none").lower()
    outlier_k = data.get("outlier_k")
    outlier_k_mad = data.get("outlier_k_mad")
    if outlier_k is not None:
        try: outlier_k = float(outlier_k)
        except Exception: outlier_k = None
    if outlier_k_mad is not None:
        try: outlier_k_mad = float(outlier_k_mad)
        except Exception: outlier_k_mad = None

    try:
        import FinanceDataReader as fdr
    except Exception as e:
        return jsonify({"status": "error",
                        "message": f"FinanceDataReader 미설치: {str(e)[:120]}"}), 200

    # 종목명·시총 매핑(가능하면)
    name_cap = {}
    try:
        for r in _stock_listing():
            name_cap[r["code"]] = (r["name"], r["marcap"])
    except Exception:
        pass

    series, raw = {}, {}
    for code in tickers:
        try:
            dates, closes, vols = _fetch_price_series(code, start, end)  # A주=baostock, 그 외=FDR
            if len(closes) < 2:
                raise ValueError("조회 결과가 비어 있습니다.")
            nm, cap = name_cap.get(code, (code, None))
            series[code] = {"name": nm, "dates": dates, "closes": closes,
                            "volumes": vols, "cap": cap}
            raw[code] = {"name": nm, "dates": dates, "closes": closes}
        except Exception as e:
            series[code] = {"name": name_cap.get(code, (code, None))[0],
                            "closes": [], "error": str(e)[:150]}

    try:
        result = basket_volatility(series, trading_days=trading_days,
                                   log=bool(log), method=method,
                                   outlier_method=outlier_method, outlier_k=outlier_k,
                                   outlier_k_mad=outlier_k_mad)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)[:200]}), 200

    if result["sigma"] is None:
        reasons = "; ".join(
            f"{p['ticker']}: {p.get('error', '실패')}" for p in result["per_ticker"]
        )[:300]
        return jsonify({
            "status": "error",
            "message": f"유효한 변동성을 산출하지 못했습니다. 종목별 사유 → {reasons}",
            "per_ticker": result["per_ticker"],
        }), 200

    return jsonify({
        "status": "ok",
        "sigma": result["sigma"],
        "sigma_pct": round(result["sigma"] * 100, 2),
        "method": method,
        "trading_days": trading_days,
        "start": start, "end": end,
        "per_ticker": result["per_ticker"],
        "failed": result["failed"],
        "outlier_info": result.get("outlier_info"),
        "basket_trailings": result.get("basket_trailings"),  # 1y/2y/3y/5y σ 비교
        "raw": raw,
    })


@app.route("/api/volatility/upload", methods=["POST"])
def volatility_upload():
    """CSV 업로드로 변동성 산출(자동조회 fallback·재현성용).

    CSV 형식: 1열=날짜, 2열~=종목별 종가(헤더=종목명). 결측은 빈칸.
    """
    from models.volatility import basket_volatility
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "파일이 없습니다."}), 400
    raw = request.files['file'].read()
    try:
        text = raw.decode('utf-8-sig')
    except UnicodeDecodeError:
        text = raw.decode('cp949', errors='replace')
    rows = list(csv.reader(io.StringIO(text)))
    if len(rows) < 3:
        return jsonify({"status": "error", "message": "데이터 행이 부족합니다."}), 200

    header = rows[0]
    cols = header[1:]  # 종목명들
    series = {name.strip() or f"col{i}": {"name": name.strip() or f"col{i}",
                                          "dates": [], "closes": []}
              for i, name in enumerate(cols)}
    keys = list(series.keys())
    td_raw = request.form.get("trading_days") or "252"
    trading_days = "auto" if td_raw.lower() == "auto" else int(td_raw)
    method = request.form.get("method") or "median"

    for r in rows[1:]:
        if not r or not r[0].strip():
            continue
        d = r[0].strip()
        for j, key in enumerate(keys):
            cell = r[j + 1].strip() if j + 1 < len(r) else ""
            if cell == "":
                continue
            try:
                series[key]["closes"].append(float(cell.replace(",", "")))
                series[key]["dates"].append(d)
            except ValueError:
                continue

    try:
        result = basket_volatility(series, trading_days=trading_days, method=method)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)[:200]}), 200

    return jsonify({
        "status": "ok",
        "sigma": result["sigma"],
        "sigma_pct": round(result["sigma"] * 100, 2),
        "method": method, "trading_days": trading_days,
        "per_ticker": result["per_ticker"], "failed": result["failed"],
    })


# ══════════════════════════════════════════════════════════════════
#  계약서 요약 → Excel 다운로드 (inputs/rcps 요약.xlsx 스타일)
# ══════════════════════════════════════════════════════════════════
def _fmt_num_kr(v, suffix=""):
    """숫자면 천단위 콤마 + 단위, 아니면 원본."""
    if v in (None, ""):
        return None
    try:
        n = float(str(v).replace(",", ""))
        body = f"{int(n):,}" if n == int(n) else f"{n:,.2f}"
        return f"{body} {suffix}".strip()
    except (ValueError, TypeError):
        return str(v)


@app.route("/api/contract/export", methods=["POST"])
def contract_export_xlsx():
    """계약서 요약(JSON) → Excel 다운로드. inputs/rcps 요약.xlsx 와 동일한 섹션·행 레이아웃."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = request.get_json(force=True) or {}
    summary = data.get("summary") or data
    if not summary or not isinstance(summary, dict):
        return jsonify({"status": "error", "message": "summary JSON이 필요합니다."}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "계약서 요약"

    F_TITLE = Font(bold=True, size=14)
    F_HEAD = Font(bold=True, size=11, color="FFFFFF")
    F_SUB = Font(bold=True, size=10)
    FILL_HEAD = PatternFill("solid", fgColor="3182F6")
    FILL_KEY = PatternFill("solid", fgColor="F2F4F6")
    thin = Side(style="thin", color="DDDDDD")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    WRAP = Alignment(wrap_text=True, vertical="top")

    row = 1
    ws.cell(row=row, column=2, value="RCPS 계약서 요약").font = F_TITLE
    row += 2

    def section_header(text):
        nonlocal row
        c = ws.cell(row=row, column=2, value=text)
        c.font = F_HEAD; c.fill = FILL_HEAD
        c.alignment = Alignment(vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        ws.row_dimensions[row].height = 22
        row += 1

    def kv(k, v):
        nonlocal row
        kc = ws.cell(row=row, column=2, value=k)
        kc.font = F_SUB; kc.fill = FILL_KEY; kc.alignment = WRAP; kc.border = BORDER
        vc = ws.cell(row=row, column=3, value=v if v not in (None, "") else "—")
        vc.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
        for col in range(2, 8):
            ws.cell(row=row, column=col).border = BORDER
        # 텍스트 길이에 따라 행 높이 조정
        if isinstance(v, str) and ("\n" in v or len(v) > 60):
            ws.row_dimensions[row].height = max(18, min(120, 15 + 14 * (v.count("\n") + len(v) // 60)))
        row += 1

    # ── 1. 발행조건 (고정 키) ──
    s1 = summary.get("section1_발행조건", {}) or {}
    section_header("1. 발행조건")
    adj = s1.get("전환가액조정") or []
    adj_text = "\n".join("· " + str(x) for x in adj) if adj else ""
    conv_ratio = s1.get("전환비율") or ""
    conv_combined = (conv_ratio + ("\n" + adj_text if adj_text else "")).strip()

    for k, v in [
        ("종류", s1.get("종류")),
        ("우선주의 종류", s1.get("우선주의_종류")),
        ("우선주 의결권", s1.get("우선주_의결권")),
        ("발행일", s1.get("발행일")),
        ("우선주 주식수", _fmt_num_kr(s1.get("주식수"), "주")),
        ("주당 발행금액", _fmt_num_kr(s1.get("주당발행금액"), "원")),
        ("총 발행금액", _fmt_num_kr(s1.get("총발행금액"), "원")),
        ("1주당 액면가액", _fmt_num_kr(s1.get("액면가액"), "원")),
        ("존속기간", s1.get("존속기간")),
        ("전환비율 및 종류", conv_combined),
        ("전환청구기간", s1.get("전환청구기간")),
        ("상환청구기간", s1.get("상환청구기간")),
        ("상환가액", s1.get("상환가액")),
        ("우선배당률", s1.get("우선배당률")),
    ]:
        kv(k, v)
    row += 1

    # ── 2+. analysis_sections (AI 자율 구성) ──
    secs = summary.get("analysis_sections") or []
    for i, sec in enumerate(secs):
        section_header(f"{i + 2}. {sec.get('title') or '분석'}")
        st = sec.get("type")
        if st == "table" and isinstance(sec.get("columns"), list) and isinstance(sec.get("rows"), list):
            for j, col in enumerate(sec["columns"]):
                c = ws.cell(row=row, column=2 + j, value=str(col))
                c.font = F_SUB; c.fill = FILL_KEY; c.alignment = WRAP; c.border = BORDER
            row += 1
            for r in sec["rows"]:
                cells = r if isinstance(r, list) else [r]
                for j, cellv in enumerate(cells):
                    c = ws.cell(row=row, column=2 + j, value=("" if cellv is None else str(cellv)))
                    c.alignment = WRAP; c.border = BORDER
                row += 1
        elif st == "list" and isinstance(sec.get("items"), list):
            for it in sec["items"]:
                c = ws.cell(row=row, column=2, value="· " + str(it))
                c.alignment = WRAP
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
                row += 1
        elif sec.get("body"):
            c = ws.cell(row=row, column=2, value=str(sec["body"]))
            c.alignment = WRAP
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
            body = str(sec["body"])
            ws.row_dimensions[row].height = max(20, min(200, 15 + 14 * (body.count("\n") + len(body) // 70)))
            row += 1
        if sec.get("note"):
            c = ws.cell(row=row, column=2, value="※ " + str(sec["note"]))
            c.font = Font(size=9, color="888888"); c.alignment = WRAP
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
            row += 1
        row += 1

    # ── 담보제공자산 (있으면) ──
    if summary.get("담보제공자산"):
        section_header(f"{len(secs) + 2}. 담보제공자산")
        c = ws.cell(row=row, column=2, value=str(summary["담보제공자산"]))
        c.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        row += 1

    # 열 너비
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 24
    for col in range(3, 8):
        ws.column_dimensions[get_column_letter(col)].width = 22

    import tempfile, datetime
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name); tmp.close()
    fname = f"contract_summary_{datetime.date.today().isoformat()}.xlsx"
    return send_file(tmp.name, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ══════════════════════════════════════════════════════════════════
#  BDT 단기금리 이항트리 — 풋채권가치 교차검증
# ══════════════════════════════════════════════════════════════════
@app.route("/api/bdt", methods=["POST"])
def bdt_evaluate():
    """BDT 트리로 풋옵션부 채권가치를 교차평가(TF 결과와 비교용).

    body: /api/evaluate 와 동일 + rate_vol(연 단기금리 변동성, 소수).
    RD(신용조정) 곡선을 BDT 트리에 캘리브레이션 후 후방귀납으로
    풋채권가치(pbv)·일반채권가치(bv)·풋옵션가치(pbv−bv) 산정.
    """
    data = request.get_json(force=True) or {}
    try:
        params = parse_params(data["params"])    # /api/evaluate 와 동일한 body 구조 사용
    except Exception as e:
        return jsonify({"status": "error", "message": f"파라미터 오류: {str(e)[:150]}"}), 200
    try:
        sigma = float(data.get("rate_vol") or 0)
    except (TypeError, ValueError):
        sigma = 0.0
    if sigma <= 0:
        return jsonify({"status": "error",
                        "message": "금리 변동성(rate_vol)이 필요합니다 (예: 0.15 = 15%)."}), 200
    rd_spot = data.get("rd_spot")
    if not rd_spot:
        return jsonify({"status": "error",
                        "message": "RD 곡선(rd_spot)이 필요합니다. 이자율 부트스트래핑 탭을 먼저 채우세요."}), 200
    if not params.has_put:
        return jsonify({"status": "error",
                        "message": "풋옵션이 없는 RCPS는 BDT 풋채권 교차검증이 의미가 없습니다."}), 200

    T = params.T
    steps = int(data.get("steps") or max(int(round(T * 12)), 12))
    dt = T / steps

    from models.binomial_v2 import _coupon_schedule
    from models.bdt import evaluate_bdt_bond
    coupon_cf = _coupon_schedule(params, steps, dt)

    # 풋 스케줄: 4개 모형 통일 컨벤션(round) 사용 — TF/GS와 풋 활성 step 정확 일치
    put_step = params.date_to_step(params.put_start, steps)
    put_schedule = {i: float(params.put_exercise_price(i * dt))
                    for i in range(put_step, steps + 1)}

    try:
        out = evaluate_bdt_bond(rd_spot, T, steps, sigma, float(params.face_value),
                                coupon_cf, put_schedule)
    except Exception as e:
        return jsonify({"status": "error", "message": f"BDT 평가 실패: {str(e)[:200]}"}), 200

    # σ_r 해석 안내 — 사용자가 주식변동성과 혼동하지 않도록
    interp = ("σ_r은 단기금리 lognormal 변동성입니다. 주식 변동성(보통 30~60%)과 "
              "다른 개념으로, 한국 단기금리(CD91 기준) lognormal σ는 통상 5~15% 범위입니다.")
    sigma_warn = None
    if sigma > 0.5:
        sigma_warn = f"σ_r={sigma*100:.1f}% — 50% 초과. 입력값이 비정상적으로 큽니다 (주식 변동성을 잘못 입력했을 가능성)."
    elif sigma < 0.03:
        sigma_warn = f"σ_r={sigma*100:.2f}% — 3% 미만. 한국 시장 통상 범위(5~15%) 하회."

    # σ_r 민감도 — 사용자 입력 σ를 기준으로 5/8/10/12/15% 자동 산출
    # 평가법인 워크페이퍼 표준: 금리 옵션 σ는 시장 관측 어려운 가정 → 민감도 분석 필수
    sens_sigmas = [0.05, 0.08, 0.10, 0.12, 0.15]
    sigma_sensitivity = []
    base_pbv = out.get("put_bond_value")
    for s in sens_sigmas:
        try:
            r = evaluate_bdt_bond(rd_spot, T, steps, s, float(params.face_value),
                                  coupon_cf, put_schedule)
            sigma_sensitivity.append({
                "sigma": s,
                "sigma_pct": round(s * 100, 1),
                "put_bond_value": r.get("put_bond_value"),
                "put_option_value": r.get("put_option_value"),
                "diff_from_base": (r.get("put_bond_value") - base_pbv) if base_pbv else None,
            })
        except Exception:
            continue

    out.update({
        "status": "ok",
        "T": T,
        "n_steps_used": steps,
        "put_step": put_step,
        "rate_vol": sigma,
        "model": "Black-Derman-Toy (constant sigma, q=0.5)",
        "interpretation": interp,
        "sigma_warning": sigma_warn,
        "sigma_sensitivity": sigma_sensitivity,
    })
    return jsonify(out)


# ══════════════════════════════════════════════════════════════════
#  WACC 베타 자동 회귀산정 (CAPM)
# ══════════════════════════════════════════════════════════════════
@app.route("/api/wacc/beta", methods=["POST"])
def wacc_beta():
    """유사기업 베타 자동 산정 — 시장지수에 대한 단순선형회귀(OLS).

    body: {tickers:[...], market:'KS11', start, end, adjustment:'raw'|'blume'}
    """
    from models.volatility import clean_closes
    import numpy as np

    data = request.get_json(force=True) or {}
    tickers = [str(t).strip() for t in (data.get("tickers") or []) if str(t).strip()]
    if not tickers:
        return jsonify({"status": "error", "message": "종목이 비어있습니다."}), 200
    market = (data.get("market") or "KS11").strip()
    start = data.get("start") or None
    end = data.get("end") or None
    adjust = data.get("adjustment", "raw")
    frequency = (data.get("frequency") or "daily").lower()    # 'daily' or 'weekly'

    try:
        import FinanceDataReader as fdr
    except Exception:
        return jsonify({"status": "error", "message": "FinanceDataReader 미설치"}), 200

    # 시장지수 가격
    try:
        m_df = fdr.DataReader(market, start, end)
        if m_df is None or m_df.empty or 'Close' not in m_df.columns:
            raise ValueError("시장지수 조회 결과가 비어 있습니다")
        m_dates = [d.strftime('%Y-%m-%d') for d in m_df.index]
        m_closes = [float(x) for x in m_df['Close'].tolist()]
        m_dates, m_closes, _ = clean_closes(m_dates, m_closes)
        if len(m_closes) < 30:
            raise ValueError("시장지수 관측치 부족")
        m_map = dict(zip(m_dates, m_closes))
    except Exception as e:
        return jsonify({"status": "error",
                        "message": f"시장지수({market}) 조회 실패: {str(e)[:150]}"}), 200

    results = []
    for code in tickers:
        try:
            dates, closes, _ = _fetch_price_series(code, start, end)
            t_dates, tc, _ = clean_closes(dates, closes)
            if len(tc) < 30:
                raise ValueError("종목 관측치 부족 (<30)")
            t_map = dict(zip(t_dates, tc))
            common = sorted(set(m_map.keys()) & set(t_map.keys()))
            min_obs = 30
            # 주별 옵션: ISO 주차별 마지막 거래일만 추출 (Friday close 관행 근사)
            if frequency == "weekly":
                from datetime import date as _dt
                week_last = {}
                for d in common:
                    try:
                        yr, wk, _ = _dt.fromisoformat(d).isocalendar()
                        week_last[(yr, wk)] = d   # 후속 일자가 덮어씀 → 마지막 일
                    except Exception:
                        pass
                common = sorted(week_last.values())
            if len(common) < min_obs:
                raise ValueError(f"{'주별' if frequency=='weekly' else '공통'} 관측치 부족 ({len(common)}<{min_obs})")
            mc = np.array([m_map[d] for d in common], dtype=float)
            tc_a = np.array([t_map[d] for d in common], dtype=float)
            m_r = np.diff(np.log(mc))
            t_r = np.diff(np.log(tc_a))
            cov = float(np.cov(t_r, m_r, ddof=1)[0, 1])
            var_m = float(np.var(m_r, ddof=1))
            if var_m <= 0:
                raise ValueError("시장 분산 0")
            beta_raw = cov / var_m
            corr = float(np.corrcoef(t_r, m_r)[0, 1])
            r2 = corr * corr
            beta_adj = 0.67 * beta_raw + 0.33 if adjust == "blume" else beta_raw
            # 산점도용 좌표 (시장수익률, 종목수익률) — 모달 시각화용
            # 큰 데이터 방지: 500개 초과 시 균등 샘플링
            pts_n = len(m_r)
            if pts_n > 500:
                idx = np.linspace(0, pts_n - 1, 500).astype(int)
                m_r_s = m_r[idx]; t_r_s = t_r[idx]
            else:
                m_r_s = m_r; t_r_s = t_r
            scatter = [[float(x), float(y)] for x, y in zip(m_r_s, t_r_s)]
            # 회귀선 끝점 (시장수익률 min·max)
            mr_min, mr_max = float(m_r.min()), float(m_r.max())
            alpha = float(t_r.mean() - beta_raw * m_r.mean())
            line = [[mr_min, alpha + beta_raw * mr_min],
                    [mr_max, alpha + beta_raw * mr_max]]

            results.append({
                "ticker": code,
                "raw_beta": round(beta_raw, 4),
                "adjusted_beta": round(beta_adj, 4),
                "r2": round(r2, 4),
                "n_obs": int(len(common) - 1),
                "alpha": round(alpha, 6),
                "scatter": scatter,        # 산점도 점들
                "regression_line": line,   # 회귀선 두 끝점
                "start": common[0] if common else None,
                "end": common[-1] if common else None,
                "frequency": frequency,
                "market": market,
            })
        except Exception as e:  # noqa: BLE001
            results.append({"ticker": code, "error": str(e)[:150]})

    return jsonify({
        "status": "ok",
        "market": market,
        "adjustment": adjust,
        "frequency": frequency,
        "start": start, "end": end,
        "results": results,
    })


# ══════════════════════════════════════════════════════════════════
#  유사기업 D/E 자동 조회 (yfinance — Yahoo 재무제표 + 시가총액)
# ══════════════════════════════════════════════════════════════════
def _to_yahoo_symbol(code, listing_map):
    """6자리 KR 코드 → .KS/.KQ 부착. 이미 접미사 있거나 영문 티커는 그대로."""
    c = (code or "").strip()
    if not c:
        return c
    # 이미 접미사 있음(.MI/.HK/...) 또는 영문 포함 → 그대로
    if "." in c or any(ch.isalpha() for ch in c):
        return c
    # 순수 6자리 → 시장 매핑
    market = (listing_map.get(c) or "KOSPI").upper()
    return c + (".KQ" if market == "KOSDAQ" else ".KS")


def _find_bs_at(ticker_obj, target_date_str):
    """target_date 이하 가장 가까운 period-end의 Total Debt·Equity 추출.
    분기 우선, 없으면 연간. (period_end_str, total_debt, total_equity) 또는 None.
    """
    try:
        import pandas as pd
    except Exception:
        return None
    target = pd.Timestamp(target_date_str)
    for bs_attr in ('quarterly_balance_sheet', 'balance_sheet'):
        bs = getattr(ticker_obj, bs_attr, None)
        if bs is None or bs.empty or 'Total Debt' not in bs.index:
            continue
        # period_end ≤ target & Total Debt 값 존재
        valid_cols = []
        for c in bs.columns:
            try:
                if pd.Timestamp(c) <= target and pd.notna(bs.loc['Total Debt', c]):
                    valid_cols.append(c)
            except Exception:
                continue
        if not valid_cols:
            continue
        best = max(valid_cols)
        td = float(bs.loc['Total Debt', best])
        eq = None
        for eq_row in ('Stockholders Equity', 'Common Stock Equity', 'Total Equity Gross Minority Interest'):
            try:
                if eq_row in bs.index and pd.notna(bs.loc[eq_row, best]):
                    eq = float(bs.loc[eq_row, best]); break
            except Exception:
                continue
        return (str(best.date()), td, eq)
    return None


@app.route("/api/wacc/de", methods=["POST"])
def wacc_de():
    """유사기업 D/E·총부채·시가총액·Yahoo 베타 자동 조회.

    body: {tickers:[...], date:'YYYY-MM-DD' (선택)}.
    date 지정 시 분기/연간 재무상태표에서 그 일자 이하 가장 가까운 period의 값 사용.
    시가총액은 FDR로 그 일자의 종가 × 현재 발행주식수(근사).
    date 미지정 시 현 시점 스냅샷(.info) 사용.
    """
    data = request.get_json(force=True) or {}
    tickers = [str(t).strip() for t in (data.get("tickers") or []) if str(t).strip()]
    if not tickers:
        return jsonify({"status": "error", "message": "종목이 비어있습니다."}), 200
    date_str = (data.get("date") or "").strip() or None
    try:
        import yfinance as yf
    except Exception:
        return jsonify({"status": "error", "message": "yfinance 미설치"}), 200

    listing_map = {}
    try:
        for r in _stock_listing():
            listing_map[r["code"]] = r["market"]
    except Exception:
        pass

    results = []
    for code in tickers:
        ysym = _to_yahoo_symbol(code, listing_map)
        try:
            t = yf.Ticker(ysym)
            info = t.info or {}
            name = info.get('shortName') or info.get('longName') or code
            shares = info.get('sharesOutstanding')
            beta_yahoo = info.get('beta')
            # 매출액 (사이즈 프리미엄 보조지표). info.totalRevenue 는 TTM(최근 4분기). 평가기준일 매칭은 약함.
            total_revenue = info.get('totalRevenue')
            # 국가 코드 — 다국적 peer Hamada unlever에 국가별 세율 적용 (WACC M4)
            country = info.get('country') or info.get('countryRegion') or None

            td = eq = mc = period_end = None
            td_source = mc_source = None
            mc_as_of = None
            if date_str:
                bs = _find_bs_at(t, date_str)
                if bs:
                    period_end, td, eq = bs
                    td_source = 'bs'  # 평가기준일 ≤ period_end 매칭
                # 평가기준일 종가 × 현재 발행주식수 → 시총 근사
                try:
                    dates, closes, _ = _fetch_price_series(code, (datetime.fromisoformat(date_str) - timedelta(days=14)).date().isoformat(), (datetime.fromisoformat(date_str) + timedelta(days=2)).date().isoformat())
                    pick = None
                    for dt, c in zip(dates, closes):
                        if dt <= date_str:
                            pick = (dt, c)
                    if pick and shares:
                        mc = pick[1] * shares
                        mc_source = 'fdr'  # 평가기준일 종가 × 현 sharesOutstanding
                        mc_as_of = pick[0]
                except Exception:
                    pass
            # 현 시점 폴백 (date 없거나 historical 조회 실패)
            if td is None:
                td = info.get('totalDebt')
                if td is not None:
                    td_source = 'info'  # 현시점 스냅샷 폴백
            if mc is None:
                mc = info.get('marketCap')
                if mc is not None:
                    mc_source = 'info'
            # D/E 산정 — peer set 내 비교가능성 보장:
            # ① BS 조회 성공(td & eq)이면 계산값 사용, ② BS 폴백 시에도 동일 계산 시도,
            # ③ 어느 한 쪽이라도 없으면 None (yfinance.info.debtToEquity 폴백 제거 —
            #    info의 D/E는 부채 정의가 회사마다 달라 peer set 일관성 깨짐)
            de_book = round(td / eq * 100, 2) if (td and eq) else None
            de_market = round(td / mc * 100, 2) if (td and mc and mc > 0) else None
            # 시점 mismatch 경고: BS 분기말과 시총 조회일 차이가 90일 초과 시 플래그
            de_warn = None
            try:
                if period_end and mc_as_of:
                    from datetime import date as _date
                    bs_d = _date.fromisoformat(str(period_end)[:10])
                    mc_d = _date.fromisoformat(str(mc_as_of)[:10])
                    gap = abs((mc_d - bs_d).days)
                    if gap > 90:
                        de_warn = f"BS({period_end})와 시총조회일({mc_as_of}) 차이 {gap}일 — D/E 시점 mismatch 가능"
            except Exception:
                pass

            results.append({
                "ticker": code,
                "yahoo_symbol": ysym,
                "name": name,
                "de_book": de_book,
                "de_market": de_market,
                "total_debt": td,
                "stockholders_equity": eq,
                "market_cap": mc,
                "total_revenue": total_revenue,
                "yahoo_beta": beta_yahoo,
                "country": country,         # 다국적 peer 국가별 세율 적용 (M4)
                "as_of": period_end,
                "mc_as_of": mc_as_of,
                "td_source": td_source,
                "mc_source": mc_source,
                "de_warn": de_warn,
                "requested_date": date_str,
            })
        except Exception as e:  # noqa: BLE001
            results.append({"ticker": code, "yahoo_symbol": ysym, "error": str(e)[:150]})

    return jsonify({"status": "ok", "results": results, "date": date_str})


# ── 시작 시 종목목록 캐시 백그라운드 프리로드 ───────────────────────────
# Render Singapore 리전에서 첫 peer 검색이 10~30초 걸리던 문제 해결.
# 서버 import 시 별도 스레드로 _stock_listing() 호출 → 캐시 미리 채워둠.
# UptimeRobot keep-alive로 인스턴스 깨어 있는 동안 캐시도 유지됨.
def _preload_stock_listing():
    try:
        import time
        t0 = time.time()
        _stock_listing()
        print(f"[preload] stock listing cached in {time.time()-t0:.1f}s", flush=True)
    except Exception as e:
        print(f"[preload] stock listing failed: {e}", flush=True)

import threading as _threading
_threading.Thread(target=_preload_stock_listing, daemon=True).start()


# ══════════════════════════════════════════════════════════════
#  EV/EBITDA 자동 조회 (yfinance) — 비상장 RCPS DCF Exit Multiple 산출 보조
# ══════════════════════════════════════════════════════════════
def _find_ebitda_at(ticker_obj, target_date_str):
    """target_date 이하 가장 가까운 period의 EBITDA 추출. 분기 우선, 없으면 연간."""
    try:
        import pandas as pd
    except Exception:
        return None
    target = pd.Timestamp(target_date_str)
    for fin_attr in ('quarterly_financials', 'financials'):
        fin = getattr(ticker_obj, fin_attr, None)
        if fin is None or fin.empty:
            continue
        # EBITDA 직접 찾기
        ebitda_row = None
        for key in ('EBITDA', 'Normalized EBITDA'):
            if key in fin.index:
                ebitda_row = key; break
        # EBITDA 없으면 영업이익 + 감가상각비로 계산
        if ebitda_row is None:
            op_inc = None; dep = None
            for k in ('Operating Income', 'EBIT'):
                if k in fin.index: op_inc = k; break
            for k in ('Reconciled Depreciation', 'Depreciation', 'Depreciation And Amortization'):
                if k in fin.index: dep = k; break
            if op_inc is None: continue
        # 가장 가까운 period
        valid_cols = []
        for c in fin.columns:
            try:
                if pd.Timestamp(c) <= target:
                    if ebitda_row:
                        if pd.notna(fin.loc[ebitda_row, c]):
                            valid_cols.append((c, float(fin.loc[ebitda_row, c])))
                    else:
                        oi = fin.loc[op_inc, c] if op_inc in fin.index else None
                        dp = fin.loc[dep, c] if dep and dep in fin.index else 0
                        if pd.notna(oi):
                            valid_cols.append((c, float(oi) + (float(dp) if pd.notna(dp) else 0)))
            except Exception:
                continue
        if not valid_cols:
            continue
        best = max(valid_cols, key=lambda x: x[0])
        # 분기 데이터는 4개 분기 합산 (TTM)으로 연환산
        if fin_attr == 'quarterly_financials':
            ttm_cols = [v for c, v in valid_cols if (target - pd.Timestamp(c)).days <= 400]
            if len(ttm_cols) >= 4:
                return (str(best[0].date()), sum(ttm_cols[:4]), 'quarterly_ttm')
            elif ttm_cols:
                return (str(best[0].date()), sum(ttm_cols) * (4.0 / len(ttm_cols)), 'quarterly_annualized')
        return (str(best[0].date()), best[1], fin_attr)
    return None


@app.route("/api/peer_search", methods=["GET"])
def peer_search():
    """회사명·티커로 비교회사 검색 — yfinance.Search 활용 (해외 포함).

    국내 종목은 _stock_listing 명단 우선 매칭 (FDR 기반, 정확함).
    해외 종목은 yfinance.Search로 회사명 → ticker (예: 'Intercos' → 'ICOS.MI').
    """
    q = (request.args.get("q") or "").strip()
    if not q:
        return jsonify({"status": "ok", "results": []})
    results = []
    seen = set()

    # 1) 국내 종목 명단 (KRX FDR) — 회사명·코드 부분일치
    try:
        listing = _stock_listing()
        ql = q.lower()
        for r in listing:
            nm = (r.get("name") or "")
            code = r.get("code") or ""
            if (ql in nm.lower()) or (q == code) or code.startswith(q):
                key = code
                if key in seen: continue
                seen.add(key)
                results.append({
                    "code": code,
                    "name": nm,
                    "market": r.get("market", "KOSPI"),
                    "exchange": "KRX",
                })
                if len(results) >= 15: break
    except Exception:
        pass

    # 2) yfinance Search — 해외 종목 (회사명 검색)
    if len(results) < 10:
        try:
            import yfinance as yf
            s = yf.Search(q, max_results=10)
            for item in (s.quotes or []):
                sym = item.get("symbol") or ""
                if not sym or sym in seen: continue
                # 미국 ticker는 점 없음 (AAPL), 해외는 .MI/.PA/.T/.HK 등
                # 한국 종목은 위에서 이미 처리했으므로 .KS/.KQ는 건너뜀
                if sym.endswith(".KS") or sym.endswith(".KQ"): continue
                seen.add(sym)
                results.append({
                    "code": sym,
                    "name": item.get("shortname") or item.get("longname") or sym,
                    "market": item.get("exchange") or "",
                    "exchange": item.get("exchDisp") or item.get("exchange") or "",
                })
                if len(results) >= 15: break
        except Exception:
            pass

    return jsonify({"status": "ok", "results": results})


@app.route("/api/dcf/peer_multiples", methods=["POST"])
def dcf_peer_multiples():
    """비교회사 EV/EBITDA 자동 산출 (평가기준일 기준).

    EV = 시총 + 순차입금 (단순화: 총차입금 − 현금)
    EBITDA = 직접 항목 우선, 없으면 영업이익 + 감가상각비
    배수 = EV / EBITDA

    body: {tickers:[...], date:'YYYY-MM-DD' (선택)}
    """
    data = request.get_json(force=True) or {}
    tickers = [str(t).strip() for t in (data.get("tickers") or []) if str(t).strip()]
    if not tickers:
        return jsonify({"status": "error", "message": "종목이 비어있습니다."}), 200
    date_str = (data.get("date") or "").strip() or None
    try:
        import yfinance as yf
        import pandas as pd
    except Exception:
        return jsonify({"status": "error", "message": "yfinance/pandas 미설치"}), 200

    listing_map = {}
    name_to_code = {}  # 한글·영문 회사명 → ticker 코드 매핑
    try:
        for r in _stock_listing():
            listing_map[r["code"]] = r["market"]
            nm = (r.get("name") or "").strip()
            if nm:
                name_to_code[nm] = r["code"]
                # 약식 매칭(공백 제거)도 등록
                name_to_code[nm.replace(" ", "")] = r["code"]
    except Exception:
        pass

    def _resolve(token):
        """입력 토큰을 ticker 코드로 변환 — 회사명·6자리코드·yfinance ticker 모두 허용."""
        t = (token or "").strip()
        if not t: return t
        # 1) 회사명 직접 매칭
        if t in name_to_code: return name_to_code[t]
        # 2) 공백 제거 매칭
        if t.replace(" ", "") in name_to_code:
            return name_to_code[t.replace(" ", "")]
        # 3) 부분 일치 (한글 이름 일부만 입력한 경우)
        if any(ord(c) > 127 for c in t):  # 한글 등 비-ASCII
            for nm, code in name_to_code.items():
                if t in nm:
                    return code
        # 4) 6자리 코드 또는 영문 ticker — 그대로
        return t

    results = []
    for raw in tickers:
        code = _resolve(raw)
        ysym = _to_yahoo_symbol(code, listing_map)
        try:
            t = yf.Ticker(ysym)
            info = t.info or {}
            name = info.get('shortName') or info.get('longName') or code

            # 1) 시총·순차입금 (date 매칭 또는 현시점)
            mc = info.get('marketCap')
            net_debt = None
            period_end = None
            if date_str:
                bs = _find_bs_at(t, date_str)
                if bs:
                    period_end, td, _eq = bs
                    # 현금 추출
                    cash = None
                    try:
                        for bs_attr in ('quarterly_balance_sheet', 'balance_sheet'):
                            bs_df = getattr(t, bs_attr, None)
                            if bs_df is None or bs_df.empty: continue
                            for cash_row in ('Cash And Cash Equivalents', 'Cash Cash Equivalents And Short Term Investments'):
                                if cash_row in bs_df.index:
                                    col = pd.Timestamp(period_end)
                                    if col in bs_df.columns and pd.notna(bs_df.loc[cash_row, col]):
                                        cash = float(bs_df.loc[cash_row, col]); break
                            if cash is not None: break
                    except Exception: pass
                    net_debt = td - (cash or 0)
                # 평가기준일 종가 × 발행주식수로 시총 재계산
                try:
                    import FinanceDataReader as fdr
                    px = fdr.DataReader(code, date_str, date_str)
                    if not px.empty:
                        close = float(px['Close'].iloc[-1])
                        shares = info.get('sharesOutstanding')
                        if shares:
                            mc = close * shares
                except Exception:
                    pass
            else:
                # 현시점: info에서 직접
                td_info = info.get('totalDebt')
                cash_info = info.get('totalCash')
                if td_info is not None:
                    net_debt = td_info - (cash_info or 0)

            # 2) EBITDA
            ebitda = None
            ebitda_period = None
            ebitda_source = None
            if date_str:
                eb = _find_ebitda_at(t, date_str)
                if eb:
                    ebitda_period, ebitda, ebitda_source = eb
            if ebitda is None:
                # 현시점 fallback
                ebitda = info.get('ebitda')
                ebitda_source = 'info'

            # 3) EV·배수 산출
            ev = None; multiple = None
            if mc is not None and net_debt is not None:
                ev = mc + net_debt
            if ev is not None and ebitda and ebitda > 0:
                multiple = ev / ebitda

            results.append({
                "input": raw,           # 사용자가 입력한 원본 (회사명·코드)
                "ticker": code,         # 매핑된 ticker 코드
                "name": name,
                "market_cap": mc,
                "net_debt": net_debt,
                "ebitda": ebitda,
                "ebitda_period": ebitda_period,
                "ebitda_source": ebitda_source,
                "enterprise_value": ev,
                "ev_ebitda": multiple,
                "period_end": period_end,
            })
        except Exception as e:
            results.append({"input": raw, "ticker": code, "error": str(e)[:150]})

    # 통계 (정상값만)
    valid_multiples = [r["ev_ebitda"] for r in results if r.get("ev_ebitda") and 0 < r["ev_ebitda"] < 100]
    median = None; mean = None
    if valid_multiples:
        s = sorted(valid_multiples)
        n = len(s)
        median = s[n//2] if n % 2 else (s[n//2-1] + s[n//2]) / 2
        mean = sum(s) / n

    return jsonify({
        "status": "ok",
        "date": date_str,
        "peers": results,
        "summary": {
            "n_valid": len(valid_multiples),
            "median": median,
            "mean": mean,
        },
    })


# ══════════════════════════════════════════════════════════════
#  상세 Excel 다운로드 라우트 (DCF · WACC · 부트스트래핑)
# ══════════════════════════════════════════════════════════════
@app.route("/api/dcf/export", methods=["POST"])
def dcf_export():
    """DCF 상세 Excel — 가정·연도별 FCFF·평가결과 3 시트."""
    data = request.get_json(force=True) or {}
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            generate_dcf_xlsx(data, tmp.name)
        return send_file(tmp.name, as_attachment=True,
                         download_name=f"DCF상세_{data.get('valuation_date','')}.xlsx")
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)[:300]}), 200


@app.route("/api/wacc/export", methods=["POST"])
def wacc_export():
    """WACC 상세 Excel — WACC결과·CAPM_Ke·유사기업 3 시트."""
    data = request.get_json(force=True) or {}
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            generate_wacc_xlsx(data, tmp.name)
        return send_file(tmp.name, as_attachment=True,
                         download_name=f"WACC상세_{data.get('valuation_date','')}.xlsx")
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)[:300]}), 200


@app.route("/api/bootstrap/export", methods=["POST"])
def bootstrap_export():
    """부트스트래핑 곡선 Excel — Rf·Rd·이자율DATA 3 시트."""
    data = request.get_json(force=True) or {}
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            generate_bootstrap_xlsx(data, tmp.name)
        return send_file(tmp.name, as_attachment=True,
                         download_name=f"부트스트래핑_{data.get('valuation_date','')}.xlsx")
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)[:300]}), 200


@app.route("/api/volatility/export", methods=["POST"])
def volatility_export():
    """변동성 산정 결과 Excel — 산정조건·유사기업σ·조회실패·원자료 시트."""
    data = request.get_json(force=True) or {}
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            generate_volatility_xlsx(data, tmp.name)
        return send_file(tmp.name, as_attachment=True,
                         download_name=f"변동성평가_{data.get('valuation_date','')}.xlsx")
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)[:300]}), 200


if __name__ == "__main__":
    _DEBUG = os.environ.get("FLASK_DEBUG", "0") == "1"
    _HOST = os.environ.get("FLASK_HOST", "127.0.0.1")
    _PORT = int(os.environ.get("FLASK_PORT", "5000"))
    print(f"RCPS 평가툴 서버 시작: http://{_HOST}:{_PORT}  (debug={_DEBUG})")
    app.run(debug=_DEBUG, port=_PORT, host=_HOST, use_reloader=False)
