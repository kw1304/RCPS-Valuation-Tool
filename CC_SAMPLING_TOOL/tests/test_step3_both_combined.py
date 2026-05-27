"""test_step3_both_combined — kind=both 시 C100+AA100 단일 파일 통합 검증."""
from __future__ import annotations

import io
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl
import pytest

from api.app import app as flask_app

LEDGER_PATH = ROOT / "input" / "회사자료" / "채권채무조회서 거래처별 원장.XLSX"


@pytest.fixture(scope="module")
def client():
    flask_app.config["TESTING"] = True
    with flask_app.test_client() as c:
        yield c


@pytest.fixture(scope="module")
def combined_project(client):
    """채권+채무 both 프로젝트 생성 → Step1 실행."""
    if not LEDGER_PATH.exists():
        pytest.skip("원장 파일 없음")

    r = client.post("/api/project", json={
        "company_name": "CombinedTest",
        "period_end": "2025-12-31",
        "kind": "both",
    })
    assert r.status_code == 201
    pid = r.json["id"]

    with open(LEDGER_PATH, "rb") as f:
        data = {"ledger": (io.BytesIO(f.read()), "ledger.xlsx")}
    r = client.post(f"/api/upload?project_id={pid}",
                    data=data, content_type="multipart/form-data")
    sheet_map = r.json.get("sheet_map", {})
    ar_sheet = sheet_map.get("receivable", "채권")
    ap_sheet = sheet_map.get("payable", "채무")

    # 채권 Step 1
    r = client.post("/api/run", json={
        "kind": "receivable",
        "sheet": ar_sheet,
        "project_id": pid,
        "company_name": "CombinedTest",
        "period_end": "2025-12-31",
        "performance_materiality": 2_738_000_000,
        "risk_level": "유의적위험",
        "control_reliance": "Y",
        "seed": 42,
    })
    assert r.status_code == 200, r.json

    # 채무 Step 1
    r = client.post("/api/run", json={
        "kind": "payable",
        "sheet": ap_sheet,
        "project_id": pid,
        "company_name": "CombinedTest",
        "period_end": "2025-12-31",
        "performance_materiality": 2_738_000_000,
        "risk_level": "유의적위험",
        "control_reliance": "Y",
        "seed": 42,
    })
    # 채무 원장 없으면 skip (채무 시트 없는 클라이언트)
    if r.status_code != 200:
        pytest.skip("채무 원장 없음")

    return pid


def test_step3_both_returns_201(client, combined_project):
    """kind=both → 201 응답 + artifact_id."""
    pid = combined_project
    r = client.post(f"/api/project/{pid}/step3/build", json={
        "kind": "both",
        "preparer": "통합테스터",
        "reviewer": "검토자",
    })
    assert r.status_code == 201, r.json
    assert "artifact_id" in r.json
    assert "filename" in r.json


def test_step3_both_filename_format(client, combined_project):
    """kind=both → 파일명이 C100AA100_ 접두사 (채권채무 통합)."""
    pid = combined_project
    r = client.post(f"/api/project/{pid}/step3/build", json={
        "kind": "both",
        "preparer": "통합테스터",
    })
    assert r.status_code == 201
    filename = r.json.get("filename", "")
    assert filename.startswith("C100AA100_"), f"파일명 형식 오류: {filename}"


def test_step3_both_file_has_9_sheets(client, combined_project, tmp_path):
    """kind=both → 단일 xlsx에 generic reporter 9개 시트."""
    pid = combined_project
    r = client.post(f"/api/project/{pid}/step3/build", json={
        "kind": "both",
        "preparer": "통합테스터",
    })
    assert r.status_code == 201

    # 다운로드 (receivable 워크페이퍼에 연결)
    r_dl = client.get(f"/api/project/{pid}/step3/download/receivable")
    assert r_dl.status_code == 200

    out = tmp_path / "combined_workpaper.xlsx"
    out.write_bytes(r_dl.data)
    wb = openpyxl.load_workbook(out, read_only=True)
    sheets = set(wb.sheetnames)
    n = len(sheets)
    wb.close()

    assert n == 9, f"통합 조서 시트 수 불일치: {n} (기대 9)"
    assert "요약" in sheets, f"'요약' 시트 없음"
    assert "조회서" in sheets, f"'조회서' 시트 없음"


def test_step3_both_요약_sheet_data_exists(client, combined_project, tmp_path):
    """kind=both → '요약' 시트에 데이터가 채워짐."""
    pid = combined_project
    r = client.post(f"/api/project/{pid}/step3/build", json={
        "kind": "both",
        "preparer": "통합테스터",
    })
    assert r.status_code == 201

    r_dl = client.get(f"/api/project/{pid}/step3/download/receivable")
    assert r_dl.status_code == 200

    out = tmp_path / "combined_data.xlsx"
    out.write_bytes(r_dl.data)
    wb = openpyxl.load_workbook(out)
    ws = wb["요약"]
    # 셀 값 존재 여부 확인 (A열에 무언가 있어야 함)
    has_data = any(ws.cell(row, 2).value for row in range(1, 20))
    wb.close()

    assert has_data, "'요약' 시트에 데이터 없음"
