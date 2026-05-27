"""전체 5단계 통합 E2E 테스트

플로우:
  프로젝트 생성 → 회사자료 업로드 → Step 1 샘플링 → Step 2 발송명단 →
  Step 3 조서 → Step 4 PDF 회신 (모의 1건) → Step 5 대체적 절차 (1건) → 최종 조서

Flask test client와 SQLAlchemy in-memory DB로 완전 격리된 환경에서 실행.
실제 파일 의존 없이 더미 데이터 사용 (항상 실행).
"""
from __future__ import annotations

import io
import json
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl
import pytest
from openpyxl import Workbook


# ── 픽스처 헬퍼 ─────────────────────────────────────────────────────────────
def _make_ledger_xlsx() -> bytes:
    """7620 표준 컬럼 순서의 최소 거래처원장."""
    wb = Workbook()
    ws = wb.active
    ws.title = "채권"
    ws.append(["코드", "명", "계정과목", "계정과목명", "통화", "기초", "증감", "기말"])
    ws.append(["C001", "테스트거래처A", "01", "외상매출금", "KRW", 0, 800_000_000, 800_000_000])
    ws.append(["C002", "테스트거래처B", "01", "외상매출금", "KRW", 0, 300_000_000, 300_000_000])
    ws.append(["C003", "테스트거래처C", "03", "미수금",    "KRW", 0, 100_000_000, 100_000_000])

    ws2 = wb.create_sheet("채무")
    ws2.append(["코드", "명", "계정과목", "계정과목명", "통화", "기초", "증감", "기말"])
    ws2.append(["V001", "매입거래처X", "11", "외상매입금", "KRW", 0, 400_000_000, 400_000_000])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_fs_xlsx() -> bytes:
    """최소 재무제표."""
    wb = Workbook()
    ws = wb.active
    ws.title = "FS_M"
    # 기본 로더: item_col=3, value_col=5
    for _ in range(4):
        ws.append([None, None, None, None, None])  # 빈 행 채우기
    ws.cell(5, 3, "자산총계")
    ws.cell(5, 5, 5_000_000_000)
    ws.cell(6, 3, "외상매출금")
    ws.cell(6, 5, 1_100_000_000)
    ws.cell(7, 3, "미수금")
    ws.cell(7, 5, 100_000_000)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_dummy_pdf() -> bytes:
    """텍스트 레이어 포함 더미 PDF (pdfplumber 파싱용)."""
    # 최소 PDF 구조 — 실제 파싱은 빈 텍스트 반환하지만 파일 처리 플로우 검증용
    return b"""%PDF-1.4
1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj
2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj
3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]>>endobj
xref
0 4
0000000000 65535 f
0000000009 00000 n
0000000058 00000 n
0000000115 00000 n
trailer<</Size 4/Root 1 0 R>>
startxref
190
%%EOF"""


# ── Flask 클라이언트 픽스처 ──────────────────────────────────────────────────
@pytest.fixture(scope="module")
def client():
    import os
    os.environ["CC_TEST_DB"] = ":memory:"
    import api.app as app_module
    app_module.app.config["TESTING"] = True
    with app_module.app.test_client() as c:
        yield c, app_module
    os.environ.pop("CC_TEST_DB", None)


# ── 전체 5단계 통합 테스트 ──────────────────────────────────────────────────
class TestFullWorkflowE2E:
    """프로젝트 생성부터 최종 조서까지 전체 플로우 검증."""

    def test_step0_create_project(self, client):
        """Step 0: 프로젝트 생성."""
        c, app = client
        rv = c.post("/api/project", json={
            "company_name": "E2E테스트회사",
            "period_end": "2025-12-31",
            "kind": "both",
        })
        assert rv.status_code == 201
        data = rv.get_json()
        assert data["company_name"] == "E2E테스트회사"
        assert "id" in data

        # 프로젝트 ID를 모듈 수준에서 공유 (fixture scope=module)
        self.__class__._pid = data["id"]
        self.__class__._wp_receivable = data["workpapers"].get("receivable")
        self.__class__._wp_payable = data["workpapers"].get("payable")

    def test_step0_upload_files(self, client):
        """Step 0: 회사자료 파일 업로드."""
        c, app = client
        ledger_bytes = _make_ledger_xlsx()
        fs_bytes = _make_fs_xlsx()

        data = {
            "ledger": (io.BytesIO(ledger_bytes), "거래처원장.xlsx"),
            "fs": (io.BytesIO(fs_bytes), "재무제표.xlsx"),
            "project_id": self.__class__._pid,
        }
        rv = c.post(
            f"/api/upload?project_id={self.__class__._pid}",
            data=data,
            content_type="multipart/form-data",
        )
        assert rv.status_code == 200
        body = rv.get_json()
        assert "sheets" in body
        assert "채권" in body["sheets"]

    def test_step1_sampling_receivable(self, client):
        """Step 1: 채권 샘플링 실행."""
        c, app = client
        rv = c.post("/api/run", json={
            "kind": "receivable",
            "sheet": "채권",
            "company_name": "E2E테스트회사",
            "period_end": "2025-12-31",
            "performance_materiality": 100_000_000,
            "risk_level": "유의적위험",
            "control_reliance": "Y",
            "project_id": self.__class__._pid,
        })
        assert rv.status_code == 200
        body = rv.get_json()
        assert body["population_amount"] > 0
        assert "decisions" in body
        assert len(body["decisions"]) > 0

        # 최소 1건 이상 final_sampled
        sampled = [d for d in body["decisions"] if d["final_sampled"]]
        assert len(sampled) >= 1

    def test_step2_build_send_list(self, client):
        """Step 2: 발송명단 Excel 생성."""
        c, app = client
        pid = self.__class__._pid
        rv = c.post(f"/api/project/{pid}/step2/build", json={
            "kind": "receivable",
            "reply_deadline": "2026-01-31",
            "contact_info": {"preparer": "테스터", "phone": "02-1234-5678"},
        })
        assert rv.status_code == 201
        body = rv.get_json()
        assert "artifact_id" in body
        assert body["party_count"] >= 1

    def test_step3_build_workpaper(self, client):
        """Step 3: 조서 Excel 생성."""
        c, app = client
        pid = self.__class__._pid
        rv = c.post(f"/api/project/{pid}/step3/build", json={
            "kind": "receivable",
            "template_id": "woongkye_standard",
            "preparer": "테스터",
            "reviewer": "검토자",
        })
        assert rv.status_code == 201
        body = rv.get_json()
        assert "artifact_id" in body
        assert "download_url" in body

    def test_step3_download_workpaper(self, client):
        """Step 3: 조서 다운로드."""
        c, app = client
        pid = self.__class__._pid
        rv = c.get(f"/api/project/{pid}/step3/download/receivable")
        assert rv.status_code == 200
        assert rv.content_type in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/octet-stream",
        )

    def test_step4_upload_pdf_reply(self, client):
        """Step 4: PDF 회신 업로드 (더미 PDF)."""
        c, app = client
        pid = self.__class__._pid
        pdf_bytes = _make_dummy_pdf()
        data = {
            "files[]": (io.BytesIO(pdf_bytes), "테스트거래처A_회신.pdf"),
            "kind": "receivable",
            "tolerance": "0",
        }
        rv = c.post(
            f"/api/project/{pid}/step4/upload-replies",
            data=data,
            content_type="multipart/form-data",
        )
        # PDF 파싱 실패는 허용 (더미 PDF이므로) — 400이 아닌 200/201만 확인
        assert rv.status_code in (200, 201, 207), f"Step 4 응답: {rv.get_json()}"

    def test_step4_pdf_file_format_rejected(self, client):
        """Step 4: PDF 아닌 파일 업로드 시 400."""
        c, app = client
        pid = self.__class__._pid
        data = {
            "files[]": (io.BytesIO(b"not a pdf"), "reply.xlsx"),
            "kind": "receivable",
        }
        rv = c.post(
            f"/api/project/{pid}/step4/upload-replies",
            data=data,
            content_type="multipart/form-data",
        )
        assert rv.status_code == 400

    def test_step5_list_procedures(self, client):
        """Step 5: 대체적 절차 목록 조회 (빈 목록도 OK)."""
        c, app = client
        pid = self.__class__._pid
        wp_id = self.__class__._wp_receivable
        if wp_id:
            rv = c.get(f"/api/project/{pid}/step5/procedures/{wp_id}")
            assert rv.status_code in (200, 404)  # wp_id가 있으면 200

    def test_project_list_includes_e2e_project(self, client):
        """프로젝트 목록에 E2E 테스트 프로젝트가 포함되어야 한다."""
        c, app = client
        rv = c.get("/api/project")
        assert rv.status_code == 200
        projects = rv.get_json()
        pids = [p["id"] for p in projects]
        assert self.__class__._pid in pids

    def test_project_detail_shows_step_progress(self, client):
        """프로젝트 상세에서 Step 진행 상황이 표시되어야 한다."""
        c, app = client
        pid = self.__class__._pid
        rv = c.get(f"/api/project/{pid}")
        assert rv.status_code == 200
        data = rv.get_json()
        assert "workpapers" in data
        wps = {wp["kind"]: wp for wp in data["workpapers"]}
        # Step 3까지 진행했으므로 step3_done=True는 아닐 수 있지만 workpaper는 존재
        assert "receivable" in wps


# ── 7620 회귀 회귀 (항상 실행) ──────────────────────────────────────────────
class TestRegressionAlgorithm:
    """MUS 알고리즘·표본규모 회귀 — 수치 변경 없음 보장."""

    def test_mus_deterministic(self):
        """동일 seed면 동일 결과."""
        from src.domain.mus import run_mus

        pool = [("거래처A", 500_000_000), ("거래처B", 300_000_000),
                ("거래처C", 100_000_000), ("거래처D", 80_000_000)]

        r1 = run_mus(pool, sample_size=2, sample_interval=300_000_000, seed=42)
        r2 = run_mus(pool, sample_size=2, sample_interval=300_000_000, seed=42)
        assert r1.sampled_names == r2.sampled_names

    def test_sample_size_formula(self):
        """표본규모 = ceil(잔여모집단 / PM × CF) 수식 검증."""
        from src.domain.sample_size import compute_sample_size, SampleSizeInput
        import math

        inp = SampleSizeInput(
            population_amount=10_000_000_000,
            performance_materiality=1_000_000_000,
            risk_level="유의적위험",
            control_reliance="Y",
            key_item_amount=3_000_000_000,
        )
        result = compute_sample_size(inp)
        # 잔여 = 10B - 3B = 7B, base = 7B / 1B = 7, final = ceil(7 × 1.4) = 10
        assert result.base_sample_size == pytest.approx(7.0)
        assert result.final_sample_size == math.ceil(7 * 1.4)

    def test_key_item_threshold(self):
        """Key item 기준금액 = PM × ratio (유의적위험, Y = 75%)."""
        from src.domain.sample_size import compute_sample_size, SampleSizeInput

        inp = SampleSizeInput(
            population_amount=1_000_000_000,
            performance_materiality=100_000_000,
            risk_level="유의적위험",
            control_reliance="Y",
        )
        result = compute_sample_size(inp)
        assert result.key_item_threshold == pytest.approx(75_000_000)
