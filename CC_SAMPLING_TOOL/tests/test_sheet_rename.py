"""Task 1: 시트 이름 변경 검증 — '조회서' → '샘플링 거래처 내역'."""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl
import pytest

from src.infrastructure.report.generic_reporter import (
    EXPECTED_GENERIC_SHEETS,
    EXPECTED_SHEET_COUNT,
    build_combined_report,
    build_generic_report,
)
from src.domain.mus import MUSResult, MUSSelection
from src.domain.population import CompletenessCheck, PartyDecision
from src.domain.sample_size import SampleSizeResult
from src.infrastructure.report.generic_reporter import (
    KindData, ReportContext,
)


def _ctx(kind="receivable") -> ReportContext:
    return ReportContext(
        company_name="시트이름테스트",
        period_end=date(2025, 12, 31),
        kind=kind,
        preparer="작성자",
        reviewer="검토자",
    )


def _completeness() -> CompletenessCheck:
    return CompletenessCheck(
        by_group=[{"group": "외상매출금", "ledger": 1_000_000, "fs": 1_000_000, "diff": 0}],
        total_ledger=1_000_000, total_fs=1_000_000, total_diff=0,
    )


def _size() -> SampleSizeResult:
    return SampleSizeResult(
        key_item_threshold=100_000,
        key_item_ratio=0.5,
        confidence_factor=1.6,
        base_sample_size=5.0,
        final_sample_size=8,
        sample_interval=112_500,
        remaining_population=900_000,
    )


def _decisions() -> list[PartyDecision]:
    return [
        PartyDecision(name="거래처A", balance=500_000, is_key_item=True,
                      is_representative=False, is_related_party=False,
                      is_excluded=False, final_sampled=True),
        PartyDecision(name="거래처B", balance=200_000, is_key_item=False,
                      is_representative=True, is_related_party=False,
                      is_excluded=False, final_sampled=True),
    ]


def _mus() -> MUSResult:
    return MUSResult(
        sample_interval=112_500,
        random_start=50_000,
        selections=[
            MUSSelection(name="거래처A", balance=500_000, cumulative=500_000,
                         selections=4, remainder_after=50_000, hit=True),
        ],
        sampled_names=["거래처A"],
    )


def _kd(kind="receivable") -> KindData:
    return KindData(
        ctx=_ctx(kind),
        completeness=_completeness(),
        size_result=_size(),
        decisions=_decisions(),
        mus_result=_mus(),
        performance_materiality=200_000,
        population_amount=1_000_000,
    )


def test_sampling_party_list_sheet_exists(tmp_path):
    """'샘플링 거래처 내역' 시트가 존재해야 한다."""
    out = tmp_path / "sheet_rename.xlsx"
    build_combined_report(out, receivable=_kd(), payable=None)

    wb = openpyxl.load_workbook(out, read_only=True)
    sheets = set(wb.sheetnames)
    wb.close()

    assert "샘플링 거래처 내역" in sheets, (
        f"'샘플링 거래처 내역' 시트 없음 — 실제 시트: {sheets}"
    )


def test_confirmation_sheet_exists(tmp_path):
    """'조회서' 시트가 존재해야 한다 (신규 UploadGuide 기반 시트)."""
    out = tmp_path / "sheet_rename.xlsx"
    build_combined_report(out, receivable=_kd(), payable=None)

    wb = openpyxl.load_workbook(out, read_only=True)
    sheets = set(wb.sheetnames)
    wb.close()

    assert "조회서" in sheets, f"'조회서' 시트 없음 — 실제 시트: {sheets}"


def test_old_confirmation_name_gone(tmp_path):
    """기존 '조회서'가 발송 명단(Control Sheet) 역할로 남아있지 않음.
    즉 '샘플링 거래처 내역'이 그 역할을 대체함."""
    out = tmp_path / "sheet_rename.xlsx"
    build_combined_report(out, receivable=_kd(), payable=None)

    wb = openpyxl.load_workbook(out, data_only=True)
    # 샘플링 거래처 내역에 거래처A가 있어야 함
    ws = wb["샘플링 거래처 내역"]
    values = {ws.cell(r, c).value
              for r in range(1, ws.max_row + 1)
              for c in range(1, ws.max_column + 1)}
    wb.close()

    assert "거래처A" in values, "샘플링 거래처 내역에 발송 명단 데이터 없음"


def test_total_sheet_count(tmp_path):
    """build_combined_report → EXPECTED_SHEET_COUNT(10)개 시트."""
    out = tmp_path / "count_check.xlsx"
    build_combined_report(out, receivable=_kd(), payable=_kd("payable"))

    wb = openpyxl.load_workbook(out, read_only=True)
    n = len(wb.sheetnames)
    actual = set(wb.sheetnames)
    wb.close()

    assert n == EXPECTED_SHEET_COUNT, (
        f"시트 수 불일치: {n} (기대 {EXPECTED_SHEET_COUNT}) — {actual}"
    )


def test_expected_sheets_constant_includes_both(tmp_path):
    """EXPECTED_GENERIC_SHEETS에 '샘플링 거래처 내역'과 '조회서' 모두 포함."""
    assert "샘플링 거래처 내역" in EXPECTED_GENERIC_SHEETS
    assert "조회서" in EXPECTED_GENERIC_SHEETS
