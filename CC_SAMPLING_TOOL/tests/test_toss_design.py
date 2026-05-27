"""
test_toss_design.py — Toss 디자인 핵심 요소 검증

검증:
  - 헤더 행: 흰 배경(FFFFFF) + 파란 bottom border(3182F6)
  - KPI 값: 14pt bold 파란 글씨
  - 발송제외: strikethrough (줄긋기)
  - navy 배경(44546A) 부재 확인 (구 7620 스타일 완전 제거)
  - 노랑 배경(FFFFCC) 부재 확인 (구 입력값 강조 스타일 제거)
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl
import pytest

from src.infrastructure.report.generic_reporter import (
    PartyContactInfo,
    ReportContext,
    build_generic_report,
    TOSS_ACCENT,
    TOSS_BG_SUB,
    TOSS_WHITE,
)
from src.domain.mus import MUSResult, MUSSelection
from src.domain.population import CompletenessCheck, PartyDecision
from src.domain.sample_size import SampleSizeResult


def _ctx() -> ReportContext:
    return ReportContext(
        company_name="Toss디자인테스트",
        period_end=date(2025, 12, 31),
        kind="receivable",
        preparer="작성자",
        reviewer="검토자",
        workpaper_no_prefix="C100",
    )


def _make_wb(tmp_path) -> openpyxl.Workbook:
    out = tmp_path / "toss_design.xlsx"
    build_generic_report(
        out_path=out,
        ctx=_ctx(),
        completeness=CompletenessCheck(
            by_group=[{"group": "외상매출금", "ledger": 2_000_000, "fs": 2_000_000, "diff": 0, "note": ""}],
            total_ledger=2_000_000, total_fs=2_000_000, total_diff=0,
        ),
        size_result=SampleSizeResult(
            key_item_threshold=400_000,
            key_item_ratio=0.5,
            confidence_factor=1.4,
            base_sample_size=6,
            final_sample_size=6,
            sample_interval=300_000,
            remaining_population=1_800_000,
        ),
        decisions=[
            PartyDecision(
                name="키거래처", balance=800_000,
                is_key_item=True, is_representative=False,
                is_related_party=False, is_excluded=False,
                final_sampled=True,
                by_account={"외상매출금": 800_000},
            ),
            PartyDecision(
                name="일반거래처", balance=300_000,
                is_key_item=False, is_representative=True,
                is_related_party=False, is_excluded=False,
                final_sampled=True,
                by_account={"미수금": 300_000},
            ),
            PartyDecision(
                name="제외거래처", balance=50_000,
                is_key_item=False, is_representative=False,
                is_related_party=False, is_excluded=True,
                final_sampled=False,
                exclusion_reason="해산",
                by_account={"외상매출금": 50_000},
            ),
        ],
        mus_result=MUSResult(
            sample_interval=300_000,
            random_start=100_000,
            selections=[
                MUSSelection(
                    name="일반거래처", balance=300_000, cumulative=300_000,
                    selections=1, remainder_after=0, hit=True,
                ),
            ],
            sampled_names=["일반거래처"],
        ),
        performance_materiality=800_000,
        population_amount=2_000_000,
    )
    return openpyxl.load_workbook(out)


def test_header_cells_white_background(tmp_path):
    """조회서 헤더 행: 흰 배경(FFFFFF) 확인."""
    wb = _make_wb(tmp_path)
    ws = wb["샘플링 거래처 내역"]

    white_header_found = any(
        cell.fill and cell.fill.fgColor
        and cell.fill.fgColor.rgb.upper().endswith("FFFFFF")
        and cell.border and cell.border.bottom
        and cell.border.bottom.color
        and cell.border.bottom.color.rgb.upper().endswith(TOSS_ACCENT.upper())
        for row in ws.iter_rows()
        for cell in row
    )
    assert white_header_found, "조회서 헤더 셀에 흰 배경 + 파란 bottom border 없음"


def test_header_cells_no_navy_background(tmp_path):
    """구 7620 navy(44546A) 배경 셀이 존재하지 않아야 한다."""
    wb = _make_wb(tmp_path)
    navy_cells = [
        f"{sheet_name}!{cell.coordinate}"
        for sheet_name in wb.sheetnames
        for row in wb[sheet_name].iter_rows()
        for cell in row
        if cell.fill and cell.fill.fgColor
        and cell.fill.fgColor.rgb.upper().endswith("44546A")
    ]
    assert not navy_cells, f"navy(44546A) 배경 셀 발견 — Toss 스타일 미적용: {navy_cells[:5]}"


def test_no_input_yellow_background(tmp_path):
    """구 7620 노랑(FFFFCC) 배경 셀이 존재하지 않아야 한다."""
    wb = _make_wb(tmp_path)
    yellow_cells = [
        f"{sheet_name}!{cell.coordinate}"
        for sheet_name in wb.sheetnames
        for row in wb[sheet_name].iter_rows()
        for cell in row
        if cell.fill and cell.fill.fgColor
        and cell.fill.fgColor.rgb.upper().endswith("FFFFCC")
    ]
    assert not yellow_cells, f"노랑(FFFFCC) 배경 셀 발견 — 구 7620 스타일 잔재: {yellow_cells[:5]}"


def test_kpi_value_font_size_and_color(tmp_path):
    """요약 시트 KPI 값: 14pt bold 파란(3182F6)이어야 한다."""
    wb = _make_wb(tmp_path)
    ws = wb["요약"]

    kpi_found = any(
        cell.font and cell.font.size == 14
        and cell.font.bold
        and cell.font.color.rgb.upper().endswith(TOSS_ACCENT.upper())
        for row in ws.iter_rows()
        for cell in row
        if cell.value and isinstance(cell.value, (int, float)) and cell.value > 0
    )
    assert kpi_found, "요약 시트에 14pt bold 파란 KPI 값 셀 없음"


def test_excluded_party_strikethrough_in_confirmation(tmp_path):
    """조회서: 발송제외 거래처(제외거래처) 이름 셀이 strikethrough이어야 한다."""
    wb = _make_wb(tmp_path)
    ws = wb["샘플링 거래처 내역"]

    strike_found = any(
        cell.font and cell.font.strike
        for row in ws.iter_rows()
        for cell in row
        if cell.value == "제외거래처"
    )
    assert strike_found, "발송제외 거래처 이름 셀에 strikethrough 없음"


def test_total_row_bg_sub_fill(tmp_path):
    """합계 행: F9FAFB(TOSS_BG_SUB) 배경 + bold 확인."""
    wb = _make_wb(tmp_path)
    ws = wb["샘플링 거래처 내역"]

    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "총합계":
                assert cell.font.bold, "총합계 셀 bold 아님"
                assert (cell.fill and cell.fill.fgColor
                        and cell.fill.fgColor.rgb.upper().endswith(TOSS_BG_SUB.upper())), (
                    f"총합계 배경 불일치: {cell.fill.fgColor.rgb}"
                )
                return
    pytest.fail("조회서 '총합계' 셀 없음")
