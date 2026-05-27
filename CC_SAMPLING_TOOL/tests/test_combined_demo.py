"""
test_combined_demo.py — kind=both 통합 11시트 조서 검증

검증 항목:
  1. build_combined_report → 11시트 단일 파일
  2. 1번 시트 = "샘플링 요약"
  3. 시트 순서가 desired_order와 일치
  4. C100 시리즈(채권)에 채권 데이터 존재
  5. AA100 시리즈(채무)에 채무 데이터 존재
  6. C100A·대체적 절차 시트 존재
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl
import pytest

from src.infrastructure.report.generic_reporter import (
    AlternativeProcedureEntry,
    ConfirmationReplyInfo,
    ExclusionRow,
    KindData,
    PartyContactInfo,
    ReportContext,
    build_combined_report,
)
from src.domain.mus import MUSResult, MUSSelection
from src.domain.population import CompletenessCheck, PartyDecision
from src.domain.sample_size import SampleSizeResult


# ─────────────────────────────────────────────────────────────
# 픽스처
# ─────────────────────────────────────────────────────────────

def _completeness() -> CompletenessCheck:
    return CompletenessCheck(
        by_group=[
            {"group": "외상매출금", "ledger": 3_000_000, "fs": 3_000_000, "diff": 0, "note": ""},
        ],
        total_ledger=3_000_000,
        total_fs=3_000_000,
        total_diff=0,
    )


def _size_result() -> SampleSizeResult:
    return SampleSizeResult(
        key_item_threshold=400_000,
        key_item_ratio=0.5,
        confidence_factor=1.4,
        base_sample_size=7,
        final_sample_size=7,
        sample_interval=400_000,
        remaining_population=2_000_000,
    )


def _mus(name: str) -> MUSResult:
    return MUSResult(
        sample_interval=400_000,
        random_start=100_000,
        selections=[
            MUSSelection(
                name=name, balance=300_000, cumulative=300_000,
                selections=1, remainder_after=100_000, hit=True,
            ),
        ],
        sampled_names=[name],
    )


def _receivable_kd() -> KindData:
    return KindData(
        ctx=ReportContext(
            company_name="통합테스트회사",
            period_end=date(2025, 12, 31),
            kind="receivable",
            preparer="작성자A",
            reviewer="검토자B",
            workpaper_no_prefix="C100",
        ),
        completeness=_completeness(),
        size_result=_size_result(),
        decisions=[
            PartyDecision(
                name="채권거래처X", balance=900_000,
                is_key_item=True, is_representative=False,
                is_related_party=False, is_excluded=False,
                final_sampled=True,
                by_account={"외상매출금": 900_000},
            ),
            PartyDecision(
                name="채권거래처Y", balance=300_000,
                is_key_item=False, is_representative=True,
                is_related_party=False, is_excluded=False,
                final_sampled=True,
                by_account={"미수금": 300_000},
            ),
        ],
        mus_result=_mus("채권거래처Y"),
        performance_materiality=800_000,
        population_amount=3_000_000,
        contacts=[
            PartyContactInfo(name="채권거래처X", country="KR",
                             contact_person="김담당", email="x@test.com"),
        ],
    )


def _payable_kd() -> KindData:
    return KindData(
        ctx=ReportContext(
            company_name="통합테스트회사",
            period_end=date(2025, 12, 31),
            kind="payable",
            preparer="작성자A",
            reviewer="검토자B",
            workpaper_no_prefix="AA100",
        ),
        completeness=_completeness(),
        size_result=_size_result(),
        decisions=[
            PartyDecision(
                name="채무거래처P", balance=600_000,
                is_key_item=True, is_representative=False,
                is_related_party=False, is_excluded=False,
                final_sampled=True,
                by_account={"외상매입금": 600_000},
            ),
            PartyDecision(
                name="채무거래처Q", balance=250_000,
                is_key_item=False, is_representative=True,
                is_related_party=False, is_excluded=False,
                final_sampled=True,
                by_account={"미지급금": 250_000},
            ),
        ],
        mus_result=_mus("채무거래처Q"),
        performance_materiality=800_000,
        population_amount=2_500_000,
    )


def _build_combined(tmp_path: Path) -> openpyxl.Workbook:
    out = tmp_path / "combined_test.xlsx"
    build_combined_report(
        out_path=out,
        receivable=_receivable_kd(),
        payable=_payable_kd(),
    )
    return openpyxl.load_workbook(out)


# ─────────────────────────────────────────────────────────────
# 1. 시트 수
# ─────────────────────────────────────────────────────────────

def test_combined_has_11_sheets(tmp_path):
    """build_combined_report → 정확히 11개 시트."""
    wb = _build_combined(tmp_path)
    assert len(wb.sheetnames) == 11, (
        f"시트 수 불일치: {len(wb.sheetnames)} — {wb.sheetnames}"
    )


# ─────────────────────────────────────────────────────────────
# 2. 1번 시트 = 샘플링 요약
# ─────────────────────────────────────────────────────────────

def test_first_sheet_is_summary(tmp_path):
    """첫 번째 시트가 '샘플링 요약'이어야 한다."""
    wb = _build_combined(tmp_path)
    assert wb.sheetnames[0] == "샘플링 요약", (
        f"첫 시트 불일치: {wb.sheetnames[0]}"
    )


# ─────────────────────────────────────────────────────────────
# 3. 시트 순서 — desired_order 전체 일치
# ─────────────────────────────────────────────────────────────

DESIRED_ORDER = [
    "샘플링 요약",
    "C100 조회서", "C100-1 표본규모 결정",
    "C100-2 Key item 추출", "C100-3 표본 추출(MUS)",
    "AA100 조회서", "AA100-1 표본규모 결정",
    "AA100-2 Key item 추출", "AA100-3 표본 추출(MUS)",
    "C100A 조회처 주소 적정성",
    "대체적 절차",
]


def test_sheet_order_matches_desired(tmp_path):
    """11개 시트 순서가 desired_order와 정확히 일치해야 한다."""
    wb = _build_combined(tmp_path)
    assert wb.sheetnames == DESIRED_ORDER, (
        f"\n실제:   {wb.sheetnames}\n기대: {DESIRED_ORDER}"
    )


# ─────────────────────────────────────────────────────────────
# 4. 채권 데이터가 C100 시리즈에 있어야 함
# ─────────────────────────────────────────────────────────────

def test_c100_sheet_contains_receivable_party(tmp_path):
    """C100 조회서에 채권거래처X가 포함되어야 한다."""
    wb = _build_combined(tmp_path)
    ws = wb["C100 조회서"]
    all_values = [
        str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    assert "채권거래처X" in all_values, "C100 조회서에 채권거래처X 없음"


def test_c100_1_sheet_contains_receivable_data(tmp_path):
    """C100-1 시트에 채권 모집단 금액(3,000,000)이 반영되어야 한다."""
    wb = _build_combined(tmp_path)
    ws = wb["C100-1 표본규모 결정"]
    all_values = [cell.value for row in ws.iter_rows() for cell in row]
    assert 3_000_000 in all_values, "C100-1에 채권 모집단 금액 없음"


# ─────────────────────────────────────────────────────────────
# 5. 채무 데이터가 AA100 시리즈에 있어야 함
# ─────────────────────────────────────────────────────────────

def test_aa100_sheet_contains_payable_party(tmp_path):
    """AA100 조회서에 채무거래처P가 포함되어야 한다."""
    wb = _build_combined(tmp_path)
    ws = wb["AA100 조회서"]
    all_values = [
        str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    assert "채무거래처P" in all_values, "AA100 조회서에 채무거래처P 없음"


def test_aa100_1_sheet_exists_and_has_payable_data(tmp_path):
    """AA100-1 시트에 채무 모집단 금액(2,500,000)이 반영되어야 한다."""
    wb = _build_combined(tmp_path)
    ws = wb["AA100-1 표본규모 결정"]
    all_values = [cell.value for row in ws.iter_rows() for cell in row]
    assert 2_500_000 in all_values, "AA100-1에 채무 모집단 금액 없음"


# ─────────────────────────────────────────────────────────────
# 6. C100A, 대체적 절차 시트 존재
# ─────────────────────────────────────────────────────────────

def test_c100a_sheet_exists(tmp_path):
    """C100A 조회처 주소 적정성 시트가 존재해야 한다."""
    wb = _build_combined(tmp_path)
    assert "C100A 조회처 주소 적정성" in wb.sheetnames


def test_alt_procedure_sheet_exists(tmp_path):
    """대체적 절차 시트가 존재해야 한다."""
    wb = _build_combined(tmp_path)
    assert "대체적 절차" in wb.sheetnames


# ─────────────────────────────────────────────────────────────
# 7. AA100 조서번호 셀 빨강
# ─────────────────────────────────────────────────────────────

def test_aa100_1_workpaper_no_red(tmp_path):
    """AA100-1 시트 I2 셀: 'AA100-1' 빨강 bold."""
    wb = _build_combined(tmp_path)
    ws = wb["AA100-1 표본규모 결정"]
    cell = ws["I2"]
    assert cell.value == "AA100-1", f"I2 값: {cell.value}"
    assert cell.font.bold, "AA100-1 조서번호 셀이 bold 아님"
    assert cell.font.color.rgb.upper().endswith("FF0000"), (
        f"AA100-1 조서번호 글씨 색: {cell.font.color.rgb}"
    )


# ─────────────────────────────────────────────────────────────
# 8. 통합 요약 시트 컨텐츠
# ─────────────────────────────────────────────────────────────

def test_summary_sheet_contains_both_amounts(tmp_path):
    """샘플링 요약: 채권+채무 합산 모집단(5,500,000)이 반영되어야 한다."""
    wb = _build_combined(tmp_path)
    ws = wb["샘플링 요약"]
    all_values = [cell.value for row in ws.iter_rows() for cell in row]
    # 3,000,000 + 2,500,000 = 5,500,000
    assert 5_500_000 in all_values, (
        f"요약 시트에 합산 모집단 5,500,000 없음. 실제 숫자 값: "
        f"{[v for v in all_values if isinstance(v, (int, float))]}"
    )


# ─────────────────────────────────────────────────────────────
# 9. 채권·채무 데이터 교차 오염 없음
# ─────────────────────────────────────────────────────────────

def test_c100_sheet_no_payable_party(tmp_path):
    """C100 조회서에 채무 전용 거래처(채무거래처P)가 없어야 한다."""
    wb = _build_combined(tmp_path)
    ws = wb["C100 조회서"]
    all_values = [
        str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    assert "채무거래처P" not in all_values, "C100 조회서에 채무 거래처 교차 오염"


def test_aa100_sheet_no_receivable_party(tmp_path):
    """AA100 조회서에 채권 전용 거래처(채권거래처X)가 없어야 한다."""
    wb = _build_combined(tmp_path)
    ws = wb["AA100 조회서"]
    all_values = [
        str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    assert "채권거래처X" not in all_values, "AA100 조회서에 채권 거래처 교차 오염"
