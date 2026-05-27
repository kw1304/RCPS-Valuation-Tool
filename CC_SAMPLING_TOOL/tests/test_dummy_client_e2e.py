"""더미 클라이언트 A E2E 통합 테스트

범용성 검증: 시트명·컬럼 순서가 7620과 다른 새 클라이언트(dummy_client_a)에 대해
Step 1 샘플링 → Step 2 발송명단 → Step 3 조서 생성 자동 동작을 확인한다.

더미 클라이언트 A 특징:
- 거래처원장: 시트명 "매출원장"/"매입원장" (채권/채무 변형)
- 재무제표: 시트명 "BS" (FS_M 대신), 컬럼 [No, 구분, 항목명, 전기금액, 당기금액]
- 특관자: 시트명 "관계회사" + 2열 구조
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import pytest
import pandas as pd

FIXTURES = ROOT / "tests" / "fixtures" / "dummy_client_a"
LEDGER_PATH = FIXTURES / "거래처원장.xlsx"
FS_PATH = FIXTURES / "재무제표.xlsx"
RP_PATH = FIXTURES / "특관자.xlsx"

pytestmark = pytest.mark.skipif(
    not LEDGER_PATH.exists(),
    reason="dummy_client_a 픽스처 없음",
)


# ── 1. 시트명 자동 감지 ──────────────────────────────────────────────────────
class TestDummyLedgerSheetDetect:
    def test_receivable_sheet_detected_as_매출원장(self):
        """'매출원장' 시트가 채권 시트로 자동 감지되어야 한다."""
        import openpyxl
        from src.infrastructure.schemas.ledger_schema import detect_ledger_sheets

        wb = openpyxl.load_workbook(LEDGER_PATH, read_only=True, data_only=True)
        result = detect_ledger_sheets(wb.sheetnames)
        wb.close()

        assert result["receivable"] == "매출원장", (
            f"채권 시트 감지 실패: {result}"
        )

    def test_payable_sheet_detected_as_매입원장(self):
        """'매입원장' 시트가 채무 시트로 자동 감지되어야 한다."""
        import openpyxl
        from src.infrastructure.schemas.ledger_schema import detect_ledger_sheets

        wb = openpyxl.load_workbook(LEDGER_PATH, read_only=True, data_only=True)
        result = detect_ledger_sheets(wb.sheetnames)
        wb.close()

        assert result["payable"] == "매입원장", (
            f"채무 시트 감지 실패: {result}"
        )


# ── 2. 컬럼 자동 감지 ──────────────────────────────────────────────────────
class TestDummyLedgerColumnDetect:
    def test_ending_balance_col_detected(self):
        """컬럼 순서 변형 시에도 '기말잔액' 컬럼이 감지되어야 한다."""
        from src.infrastructure.schemas.ledger_schema import detect_ledger_columns

        df = pd.read_excel(LEDGER_PATH, sheet_name="매출원장", header=0)
        col_map = detect_ledger_columns(df)

        assert col_map["ending"] is not None, "기말잔액 컬럼 감지 실패"
        assert col_map["name_col"] is not None, "거래처명 컬럼 감지 실패"

    def test_ending_col_points_to_correct_index(self):
        """감지된 기말잔액 인덱스가 실제 700,000,000 등 숫자를 반환해야 한다."""
        from src.infrastructure.schemas.ledger_schema import detect_ledger_columns

        df = pd.read_excel(LEDGER_PATH, sheet_name="매출원장", header=0)
        col_map = detect_ledger_columns(df)
        ending_idx = col_map["ending"]

        assert ending_idx is not None
        vals = df.iloc[:, ending_idx].dropna().tolist()
        assert any(isinstance(v, (int, float)) and v > 0 for v in vals), (
            f"기말잔액 컬럼이 유효한 숫자를 포함하지 않음: {vals[:3]}"
        )


# ── 3. 재무제표 자동 감지 ────────────────────────────────────────────────────
class TestDummyFsAutoDetect:
    def test_fs_sheet_detected_as_bs(self):
        """'BS' 시트명이 재무제표로 자동 감지되어야 한다."""
        import openpyxl
        from src.infrastructure.schemas.fs_schema import detect_fs_sheet

        wb = openpyxl.load_workbook(FS_PATH, read_only=True, data_only=True)
        result = detect_fs_sheet(wb.sheetnames)
        wb.close()

        assert result == "BS", f"BS 시트 감지 실패: {result}"

    def test_fs_loads_당기금액(self):
        """load_fs_amounts가 '당기금액' 컬럼을 자동 선택해야 한다."""
        from src.infrastructure.loaders import load_fs_amounts

        fs = load_fs_amounts(FS_PATH)  # sheet=None → 자동 감지

        assert "자산총계" in fs, f"자산총계 누락: {list(fs.keys())}"
        # dummy_client_a BS: 자산총계 당기 = 4,000,000,000
        assert fs["자산총계"] == pytest.approx(4_000_000_000), (
            f"자산총계 불일치: {fs.get('자산총계')}"
        )

    def test_get_total_assets_from_bs(self):
        """총자산이 4,000,000,000으로 추출되어야 한다."""
        from src.infrastructure.loaders import load_fs_amounts, get_total_assets

        fs = load_fs_amounts(FS_PATH)
        total = get_total_assets(fs)

        assert total == pytest.approx(4_000_000_000)


# ── 4. 특관자 자동 감지 ─────────────────────────────────────────────────────
class TestDummyRelatedParties:
    def test_rp_sheet_detected_as_관계회사(self):
        """'관계회사' 시트가 특관자 시트로 자동 감지되어야 한다."""
        import openpyxl
        from src.infrastructure.schemas.rp_schema import detect_rp_sheet

        wb = openpyxl.load_workbook(RP_PATH, read_only=True, data_only=True)
        result = detect_rp_sheet(wb.sheetnames)
        wb.close()

        assert result == "관계회사", f"관계회사 시트 감지 실패: {result}"

    def test_rp_names_loaded(self):
        """특관자 이름 3건이 올바르게 로드되어야 한다."""
        from src.infrastructure.loaders import load_related_parties

        rp = load_related_parties(RP_PATH)

        assert "카파홀딩스" in rp
        assert "감마코퍼레이션" in rp
        assert "에타그룹" in rp


# ── 5. Step 1 샘플링 (채권) ──────────────────────────────────────────────────
class TestDummyStep1Sampling:
    def _run_receivable_sampling(self):
        from src.infrastructure.loaders import load_related_parties
        from src.infrastructure.schemas.ledger_schema import detect_ledger_sheets
        from src.orchestrator import SamplingParams, run_sampling

        # 시트 자동 감지
        import openpyxl
        wb = openpyxl.load_workbook(LEDGER_PATH, read_only=True, data_only=True)
        sheet_map = detect_ledger_sheets(wb.sheetnames)
        wb.close()

        df = pd.read_excel(LEDGER_PATH, sheet_name=sheet_map["receivable"])
        rp = load_related_parties(RP_PATH)

        params = SamplingParams(
            company_name="더미클라이언트A",
            period_end=date(2025, 12, 31),
            kind="receivable",
            performance_materiality=80_000_000,   # 총자산 4B × 2% = 80M
            risk_level="유의적위험",
            control_reliance="Y",
            related_parties=rp,
            force_include_related=True,
            random_seed=42,
        )
        return run_sampling(df, params), params

    def test_population_is_positive(self):
        """모집단 잔액이 0보다 커야 한다."""
        result, _ = self._run_receivable_sampling()
        assert result.population_amount > 0

    def test_decisions_have_correct_types(self):
        """decisions 목록에 Key item 또는 Representative 가 포함되어야 한다."""
        result, _ = self._run_receivable_sampling()
        final_sampled = [d for d in result.decisions if d.final_sampled]
        assert len(final_sampled) > 0, "최종 샘플링 결과가 0건"

    def test_related_party_forced_included(self):
        """특관자(카파홀딩스)는 final_sampled=True 이어야 한다."""
        result, _ = self._run_receivable_sampling()
        카파 = next((d for d in result.decisions if d.name == "카파홀딩스"), None)
        assert 카파 is not None, "카파홀딩스가 decisions에 없음"
        assert 카파.is_related_party, "카파홀딩스가 특관자로 분류 안됨"
        assert 카파.final_sampled, "카파홀딩스가 final_sampled=False"

    def test_size_result_consistent(self):
        """SizeResult의 final_sample_size가 1 이상이어야 한다."""
        result, _ = self._run_receivable_sampling()
        assert result.size_result.final_sample_size >= 1


# ── 6. Step 3 조서 생성 ─────────────────────────────────────────────────────
class TestDummyStep3Workpaper:
    def test_workpaper_generated(self, tmp_path):
        """dummy_client_a 채권 조서가 오류 없이 생성되어야 한다."""
        import openpyxl
        from src.infrastructure.loaders import load_related_parties
        from src.infrastructure.schemas.ledger_schema import detect_ledger_sheets
        from src.orchestrator import SamplingParams, run_sampling, write_report

        wb = openpyxl.load_workbook(LEDGER_PATH, read_only=True, data_only=True)
        sheet_map = detect_ledger_sheets(wb.sheetnames)
        wb.close()

        df = pd.read_excel(LEDGER_PATH, sheet_name=sheet_map["receivable"])
        rp = load_related_parties(RP_PATH)

        params = SamplingParams(
            company_name="더미클라이언트A",
            period_end=date(2025, 12, 31),
            kind="receivable",
            performance_materiality=80_000_000,
            risk_level="유의적위험",
            control_reliance="Y",
            related_parties=rp,
            force_include_related=True,
            random_seed=42,
            preparer="테스터",
            reviewer="검토자",
        )

        result = run_sampling(df, params)
        out_path = tmp_path / "C100_더미클라이언트A.xlsx"
        write_report(result, params, out_path)

        assert out_path.exists(), "조서 파일이 생성되지 않음"
        wb2 = openpyxl.load_workbook(out_path, read_only=True, data_only=True)
        assert len(wb2.sheetnames) >= 1, "조서 시트가 없음"
        wb2.close()

    def test_payable_workpaper_generated(self, tmp_path):
        """dummy_client_a 채무 조서도 생성되어야 한다."""
        import openpyxl
        from src.infrastructure.loaders import load_related_parties
        from src.infrastructure.schemas.ledger_schema import detect_ledger_sheets
        from src.orchestrator import SamplingParams, run_sampling, write_report

        wb = openpyxl.load_workbook(LEDGER_PATH, read_only=True, data_only=True)
        sheet_map = detect_ledger_sheets(wb.sheetnames)
        wb.close()

        df = pd.read_excel(LEDGER_PATH, sheet_name=sheet_map["payable"])
        rp = load_related_parties(RP_PATH)

        params = SamplingParams(
            company_name="더미클라이언트A",
            period_end=date(2025, 12, 31),
            kind="payable",
            performance_materiality=80_000_000,
            risk_level="유의적위험",
            control_reliance="Y",
            related_parties=rp,
            force_include_related=True,
            random_seed=42,
            preparer="테스터",
            reviewer="검토자",
        )

        result = run_sampling(df, params)
        out_path = tmp_path / "AA100_더미클라이언트A.xlsx"
        write_report(result, params, out_path)

        assert out_path.exists(), "채무 조서 파일이 생성되지 않음"


# ── 7. 범용성 시트명 변형 테스트 ────────────────────────────────────────────
class TestUniversalSheetVariants:
    """다양한 시트명 변형이 감지되는지 확인."""

    @pytest.mark.parametrize("sheet_name,expected_key", [
        ("채권", "receivable"),
        ("AR", "receivable"),
        ("매출원장", "receivable"),
        ("매출채권", "receivable"),
        ("receivable", "receivable"),
        ("채무", "payable"),
        ("AP", "payable"),
        ("매입원장", "payable"),
        ("매입채무", "payable"),
        ("payable", "payable"),
    ])
    def test_sheet_name_variants(self, sheet_name, expected_key):
        """시트명 변형들이 올바른 key로 감지되어야 한다."""
        from src.infrastructure.schemas.ledger_schema import detect_ledger_sheets

        # 두 시트(채권+채무) 페어를 만들어 감지
        if expected_key == "receivable":
            sheets = [sheet_name, "채무"]
        else:
            sheets = ["채권", sheet_name]

        result = detect_ledger_sheets(sheets)
        assert result[expected_key] == sheet_name, (
            f"'{sheet_name}' → '{expected_key}' 감지 실패: {result}"
        )

    @pytest.mark.parametrize("sheet_name", [
        "FS_M", "BS", "재무상태표", "재무제표", "Balance Sheet",
    ])
    def test_fs_sheet_variants(self, sheet_name):
        """재무제표 시트명 변형들이 감지되어야 한다."""
        from src.infrastructure.schemas.fs_schema import detect_fs_sheet

        result = detect_fs_sheet([sheet_name, "기타시트"])
        assert result == sheet_name, f"'{sheet_name}' FS 시트 감지 실패"

    @pytest.mark.parametrize("sheet_name", [
        "특관자리스트", "특수관계자리스트", "관계회사", "Related Parties", "특관자",
    ])
    def test_rp_sheet_variants(self, sheet_name):
        """특관자 시트명 변형들이 감지되어야 한다."""
        from src.infrastructure.schemas.rp_schema import detect_rp_sheet

        result = detect_rp_sheet([sheet_name, "기타시트"])
        assert result == sheet_name, f"'{sheet_name}' RP 시트 감지 실패"
