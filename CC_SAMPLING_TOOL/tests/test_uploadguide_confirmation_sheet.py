"""Task 3: '조회서' 시트 구조 — UploadGuide C1~C15 + 회신 컬럼 C16~C21."""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl
import pytest

from src.infrastructure.report.generic_reporter import (
    KindData, ReportContext,
    ConfirmationReplyInfo, AlternativeProcedureEntry,
    build_combined_report,
)
from src.infrastructure.loaders import UploadGuideData, PartyContact
from src.domain.mus import MUSResult, MUSSelection
from src.domain.population import CompletenessCheck, PartyDecision
from src.domain.sample_size import SampleSizeResult


def _ctx() -> ReportContext:
    return ReportContext(
        company_name="조회서테스트",
        period_end=date(2025, 12, 31),
        kind="both",
        preparer="작성자",
        reviewer="검토자",
    )


def _completeness() -> CompletenessCheck:
    return CompletenessCheck(
        by_group=[{"group": "외상매출금", "ledger": 1_000_000, "fs": 1_000_000, "diff": 0}],
        total_ledger=1_000_000, total_fs=1_000_000, total_diff=0,
    )


def _size() -> SampleSizeResult:
    return SampleSizeResult(
        key_item_threshold=100_000,
        key_item_ratio=0.5,
        confidence_factor=1.6,
        base_sample_size=5.0,
        final_sample_size=3,
        sample_interval=120_000,
        remaining_population=700_000,
    )


def _decisions() -> list[PartyDecision]:
    return [
        PartyDecision(name="(주)알파상사", balance=500_000, is_key_item=True,
                      is_representative=False, is_related_party=False,
                      is_excluded=False, final_sampled=True),
        PartyDecision(name="베타코퍼레이션", balance=200_000, is_key_item=False,
                      is_representative=True, is_related_party=False,
                      is_excluded=False, final_sampled=True),
        PartyDecision(name="감마인터내셔널", balance=100_000, is_key_item=False,
                      is_representative=True, is_related_party=False,
                      is_excluded=False, final_sampled=True),
    ]


def _mus() -> MUSResult:
    return MUSResult(
        sample_interval=120_000,
        random_start=60_000,
        selections=[
            MUSSelection(name="(주)알파상사", balance=500_000, cumulative=500_000,
                         selections=4, remainder_after=60_000, hit=True),
        ],
        sampled_names=["(주)알파상사"],
    )


def _kd() -> KindData:
    return KindData(
        ctx=_ctx(),
        completeness=_completeness(),
        size_result=_size(),
        decisions=_decisions(),
        mus_result=_mus(),
        performance_materiality=200_000,
        population_amount=1_000_000,
    )


def _upload_guide_data() -> UploadGuideData:
    """3개 거래처, 각 1건 계정과목."""
    return UploadGuideData(
        send_targets=[
            PartyContact(
                name="(주)알파상사",
                country="국내",
                business_no="123-45-67890",
                ceo_name="김대표",
                contact_person="이담당",
                phone="02-1234-5678",
                email="alpha@test.com",
                accounts=[("외상매출금", "KRW", 500_000)],
            ),
            PartyContact(
                name="베타코퍼레이션",
                country="국내",
                business_no="234-56-78901",
                ceo_name="박사장",
                contact_person="최담당",
                phone="02-2345-6789",
                email="beta@test.com",
                accounts=[("미지급금", "KRW", 200_000)],
            ),
            PartyContact(
                name="감마인터내셔널",
                country="해외",
                business_no="",
                ceo_name="John",
                contact_person="Jane",
                phone="+1-555-0100",
                email="",  # 이메일 없음 → 필수항목 미완료
                accounts=[("외상매출금", "USD", 100_000)],
            ),
        ],
        excluded=[],
    )


def _replies() -> list[ConfirmationReplyInfo]:
    return [
        ConfirmationReplyInfo(
            party_name="(주)알파상사",
            status="matched",
            extracted_balance=500_000,
            reply_date="2026-01-15",
        ),
    ]


def _alt_procs() -> list[AlternativeProcedureEntry]:
    return [
        AlternativeProcedureEntry(
            party_name="베타코퍼레이션",
            reason="미회신",
            ledger_balance=200_000,
            procedure_type="후속입금",
            evidence_names=["invoice_beta.pdf"],
            covered_amount=200_000,
            coverage_ratio=1.0,
            conclusion="충분",
        ),
    ]


def _build_report(tmp_path, ug=True) -> openpyxl.Workbook:
    out = tmp_path / "ug_confirmation.xlsx"
    build_combined_report(
        out,
        receivable=_kd(),
        payable=None,
        upload_guide_data=_upload_guide_data() if ug else None,
    )
    return openpyxl.load_workbook(out, data_only=True)


def test_confirmation_sheet_exists(tmp_path):
    """'조회서' 시트가 존재해야 한다."""
    wb = _build_report(tmp_path)
    assert "조회서" in wb.sheetnames, f"'조회서' 없음 — {wb.sheetnames}"
    wb.close()


def test_confirmation_sheet_21_columns(tmp_path):
    """'조회서' 시트는 C1~C21까지 21개 헤더 컬럼을 가져야 한다."""
    wb = _build_report(tmp_path)
    ws = wb["조회서"]
    # 헤더 행 탐색 (처음 15행 이내)
    header_row = None
    for r in range(1, 15):
        if ws.cell(r, 1).value == "채권채무구분":
            header_row = r
            break
    wb.close()

    assert header_row is not None, "헤더 행(채권채무구분) 미발견"
    # 21번째 컬럼 헤더 확인
    wb2 = _build_report(tmp_path)
    ws2 = wb2["조회서"]
    col21 = ws2.cell(header_row, 21).value
    wb2.close()
    assert col21 == "비고", f"21번째 컬럼 헤더 불일치: '{col21}'"


def test_confirmation_sheet_headers_c1_to_c15(tmp_path):
    """UploadGuide C1~C15 헤더가 순서대로 존재해야 한다."""
    expected_headers = [
        "채권채무구분", "계정과목명", "통화", "조회금액", "통화2", "조회금액2",
        "거래처명", "국가", "거래처구분", "사업자번호", "대표자명",
        "담당자명", "전화번호", "이메일", "필수항목",
    ]
    wb = _build_report(tmp_path)
    ws = wb["조회서"]

    header_row = None
    for r in range(1, 15):
        if ws.cell(r, 1).value == "채권채무구분":
            header_row = r
            break
    wb.close()

    assert header_row is not None, "헤더 행 미발견"

    wb2 = _build_report(tmp_path)
    ws2 = wb2["조회서"]
    actual_headers = [ws2.cell(header_row, c).value for c in range(1, 16)]
    wb2.close()

    for i, (exp, act) in enumerate(zip(expected_headers, actual_headers), 1):
        assert exp == act, f"C{i} 헤더 불일치: 기대='{exp}', 실제='{act}'"


def test_confirmation_sheet_reply_headers_c16_c21(tmp_path):
    """C16~C21 회신 컬럼 헤더: 회신상태, 회신금액, 차이, 일치여부, 회신일자, 비고."""
    expected = ["회신상태", "회신금액", "차이", "일치여부", "회신일자", "비고"]
    wb = _build_report(tmp_path)
    ws = wb["조회서"]

    header_row = None
    for r in range(1, 15):
        if ws.cell(r, 16).value == "회신상태":
            header_row = r
            break
    wb.close()

    assert header_row is not None, "회신상태 헤더 미발견"

    wb2 = _build_report(tmp_path)
    ws2 = wb2["조회서"]
    actual = [ws2.cell(header_row, c).value for c in range(16, 22)]
    wb2.close()

    for i, (exp, act) in enumerate(zip(expected, actual), 16):
        assert exp == act, f"C{i} 헤더 불일치: 기대='{exp}', 실제='{act}'"


def test_confirmation_sheet_party_names_in_data(tmp_path):
    """UploadGuide 거래처명이 '조회서' 시트 데이터에 포함되어야 한다."""
    wb = _build_report(tmp_path)
    ws = wb["조회서"]
    all_values = {ws.cell(r, c).value
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)}
    wb.close()

    assert "(주)알파상사" in all_values, "(주)알파상사 미포함"
    assert "베타코퍼레이션" in all_values, "베타코퍼레이션 미포함"
    assert "감마인터내셔널" in all_values, "감마인터내셔널 미포함"


def test_confirmation_sheet_no_uploadguide_fallback(tmp_path):
    """UploadGuide 없으면 '조회서'에 미제공 메시지가 표시되어야 한다."""
    wb = _build_report(tmp_path, ug=False)
    ws = wb["조회서"]
    all_values = [ws.cell(r, c).value
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)]
    wb.close()

    has_placeholder = any(
        isinstance(v, str) and "미제공" in v
        for v in all_values if v is not None
    )
    assert has_placeholder, "UploadGuide 미제공 시 플레이스홀더 메시지 없음"


def test_confirmation_sheet_reply_matched_shows_원본(tmp_path):
    """회신 matched 거래처는 회신상태='원본'으로 표시되어야 한다."""
    out = tmp_path / "ug_reply.xlsx"
    kd = _kd()
    kd.pdf_replies = _replies()
    build_combined_report(
        out,
        receivable=kd,
        payable=None,
        upload_guide_data=_upload_guide_data(),
    )
    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["조회서"]
    all_values = {ws.cell(r, c).value
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)}
    wb.close()

    assert "원본" in all_values, "회신 matched 거래처의 회신상태 '원본' 미표시"


def test_confirmation_sheet_alt_proc_shows_대체적(tmp_path):
    """대체적 절차 거래처는 회신상태='대체적'으로 표시되어야 한다."""
    out = tmp_path / "ug_alt.xlsx"
    kd = _kd()
    kd.alt_procedures = _alt_procs()
    build_combined_report(
        out,
        receivable=kd,
        payable=None,
        upload_guide_data=_upload_guide_data(),
    )
    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["조회서"]
    all_values = {ws.cell(r, c).value
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)}
    wb.close()

    assert "대체적" in all_values, "대체적 절차 거래처의 회신상태 '대체적' 미표시"


def test_confirmation_sheet_no_reply_shows_미회신(tmp_path):
    """회신 없는 거래처(감마인터내셔널)는 '미회신'으로 표시되어야 한다."""
    out = tmp_path / "ug_noreply.xlsx"
    build_combined_report(
        out,
        receivable=_kd(),
        payable=None,
        upload_guide_data=_upload_guide_data(),
    )
    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["조회서"]
    all_values = {ws.cell(r, c).value
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)}
    wb.close()

    assert "미회신" in all_values, "미회신 거래처 상태 '미회신' 미표시"
