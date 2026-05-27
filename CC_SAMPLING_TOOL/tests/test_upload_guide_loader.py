"""UploadGuide 로더 단위 테스트 — Sheet1 파싱, 발송제외 시트 파싱."""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl
import pytest

from src.infrastructure.loaders import (
    UploadGuideData,
    PartyContact,
    ExcludedParty,
    load_upload_guide,
)


# ─────────────────────────────────────────────────────────────
# 인메모리 픽스처 생성 헬퍼
# ─────────────────────────────────────────────────────────────

def _make_upload_guide_xlsx(tmp_path: Path) -> Path:
    """테스트용 UploadGuide xlsx 생성.

    Sheet1: 3개 거래처 (발송 대상)
    채채대상이나 발송대상제외: 1개 거래처 (발송 제외)
    """
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"

    # 헤더 행 (UploadGuide 실제 컬럼명 기준)
    headers = [
        "채권채무구분", "계정과목명", "통화1", "조회금액1",
        "통화2", "조회금액2", "거래처명", "국가",
        "거래처 구분", "사업자번호", "대표자명", "담당자명",
        "담당자 유선전화번호", "담당자 이메일", "필수항목 기재여부",
    ]
    ws1.append(headers)

    # 데이터 행
    ws1.append(["채권", "외상매출금", "KRW", 1_000_000, "", "", "거래처Alpha",
                "KR", "일반", "111-22-33333", "김대표", "이담당",
                "02-1111-2222", "alpha@example.com", "Y"])
    ws1.append(["채권", "미수금", "USD", 50_000, "KRW", 70_000_000, "거래처Beta",
                "US", "외국법인", "", "Smith", "Johnson",
                "+1-555-0100", "beta@example.com", "Y"])
    ws1.append(["채무", "외상매입금", "KRW", 500_000, "", "", "거래처Gamma",
                "CN", "일반", "222-33-44444", "王总", "张担当",
                "+86-10-1234", "gamma@example.com", "Y"])

    # 발송제외 시트
    ws2 = wb.create_sheet("채채대상이나 발송대상제외")
    excl_headers = ["채권채무구분", "계정과목명", "통화", "조회금액", "거래처명"]
    ws2.append(excl_headers)
    ws2.append(["채권", "외상매출금", "KRW", 200_000, "발송제외거래처X"])

    path = tmp_path / "UploadGuide_test.xlsx"
    wb.save(path)
    return path


# ─────────────────────────────────────────────────────────────
# 테스트
# ─────────────────────────────────────────────────────────────

def test_load_upload_guide_returns_correct_types(tmp_path):
    """load_upload_guide 반환 타입 확인."""
    path = _make_upload_guide_xlsx(tmp_path)
    result = load_upload_guide(path)
    assert isinstance(result, UploadGuideData)
    assert isinstance(result.send_targets, list)
    assert isinstance(result.excluded, list)


def test_send_targets_count(tmp_path):
    """Sheet1 에서 3개 거래처 파싱."""
    path = _make_upload_guide_xlsx(tmp_path)
    result = load_upload_guide(path)
    assert len(result.send_targets) == 3, f"send_targets 수 오류: {len(result.send_targets)}"


def test_send_target_fields(tmp_path):
    """거래처Alpha 필드 정확성 확인."""
    path = _make_upload_guide_xlsx(tmp_path)
    result = load_upload_guide(path)
    contact_map = result.contact_map()

    assert "거래처Alpha" in contact_map
    ct = contact_map["거래처Alpha"]
    assert ct.country == "KR"
    assert ct.business_no == "111-22-33333"
    assert ct.ceo_name == "김대표"
    assert ct.contact_person == "이담당"
    assert ct.phone == "02-1111-2222"
    assert ct.email == "alpha@example.com"


def test_send_target_accounts(tmp_path):
    """거래처Alpha 계정과목 정보 파싱."""
    path = _make_upload_guide_xlsx(tmp_path)
    result = load_upload_guide(path)
    ct = result.contact_map()["거래처Alpha"]
    # (계정과목명, 통화, 금액) 튜플 존재 확인
    assert len(ct.accounts) >= 1
    acct_names = [a[0] for a in ct.accounts]
    assert "외상매출금" in acct_names


def test_excluded_parties_count(tmp_path):
    """발송제외 시트에서 1개 파싱."""
    path = _make_upload_guide_xlsx(tmp_path)
    result = load_upload_guide(path)
    assert len(result.excluded) == 1, f"excluded 수 오류: {len(result.excluded)}"


def test_excluded_party_name(tmp_path):
    """발송제외 거래처명 정확성 확인."""
    path = _make_upload_guide_xlsx(tmp_path)
    result = load_upload_guide(path)
    excl_names = result.excluded_names()
    assert "발송제외거래처X" in excl_names


def test_excluded_names_set(tmp_path):
    """excluded_names() → set 반환."""
    path = _make_upload_guide_xlsx(tmp_path)
    result = load_upload_guide(path)
    names = result.excluded_names()
    assert isinstance(names, set)


def test_contact_map_fast_lookup(tmp_path):
    """contact_map() — 거래처명으로 빠른 조회."""
    path = _make_upload_guide_xlsx(tmp_path)
    result = load_upload_guide(path)
    cmap = result.contact_map()
    assert "거래처Beta" in cmap
    assert "거래처Gamma" in cmap


def test_empty_file_returns_empty_data(tmp_path):
    """빈 Excel — 예외 없이 빈 UploadGuideData 반환."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["채권채무구분", "계정과목명", "통화1", "조회금액1", "거래처명"])  # 헤더만
    path = tmp_path / "empty_ug.xlsx"
    wb.save(path)

    result = load_upload_guide(path)
    assert len(result.send_targets) == 0
    assert len(result.excluded) == 0


def test_no_exclusion_sheet(tmp_path):
    """발송제외 시트 없음 — excluded 비어있음."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["채권채무구분", "계정과목명", "통화1", "조회금액1",
                "통화2", "조회금액2", "거래처명", "국가",
                "거래처 구분", "사업자번호", "대표자명", "담당자명",
                "담당자 유선전화번호", "담당자 이메일"])
    ws1.append(["채권", "외상매출금", "KRW", 100_000, "", "", "거래처Z",
                "KR", "일반", "", "", "", "", "z@example.com"])
    path = tmp_path / "no_excl.xlsx"
    wb.save(path)

    result = load_upload_guide(path)
    assert len(result.send_targets) == 1
    assert len(result.excluded) == 0
