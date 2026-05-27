"""
test_visual_fidelity.py — 7620 양식 시각 정합성 검증

검증 항목:
  1. 모든 시트의 데이터 셀 폰트가 "맑은 고딕"
  2. C100-1 PM·Key item 기준금액 셀 → FILL_INPUT_YELLOW (FFFFCC)
  3. C100 조회서 헤더 행 → navy 배경(44546A) + 흰 글씨(FFFFFF)
  4. 컬럼 너비 — C100-1 A=33.85, B=19.28
  5. 조서번호 셀 I2 → 빨강 글씨 (FF0000) + bold
  6. 소제목 행 → D6DCE5 배경
  7. Key item 행 → FFF2CC 배경
  8. MUS hit 행 → C6EFCE 배경
  9. 합계 행 → D6DCE5 배경 + bold
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl
import pytest

from src.infrastructure.report.generic_reporter import (
    AlternativeProcedureEntry,
    ConfirmationReplyInfo,
    ExclusionRow,
    PartyContactInfo,
    ReportContext,
    build_generic_report,
    FONT_NAME,
    FILL_INPUT_YELLOW,
    FILL_HEADER_NAVY,
    FILL_SUBHEADER,
    FILL_KEY_ITEM,
    FILL_SAMPLED,
    FILL_TOTAL_ROW,
    FONT_RED_BOLD,
)
from src.domain.mus import MUSResult, MUSSelection
from src.domain.population import CompletenessCheck, PartyDecision
from src.domain.sample_size import SampleSizeResult


# ─────────────────────────────────────────────────────────────
# 공통 픽스처
# ─────────────────────────────────────────────────────────────

def _ctx() -> ReportContext:
    return ReportContext(
        company_name="시각테스트회사",
        period_end=date(2025, 12, 31),
        kind="receivable",
        preparer="홍길동",
        reviewer="이검토",
        prep_date=date(2026, 1, 10),
        review_date=date(2026, 1, 15),
        workpaper_no_prefix="C100",
    )


def _completeness() -> CompletenessCheck:
    return CompletenessCheck(
        by_group=[
            {"group": "외상매출금", "ledger": 5_000_000, "fs": 5_000_000, "diff": 0, "note": ""},
        ],
        total_ledger=5_000_000,
        total_fs=5_000_000,
        total_diff=0,
    )


def _size_result() -> SampleSizeResult:
    return SampleSizeResult(
        key_item_threshold=500_000,
        key_item_ratio=0.5,
        confidence_factor=1.4,
        base_sample_size=8,
        final_sample_size=8,
        sample_interval=600_000,
        remaining_population=4_000_000,
    )


def _decisions() -> list[PartyDecision]:
    return [
        PartyDecision(
            name="대형거래처A", balance=1_200_000,
            is_key_item=True, is_representative=False,
            is_related_party=False, is_excluded=False,
            final_sampled=True,
            by_account={"외상매출금": 1_200_000},
        ),
        PartyDecision(
            name="대형거래처B", balance=800_000,
            is_key_item=True, is_representative=False,
            is_related_party=False, is_excluded=False,
            final_sampled=True,
            by_account={"외상매출금": 800_000},
        ),
        PartyDecision(
            name="중소거래처C", balance=350_000,
            is_key_item=False, is_representative=True,
            is_related_party=False, is_excluded=False,
            final_sampled=True,
            by_account={"미수금": 350_000},
        ),
        PartyDecision(
            name="제외거래처D", balance=100_000,
            is_key_item=False, is_representative=False,
            is_related_party=False, is_excluded=True,
            final_sampled=False,
            exclusion_reason="해산",
            by_account={"외상매출금": 100_000},
        ),
    ]


def _mus_result() -> MUSResult:
    return MUSResult(
        sample_interval=600_000,
        random_start=150_000,
        selections=[
            MUSSelection(
                name="중소거래처C", balance=350_000, cumulative=350_000,
                selections=1, remainder_after=200_000, hit=True,
            ),
            MUSSelection(
                name="소형거래처E", balance=80_000, cumulative=430_000,
                selections=0, remainder_after=120_000, hit=False,
            ),
        ],
        sampled_names=["중소거래처C"],
    )


def _build(tmp_path: Path) -> openpyxl.Workbook:
    out = tmp_path / "visual_test.xlsx"
    build_generic_report(
        out_path=out,
        ctx=_ctx(),
        completeness=_completeness(),
        size_result=_size_result(),
        decisions=_decisions(),
        mus_result=_mus_result(),
        performance_materiality=1_000_000,
        population_amount=5_000_000,
        contacts=[
            PartyContactInfo(
                name="대형거래처A",
                country="KR",
                contact_person="김담당",
                email="a@example.com",
            )
        ],
    )
    return openpyxl.load_workbook(out)


# ─────────────────────────────────────────────────────────────
# 1. 폰트 이름 검증
# ─────────────────────────────────────────────────────────────

def test_all_data_cells_use_malgun_gothic(tmp_path):
    """생성된 모든 시트 데이터 셀의 폰트 이름이 '맑은 고딕'이어야 한다."""
    wb = _build(tmp_path)
    violations = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and cell.font and cell.font.name:
                    if cell.font.name != FONT_NAME:
                        violations.append(
                            f"{sheet_name}!{cell.coordinate}: font={cell.font.name}"
                        )
    assert not violations, f"비맑은고딕 셀 발견: {violations[:10]}"


# ─────────────────────────────────────────────────────────────
# 2. 입력값 강조 셀 (PM, Key item 기준금액) — FFFFCC
# ─────────────────────────────────────────────────────────────

def test_c100_1_pm_cell_has_input_yellow_fill(tmp_path):
    """C100-1 시트: PM 값 셀이 FFFFCC(연노랑) 배경이어야 한다."""
    wb = _build(tmp_path)
    ws = wb["C100-1 표본규모 결정"]

    # 수행중요성(PM) 라벨이 있는 행의 B열 값 셀 확인
    found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "수행중요성 (PM)":
                # 같은 행 B열
                val_cell = ws.cell(cell.row, 2)
                assert val_cell.fill.fgColor.rgb.upper().endswith("FFFFCC"), (
                    f"PM 셀 배경색 불일치: {val_cell.fill.fgColor.rgb}"
                )
                found = True
    assert found, "C100-1에 '수행중요성 (PM)' 라벨이 없음"


def test_c100_1_key_item_threshold_cell_has_input_yellow(tmp_path):
    """C100-1 시트: Key item 기준금액 셀이 FFFFCC 배경이어야 한다."""
    wb = _build(tmp_path)
    ws = wb["C100-1 표본규모 결정"]

    found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "Key item 기준금액" in str(cell.value) and "PM" in str(cell.value):
                val_cell = ws.cell(cell.row, 2)
                assert val_cell.fill.fgColor.rgb.upper().endswith("FFFFCC"), (
                    f"Key item 기준금액 셀 배경 불일치: {val_cell.fill.fgColor.rgb}"
                )
                found = True
    assert found, "C100-1에 'Key item 기준금액' 라벨이 없음"


# ─────────────────────────────────────────────────────────────
# 3. Control sheet 헤더 — navy(44546A) + 흰 글씨(FFFFFF)
# ─────────────────────────────────────────────────────────────

def test_c100_control_sheet_header_navy_bg(tmp_path):
    """C100 조회서: 테이블 헤더 행 배경이 44546A(navy)이어야 한다."""
    wb = _build(tmp_path)
    ws = wb["C100 조회서"]

    navy_found = False
    for row in ws.iter_rows():
        for cell in row:
            if (cell.fill and cell.fill.fgColor
                    and cell.fill.fgColor.rgb.upper().endswith("44546A")):
                navy_found = True
                # 흰 글씨 확인
                if cell.font and cell.font.color:
                    assert cell.font.color.rgb.upper().endswith("FFFFFF"), (
                        f"헤더 글씨 색 불일치: {cell.font.color.rgb}"
                    )
    assert navy_found, "C100 조회서에 navy 헤더 셀이 없음"


def test_c100_1_header_navy(tmp_path):
    """C100-1 표본규모 결정: 테이블 헤더가 navy 배경이어야 한다."""
    wb = _build(tmp_path)
    ws = wb["C100-1 표본규모 결정"]
    navy_found = any(
        cell.fill and cell.fill.fgColor
        and cell.fill.fgColor.rgb.upper().endswith("44546A")
        for row in ws.iter_rows()
        for cell in row
    )
    assert navy_found, "C100-1에 navy 헤더 셀이 없음"


# ─────────────────────────────────────────────────────────────
# 4. 컬럼 너비 검증
# ─────────────────────────────────────────────────────────────

def test_c100_1_column_widths(tmp_path):
    """C100-1: A열 ≈ 33.85, B열 ≈ 19.28 (오차 ±0.5)."""
    wb = _build(tmp_path)
    ws = wb["C100-1 표본규모 결정"]

    w_a = ws.column_dimensions["A"].width
    w_b = ws.column_dimensions["B"].width

    assert abs(w_a - 33.85) < 0.5, f"A열 너비 불일치: {w_a}"
    assert abs(w_b - 19.28) < 0.5, f"B열 너비 불일치: {w_b}"


def test_c100_control_col_b_width(tmp_path):
    """C100 조회서: B열(거래처명) 너비 ≈ 33.85."""
    wb = _build(tmp_path)
    ws = wb["C100 조회서"]
    w_b = ws.column_dimensions["B"].width
    assert abs(w_b - 33.85) < 0.5, f"B열 너비 불일치: {w_b}"


def test_c100_control_col_a_width(tmp_path):
    """C100 조회서: A열(No) 너비 ≈ 4.0."""
    wb = _build(tmp_path)
    ws = wb["C100 조회서"]
    w_a = ws.column_dimensions["A"].width
    assert abs(w_a - 4.0) < 0.5, f"A열 너비 불일치: {w_a}"


# ─────────────────────────────────────────────────────────────
# 5. 조서번호 셀 — I2 빨강(FF0000) + bold
# ─────────────────────────────────────────────────────────────

def test_workpaper_no_cell_red_bold(tmp_path):
    """C100-1 시트 I2 셀: 조서번호 값이 빨강(FF0000) + bold이어야 한다."""
    wb = _build(tmp_path)
    ws = wb["C100-1 표본규모 결정"]

    cell = ws["I2"]
    assert cell.value == "C100-1", f"I2 값 불일치: {cell.value}"
    assert cell.font.bold, "조서번호 셀이 bold가 아님"
    assert cell.font.color.rgb.upper().endswith("FF0000"), (
        f"조서번호 글씨 색 불일치: {cell.font.color.rgb}"
    )


def test_summary_sheet_wp_no_red(tmp_path):
    """샘플링 요약 시트 I2 셀: 조서번호가 빨강이어야 한다."""
    wb = _build(tmp_path)
    ws = wb["샘플링 요약"]
    cell = ws["I2"]
    assert cell.font.color.rgb.upper().endswith("FF0000"), (
        f"요약 시트 I2 글씨색 불일치: {cell.font.color.rgb}"
    )


# ─────────────────────────────────────────────────────────────
# 6. 소제목 행 — D6DCE5 배경
# ─────────────────────────────────────────────────────────────

def test_subheader_fill_d6dce5(tmp_path):
    """C100-1 소제목 행(감사목적 등)이 D6DCE5 배경이어야 한다."""
    wb = _build(tmp_path)
    ws = wb["C100-1 표본규모 결정"]

    subheader_found = any(
        cell.fill and cell.fill.fgColor
        and cell.fill.fgColor.rgb.upper().endswith("D6DCE5")
        for row in ws.iter_rows()
        for cell in row
        if cell.value and "감사목적" in str(cell.value)
    )
    assert subheader_found, "C100-1 '감사목적' 소제목 행에 D6DCE5 배경 없음"


# ─────────────────────────────────────────────────────────────
# 7. Key item 행 — FFF2CC 배경
# ─────────────────────────────────────────────────────────────

def test_key_item_row_fill_fff2cc(tmp_path):
    """C100 조회서: Key item 거래처 행이 FFF2CC 배경이어야 한다."""
    wb = _build(tmp_path)
    ws = wb["C100 조회서"]

    # "대형거래처A" or "대형거래처B" 가 FFF2CC 배경인지 확인
    ki_fill_found = any(
        cell.fill and cell.fill.fgColor
        and cell.fill.fgColor.rgb.upper().endswith("FFF2CC")
        for row in ws.iter_rows()
        for cell in row
        if cell.value in ("대형거래처A", "대형거래처B")
    )
    assert ki_fill_found, "C100 조회서 Key item 행에 FFF2CC 배경 없음"


# ─────────────────────────────────────────────────────────────
# 8. MUS hit 행 — C6EFCE 배경
# ─────────────────────────────────────────────────────────────

def test_mus_hit_row_fill_c6efce(tmp_path):
    """C100-3: MUS hit=True 행이 C6EFCE 배경이어야 한다."""
    wb = _build(tmp_path)
    ws = wb["C100-3 표본 추출(MUS)"]

    hit_fill_found = any(
        cell.fill and cell.fill.fgColor
        and cell.fill.fgColor.rgb.upper().endswith("C6EFCE")
        for row in ws.iter_rows()
        for cell in row
        if cell.value == "중소거래처C"
    )
    assert hit_fill_found, "C100-3 MUS hit 행에 C6EFCE 배경 없음"


# ─────────────────────────────────────────────────────────────
# 9. 합계 행 — D6DCE5 배경 + bold
# ─────────────────────────────────────────────────────────────

def test_total_row_fill_and_bold(tmp_path):
    """C100 조회서: '총합계' 셀이 D6DCE5 배경 + bold이어야 한다."""
    wb = _build(tmp_path)
    ws = wb["C100 조회서"]

    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "총합계":
                assert cell.font and cell.font.bold, "총합계 셀이 bold가 아님"
                assert (cell.fill and cell.fill.fgColor
                        and cell.fill.fgColor.rgb.upper().endswith("D6DCE5")), (
                    f"총합계 셀 배경 불일치: {cell.fill.fgColor.rgb if cell.fill else None}"
                )
                return
    pytest.fail("C100 조회서에 '총합계' 셀이 없음")


# ─────────────────────────────────────────────────────────────
# 10. 헤더 블록 구조 검증 (R1~R3)
# ─────────────────────────────────────────────────────────────

def test_doc_header_structure(tmp_path):
    """C100-1: A1=회사명, A2=제목(bold), I1=조서번호 라벨, I2=C100-1(빨강)."""
    wb = _build(tmp_path)
    ws = wb["C100-1 표본규모 결정"]

    assert ws["A1"].value == "시각테스트회사", f"A1 값: {ws['A1'].value}"
    assert ws["A2"].font.bold, "A2 제목 셀이 bold가 아님"
    assert ws["I1"].value == "조서번호", f"I1 값: {ws['I1'].value}"
    assert ws["I2"].value == "C100-1", f"I2 값: {ws['I2'].value}"


def test_doc_header_preparer_reviewer(tmp_path):
    """C100-1: E1='작성자:', E2='검토자:', F1=작성자명, F2=검토자명."""
    wb = _build(tmp_path)
    ws = wb["C100-1 표본규모 결정"]

    assert ws["E1"].value == "작성자:", f"E1: {ws['E1'].value}"
    assert ws["E2"].value == "검토자:", f"E2: {ws['E2'].value}"
    assert ws["F1"].value == "홍길동", f"F1: {ws['F1'].value}"
    assert ws["F2"].value == "이검토", f"F2: {ws['F2'].value}"


# ─────────────────────────────────────────────────────────────
# 11. 숫자 포맷 검증
# ─────────────────────────────────────────────────────────────

def test_amount_cells_use_numfmt_int(tmp_path):
    """C100-1 표본규모 결정: 금액 셀에 NUMFMT_INT 포맷 적용 여부 확인."""
    from src.infrastructure.report.generic_reporter import NUMFMT_INT
    wb = _build(tmp_path)
    ws = wb["C100-1 표본규모 결정"]

    # PM 값 셀 포맷 확인
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "수행중요성 (PM)":
                val_cell = ws.cell(cell.row, 2)
                assert val_cell.number_format == NUMFMT_INT, (
                    f"PM 셀 포맷 불일치: {val_cell.number_format}"
                )
                return
    pytest.fail("PM 셀 미발견")


# ─────────────────────────────────────────────────────────────
# 12. 발송제외 행 — F4CCCC 배경
# ─────────────────────────────────────────────────────────────

def test_excluded_row_fill_in_c100_2(tmp_path):
    """C100-2: 발송제외 거래처 행이 F4CCCC(연빨강) 배경이어야 한다."""
    wb = _build(tmp_path)
    ws = wb["C100-2 Key item 추출"]

    excl_fill_found = any(
        cell.fill and cell.fill.fgColor
        and cell.fill.fgColor.rgb.upper().endswith("F4CCCC")
        for row in ws.iter_rows()
        for cell in row
        if cell.value == "제외거래처D"
    )
    assert excl_fill_found, "C100-2 발송제외 거래처 행에 F4CCCC 배경 없음"
