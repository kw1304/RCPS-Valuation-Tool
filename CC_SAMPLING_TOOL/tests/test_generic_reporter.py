"""generic_reporter.py — 9시트 생성·데이터 정합성 단위 테스트."""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl
import pytest

from src.infrastructure.report.generic_reporter import (
    AlternativeProcedureEntry,
    ConfirmationReplyInfo,
    ExclusionRow,
    PartyContactInfo,
    ReportContext,
    EXPECTED_GENERIC_SHEETS,
    EXPECTED_SHEET_COUNT,
    build_generic_report,
)
from src.domain.mus import MUSResult, MUSSelection
from src.domain.population import CompletenessCheck, PartyDecision
from src.domain.sample_size import SampleSizeResult


# ─────────────────────────────────────────────────────────────
# 픽스처
# ─────────────────────────────────────────────────────────────

def _make_ctx(kind="receivable") -> ReportContext:
    return ReportContext(
        company_name="테스트회사",
        period_end=date(2025, 12, 31),
        kind=kind,
        preparer="작성자A",
        reviewer="검토자B",
        workpaper_no_prefix="C100" if kind != "payable" else "AA100",
    )


def _make_completeness() -> CompletenessCheck:
    return CompletenessCheck(
        by_group=[
            {"group": "외상매출금", "ledger": 1_000_000, "fs": 1_000_000, "diff": 0, "note": ""},
            {"group": "미수금",     "ledger": 500_000,   "fs": 490_000,   "diff": 10_000, "note": "조정"},
        ],
        total_ledger=1_500_000,
        total_fs=1_490_000,
        total_diff=10_000,
    )


def _make_size_result() -> SampleSizeResult:
    return SampleSizeResult(
        key_item_threshold=300_000,
        key_item_ratio=0.5,
        confidence_factor=1.4,
        base_sample_size=10,
        final_sample_size=10,
        sample_interval=120_000,
        remaining_population=1_200_000,
    )


def _make_decisions() -> list[PartyDecision]:
    return [
        PartyDecision(
            name="거래처A", balance=500_000,
            is_key_item=True, is_representative=False,
            is_related_party=False, is_excluded=False,
            final_sampled=True,
            by_account={"외상매출금": 500_000},
        ),
        PartyDecision(
            name="거래처B", balance=200_000,
            is_key_item=False, is_representative=True,
            is_related_party=False, is_excluded=False,
            final_sampled=True,
            by_account={"미수금": 200_000},
        ),
        PartyDecision(
            name="거래처C", balance=150_000,
            is_key_item=False, is_representative=False,
            is_related_party=False, is_excluded=True,
            final_sampled=False,
            exclusion_reason="해산",
            by_account={"외상매출금": 150_000},
        ),
    ]


def _make_mus_result() -> MUSResult:
    return MUSResult(
        sample_interval=120_000,
        random_start=45_000,
        selections=[
            MUSSelection(
                name="거래처B", balance=200_000, cumulative=200_000,
                selections=1, remainder_after=35_000, hit=True,
            ),
        ],
        sampled_names=["거래처B"],
    )


# ─────────────────────────────────────────────────────────────
# 테스트
# ─────────────────────────────────────────────────────────────

def test_build_generic_report_creates_9_sheets(tmp_path):
    """build_generic_report → 정확히 9개 시트 생성."""
    out = tmp_path / "test_workpaper.xlsx"
    build_generic_report(
        out_path=out,
        ctx=_make_ctx(),
        completeness=_make_completeness(),
        size_result=_make_size_result(),
        decisions=_make_decisions(),
        mus_result=_make_mus_result(),
        performance_materiality=200_000,
        population_amount=1_500_000,
    )
    assert out.exists(), "출력 파일 미생성"

    wb = openpyxl.load_workbook(out, read_only=True)
    sheets = set(wb.sheetnames)
    wb.close()

    assert len(sheets) == EXPECTED_SHEET_COUNT, f"시트 수 불일치: {len(sheets)} — {sheets}"


def test_build_generic_report_sheet_names_correct(tmp_path):
    """10개 시트 이름이 기대값과 일치."""
    out = tmp_path / "test_sheets.xlsx"
    build_generic_report(
        out_path=out,
        ctx=_make_ctx(),
        completeness=_make_completeness(),
        size_result=_make_size_result(),
        decisions=_make_decisions(),
        mus_result=_make_mus_result(),
        performance_materiality=200_000,
        population_amount=1_500_000,
    )

    wb = openpyxl.load_workbook(out, read_only=True)
    actual = set(wb.sheetnames)
    wb.close()

    missing = EXPECTED_GENERIC_SHEETS - actual
    assert not missing, f"누락 시트: {missing}"


def test_build_generic_report_no_template_copy(tmp_path):
    """템플릿 복사 없이 생성 — 21개 시트를 갖지 않음."""
    out = tmp_path / "test_no_template.xlsx"
    build_generic_report(
        out_path=out,
        ctx=_make_ctx(),
        completeness=_make_completeness(),
        size_result=_make_size_result(),
        decisions=_make_decisions(),
        mus_result=_make_mus_result(),
        performance_materiality=200_000,
        population_amount=1_500_000,
    )

    wb = openpyxl.load_workbook(out, read_only=True)
    sheet_count = len(wb.sheetnames)
    wb.close()

    assert sheet_count < 21, "21개 시트 — 템플릿 복사 방식이 잔존할 가능성"


def test_pm_value_in_workbook(tmp_path):
    """PM 값이 생성된 조서(표본규모 산출 또는 요약)에 존재."""
    out = tmp_path / "test_summary.xlsx"
    pm = 200_000
    build_generic_report(
        out_path=out,
        ctx=_make_ctx(),
        completeness=_make_completeness(),
        size_result=_make_size_result(),
        decisions=_make_decisions(),
        mus_result=_make_mus_result(),
        performance_materiality=pm,
        population_amount=1_500_000,
    )

    wb = openpyxl.load_workbook(out, data_only=True)
    # PM은 표본규모 산출 시트 또는 요약 시트에 존재
    found = False
    for sheet_name in ["표본규모 산출", "요약"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_values = [ws.cell(r, c).value for r in range(1, ws.max_row + 1)
                          for c in range(1, ws.max_column + 1)]
            if pm in all_values:
                found = True
                break
    wb.close()

    assert found, f"PM({pm}) 값이 조서 어느 시트에도 없음"


def test_confirmation_sheet_has_final_sampled_parties(tmp_path):
    """조회서 — final_sampled=True 거래처만 seq 번호 표시."""
    out = tmp_path / "test_confirmation.xlsx"
    build_generic_report(
        out_path=out,
        ctx=_make_ctx(),
        completeness=_make_completeness(),
        size_result=_make_size_result(),
        decisions=_make_decisions(),
        mus_result=_make_mus_result(),
        performance_materiality=200_000,
        population_amount=1_500_000,
    )

    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["샘플링 거래처 내역"]
    all_values = [ws.cell(r, c).value for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)]
    wb.close()

    assert "거래처A" in all_values, "final_sampled 거래처A 미포함"
    assert "거래처B" in all_values, "final_sampled 거래처B 미포함"


def test_completeness_sheet_has_data(tmp_path):
    """모집단 완전성 시트 — 완전성 데이터 반영."""
    out = tmp_path / "test_completeness.xlsx"
    build_generic_report(
        out_path=out,
        ctx=_make_ctx(),
        completeness=_make_completeness(),
        size_result=_make_size_result(),
        decisions=_make_decisions(),
        mus_result=_make_mus_result(),
        performance_materiality=200_000,
        population_amount=1_500_000,
    )

    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["모집단 완전성"]
    all_values = [ws.cell(r, c).value for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)]
    wb.close()

    assert "외상매출금" in all_values, "외상매출금 그룹 미표시"
    assert 1_000_000 in all_values, "외상매출금 잔액 1,000,000 미반영"


def test_mus_sheet_has_data(tmp_path):
    """MUS 추출 내역 시트 — MUS 추출 결과 반영."""
    out = tmp_path / "test_mus.xlsx"
    mus = _make_mus_result()
    build_generic_report(
        out_path=out,
        ctx=_make_ctx(),
        completeness=_make_completeness(),
        size_result=_make_size_result(),
        decisions=_make_decisions(),
        mus_result=mus,
        performance_materiality=200_000,
        population_amount=1_500_000,
    )

    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["MUS 추출 내역"]
    all_values = [ws.cell(r, c).value for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)]
    wb.close()

    assert "거래처B" in all_values, "MUS hit 거래처B 미포함"
    assert mus.random_start in all_values, f"임의출발점({mus.random_start}) 미반영"


def test_contacts_in_address_sheet(tmp_path):
    """UploadGuide 연락처 → 주소 적정성 시트 반영."""
    out = tmp_path / "test_address.xlsx"
    contacts = [
        PartyContactInfo(
            name="거래처A", country="KR", business_no="123-45-67890",
            ceo_name="김대표", contact_person="이담당", phone="02-1234-5678",
            email="test@example.com",
        )
    ]
    build_generic_report(
        out_path=out,
        ctx=_make_ctx(),
        completeness=_make_completeness(),
        size_result=_make_size_result(),
        decisions=_make_decisions(),
        mus_result=_make_mus_result(),
        performance_materiality=200_000,
        population_amount=1_500_000,
        contacts=contacts,
    )

    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["주소 적정성"]
    all_values = [ws.cell(r, c).value for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)]
    wb.close()

    assert "test@example.com" in all_values, "이메일 미반영"
    assert "이담당" in all_values, "담당자명 미반영"


def test_alt_procedures_in_sheet(tmp_path):
    """대체적 절차 데이터 → 대체적 절차 시트 반영."""
    out = tmp_path / "test_alt.xlsx"
    alt = [
        AlternativeProcedureEntry(
            party_name="거래처D", reason="미회신",
            ledger_balance=100_000, procedure_type="후속입금",
            evidence_names=["invoice_001.pdf"],
            covered_amount=100_000, coverage_ratio=1.0,
            conclusion="충분", auditor_notes="후속 입금 확인",
        )
    ]
    build_generic_report(
        out_path=out,
        ctx=_make_ctx(),
        completeness=_make_completeness(),
        size_result=_make_size_result(),
        decisions=_make_decisions(),
        mus_result=_make_mus_result(),
        performance_materiality=200_000,
        population_amount=1_500_000,
        alt_procedures=alt,
    )

    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["대체적 절차"]
    all_values = [ws.cell(r, c).value for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)]
    wb.close()

    assert "거래처D" in all_values, "대체적 절차 거래처D 미포함"
    assert "충분" in all_values, "결론 '충분' 미반영"


def test_payable_prefix_generates_9_sheets(tmp_path):
    """채무(payable) 조서 — 9시트 생성."""
    out = tmp_path / "test_payable.xlsx"
    ctx = _make_ctx(kind="payable")
    build_generic_report(
        out_path=out,
        ctx=ctx,
        completeness=_make_completeness(),
        size_result=_make_size_result(),
        decisions=_make_decisions(),
        mus_result=_make_mus_result(),
        performance_materiality=200_000,
        population_amount=1_500_000,
    )

    wb = openpyxl.load_workbook(out, read_only=True)
    sheets = set(wb.sheetnames)
    wb.close()

    assert len(sheets) == EXPECTED_SHEET_COUNT, f"채무 조서 시트 수 불일치: {len(sheets)}"
    assert EXPECTED_GENERIC_SHEETS == sheets, f"시트 이름 불일치: {sheets}"
