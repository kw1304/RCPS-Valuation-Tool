"""Step 3 kind='both' — 채권+채무 통합 조서 생성 테스트."""
from __future__ import annotations

import io
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl
import pytest

from api.app import app as flask_app


LEDGER_PATH = ROOT / "input" / "회사자료" / "채권채무조회서 거래처별 원장.XLSX"


@pytest.fixture(scope="module")
def client():
    flask_app.config["TESTING"] = True
    with flask_app.test_client() as c:
        yield c


@pytest.fixture(scope="module")
def both_project(client):
    """채권+채무 프로젝트 생성 → 업로드 → 채권 Step1."""
    if not LEDGER_PATH.exists():
        pytest.skip("원장 파일 없음")

    r = client.post("/api/project", json={
        "company_name": "BothKindTest",
        "period_end": "2025-12-31",
        "kind": "both",
    })
    assert r.status_code == 201
    pid = r.json["id"]

    with open(LEDGER_PATH, "rb") as f:
        data = {"ledger": (io.BytesIO(f.read()), "ledger.xlsx")}
    r = client.post(f"/api/upload?project_id={pid}",
                    data=data, content_type="multipart/form-data")
    sheet_map = r.json.get("sheet_map", {})
    ar_sheet = sheet_map.get("receivable", "채권")
    ap_sheet = sheet_map.get("payable", "채무")

    # 채권 Step 1
    r = client.post("/api/run", json={
        "kind": "receivable",
        "sheet": ar_sheet,
        "project_id": pid,
        "company_name": "BothKindTest",
        "period_end": "2025-12-31",
        "performance_materiality": 2_738_000_000,
        "risk_level": "유의적위험",
        "control_reliance": "Y",
        "seed": 42,
    })
    assert r.status_code == 200, r.json

    return {"pid": pid, "ar_sheet": ar_sheet, "ap_sheet": ap_sheet}


def test_step3_receivable_has_9_sheets(client, both_project, tmp_path):
    """채권 조서 — 9개 시트 (generic reporter)."""
    pid = both_project["pid"]
    r = client.post(f"/api/project/{pid}/step3/build", json={
        "kind": "receivable",
        "preparer": "테스터",
        "reviewer": "검토자",
    })
    assert r.status_code == 201, r.json

    r_dl = client.get(f"/api/project/{pid}/step3/download/receivable")
    assert r_dl.status_code == 200

    out = tmp_path / "ar_workpaper.xlsx"
    out.write_bytes(r_dl.data)
    wb = openpyxl.load_workbook(out, read_only=True)
    sheets = wb.sheetnames
    wb.close()

    assert len(sheets) == 9, f"채권 조서 시트 수 불일치: {len(sheets)} — {sheets}"


def test_step3_receivable_has_요약_sheet(client, both_project, tmp_path):
    """채권 조서 — '요약' 시트 존재 (generic reporter)."""
    pid = both_project["pid"]
    r_dl = client.get(f"/api/project/{pid}/step3/download/receivable")
    assert r_dl.status_code == 200

    out = tmp_path / "ar_workpaper2.xlsx"
    out.write_bytes(r_dl.data)
    wb = openpyxl.load_workbook(out, read_only=True)
    sheets = wb.sheetnames
    wb.close()

    assert "요약" in sheets, f"'요약' 시트 없음: {sheets}"


def test_step3_build_returns_filename(client, both_project):
    """채권 step3/build → filename 반환."""
    pid = both_project["pid"]
    r = client.post(f"/api/project/{pid}/step3/build", json={
        "kind": "receivable",
        "preparer": "테스터",
    })
    assert r.status_code == 201
    assert "filename" in r.json
    assert r.json["filename"].endswith(".xlsx")


def test_step3_build_returns_artifact_id(client, both_project):
    """step3/build 응답에 artifact_id 포함."""
    pid = both_project["pid"]
    r = client.post(f"/api/project/{pid}/step3/build", json={
        "kind": "receivable",
        "preparer": "테스터",
    })
    assert r.status_code == 201
    assert "artifact_id" in r.json, "artifact_id 누락"
    assert r.json["artifact_id"]  # 비어있지 않음
