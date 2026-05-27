"""
test_unified_workpaper.py — 9시트 통합 조서 구조 검증

검증:
  - build_combined_report → 정확히 9개 시트
  - 9개 시트 이름 정확
  - 조회서 시트에 채권+채무 거래처 모두 포함
  - 채권만(receivable) → 9시트, 채무 컬럼 비움
  - 채무만(payable) → 9시트
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl
import pytest

from src.infrastructure.report.generic_reporter import (
    KindData,
    ReportContext,
    PartyContactInfo,
    EXPECTED_GENERIC_SHEETS,
    EXPECTED_SHEET_COUNT,
    build_combined_report,
    build_generic_report,
)
from src.domain.mus import MUSResult, MUSSelection
from src.domain.population import CompletenessCheck, PartyDecision
from src.domain.sample_size import SampleSizeResult


def _make_ctx(kind: str, prefix: str) -> ReportContext:
    return ReportContext(
        company_name="통합테스트회사",
        period_end=date(2025, 12, 31),
        kind=kind,
        preparer="통합작성자",
        reviewer="통합검토자",
        workpaper_no_prefix=prefix,
    )


def _make_completeness() -> CompletenessCheck:
    return CompletenessCheck(
        by_group=[{"group": "외상매출금", "ledger": 1_000_000, "fs": 1_000_000, "diff": 0, "note": ""}],
        total_ledger=1_000_000, total_fs=1_000_000, total_diff=0,
    )


def _make_size() -> SampleSizeResult:
    return SampleSizeResult(
        key_item_threshold=300_000,
        key_item_ratio=0.5,
        confidence_factor=1.4,
        base_sample_size=5,
        final_sample_size=5,
        sample_interval=150_000,
        remaining_population=750_000,
    )


def _make_mus() -> MUSResult:
    return MUSResult(
        sample_interval=150_000,
        random_start=50_000,
        selections=[
            MUSSelection(
                name="MUS거래처", balance=200_000, cumulative=200_000,
                selections=1, remainder_after=0, hit=True,
            ),
        ],
        sampled_names=["MUS거래처"],
    )


def _ar_kd() -> KindData:
    return KindData(
        ctx=_make_ctx("receivable", "C100"),
        completeness=_make_completeness(),
        size_result=_make_size(),
        decisions=[
            PartyDecision(
                name="채권거래처A", balance=500_000,
                is_key_item=True, is_representative=False,
                is_related_party=False, is_excluded=False,
                final_sampled=True,
                by_account={"외상매출금": 500_000},
            ),
            PartyDecision(
                name="MUS거래처", balance=200_000,
                is_key_item=False, is_representative=True,
                is_related_party=False, is_excluded=False,
                final_sampled=True,
                by_account={"미수금": 200_000},
            ),
        ],
        mus_result=_make_mus(),
        performance_materiality=600_000,
        population_amount=1_000_000,
    )


def _ap_kd() -> KindData:
    return KindData(
        ctx=_make_ctx("payable", "AA100"),
        completeness=_make_completeness(),
        size_result=_make_size(),
        decisions=[
            PartyDecision(
                name="채무거래처X", balance=400_000,
                is_key_item=True, is_representative=False,
                is_related_party=False, is_excluded=False,
                final_sampled=True,
                by_account={"외상매입금": 400_000},
            ),
            PartyDecision(
                name="채무거래처Y", balance=150_000,
                is_key_item=False, is_representative=True,
                is_related_party=False, is_excluded=False,
                final_sampled=True,
                by_account={"미지급금": 150_000},
            ),
        ],
        mus_result=_make_mus(),
        performance_materiality=600_000,
        population_amount=1_000_000,
    )


# ─────────────────────────────────────────────────────────────
# 테스트
# ─────────────────────────────────────────────────────────────

def test_combined_report_has_9_sheets(tmp_path):
    """build_combined_report(both) → 정확히 9개 시트."""
    out = tmp_path / "combined.xlsx"
    build_combined_report(out, receivable=_ar_kd(), payable=_ap_kd())

    wb = openpyxl.load_workbook(out, read_only=True)
    n = len(wb.sheetnames)
    sheets = set(wb.sheetnames)
    wb.close()

    assert n == EXPECTED_SHEET_COUNT, f"시트 수 불일치: {n} — {sheets}"


def test_combined_report_correct_sheet_names(tmp_path):
    """build_combined_report → 9개 시트 이름이 기대값과 일치."""
    out = tmp_path / "combined_names.xlsx"
    build_combined_report(out, receivable=_ar_kd(), payable=_ap_kd())

    wb = openpyxl.load_workbook(out, read_only=True)
    actual = set(wb.sheetnames)
    wb.close()

    missing = EXPECTED_GENERIC_SHEETS - actual
    extra   = actual - EXPECTED_GENERIC_SHEETS
    assert not missing, f"누락 시트: {missing}"
    assert not extra,   f"예상 외 시트: {extra}"


def test_confirmation_sheet_has_both_ar_ap_parties(tmp_path):
    """조회서 시트에 채권·채무 거래처 모두 포함."""
    out = tmp_path / "both_parties.xlsx"
    build_combined_report(out, receivable=_ar_kd(), payable=_ap_kd())

    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["조회서"]
    all_values = {ws.cell(r, c).value for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)}
    wb.close()

    assert "채권거래처A" in all_values, "채권거래처A 미포함"
    assert "채무거래처X" in all_values, "채무거래처X 미포함"
    assert "채무거래처Y" in all_values, "채무거래처Y 미포함"


def test_receivable_only_has_9_sheets(tmp_path):
    """채권만(receivable only) → 9시트."""
    out = tmp_path / "ar_only.xlsx"
    build_combined_report(out, receivable=_ar_kd(), payable=None)

    wb = openpyxl.load_workbook(out, read_only=True)
    n = len(wb.sheetnames)
    wb.close()
    assert n == EXPECTED_SHEET_COUNT, f"채권 단독 시트 수 불일치: {n}"


def test_payable_only_has_9_sheets(tmp_path):
    """채무만(payable only) → 9시트."""
    out = tmp_path / "ap_only.xlsx"
    build_combined_report(out, receivable=None, payable=_ap_kd())

    wb = openpyxl.load_workbook(out, read_only=True)
    n = len(wb.sheetnames)
    wb.close()
    assert n == EXPECTED_SHEET_COUNT, f"채무 단독 시트 수 불일치: {n}"


def test_combined_summary_has_total_population(tmp_path):
    """요약 시트: 채권+채무 합산 모집단 금액이 존재."""
    out = tmp_path / "summary_total.xlsx"
    build_combined_report(out, receivable=_ar_kd(), payable=_ap_kd())

    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["요약"]
    all_values = {ws.cell(r, c).value for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)}
    wb.close()

    # 채권 1,000,000 + 채무 1,000,000 = 2,000,000
    assert 2_000_000 in all_values, "통합 모집단 2,000,000 미반영"
