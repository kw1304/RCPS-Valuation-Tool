"""범용성 통합 커버리지 테스트

다양한 시트명·컬럼 순서·재무제표 형식 변형에 대해
자동 감지 로직이 올바르게 동작하는지 확인합니다.

커버 케이스:
- 채권 시트명 변형 (채권/매출채권/매출원장/AR/Receivable)
- 채무 시트명 변형 (채무/매입채무/매입원장/AP/Payable)
- 재무제표 시트명 변형 (FS_M/BS/재무상태표/Balance Sheet)
- 특관자 시트명 변형 (특관자리스트/관계회사/Related Parties)
- 컬럼 순서 변형 (7620 표준 vs dummy_client_a 변형)
- load_fs_amounts 당기 컬럼 선택 (전기/당기 병렬 컬럼)
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import pytest
import pandas as pd
import openpyxl
from openpyxl import Workbook


# ── 시트명 변형 — 채권 ──────────────────────────────────────────────────────
@pytest.mark.parametrize("ar_sheet", [
    "채권", "채권원장", "매출채권", "매출원장", "AR", "ar",
    "Receivable", "receivable", "외상매출금원장",
])
def test_receivable_sheet_detection(ar_sheet):
    from src.infrastructure.schemas.ledger_schema import detect_ledger_sheets
    result = detect_ledger_sheets([ar_sheet, "채무"])
    assert result["receivable"] == ar_sheet, f"'{ar_sheet}' 채권 감지 실패: {result}"


# ── 시트명 변형 — 채무 ──────────────────────────────────────────────────────
@pytest.mark.parametrize("ap_sheet", [
    "채무", "채무원장", "매입채무", "매입원장", "AP", "ap",
    "Payable", "payable", "외상매입금원장",
])
def test_payable_sheet_detection(ap_sheet):
    from src.infrastructure.schemas.ledger_schema import detect_ledger_sheets
    result = detect_ledger_sheets(["채권", ap_sheet])
    assert result["payable"] == ap_sheet, f"'{ap_sheet}' 채무 감지 실패: {result}"


# ── 시트명 변형 — 재무제표 ───────────────────────────────────────────────────
@pytest.mark.parametrize("fs_sheet", [
    "FS_M", "fs_m", "BS", "재무상태표", "재무제표", "Balance Sheet", "BalanceSheet",
])
def test_fs_sheet_detection(fs_sheet):
    from src.infrastructure.schemas.fs_schema import detect_fs_sheet
    result = detect_fs_sheet([fs_sheet, "기타시트"])
    assert result == fs_sheet, f"'{fs_sheet}' FS 시트 감지 실패"


# ── 시트명 변형 — 특관자 ────────────────────────────────────────────────────
@pytest.mark.parametrize("rp_sheet", [
    "특관자리스트", "특수관계자리스트", "특관자", "관계회사", "관계사",
    "Related Parties", "RelatedParties",
])
def test_rp_sheet_detection(rp_sheet):
    from src.infrastructure.schemas.rp_schema import detect_rp_sheet
    result = detect_rp_sheet([rp_sheet, "기타시트"])
    assert result == rp_sheet, f"'{rp_sheet}' RP 시트 감지 실패"


# ── 컬럼 순서 변형 ──────────────────────────────────────────────────────────
def _make_ledger_df(columns, row):
    """주어진 컬럼 순서로 DataFrame 생성."""
    return pd.DataFrame([row], columns=columns)


@pytest.mark.parametrize("columns,row,expected_ending_val", [
    # 7620 표준: 코드|명|계정과목|계정과목명|통화|기초|증감|기말
    (
        ["코드", "명", "계정과목", "계정과목명", "통화", "기초", "증감", "기말"],
        ["C001", "테스트", "외상매출금", "외상매출금", "KRW", 0, 100, 100],
        100,
    ),
    # dummy_client_a 변형: 거래처명|거래처코드|계정과목명|통화|기초잔액|당기증감|기말잔액
    (
        ["거래처명", "거래처코드", "계정과목명", "통화", "기초잔액", "당기증감", "기말잔액", "비고"],
        ["테스트거래처", "C001", "외상매출금", "KRW", 0, 200, 200, ""],
        200,
    ),
    # AR 변형: Name|Code|Account|Currency|Opening Balance|Change|Closing Balance
    (
        ["Name", "Code", "Account", "Currency", "Opening Balance", "Change", "Closing Balance"],
        ["Test Co", "V001", "외상매출금", "KRW", 0, 300, 300],
        300,
    ),
])
def test_column_detection_variants(columns, row, expected_ending_val):
    from src.infrastructure.schemas.ledger_schema import detect_ledger_columns

    df = _make_ledger_df(columns, row)
    col_map = detect_ledger_columns(df)

    assert col_map["ending"] is not None, f"기말잔액 컬럼 감지 실패: {col_map}"
    detected_val = df.iloc[0, col_map["ending"]]
    assert detected_val == expected_ending_val, (
        f"기말잔액 값 불일치: 기대={expected_ending_val}, 감지={detected_val}"
    )


# ── load_fs_amounts 당기 컬럼 선택 ──────────────────────────────────────────
def _make_fs_xlsx(columns, rows, sheet_name="FS_M"):
    """메모리 내 Excel 파일 생성 (임시 파일 반환)."""
    import tempfile
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(columns)
    for row in rows:
        ws.append(row)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


def test_fs_amounts_selects_당기_column(tmp_path):
    """전기/당기 컬럼이 모두 있을 때 당기금액이 선택되어야 한다."""
    from src.infrastructure.loaders import load_fs_amounts

    wb = Workbook()
    ws = wb.active
    ws.title = "BS"
    ws.append(["No", "구분", "항목명", "전기금액", "당기금액"])
    ws.append([1, "자산", "자산총계", 3_000_000_000, 4_000_000_000])
    ws.append([2, "자산", "외상매출금", 600_000_000, 720_000_000])

    path = tmp_path / "fs.xlsx"
    wb.save(str(path))

    fs = load_fs_amounts(str(path))
    assert "자산총계" in fs
    assert fs["자산총계"] == pytest.approx(4_000_000_000), (
        f"당기금액 선택 실패 — 전기금액이 선택됨: {fs['자산총계']}"
    )


def test_fs_amounts_single_value_column(tmp_path):
    """금액 컬럼이 하나만 있을 때 정상 처리되어야 한다."""
    from src.infrastructure.loaders import load_fs_amounts

    wb = Workbook()
    ws = wb.active
    ws.title = "FS_M"
    ws.append(["항목명", "금액"])
    ws.append(["자산총계", 5_000_000_000])
    ws.append(["외상매출금", 800_000_000])

    path = tmp_path / "fs2.xlsx"
    wb.save(str(path))

    fs = load_fs_amounts(str(path))
    assert fs.get("자산총계") == pytest.approx(5_000_000_000)


# ── 전체 샘플링 플로우 (컬럼 변형) ──────────────────────────────────────────
def test_sampling_with_column_variant(tmp_path):
    """컬럼 순서가 7620 표준과 다른 원장으로도 샘플링이 정상 동작해야 한다."""
    from src.orchestrator import SamplingParams, run_sampling
    import datetime

    # dummy_client_a 컬럼 순서 재현
    columns = ["거래처명", "거래처코드", "계정과목명", "통화", "기초잔액", "당기증감", "기말잔액"]
    data = [
        ["거래처A", "C001", "외상매출금", "KRW", 0, 500_000_000, 500_000_000],
        ["거래처B", "C002", "외상매출금", "KRW", 0, 300_000_000, 300_000_000],
        ["거래처C", "C003", "미수금",    "KRW", 0, 120_000_000, 120_000_000],
        ["거래처D", "C004", "외상매출금", "KRW", 0, 80_000_000,  80_000_000],
        ["거래처E", "C005", "외상매출금", "KRW", 0, 50_000_000,  50_000_000],
    ]
    df = pd.DataFrame(data, columns=columns)

    # PM=500M으로 설정 → Key item threshold = 375M → 거래처A만 Key item
    # 나머지 4건은 MUS pool에 들어가야 함
    params = SamplingParams(
        company_name="컬럼변형테스트",
        period_end=datetime.date(2025, 12, 31),
        kind="receivable",
        performance_materiality=500_000_000,
        risk_level="유의적위험",
        control_reliance="Y",
        random_seed=42,
    )

    result = run_sampling(df, params)
    assert result.population_amount > 0, "모집단 잔액이 0 — 컬럼 감지 실패"
    # 적어도 Key item(거래처A) 1건은 final_sampled 되어야 함
    assert any(d.final_sampled for d in result.decisions), "최종 샘플링 결과가 0건"


def test_sampling_with_standard_7620_columns(tmp_path):
    """7620 표준 컬럼 순서에서도 정상 동작 (회귀)."""
    from src.orchestrator import SamplingParams, run_sampling
    import datetime

    # 7620 표준 컬럼 순서
    columns = ["코드", "명", "계정과목", "계정과목명", "통화", "기초", "증감", "기말"]
    data = [
        ["C001", "거래처X", "01", "외상매출금", "KRW", 0, 400_000_000, 400_000_000],
        ["C002", "거래처Y", "01", "외상매출금", "KRW", 0, 200_000_000, 200_000_000],
        ["C003", "거래처Z", "02", "미수금",    "KRW", 0, 100_000_000, 100_000_000],
    ]
    df = pd.DataFrame(data, columns=columns)

    params = SamplingParams(
        company_name="7620표준테스트",
        period_end=datetime.date(2025, 12, 31),
        kind="receivable",
        performance_materiality=40_000_000,
        risk_level="유의적위험",
        control_reliance="Y",
        random_seed=42,
    )

    result = run_sampling(df, params)
    assert result.population_amount == pytest.approx(700_000_000)


# ── 한자·중국어 거래처명 처리 ────────────────────────────────────────────────
def test_chinese_party_name_handled():
    """한자·중국어 거래처명이 포함된 원장도 정상 처리되어야 한다."""
    from src.orchestrator import SamplingParams, run_sampling
    import datetime

    columns = ["거래처명", "거래처코드", "계정과목명", "통화", "기초잔액", "당기증감", "기말잔액"]
    data = [
        ["三星電子", "C001", "외상매출금", "KRW", 0, 500_000_000, 500_000_000],
        ["上海贸易", "C002", "외상매출금", "CNY", 0, 100_000_000, 100_000_000],
        ["한국거래처", "C003", "외상매출금", "KRW", 0, 200_000_000, 200_000_000],
    ]
    df = pd.DataFrame(data, columns=columns)

    params = SamplingParams(
        company_name="국제거래테스트",
        period_end=datetime.date(2025, 12, 31),
        kind="receivable",
        performance_materiality=400_000_000,
        risk_level="보통",
        control_reliance="Y",
        random_seed=42,
    )

    result = run_sampling(df, params)
    party_names = [d.name for d in result.decisions]
    assert "三星電子" in party_names, "한자 거래처명이 처리되지 않음"
    assert "上海贸易" in party_names, "중국어 거래처명이 처리되지 않음"
    assert result.population_amount > 0
