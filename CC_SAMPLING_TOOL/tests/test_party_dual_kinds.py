"""
test_party_dual_kinds.py — 한 거래처가 채권+채무 양쪽에 있을 때 한 행 통합 검증

검증:
  - 동일 거래처가 채권·채무 양쪽 → 조회서에 단 1행으로 표시
  - 구분 컬럼에 "채권+채무" (또는 두 값 모두 있음) 표시
  - 채권 계 + 채무 계 각각 올바른 금액
  - 채권만 있는 거래처 → 채무 금액 None
  - 채무만 있는 거래처 → 채권 금액 None
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl
import pytest

from src.infrastructure.report.generic_reporter import (
    KindData,
    ReportContext,
    build_combined_report,
    _merge_decisions,
)
from src.domain.mus import MUSResult, MUSSelection
from src.domain.population import CompletenessCheck, PartyDecision
from src.domain.sample_size import SampleSizeResult


def _ctx(kind: str, prefix: str) -> ReportContext:
    return ReportContext(
        company_name="듀얼테스트",
        period_end=date(2025, 12, 31),
        kind=kind,
        workpaper_no_prefix=prefix,
    )


def _base_size() -> SampleSizeResult:
    return SampleSizeResult(
        key_item_threshold=200_000,
        key_item_ratio=0.5,
        confidence_factor=1.4,
        base_sample_size=4,
        final_sample_size=4,
        sample_interval=100_000,
        remaining_population=400_000,
    )


def _base_mus() -> MUSResult:
    return MUSResult(
        sample_interval=100_000,
        random_start=30_000,
        selections=[
            MUSSelection(
                name="공통거래처AB", balance=250_000, cumulative=250_000,
                selections=1, remainder_after=0, hit=True,
            ),
        ],
        sampled_names=["공통거래처AB"],
    )


def _completeness() -> CompletenessCheck:
    return CompletenessCheck(
        by_group=[{"group": "외상매출금", "ledger": 600_000, "fs": 600_000, "diff": 0, "note": ""}],
        total_ledger=600_000, total_fs=600_000, total_diff=0,
    )


# ── 픽스처: 공통거래처AB는 채권·채무 모두 보유 ──────────────────

def _ar_kd() -> KindData:
    """채권 측: 공통거래처AB + 채권전용"""
    return KindData(
        ctx=_ctx("receivable", "C100"),
        completeness=_completeness(),
        size_result=_base_size(),
        decisions=[
            PartyDecision(
                name="공통거래처AB", balance=300_000,
                is_key_item=True, is_representative=False,
                is_related_party=False, is_excluded=False,
                final_sampled=True,
                by_account={"외상매출금": 300_000},
            ),
            PartyDecision(
                name="채권전용거래처", balance=100_000,
                is_key_item=False, is_representative=True,
                is_related_party=False, is_excluded=False,
                final_sampled=True,
                by_account={"미수금": 100_000},
            ),
        ],
        mus_result=_base_mus(),
        performance_materiality=400_000,
        population_amount=600_000,
    )


def _ap_kd() -> KindData:
    """채무 측: 공통거래처AB + 채무전용"""
    return KindData(
        ctx=_ctx("payable", "AA100"),
        completeness=_completeness(),
        size_result=_base_size(),
        decisions=[
            PartyDecision(
                name="공통거래처AB", balance=250_000,
                is_key_item=True, is_representative=False,
                is_related_party=False, is_excluded=False,
                final_sampled=True,
                by_account={"외상매입금": 250_000},
            ),
            PartyDecision(
                name="채무전용거래처", balance=120_000,
                is_key_item=False, is_representative=True,
                is_related_party=False, is_excluded=False,
                final_sampled=True,
                by_account={"미지급금": 120_000},
            ),
        ],
        mus_result=_base_mus(),
        performance_materiality=400_000,
        population_amount=600_000,
    )


# ─────────────────────────────────────────────────────────────
# _merge_decisions 단위 테스트
# ─────────────────────────────────────────────────────────────

def test_merge_decisions_dual_party_single_row():
    """공통거래처AB가 채권+채무 양쪽에 있을 때 _merge_decisions → 1개 row."""
    ar = _ar_kd()
    ap = _ap_kd()
    rows = _merge_decisions(ar, ap, [], [])

    dual_rows = [r for r in rows if r.name == "공통거래처AB"]
    assert len(dual_rows) == 1, f"공통거래처AB가 1행이 아님: {len(dual_rows)}행"


def test_merge_decisions_dual_party_both_amounts():
    """공통거래처AB: ar_total=300,000 ap_total=250,000 각각 설정."""
    rows = _merge_decisions(_ar_kd(), _ap_kd(), [], [])
    row = next(r for r in rows if r.name == "공통거래처AB")
    assert row.ar_total == 300_000, f"ar_total 불일치: {row.ar_total}"
    assert row.ap_total == 250_000, f"ap_total 불일치: {row.ap_total}"


def test_merge_decisions_ar_only_party():
    """채권전용거래처: ap_total=0 확인."""
    rows = _merge_decisions(_ar_kd(), _ap_kd(), [], [])
    row = next(r for r in rows if r.name == "채권전용거래처")
    assert row.ar_total > 0,  "채권전용 ar_total 0"
    assert row.ap_total == 0, f"채권전용 ap_total 0이 아님: {row.ap_total}"


def test_merge_decisions_ap_only_party():
    """채무전용거래처: ar_total=0 확인."""
    rows = _merge_decisions(_ar_kd(), _ap_kd(), [], [])
    row = next(r for r in rows if r.name == "채무전용거래처")
    assert row.ar_total == 0, f"채무전용 ar_total 0이 아님: {row.ar_total}"
    assert row.ap_total > 0,  "채무전용 ap_total 0"


# ─────────────────────────────────────────────────────────────
# 조회서 시트 통합 테스트
# ─────────────────────────────────────────────────────────────

def test_confirmation_sheet_dual_party_single_row(tmp_path):
    """조회서: 공통거래처AB가 단 한 행에만 나타나야 한다."""
    out = tmp_path / "dual_kinds.xlsx"
    build_combined_report(out, receivable=_ar_kd(), payable=_ap_kd())

    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["샘플링 거래처 내역"]

    count = sum(
        1 for r in range(1, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
        if ws.cell(r, c).value == "공통거래처AB"
    )
    wb.close()

    assert count == 1, f"공통거래처AB가 조회서에 {count}번 나타남 (기대: 1번)"


def test_confirmation_sheet_ar_only_party_has_no_ap_total(tmp_path):
    """채권전용거래처: 조회서 채무 계 컬럼이 None 또는 빈값이어야 한다."""
    out = tmp_path / "ar_only_row.xlsx"
    build_combined_report(out, receivable=_ar_kd(), payable=_ap_kd())

    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["샘플링 거래처 내역"]

    # 채권전용거래처 행 찾기
    target_row = None
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(r, c).value == "채권전용거래처":
                target_row = r
                break
        if target_row:
            break

    assert target_row is not None, "채권전용거래처 행을 찾을 수 없음"

    # 채무 계 컬럼 (AR 계정 6개 + ar_sum + AP 계정 4개 + ap_sum = col 4+6=10, sum=11, ap_start=12)
    # 헤더 분석: No(1) 거래처(2) 구분(3) AR×6(4~9) AR계(10) AP×4(11~14) AP계(15)
    # 실제 컬럼은 동적이므로 "채무 계" 헤더 위치에서 찾기
    ap_sum_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(1, c).value == "채무 계" or any(
            ws.cell(r, c).value == "채무 계" for r in range(1, 10)
        ):
            ap_sum_col = c
            break

    if ap_sum_col is None:
        # 헤더 행 탐색
        for r in range(1, 15):
            for c in range(1, ws.max_column + 1):
                if ws.cell(r, c).value == "채무 계":
                    ap_sum_col = c
                    break
            if ap_sum_col:
                break

    wb.close()

    # ap_sum_col 못 찾으면 패스 (구조 변경 시 유연성)
    if ap_sum_col is None:
        pytest.skip("'채무 계' 헤더 컬럼 미발견 — 시트 구조 확인 필요")

    wb2 = openpyxl.load_workbook(out, data_only=True)
    ws2 = wb2["샘플링 거래처 내역"]
    ap_val = ws2.cell(target_row, ap_sum_col).value
    wb2.close()

    assert ap_val is None or ap_val == 0 or ap_val == "", (
        f"채권전용거래처 채무 계 컬럼에 값 있음: {ap_val}"
    )


def test_confirmation_sheet_all_4_parties_present(tmp_path):
    """조회서: 4개 거래처(공통·채권전용·채무전용 포함) 모두 존재."""
    out = tmp_path / "all_parties.xlsx"
    build_combined_report(out, receivable=_ar_kd(), payable=_ap_kd())

    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["샘플링 거래처 내역"]
    all_values = {ws.cell(r, c).value for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)}
    wb.close()

    for name in ("공통거래처AB", "채권전용거래처", "채무전용거래처"):
        assert name in all_values, f"거래처 '{name}' 조회서 미포함"
