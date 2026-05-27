"""발송명단 빌더 단위 테스트 (Week 2)."""
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl
import pandas as pd
import pytest

from src.domain.population import PartyDecision, aggregate_by_party, classify_parties, load_ledger_rows
from src.domain.sample_size import SampleSizeInput, compute_sample_size
from src.domain.mus import run_mus
from src.infrastructure.confirmations.send_list_builder import build_send_list


LEDGER_PATH = ROOT / "input" / "회사자료" / "채권채무조회서 거래처별 원장.XLSX"
PERFORMANCE_MATERIALITY = 2_738_000_000


@pytest.fixture(scope="module")
def sampled_decisions():
    """실제 원장 데이터로 Step 1 실행 → final_sampled decisions 반환."""
    if not LEDGER_PATH.exists():
        pytest.skip("원장 파일 없음")

    df = pd.read_excel(LEDGER_PATH, sheet_name="채권")
    rows = load_ledger_rows(df, kind="receivable")
    parties = aggregate_by_party(rows, kind="receivable", sign_normalize=True)

    pm = PERFORMANCE_MATERIALITY
    ratio = 0.75
    threshold = pm * ratio

    decisions = classify_parties(
        parties=parties,
        key_item_threshold=threshold,
        related_party_names=set(),
        excluded_parties={"helloBiome safe": "제외"},
    )

    pool = [(d.name, d.balance) for d in decisions if not d.is_excluded and not d.is_key_item and d.balance > 0]
    pool.sort(key=lambda x: x[0])

    size_result = compute_sample_size(SampleSizeInput(
        population_amount=sum(d.balance for d in decisions if not d.is_excluded),
        performance_materiality=pm,
        risk_level="유의적위험",
        control_reliance="Y",
        key_item_amount=sum(d.balance for d in decisions if d.is_key_item),
    ))

    mus_result = run_mus(
        pool=pool,
        sample_size=size_result.final_sample_size,
        sample_interval=size_result.sample_interval,
        seed=42,
    )

    sampled_set = set(mus_result.sampled_names)
    for d in decisions:
        if d.name in sampled_set:
            d.is_representative = True
        if d.is_key_item or d.is_representative:
            d.final_sampled = True

    return decisions


@pytest.fixture
def send_list_path(tmp_path, sampled_decisions):
    """발송명단 Excel 생성 후 경로 반환."""
    out = tmp_path / "발송명단_테스트.xlsx"
    build_send_list(
        out_path=out,
        project_info={
            "company_name": "테스트회사",
            "period_end": "2025-12-31",
            "audit_firm": "웅계회계법인",
            "preparer": "이슬기",
        },
        decisions=sampled_decisions,
        kind="receivable",
        reply_deadline=date(2026, 2, 28),
        contact_info={"email": "audit@wc.com", "phone": "02-1234-5678", "address": "서울"},
        party_contacts={},
    )
    return out


def test_send_list_file_exists(send_list_path):
    """발송명단 파일 생성됨."""
    assert send_list_path.exists()


def test_send_list_has_three_sheets(send_list_path):
    """시트 3개: 발송명단·송부 안내문·거래처별 회신서 양식."""
    wb = openpyxl.load_workbook(send_list_path, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    assert "발송명단" in sheets, f"발송명단 시트 없음 — 현재: {sheets}"
    assert "송부 안내문" in sheets, f"송부 안내문 시트 없음"
    assert "거래처별 회신서 양식" in sheets, f"거래처별 회신서 양식 시트 없음"


def test_send_list_sheet_headers(send_list_path):
    """발송명단 시트 헤더 컬럼 검증 (No/거래처명/.../비고)."""
    wb = openpyxl.load_workbook(send_list_path, data_only=True)
    ws = wb["발송명단"]
    # 헤더는 5행
    headers = [ws.cell(row=5, column=c).value for c in range(1, 10)]
    wb.close()
    assert headers[0] == "No", f"첫 헤더 오류: {headers[0]}"
    assert headers[1] == "거래처명", f"두 번째 헤더 오류: {headers[1]}"
    assert "기말잔액" in str(headers[4]), f"잔액 헤더 오류: {headers[4]}"
    assert headers[8] == "비고", f"마지막 헤더 오류: {headers[8]}"


def test_send_list_data_row_count(send_list_path, sampled_decisions):
    """데이터 행 수 == final_sampled 거래처 수."""
    final_count = sum(1 for d in sampled_decisions if d.final_sampled)
    wb = openpyxl.load_workbook(send_list_path, data_only=True)
    ws = wb["발송명단"]
    wb.close()
    # 데이터는 6행~, 합계행 제외 → max_row - 6 (6=데이터시작) - 1(합계)
    # No 컬럼(1)이 숫자인 행 카운트
    data_rows = 0
    for row in ws.iter_rows(min_row=6, values_only=True):
        if isinstance(row[0], (int, float)):
            data_rows += 1
    assert data_rows == final_count, f"데이터 행 수 불일치: {data_rows} != {final_count}"


def test_send_list_confirmation_forms_count(send_list_path, sampled_decisions):
    """거래처별 회신서 양식: 거래처 수(페이지) 만큼 데이터 존재 (제목 행 개수로 검증)."""
    final_count = sum(1 for d in sampled_decisions if d.final_sampled)
    wb = openpyxl.load_workbook(send_list_path, data_only=True)
    ws = wb["거래처별 회신서 양식"]
    wb.close()

    # 각 거래처 양식 제목은 "채권채무조회서" 문자열 포함
    title_count = 0
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and "채권채무조회서" in cell:
                title_count += 1
                break

    assert title_count == final_count, f"회신서 페이지 수 불일치: {title_count} != {final_count}"


def test_send_list_reply_deadline_appears(send_list_path):
    """송부 안내문에 회신 기한이 포함됨."""
    wb = openpyxl.load_workbook(send_list_path, data_only=True)
    ws = wb["송부 안내문"]
    wb.close()
    found = False
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell and "2026" in str(cell):
                found = True
                break
    assert found, "송부 안내문에 회신 기한(2026년) 미포함"
