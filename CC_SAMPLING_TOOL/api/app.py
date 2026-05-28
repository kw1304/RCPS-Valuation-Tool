"""웅계사's CC Sampling Tool — Flask 서버 (Week 3: PDF 회신 추출·매칭·차이판정 API)"""
from __future__ import annotations

import json
import logging
import os
import sys
import uuid
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import pandas as pd
import yaml
from flask import Flask, jsonify, request, send_file, send_from_directory

from src.domain.population import (
    ACCOUNT_GROUP_MAP_PAYABLE,
    ACCOUNT_GROUP_MAP_RECEIVABLE,
    PartyDecision,
    aggregate_by_party,
    load_ledger_rows,
)
from src.domain.sample_size import (
    CONFIDENCE_FACTOR_MATRIX,
    KEY_ITEM_RATIO_MATRIX,
)
from src.infrastructure.loaders import (
    get_total_assets, load_fs_amounts, load_related_parties, load_upload_guide,
    UploadGuideData, PartyContact,
)
from src.infrastructure.report.generic_reporter import (
    PartyContactInfo, ExclusionRow, ConfirmationReplyInfo, AlternativeProcedureEntry,
)
from src.infrastructure.schemas.ledger_schema import detect_ledger_sheets
from src.infrastructure.persistence import (
    AlternativeProcedureRepository,
    ArtifactRepository,
    AuditTrail,
    ConfirmationReplyRepository,
    ProjectRepository,
    WorkpaperRepository,
    init_db,
    get_session,
)
from src.infrastructure.pdf import extract_text, parse_confirmation
from src.infrastructure.pdf.form_detector import detect_form
from src.infrastructure.pdf.pattern_library import get_patterns
from src.infrastructure.pdf.parser import parse_confirmation_v2
from src.domain.matching import match_party
from src.domain.reconciliation import reconcile, reconcile_v2
from src.domain.currency import CurrencyResolver
from src.infrastructure.report.template_registry import list_templates, get_template
from src.infrastructure.confirmations.send_list_builder import build_send_list
from src.orchestrator import SamplingParams, run_sampling, write_report

ROOT = Path(__file__).resolve().parent.parent
FRONTEND = ROOT / "frontend"
UPLOAD_DIR = ROOT / "input" / "_uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

app = Flask(__name__, static_folder=str(FRONTEND))
log = logging.getLogger("cc_sampling")

# DB 초기화 (첫 실행 시 테이블 생성)
init_db()


# ── 전역 에러 핸들러 ──────────────────────────────────────────────────────────
@app.errorhandler(Exception)
def handle_unhandled_exception(e):
    """처리되지 않은 예외 → JSON 500 응답 (trace_id 포함)."""
    trace_id = str(uuid.uuid4())
    log.exception(f"[trace_id={trace_id}] Unhandled exception: {e}")
    return jsonify({"error": "내부 서버 오류가 발생했습니다", "trace_id": trace_id}), 500


@app.errorhandler(400)
def handle_bad_request(e):
    return jsonify({"error": str(e.description) if hasattr(e, "description") else "잘못된 요청"}), 400


@app.errorhandler(404)
def handle_not_found(e):
    return jsonify({"error": "요청한 리소스를 찾을 수 없습니다"}), 404


# ── 파일 확장자 유효성 검증 헬퍼 ──────────────────────────────────────────────
_ALLOWED_EXCEL_EXTS = {".xlsx", ".xls", ".xlsm"}
_ALLOWED_PDF_EXT = {".pdf"}


def _validate_file_ext(filename: str, allowed: set[str]) -> bool:
    """파일명 확장자가 허용 목록에 있는지 확인."""
    return Path(filename).suffix.lower() in allowed

# ── in-memory STATE — Week 1 동안 DB와 병행 운영 ──────────────────────────
# 기존 /api/upload, /api/run, /api/download 는 project_id 연동으로 확장.
# project_id 없는 호출은 deprecation warning 로그 + 임시 프로젝트 자동 생성.
STATE: dict = {
    "ledger_path": None,
    "fs_path": None,
    "rp_path": None,
    "upload_guide_path": None,   # UploadGuide xlsx 경로
    "upload_guide_data": None,   # UploadGuideData 객체 캐시
    "last_result": {},   # {"receivable": {"result": ..., "params": ...}, ...}
    "current_project_id": None,
    "workpaper_ids": {},  # {kind: workpaper_id}
    "kind": "receivable",
    "sheets": {},        # {"receivable": sheet_name, "payable": sheet_name}
}


def _get_or_create_temp_project(session, kind: str = "both") -> tuple[str, str]:
    """project_id 없는 legacy 호출용 임시 프로젝트 생성.
    (project_id, workpaper_id) 반환.
    """
    proj_repo = ProjectRepository(session)
    wp_repo = WorkpaperRepository(session)

    # STATE 에 이미 있으면 재사용
    if STATE.get("current_project_id"):
        pid = STATE["current_project_id"]
        proj = proj_repo.get(pid)
        if proj and proj.status != "archived":
            wp = wp_repo.get_or_create(pid, kind)
            return pid, wp.id

    proj = proj_repo.create(
        company_name="임시프로젝트",
        period_end=date.today().isoformat(),
        kind="both",
    )
    STATE["current_project_id"] = proj.id
    wp = wp_repo.get_or_create(proj.id, kind)
    STATE["workpaper_ids"][kind] = wp.id
    return proj.id, wp.id


# ─────────────────────────────────────────────────────────────
# 헬스체크 + 정적 파일
# ─────────────────────────────────────────────────────────────
@app.route("/healthz")
def healthz():
    return {"status": "ok"}, 200


@app.route("/")
def index():
    return send_from_directory(FRONTEND, "index.html")


@app.route("/<path:p>")
def static_files(p):
    return send_from_directory(FRONTEND, p)


# ─────────────────────────────────────────────────────────────
# Project CRUD
# ─────────────────────────────────────────────────────────────
@app.route("/api/project", methods=["POST"])
def create_project():
    """신규 프로젝트 생성.
    body: {company_name, period_end, kind, audit_firm, preparer_email}
    """
    data = request.json or {}
    company_name = data.get("company_name", "").strip()
    period_end = data.get("period_end", date.today().isoformat())
    kind = data.get("kind", "both")
    audit_firm = data.get("audit_firm", "")
    preparer_email = data.get("preparer_email", "")

    if not company_name:
        return jsonify({"error": "company_name 필수"}), 400

    with get_session() as s:
        repo = ProjectRepository(s)
        proj = repo.create(
            company_name=company_name,
            period_end=period_end,
            kind=kind,
            audit_firm=audit_firm,
            created_by_email=preparer_email,
        )
        # 채권·채무 워크페이퍼 미리 생성
        wp_repo = WorkpaperRepository(s)
        wps = {}
        if kind in ("receivable", "both"):
            wp = wp_repo.get_or_create(proj.id, "receivable")
            wps["receivable"] = wp.id
        if kind in ("payable", "both"):
            wp = wp_repo.get_or_create(proj.id, "payable")
            wps["payable"] = wp.id

        # 현재 세션 업데이트
        STATE["current_project_id"] = proj.id
        STATE["workpaper_ids"] = wps

        return jsonify({
            "id": proj.id,
            "company_name": proj.company_name,
            "period_end": proj.period_end,
            "kind": proj.kind,
            "audit_firm": proj.audit_firm,
            "status": proj.status,
            "workpapers": wps,
            "created_at": proj.created_at.isoformat(),
        }), 201


@app.route("/api/project", methods=["GET"])
def list_projects():
    """프로젝트 목록 (최근 수정 순, archived 제외)."""
    with get_session() as s:
        repo = ProjectRepository(s)
        projects = repo.list_all()
        return jsonify([_serialize_project(p) for p in projects])


@app.route("/api/project/<pid>", methods=["GET"])
def get_project(pid: str):
    """단일 프로젝트 상세 + 워크페이퍼 + 상태."""
    with get_session() as s:
        proj = ProjectRepository(s).get(pid)
        if proj is None:
            return jsonify({"error": "not found"}), 404
        wp_repo = WorkpaperRepository(s)
        workpapers = []
        for wp in proj.workpapers:
            workpapers.append({
                "id": wp.id,
                "kind": wp.kind,
                "step1_done": wp.step1_completed_at is not None,
                "step2_done": wp.step2_completed_at is not None,
                "step3_done": wp.step3_completed_at is not None,
                "step4_done": wp.step4_completed_at is not None,
                "step5_done": wp.step5_completed_at is not None,
                "updated_at": wp.updated_at.isoformat() if wp.updated_at else None,
            })
        result = _serialize_project(proj)
        result["workpapers"] = workpapers
        return jsonify(result)


@app.route("/api/project/<pid>", methods=["PATCH"])
def update_project(pid: str):
    """프로젝트 메타정보 수정."""
    data = request.json or {}
    user_email = data.pop("user_email", "")
    with get_session() as s:
        proj = ProjectRepository(s).update(pid, user_email=user_email, **data)
        if proj is None:
            return jsonify({"error": "not found"}), 404
        return jsonify(_serialize_project(proj))


@app.route("/api/project/<pid>", methods=["DELETE"])
def delete_project(pid: str):
    """Soft delete — status=archived. Artifact 파일은 유지."""
    user_email = request.args.get("user_email", "")
    with get_session() as s:
        ok = ProjectRepository(s).soft_delete(pid, user_email=user_email)
        if not ok:
            return jsonify({"error": "not found"}), 404
        return jsonify({"deleted": pid})


@app.route("/api/project/<pid>/activate", methods=["POST"])
def activate_project(pid: str):
    """프로젝트를 현재 세션으로 활성화 (브라우저 셀렉터 전환용)."""
    with get_session() as s:
        proj = ProjectRepository(s).get(pid)
        if proj is None or proj.status == "archived":
            return jsonify({"error": "not found"}), 404
        wp_repo = WorkpaperRepository(s)
        wps = {wp.kind: wp.id for wp in proj.workpapers}

        STATE["current_project_id"] = pid
        STATE["workpaper_ids"] = wps
        # STATE 파일 경로 초기화 (다른 프로젝트 파일이 남아있을 수 있음)
        STATE["ledger_path"] = None
        STATE["fs_path"] = None
        STATE["rp_path"] = None
        STATE["last_result"] = {}

        return jsonify({
            "activated": pid,
            "company_name": proj.company_name,
            "period_end": proj.period_end,
            "workpapers": wps,
        })


# ─────────────────────────────────────────────────────────────
# 1. 파일 업로드 (기존 + project_id 확장)
# ─────────────────────────────────────────────────────────────
@app.route("/api/upload", methods=["POST"])
def upload():
    """ledger + fs + rp 파일 업로드. multipart form-data.
    project_id 쿼리 파라미터 옵션: ?project_id=<uuid>
    """
    project_id = request.args.get("project_id") or request.form.get("project_id")
    result = {}

    for kind in ("ledger", "fs", "rp"):
        f = request.files.get(kind)
        if f and f.filename:
            # 파일 형식 검증 — Excel(.xlsx/.xls/.xlsm)만 허용
            if not _validate_file_ext(f.filename, _ALLOWED_EXCEL_EXTS):
                ext = Path(f.filename).suffix.lower()
                return jsonify({
                    "error": f"'{kind}' 파일 형식 오류: {ext} 파일은 지원하지 않습니다. "
                             f"Excel 파일(.xlsx/.xls/.xlsm)을 업로드하세요."
                }), 400
            path = UPLOAD_DIR / f"_{kind}.xlsx"
            f.save(path)
            STATE[f"{kind}_path"] = str(path)
            result[kind] = path.name

            # Artifact DB 저장
            if project_id:
                try:
                    with get_session() as s:
                        art_repo = ArtifactRepository(s)
                        wp_id = STATE["workpaper_ids"].get(
                            "receivable" if kind in ("ledger", "fs") else "receivable"
                        )
                        art = art_repo.save_file(
                            project_id=project_id,
                            kind=kind,
                            source_path=path,
                            filename=f.filename,
                            workpaper_id=wp_id,
                        )
                        result[f"{kind}_artifact_id"] = art.id
                except Exception as e:
                    log.warning(f"Artifact 저장 실패 (비치명적): {e}")

    # UploadGuide 업로드 (선택) — 거래처 연락처·발송제외 자동 추출
    ug_file = request.files.get("upload_guide")
    if ug_file and ug_file.filename:
        if not _validate_file_ext(ug_file.filename, _ALLOWED_EXCEL_EXTS):
            ext = Path(ug_file.filename).suffix.lower()
            return jsonify({
                "error": f"'upload_guide' 파일 형식 오류: {ext}. Excel 파일(.xlsx)을 업로드하세요."
            }), 400
        ug_path = UPLOAD_DIR / "_upload_guide.xlsx"
        ug_file.save(ug_path)
        STATE["upload_guide_path"] = str(ug_path)
        STATE["upload_guide_data"] = None  # 캐시 초기화
        try:
            ug_data = load_upload_guide(ug_path)
            STATE["upload_guide_data"] = ug_data
            result["upload_guide"] = ug_file.filename
            result["upload_guide_send_targets"] = len(ug_data.send_targets)
            result["upload_guide_excluded"] = len(ug_data.excluded)
            result["upload_guide_excluded_names"] = sorted(ug_data.excluded_names())
        except Exception as e:
            log.warning(f"UploadGuide 파싱 실패 (비치명적): {e}")
            result["upload_guide_warning"] = f"UploadGuide 파싱 실패: {e!s}"

    # project_id 없는 레거시 호출 — 임시 프로젝트 자동 생성
    if not project_id and (STATE.get("ledger_path") or STATE.get("fs_path")):
        log.warning("project_id 없는 /api/upload 호출 — 임시 프로젝트 생성 (deprecated)")

    # ledger 시트 감지
    if STATE.get("ledger_path"):
        try:
            wb = openpyxl.load_workbook(STATE["ledger_path"], read_only=True, data_only=True)
            sheets = wb.sheetnames
            wb.close()
            result["sheets"] = sheets
            sheet_map = detect_ledger_sheets(sheets)
            result["sheet_map"] = sheet_map
            STATE["sheets"] = sheet_map
        except Exception as e:
            log.warning(f"거래처원장 시트 감지 실패: {e}")
            result["sheets_warning"] = f"시트 자동 감지 실패: {e!s}. 시트명을 직접 지정하세요."

    # 재무제표 자동 — 총자산
    if STATE.get("fs_path"):
        try:
            fs = load_fs_amounts(STATE["fs_path"])
            if not fs:
                result["fs_warning"] = "재무제표에서 금액 데이터를 읽지 못했습니다. 파일 구조를 확인하세요."
            else:
                result["total_assets"] = get_total_assets(fs)
                result["fs_amounts"] = {
                    k: v for k, v in fs.items() if any(g in k for g in [
                        "외상매출금", "받을어음", "미수금", "선급금", "대여금",
                        "임차보증금", "기타보증금", "외상매입금", "지급어음",
                        "미지급금", "선수금", "임대보증금",
                    ])
                }
        except Exception as e:
            log.warning(f"재무제표 로드 실패: {e}")
            result["fs_warning"] = f"재무제표 파싱 실패: {e!s}"

    if STATE.get("rp_path"):
        try:
            wb = openpyxl.load_workbook(STATE["rp_path"], read_only=True, data_only=True)
            wb.close()
            rp = load_related_parties(STATE["rp_path"])
            result["related_parties"] = sorted(rp)
        except Exception as e:
            log.warning(f"특관자리스트 로드 실패: {e}")
            result["rp_warning"] = f"특관자리스트 파싱 실패: {e!s}"

    return jsonify(result)


# ─────────────────────────────────────────────────────────────
# 2. 모집단 미리보기
# ─────────────────────────────────────────────────────────────
@app.route("/api/inspect", methods=["POST"])
def inspect():
    data = request.json or {}
    kind = data.get("kind", "receivable")
    sheet = data.get("sheet")

    if not STATE.get("ledger_path"):
        return jsonify({"error": "ledger not uploaded"}), 400

    df = pd.read_excel(STATE["ledger_path"], sheet_name=sheet)
    rows = load_ledger_rows(df, kind=kind)
    parties = aggregate_by_party(rows, kind=kind, sign_normalize=True)

    group_totals: dict[str, float] = {}
    for pb in parties.values():
        for g, amt in pb.by_account.items():
            group_totals[g] = group_totals.get(g, 0.0) + amt

    return jsonify({
        "party_count": len(parties),
        "total": sum(pb.total for pb in parties.values()),
        "groups": [{"name": g, "amount": v} for g, v in sorted(group_totals.items(), key=lambda x: -x[1])],
    })


# ─────────────────────────────────────────────────────────────
# 3. 샘플링 실행
# ─────────────────────────────────────────────────────────────
def _run_single_kind(data: dict, kind: str, sheet: str | None,
                     project_id: str, related_parties: set) -> tuple:
    """단일 kind(채권 or 채무) 샘플링 실행 → (serialized_dict | None, error_str | None)."""
    if not STATE.get("ledger_path"):
        return None, "거래처원장 파일을 먼저 업로드하세요"

    actual_sheet = sheet or STATE.get("sheets", {}).get(kind)
    if not actual_sheet:
        # 시트가 없으면 에러가 아니라 skip (both 모드에서 한쪽만 있어도 동작)
        return None, f"{kind} 시트 미감지 — 원장에 해당 시트가 없습니다"

    pm_raw = data.get("performance_materiality", 0)
    try:
        pm = float(pm_raw)
    except (TypeError, ValueError):
        return None, f"수행 중요성(performance_materiality) 값이 잘못되었습니다: {pm_raw}"
    if pm <= 0:
        return None, "수행 중요성(PM)은 0보다 커야 합니다"

    try:
        df = pd.read_excel(STATE["ledger_path"], sheet_name=actual_sheet)
    except Exception as e:
        return None, f"거래처원장 파일 읽기 실패 ({kind}): {e!s}"

    if df.empty:
        return None, f"거래처원장이 비어 있습니다 ({kind})"

    params = SamplingParams(
        company_name=data.get("company_name", ""),
        period_end=date.fromisoformat(data.get("period_end", "2025-12-31")),
        kind=kind,
        performance_materiality=pm,
        risk_level=data.get("risk_level", "유의적위험"),
        control_reliance=data.get("control_reliance", "Y"),
        key_item_ratio_override=_f(data.get("key_item_ratio")),
        confidence_factor_override=_f(data.get("confidence_factor")),
        fs_amounts_by_group=data.get("fs_amounts_by_group") or {},
        completeness_notes=data.get("completeness_notes") or {},
        excluded_parties=data.get("excluded_parties") or {},
        related_parties=related_parties,
        force_include_related=bool(data.get("force_include_related", True)),
        random_seed=_i(data.get("seed", 42)),
        preparer=data.get("preparer", ""),
        reviewer=data.get("reviewer", ""),
    )

    result = run_sampling(df, params)

    if result.population_amount <= 0:
        kind_label = "채권" if kind == "receivable" else "채무"
        return None, (
            f"{kind_label} 모집단 잔액이 0입니다. "
            f"올바른 시트를 선택했는지 확인하세요. (시트: {actual_sheet})"
        )

    serialized = _serialize_result(result, params)
    STATE["last_result"][kind] = {"result": result, "params": params}

    # Workpaper DB 저장
    with get_session() as s:
        wp_repo = WorkpaperRepository(s)
        wp = wp_repo.get_or_create(project_id, kind)
        STATE["workpaper_ids"][kind] = wp.id
        params_dict = {
            "company_name": params.company_name,
            "period_end": str(params.period_end),
            "kind": params.kind,
            "performance_materiality": params.performance_materiality,
            "risk_level": params.risk_level,
            "control_reliance": params.control_reliance,
            "preparer": params.preparer,
            "reviewer": params.reviewer,
        }
        wp_repo.save_sampling_result(wp.id, params_dict, serialized)
        _audit(s, "step1_sampling", "Workpaper", wp.id, project_id,
               after={
                   "kind": kind,
                   "final_sample_size": result.size_result.final_sample_size,
                   "population_amount": result.population_amount,
               })

    return serialized, None


@app.route("/api/run", methods=["POST"])
def run():
    """샘플링 실행.

    kind="both"(default) → 채권·채무 양쪽 자동 실행.
    kind="receivable" or "payable" → 단일 실행 (하위 호환).
    """
    data = request.json or {}
    kind = data.get("kind", "both")   # default = both (자동)
    sheet = data.get("sheet")
    project_id = data.get("project_id") or STATE.get("current_project_id")

    if not STATE.get("ledger_path"):
        return jsonify({"error": "거래처원장 파일을 먼저 업로드하세요 (Step 0)"}), 400

    if project_id is None:
        log.warning("project_id 없는 /api/run 호출 — 임시 프로젝트 생성 (deprecated)")
        with get_session() as s:
            project_id, _ = _get_or_create_temp_project(s, kind)

    # 특관자 fallback
    related_parties_data = data.get("related_parties")
    if related_parties_data:
        related_parties = set(related_parties_data)
    elif STATE.get("rp_path"):
        try:
            related_parties = load_related_parties(STATE["rp_path"])
        except Exception:
            related_parties = set()
    else:
        related_parties = set()

    # ── kind="both": 채권+채무 양쪽 자동 실행 ────────────────────────────
    if kind == "both":
        results: dict[str, dict] = {}
        errors: dict[str, str] = {}
        for chk_kind in ("receivable", "payable"):
            chk_sheet = sheet or STATE.get("sheets", {}).get(chk_kind)
            serialized, err = _run_single_kind(
                data=data, kind=chk_kind, sheet=chk_sheet,
                project_id=project_id, related_parties=related_parties,
            )
            if err:
                errors[chk_kind] = err
            else:
                results[chk_kind] = serialized

        if not results:
            # 양쪽 모두 실패
            return jsonify({
                "error": "채권·채무 샘플링 모두 실패했습니다",
                "receivable_error": errors.get("receivable"),
                "payable_error": errors.get("payable"),
            }), 400

        combined_final = sum(
            (r.get("size", {}).get("final_sample_size", 0) for r in results.values()), 0
        )
        combined_population = sum(
            (r.get("population_amount", 0) for r in results.values()), 0.0
        )

        response: dict = {
            "kind": "both",
            "receivable": results.get("receivable"),
            "payable": results.get("payable"),
            "combined": {
                "total_population": combined_population,
                "total_final_sample_size": combined_final,
                "receivable_sample": results.get("receivable", {}).get("size", {}).get("final_sample_size", 0),
                "payable_sample": results.get("payable", {}).get("size", {}).get("final_sample_size", 0),
            },
        }
        if errors:
            response["warnings"] = errors
        return jsonify(response)

    # ── 단일 kind (deprecated, 하위 호환) ────────────────────────────────
    pm_raw = data.get("performance_materiality", 0)
    try:
        pm = float(pm_raw)
    except (TypeError, ValueError):
        return jsonify({"error": f"수행 중요성(performance_materiality) 값이 잘못되었습니다: {pm_raw}"}), 400
    if pm <= 0:
        return jsonify({"error": "수행 중요성(PM)은 0보다 커야 합니다"}), 400

    serialized, err = _run_single_kind(
        data=data, kind=kind, sheet=sheet,
        project_id=project_id, related_parties=related_parties,
    )
    if err:
        return jsonify({"error": err, "population_amount": 0}), 400

    return jsonify(serialized)


# ─────────────────────────────────────────────────────────────
# 4. 조서 다운로드
# ─────────────────────────────────────────────────────────────
@app.route("/api/download/<kind>")
def download(kind):
    cached = STATE["last_result"].get(kind)
    if not cached:
        return jsonify({"error": "no result"}), 404

    result = cached["result"]
    params = cached["params"]
    prefix = "C100" if kind == "receivable" else "AA100"
    fname = f"{prefix}_{params.company_name}_{params.period_end}.xlsx"
    out_path = ROOT / "output" / fname
    write_report(result, params, out_path)

    # Artifact DB 저장
    project_id = STATE.get("current_project_id")
    if project_id:
        try:
            with get_session() as s:
                wp_id = STATE["workpaper_ids"].get(kind)
                art = ArtifactRepository(s).save_file(
                    project_id=project_id,
                    kind="workpaper",
                    source_path=out_path,
                    filename=fname,
                    workpaper_id=wp_id,
                )
        except Exception as e:
            log.warning(f"조서 Artifact 저장 실패 (비치명적): {e}")

    return send_file(out_path, as_attachment=True, download_name=fname)


# ─────────────────────────────────────────────────────────────
# 5. 매트릭스 메타정보
# ─────────────────────────────────────────────────────────────
@app.route("/api/matrix")
def matrix():
    return jsonify({
        "key_item_ratio": {f"{k[0]}|{k[1]}": v for k, v in KEY_ITEM_RATIO_MATRIX.items()},
        "confidence_factor": {f"{k[0]}|{k[1]}": v for k, v in CONFIDENCE_FACTOR_MATRIX.items()},
    })


# ─────────────────────────────────────────────────────────────
# 6. Template 목록
# ─────────────────────────────────────────────────────────────
@app.route("/api/templates")
def get_templates():
    """등록된 조서 양식 목록 반환 (Step 3 드롭다운용)."""
    templates = list_templates()
    return jsonify([
        {"id": t.id, "name": t.name, "firm_name": t.firm_name}
        for t in templates
    ])


# ─────────────────────────────────────────────────────────────
# 7. Step 2 — 발송명단
# ─────────────────────────────────────────────────────────────
@app.route("/api/project/<pid>/step2/build", methods=["POST"])
def step2_build(pid: str):
    """발송명단 Excel 생성 + Artifact 저장.

    body: {kind, reply_deadline, contact_info, party_contacts}
    """
    with get_session() as s:
        proj = ProjectRepository(s).get(pid)
        if proj is None or proj.status == "archived":
            return jsonify({"error": "not found"}), 404

        data = request.json or {}
        kind = data.get("kind", STATE.get("kind") or "receivable")
        reply_deadline_str = data.get("reply_deadline")
        contact_info = data.get("contact_info") or {}
        party_contacts = data.get("party_contacts") or {}

        # Step 1 결과에서 decisions 복원
        # STATE cache는 현재 활성 프로젝트 데이터일 때만 사용 (다른 프로젝트 데이터 오염 방지)
        cached = STATE["last_result"].get(kind) if STATE.get("current_project_id") == pid else None
        if not cached:
            # DB에서 복원 — 해당 pid/kind workpaper의 sampling_result
            wp_repo = WorkpaperRepository(s)
            wp = wp_repo.get_or_create(pid, kind)
            if not wp.sampling_result:
                return jsonify({"error": "Step 1 샘플링을 먼저 실행하세요"}), 400
            result_dict = json.loads(wp.sampling_result)
            decisions = [
                PartyDecision(
                    name=d["name"],
                    balance=d["balance"],
                    is_excluded=d["is_excluded"],
                    is_related_party=d["is_related_party"],
                    is_key_item=d["is_key_item"],
                    is_representative=d["is_representative"],
                    final_sampled=d["final_sampled"],
                    exclusion_reason=d.get("exclusion_reason"),
                )
                for d in result_dict.get("decisions", [])
            ]
        else:
            decisions = cached["result"].decisions
            result_dict = _serialize_result(cached["result"], cached["params"])

        # 파싱
        try:
            reply_deadline = (
                date.fromisoformat(reply_deadline_str) if reply_deadline_str else None
            )
        except ValueError:
            reply_deadline = None

        project_info = {
            "company_name": proj.company_name,
            "period_end": proj.period_end,
            "audit_firm": proj.audit_firm,
            "preparer": contact_info.get("preparer", ""),
        }

        kind_label = "채권" if kind == "receivable" else "채무"
        fname = f"발송명단_{proj.company_name}_{proj.period_end}_{kind_label}.xlsx"
        out_dir = ROOT / "output"
        out_dir.mkdir(exist_ok=True)
        out_path = out_dir / fname

        build_send_list(
            out_path=out_path,
            project_info=project_info,
            decisions=decisions,
            kind=kind,
            reply_deadline=reply_deadline,
            contact_info=contact_info,
            party_contacts=party_contacts,
        )

        # Artifact 저장
        art_repo = ArtifactRepository(s)
        wp_repo = WorkpaperRepository(s)
        wp = wp_repo.get_or_create(pid, kind)
        art = art_repo.save_file(
            project_id=pid,
            kind="send_list",
            source_path=out_path,
            filename=fname,
            workpaper_id=wp.id,
        )
        wp.send_list_artifact_id = art.id
        wp.updated_at = datetime.now(timezone.utc)

        # AuditTrail
        _audit(s, "step2_build_send_list", "Workpaper", wp.id, pid,
               after={"kind": kind, "filename": fname, "parties": len([d for d in decisions if d.final_sampled])})

        return jsonify({
            "artifact_id": art.id,
            "download_url": f"/api/project/{pid}/step2/download/{kind}",
            "filename": fname,
            "party_count": len([d for d in decisions if d.final_sampled]),
        }), 201


@app.route("/api/project/<pid>/step2/download/<kind>")
def step2_download(pid: str, kind: str):
    """생성된 발송명단 다운로드."""
    with get_session() as s:
        wp_repo = WorkpaperRepository(s)
        wp = wp_repo.get_or_create(pid, kind)
        if not wp.send_list_artifact_id:
            return jsonify({"error": "발송명단 미생성 — 먼저 build 실행"}), 404
        art = ArtifactRepository(s).get(wp.send_list_artifact_id)
        if art is None or not Path(art.stored_path).exists():
            return jsonify({"error": "파일 없음"}), 404
        return send_file(
            art.stored_path,
            as_attachment=True,
            download_name=art.filename,
        )


@app.route("/api/project/<pid>/step2/mark-sent", methods=["POST"])
def step2_mark_sent(pid: str):
    """발송명단 회사 송부 완료 기록."""
    data = request.json or {}
    kind = data.get("kind", "receivable")
    with get_session() as s:
        proj = ProjectRepository(s).get(pid)
        if proj is None:
            return jsonify({"error": "not found"}), 404
        wp_repo = WorkpaperRepository(s)
        wp = wp_repo.get_or_create(pid, kind)
        wp.step2_completed_at = datetime.now(timezone.utc)
        wp.updated_at = datetime.now(timezone.utc)
        _audit(s, "step2_mark_sent", "Workpaper", wp.id, pid, after={"kind": kind})
        return jsonify({"ok": True, "step2_completed_at": wp.step2_completed_at.isoformat()})


# ─────────────────────────────────────────────────────────────
# 8. Step 3 — 조서 생성
# ─────────────────────────────────────────────────────────────
def _step3_build_both(pid: str, proj, data: dict, session):
    """kind=both: 채권·채무 양쪽 DB 복원 → 단일 통합 파일 생성.

    STATE에 의존하지 않고 wp.sampling_result / wp.sampling_params 만 사용.
    """
    from src.infrastructure.report.generic_reporter import (
        KindData, ReportContext, build_combined_report,
    )
    from src.domain.mus import MUSResult, MUSSelection
    from src.domain.sample_size import SampleSizeResult
    from src.domain.population import CompletenessCheck

    preparer = data.get("preparer", "")
    reviewer = data.get("reviewer", "")
    template_id = data.get("template_id", "woongkye_standard")

    wp_repo = WorkpaperRepository(session)
    art_repo = ArtifactRepository(session)
    reply_repo = ConfirmationReplyRepository(session)
    proc_repo = AlternativeProcedureRepository(session)

    kind_datas: dict[str, KindData] = {}

    for chk_kind in ("receivable", "payable"):
        wp = wp_repo.get_or_create(pid, chk_kind)
        if not wp.sampling_result or not wp.sampling_params:
            return jsonify({"error": f"Step 1 샘플링을 먼저 실행하세요 ({chk_kind})"}), 400

        # ── params 복원 ───────────────────────────────────────────────
        pd_raw = json.loads(wp.sampling_params)
        params = SamplingParams(
            company_name=pd_raw.get("company_name", proj.company_name),
            period_end=date.fromisoformat(pd_raw.get("period_end", proj.period_end)),
            kind=chk_kind,
            performance_materiality=float(pd_raw.get("performance_materiality", 0)),
            risk_level=pd_raw.get("risk_level", "유의적위험"),
            control_reliance=pd_raw.get("control_reliance", "Y"),
            preparer=preparer or pd_raw.get("preparer", ""),
            reviewer=reviewer or pd_raw.get("reviewer", ""),
        )

        # ── sampling_result 역직렬화 ──────────────────────────────────
        sr = json.loads(wp.sampling_result)

        decisions = [
            PartyDecision(
                name=d["name"], balance=d["balance"],
                is_excluded=d["is_excluded"], is_related_party=d["is_related_party"],
                is_key_item=d["is_key_item"], is_representative=d["is_representative"],
                final_sampled=d["final_sampled"], exclusion_reason=d.get("exclusion_reason"),
            )
            for d in sr.get("decisions", [])
        ]

        sz = sr.get("size", {})
        size_result = SampleSizeResult(
            key_item_threshold=sz.get("key_item_threshold", 0),
            key_item_ratio=sz.get("key_item_ratio", 0),
            confidence_factor=sz.get("confidence_factor", 0),
            base_sample_size=sz.get("base_sample_size", 0),
            final_sample_size=sz.get("final_sample_size", 0),
            sample_interval=sz.get("sample_interval", 0),
            remaining_population=sz.get("remaining_population", 0),
        )

        from src.domain.population import CompletenessCheck as _CC
        comp_raw = sr.get("completeness", {})
        completeness = _CC(
            by_group=comp_raw.get("rows", []),
            total_ledger=comp_raw.get("total_ledger", 0),
            total_fs=comp_raw.get("total_fs", 0),
            total_diff=comp_raw.get("total_diff", 0),
        )

        mus_raw = sr.get("mus", {})
        selections = [
            MUSSelection(
                name=sel["name"], balance=sel["balance"], cumulative=sel["cumulative"],
                selections=sel["selections"], remainder_after=sel["remainder_after"],
                hit=sel["hit"],
            )
            for sel in mus_raw.get("selections", [])
        ]
        mus_result = MUSResult(
            sample_interval=mus_raw.get("sample_interval", 0),
            random_start=mus_raw.get("random_start", 0),
            selections=selections,
            sampled_names=mus_raw.get("sampled_names", []),
        )

        # ── UploadGuide 연락처 ────────────────────────────────────────
        contacts: list[PartyContactInfo] = []
        exclusion_rows_list: list[ExclusionRow] = []
        ug_data = STATE.get("upload_guide_data")
        if ug_data is None and STATE.get("upload_guide_path"):
            try:
                ug_data = load_upload_guide(STATE["upload_guide_path"])
                STATE["upload_guide_data"] = ug_data
            except Exception as e:
                log.warning(f"UploadGuide 재로드 실패: {e}")
        if ug_data:
            contacts = [
                PartyContactInfo(
                    name=ct.name, country=ct.country, business_no=ct.business_no,
                    ceo_name=ct.ceo_name, contact_person=ct.contact_person,
                    phone=ct.phone, email=ct.email,
                )
                for ct in ug_data.send_targets
            ]
            exclusion_rows_list = [
                ExclusionRow(
                    name=ex.name, account_name=ex.account_name,
                    currency=ex.currency, amount=ex.amount, kind=ex.kind,
                )
                for ex in ug_data.excluded
            ]

        # ── PDF 회신 ─────────────────────────────────────────────────
        pdf_replies_list: list[ConfirmationReplyInfo] = []
        try:
            replies_db = reply_repo.list_by_workpaper(wp.id)
            for rep in replies_db:
                pdf_replies_list.append(ConfirmationReplyInfo(
                    party_name=rep.party_name_matched or rep.party_name_raw,
                    status=rep.status,
                    extracted_balance=rep.extracted_balance,
                    reply_date=rep.reply_date,
                ))
        except Exception as e:
            log.warning(f"PDF 회신 로드 실패 ({chk_kind}): {e}")

        # ── 대체적 절차 ───────────────────────────────────────────────
        alt_procs_list: list[AlternativeProcedureEntry] = []
        try:
            procs_db = proc_repo.list_by_workpaper(wp.id)
            for proc in procs_db:
                ev_ids: list[str] = []
                if proc.evidence_artifact_ids:
                    try:
                        ev_ids = json.loads(proc.evidence_artifact_ids)
                    except Exception:
                        ev_ids = []
                alt_procs_list.append(AlternativeProcedureEntry(
                    party_name=proc.party_name, reason=proc.reason,
                    ledger_balance=proc.ledger_balance, procedure_type=proc.procedure_type,
                    evidence_names=ev_ids, covered_amount=proc.covered_amount,
                    coverage_ratio=proc.coverage_ratio, conclusion=proc.conclusion,
                    auditor_notes=proc.auditor_notes,
                ))
        except Exception as e:
            log.warning(f"대체적 절차 로드 실패 ({chk_kind}): {e}")

        prefix = "C100" if chk_kind == "receivable" else "AA100"
        ctx = ReportContext(
            company_name=params.company_name,
            period_end=params.period_end,
            kind=chk_kind,
            preparer=params.preparer,
            reviewer=params.reviewer,
            workpaper_no_prefix=prefix,
        )

        kind_datas[chk_kind] = KindData(
            ctx=ctx,
            completeness=completeness,
            size_result=size_result,
            decisions=decisions,
            mus_result=mus_result,
            performance_materiality=params.performance_materiality,
            population_amount=sr.get("population_amount", 0.0),
            contacts=contacts or None,
            exclusion_rows=exclusion_rows_list or None,
            pdf_replies=pdf_replies_list or None,
            alt_procedures=alt_procs_list or None,
        )

    # ── 통합 파일 생성 ────────────────────────────────────────────────
    # ReportContext.kind = "both" 로 덮어써서 헤더에 "채권채무 조회 통합" 표시
    ar_kd = kind_datas.get("receivable")
    ap_kd = kind_datas.get("payable")
    if ar_kd:
        ar_kd.ctx.kind = "both"
    elif ap_kd:
        ap_kd.ctx.kind = "both"

    base_kd = ar_kd or ap_kd
    base_params = json.loads(wp_repo.get_or_create(pid, "receivable" if ar_kd else "payable").sampling_params)

    fname = f"C100AA100_{proj.company_name}_{proj.period_end}.xlsx"
    out_path = ROOT / "output" / fname
    ug_data_for_report = STATE.get("upload_guide_data")
    if ug_data_for_report is None and STATE.get("upload_guide_path"):
        try:
            ug_data_for_report = load_upload_guide(STATE["upload_guide_path"])
            STATE["upload_guide_data"] = ug_data_for_report
        except Exception as e:
            log.warning(f"UploadGuide 재로드 실패 (step3_both): {e}")
    build_combined_report(out_path, receivable=ar_kd, payable=ap_kd,
                          upload_guide_data=ug_data_for_report)

    # ── Artifact 저장 (채권 workpaper 기준) ──────────────────────────
    base_wp = wp_repo.get_or_create(pid, "receivable" if ar_kd else "payable")
    art = art_repo.save_file(
        project_id=pid,
        kind="workpaper",
        source_path=out_path,
        filename=fname,
        workpaper_id=base_wp.id,
    )
    base_wp.workpaper_artifact_id = art.id

    from datetime import datetime, timezone
    base_wp.updated_at = datetime.now(timezone.utc)

    _audit(session, "step3_export_workpaper_both", "Workpaper", base_wp.id, pid,
           after={"kind": "both", "template_id": template_id, "filename": fname})

    return jsonify({
        "artifact_id": art.id,
        "download_url": f"/api/project/{pid}/step3/download/receivable",
        "filename": fname,
        "kind": "both",
    }), 201


@app.route("/api/project/<pid>/step3/build", methods=["POST"])
def step3_build(pid: str):
    """조서 Excel 생성 + Artifact 저장.

    body: {kind, template_id, preparer, reviewer}
    kind="both" → 채권+채무 단일 통합 파일
    """
    with get_session() as s:
        proj = ProjectRepository(s).get(pid)
        if proj is None or proj.status == "archived":
            return jsonify({"error": "not found"}), 404

        data = request.json or {}
        kind = data.get("kind", "both")  # default = both — 단일 통합 파일

        # kind=both: 채권+채무 단일 통합 파일 빌드
        if kind == "both":
            return _step3_build_both(pid, proj, data, s)
        template_id = data.get("template_id", "woongkye_standard")
        preparer = data.get("preparer", "")
        reviewer = data.get("reviewer", "")

        # Step 1 결과 복원
        cached = STATE["last_result"].get(kind)
        if not cached:
            wp_repo = WorkpaperRepository(s)
            wp = wp_repo.get_or_create(pid, kind)
            if not wp.sampling_result or not wp.sampling_params:
                return jsonify({"error": "Step 1 샘플링을 먼저 실행하세요"}), 400
            params_dict = json.loads(wp.sampling_params)
            params = SamplingParams(
                company_name=params_dict.get("company_name", proj.company_name),
                period_end=date.fromisoformat(params_dict.get("period_end", proj.period_end)),
                kind=kind,
                performance_materiality=float(params_dict.get("performance_materiality", 0)),
                risk_level=params_dict.get("risk_level", "유의적위험"),
                control_reliance=params_dict.get("control_reliance", "Y"),
                preparer=preparer or params_dict.get("preparer", ""),
                reviewer=reviewer or params_dict.get("reviewer", ""),
            )
            # DB 재실행 없이 임시 빌드를 위해 ledger 재로드 필요 — STATE 미캐시 상황
            # 이미 STATE에 ledger가 있으면 재실행, 없으면 에러
            if not STATE.get("ledger_path"):
                return jsonify({"error": "서버 재기동 후 Step 1을 다시 실행하세요"}), 400
            df = pd.read_excel(STATE["ledger_path"], sheet_name=STATE["sheets"].get(kind))
            result = run_sampling(df, params)
        else:
            result = cached["result"]
            params = cached["params"]
            if preparer:
                params.preparer = preparer
            if reviewer:
                params.reviewer = reviewer

        # ── UploadGuide 연락처 로드 ──────────────────────────────
        contacts: list[PartyContactInfo] = []
        exclusion_rows_list: list[ExclusionRow] = []
        ug_data: UploadGuideData | None = STATE.get("upload_guide_data")
        if ug_data is None and STATE.get("upload_guide_path"):
            try:
                ug_data = load_upload_guide(STATE["upload_guide_path"])
                STATE["upload_guide_data"] = ug_data
            except Exception as e:
                log.warning(f"UploadGuide 재로드 실패: {e}")
        if ug_data:
            contacts = [
                PartyContactInfo(
                    name=ct.name,
                    country=ct.country,
                    business_no=ct.business_no,
                    ceo_name=ct.ceo_name,
                    contact_person=ct.contact_person,
                    phone=ct.phone,
                    email=ct.email,
                )
                for ct in ug_data.send_targets
            ]
            exclusion_rows_list = [
                ExclusionRow(
                    name=ex.name,
                    account_name=ex.account_name,
                    currency=ex.currency,
                    amount=ex.amount,
                    kind=ex.kind,
                )
                for ex in ug_data.excluded
            ]

        # ── PDF 회신 결과 로드 (Step 4 완료 시) ─────────────────
        pdf_replies_list: list[ConfirmationReplyInfo] = []
        try:
            with get_session() as s2:
                wp_repo2 = WorkpaperRepository(s2)
                wp2 = wp_repo2.get_or_create(pid, kind)
                from src.infrastructure.persistence import ConfirmationReplyRepository
                reply_repo2 = ConfirmationReplyRepository(s2)
                replies_db = reply_repo2.list_by_workpaper(wp2.id)
                for rep in replies_db:
                    pdf_replies_list.append(ConfirmationReplyInfo(
                        party_name=rep.party_name_matched or rep.party_name_raw,
                        status=rep.status,
                        extracted_balance=rep.extracted_balance,
                        reply_date=rep.reply_date,
                    ))
        except Exception as e:
            log.warning(f"PDF 회신 로드 실패 (비치명적): {e}")

        # ── 대체적 절차 로드 (Step 5 완료 시) ──────────────────
        alt_procs_list: list[AlternativeProcedureEntry] = []
        try:
            with get_session() as s3:
                from src.infrastructure.persistence import AlternativeProcedureRepository
                wp_repo3 = WorkpaperRepository(s3)
                wp3 = wp_repo3.get_or_create(pid, kind)
                proc_repo3 = AlternativeProcedureRepository(s3)
                procs_db = proc_repo3.list_by_workpaper(wp3.id)
                for proc in procs_db:
                    import json as _json2
                    ev_ids: list[str] = []
                    if proc.evidence_artifact_ids:
                        try:
                            ev_ids = _json2.loads(proc.evidence_artifact_ids)
                        except Exception:
                            ev_ids = []
                    # evidence names — artifact ID만 있으므로 ID를 이름 대용으로 표시
                    alt_procs_list.append(AlternativeProcedureEntry(
                        party_name=proc.party_name,
                        reason=proc.reason,
                        ledger_balance=proc.ledger_balance,
                        procedure_type=proc.procedure_type,
                        evidence_names=ev_ids,
                        covered_amount=proc.covered_amount,
                        coverage_ratio=proc.coverage_ratio,
                        conclusion=proc.conclusion,
                        auditor_notes=proc.auditor_notes,
                    ))
        except Exception as e:
            log.warning(f"대체적 절차 로드 실패 (비치명적): {e}")

        prefix = "C100" if kind == "receivable" else "AA100"
        fname = f"{prefix}_{proj.company_name}_{proj.period_end}.xlsx"
        out_path = ROOT / "output" / fname
        write_report(
            result, params, out_path,
            template_id=template_id,
            contacts=contacts or None,
            exclusion_rows=exclusion_rows_list or None,
            pdf_replies=pdf_replies_list or None,
            alt_procedures=alt_procs_list or None,
        )

        # Artifact 저장
        art_repo = ArtifactRepository(s)
        wp_repo = WorkpaperRepository(s)
        wp = wp_repo.get_or_create(pid, kind)
        art = art_repo.save_file(
            project_id=pid,
            kind="workpaper",
            source_path=out_path,
            filename=fname,
            workpaper_id=wp.id,
        )
        wp.workpaper_artifact_id = art.id
        wp.updated_at = datetime.now(timezone.utc)

        _audit(s, "step3_export_workpaper", "Workpaper", wp.id, pid,
               after={"kind": kind, "template_id": template_id, "filename": fname})

        return jsonify({
            "artifact_id": art.id,
            "download_url": f"/api/project/{pid}/step3/download/{kind}",
            "filename": fname,
        }), 201


@app.route("/api/project/<pid>/step3/download/<kind>")
def step3_download(pid: str, kind: str):
    """생성된 조서 다운로드."""
    with get_session() as s:
        wp_repo = WorkpaperRepository(s)
        wp = wp_repo.get_or_create(pid, kind)
        if not wp.workpaper_artifact_id:
            return jsonify({"error": "조서 미생성 — 먼저 build 실행"}), 404
        art = ArtifactRepository(s).get(wp.workpaper_artifact_id)
        if art is None or not Path(art.stored_path).exists():
            return jsonify({"error": "파일 없음"}), 404
        return send_file(
            art.stored_path,
            as_attachment=True,
            download_name=art.filename,
        )


@app.route("/api/project/<pid>/step3/mark-done", methods=["POST"])
def step3_mark_done(pid: str):
    """조서 작성 완료 기록."""
    data = request.json or {}
    kind = data.get("kind", "receivable")
    with get_session() as s:
        proj = ProjectRepository(s).get(pid)
        if proj is None:
            return jsonify({"error": "not found"}), 404
        wp_repo = WorkpaperRepository(s)
        wp = wp_repo.get_or_create(pid, kind)
        wp.step3_completed_at = datetime.now(timezone.utc)
        wp.updated_at = datetime.now(timezone.utc)
        _audit(s, "step3_mark_done", "Workpaper", wp.id, pid, after={"kind": kind})
        return jsonify({"ok": True, "step3_completed_at": wp.step3_completed_at.isoformat()})


# ─────────────────────────────────────────────────────────────
# Step 4: PDF 회신 업로드·자동처리·수동보정·완료 기록
# ─────────────────────────────────────────────────────────────
@app.route("/api/project/<pid>/step4/upload-replies", methods=["POST"])
def step4_upload_replies(pid: str):
    """PDF 회신 다중 업로드 → 자동 추출·매칭·차이판정 → ConfirmationReply 생성.

    multipart form-data: files[]=<pdf>, kind=receivable|payable, tolerance=0
    """
    with get_session() as s:
        proj = ProjectRepository(s).get(pid)
        if proj is None or proj.status == "archived":
            return jsonify({"error": "not found"}), 404

        kind_raw = request.form.get("kind", "auto")
        # "auto" 또는 "both" → receivable을 Artifact FK 기준으로, 실제 분류는 양쪽 자동
        kind = "receivable" if kind_raw in ("auto", "both") else kind_raw
        try:
            tolerance = float(request.form.get("tolerance", "0") or "0")
        except ValueError:
            tolerance = 0.0

        # Step 1 결과에서 candidates (거래처명) + ledger_balance 가져오기
        # Task 1: 한 거래처가 채권·채무 양쪽 final_sampled 일 수 있으므로 양쪽 합집합
        wp_repo = WorkpaperRepository(s)
        wp = wp_repo.get_or_create(pid, kind)

        # 양쪽 kind 합집합 candidates (name → {balance, kind})
        candidates_full: dict[str, dict] = {}  # name → {balance, kind}
        for chk_kind in ("receivable", "payable"):
            chk_wp = wp_repo.get_or_create(pid, chk_kind)
            if chk_wp.sampling_result:
                result_dict = json.loads(chk_wp.sampling_result)
                for d in result_dict.get("decisions", []):
                    if d.get("final_sampled"):
                        name = d["name"]
                        bal = float(d.get("balance", 0))
                        if name not in candidates_full:
                            candidates_full[name] = {"balance": bal, "kind": chk_kind}
                        else:
                            # 이미 있으면 kind="both" 표시
                            candidates_full[name]["kind"] = "both"

        candidates: list[str] = list(candidates_full.keys())
        ledger_map: dict[str, float] = {n: v["balance"] for n, v in candidates_full.items()}
        kind_map: dict[str, str] = {n: v["kind"] for n, v in candidates_full.items()}

        # 요청된 kind wp (Artifact FK용)
        # wp는 이미 위에서 get_or_create(pid, kind) 완료

        files = request.files.getlist("files[]")
        if not files:
            return jsonify({"error": "업로드된 PDF 파일이 없습니다"}), 400

        art_repo = ArtifactRepository(s)
        reply_repo = ConfirmationReplyRepository(s)

        # PDF 형식 검증
        non_pdf = [f.filename for f in files if f.filename and not _validate_file_ext(f.filename, _ALLOWED_PDF_EXT)]
        if non_pdf:
            return jsonify({
                "error": f"PDF 파일이 아닌 파일이 포함되어 있습니다: {', '.join(non_pdf)}. PDF(.pdf)만 업로드하세요."
            }), 400

        results = []
        for f in files:
            if not f.filename:
                continue
            filename = f.filename

            # 1) Artifact 저장
            pdf_bytes = f.read()
            if len(pdf_bytes) == 0:
                results.append({"filename": filename, "status": "error", "error": "빈 파일입니다"})
                continue
            artifact = art_repo.save_bytes(
                project_id=pid,
                kind="pdf_reply",
                content=pdf_bytes,
                filename=filename,
                workpaper_id=wp.id,
            )

            # 2) 텍스트 추출 (tables 포함)
            from pathlib import Path as _Path
            extract_result = extract_text(_Path(artifact.stored_path))
            ocr_warnings = extract_result.warnings

            # 3a) 양식 분류
            file_meta = {"filename": filename}
            form_profile = detect_form(
                extract_result.full_text,
                tables=extract_result.tables or None,
                file_meta=file_meta,
            )
            patterns = get_patterns(form_profile.form_id)

            # 3b) Parser v2 (declared_match + per_account_rows + original_currency)
            parsed = parse_confirmation_v2(
                extract_result.full_text,
                tables=extract_result.tables or None,
                patterns=patterns,
                filename_hint=filename,
            )

            # 4) 거래처 매칭 v2 (UploadGuide 동적 alias + CJK + 사업자번호)
            raw_name = parsed.extracted_name or filename
            ug_data = STATE.get("upload_guide_data")
            if candidates:
                match_result = match_party(
                    raw_name,
                    candidates,
                    upload_guide_data=ug_data,
                    business_no=None,   # PDF에서 사업자번호 추출 미구현 → None
                    filename_hint=filename,
                )
                matched_name = match_result.matched_name
                match_conf = match_result.confidence
                match_method = match_result.method
            else:
                matched_name = None
                match_conf = 0.0
                match_method = "failed"
                match_result = None

            # 4b) 매칭 실패 시 normalize 재시도 (match_party 임계값 미달 케이스 구제)
            # match_party가 top3_candidates를 반환한 경우 가장 높은 후보의 normalize가
            # raw_name normalize와 동일하면 매칭 성공으로 처리.
            if matched_name is None and match_result and match_result.candidates:
                from src.domain.matching import _normalize as _match_norm
                raw_norm = _match_norm(raw_name)
                for cand in match_result.candidates:
                    cand_name = cand if isinstance(cand, str) else (cand.get("name") if isinstance(cand, dict) else str(cand))
                    if _match_norm(cand_name) == raw_norm:
                        matched_name = cand_name
                        match_conf = 0.85
                        match_method = "normalize_retry"
                        break

            # 5) 차이 판정 v2
            ledger_bal = ledger_map.get(matched_name) if matched_name else None

            # UploadGuide 거래처 행 조회
            upload_guide_row = None
            if ug_data and matched_name:
                upload_guide_row = ug_data.contact_map().get(matched_name)

            currency_resolver = CurrencyResolver(ug_data, STATE.get("manual_rates"))

            recon = reconcile_v2(
                ledger_balance_krw=ledger_bal if ledger_bal is not None else 0.0,
                parsed_reply=parsed,
                upload_guide_row=upload_guide_row,
                currency_resolver=currency_resolver,
                tolerance=tolerance,
                tolerance_pct=0.01,
            )
            status = recon.status
            if matched_name is None:
                status = "needs_review"

            # 5b) declared=False 이지만 차이 5% 이내 → matched 자동 보정
            # (PDF 자체 선언이 틀린 경우 — 감사 기준상 실질 잔액 일치가 우선)
            if (
                status == "mismatch"
                and parsed.declared_match is False
                and recon.difference_pct is not None
                and recon.difference_pct <= 0.05
            ):
                status = "matched"
                log.info(
                    "declared=False 이지만 차이 %.1f%% ≤ 5%% → matched 자동 보정: %s",
                    (recon.difference_pct or 0) * 100,
                    raw_name,
                )

            # top3 candidates (매칭 실패 시)
            top3_json = None
            if match_result and match_result.method == "failed" and match_result.candidates:
                top3_json = json.dumps(match_result.candidates[:3], ensure_ascii=False)

            # per_account_findings JSON 직렬화
            per_acct_json = None
            if recon.per_account_findings:
                per_acct_json = json.dumps(recon.per_account_findings, ensure_ascii=False)

            # 6) DB 저장 — 양쪽 표 자동 분리 + 거래처 양쪽 등록
            # PDF per_account_rows 섹션 분석: 실제 채권/채무 표가 있는 쪽 파악
            pdf_has_receivable = any(
                row.section == "receivable" for row in (parsed.per_account_rows or [])
            )
            pdf_has_payable = any(
                row.section == "payable" for row in (parsed.per_account_rows or [])
            )
            # per_account_rows 없으면 종합 extracted_balance 섹션 추정
            if not parsed.per_account_rows:
                pdf_has_receivable = parsed.receivable_total is not None
                pdf_has_payable = parsed.payable_total is not None

            # 채권/채무 표별 회신금액 분리
            recv_balance = parsed.receivable_total
            payb_balance = parsed.payable_total

            # 거래처 kind 판정 (Step 1 final_sampled 기반)
            matched_kinds: list[str] = []
            if matched_name:
                party_kind = kind_map.get(matched_name, kind)
                if party_kind == "both":
                    matched_kinds = ["receivable", "payable"]
                else:
                    # PDF에 반대쪽 표도 있으면 추가 등록 시도
                    matched_kinds = [party_kind]
                    if party_kind == "receivable" and pdf_has_payable and payb_balance is not None:
                        matched_kinds.append("payable")
                    elif party_kind == "payable" and pdf_has_receivable and recv_balance is not None:
                        matched_kinds.insert(0, "receivable")
            else:
                matched_kinds = [kind]  # 매칭 실패 → 요청 kind에만 저장

            for target_kind in matched_kinds:
                # target_kind에 실제 해당 거래처가 sampled 되어 있는지 확인
                if matched_name and matched_name not in ledger_map:
                    pass  # 매칭 실패 케이스 — 그냥 진행

                target_wp = wp_repo.get_or_create(pid, target_kind)
                # 같은 거래처·같은 PDF 중복 방지 (이미 존재하면 skip)
                existing_replies = reply_repo.list_by_workpaper(target_wp.id)
                already_exists = any(
                    r.pdf_artifact_id == artifact.id and r.party_name_matched == matched_name
                    for r in existing_replies
                )
                if already_exists:
                    continue

                # target_kind 의 회신금액 선택 (표 섹션 기반)
                if target_kind == "receivable" and recv_balance is not None:
                    target_extracted_balance = recv_balance
                elif target_kind == "payable" and payb_balance is not None:
                    target_extracted_balance = payb_balance
                else:
                    # 하위 호환: 단일 extracted_balance 사용
                    target_extracted_balance = parsed.extracted_balance

                # target_kind의 장부 잔액 재조회
                target_ledger_bal = ledger_map.get(matched_name) if matched_name else ledger_bal
                if target_ledger_bal is None and matched_name:
                    target_wp_sr = target_wp.sampling_result
                    if target_wp_sr:
                        _sr = json.loads(target_wp_sr)
                        for _d in _sr.get("decisions", []):
                            if _d.get("final_sampled") and _d["name"] == matched_name:
                                target_ledger_bal = float(_d.get("balance", 0))
                                break

                # 차이 재계산 (target_kind별 금액 기준)
                if target_extracted_balance is not None and target_ledger_bal is not None:
                    diff = target_ledger_bal - target_extracted_balance
                    diff_pct = abs(diff / target_ledger_bal) if target_ledger_bal else None
                    if abs(diff) <= tolerance or (diff_pct is not None and diff_pct <= 0.01):
                        target_status = "matched"
                    else:
                        target_status = "mismatch"
                else:
                    diff = recon.difference
                    diff_pct = recon.difference_pct
                    target_status = status

                if matched_name is None:
                    target_status = "needs_review"

                reply = reply_repo.create(
                    workpaper_id=target_wp.id,
                    pdf_artifact_id=artifact.id,
                    party_name_raw=raw_name,
                    party_name_matched=matched_name,
                    party_match_confidence=match_conf,
                    party_match_method=match_method,
                    extracted_balance=target_extracted_balance,
                    extracted_balance_currency=parsed.balance_currency,
                    reply_date=parsed.reply_date,
                    ledger_balance=target_ledger_bal,
                    difference=diff,
                    difference_pct=diff_pct,
                    status=target_status,
                    extraction_method=extract_result.method,
                    extraction_confidence=extract_result.confidence,
                    notes="; ".join(ocr_warnings) if ocr_warnings else None,
                    # v2 확장 필드
                    declared_match=parsed.declared_match,
                    per_account_findings=per_acct_json,
                    original_currency=parsed.original_currency,
                    decision_basis=recon.decision_basis,
                    top3_candidates=top3_json,
                )

                _audit(s, "step4_upload_reply", "ConfirmationReply", reply.id, pid,
                       after={
                           "filename": filename,
                           "status": target_status,
                           "extracted_balance": target_extracted_balance,
                           "party_name_raw": raw_name,
                           "party_name_matched": matched_name,
                           "target_kind": target_kind,
                       })

                results.append({**_serialize_reply(reply, ocr_warnings), "kind": target_kind})

        return jsonify(results), 201


@app.route("/api/project/<pid>/step4/replies", methods=["GET"])
def step4_list_replies(pid: str):
    """워크페이퍼별 회신 일람 조회."""
    kind = request.args.get("kind", "receivable")
    with get_session() as s:
        proj = ProjectRepository(s).get(pid)
        if proj is None:
            return jsonify({"error": "not found"}), 404
        wp_repo = WorkpaperRepository(s)
        wp = wp_repo.get_or_create(pid, kind)
        reply_repo = ConfirmationReplyRepository(s)
        replies = reply_repo.list_by_workpaper(wp.id)
        return jsonify([_serialize_reply(r) for r in replies])


@app.route("/api/project/<pid>/step4/reply/<reply_id>", methods=["PATCH"])
def step4_patch_reply(pid: str, reply_id: str):
    """회신 수동 보정 — party_name, extracted_balance, reply_date, status."""
    data = request.json or {}
    with get_session() as s:
        proj = ProjectRepository(s).get(pid)
        if proj is None:
            return jsonify({"error": "not found"}), 404

        reply_repo = ConfirmationReplyRepository(s)
        reply = reply_repo.update_reviewer_confirmation(
            reply_id=reply_id,
            reviewer_confirmed_status="overridden",
            party_name_matched=data.get("party_name_matched"),
            extracted_balance=_f(data.get("extracted_balance")),
            reply_date=data.get("reply_date"),
            status=data.get("status"),
            notes=data.get("notes"),
        )
        if reply is None:
            return jsonify({"error": "reply not found"}), 404

        _audit(s, "step4_reviewer_override", "ConfirmationReply", reply_id, pid,
               after={k: v for k, v in data.items() if k in (
                   "party_name_matched", "extracted_balance", "reply_date", "status", "notes"
               )})

        return jsonify(_serialize_reply(reply))


@app.route("/api/project/<pid>/step4/mark-done", methods=["POST"])
def step4_mark_done(pid: str):
    """Step 4 완료 기록."""
    data = request.json or {}
    kind = data.get("kind", "receivable")
    with get_session() as s:
        proj = ProjectRepository(s).get(pid)
        if proj is None:
            return jsonify({"error": "not found"}), 404
        wp_repo = WorkpaperRepository(s)
        wp = wp_repo.get_or_create(pid, kind)
        wp.step4_completed_at = datetime.now(timezone.utc)
        wp.updated_at = datetime.now(timezone.utc)
        _audit(s, "step4_mark_done", "Workpaper", wp.id, pid, after={"kind": kind})
        return jsonify({"ok": True, "step4_completed_at": wp.step4_completed_at.isoformat()})


# ─────────────────────────────────────────────────────────────
# Step 5: 대체적 절차 — 미회신·차이 자동식별, 증빙 업로드, AlternativeProcedure CRUD
# ─────────────────────────────────────────────────────────────

@app.route("/api/project/<pid>/step5/pending", methods=["GET"])
def step5_pending(pid: str):
    """미회신·차이 거래처 자동 식별.

    미회신 = final_sampled 거래처 중 ConfirmationReply 가 없는 것
    차이   = ConfirmationReply.status == "mismatch"
    """
    kind = request.args.get("kind", "receivable")
    with get_session() as s:
        proj = ProjectRepository(s).get(pid)
        if proj is None or proj.status == "archived":
            return jsonify({"error": "not found"}), 404

        wp_repo = WorkpaperRepository(s)
        wp = wp_repo.get_or_create(pid, kind)

        # Step 1 결과에서 최종 선택 거래처
        sampled_parties: dict[str, float] = {}  # party_name → ledger_balance
        if wp.sampling_result:
            result_dict = json.loads(wp.sampling_result)
            for d in result_dict.get("decisions", []):
                if d.get("final_sampled"):
                    sampled_parties[d["name"]] = float(d.get("balance", 0))

        # 회신 현황
        reply_repo = ConfirmationReplyRepository(s)
        replies = reply_repo.list_by_workpaper(wp.id)
        replied_map: dict[str, str] = {}  # matched_name → status
        for r in replies:
            if r.party_name_matched:
                replied_map[r.party_name_matched] = r.status

        pending: list[dict] = []
        for party, balance in sampled_parties.items():
            status = replied_map.get(party)
            if status is None:
                reason = "미회신"
            elif status == "mismatch":
                reason = "차이"
            else:
                continue  # 정상 회신 → 대체절차 불필요

            # 회신값·차이 조회
            reply_balance: float | None = None
            difference: float | None = None
            for r in replies:
                if r.party_name_matched == party and r.status == "mismatch":
                    reply_balance = r.extracted_balance
                    difference = r.difference
                    break

            pending.append({
                "party_name": party,
                "reason": reason,
                "ledger_balance": balance,
                "reply_balance": reply_balance,
                "difference": difference,
            })

        return jsonify({
            "kind": kind,
            "workpaper_id": wp.id,
            "pending": pending,
            "total": len(pending),
        })


@app.route("/api/project/<pid>/step5/upload-evidence", methods=["POST"])
def step5_upload_evidence(pid: str):
    """단일 거래처 증빙 다중 업로드 → 자동 추출 → AlternativeProcedure 생성/갱신.

    multipart form-data:
      files[]        — 증빙 파일 (1개 이상)
      party_name     — 거래처명
      kind           — receivable | payable
      procedure_type — 후속입금 | 매출증빙 | 발주서대조 | 후속송금 | 기타
      auditor_notes  — 감사인 메모
      ledger_balance — 장부가 (선택)
      reply_balance  — 회신금액 (선택)
    """
    from src.infrastructure.evidence.extractor import extract_evidence
    import tempfile

    with get_session() as s:
        proj = ProjectRepository(s).get(pid)
        if proj is None or proj.status == "archived":
            return jsonify({"error": "not found"}), 404

        kind = request.form.get("kind", "receivable")
        party_name = (request.form.get("party_name") or "").strip()
        procedure_type = request.form.get("procedure_type", "기타")
        auditor_notes = request.form.get("auditor_notes", "")
        ledger_balance = _f(request.form.get("ledger_balance"))
        reply_balance = _f(request.form.get("reply_balance"))

        if not party_name:
            return jsonify({"error": "party_name 필수"}), 400

        files = request.files.getlist("files[]")
        if not files:
            return jsonify({"error": "업로드 파일 없음"}), 400

        wp_repo = WorkpaperRepository(s)
        wp = wp_repo.get_or_create(pid, kind)
        art_repo = ArtifactRepository(s)
        proc_repo = AlternativeProcedureRepository(s)

        # 기존 AlternativeProcedure 조회 (동일 거래처면 갱신)
        proc = proc_repo.get_by_party(wp.id, party_name)

        # 기존 artifact_ids 유지
        existing_ids: list[str] = []
        if proc and proc.evidence_artifact_ids:
            try:
                existing_ids = json.loads(proc.evidence_artifact_ids)
            except Exception:
                existing_ids = []

        # 파일 저장 + 추출
        new_ids: list[str] = []
        extracts = []
        for f in files:
            if not f.filename:
                continue
            file_bytes = f.read()
            art = art_repo.save_bytes(
                project_id=pid,
                kind="evidence",
                content=file_bytes,
                filename=f.filename,
                workpaper_id=wp.id,
            )
            new_ids.append(art.id)

            # 자동 추출
            stored_path = Path(art.stored_path)
            ex = extract_evidence(stored_path)
            extracts.append({
                "artifact_id": art.id,
                "filename": f.filename,
                "document_type": ex.document_type,
                "extracted_amount": ex.extracted_amount,
                "extracted_currency": ex.extracted_currency,
                "extracted_date": ex.extracted_date.isoformat() if ex.extracted_date else None,
                "extraction_method": ex.extraction_method,
                "confidence": ex.confidence,
            })

        all_ids = existing_ids + new_ids

        # 커버리지 계산
        covered_amount = _compute_covered_amount(extracts)
        coverage_ratio: float | None = None
        if ledger_balance and ledger_balance > 0 and covered_amount is not None:
            coverage_ratio = min(1.0, covered_amount / ledger_balance)

        # 결론 자동 판정
        conclusion = _auto_conclusion(coverage_ratio)

        if proc is None:
            # 신규 생성
            # reason 추론: Step 5 pending 에서 조회
            reason = _infer_reason(s, wp.id, party_name)
            proc = proc_repo.create(
                workpaper_id=wp.id,
                party_name=party_name,
                reason=reason,
                ledger_balance=ledger_balance,
                reply_balance=reply_balance,
                difference=(ledger_balance - reply_balance) if (ledger_balance and reply_balance) else None,
                procedure_type=procedure_type,
                evidence_artifact_ids=all_ids,
                covered_amount=covered_amount,
                coverage_ratio=coverage_ratio,
                conclusion=conclusion,
                auditor_notes=auditor_notes or None,
                status="pending",
            )
        else:
            # 갱신
            new_covered = (proc.covered_amount or 0.0) + (covered_amount or 0.0)
            new_lb = ledger_balance or proc.ledger_balance
            new_ratio: float | None = None
            if new_lb and new_lb > 0:
                new_ratio = min(1.0, new_covered / new_lb)
            proc_repo.update(
                proc.id,
                evidence_artifact_ids=all_ids,
                covered_amount=new_covered,
                coverage_ratio=new_ratio,
                conclusion=_auto_conclusion(new_ratio),
                procedure_type=procedure_type or proc.procedure_type,
                auditor_notes=auditor_notes or proc.auditor_notes,
            )

        _audit(s, "step5_upload_evidence", "AlternativeProcedure", proc.id, pid,
               after={
                   "party_name": party_name,
                   "files": len(new_ids),
                   "covered_amount": covered_amount,
                   "coverage_ratio": coverage_ratio,
                   "conclusion": conclusion,
               })

        return jsonify({
            "procedure_id": proc.id,
            "party_name": party_name,
            "evidence_count": len(all_ids),
            "extracts": extracts,
            "covered_amount": covered_amount,
            "coverage_ratio": coverage_ratio,
            "conclusion": conclusion,
        }), 201


@app.route("/api/project/<pid>/step5/upload-folder", methods=["POST"])
def step5_upload_folder(pid: str):
    """BC-N 폴더 zip 업로드 OR 다중 파일 (폴더명 자동 매핑).

    multipart form-data:
      folder_zip     — zip 파일 (선택)
      files[]        — 다중 파일 (선택; folder_zip 없을 때)
      folder_name    — 폴더명 힌트 (예: BC-14_New Future International Trade Co)
      kind           — receivable | payable
    """
    import zipfile, tempfile

    with get_session() as s:
        proj = ProjectRepository(s).get(pid)
        if proj is None or proj.status == "archived":
            return jsonify({"error": "not found"}), 404

        kind = request.form.get("kind", "receivable")
        folder_name_hint = request.form.get("folder_name", "")

        wp_repo = WorkpaperRepository(s)
        wp = wp_repo.get_or_create(pid, kind)
        art_repo = ArtifactRepository(s)
        proc_repo = AlternativeProcedureRepository(s)

        from src.infrastructure.evidence.aggregator import aggregate_folder, parse_bc_folder_name
        from src.infrastructure.evidence.extractor import extract_evidence

        tmp_dir = Path(tempfile.mkdtemp())
        results: list[dict] = []

        try:
            zip_file = request.files.get("folder_zip")
            if zip_file and zip_file.filename:
                # zip 해제
                zip_bytes = zip_file.read()
                zip_path = tmp_dir / "upload.zip"
                zip_path.write_bytes(zip_bytes)
                with zipfile.ZipFile(zip_path, "r") as zf:
                    zf.extractall(tmp_dir / "extracted")
                work_dir = tmp_dir / "extracted"
            else:
                # 다중 파일 → 임시 폴더에 저장
                work_dir = tmp_dir / (folder_name_hint or "evidence")
                work_dir.mkdir(parents=True, exist_ok=True)
                for f in request.files.getlist("files[]"):
                    if f.filename:
                        dst = work_dir / f.filename
                        dst.write_bytes(f.read())

            # BC-{N} 하위 폴더 탐색
            bc_folders = [d for d in work_dir.iterdir() if d.is_dir() and d.name.startswith("BC-")]
            if not bc_folders:
                bc_folders = [work_dir]  # 단일 폴더

            for folder in bc_folders:
                bc_nums, party = parse_bc_folder_name(folder.name)
                if not party and folder_name_hint:
                    _, party = parse_bc_folder_name(folder_name_hint)
                if not party:
                    party = folder.name

                # Task 2: BC kind 자동 식별 — 채권·채무 양쪽 candidates 합집합
                candidates_list: list[str] = []
                candidates_kind_map: dict[str, str] = {}  # name → kind
                for chk_kind in ("receivable", "payable"):
                    chk_wp2 = wp_repo.get_or_create(pid, chk_kind)
                    if chk_wp2.sampling_result:
                        _sr2 = json.loads(chk_wp2.sampling_result)
                        for _d2 in _sr2.get("decisions", []):
                            if _d2.get("final_sampled"):
                                _name = _d2["name"]
                                if _name not in candidates_kind_map:
                                    candidates_kind_map[_name] = chk_kind
                                    candidates_list.append(_name)
                                else:
                                    candidates_kind_map[_name] = "both"
                ug_data = STATE.get("upload_guide_data")
                currency_resolver = CurrencyResolver(ug_data, STATE.get("manual_rates"))

                agg = aggregate_folder(
                    folder,
                    party_name=party,
                    final_sampled_candidates=candidates_list or None,
                    upload_guide_data=ug_data,
                    currency_resolver=currency_resolver,
                    ledger_balance_krw=_f(request.form.get("ledger_balance")),
                )

                # Task 2: BC kind 자동 식별 — 매칭된 거래처가 어느 kind인지 결정
                matched_bc_name = agg.matched_party_name or party
                matched_bc_kind = candidates_kind_map.get(matched_bc_name, kind)
                # "both"면 요청 kind를 fallback으로 사용 (양쪽 모두 등록은 아래서 처리)
                target_kinds_bc: list[str] = (
                    ["receivable", "payable"] if matched_bc_kind == "both"
                    else [matched_bc_kind]
                )

                ledger_balance = _f(request.form.get("ledger_balance"))

                for target_bc_kind in target_kinds_bc:
                    target_bc_wp = wp_repo.get_or_create(pid, target_bc_kind)

                    # target kind 장부 잔액 조회
                    target_bc_lb = ledger_balance
                    if target_bc_lb is None and matched_bc_name:
                        _sr_bc = target_bc_wp.sampling_result
                        if _sr_bc:
                            for _d_bc in json.loads(_sr_bc).get("decisions", []):
                                if _d_bc.get("final_sampled") and _d_bc["name"] == matched_bc_name:
                                    target_bc_lb = float(_d_bc.get("balance", 0))
                                    break

                    # 각 파일을 Artifact 저장
                    artifact_ids: list[str] = []
                    for ex in agg.extracts:
                        if ex.file_path.exists():
                            art = art_repo.save_file(
                                project_id=pid,
                                kind="evidence",
                                source_path=ex.file_path,
                                filename=ex.file_path.name,
                                workpaper_id=target_bc_wp.id,
                            )
                            artifact_ids.append(art.id)

                    # AlternativeProcedure 생성/갱신
                    coverage_ratio: float | None = None
                    if target_bc_lb and target_bc_lb > 0 and agg.total_amount:
                        coverage_ratio = min(1.0, agg.total_amount / target_bc_lb)

                    conclusion = _auto_conclusion(coverage_ratio)
                    reason = _infer_reason(s, target_bc_wp.id, matched_bc_name)

                    proc = proc_repo.get_by_party(target_bc_wp.id, matched_bc_name)
                    if proc is None:
                        proc = proc_repo.create(
                            workpaper_id=target_bc_wp.id,
                            party_name=matched_bc_name,
                            reason=reason,
                            ledger_balance=target_bc_lb,
                            procedure_type="auto_detected",
                            evidence_artifact_ids=artifact_ids,
                            covered_amount=agg.total_amount,
                            coverage_ratio=coverage_ratio,
                            conclusion=conclusion,
                            status="pending",
                        )
                    else:
                        proc_repo.update(
                            proc.id,
                            evidence_artifact_ids=artifact_ids,
                            covered_amount=agg.total_amount,
                            coverage_ratio=coverage_ratio,
                            conclusion=conclusion,
                        )

                    _audit(s, "step5_upload_folder", "AlternativeProcedure", proc.id, pid,
                           after={
                               "party_name": matched_bc_name,
                               "bc_numbers": agg.bc_numbers,
                               "total_amount": agg.total_amount,
                               "total_currency": agg.total_currency,
                               "success_count": agg.success_count,
                               "failed_count": agg.failed_count,
                               "target_kind": target_bc_kind,
                           })

                    results.append({
                        "procedure_id": proc.id,
                        "party_name": matched_bc_name,
                        "detected_kind": target_bc_kind,
                        "bc_numbers": agg.bc_numbers,
                        "total_files": agg.total_files,
                        "success_count": agg.success_count,
                        "failed_count": agg.failed_count,
                        "total_amount": agg.total_amount,
                        "total_currency": agg.total_currency,
                        "amounts_by_currency": agg.amounts_by_currency,
                        "conclusion": agg.conclusion,
                        # Week 2 확장
                        "matched_party_name": agg.matched_party_name,
                        "match_confidence": agg.match_confidence,
                        "match_candidates": agg.match_candidates,
                        "covered_amount_krw": agg.covered_amount_krw,
                        "coverage_ratio": agg.coverage_ratio,
                        "low_confidence_files": [str(f.name) for f in agg.low_confidence_files],
                    })

        finally:
            import shutil
            shutil.rmtree(tmp_dir, ignore_errors=True)

        return jsonify(results), 201


@app.route("/api/project/<pid>/step5/procedures", methods=["GET"])
def step5_list_procedures(pid: str):
    """AlternativeProcedure 일람."""
    kind = request.args.get("kind", "receivable")
    with get_session() as s:
        proj = ProjectRepository(s).get(pid)
        if proj is None:
            return jsonify({"error": "not found"}), 404
        wp_repo = WorkpaperRepository(s)
        wp = wp_repo.get_or_create(pid, kind)
        proc_repo = AlternativeProcedureRepository(s)
        procs = proc_repo.list_by_workpaper(wp.id)
        return jsonify([_serialize_procedure(p) for p in procs])


@app.route("/api/project/<pid>/step5/procedure/<proc_id>", methods=["PATCH"])
def step5_patch_procedure(pid: str, proc_id: str):
    """AlternativeProcedure 수동 보정.

    body: {procedure_type, covered_amount, conclusion, auditor_notes, status}
    """
    data = request.json or {}
    with get_session() as s:
        proj = ProjectRepository(s).get(pid)
        if proj is None:
            return jsonify({"error": "not found"}), 404

        proc_repo = AlternativeProcedureRepository(s)
        proc = proc_repo.update(proc_id, **{
            k: v for k, v in data.items()
            if k in ("procedure_type", "covered_amount", "conclusion",
                     "auditor_notes", "status", "reason",
                     "ledger_balance", "reply_balance")
        })
        if proc is None:
            return jsonify({"error": "procedure not found"}), 404

        # coverage_ratio 재계산
        if proc.ledger_balance and proc.ledger_balance > 0 and proc.covered_amount:
            proc.coverage_ratio = min(1.0, proc.covered_amount / proc.ledger_balance)

        _audit(s, "step5_manual_update", "AlternativeProcedure", proc_id, pid,
               after={k: v for k, v in data.items()})

        return jsonify(_serialize_procedure(proc))


@app.route("/api/project/<pid>/step5/parse-reconciliation", methods=["POST"])
def step5_parse_reconciliation(pid: str):
    """불일치 소명 xlsx 업로드 → 자동 인식.

    multipart: file=<xlsx>
    """
    import tempfile
    from src.infrastructure.evidence.reconciliation_parser import parse_reconciliation_xlsx

    with get_session() as s:
        proj = ProjectRepository(s).get(pid)
        if proj is None or proj.status == "archived":
            return jsonify({"error": "not found"}), 404

        f = request.files.get("file")
        if not f or not f.filename:
            return jsonify({"error": "파일 없음"}), 400

        kind = request.form.get("kind", "payable")
        wp_repo = WorkpaperRepository(s)
        wp = wp_repo.get_or_create(pid, kind)
        art_repo = ArtifactRepository(s)

        # Artifact 저장
        file_bytes = f.read()
        art = art_repo.save_bytes(
            project_id=pid,
            kind="reconciliation",
            content=file_bytes,
            filename=f.filename,
            workpaper_id=wp.id,
        )

        # 파싱
        try:
            sheets = parse_reconciliation_xlsx(Path(art.stored_path))
        except Exception as e:
            return jsonify({"error": f"파싱 실패: {e}"}), 400

        summary = sheets.summary_by_party()

        # Task 4: 불일치 소명 자동 통합
        # 1) 회신-불일치 시트 → ConfirmationReply.notes + status 갱신
        # 2) 거래처별 시트 → AlternativeProcedure 증빙/소명 자동 등록
        proc_repo = AlternativeProcedureRepository(s)
        reply_repo2 = ConfirmationReplyRepository(s)
        reconcile_created = 0
        reconcile_updated_reply = 0

        for party_name, v in summary.items():
            reasons = v.get("reasons", [])
            reason_str = "; ".join(reasons) if reasons else "불일치 소명"
            has_reason = bool(reasons)

            # ConfirmationReply 갱신 — 채권·채무 모두 탐색
            for chk_kind in ("receivable", "payable"):
                chk_wp3 = wp_repo.get_or_create(pid, chk_kind)
                replies3 = reply_repo2.list_by_workpaper(chk_wp3.id)
                for rep3 in replies3:
                    if rep3.party_name_matched == party_name:
                        # 사유가 명시되어 있으면 matched로 상향 가능
                        new_status = ("matched" if has_reason else rep3.status)
                        new_notes = f"[불일치소명] {reason_str}"
                        if rep3.notes:
                            new_notes = rep3.notes + " | " + new_notes
                        reply_repo2.update_reviewer_confirmation(
                            rep3.id,
                            reviewer_confirmed_status="reconciliation_applied",
                            status=new_status,
                            notes=new_notes,
                        )
                        reconcile_updated_reply += 1

            # AlternativeProcedure — 거래처별 시트 있으면 자동 생성/갱신
            if party_name in sheets.party_details:
                detail_rows = sheets.party_details[party_name]
                total_func = sum(
                    abs(r.amount_func) for r in detail_rows if r.amount_func is not None
                )
                for chk_kind in ("receivable", "payable"):
                    chk_wp4 = wp_repo.get_or_create(pid, chk_kind)
                    # 해당 kind의 sampled인지 확인
                    _is_sampled = False
                    if chk_wp4.sampling_result:
                        for _dd in json.loads(chk_wp4.sampling_result).get("decisions", []):
                            if _dd.get("final_sampled") and _dd["name"] == party_name:
                                _is_sampled = True
                                _lb = float(_dd.get("balance", 0))
                                break
                    if not _is_sampled:
                        continue

                    proc4 = proc_repo.get_by_party(chk_wp4.id, party_name)
                    if proc4 is None:
                        proc4 = proc_repo.create(
                            workpaper_id=chk_wp4.id,
                            party_name=party_name,
                            reason="차이",
                            ledger_balance=_lb if "_lb" in dir() else None,
                            procedure_type="불일치소명",
                            covered_amount=total_func if total_func else None,
                            conclusion="needs_review",
                            auditor_notes=reason_str,
                            status="pending",
                        )
                        reconcile_created += 1
                    else:
                        proc_repo.update(
                            proc4.id,
                            procedure_type="불일치소명",
                            covered_amount=total_func if total_func else proc4.covered_amount,
                            auditor_notes=reason_str,
                        )
                        reconcile_created += 1

        _audit(s, "step5_parse_reconciliation", "Artifact", art.id, pid,
               after={
                   "filename": f.filename,
                   "mismatch_rows": len(sheets.mismatch_rows),
                   "party_sheets": len(sheets.party_details),
                   "parties": list(summary.keys()),
                   "reconcile_created": reconcile_created,
                   "reconcile_updated_reply": reconcile_updated_reply,
               })

        return jsonify({
            "artifact_id": art.id,
            "mismatch_rows": len(sheets.mismatch_rows),
            "party_sheets": list(sheets.party_details.keys()),
            "summary_by_party": {
                party: {
                    "currency": v.get("currency"),
                    "sent_amount": v.get("sent_amount"),
                    "reply_amount": v.get("reply_amount"),
                    "difference": v.get("difference"),
                    "reasons": v.get("reasons", []),
                    "account_types": v.get("account_types", []),
                    "detail_rows": len(sheets.party_details.get(party, [])),
                }
                for party, v in summary.items()
            },
            "mismatch_detail": [
                {
                    "party_name": r.party_name,
                    "currency": r.currency,
                    "account_type": r.account_type,
                    "sent_amount": r.sent_amount,
                    "reply_amount": r.reply_amount,
                    "difference": r.difference,
                    "reason": r.reason,
                    "matched": r.matched,
                }
                for r in sheets.mismatch_rows
            ],
        })


@app.route("/api/project/<pid>/step5/match-bc-folder", methods=["POST"])
def step5_match_bc_folder(pid: str):
    """BC 폴더 거래처 매핑 확정 — 사용자가 선택한 party_name을 AlternativeProcedure에 저장.

    body: {procedure_id, party_name}
    - AlternativeProcedure.party_name 갱신
    - configs/party_aliases.yaml 의 pending_aliases 에 자동 저장
    """
    data = request.json or {}
    procedure_id = data.get("procedure_id", "").strip()
    confirmed_party = data.get("party_name", "").strip()

    if not procedure_id or not confirmed_party:
        return jsonify({"error": "procedure_id, party_name 필수"}), 400

    with get_session() as s:
        proj = ProjectRepository(s).get(pid)
        if proj is None or proj.status == "archived":
            return jsonify({"error": "not found"}), 404

        proc_repo = AlternativeProcedureRepository(s)
        proc = proc_repo.update(procedure_id, party_name=confirmed_party)
        if proc is None:
            return jsonify({"error": "procedure not found"}), 404

        # alias 사전 자동 저장 (pending_aliases 섹션)
        _save_pending_alias(proc.party_name, confirmed_party)

        _audit(s, "step5_match_bc_folder", "AlternativeProcedure", procedure_id, pid,
               after={"party_name": confirmed_party})

        return jsonify({
            "procedure_id": procedure_id,
            "party_name": confirmed_party,
            "alias_saved": True,
        })


def _save_pending_alias(raw_name: str, canonical_name: str) -> None:
    """configs/party_aliases.yaml 의 pending_aliases 섹션에 alias 추가.

    raw_name != canonical_name 인 경우에만 저장 (동일하면 불필요).
    """
    if not raw_name or not canonical_name or raw_name == canonical_name:
        return

    alias_path = ROOT / "configs" / "party_aliases.yaml"
    try:
        if alias_path.exists():
            with open(alias_path, encoding="utf-8") as f:
                config = yaml.safe_load(f) or {}
        else:
            config = {}

        if "pending_aliases" not in config or not isinstance(config["pending_aliases"], dict):
            config["pending_aliases"] = {}

        existing = config["pending_aliases"].get(canonical_name, [])
        if not isinstance(existing, list):
            existing = [str(existing)]
        if raw_name not in existing:
            existing.append(raw_name)
        config["pending_aliases"][canonical_name] = existing

        alias_path.parent.mkdir(parents=True, exist_ok=True)
        with open(alias_path, "w", encoding="utf-8") as f:
            yaml.dump(config, f, allow_unicode=True, default_flow_style=False)

        # 매칭 캐시 초기화 (다음 요청에서 신규 alias 반영)
        from src.domain.matching import reload_aliases
        reload_aliases()

    except Exception as e:
        log.warning("pending_alias 저장 실패 (비치명적): %s", e)


@app.route("/api/project/<pid>/step5/auto-identify-pending", methods=["POST"])
def step5_auto_identify_pending(pid: str):
    """미회신 거래처 자동 식별 + AlternativeProcedure placeholder 생성.

    채권·채무 final_sampled 중 ConfirmationReply가 matched/mismatch 없는 거래처를
    '미회신'으로 자동 등록. 이미 AlternativeProcedure 있으면 update.

    body: {} (파라미터 없음 — 채권·채무 양쪽 자동 처리)
    """
    with get_session() as s:
        proj = ProjectRepository(s).get(pid)
        if proj is None or proj.status == "archived":
            return jsonify({"error": "not found"}), 404

        wp_repo = WorkpaperRepository(s)
        reply_repo = ConfirmationReplyRepository(s)
        proc_repo = AlternativeProcedureRepository(s)

        created_count = 0
        updated_count = 0
        skipped_count = 0
        report: list[dict] = []

        for chk_kind in ("receivable", "payable"):
            chk_wp = wp_repo.get_or_create(pid, chk_kind)
            if not chk_wp.sampling_result:
                continue

            result_dict = json.loads(chk_wp.sampling_result)
            sampled_parties: dict[str, float] = {}
            for d in result_dict.get("decisions", []):
                if d.get("final_sampled"):
                    sampled_parties[d["name"]] = float(d.get("balance", 0))

            # 회신 현황 — matched/mismatch/needs_review 모두 회신 있음으로 처리
            # needs_review 는 추출 실패 등이지 회신 자체가 없는 것이 아님 → placeholder 생성 skip
            replies = reply_repo.list_by_workpaper(chk_wp.id)
            from src.domain.matching import _normalize as _norm
            # normalize 기반 replied_set:
            #   - party_name_matched (매칭 성공 회신)
            #   - party_name_raw (매칭 실패 회신 포함) — PDF는 실제로 존재하므로 placeholder 중복 방지
            replied_set: set[str] = set()
            replied_norms: set[str] = set()
            for r in replies:
                # 매칭 성공한 회신: matched/mismatch/needs_review 모두 회신 있음
                if r.party_name_matched and r.status in ("matched", "mismatch", "needs_review"):
                    replied_set.add(r.party_name_matched)
                    replied_norms.add(_norm(r.party_name_matched))
                # 매칭 실패(party_name_matched=None)이더라도 PDF 자체가 존재하면
                # party_name_raw normalize 가 sampled party normalize 와 일치하는지 확인
                if r.party_name_raw:
                    replied_norms.add(_norm(r.party_name_raw))

            # 이미 AlternativeProcedure 가 사용자 등록(evidence 있음)이면 skip
            existing_procs_with_evidence: set[str] = set()
            for proc in proc_repo.list_by_workpaper(chk_wp.id):
                if proc.evidence_artifact_ids:
                    try:
                        ev_ids_check = json.loads(proc.evidence_artifact_ids)
                        if ev_ids_check:
                            existing_procs_with_evidence.add(proc.party_name)
                    except Exception:
                        pass

            for party_name, ledger_balance in sampled_parties.items():
                # 원본 이름 매칭 또는 normalize 후 매칭
                if party_name in replied_set or _norm(party_name) in replied_norms:
                    skipped_count += 1
                    continue  # 이미 회신 처리됨

                # 사용자가 이미 증빙을 등록한 대체적 절차가 있으면 skip
                if party_name in existing_procs_with_evidence:
                    skipped_count += 1
                    continue

                existing_proc = proc_repo.get_by_party(chk_wp.id, party_name)
                if existing_proc is None:
                    proc_repo.create(
                        workpaper_id=chk_wp.id,
                        party_name=party_name,
                        reason="미회신",
                        ledger_balance=ledger_balance,
                        procedure_type="미정",
                        status="pending",
                        conclusion="needs_review",
                    )
                    created_count += 1
                    action = "created"
                else:
                    # 기존 레코드 reason만 보정 (evidence 유지)
                    if existing_proc.reason not in ("미회신", "차이"):
                        proc_repo.update(existing_proc.id, reason="미회신")
                    updated_count += 1
                    action = "updated"

                report.append({
                    "party_name": party_name,
                    "kind": chk_kind,
                    "ledger_balance": ledger_balance,
                    "action": action,
                })

        _audit(s, "step5_auto_identify_pending", "Project", pid, pid,
               after={
                   "created": created_count,
                   "updated": updated_count,
                   "skipped": skipped_count,
               })

        return jsonify({
            "created": created_count,
            "updated": updated_count,
            "skipped_already_replied": skipped_count,
            "pending_parties": report,
        }), 201


@app.route("/api/project/<pid>/step5/mark-done", methods=["POST"])
def step5_mark_done(pid: str):
    """Step 5 완료 기록."""
    data = request.json or {}
    kind = data.get("kind", "receivable")
    with get_session() as s:
        proj = ProjectRepository(s).get(pid)
        if proj is None:
            return jsonify({"error": "not found"}), 404
        wp_repo = WorkpaperRepository(s)
        wp = wp_repo.get_or_create(pid, kind)
        wp.step5_completed_at = datetime.now(timezone.utc)
        wp.updated_at = datetime.now(timezone.utc)
        _audit(s, "step5_mark_done", "Workpaper", wp.id, pid, after={"kind": kind})
        return jsonify({"ok": True, "step5_completed_at": wp.step5_completed_at.isoformat()})


# ─────────────────────────────────────────────────────────────
# 9. AuditTrail 조회
# ─────────────────────────────────────────────────────────────
@app.route("/api/project/<pid>/audit-trail")
def get_audit_trail(pid: str):
    """프로젝트 변경 이력 조회."""
    from sqlalchemy import select
    with get_session() as s:
        stmt = (
            select(AuditTrail)
            .where(AuditTrail.project_id == pid)
            .order_by(AuditTrail.timestamp.desc())
        )
        trails = list(s.execute(stmt).scalars())
        return jsonify([
            {
                "id": t.id,
                "timestamp": t.timestamp.isoformat(),
                "user_email": t.user_email,
                "action": t.action,
                "entity_type": t.entity_type,
                "entity_id": t.entity_id,
                "notes": t.notes,
                "after_value": t.after_value,
            }
            for t in trails
        ])


# ─────────────────────────────────────────────────────────────
# helpers
# ─────────────────────────────────────────────────────────────
def _audit(session, action: str, entity_type: str, entity_id: str | None,
           project_id: str | None, before=None, after=None, notes: str | None = None,
           user_email: str = "") -> None:
    """AuditTrail 레코드 추가 — 모든 중요 변경 지점에서 호출."""
    trail = AuditTrail(
        project_id=project_id,
        user_email=user_email,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        before_value=json.dumps(before, ensure_ascii=False) if before is not None else None,
        after_value=json.dumps(after, ensure_ascii=False) if after is not None else None,
        notes=notes,
    )
    session.add(trail)


def _serialize_project(proj) -> dict:
    return {
        "id": proj.id,
        "company_name": proj.company_name,
        "audit_firm": proj.audit_firm,
        "period_end": proj.period_end,
        "kind": proj.kind,
        "status": proj.status,
        "created_at": proj.created_at.isoformat() if proj.created_at else None,
        "updated_at": proj.updated_at.isoformat() if proj.updated_at else None,
        "created_by_email": proj.created_by_email,
    }


def _f(v, default=None):
    if v in (None, "", "null"):
        return default
    try:
        return float(v)
    except Exception:
        return default


def _i(v, default=None):
    if v in (None, "", "null"):
        return default
    try:
        return int(v)
    except Exception:
        return default


def _serialize_reply(reply, warnings: list | None = None) -> dict:
    return {
        "id": reply.id,
        "workpaper_id": reply.workpaper_id,
        "pdf_artifact_id": reply.pdf_artifact_id,
        "party_name_raw": reply.party_name_raw,
        "party_name_matched": reply.party_name_matched,
        "party_match_confidence": reply.party_match_confidence,
        "party_match_method": reply.party_match_method,
        "extracted_balance": reply.extracted_balance,
        "extracted_balance_currency": reply.extracted_balance_currency,
        "reply_date": reply.reply_date,
        "ledger_balance": reply.ledger_balance,
        "difference": reply.difference,
        "difference_pct": reply.difference_pct,
        "status": reply.status,
        "reviewer_confirmed_status": reply.reviewer_confirmed_status,
        "extraction_method": reply.extraction_method,
        "extraction_confidence": reply.extraction_confidence,
        "notes": reply.notes,
        "created_at": reply.created_at.isoformat() if reply.created_at else None,
        "warnings": warnings or [],
        # v2 확장 필드
        "declared_match": getattr(reply, "declared_match", None),
        "original_currency": getattr(reply, "original_currency", "KRW"),
        "decision_basis": getattr(reply, "decision_basis", None),
        "per_account_findings": (
            json.loads(reply.per_account_findings)
            if getattr(reply, "per_account_findings", None) else []
        ),
        "top3_candidates": (
            json.loads(reply.top3_candidates)
            if getattr(reply, "top3_candidates", None) else []
        ),
    }


def _serialize_procedure(proc) -> dict:
    import json as _json
    ids: list[str] = []
    if proc.evidence_artifact_ids:
        try:
            ids = _json.loads(proc.evidence_artifact_ids)
        except Exception:
            ids = []
    return {
        "id": proc.id,
        "workpaper_id": proc.workpaper_id,
        "party_name": proc.party_name,
        "reason": proc.reason,
        "ledger_balance": proc.ledger_balance,
        "reply_balance": proc.reply_balance,
        "difference": proc.difference,
        "procedure_type": proc.procedure_type,
        "evidence_artifact_ids": ids,
        "evidence_count": len(ids),
        "covered_amount": proc.covered_amount,
        "coverage_ratio": proc.coverage_ratio,
        "conclusion": proc.conclusion,
        "auditor_notes": proc.auditor_notes,
        "status": proc.status,
        "created_at": proc.created_at.isoformat() if proc.created_at else None,
        "updated_at": proc.updated_at.isoformat() if proc.updated_at else None,
    }


def _compute_covered_amount(extracts: list[dict]) -> float | None:
    """추출 결과 목록에서 KRW 금액 합산 (통화 혼재 시 KRW 우선)."""
    amounts_by_cur: dict[str, float] = {}
    for ex in extracts:
        amt = ex.get("extracted_amount")
        cur = ex.get("extracted_currency") or "KRW"
        if amt and amt > 0:
            amounts_by_cur[cur] = amounts_by_cur.get(cur, 0.0) + float(amt)
    if not amounts_by_cur:
        return None
    if "KRW" in amounts_by_cur:
        return amounts_by_cur["KRW"]
    return sum(amounts_by_cur.values())


def _auto_conclusion(coverage_ratio: float | None) -> str:
    """커버리지 비율 → 결론 자동 판정.

    감사실무 기준:
      ≥ 0.95 → 충분
      ≥ 0.50 → 부분
      < 0.50 → 미해소
      None   → needs_review
    """
    if coverage_ratio is None:
        return "needs_review"
    if coverage_ratio >= 0.95:
        return "충분"
    if coverage_ratio >= 0.50:
        return "부분"
    return "미해소"


def _infer_reason(session, workpaper_id: str, party_name: str) -> str:
    """ConfirmationReply 에서 거래처 reason 추론.

    매칭된 회신이 없으면 → "미회신"
    mismatch 있으면 → "차이"
    """
    from sqlalchemy import select
    from src.infrastructure.persistence.models import ConfirmationReply
    stmt = (
        select(ConfirmationReply)
        .where(ConfirmationReply.workpaper_id == workpaper_id)
        .where(ConfirmationReply.party_name_matched == party_name)
    )
    reply = session.execute(stmt).scalar_one_or_none()
    if reply is None:
        return "미회신"
    if reply.status == "mismatch":
        return "차이"
    return "기타"


def _serialize_result(result, params):
    size = result.size_result
    return {
        "population_amount": result.population_amount,
        "completeness": {
            "rows": [
                {"group": r["group"], "ledger": r["ledger"], "fs": r["fs"],
                 "diff": r["diff"], "note": r["note"]}
                for r in result.completeness.by_group
            ],
            "total_ledger": result.completeness.total_ledger,
            "total_fs": result.completeness.total_fs,
            "total_diff": result.completeness.total_diff,
        },
        "size": {
            "key_item_threshold": size.key_item_threshold,
            "key_item_ratio": size.key_item_ratio,
            "confidence_factor": size.confidence_factor,
            "base_sample_size": size.base_sample_size,
            "final_sample_size": size.final_sample_size,
            "sample_interval": size.sample_interval,
            "remaining_population": size.remaining_population,
        },
        "decisions": [
            {
                "name": d.name, "balance": d.balance,
                "is_excluded": d.is_excluded, "is_related_party": d.is_related_party,
                "is_key_item": d.is_key_item, "is_representative": d.is_representative,
                "final_sampled": d.final_sampled, "exclusion_reason": d.exclusion_reason,
            }
            for d in result.decisions
        ],
        "mus": {
            "sample_interval": result.mus_result.sample_interval,
            "random_start": result.mus_result.random_start,
            "selections": [
                {
                    "name": s.name, "balance": s.balance, "cumulative": s.cumulative,
                    "selections": s.selections, "remainder_after": s.remainder_after, "hit": s.hit,
                }
                for s in result.mus_result.selections
            ],
            "sampled_names": result.mus_result.sampled_names,
        },
    }


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=8520, debug=False, use_reloader=False)
