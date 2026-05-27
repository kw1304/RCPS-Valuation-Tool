"""불일치 소명 xlsx 파서 — Week 5.

BC-5,12_채권채무조회서_불일치 소명.xlsx 구조:
  시트 "회신-불일치": 거래처별 차이 요약 (1행 헤더)
    컬럼: 기업명, 조회차수, 사업자등록번호, 담당자명, 담당자이메일,
          일치여부, 통화, 구분, 발송금액, 회신금액, 차이금액, 최초발송일시, 회신일시, 사유

  거래처별 시트 (예: "COSMAX USA", "COSMAX California"):
    헤더 행 인식: 2행 (row 1=법인명, row 2=컬럼헤더)
    컬럼: NO, 매입법인코드, 매입법인명, 거래유형코드, 거래유형명,
          대사구분, 일치키, ..., 거래통화, 거래통화금액, 기능통화, 기능통화금액,
          ..., 적요, 거래일자, 회계일자, ...

반환: ReconciliationSheets
  .mismatch_rows: 회신-불일치 시트 데이터
  .party_details: {거래처명: [거래 단위 명세]}
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

log = logging.getLogger("cc_sampling.evidence.reconciliation_parser")

# 회신-불일치 시트명
_MISMATCH_SHEET = "회신-불일치"

# 거래처별 시트에서 제외할 시트명 (비 거래처 시트)
_SKIP_SHEETS = {_MISMATCH_SHEET}

# 회신-불일치 컬럼 인덱스 (헤더 기반 동적 탐색 — 하드코딩 방지)
_MISMATCH_COLS = {
    "party_name":     ["기업명"],
    "round":          ["조회차수"],
    "tax_id":         ["사업자등록번호"],
    "contact_name":   ["담당자명"],
    "contact_email":  ["담당자이메일"],
    "matched":        ["일치여부"],
    "currency":       ["통화"],
    "account_type":   ["구분"],
    "sent_amount":    ["발송금액"],
    "reply_amount":   ["회신금액"],
    "difference":     ["차이금액"],
    "sent_at":        ["최초발송일시"],
    "replied_at":     ["회신일시"],
    "reason":         ["사유"],
}

# 거래처별 시트 컬럼 인덱스
_DETAIL_COLS = {
    "entity_code":    ["매입법인코드"],
    "entity_name":    ["매입법인명"],
    "tx_type_code":   ["거래유형코드"],
    "tx_type_name":   ["거래유형명"],
    "reconcile_type": ["대사구분"],
    "match_key":      ["일치키"],
    "acct_code":      ["법인계정코드"],
    "acct_name":      ["법인계정명"],
    "currency":       ["거래통화"],
    "amount_fx":      ["거래통화금액"],
    "func_currency":  ["기능통화"],
    "amount_func":    ["기능통화금액"],
    "description":    ["적요"],
    "tx_date":        ["거래일자"],
    "acct_date":      ["회계일자"],
    "invoice_no":     ["Invoice no"],
}


@dataclass
class MismatchRow:
    """회신-불일치 시트 한 행."""
    party_name: str
    round: Optional[int]
    tax_id: Optional[str]
    contact_name: Optional[str]
    contact_email: Optional[str]
    matched: Optional[str]          # "Y" | "N" | ""
    currency: Optional[str]
    account_type: Optional[str]
    sent_amount: Optional[float]
    reply_amount: Optional[float]
    difference: Optional[float]
    sent_at: Optional[datetime]
    replied_at: Optional[datetime]
    reason: Optional[str]


@dataclass
class DetailRow:
    """거래처별 시트 한 거래 행."""
    entity_code: Optional[str]
    entity_name: Optional[str]
    tx_type_code: Optional[str]
    tx_type_name: Optional[str]
    reconcile_type: Optional[str]   # "수동일치" | "대사제외"
    match_key: Optional[str]
    acct_code: Optional[str]
    acct_name: Optional[str]
    currency: Optional[str]
    amount_fx: Optional[float]
    func_currency: Optional[str]
    amount_func: Optional[float]
    description: Optional[str]
    tx_date: Optional[str]
    acct_date: Optional[str]
    invoice_no: Optional[str]


@dataclass
class ReconciliationSheets:
    """불일치 소명 xlsx 전체 파싱 결과."""
    source_path: Path
    mismatch_rows: list[MismatchRow] = field(default_factory=list)
    party_details: dict[str, list[DetailRow]] = field(default_factory=dict)
    # {거래처명: [DetailRow, ...]}

    @property
    def mismatch_parties(self) -> list[str]:
        """불일치 거래처명 중복 제거 목록."""
        seen: dict[str, bool] = {}
        for row in self.mismatch_rows:
            if row.party_name:
                seen[row.party_name] = True
        return list(seen.keys())

    def summary_by_party(self) -> dict[str, dict]:
        """거래처명 → {통화, 발송금액, 회신금액, 차이금액, 사유} 요약."""
        result: dict[str, dict] = {}
        for row in self.mismatch_rows:
            key = row.party_name
            if not key:
                continue
            if key not in result:
                result[key] = {
                    "currency": row.currency,
                    "sent_amount": 0.0,
                    "reply_amount": 0.0,
                    "difference": 0.0,
                    "reasons": [],
                    "account_types": [],
                }
            entry = result[key]
            if row.sent_amount:
                entry["sent_amount"] = (entry["sent_amount"] or 0.0) + float(row.sent_amount)
            if row.reply_amount:
                entry["reply_amount"] = (entry["reply_amount"] or 0.0) + float(row.reply_amount)
            if row.reason and row.reason not in entry["reasons"]:
                entry["reasons"].append(row.reason)
            if row.account_type and row.account_type not in entry["account_types"]:
                entry["account_types"].append(row.account_type)
        # 차이 재계산
        for entry in result.values():
            entry["difference"] = (entry["sent_amount"] or 0.0) - (entry["reply_amount"] or 0.0)
        return result


# ─── 헬퍼 ───────────────────────────────────────────────────────────────────

def _build_col_map(headers: tuple, col_defs: dict[str, list[str]]) -> dict[str, int]:
    """헤더 행에서 필드명 → 열 인덱스 매핑 생성."""
    col_map: dict[str, int] = {}
    for field_key, candidates in col_defs.items():
        for i, h in enumerate(headers):
            h_str = str(h or "").strip()
            if h_str in candidates:
                col_map[field_key] = i
                break
    return col_map


def _cell(row: tuple, idx: Optional[int], default=None):
    if idx is None or idx >= len(row):
        return default
    v = row[idx]
    if v is None or str(v).strip() == "":
        return default
    return v


def _to_float(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return None


def _to_datetime(v) -> Optional[datetime]:
    if isinstance(v, datetime):
        return v
    return None


# ─── 공개 진입점 ────────────────────────────────────────────────────────────

def parse_reconciliation_xlsx(path: Path) -> ReconciliationSheets:
    """불일치 소명 xlsx 파일을 파싱해 ReconciliationSheets 반환.

    지원:
      - "회신-불일치" 시트 → MismatchRow 목록
      - 거래처별 시트 (non-"회신-불일치") → DetailRow 목록

    Raises:
        FileNotFoundError: 파일 없음
        ValueError: 지원 포맷이 아닌 xlsx
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(path)

    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise ImportError("openpyxl 미설치: pip install openpyxl")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    result = ReconciliationSheets(source_path=path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        if sheet_name == _MISMATCH_SHEET:
            _parse_mismatch_sheet(rows, result)
        elif sheet_name not in _SKIP_SHEETS:
            _parse_detail_sheet(sheet_name, rows, result)

    wb.close()
    log.info(
        "불일치 소명 파싱 완료: mismatch=%d rows, parties=%d",
        len(result.mismatch_rows), len(result.party_details),
    )
    return result


def _parse_mismatch_sheet(rows: list[tuple], result: ReconciliationSheets) -> None:
    """회신-불일치 시트 파싱."""
    if len(rows) < 2:
        return

    # 1행: 헤더
    header = rows[0]
    col_map = _build_col_map(header, _MISMATCH_COLS)

    for row in rows[1:]:
        if not any(v is not None and str(v).strip() for v in row):
            continue  # 빈 행 skip

        party = str(_cell(row, col_map.get("party_name"), "") or "").strip()
        if not party:
            continue

        result.mismatch_rows.append(MismatchRow(
            party_name=party,
            round=int(_cell(row, col_map.get("round"), 0) or 0) or None,
            tax_id=str(_cell(row, col_map.get("tax_id"), "") or "").strip() or None,
            contact_name=str(_cell(row, col_map.get("contact_name"), "") or "").strip() or None,
            contact_email=str(_cell(row, col_map.get("contact_email"), "") or "").strip() or None,
            matched=str(_cell(row, col_map.get("matched"), "") or "").strip() or None,
            currency=str(_cell(row, col_map.get("currency"), "") or "").strip() or None,
            account_type=str(_cell(row, col_map.get("account_type"), "") or "").strip() or None,
            sent_amount=_to_float(_cell(row, col_map.get("sent_amount"))),
            reply_amount=_to_float(_cell(row, col_map.get("reply_amount"))),
            difference=_to_float(_cell(row, col_map.get("difference"))),
            sent_at=_to_datetime(_cell(row, col_map.get("sent_at"))),
            replied_at=_to_datetime(_cell(row, col_map.get("replied_at"))),
            reason=str(_cell(row, col_map.get("reason"), "") or "").strip() or None,
        ))


def _parse_detail_sheet(sheet_name: str, rows: list[tuple], result: ReconciliationSheets) -> None:
    """거래처별 시트 파싱.

    BC-5,12 패턴:
      - row 0: 법인명 헤더 (예: ("매입법인", None, ...))
      - row 1: 컬럼 헤더 (예: ("NO", "매입법인코드", ...))
      - row 2+: 데이터
    """
    # 헤더 행 탐색 (최대 5행 내에서 "NO" 또는 "매입법인코드" 찾기)
    header_row_idx: Optional[int] = None
    for i, row in enumerate(rows[:5]):
        row_text = " ".join(str(v or "") for v in row)
        if "매입법인코드" in row_text or ("NO" in row_text and "거래유형" in row_text):
            header_row_idx = i
            break

    if header_row_idx is None:
        log.debug("거래처 시트 '%s' 헤더 미인식 — skip", sheet_name)
        return

    header = rows[header_row_idx]
    col_map = _build_col_map(header, _DETAIL_COLS)

    details: list[DetailRow] = []
    for row in rows[header_row_idx + 1:]:
        if not any(v is not None and str(v).strip() for v in row):
            continue

        amt_fx = _to_float(_cell(row, col_map.get("amount_fx")))
        amt_func = _to_float(_cell(row, col_map.get("amount_func")))

        # 금액도 없고 대사구분도 없으면 데이터 행 아님
        if amt_fx is None and amt_func is None:
            continue

        # 날짜 처리 (datetime 또는 문자열)
        def _to_date_str(v) -> Optional[str]:
            if v is None:
                return None
            if isinstance(v, datetime):
                return v.strftime("%Y-%m-%d")
            return str(v).strip() or None

        details.append(DetailRow(
            entity_code=str(_cell(row, col_map.get("entity_code"), "") or "").strip() or None,
            entity_name=str(_cell(row, col_map.get("entity_name"), "") or "").strip() or None,
            tx_type_code=str(_cell(row, col_map.get("tx_type_code"), "") or "").strip() or None,
            tx_type_name=str(_cell(row, col_map.get("tx_type_name"), "") or "").strip() or None,
            reconcile_type=str(_cell(row, col_map.get("reconcile_type"), "") or "").strip() or None,
            match_key=str(_cell(row, col_map.get("match_key"), "") or "").strip() or None,
            acct_code=str(_cell(row, col_map.get("acct_code"), "") or "").strip() or None,
            acct_name=str(_cell(row, col_map.get("acct_name"), "") or "").strip() or None,
            currency=str(_cell(row, col_map.get("currency"), "") or "").strip() or None,
            amount_fx=amt_fx,
            func_currency=str(_cell(row, col_map.get("func_currency"), "") or "").strip() or None,
            amount_func=amt_func,
            description=str(_cell(row, col_map.get("description"), "") or "").strip() or None,
            tx_date=_to_date_str(_cell(row, col_map.get("tx_date"))),
            acct_date=_to_date_str(_cell(row, col_map.get("acct_date"))),
            invoice_no=str(_cell(row, col_map.get("invoice_no"), "") or "").strip() or None,
        ))

    if details:
        result.party_details[sheet_name] = details
        log.debug("거래처 시트 '%s': %d 행 파싱", sheet_name, len(details))
