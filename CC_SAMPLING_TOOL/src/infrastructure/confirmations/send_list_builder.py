"""발송명단 Excel 생성기 — Step 2 핵심 산출물.

시트 구성 (감사실무 표준):
  1. 발송명단 — 거래처별 기말잔액·연락처 (회사 담당자 작성용)
  2. 송부 안내문 — 회계법인이 회사에 전달하는 안내문
  3. 거래처별 회신서 양식 — 표준 채권채무조회서 (거래처 수만큼 페이지 추가)

설계 원칙:
  - 감사실무 양식 톤: 한글 폰트, 테두리, 적절한 셀 너비
  - party_contacts 없으면 빈칸 처리 (회사 담당자가 직접 채움)
  - 거래처 수에 따라 3번 시트 페이지 자동 확장
"""
from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from src.domain.population import PartyDecision


# ── 스타일 상수 ─────────────────────────────────────────────
_FONT_BODY = Font(name="맑은 고딕", size=10)
_FONT_HEADER = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
_FONT_TITLE = Font(name="맑은 고딕", size=13, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=11, bold=True)
_FONT_SMALL = Font(name="맑은 고딕", size=9)
_FONT_SMALL_BOLD = Font(name="맑은 고딕", size=9, bold=True)

_FILL_HEADER = PatternFill("solid", fgColor="1F4E78")
_FILL_SUBHEADER = PatternFill("solid", fgColor="D6E4F0")
_FILL_LIGHT = PatternFill("solid", fgColor="F2F8FC")
_FILL_CONFIRM_HEADER = PatternFill("solid", fgColor="2E5F8A")
_FILL_SECTION = PatternFill("solid", fgColor="EBF3FA")

_SIDE_THIN = Side(style="thin", color="B0B8C1")
_SIDE_MEDIUM = Side(style="medium", color="4A90D9")
_BORDER_THIN = Border(left=_SIDE_THIN, right=_SIDE_THIN, top=_SIDE_THIN, bottom=_SIDE_THIN)
_BORDER_MEDIUM = Border(
    left=Side(style="medium", color="1F4E78"),
    right=Side(style="medium", color="1F4E78"),
    top=Side(style="medium", color="1F4E78"),
    bottom=Side(style="medium", color="1F4E78"),
)
_BORDER_HEADER_BOTTOM = Border(
    left=_SIDE_THIN, right=_SIDE_THIN,
    top=_SIDE_THIN,
    bottom=Side(style="medium", color="1F4E78"),
)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
_ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _fmt_amount(v: float) -> str:
    if v == 0:
        return "-"
    return f"{v:,.0f}"


def _set_cell(
    ws,
    row: int,
    col: int,
    value,
    font=None,
    fill=None,
    border=None,
    alignment=None,
    number_format: str | None = None,
):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment
    if number_format:
        cell.number_format = number_format
    return cell


def _merge_and_set(ws, r1, c1, r2, c2, value, font=None, fill=None, border=None, alignment=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    # 병합 셀 외곽 테두리
    if border:
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(r, c).border = border
    return cell


def build_send_list(
    out_path: Path,
    project_info: dict,
    decisions: list[PartyDecision],
    kind: str,
    reply_deadline: Optional[date] = None,
    contact_info: Optional[dict] = None,
    party_contacts: Optional[dict[str, dict]] = None,
) -> None:
    """발송명단 Excel 생성.

    Args:
        out_path: 저장 경로
        project_info: company_name, period_end, audit_firm, preparer 등
        decisions: Step 1 PartyDecision 목록 (final_sampled=True 인 것만 발송)
        kind: "receivable" | "payable"
        reply_deadline: 조회서 회신 기한
        contact_info: {"firm": ..., "email": ..., "phone": ..., "address": ...}
        party_contacts: {거래처명: {"email": ..., "phone": ..., "address": ...}}
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    final_parties = [d for d in decisions if d.final_sampled]
    final_parties.sort(key=lambda d: -d.balance)

    ci = contact_info or {}
    pc = party_contacts or {}
    pi = project_info or {}

    company_name = pi.get("company_name", "")
    period_end_str = pi.get("period_end", "")
    audit_firm = pi.get("audit_firm", ci.get("firm", ""))
    preparer = pi.get("preparer", "")
    kind_label = "채권" if kind == "receivable" else "채무"

    # 1. 발송명단 시트
    ws1 = wb.create_sheet("발송명단")
    _build_send_list_sheet(
        ws1, final_parties, company_name, period_end_str,
        audit_firm, kind_label, pc,
    )

    # 2. 송부 안내문 시트
    ws2 = wb.create_sheet("송부 안내문")
    _build_cover_letter_sheet(
        ws2, company_name, period_end_str, audit_firm,
        preparer, reply_deadline, ci, kind_label,
    )

    # 3. 거래처별 회신서 양식 시트
    ws3 = wb.create_sheet("거래처별 회신서 양식")
    _build_confirmation_forms_sheet(
        ws3, final_parties, company_name, period_end_str,
        audit_firm, kind_label, reply_deadline, ci,
    )

    wb.save(out_path)


def _build_send_list_sheet(
    ws,
    parties: list[PartyDecision],
    company_name: str,
    period_end: str,
    audit_firm: str,
    kind_label: str,
    party_contacts: dict,
) -> None:
    """Sheet 1: 발송명단."""

    # 컬럼 너비
    col_widths = [6, 28, 14, 10, 18, 8, 26, 18, 18]
    headers = ["No", "거래처명", "거래처코드", "계정과목", "기말잔액(원)", "통화", "거래처 이메일", "거래처 연락처", "비고"]

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 14
    ws.row_dimensions[5].height = 26

    # 제목 행
    _merge_and_set(
        ws, 1, 1, 1, 9,
        f"채권채무조회서 발송명단 ({kind_label})",
        font=_FONT_TITLE,
        alignment=_ALIGN_CENTER,
    )

    # 부제목
    _set_cell(ws, 2, 1, f"회사명: {company_name}", font=_FONT_SMALL_BOLD, alignment=_ALIGN_LEFT)
    _set_cell(ws, 2, 4, f"기준일: {period_end}", font=_FONT_SMALL_BOLD, alignment=_ALIGN_LEFT)
    _set_cell(ws, 2, 7, f"감사인: {audit_firm}", font=_FONT_SMALL_BOLD, alignment=_ALIGN_LEFT)
    _merge_and_set(ws, 2, 1, 2, 3, f"회사명: {company_name}", font=_FONT_SMALL_BOLD, alignment=_ALIGN_LEFT)
    _merge_and_set(ws, 2, 4, 2, 6, f"기준일: {period_end}", font=_FONT_SMALL_BOLD, alignment=_ALIGN_LEFT)
    _merge_and_set(ws, 2, 7, 2, 9, f"감사인: {audit_firm}", font=_FONT_SMALL_BOLD, alignment=_ALIGN_LEFT)

    # 안내 메모
    _merge_and_set(
        ws, 3, 1, 3, 9,
        "※ 거래처 이메일·연락처 미기재 시 회사 담당자가 직접 입력하여 반환",
        font=Font(name="맑은 고딕", size=8, color="FF0000"),
        alignment=_ALIGN_LEFT,
    )

    ws.row_dimensions[4].height = 4  # 빈 줄 간격

    # 헤더 행 (5행)
    for c, h in enumerate(headers, 1):
        _set_cell(
            ws, 5, c, h,
            font=_FONT_HEADER,
            fill=_FILL_HEADER,
            border=_BORDER_THIN,
            alignment=_ALIGN_CENTER,
        )
    ws.row_dimensions[5].height = 22

    # 데이터 행
    for i, party in enumerate(parties):
        row = 6 + i
        ws.row_dimensions[row].height = 18

        pc_info = party_contacts.get(party.name, {})
        account_label = "채권" if party.is_key_item else "채무/채권"
        fill = _FILL_LIGHT if i % 2 == 0 else None

        values = [
            i + 1,
            party.name,
            "",            # 거래처코드 (원장에 없으면 공란)
            kind_label,
            party.balance,
            "KRW",
            pc_info.get("email", ""),
            pc_info.get("phone", ""),
            "",            # 비고
        ]
        for c, v in enumerate(values, 1):
            cell = _set_cell(
                ws, row, c, v,
                font=_FONT_BODY,
                fill=fill,
                border=_BORDER_THIN,
                alignment=_ALIGN_RIGHT if c == 5 else _ALIGN_CENTER if c in (1, 6) else _ALIGN_LEFT,
            )
            if c == 5:
                cell.number_format = '#,##0'

    # 합계 행
    sum_row = 6 + len(parties)
    ws.row_dimensions[sum_row].height = 20
    _merge_and_set(ws, sum_row, 1, sum_row, 4, "합계", font=_FONT_SMALL_BOLD, fill=_FILL_SUBHEADER, alignment=_ALIGN_CENTER, border=_BORDER_THIN)
    total = sum(p.balance for p in parties)
    _set_cell(ws, sum_row, 5, total, font=_FONT_SMALL_BOLD, fill=_FILL_SUBHEADER, border=_BORDER_THIN, alignment=_ALIGN_RIGHT, number_format='#,##0')
    for c in range(6, 10):
        _set_cell(ws, sum_row, c, "", font=_FONT_SMALL_BOLD, fill=_FILL_SUBHEADER, border=_BORDER_THIN, alignment=_ALIGN_CENTER)

    # 인쇄 설정
    ws.print_area = f"A1:I{sum_row}"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.freeze_panes = "A6"


def _build_cover_letter_sheet(
    ws,
    company_name: str,
    period_end: str,
    audit_firm: str,
    preparer: str,
    reply_deadline: date | None,
    contact_info: dict,
    kind_label: str,
) -> None:
    """Sheet 2: 송부 안내문."""
    for i, w in enumerate([4, 20, 14, 14, 14, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    deadline_str = reply_deadline.strftime("%Y년 %m월 %d일") if reply_deadline else "__년 __월 __일"
    today_str = date.today().strftime("%Y년 %m월 %d일")
    firm_email = contact_info.get("email", "")
    firm_phone = contact_info.get("phone", "")
    firm_address = contact_info.get("address", "")

    row = 1

    def add_row(height=18):
        nonlocal row
        ws.row_dimensions[row].height = height
        row += 1

    # 여백
    add_row(12)

    # 수신자 (회사 담당자)
    _merge_and_set(ws, row, 2, row, 8, f"수신:  {company_name}  재무담당자 귀중", font=Font(name="맑은 고딕", size=11, bold=True), alignment=_ALIGN_LEFT)
    add_row(24)

    # 제목
    _merge_and_set(ws, row, 2, row, 8, f"채권채무 조회서 발송 요청 ({kind_label} 기준 / 기준일: {period_end})", font=_FONT_TITLE, alignment=_ALIGN_CENTER)
    add_row(36)

    add_row(8)

    # 본문
    body_lines = [
        f"귀 회사의 {period_end} 기준 {kind_label} 잔액에 대한 감사절차 수행의 일환으로, 외부 채권채무 조회를 실시합니다.",
        "",
        "아래 발송명단 시트에 기재된 거래처 담당자에게 첨부된 조회서 양식을 전달하시고,",
        f"회신 기한 내 (기한: {deadline_str}) 당 법인으로 직접 회신되도록 안내해 주시기 바랍니다.",
        "",
        "※ 감사인에게 직접 회신 요청 (귀 회사를 통한 회신 불가)",
        "※ 거래처 이메일 및 연락처를 발송명단에 기재하여 반환해 주십시오.",
    ]

    for line in body_lines:
        if line:
            _merge_and_set(ws, row, 2, row, 8, line, font=_FONT_BODY, alignment=_ALIGN_LEFT)
            add_row(18)
        else:
            add_row(10)

    add_row(12)

    # 감사인 정보 박스
    info_start = row
    _merge_and_set(ws, row, 2, row, 8, "감사인 연락처", font=_FONT_SUBTITLE, fill=_FILL_SECTION, alignment=_ALIGN_CENTER, border=_BORDER_THIN)
    add_row(22)

    contact_rows = [
        ("감사인(회계법인)", audit_firm),
        ("담당 감사인", preparer),
        ("이메일", firm_email),
        ("전화번호", firm_phone),
        ("주소", firm_address),
    ]
    for label, value in contact_rows:
        _set_cell(ws, row, 2, label, font=_FONT_SMALL_BOLD, fill=_FILL_SUBHEADER, border=_BORDER_THIN, alignment=_ALIGN_CENTER)
        _merge_and_set(ws, row, 3, row, 8, value, font=_FONT_BODY, border=_BORDER_THIN, alignment=_ALIGN_LEFT)
        add_row(18)

    add_row(20)

    # 발신 일자 및 서명
    _merge_and_set(ws, row, 2, row, 8, today_str, font=_FONT_BODY, alignment=_ALIGN_RIGHT)
    add_row(18)
    _merge_and_set(ws, row, 2, row, 8, f"{audit_firm}  담당 감사인:  {preparer}  (서명/날인)", font=Font(name="맑은 고딕", size=11, bold=True), alignment=_ALIGN_RIGHT)
    add_row(24)


def _build_confirmation_forms_sheet(
    ws,
    parties: list[PartyDecision],
    company_name: str,
    period_end: str,
    audit_firm: str,
    kind_label: str,
    reply_deadline: date | None,
    contact_info: dict,
) -> None:
    """Sheet 3: 거래처별 회신서 양식 (거래처 수만큼 페이지).

    각 거래처 = 40행 (1페이지) 단위 배치.
    """
    for i, w in enumerate([4, 14, 18, 18, 18, 18, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    PAGE_HEIGHT = 42   # 거래처 1건당 행 수 (인쇄 1페이지)
    deadline_str = reply_deadline.strftime("%Y년 %m월 %d일") if reply_deadline else "__년 __월 __일"
    firm_info = f"{audit_firm}  /  {contact_info.get('email', '')}  /  {contact_info.get('phone', '')}"

    for idx, party in enumerate(parties):
        base = idx * PAGE_HEIGHT + 1   # 이 거래처의 시작 행

        for r in range(base, base + PAGE_HEIGHT):
            ws.row_dimensions[r].height = 16

        _write_confirmation_form(
            ws, base,
            party_name=party.name,
            company_name=company_name,
            period_end=period_end,
            balance=party.balance,
            kind_label=kind_label,
            deadline_str=deadline_str,
            firm_info=firm_info,
        )

        # 인쇄 페이지 구분선 (마지막 거래처 제외)
        if idx < len(parties) - 1:
            ws.row_dimensions[base + PAGE_HEIGHT - 1].height = 2


def _write_confirmation_form(
    ws,
    base_row: int,
    party_name: str,
    company_name: str,
    period_end: str,
    balance: float,
    kind_label: str,
    deadline_str: str,
    firm_info: str,
) -> None:
    """단일 거래처 회신서 양식 작성 (1페이지 분량)."""
    r = base_row

    # 1: 양식 제목
    _merge_and_set(
        ws, r, 2, r, 7,
        f"채권채무조회서 ({kind_label}) — {party_name}",
        font=_FONT_TITLE,
        fill=PatternFill("solid", fgColor="1F4E78"),
        alignment=_ALIGN_CENTER,
        border=_BORDER_THIN,
    )
    ws.row_dimensions[r].height = 26
    r += 1

    # 2: 감사인 정보
    _merge_and_set(ws, r, 2, r, 7, f"감사인: {firm_info}", font=_FONT_SMALL, alignment=_ALIGN_LEFT)
    ws.row_dimensions[r].height = 14
    r += 1

    # 3: 빈행
    ws.row_dimensions[r].height = 8
    r += 1

    # 4-5: 기본정보 섹션
    _merge_and_set(ws, r, 2, r, 7, "[ 기본 정보 ]", font=_FONT_SMALL_BOLD, fill=_FILL_SECTION, alignment=_ALIGN_CENTER, border=_BORDER_THIN)
    ws.row_dimensions[r].height = 18
    r += 1

    info_rows = [
        ("회사명 (요청자)", company_name),
        ("거래처명 (수신자)", party_name),
        ("기준일", period_end),
        ("구분", f"{kind_label}"),
    ]
    for label, value in info_rows:
        _set_cell(ws, r, 2, label, font=_FONT_SMALL_BOLD, fill=_FILL_SUBHEADER, border=_BORDER_THIN, alignment=_ALIGN_CENTER)
        _set_cell(ws, r, 3, "", fill=_FILL_SUBHEADER, border=_BORDER_THIN)
        _merge_and_set(ws, r, 3, r, 7, value, font=_FONT_BODY, border=_BORDER_THIN, alignment=_ALIGN_LEFT)
        ws.row_dimensions[r].height = 18
        r += 1

    # 빈행
    ws.row_dimensions[r].height = 8
    r += 1

    # 잔액 확인 섹션
    _merge_and_set(ws, r, 2, r, 7, "[ 잔액 확인 ]", font=_FONT_SMALL_BOLD, fill=_FILL_SECTION, alignment=_ALIGN_CENTER, border=_BORDER_THIN)
    ws.row_dimensions[r].height = 18
    r += 1

    # 헤더
    balance_headers = ["항목", "회사 제시 잔액 (원)", "귀사 확인 잔액 (원)", "차이 금액 (원)", "확인 (O/X)", "비고"]
    for c, h in enumerate(balance_headers, 2):
        _set_cell(ws, r, c, h, font=_FONT_SMALL_BOLD, fill=_FILL_HEADER, border=_BORDER_THIN, alignment=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 20
    r += 1

    # 잔액 행
    _set_cell(ws, r, 2, f"{kind_label} 잔액", font=_FONT_BODY, border=_BORDER_THIN, alignment=_ALIGN_CENTER)
    cell = _set_cell(ws, r, 3, balance, font=_FONT_BODY, border=_BORDER_THIN, alignment=_ALIGN_RIGHT, number_format='#,##0')
    for c in range(4, 8):
        _set_cell(ws, r, c, "", font=_FONT_BODY, border=_BORDER_THIN, alignment=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 20
    r += 1

    # 빈 확인 행 2줄 (기타 항목용)
    for _ in range(2):
        for c in range(2, 8):
            _set_cell(ws, r, c, "", font=_FONT_BODY, border=_BORDER_THIN, alignment=_ALIGN_CENTER)
        ws.row_dimensions[r].height = 18
        r += 1

    # 빈행
    ws.row_dimensions[r].height = 8
    r += 1

    # 회신 안내
    notice_lines = [
        f"※ 위 잔액은 {period_end} 기준으로 {company_name}의 장부에 기록된 금액입니다.",
        f"※ 회신 기한: {deadline_str}",
        "※ 본 조회서는 감사인에게 직접 회신하여 주십시오. (회사를 통한 회신 불가)",
        "※ 차이가 있는 경우 차이 금액 및 사유를 구체적으로 기재해 주십시오.",
    ]
    for line in notice_lines:
        _merge_and_set(ws, r, 2, r, 7, line, font=Font(name="맑은 고딕", size=8, color="666666"), alignment=_ALIGN_LEFT)
        ws.row_dimensions[r].height = 14
        r += 1

    # 빈행
    ws.row_dimensions[r].height = 8
    r += 1

    # 서명 섹션
    _merge_and_set(ws, r, 2, r, 7, "[ 회신자 서명 ]", font=_FONT_SMALL_BOLD, fill=_FILL_SECTION, alignment=_ALIGN_CENTER, border=_BORDER_THIN)
    ws.row_dimensions[r].height = 18
    r += 1

    sign_rows = [
        ("회사명", ""),
        ("담당자 성명", ""),
        ("직위", ""),
        ("서명/날인", ""),
        ("회신일자", ""),
    ]
    for label, value in sign_rows:
        _set_cell(ws, r, 2, label, font=_FONT_SMALL_BOLD, fill=_FILL_SUBHEADER, border=_BORDER_THIN, alignment=_ALIGN_CENTER)
        _merge_and_set(ws, r, 3, r, 7, value, font=_FONT_BODY, border=_BORDER_THIN, alignment=_ALIGN_LEFT)
        ws.row_dimensions[r].height = 18
        r += 1

    # 하단 여백
    while r < base_row + 42:
        ws.row_dimensions[r].height = 10
        r += 1
