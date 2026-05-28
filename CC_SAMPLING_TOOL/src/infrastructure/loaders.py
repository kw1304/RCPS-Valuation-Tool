"""Excel 로더 — 회사 제시 거래처별 원장, 재무제표, 특관자리스트, UploadGuide

loaders.py 는 파일 I/O 담당. 컬럼 감지 로직은 schemas/ 패키지에 위임.
기존 호출처(시트명 직접 지정)는 완전 호환 유지.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional
import pandas as pd
import openpyxl

from .schemas.ledger_schema import (
    detect_ledger_sheets,
    detect_ledger_columns,
    detect_multi_account_sheets,
    is_multi_sheet_ledger,
)
from .schemas.fs_schema import detect_fs_sheet
from .schemas.rp_schema import detect_rp_sheet


# ─────────────────────────────────────────────────────────────
# UploadGuide 데이터 모델
# ─────────────────────────────────────────────────────────────

@dataclass
class PartyContact:
    """거래처 연락처 (UploadGuide Sheet1 행 1건)."""
    name: str
    country: str = ""
    business_no: str = ""
    ceo_name: str = ""
    contact_person: str = ""
    phone: str = ""
    email: str = ""
    # 계정과목별 잔액: [(계정과목명, 통화, 금액), ...]
    accounts: list[tuple[str, str, float]] = field(default_factory=list)


@dataclass
class ExcludedParty:
    """발송제외 거래처 (UploadGuide '채채대상이나 발송대상제외' 시트 행 1건)."""
    name: str
    account_name: str = ""
    currency: str = ""
    amount: float = 0.0
    kind: str = ""          # 채권채무구분 (채권·채무)


@dataclass
class UploadGuideData:
    """회사 제시 UploadGuide 전체 파싱 결과."""
    send_targets: list[PartyContact] = field(default_factory=list)
    excluded: list[ExcludedParty] = field(default_factory=list)

    def contact_map(self) -> dict[str, PartyContact]:
        """거래처명 → PartyContact 빠른 조회용 dict."""
        return {p.name: p for p in self.send_targets}

    def excluded_names(self) -> set[str]:
        return {e.name for e in self.excluded}


# ─────────────────────────────────────────────────────────────
# UploadGuide 컬럼 키워드 감지 (sheet1)
# ─────────────────────────────────────────────────────────────

_UG_COL_KEYWORDS: dict[str, list[str]] = {
    "kind":           ["채권채무구분", "구분"],
    "account":        ["계정과목명", "계정"],
    "currency1":      ["통화1", "통화"],
    "amount1":        ["조회금액1", "금액1", "조회금액"],
    "currency2":      ["통화2"],
    "amount2":        ["조회금액2", "금액2"],
    "party_name":     ["거래처명", "거래처"],
    "country":        ["국가"],
    "party_type":     ["거래처 구분", "거래처구분"],
    "business_no":    ["사업자번호"],
    "ceo_name":       ["대표자명", "대표자"],
    "contact_person": ["담당자명", "담당자"],
    "phone":          ["담당자 유선전화번호", "전화번호", "전화"],
    "email":          ["담당자 이메일", "이메일", "email"],
}


def _detect_ug_columns(headers: list) -> dict[str, int]:
    """UploadGuide Sheet1 헤더 행 → {필드명: 컬럼인덱스(0-based)} 감지."""
    result: dict[str, int] = {}
    for i, h in enumerate(headers):
        h_str = str(h).strip() if h else ""
        for field_name, keywords in _UG_COL_KEYWORDS.items():
            if field_name in result:
                continue
            for kw in keywords:
                if kw in h_str:
                    result[field_name] = i
                    break
    return result


def load_upload_guide(path: str | Path) -> UploadGuideData:
    """회사 제시 UploadGuide xlsx → UploadGuideData.

    Sheet1 (발송명단):
      채권채무구분 / 계정과목명 / 통화1 / 조회금액1 / 통화2 / 조회금액2
      / 거래처명 / 국가 / 거래처 구분 / 사업자번호 / 대표자명
      / 담당자명 / 담당자 유선전화번호 / 담당자 이메일 / 필수항목 기재여부

    '채채대상이나 발송대상제외' 시트 (발송제외):
      채권채무구분 / 계정과목명 / 통화 / 조회금액 / 거래처명
    """
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    wb.close()

    # ── Sheet1 파싱 ──────────────────────────────────────────
    send_targets: list[PartyContact] = []
    sheet1_name = sheets[0] if sheets else None

    if sheet1_name:
        df1 = pd.read_excel(path, sheet_name=sheet1_name, header=0, dtype=str)
        headers = list(df1.columns)
        col = _detect_ug_columns(headers)

        def _get(row, key: str, default: str = "") -> str:
            idx = col.get(key)
            if idx is None or idx >= len(row):
                return default
            v = row.iloc[idx]
            return str(v).strip() if pd.notna(v) and str(v).strip() not in ("nan", "None", "") else default

        def _get_float(row, key: str) -> float:
            idx = col.get(key)
            if idx is None or idx >= len(row):
                return 0.0
            v = row.iloc[idx]
            if pd.isna(v):
                return 0.0
            try:
                return float(str(v).replace(",", ""))
            except Exception:
                return 0.0

        for _, row in df1.iterrows():
            name = _get(row, "party_name")
            if not name:
                continue

            accounts: list[tuple[str, str, float]] = []
            acct = _get(row, "account")
            cur1 = _get(row, "currency1", "KRW")
            amt1 = _get_float(row, "amount1")
            if acct and amt1:
                accounts.append((acct, cur1, amt1))
            # 두 번째 계정 (통화2 / 금액2)
            cur2 = _get(row, "currency2")
            amt2 = _get_float(row, "amount2")
            if cur2 and amt2:
                accounts.append((acct, cur2, amt2))

            send_targets.append(PartyContact(
                name=name,
                country=_get(row, "country"),
                business_no=_get(row, "business_no"),
                ceo_name=_get(row, "ceo_name"),
                contact_person=_get(row, "contact_person"),
                phone=_get(row, "phone"),
                email=_get(row, "email"),
                accounts=accounts,
            ))

    # ── 발송제외 시트 파싱 ────────────────────────────────────
    excluded: list[ExcludedParty] = []
    EXCL_KEYWORDS = ["발송대상제외", "발송제외", "채채대상이나"]
    excl_sheet = next(
        (s for s in sheets if any(kw in s for kw in EXCL_KEYWORDS)),
        None
    )

    if excl_sheet:
        df_excl = pd.read_excel(path, sheet_name=excl_sheet, header=0, dtype=str)
        excl_headers = list(df_excl.columns)
        excl_col = _detect_ug_columns(excl_headers)
        # 단일 통화/금액 컬럼 fallback
        if "currency1" not in excl_col:
            # '통화' 키워드로 재시도
            for i, h in enumerate(excl_headers):
                h_str = str(h).strip() if h else ""
                if "통화" in h_str and "currency1" not in excl_col:
                    excl_col["currency1"] = i
                if "조회금액" in h_str and "amount1" not in excl_col:
                    excl_col["amount1"] = i

        for _, row in df_excl.iterrows():
            excl_name_idx = excl_col.get("party_name")
            if excl_name_idx is None:
                # 거래처명 컬럼 없으면 마지막 컬럼 추론
                excl_name_idx = len(excl_headers) - 1
            v = row.iloc[excl_name_idx] if excl_name_idx < len(row) else None
            name = str(v).strip() if pd.notna(v) and str(v).strip() not in ("nan", "None", "") else ""
            if not name:
                continue

            acct = ""
            acct_idx = excl_col.get("account")
            if acct_idx is not None and acct_idx < len(row):
                av = row.iloc[acct_idx]
                acct = str(av).strip() if pd.notna(av) and str(av).strip() not in ("nan", "None", "") else ""

            kind_val = ""
            kind_idx = excl_col.get("kind")
            if kind_idx is not None and kind_idx < len(row):
                kv = row.iloc[kind_idx]
                kind_val = str(kv).strip() if pd.notna(kv) and str(kv).strip() not in ("nan", "None", "") else ""

            cur = ""
            cur_idx = excl_col.get("currency1")
            if cur_idx is not None and cur_idx < len(row):
                cv = row.iloc[cur_idx]
                cur = str(cv).strip() if pd.notna(cv) and str(cv).strip() not in ("nan", "None", "") else ""

            amt = 0.0
            amt_idx = excl_col.get("amount1")
            if amt_idx is not None and amt_idx < len(row):
                av2 = row.iloc[amt_idx]
                if pd.notna(av2):
                    try:
                        amt = float(str(av2).replace(",", ""))
                    except Exception:
                        pass

            excluded.append(ExcludedParty(
                name=name, account_name=acct, currency=cur, amount=amt, kind=kind_val
            ))

    return UploadGuideData(send_targets=send_targets, excluded=excluded)


def load_ledger(
    path: str | Path,
    receivable_sheet: str | None = None,
    payable_sheet: str | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """회사 제시 거래처별 원장 → (채권 df, 채무 df)

    시트명을 지정하지 않으면 detect_ledger_sheets() 로 자동 감지.
    예상 컬럼: 코드 | 명 | 계정과목 | 계정과목명 | 통화 | 기초 | 증감 | 기말
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    wb.close()

    if receivable_sheet is None or payable_sheet is None:
        detected = detect_ledger_sheets(sheets)
        if receivable_sheet is None:
            receivable_sheet = detected.get("receivable") or "채권"
        if payable_sheet is None:
            payable_sheet = detected.get("payable") or "채무"

    ar = pd.read_excel(path, sheet_name=receivable_sheet, header=0)
    ap = pd.read_excel(path, sheet_name=payable_sheet, header=0)
    return ar, ap


def load_ledger_sheet(
    path: str | Path,
    sheet: str,
) -> tuple[pd.DataFrame, dict]:
    """단일 시트 로드 + 컬럼 자동 감지 결과 반환.

    Returns:
        (df, col_map) — col_map 은 detect_ledger_columns() 결과.
    """
    df = pd.read_excel(path, sheet_name=sheet, header=0)
    col_map = detect_ledger_columns(df)
    return df, col_map


def load_multi_sheet_ledger(
    path: str | Path,
    kind: str = "receivable",
) -> tuple[pd.DataFrame, dict]:
    """다중 계정과목 시트 방식의 원장을 통합 DataFrame으로 로드.

    코스맥스네오 양식처럼 시트별로 계정과목이 분리된 경우 사용.
    각 시트를 로드한 뒤 "account_name_injected" 컬럼에 시트명을 주입하여 통합.

    Returns:
        (merged_df, col_map)
        col_map: 첫 번째 시트에서 감지한 컬럼 인덱스 — 모든 시트 동일 구조 가정.
    """
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheetnames = wb.sheetnames
    wb.close()

    multi = detect_multi_account_sheets(sheetnames)
    sheets = multi.get(kind, [])
    if not sheets:
        raise ValueError(
            f"다중 시트 원장에서 kind='{kind}' 해당 시트를 찾을 수 없음. "
            f"사용 가능한 시트: {sheetnames}"
        )

    col_map: dict = {}
    parts: list[pd.DataFrame] = []
    for sn in sheets:
        df = pd.read_excel(path, sheet_name=sn, header=0)
        if not col_map:
            col_map = detect_ledger_columns(df)
        # 시트명을 계정과목명 override 컬럼으로 주입
        df = df.copy()
        df["_sheet_account_name"] = sn
        parts.append(df)

    merged = pd.concat(parts, ignore_index=True)
    return merged, col_map


def load_related_parties(path: str | Path, sheet: str | None = None) -> set[str]:
    """특관자리스트 → 이름 집합. sheet=None 이면 detect_rp_sheet() 자동 감지."""
    wb = openpyxl.load_workbook(path, data_only=True)

    # 시트명 자동 감지
    if sheet is None:
        sheet = detect_rp_sheet(wb.sheetnames)
    if sheet is None or sheet not in wb.sheetnames:
        # 마지막 fallback: 첫 번째 시트
        sheet = wb.sheetnames[0] if wb.sheetnames else None
    if sheet is None or sheet not in wb.sheetnames:
        return set()

    ws = wb[sheet]
    names: set[str] = set()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and isinstance(v, str) and v.strip():
            names.add(v.strip())
    return names


def load_fs_amounts(
    path: str | Path,
    sheet: str | None = None,
    item_col: int | None = None,
    value_col: int | None = None,
) -> dict[str, float]:
    """재무제표(자산·부채) → {계정명: 잔액}

    시트명·컬럼 위치를 지정하지 않으면 자동 감지:
    - 시트명: detect_fs_sheet() — FS_M / BS / 재무상태표 / 재무제표 등
    - 컬럼: 헤더 행 키워드 감지 — 항목·계정명 컬럼 + 금액·잔액 컬럼 (당기 우선)
    - 자동 감지 실패 시 기본값: item_col=3, value_col=5 (7620 호환)
    """
    from .schemas.fs_schema import detect_fs_sheet, detect_fs_columns

    wb = openpyxl.load_workbook(path, data_only=True)

    # 시트 자동 감지
    if sheet is None:
        sheet = detect_fs_sheet(wb.sheetnames) or "FS_M"
    if sheet not in wb.sheetnames:
        return {}

    ws = wb[sheet]

    # 컬럼 자동 감지 — 첫 번째 데이터 행 키워드 기반
    if item_col is None or value_col is None:
        header_vals = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        detected = detect_fs_columns(wb.sheetnames, header_vals)
        # 당기 금액 우선 (value_col이 여러 개 감지된 경우 마지막 = 당기)
        if item_col is None:
            item_col = detected.get("item_col") or 3
        if value_col is None:
            value_col = detected.get("value_col") or 5

    result: dict[str, float] = {}
    for r in range(1, ws.max_row + 1):
        item = ws.cell(r, item_col).value
        val = ws.cell(r, value_col).value
        if isinstance(item, str) and item.strip() and isinstance(val, (int, float)):
            result[item.strip()] = float(val)
    return result


def get_total_assets(fs_amounts: dict[str, float]) -> float:
    """총자산 추출 — 명칭 변형 대응"""
    for key in ("자산총계", "자산 총계", "총자산"):
        if key in fs_amounts:
            return fs_amounts[key]
    # fallback: 유동자산 + 비유동자산
    유동 = fs_amounts.get("유동자산", 0.0)
    비유동 = fs_amounts.get("비유동자산", 0.0)
    return 유동 + 비유동
