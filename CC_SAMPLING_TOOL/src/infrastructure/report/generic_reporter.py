"""
generic_reporter.py — Toss 디자인 + 10시트 단일 통합 조서 리포터

시트 구성:
  1. 요약               — KPI 카드 + 단계 진행 현황
  2. 샘플링 거래처 내역  — C100·AA100 채권/채무 통합 (한 표, 50건 발송 명단)
  3. 조회서             — UploadGuide 양식 (C1~C15) + 회신 상태 컬럼 (C16~C21)
  4. 표본규모 산출       — 채권/채무 섹션 2개
  5. 모집단 완전성       — 완전성 대사 + 발송제외
  6. Key item 매트릭스  — 채권/채무 통합
  7. MUS 추출 내역       — 채권/채무 별 표
  8. 주소 적정성         — UploadGuide 연락처
  9. 회신 추적           — PDF 회신 일치/불일치
  10. 대체적 절차        — 미회신·불일치 + 증빙

디자인: Toss 컨셉 — 흰 배경, #3182F6 accent, 얇은 회색 테두리
템플릿 복사 없음. openpyxl 직접 스타일링.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.domain.mus import MUSResult
from src.domain.population import CompletenessCheck, PartyDecision
from src.domain.sample_size import SampleSizeResult
from src.infrastructure.loaders import UploadGuideData, PartyContact


# ─────────────────────────────────────────────────────────────
# 기대 시트 이름 집합 (테스트·검증용)
# ─────────────────────────────────────────────────────────────
EXPECTED_GENERIC_SHEETS: set[str] = {
    "요약",
    "샘플링 거래처 내역",
    "조회서",
    "표본규모 산출",
    "모집단 완전성",
    "Key item 매트릭스",
    "MUS 추출 내역",
    "주소 적정성",
    "회신 추적",
    "대체적 절차",
}

EXPECTED_SHEET_COUNT = 10


# ─────────────────────────────────────────────────────────────
# Toss 디자인 색상 상수
# ─────────────────────────────────────────────────────────────

FONT_NAME = "맑은 고딕"

# 색상 코드 (# 없는 hex RGB)
TOSS_ACCENT   = "3182F6"   # 파란 accent (헤더·KPI 강조)
TOSS_ACCENT_D = "1B64DA"   # accent dark
TOSS_TEXT     = "191F28"   # 기본 텍스트
TOSS_TEXT2    = "4E5968"   # 보조 텍스트
TOSS_TEXT3    = "8B95A1"   # 흐린 텍스트
TOSS_BORDER   = "E5E8EB"   # 얇은 테두리 회색
TOSS_BG_SUB   = "F9FAFB"   # 합계행·카드 배경
TOSS_GREEN    = "00C073"   # 긍정/일치
TOSS_RED      = "F04452"   # 부정/불일치
TOSS_AMBER    = "FF9F00"   # 경고
TOSS_GOLD     = "FFD966"   # Key item 좌측 마커
TOSS_WHITE    = "FFFFFF"

# ─── 폰트 ───────────────────────────────────────────────────
FONT_KPI_VALUE = Font(name=FONT_NAME, size=14, bold=True, color=TOSS_ACCENT)
FONT_KPI_LABEL = Font(name=FONT_NAME, size=9,  color=TOSS_TEXT3)
FONT_HEADER    = Font(name=FONT_NAME, size=10, bold=True, color=TOSS_TEXT)
FONT_BODY      = Font(name=FONT_NAME, size=10, color=TOSS_TEXT)
FONT_BOLD      = Font(name=FONT_NAME, size=10, bold=True, color=TOSS_TEXT)
FONT_ACCENT    = Font(name=FONT_NAME, size=10, bold=True, color=TOSS_ACCENT)
FONT_FADED     = Font(name=FONT_NAME, size=10, color=TOSS_TEXT3)
FONT_FADED_STRIKE = Font(name=FONT_NAME, size=10, color=TOSS_TEXT3, strikethrough=True)
FONT_GREEN     = Font(name=FONT_NAME, size=10, bold=True, color=TOSS_GREEN)
FONT_RED       = Font(name=FONT_NAME, size=10, bold=True, color=TOSS_RED)
FONT_AMBER     = Font(name=FONT_NAME, size=10, bold=True, color=TOSS_AMBER)

# ─── 채우기 ─────────────────────────────────────────────────
FILL_WHITE   = PatternFill("solid", fgColor=TOSS_WHITE)
FILL_BG_SUB  = PatternFill("solid", fgColor=TOSS_BG_SUB)   # 합계행·소제목

# ─── 테두리 ─────────────────────────────────────────────────
_THIN_SIDE       = Side(style="thin",   color=TOSS_BORDER)
_THICK_SIDE      = Side(style="medium", color=TOSS_BORDER)
_ACCENT_SIDE     = Side(style="medium", color=TOSS_ACCENT)
_GOLD_SIDE       = Side(style="thick",  color=TOSS_GOLD)
_GREEN_SIDE      = Side(style="thick",  color=TOSS_GREEN)
_RED_SIDE        = Side(style="thick",  color=TOSS_RED)

BORDER_LIGHT     = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE, bottom=_THIN_SIDE,
)
BORDER_HEADER    = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE, bottom=Side(style="medium", color=TOSS_ACCENT),
)
BORDER_TOTAL     = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=Side(style="medium", color=TOSS_BORDER),
    bottom=_THIN_SIDE,
)
BORDER_SUBHEADER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE,
    bottom=Side(style="medium", color=TOSS_BORDER),
)

def _left_ki_border() -> Border:
    """Key item 좌측 강조 (GOLD 두꺼운 좌측선)."""
    return Border(
        left=_GOLD_SIDE, right=_THIN_SIDE,
        top=_THIN_SIDE, bottom=_THIN_SIDE,
    )

def _left_rep_border() -> Border:
    """MUS hit 좌측 강조 (GREEN)."""
    return Border(
        left=_GREEN_SIDE, right=_THIN_SIDE,
        top=_THIN_SIDE, bottom=_THIN_SIDE,
    )

def _left_excl_border() -> Border:
    """발송제외 좌측 표시."""
    return Border(
        left=_RED_SIDE, right=_THIN_SIDE,
        top=_THIN_SIDE, bottom=_THIN_SIDE,
    )

# ─── 숫자 포맷 ──────────────────────────────────────────────
NUMFMT_INT  = '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-'
NUMFMT_DATE = "yyyy-mm-dd"
NUMFMT_PCT  = "0.0%"

# ─────────────────────────────────────────────────────────────
# 하위 호환용 — 구버전 테스트가 임포트하는 상수 (폐기 예정)
# ─────────────────────────────────────────────────────────────
FILL_INPUT_YELLOW = PatternFill("solid", fgColor="FFFFCC")   # 하위 호환
FILL_HEADER_NAVY  = PatternFill("solid", fgColor="44546A")   # 하위 호환
FILL_SUBHEADER    = PatternFill("solid", fgColor="D6DCE5")   # 하위 호환
FILL_KEY_ITEM     = PatternFill("solid", fgColor="FFF2CC")   # 하위 호환
FILL_SAMPLED      = PatternFill("solid", fgColor="C6EFCE")   # 하위 호환
FILL_TOTAL_ROW    = PatternFill("solid", fgColor="D6DCE5")   # 하위 호환
FONT_RED_BOLD     = Font(name=FONT_NAME, size=10, bold=True, color="FF0000")  # 하위 호환


# ─────────────────────────────────────────────────────────────
# 데이터 클래스
# ─────────────────────────────────────────────────────────────

@dataclass
class ReportContext:
    company_name: str
    period_end: date
    kind: str                          # "receivable" | "payable" | "both"
    preparer: str = ""
    reviewer: str = ""
    prep_date: date | None = None
    review_date: date | None = None
    workpaper_no_prefix: str = "C100"


@dataclass
class PartyContactInfo:
    """거래처 연락처 — UploadGuide 또는 수동 입력."""
    name: str
    country: str = ""
    business_no: str = ""
    ceo_name: str = ""
    contact_person: str = ""
    phone: str = ""
    email: str = ""


@dataclass
class ExclusionRow:
    """발송제외 거래처 행."""
    name: str
    account_name: str = ""
    currency: str = ""
    amount: float = 0.0
    kind: str = ""


@dataclass
class ConfirmationReplyInfo:
    """PDF 회신 결과 요약 — Step 4 완료 시."""
    party_name: str
    status: str           # "matched" | "mismatch" | "needs_review" | "미회신"
    extracted_balance: float | None = None
    reply_date: str | None = None


@dataclass
class AlternativeProcedureEntry:
    """대체적 절차 — Step 5 완료 시."""
    party_name: str
    reason: str
    ledger_balance: float | None
    procedure_type: str
    evidence_names: list[str]
    covered_amount: float | None
    coverage_ratio: float | None
    conclusion: str
    auditor_notes: str | None = None


# ─────────────────────────────────────────────────────────────
# KindData — 채권/채무 데이터 묶음
# ─────────────────────────────────────────────────────────────

@dataclass
class KindData:
    """채권/채무 한쪽에 필요한 모든 데이터 묶음."""
    ctx: ReportContext
    completeness: CompletenessCheck
    size_result: SampleSizeResult
    decisions: list[PartyDecision]
    mus_result: MUSResult
    performance_materiality: float
    population_amount: float
    contacts: list[PartyContactInfo] | None = None
    exclusion_rows: list[ExclusionRow] | None = None
    pdf_replies: list[ConfirmationReplyInfo] | None = None
    alt_procedures: list[AlternativeProcedureEntry] | None = None


# ─────────────────────────────────────────────────────────────
# 공통 스타일 헬퍼
# ─────────────────────────────────────────────────────────────

def _set_col_width(ws: Worksheet, col: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width


def _al(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _apply(cell, font=None, fill=None, border=None, alignment=None,
           number_format: str | None = None) -> None:
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if border is not None:
        cell.border = border
    if alignment is not None:
        cell.alignment = alignment
    if number_format is not None:
        cell.number_format = number_format


def _row_height(ws: Worksheet, row: int, h: float) -> None:
    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, h)


# ── Toss 스타일 셀 헬퍼 ──────────────────────────────────────

def _header_cell(ws: Worksheet, row: int, col: int, text: str,
                 col_end: int | None = None) -> None:
    """테이블 헤더: 흰 배경 + 파란 bottom border + bold 검정."""
    c = ws.cell(row, col, text)
    _apply(c, font=FONT_HEADER, fill=FILL_WHITE,
           border=BORDER_HEADER, alignment=_al("center", wrap=True))
    if col_end and col_end > col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col_end)
    _row_height(ws, row, 22)


def _subheader_row(ws: Worksheet, row: int, text: str,
                   col_start: int = 1, col_end: int = 9) -> None:
    """소제목 구분선: 옅은 회색 배경 + bold + 굵은 bottom border."""
    c = ws.cell(row, col_start, text)
    _apply(c, font=FONT_BOLD, fill=FILL_BG_SUB,
           border=BORDER_SUBHEADER, alignment=_al("left"))
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start,
                       end_row=row, end_column=col_end)
    _row_height(ws, row, 22)


def _text_cell(ws: Worksheet, row: int, col: int, value,
               font=None, border=None, align: str = "left",
               is_ki: bool = False, is_rep: bool = False,
               is_excl: bool = False) -> None:
    c = ws.cell(row, col, value)
    b = border or (
        _left_ki_border()   if is_ki   else
        _left_rep_border()  if is_rep  else
        _left_excl_border() if is_excl else
        BORDER_LIGHT
    )
    f = font or (
        FONT_BOLD if is_ki else
        FONT_BODY
    )
    _apply(c, font=f, fill=FILL_WHITE, border=b, alignment=_al(align))
    _row_height(ws, row, 22)


def _num_cell(ws: Worksheet, row: int, col: int, value,
              font=None, border=None,
              fmt: str = NUMFMT_INT,
              is_ki: bool = False, is_rep: bool = False) -> None:
    c = ws.cell(row, col, value)
    b = border or (
        _left_ki_border()  if is_ki  else
        _left_rep_border() if is_rep else
        BORDER_LIGHT
    )
    f = font or (FONT_BOLD if is_ki else FONT_BODY)
    _apply(c, font=f, fill=FILL_WHITE, border=b,
           alignment=_al("right"), number_format=fmt)
    _row_height(ws, row, 22)


def _pct_cell(ws: Worksheet, row: int, col: int, value: float | None,
              font=None) -> None:
    if value is not None:
        c = ws.cell(row, col, value)
        _apply(c, font=font or FONT_BODY, fill=FILL_WHITE, border=BORDER_LIGHT,
               alignment=_al("center"), number_format=NUMFMT_PCT)
    else:
        _text_cell(ws, row, col, "", align="center")


def _accent_num_cell(ws: Worksheet, row: int, col: int, value,
                     fmt: str = NUMFMT_INT) -> None:
    """파란 bold 숫자 — PM, Key item 기준금액 등 강조."""
    c = ws.cell(row, col, value)
    _apply(c, font=FONT_ACCENT, fill=FILL_WHITE, border=BORDER_LIGHT,
           alignment=_al("right"), number_format=fmt)
    _row_height(ws, row, 22)


def _total_row_cell(ws: Worksheet, row: int, col: int, value,
                    is_num: bool = False, fmt: str = NUMFMT_INT) -> None:
    """합계 행: 옅은 회색 배경 + bold + 상단 굵은 border."""
    c = ws.cell(row, col, value)
    _apply(c, font=FONT_BOLD, fill=FILL_BG_SUB, border=BORDER_TOTAL,
           alignment=_al("right" if is_num else "left"),
           number_format=fmt if is_num else None)
    _row_height(ws, row, 22)


def _kpi_card(ws: Worksheet, row_top: int, col_left: int,
              label: str, value, sub: str = "",
              col_span: int = 3, row_span: int = 4,
              fmt: str = NUMFMT_INT) -> None:
    """KPI 카드 — 병합 셀 + 두꺼운 accent border."""
    col_right = col_left + col_span - 1
    row_bot   = row_top + row_span - 1

    # 카드 테두리 (전체 병합 범위)
    card_border = Border(
        left=Side(style="medium", color=TOSS_BORDER),
        right=Side(style="medium", color=TOSS_BORDER),
        top=Side(style="medium", color=TOSS_BORDER),
        bottom=Side(style="medium", color=TOSS_BORDER),
    )

    # 라벨 셀 (row_top)
    lbl = ws.cell(row_top, col_left, label)
    _apply(lbl, font=FONT_KPI_LABEL, fill=FILL_WHITE,
           border=card_border, alignment=_al("left", "center"))

    # 값 셀 (row_top+1)
    val_cell = ws.cell(row_top + 1, col_left, value)
    _apply(val_cell, font=FONT_KPI_VALUE, fill=FILL_WHITE,
           border=card_border,
           alignment=_al("left", "center"),
           number_format=fmt)

    # 부가 텍스트 셀 (row_top+2)
    if sub:
        sub_cell = ws.cell(row_top + 2, col_left, sub)
        _apply(sub_cell, font=FONT_KPI_LABEL, fill=FILL_WHITE,
               border=card_border, alignment=_al("left", "center", wrap=True))

    # 병합
    if col_span > 1:
        for r in range(row_top, row_bot + 1):
            ws.merge_cells(start_row=r, start_column=col_left,
                           end_row=r, end_column=col_right)

    # 행 높이
    for r_idx in range(row_top, row_bot + 1):
        _row_height(ws, r_idx, 22)


# ── 조서 헤더 블록 ───────────────────────────────────────────

def _write_doc_header(
    ws: Worksheet,
    company: str,
    title: str,
    period_end: date,
    preparer: str,
    reviewer: str,
    prep_date: date,
    review_date: date,
    wp_no: str,
    last_col: int = 9,
) -> None:
    """조서 공통 헤더 R1~R3."""
    _row_height(ws, 1, 20)
    _row_height(ws, 2, 20)
    _row_height(ws, 3, 18)
    _row_height(ws, 4, 6)

    c = ws.cell(1, 1, company)
    _apply(c, font=FONT_BODY, border=BORDER_LIGHT, alignment=_al("left"))
    c = ws.cell(2, 1, title)
    _apply(c, font=FONT_BOLD, border=BORDER_LIGHT, alignment=_al("left"))
    c = ws.cell(3, 1, period_end)
    _apply(c, font=FONT_BODY, border=BORDER_LIGHT,
           alignment=_al("left"), number_format=NUMFMT_DATE)

    for row, label in ((1, "작성자:"), (2, "검토자:")):
        c = ws.cell(row, 5, label)
        _apply(c, font=FONT_BODY, border=BORDER_LIGHT, alignment=_al("right"))

    f1 = ws.cell(1, 6, preparer)
    _apply(f1, font=FONT_BODY, border=BORDER_LIGHT, alignment=_al("left"))
    f2 = ws.cell(2, 6, reviewer)
    _apply(f2, font=FONT_BODY, border=BORDER_LIGHT, alignment=_al("left"))

    for row in (1, 2):
        c = ws.cell(row, 7, "일자:")
        _apply(c, font=FONT_BODY, border=BORDER_LIGHT, alignment=_al("right"))

    h1 = ws.cell(1, 8, prep_date)
    _apply(h1, font=FONT_BODY, border=BORDER_LIGHT,
           alignment=_al("center"), number_format=NUMFMT_DATE)
    h2 = ws.cell(2, 8, review_date)
    _apply(h2, font=FONT_BODY, border=BORDER_LIGHT,
           alignment=_al("center"), number_format=NUMFMT_DATE)

    # 조서번호 — Toss: accent bold 파란색
    c = ws.cell(1, 9, "조서번호")
    _apply(c, font=FONT_BODY, border=BORDER_LIGHT, alignment=_al("center"))
    c = ws.cell(2, 9, wp_no)
    _apply(c, font=FONT_ACCENT, border=BORDER_LIGHT, alignment=_al("center"))

    for row in range(1, 4):
        for col in range(2, 5):
            c = ws.cell(row, col)
            if not c.value:
                _apply(c, font=FONT_BODY, border=BORDER_LIGHT, alignment=_al("left"))
    for col in range(2, last_col + 1):
        c = ws.cell(3, col)
        if not c.value:
            _apply(c, font=FONT_BODY, border=BORDER_LIGHT, alignment=_al("left"))


# ─────────────────────────────────────────────────────────────
# 메인 진입점 — 단일 kind
# ─────────────────────────────────────────────────────────────

def build_generic_report(
    out_path: str | Path,
    ctx: ReportContext,
    completeness: CompletenessCheck,
    size_result: SampleSizeResult,
    decisions: list[PartyDecision],
    mus_result: MUSResult,
    performance_materiality: float,
    population_amount: float,
    contacts: list[PartyContactInfo] | None = None,
    exclusion_rows: list[ExclusionRow] | None = None,
    pdf_replies: list[ConfirmationReplyInfo] | None = None,
    alt_procedures: list[AlternativeProcedureEntry] | None = None,
    upload_guide_data: UploadGuideData | None = None,
) -> None:
    """단일 kind → 10시트 조서 생성 (채권 또는 채무)."""
    kd = KindData(
        ctx=ctx,
        completeness=completeness,
        size_result=size_result,
        decisions=decisions,
        mus_result=mus_result,
        performance_materiality=performance_materiality,
        population_amount=population_amount,
        contacts=contacts,
        exclusion_rows=exclusion_rows,
        pdf_replies=pdf_replies,
        alt_procedures=alt_procedures,
    )
    build_combined_report(
        out_path,
        receivable=kd if ctx.kind != "payable" else None,
        payable=kd if ctx.kind == "payable" else None,
        upload_guide_data=upload_guide_data,
    )


def build_combined_report(
    out_path: str | Path,
    receivable: KindData | None,
    payable: KindData | None,
    upload_guide_data: UploadGuideData | None = None,
) -> None:
    """채권+채무 단일 워크북 — 10시트 통합 출력."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    base_kd   = receivable or payable
    base_ctx  = base_kd.ctx
    ar_kd     = receivable
    ap_kd     = payable

    # ── 합산 데이터 준비 ──────────────────────────────────────
    all_decisions: list[PartyDecision] = []
    all_contacts:  list[PartyContactInfo] = []
    all_replies:   list[ConfirmationReplyInfo] = []
    all_alt_procs: list[AlternativeProcedureEntry] = []
    total_pop  = 0.0
    total_pm   = base_kd.performance_materiality

    for kd in (receivable, payable):
        if kd is None:
            continue
        all_decisions.extend(kd.decisions)
        if kd.contacts:
            all_contacts.extend(kd.contacts)
        if kd.pdf_replies:
            all_replies.extend(kd.pdf_replies)
        if kd.alt_procedures:
            all_alt_procs.extend(kd.alt_procedures)
        total_pop += kd.population_amount

    _build_sheet_summary(wb, base_ctx, ar_kd, ap_kd,
                         all_decisions, all_replies, all_alt_procs,
                         total_pop, total_pm)
    _build_sheet_sampling_party_list(wb, base_ctx, ar_kd, ap_kd,
                                     all_contacts, all_replies)
    _build_sheet_uploadguide_confirmation(wb, base_ctx, upload_guide_data,
                                          all_replies, all_alt_procs)
    _build_sheet_sample_size(wb, base_ctx, ar_kd, ap_kd)
    _build_sheet_completeness(wb, base_ctx, ar_kd, ap_kd)
    _build_sheet_key_item_matrix(wb, base_ctx, ar_kd, ap_kd, total_pm)
    _build_sheet_mus_detail(wb, base_ctx, ar_kd, ap_kd)
    _build_sheet_address(wb, base_ctx, all_contacts)
    _build_sheet_reply_tracking(wb, base_ctx, all_replies, all_decisions)
    _build_sheet_alt_procedures(wb, base_ctx, all_alt_procs)

    wb.save(out_path)


# ─────────────────────────────────────────────────────────────
# Sheet 1: 요약
# ─────────────────────────────────────────────────────────────

def _build_sheet_summary(
    wb,
    ctx: ReportContext,
    ar_kd: KindData | None,
    ap_kd: KindData | None,
    all_decisions: list[PartyDecision],
    all_replies: list[ConfirmationReplyInfo],
    alt_procs: list[AlternativeProcedureEntry],
    total_pop: float,
    pm: float,
) -> None:
    ws = wb.create_sheet("요약")

    for col, w in [(1, 4), (2, 20), (3, 18), (4, 18), (5, 18),
                   (6, 18), (7, 18), (8, 14), (9, 14)]:
        _set_col_width(ws, col, w)

    prep_date   = ctx.prep_date   or date.today()
    review_date = ctx.review_date or date.today()
    kind_label  = {"receivable": "채권 조회", "payable": "채무 조회",
                   "both": "채권채무 조회 통합"}.get(ctx.kind, ctx.kind)

    _write_doc_header(
        ws, ctx.company_name, f"샘플링 결과 요약 — {kind_label}",
        ctx.period_end, ctx.preparer, ctx.reviewer,
        prep_date, review_date, wp_no="요약",
    )

    r = 5

    # ── KPI 카드 4개 ──────────────────────────────────────────
    _subheader_row(ws, r, "핵심 KPI", col_end=9)
    r += 1

    ki_all    = [d for d in all_decisions if d.is_key_item and not d.is_excluded]
    final_all = [d for d in all_decisions if d.final_sampled and not d.is_excluded]
    ar_size   = ar_kd.size_result if ar_kd else None
    ap_size   = ap_kd.size_result if ap_kd else None
    ar_pop    = ar_kd.population_amount if ar_kd else 0.0
    ap_pop    = ap_kd.population_amount if ap_kd else 0.0

    ar_sub = f"채권 {ar_pop/1e8:.0f}억" if ar_kd else ""
    ap_sub = f"채무 {ap_pop/1e8:.0f}억" if ap_kd else ""
    pop_sub = (ar_sub + ("  ·  " + ap_sub if ap_sub else "")) if ar_sub else ap_sub

    ki_threshold = (ar_size or ap_size).key_item_threshold if (ar_size or ap_size) else 0
    final_n = len(final_all)
    rep_size = (ar_size or ap_size)
    ki_threshold_sub = f"PM × {(rep_size.key_item_ratio * 100):.0f}%" if rep_size else ""

    cards = [
        ("모집단",             total_pop,      pop_sub,       2),
        ("수행중요성 (PM)",    pm,             "기준 금액",    5),
        ("Key item 기준금액", ki_threshold,   ki_threshold_sub, 2),
        ("최종 샘플링 건수",   final_n,        "Key+Rep+특관자", 5),
    ]
    _row_height(ws, r, 22)
    _row_height(ws, r + 1, 28)
    _row_height(ws, r + 2, 20)

    for label, value, sub, col_start in cards:
        fmt = NUMFMT_INT if isinstance(value, float) else "0"
        _kpi_card(ws, r, col_start, label, value, sub=sub,
                  col_span=2, row_span=3, fmt=fmt)

    r += 4

    # ── 채권/채무 분포 표 ─────────────────────────────────────
    _subheader_row(ws, r, "채권·채무 분포 현황", col_end=9)
    r += 1

    for col_idx, h in enumerate(["구분", "모집단 건수", "모집단 금액", "Key item",
                                  "Rep", "특관자", "발송제외", "최종 샘플"], 1):
        _header_cell(ws, r, col_idx, h)
    r += 1

    dist_rows = []
    if ar_kd:
        d_list = ar_kd.decisions
        dist_rows.append(("채권", d_list, ar_kd.population_amount))
    if ap_kd:
        d_list = ap_kd.decisions
        dist_rows.append(("채무", d_list, ap_kd.population_amount))
    if ar_kd and ap_kd:
        dist_rows.append(("합계", all_decisions, total_pop))

    for kind_lbl, d_list, pop_amt in dist_rows:
        is_total = kind_lbl == "합계"
        incl = [d for d in d_list if not d.is_excluded]
        ki_n   = len([d for d in incl if d.is_key_item])
        rep_n  = len([d for d in incl if d.is_representative and not d.is_key_item])
        rp_n   = len([d for d in incl if d.is_related_party])
        excl_n = len([d for d in d_list if d.is_excluded])
        fin_n  = len([d for d in d_list if d.final_sampled and not d.is_excluded])

        if is_total:
            _total_row_cell(ws, r, 1, kind_lbl)
            _total_row_cell(ws, r, 2, len(incl), is_num=True)
            _total_row_cell(ws, r, 3, pop_amt, is_num=True)
            _total_row_cell(ws, r, 4, ki_n,   is_num=True)
            _total_row_cell(ws, r, 5, rep_n,  is_num=True)
            _total_row_cell(ws, r, 6, rp_n,   is_num=True)
            _total_row_cell(ws, r, 7, excl_n, is_num=True)
            _total_row_cell(ws, r, 8, fin_n,  is_num=True)
        else:
            _text_cell(ws, r, 1, kind_lbl, font=FONT_BOLD)
            _num_cell(ws, r, 2, len(incl))
            _num_cell(ws, r, 3, pop_amt)
            _num_cell(ws, r, 4, ki_n)
            _num_cell(ws, r, 5, rep_n)
            _num_cell(ws, r, 6, rp_n)
            _num_cell(ws, r, 7, excl_n)
            _num_cell(ws, r, 8, fin_n)
        r += 1

    # ── Step 진행 현황 ────────────────────────────────────────
    r += 1
    _subheader_row(ws, r, "단계별 진행 현황", col_end=9)
    r += 1

    steps = [
        ("Step 1 — 모집단 구성 및 표본규모 결정", True),
        ("Step 2 — UploadGuide 주소 확인",   bool(all_decisions)),
        ("Step 3 — 조서 생성",               True),
        ("Step 4 — 조회서 회신 처리",         bool(all_replies)),
        ("Step 5 — 대체적 절차",             bool(alt_procs)),
    ]
    for col_idx, h in enumerate(["단계", "완료 여부"], 1):
        _header_cell(ws, r, col_idx, h)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
    r += 1

    for step_label, done in steps:
        _text_cell(ws, r, 1, step_label)
        status_cell = ws.cell(r, 2, "완료" if done else "미완료")
        f = FONT_GREEN if done else FONT_FADED
        _apply(status_cell, font=f, fill=FILL_WHITE, border=BORDER_LIGHT,
               alignment=_al("left"))
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
        r += 1


# ─────────────────────────────────────────────────────────────
# Sheet 2: 샘플링 거래처 내역 (C100·AA100 통합 — 50건 발송 명단)
# ─────────────────────────────────────────────────────────────

_AR_ACCOUNTS = ["외상매출금", "받을어음", "미수금", "선급금", "임차보증금", "장기대여금"]
_AP_ACCOUNTS = ["외상매입금", "지급어음(외담대외상매입금)", "미지급금", "임대보증금"]


@dataclass
class _UnifiedPartyRow:
    """채권+채무 통합 거래처 행."""
    name: str
    ar_by_account: dict[str, float] = field(default_factory=dict)
    ap_by_account: dict[str, float] = field(default_factory=dict)
    ar_total: float = 0.0
    ap_total: float = 0.0
    is_key_item: bool = False
    is_representative: bool = False
    is_related_party: bool = False
    is_excluded: bool = False
    final_sampled: bool = False
    contact: PartyContactInfo | None = None
    reply_status: str = "미회신"


def _merge_decisions(
    ar_kd: KindData | None,
    ap_kd: KindData | None,
    contacts: list[PartyContactInfo],
    replies: list[ConfirmationReplyInfo],
) -> list[_UnifiedPartyRow]:
    """채권/채무 decisions 를 거래처명 기준으로 통합 → _UnifiedPartyRow 리스트.

    동일 거래처가 채권·채무 양쪽에 다른 표기로 존재할 때
    (예: "COSMAX INC" vs "COSMAX. INC") normalize 기반으로 한 행에 통합한다.
    canonical(대표 이름)은 먼저 등장한 이름 사용.
    """
    contact_map = {c.name: c for c in contacts}
    # normalize 기반 contact 역방향 (표기 차이 흡수)
    contact_norm_map = {_normalize_for_match(c.name): c for c in contacts}
    reply_map   = {rep.party_name: rep.status for rep in replies}
    # normalize 기반 reply 역방향
    reply_norm_map: dict[str, str] = {}
    for rep in replies:
        pnorm = _normalize_for_match(rep.party_name)
        # 우선순위: matched > mismatch > needs_review > 미회신
        status_priority = {"matched": 4, "mismatch": 3, "needs_review": 2, "미회신": 1}
        existing = reply_norm_map.get(pnorm)
        if existing is None or status_priority.get(rep.status, 0) > status_priority.get(existing, 0):
            reply_norm_map[pnorm] = rep.status

    rows: dict[str, _UnifiedPartyRow] = {}
    # normalized_name → canonical_name (이미 rows에 등록된 이름)
    norm_to_canonical: dict[str, str] = {}

    def _get_or_create(name: str) -> _UnifiedPartyRow:
        norm = _normalize_for_match(name)
        # 이미 같은 normalize 이름으로 등록된 canonical이 있으면 그쪽으로 통합
        if norm in norm_to_canonical:
            canonical = norm_to_canonical[norm]
            return rows[canonical]
        # 새 행 생성
        contact = contact_map.get(name) or contact_norm_map.get(norm)
        reply_status = reply_map.get(name) or reply_norm_map.get(norm, "미회신")
        rows[name] = _UnifiedPartyRow(
            name=name,
            contact=contact,
            reply_status=reply_status,
        )
        norm_to_canonical[norm] = name
        return rows[name]

    if ar_kd:
        for d in ar_kd.decisions:
            row = _get_or_create(d.name)
            row.ar_by_account = dict(d.by_account)
            row.ar_total = d.balance
            row.is_key_item     |= d.is_key_item
            row.is_representative |= d.is_representative
            row.is_related_party  |= d.is_related_party
            row.is_excluded       &= d.is_excluded  # 한쪽만 발송 가능 → AND
            row.final_sampled     |= d.final_sampled

    if ap_kd:
        for d in ap_kd.decisions:
            row = _get_or_create(d.name)
            row.ap_by_account = dict(d.by_account)
            row.ap_total = d.balance
            row.is_key_item       |= d.is_key_item
            row.is_representative |= d.is_representative
            row.is_related_party  |= d.is_related_party
            row.final_sampled     |= d.final_sampled

    # 한쪽에만 있는 발송제외 처리: ar_kd or ap_kd에서 excluded면 표시
    if ar_kd:
        for d in ar_kd.decisions:
            if d.is_excluded and d.name in rows:
                # 채무쪽도 없으면 excluded
                if rows[d.name].ap_total == 0:
                    rows[d.name].is_excluded = True
    if ap_kd:
        for d in ap_kd.decisions:
            if d.is_excluded and d.name in rows:
                if rows[d.name].ar_total == 0:
                    rows[d.name].is_excluded = True

    result = list(rows.values())
    # 최종샘플 + 발송대상 우선, 총 잔액 내림차순
    result.sort(key=lambda r: (
        not r.final_sampled,
        r.is_excluded,
        -(r.ar_total + r.ap_total),
    ))
    return result


def _build_sheet_confirmation(
    wb,
    ctx: ReportContext,
    ar_kd: KindData | None,
    ap_kd: KindData | None,
    contacts: list[PartyContactInfo],
    replies: list[ConfirmationReplyInfo],
) -> None:
    """하위 호환 — _build_sheet_sampling_party_list 로 위임."""
    _build_sheet_sampling_party_list(wb, ctx, ar_kd, ap_kd, contacts, replies)


def _build_sheet_sampling_party_list(
    wb,
    ctx: ReportContext,
    ar_kd: KindData | None,
    ap_kd: KindData | None,
    contacts: list[PartyContactInfo],
    replies: list[ConfirmationReplyInfo],
) -> None:
    ws = wb.create_sheet("샘플링 거래처 내역")

    # 컬럼 너비
    col_widths = {1: 4, 2: 30, 3: 10}
    # AR 계정 컬럼
    ar_start = 4
    for i in range(len(_AR_ACCOUNTS)):
        col_widths[ar_start + i] = 13
    ar_sum_col = ar_start + len(_AR_ACCOUNTS)
    # AP 계정 컬럼
    ap_start = ar_sum_col + 1
    for i in range(len(_AP_ACCOUNTS)):
        col_widths[ap_start + i] = 13
    ap_sum_col = ap_start + len(_AP_ACCOUNTS)
    col_total   = ap_sum_col + 1
    col_cls     = col_total + 1
    col_reply   = col_cls + 1
    col_email   = col_reply + 1

    col_widths[ar_sum_col] = 13
    col_widths[ap_sum_col] = 13
    col_widths[col_total]  = 13
    col_widths[col_cls]    = 18
    col_widths[col_reply]  = 12
    col_widths[col_email]  = 22

    for c, w in col_widths.items():
        _set_col_width(ws, c, w)

    prep_date   = ctx.prep_date   or date.today()
    review_date = ctx.review_date or date.today()

    _write_doc_header(
        ws, ctx.company_name, "샘플링 거래처 내역 — 채권·채무 통합 발송 명단",
        ctx.period_end, ctx.preparer, ctx.reviewer,
        prep_date, review_date, wp_no="발송명단",
        last_col=col_email,
    )

    r = 5
    _subheader_row(ws, r, "발송 거래처 목록 (채권·채무 통합)", col_end=col_email)
    r += 1

    # 헤더 행
    _header_cell(ws, r, 1,        "No")
    _header_cell(ws, r, 2,        "거래처명")
    _header_cell(ws, r, 3,        "구분")
    for i, acct in enumerate(_AR_ACCOUNTS):
        _header_cell(ws, r, ar_start + i, acct)
    _header_cell(ws, r, ar_sum_col, "채권 계")
    for i, acct in enumerate(_AP_ACCOUNTS):
        _header_cell(ws, r, ap_start + i, acct)
    _header_cell(ws, r, ap_sum_col, "채무 계")
    _header_cell(ws, r, col_total,  "총 잔액")
    _header_cell(ws, r, col_cls,    "분류")
    _header_cell(ws, r, col_reply,  "회신 상태")
    _header_cell(ws, r, col_email,  "이메일")
    r += 1

    unified_all = _merge_decisions(ar_kd, ap_kd, contacts, replies)
    # 조회서 시트 = 최종 발송 표본만 (final_sampled OR key item OR rep OR 특관자 OR 발송제외)
    unified = [
        r for r in unified_all
        if r.final_sampled or r.is_key_item or r.is_representative
        or r.is_related_party or r.is_excluded
    ]

    totals_ar: dict[str, float] = {a: 0.0 for a in _AR_ACCOUNTS}
    totals_ap: dict[str, float] = {a: 0.0 for a in _AP_ACCOUNTS}
    grand_total = 0.0
    seq = 0

    for row_data in unified:
        is_ki   = row_data.is_key_item and not row_data.is_excluded
        is_rep  = row_data.is_representative and not row_data.is_key_item and not row_data.is_excluded
        is_excl = row_data.is_excluded

        # 구분 텍스트
        kinds = []
        if row_data.ar_total > 0:
            kinds.append("채권")
        if row_data.ap_total > 0:
            kinds.append("채무")
        kind_str = "+".join(kinds) if kinds else "-"

        # 분류
        cls_parts = []
        if is_ki:
            cls_parts.append("Key item")
        if is_rep:
            cls_parts.append("Rep")
        if row_data.is_related_party:
            cls_parts.append("특관자")
        if is_excl:
            cls_parts.append("발송제외")
        cls_str = ", ".join(cls_parts) if cls_parts else "-"

        # 회신 상태
        status_raw = row_data.reply_status
        status_label = {
            "matched": "일치", "mismatch": "불일치",
            "needs_review": "검토필요", "미회신": "미회신",
        }.get(status_raw, status_raw)

        # 발송제외면 strikethrough, 회신 상태 폰트 색상
        name_font = FONT_FADED_STRIKE if is_excl else (FONT_BOLD if is_ki else FONT_BODY)
        reply_font = (
            FONT_GREEN  if status_label == "일치"  else
            FONT_RED    if status_label == "불일치" else
            FONT_AMBER  if status_label == "검토필요" else
            FONT_FADED
        )

        if not is_excl:
            seq += 1

        seq_val = seq if not is_excl else "-"

        # No
        c = ws.cell(r, 1, seq_val)
        _apply(c, font=FONT_FADED if is_excl else FONT_BODY,
               fill=FILL_WHITE,
               border=_left_ki_border() if is_ki else _left_excl_border() if is_excl else BORDER_LIGHT,
               alignment=_al("center"))
        _row_height(ws, r, 22)

        # 거래처명
        c = ws.cell(r, 2, row_data.name)
        _apply(c, font=name_font, fill=FILL_WHITE,
               border=BORDER_LIGHT, alignment=_al("left"))

        # 구분
        c = ws.cell(r, 3, kind_str)
        kind_font = Font(name=FONT_NAME, size=10,
                         color=TOSS_ACCENT if "채권" in kind_str and "채무" in kind_str
                         else TOSS_TEXT2)
        _apply(c, font=kind_font, fill=FILL_WHITE,
               border=BORDER_LIGHT, alignment=_al("center"))

        # AR 계정 컬럼
        for i, acct in enumerate(_AR_ACCOUNTS):
            amt = row_data.ar_by_account.get(acct, 0.0)
            c = ws.cell(r, ar_start + i, amt if amt else None)
            f = FONT_FADED if is_excl else (FONT_BOLD if is_ki and amt else FONT_BODY)
            _apply(c, font=f, fill=FILL_WHITE, border=BORDER_LIGHT,
                   alignment=_al("right"), number_format=NUMFMT_INT)
            if not is_excl:
                totals_ar[acct] = totals_ar.get(acct, 0.0) + amt

        # AR 계
        c = ws.cell(r, ar_sum_col, row_data.ar_total if row_data.ar_total else None)
        _apply(c, font=FONT_FADED if is_excl else FONT_BODY,
               fill=FILL_WHITE, border=BORDER_LIGHT,
               alignment=_al("right"), number_format=NUMFMT_INT)

        # AP 계정 컬럼
        for i, acct in enumerate(_AP_ACCOUNTS):
            amt = row_data.ap_by_account.get(acct, 0.0)
            c = ws.cell(r, ap_start + i, amt if amt else None)
            f = FONT_FADED if is_excl else (FONT_BOLD if is_ki and amt else FONT_BODY)
            _apply(c, font=f, fill=FILL_WHITE, border=BORDER_LIGHT,
                   alignment=_al("right"), number_format=NUMFMT_INT)
            if not is_excl:
                totals_ap[acct] = totals_ap.get(acct, 0.0) + amt

        # AP 계
        c = ws.cell(r, ap_sum_col, row_data.ap_total if row_data.ap_total else None)
        _apply(c, font=FONT_FADED if is_excl else FONT_BODY,
               fill=FILL_WHITE, border=BORDER_LIGHT,
               alignment=_al("right"), number_format=NUMFMT_INT)

        # 총 잔액
        total_row = row_data.ar_total + row_data.ap_total
        c = ws.cell(r, col_total, total_row if total_row else None)
        total_font = FONT_FADED if is_excl else (FONT_ACCENT if is_ki else FONT_BODY)
        _apply(c, font=total_font, fill=FILL_WHITE, border=BORDER_LIGHT,
               alignment=_al("right"), number_format=NUMFMT_INT)
        if not is_excl:
            grand_total += total_row

        # 분류
        c = ws.cell(r, col_cls, cls_str)
        cls_font = FONT_FADED if is_excl else (FONT_ACCENT if is_ki else FONT_BODY)
        _apply(c, font=cls_font, fill=FILL_WHITE, border=BORDER_LIGHT,
               alignment=_al("left"))

        # 회신 상태
        c = ws.cell(r, col_reply, status_label)
        _apply(c, font=reply_font, fill=FILL_WHITE, border=BORDER_LIGHT,
               alignment=_al("center"))

        # 이메일
        ct = row_data.contact
        c = ws.cell(r, col_email, ct.email if ct else "")
        _apply(c, font=FONT_FADED if is_excl else FONT_BODY,
               fill=FILL_WHITE, border=BORDER_LIGHT, alignment=_al("left"))

        r += 1

    # 합계 행
    _total_row_cell(ws, r, 1, "")
    _total_row_cell(ws, r, 2, "총합계")
    _total_row_cell(ws, r, 3, "")
    for i, acct in enumerate(_AR_ACCOUNTS):
        _total_row_cell(ws, r, ar_start + i, totals_ar.get(acct) or None, is_num=True)
    _total_row_cell(ws, r, ar_sum_col, sum(totals_ar.values()) or None, is_num=True)
    for i, acct in enumerate(_AP_ACCOUNTS):
        _total_row_cell(ws, r, ap_start + i, totals_ap.get(acct) or None, is_num=True)
    _total_row_cell(ws, r, ap_sum_col, sum(totals_ap.values()) or None, is_num=True)
    _total_row_cell(ws, r, col_total, grand_total or None, is_num=True)
    for col in (col_cls, col_reply, col_email):
        _total_row_cell(ws, r, col, "")


# ─────────────────────────────────────────────────────────────
# Sheet 3: 표본규모 산출
# ─────────────────────────────────────────────────────────────

def _build_sheet_sample_size(
    wb,
    ctx: ReportContext,
    ar_kd: KindData | None,
    ap_kd: KindData | None,
) -> None:
    ws = wb.create_sheet("표본규모 산출")

    for c, w in [(1, 35), (2, 20), (3, 18), (4, 15), (5, 15)]:
        _set_col_width(ws, c, w)

    prep_date   = ctx.prep_date   or date.today()
    review_date = ctx.review_date or date.today()

    _write_doc_header(
        ws, ctx.company_name, "표본규모 산출 (MUS)",
        ctx.period_end, ctx.preparer, ctx.reviewer,
        prep_date, review_date, wp_no="규모산출",
    )

    r = 5
    method_text = (
        "MUS(Monetary Unit Sampling): 화폐단위를 표본단위로 사용하는 통계적 표본추출. "
        "표본간격(J) = 잔여모집단 ÷ Final sample size. (감사기준서 530)"
    )
    c_cell = ws.cell(r, 1, method_text)
    _apply(c_cell, font=FONT_FADED, fill=FILL_WHITE, border=BORDER_LIGHT,
           alignment=_al("left", wrap=True))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    _row_height(ws, r, 36)
    r += 2

    for kd, kind_label in [(ar_kd, "채권 (C100)"), (ap_kd, "채무 (AA100)")]:
        if kd is None:
            continue

        is_ap = kd.ctx.kind == "payable"
        _subheader_row(ws, r, kind_label, col_end=5)
        r += 1

        for col_idx, h in enumerate(["항목", "값", "산식", "", ""], 1):
            _header_cell(ws, r, col_idx, h)
        r += 1

        size = kd.size_result
        decisions = kd.decisions
        ki = [d for d in decisions if d.is_key_item and not d.is_excluded]
        ki_amt = sum(d.balance for d in ki)

        # 채무: 모집단 = 당기 활동량 합계 (ISA 505 완전성 검토)
        pop_label = "모집단 금액 (당기 활동량)" if is_ap else "모집단 금액"
        pop_note  = "|기초|+|증감| 합계 — ISA 505 under-statement risk" if is_ap else ""

        params = [
            (pop_label,              kd.population_amount,    pop_note,                             False),
            ("수행중요성 (PM)",       kd.performance_materiality, "",                               True),
            ("Key item 비율",        size.key_item_ratio,     "위험×통제 매트릭스",                 False, "0%"),
            ("Key item 기준금액",    size.key_item_threshold, f"PM × {size.key_item_ratio*100:.0f}%", True),
            ("Key item 건수",        len(ki),                 "",                                   False, NUMFMT_INT),
            ("Key item 금액",        ki_amt,                  "",                                   False),
            ("잔여 모집단",          size.remaining_population, "모집단 − Key item 금액",           False),
            ("Base sample size",     size.base_sample_size,   "잔여모집단 ÷ PM",                    False, "0.0"),
            ("신뢰계수 (CF)",        size.confidence_factor,  "AICPA Table A-1",                    False, "0.00"),
            ("Final sample size",    size.final_sample_size,  "⌈Base × CF⌉",                       False),
            ("표본간격 (J)",         size.sample_interval,    "잔여모집단 ÷ Final size",             False),
        ]
        for row_item in params:
            label    = row_item[0]
            value    = row_item[1]
            formula  = row_item[2] if len(row_item) > 2 else ""
            is_input = row_item[3] if len(row_item) > 3 else False
            fmt      = row_item[4] if len(row_item) > 4 else NUMFMT_INT

            _text_cell(ws, r, 1, label, font=FONT_BOLD if is_input else FONT_BODY)
            if is_input:
                _accent_num_cell(ws, r, 2, value, fmt=fmt)
            else:
                _num_cell(ws, r, 2, value, fmt=fmt)

            c_formula = ws.cell(r, 3, formula)
            _apply(c_formula, font=FONT_FADED, fill=FILL_WHITE, border=BORDER_LIGHT,
                   alignment=_al("left"))
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
            r += 1

        r += 1


# ─────────────────────────────────────────────────────────────
# Sheet 4: 모집단 완전성
# ─────────────────────────────────────────────────────────────

def _build_sheet_completeness(
    wb,
    ctx: ReportContext,
    ar_kd: KindData | None,
    ap_kd: KindData | None,
) -> None:
    ws = wb.create_sheet("모집단 완전성")

    for c, w in [(1, 4), (2, 28), (3, 18), (4, 18), (5, 15), (6, 30)]:
        _set_col_width(ws, c, w)

    prep_date   = ctx.prep_date   or date.today()
    review_date = ctx.review_date or date.today()

    _write_doc_header(
        ws, ctx.company_name, "모집단 완전성 검토",
        ctx.period_end, ctx.preparer, ctx.reviewer,
        prep_date, review_date, wp_no="완전성",
        last_col=6,
    )

    r = 5

    for kd, kind_label in [(ar_kd, "채권 (C100)"), (ap_kd, "채무 (AA100)")]:
        if kd is None:
            continue

        _subheader_row(ws, r, f"{kind_label} — 회사 명세서 vs 재무제표", col_end=6)
        r += 1

        for col_idx, h in enumerate(["No", "계정과목그룹", "회사 명세서", "재무제표", "차이", "비고"], 1):
            _header_cell(ws, r, col_idx, h)
        r += 1

        for i, row_data in enumerate(kd.completeness.by_group, 1):
            diff = row_data["diff"]
            _text_cell(ws, r, 1, i, align="center")
            _text_cell(ws, r, 2, row_data["group"])
            _num_cell(ws, r, 3, row_data["ledger"])
            _num_cell(ws, r, 4, row_data["fs"])
            diff_font = FONT_RED if abs(diff) > 0 else FONT_BODY
            c_diff = ws.cell(r, 5, diff)
            _apply(c_diff, font=diff_font, fill=FILL_WHITE, border=BORDER_LIGHT,
                   alignment=_al("right"), number_format=NUMFMT_INT)
            _row_height(ws, r, 22)
            c_note = ws.cell(r, 6, row_data.get("note", "") or "")
            _apply(c_note, font=FONT_BODY, fill=FILL_WHITE, border=BORDER_LIGHT,
                   alignment=_al("left"))
            r += 1

        # 합계 행
        _total_row_cell(ws, r, 2, "합계")
        _total_row_cell(ws, r, 3, kd.completeness.total_ledger, is_num=True)
        _total_row_cell(ws, r, 4, kd.completeness.total_fs,     is_num=True)
        diff_total = kd.completeness.total_diff
        c_tot = ws.cell(r, 5, diff_total)
        _apply(c_tot, font=FONT_RED if abs(diff_total) > 0 else FONT_BOLD,
               fill=FILL_BG_SUB, border=BORDER_TOTAL,
               alignment=_al("right"), number_format=NUMFMT_INT)
        _row_height(ws, r, 22)
        _total_row_cell(ws, r, 6, "")
        r += 2

        # 발송제외 거래처
        _subheader_row(ws, r, f"{kind_label} — 발송제외 거래처", col_end=6)
        r += 1

        for col_idx, h in enumerate(["No", "거래처명", "장부가", "제외 사유", "", ""], 1):
            _header_cell(ws, r, col_idx, h)
        r += 1

        excl_list = [d for d in kd.decisions if d.is_excluded]
        extra_excl = [ex for ex in (kd.exclusion_rows or [])
                      if ex.name not in {d.name for d in excl_list}]

        if excl_list or extra_excl:
            for seq, d in enumerate(excl_list, 1):
                c = ws.cell(r, 1, seq)
                _apply(c, font=FONT_FADED, fill=FILL_WHITE,
                       border=_left_excl_border(), alignment=_al("center"))
                _row_height(ws, r, 22)
                c = ws.cell(r, 2, d.name)
                _apply(c, font=FONT_FADED_STRIKE, fill=FILL_WHITE,
                       border=BORDER_LIGHT, alignment=_al("left"))
                _num_cell(ws, r, 3, d.balance, font=FONT_FADED)
                c = ws.cell(r, 4, d.exclusion_reason or "")
                _apply(c, font=FONT_FADED, fill=FILL_WHITE, border=BORDER_LIGHT,
                       alignment=_al("left"))
                ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
                r += 1
            for seq, ex in enumerate(extra_excl, len(excl_list) + 1):
                c = ws.cell(r, 1, seq)
                _apply(c, font=FONT_FADED, fill=FILL_WHITE,
                       border=_left_excl_border(), alignment=_al("center"))
                _row_height(ws, r, 22)
                c = ws.cell(r, 2, ex.name)
                _apply(c, font=FONT_FADED_STRIKE, fill=FILL_WHITE,
                       border=BORDER_LIGHT, alignment=_al("left"))
                _num_cell(ws, r, 3, ex.amount, font=FONT_FADED)
                c = ws.cell(r, 4, "발송대상 제외")
                _apply(c, font=FONT_FADED, fill=FILL_WHITE, border=BORDER_LIGHT,
                       alignment=_al("left"))
                ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
                r += 1
        else:
            c = ws.cell(r, 2, "발송제외 거래처 없음")
            _apply(c, font=FONT_FADED, fill=FILL_WHITE, border=BORDER_LIGHT,
                   alignment=_al("left"))
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            r += 1

        r += 1


# ─────────────────────────────────────────────────────────────
# Sheet 5: Key item 매트릭스
# ─────────────────────────────────────────────────────────────

_AR_GROUP_COLS = {
    "외상매출금": 4, "받을어음": 5, "미수금": 6, "선급금": 7,
    "장기대여금": 8, "임차보증금": 9,
}
_AP_GROUP_COLS = {
    "외상매입금": 10, "지급어음(외담대외상매입금)": 11,
    "미지급금": 12, "임대보증금": 13,
}


def _build_sheet_key_item_matrix(
    wb,
    ctx: ReportContext,
    ar_kd: KindData | None,
    ap_kd: KindData | None,
    pm: float,
) -> None:
    ws = wb.create_sheet("Key item 매트릭스")

    for c, w in [(1, 4), (2, 28), (3, 10)]:
        _set_col_width(ws, c, w)
    for c in list(_AR_GROUP_COLS.values()) + list(_AP_GROUP_COLS.values()):
        _set_col_width(ws, c, 13)
    sum_col = max(_AP_GROUP_COLS.values()) + 1
    for c, w in [(sum_col, 13), (sum_col + 1, 8), (sum_col + 2, 8),
                 (sum_col + 3, 8), (sum_col + 4, 8)]:
        _set_col_width(ws, c, w)

    last_col = sum_col + 4
    prep_date   = ctx.prep_date   or date.today()
    review_date = ctx.review_date or date.today()

    _write_doc_header(
        ws, ctx.company_name, "Key item 매트릭스 (채권·채무 통합)",
        ctx.period_end, ctx.preparer, ctx.reviewer,
        prep_date, review_date, wp_no="KI매트릭스",
        last_col=last_col,
    )

    r = 5

    # Key item 기준금액 표시
    if ar_kd:
        kd = ar_kd
        c = ws.cell(r, 1, f"채권 Key item 기준금액: {kd.size_result.key_item_threshold:,.0f} 원")
        _apply(c, font=FONT_ACCENT, fill=FILL_WHITE, border=BORDER_LIGHT,
               alignment=_al("left"))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        _row_height(ws, r, 22)
        r += 1
    if ap_kd:
        kd = ap_kd
        c = ws.cell(r, 1, f"채무 Key item 기준금액: {kd.size_result.key_item_threshold:,.0f} 원")
        _apply(c, font=FONT_ACCENT, fill=FILL_WHITE, border=BORDER_LIGHT,
               alignment=_al("left"))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        _row_height(ws, r, 22)
        r += 1

    r += 1
    _subheader_row(ws, r, "거래처별 계정과목 매트릭스", col_end=last_col)
    r += 1

    # 헤더 행 — 채권/채무 계정 그룹화 표시
    _header_cell(ws, r, 1, "No")
    _header_cell(ws, r, 2, "거래처명")
    _header_cell(ws, r, 3, "구분")
    for acct, col in _AR_GROUP_COLS.items():
        _header_cell(ws, r, col, acct)
    for acct, col in _AP_GROUP_COLS.items():
        _header_cell(ws, r, col, acct)
    _header_cell(ws, r, sum_col,     "합계")
    _header_cell(ws, r, sum_col + 1, "Key")
    _header_cell(ws, r, sum_col + 2, "Rep")
    _header_cell(ws, r, sum_col + 3, "특관")
    _header_cell(ws, r, sum_col + 4, "최종")
    r += 1

    # 모든 거래처 (채권 + 채무 통합)
    all_parties_map: dict[str, tuple[PartyDecision | None, PartyDecision | None]] = {}

    if ar_kd:
        for d in ar_kd.decisions:
            if d.name not in all_parties_map:
                all_parties_map[d.name] = (d, None)
            else:
                all_parties_map[d.name] = (d, all_parties_map[d.name][1])
    if ap_kd:
        for d in ap_kd.decisions:
            if d.name not in all_parties_map:
                all_parties_map[d.name] = (None, d)
            else:
                all_parties_map[d.name] = (all_parties_map[d.name][0], d)

    parties_sorted = sorted(
        all_parties_map.items(),
        key=lambda x: -(
            (x[1][0].balance if x[1][0] else 0) +
            (x[1][1].balance if x[1][1] else 0)
        ),
    )

    totals_ar_cols: dict[str, float] = {g: 0.0 for g in _AR_GROUP_COLS}
    totals_ap_cols: dict[str, float] = {g: 0.0 for g in _AP_GROUP_COLS}
    grand_total = 0.0

    for seq, (name, (ar_d, ap_d)) in enumerate(parties_sorted, 1):
        is_ki  = bool((ar_d and ar_d.is_key_item) or (ap_d and ap_d.is_key_item))
        is_rep = bool((ar_d and ar_d.is_representative) or (ap_d and ap_d.is_representative))
        is_rp  = bool((ar_d and ar_d.is_related_party) or (ap_d and ap_d.is_related_party))
        is_fin = bool((ar_d and ar_d.final_sampled) or (ap_d and ap_d.final_sampled))
        is_excl = bool((ar_d and ar_d.is_excluded) and (ap_d is None or ap_d.is_excluded))

        kind_str = (
            "채권+채무" if ar_d and ap_d else
            "채권" if ar_d else "채무"
        )

        total = (
            (ar_d.balance if ar_d and not ar_d.is_excluded else 0) +
            (ap_d.balance if ap_d and not ap_d.is_excluded else 0)
        )

        c = ws.cell(r, 1, seq)
        b = _left_ki_border() if is_ki else _left_rep_border() if is_rep else BORDER_LIGHT
        _apply(c, font=FONT_FADED if is_excl else FONT_BODY,
               fill=FILL_WHITE, border=b, alignment=_al("center"))
        _row_height(ws, r, 22)

        c = ws.cell(r, 2, name)
        _apply(c, font=FONT_FADED_STRIKE if is_excl else (FONT_BOLD if is_ki else FONT_BODY),
               fill=FILL_WHITE, border=BORDER_LIGHT, alignment=_al("left"))

        c = ws.cell(r, 3, kind_str)
        _apply(c, font=FONT_ACCENT if kind_str == "채권+채무" else FONT_BODY,
               fill=FILL_WHITE, border=BORDER_LIGHT, alignment=_al("center"))

        for acct, col in _AR_GROUP_COLS.items():
            amt = (ar_d.by_account.get(acct, 0.0) if ar_d else 0.0)
            c = ws.cell(r, col, amt if amt else None)
            _apply(c, font=FONT_FADED if is_excl else FONT_BODY,
                   fill=FILL_WHITE, border=BORDER_LIGHT,
                   alignment=_al("right"), number_format=NUMFMT_INT)
            if not is_excl:
                totals_ar_cols[acct] = totals_ar_cols.get(acct, 0.0) + amt

        for acct, col in _AP_GROUP_COLS.items():
            amt = (ap_d.by_account.get(acct, 0.0) if ap_d else 0.0)
            c = ws.cell(r, col, amt if amt else None)
            _apply(c, font=FONT_FADED if is_excl else FONT_BODY,
                   fill=FILL_WHITE, border=BORDER_LIGHT,
                   alignment=_al("right"), number_format=NUMFMT_INT)
            if not is_excl:
                totals_ap_cols[acct] = totals_ap_cols.get(acct, 0.0) + amt

        c = ws.cell(r, sum_col, total if total else None)
        _apply(c, font=FONT_FADED if is_excl else (FONT_ACCENT if is_ki else FONT_BODY),
               fill=FILL_WHITE, border=BORDER_LIGHT,
               alignment=_al("right"), number_format=NUMFMT_INT)
        grand_total += total

        for col_offset, flag, label_y, label_n in [
            (1, is_ki,  "Y", "N"),
            (2, is_rep, "Y", "N"),
            (3, is_rp,  "Y", "N"),
            (4, is_fin, "Y", "N"),
        ]:
            c = ws.cell(r, sum_col + col_offset, label_y if flag else label_n)
            f = FONT_ACCENT if flag else FONT_FADED
            _apply(c, font=f, fill=FILL_WHITE, border=BORDER_LIGHT,
                   alignment=_al("center"))
        r += 1

    # 합계 행
    _total_row_cell(ws, r, 2, "합계")
    for acct, col in _AR_GROUP_COLS.items():
        _total_row_cell(ws, r, col, totals_ar_cols.get(acct) or None, is_num=True)
    for acct, col in _AP_GROUP_COLS.items():
        _total_row_cell(ws, r, col, totals_ap_cols.get(acct) or None, is_num=True)
    _total_row_cell(ws, r, sum_col, grand_total or None, is_num=True)
    all_d = list(all_parties_map.values())
    _total_row_cell(ws, r, sum_col + 1, sum(1 for a, b in all_d if (a and a.is_key_item) or (b and b.is_key_item)), is_num=True)
    _total_row_cell(ws, r, sum_col + 2, sum(1 for a, b in all_d if (a and a.is_representative) or (b and b.is_representative)), is_num=True)
    _total_row_cell(ws, r, sum_col + 3, sum(1 for a, b in all_d if (a and a.is_related_party) or (b and b.is_related_party)), is_num=True)
    _total_row_cell(ws, r, sum_col + 4, sum(1 for a, b in all_d if (a and a.final_sampled) or (b and b.final_sampled)), is_num=True)


# ─────────────────────────────────────────────────────────────
# Sheet 6: MUS 추출 내역
# ─────────────────────────────────────────────────────────────

def _build_sheet_mus_detail(
    wb,
    ctx: ReportContext,
    ar_kd: KindData | None,
    ap_kd: KindData | None,
) -> None:
    ws = wb.create_sheet("MUS 추출 내역")

    for c, w in [(1, 4), (2, 28), (3, 16), (4, 16), (5, 10), (6, 16), (7, 16), (8, 8)]:
        _set_col_width(ws, c, w)

    prep_date   = ctx.prep_date   or date.today()
    review_date = ctx.review_date or date.today()

    _write_doc_header(
        ws, ctx.company_name, "MUS 추출 내역",
        ctx.period_end, ctx.preparer, ctx.reviewer,
        prep_date, review_date, wp_no="MUS",
        last_col=8,
    )

    r = 5
    method = (
        "임의출발점(r₀) ≤ J 범위 내 난수, 이후 J마다 화폐단위 선택. "
        "누적금액이 r₀ + k×J를 처음 초과하는 거래처 = hit. (감사기준서 530)"
    )
    c_cell = ws.cell(r, 1, method)
    _apply(c_cell, font=FONT_FADED, fill=FILL_WHITE, border=BORDER_LIGHT,
           alignment=_al("left", wrap=True))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    _row_height(ws, r, 36)
    r += 2

    for kd, kind_label in [(ar_kd, "채권 (C100)"), (ap_kd, "채무 (AA100)")]:
        if kd is None:
            continue

        is_ap = kd.ctx.kind == "payable"
        _subheader_row(ws, r, kind_label, col_end=8)
        r += 1

        # 채무: 당기 활동량 기준 sampling 명시 (ISA 505 완전성 검토)
        if is_ap:
            note_text = (
                "채무 sampling 기준 = 당기 활동량 (|기초| + |증감|)  "
                "— ISA 505 under-statement risk: 기말 잔액 작아도 매입활동 크면 포함"
            )
            c_note = ws.cell(r, 1, note_text)
            _apply(c_note, font=FONT_FADED, fill=FILL_WHITE, border=BORDER_LIGHT,
                   alignment=_al("left", wrap=True))
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            _row_height(ws, r, 28)
            r += 1

        size = kd.size_result
        mus  = kd.mus_result

        # 파라미터 요약
        pop_label = "모집단 (당기 활동량)" if is_ap else "잔여 모집단"
        for label, val, fmt in [
            (pop_label, size.remaining_population, NUMFMT_INT),
            ("표본규모 (N)", size.final_sample_size,  NUMFMT_INT),
            ("표본간격 (J)", size.sample_interval,    NUMFMT_INT),
            ("임의출발점 (r₀)", mus.random_start,     NUMFMT_INT),
        ]:
            _text_cell(ws, r, 1, label, font=FONT_BOLD)
            _num_cell(ws, r, 2, val, fmt=fmt)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
            r += 1

        r += 1

        # 추출 내역 표 — 채무 측 컬럼은 "활동량" 표기
        balance_col_label = "활동량" if is_ap else "잔액"
        for col_idx, h in enumerate(
            ["No", "거래처명", balance_col_label, "누적금액", "선택횟수", "표본간격", "잔여", "hit"], 1
        ):
            _header_cell(ws, r, col_idx, h)
        r += 1

        hit_total = 0
        for i, sel in enumerate(mus.selections, 1):
            is_hit = sel.hit
            if is_hit:
                hit_total += 1
            border_l = _left_rep_border() if is_hit else BORDER_LIGHT
            font_v   = FONT_BOLD if is_hit else FONT_BODY

            c = ws.cell(r, 1, i)
            _apply(c, font=font_v, fill=FILL_WHITE, border=border_l, alignment=_al("center"))
            _row_height(ws, r, 22)
            c = ws.cell(r, 2, sel.name)
            hit_name_font = Font(name=FONT_NAME, size=10, bold=True, color=TOSS_GREEN) if is_hit else FONT_BODY
            _apply(c, font=hit_name_font, fill=FILL_WHITE, border=BORDER_LIGHT, alignment=_al("left"))
            _num_cell(ws, r, 3, sel.balance,    font=font_v)
            _num_cell(ws, r, 4, sel.cumulative, font=font_v)
            c = ws.cell(r, 5, sel.selections)
            _apply(c, font=font_v, fill=FILL_WHITE, border=BORDER_LIGHT, alignment=_al("center"))
            _num_cell(ws, r, 6, size.sample_interval, font=FONT_FADED)
            _num_cell(ws, r, 7, sel.remainder_after,  font=font_v)
            hit_str = "Y" if is_hit else "N"
            hit_font = FONT_GREEN if is_hit else FONT_FADED
            c = ws.cell(r, 8, hit_str)
            _apply(c, font=hit_font, fill=FILL_WHITE, border=BORDER_LIGHT, alignment=_al("center"))
            r += 1

        # 합계 행
        _total_row_cell(ws, r, 2, "합계")
        _total_row_cell(ws, r, 3, sum(s.balance for s in mus.selections), is_num=True)
        _total_row_cell(ws, r, 5, sum(s.selections for s in mus.selections), is_num=True)
        _total_row_cell(ws, r, 8, f"{hit_total}건")
        r += 2


# ─────────────────────────────────────────────────────────────
# Sheet 7: 주소 적정성
# ─────────────────────────────────────────────────────────────

def _build_sheet_address(
    wb,
    ctx: ReportContext,
    contacts: list[PartyContactInfo],
) -> None:
    ws = wb.create_sheet("주소 적정성")

    for c, w in [(1, 4), (2, 28), (3, 10), (4, 16), (5, 13),
                 (6, 13), (7, 16), (8, 22), (9, 10), (10, 20)]:
        _set_col_width(ws, c, w)

    prep_date   = ctx.prep_date   or date.today()
    review_date = ctx.review_date or date.today()

    _write_doc_header(
        ws, ctx.company_name, "C100A 조회처 주소 적정성 확인",
        ctx.period_end, ctx.preparer, ctx.reviewer,
        prep_date, review_date, wp_no="C100A",
        last_col=10,
    )

    r = 5
    _subheader_row(ws, r, "조회서 발송 주소·연락처 적정성 검토", col_end=10)
    r += 1

    headers = ["No", "거래처명", "국가", "사업자번호", "대표자명",
               "담당자명", "전화번호", "이메일", "적정성", "비고"]
    for i, h in enumerate(headers, 1):
        _header_cell(ws, r, i, h)
    r += 1

    if not contacts:
        c_cell = ws.cell(r, 2, "UploadGuide 미제공 — 주소 정보 없음")
        _apply(c_cell, font=FONT_FADED, fill=FILL_WHITE, border=BORDER_LIGHT,
               alignment=_al("left"))
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
        return

    # 중복 거래처 제거
    seen: set[str] = set()
    unique_contacts = []
    for ct in contacts:
        if ct.name not in seen:
            seen.add(ct.name)
            unique_contacts.append(ct)

    for seq, ct in enumerate(unique_contacts, 1):
        ok = bool(ct.email or ct.phone)
        ok_font = FONT_GREEN if ok else FONT_RED
        ok_str  = "Y" if ok else "N"

        _text_cell(ws, r, 1, seq,              align="center")
        _text_cell(ws, r, 2, ct.name)
        _text_cell(ws, r, 3, ct.country,       align="center")
        _text_cell(ws, r, 4, ct.business_no)
        _text_cell(ws, r, 5, ct.ceo_name)
        _text_cell(ws, r, 6, ct.contact_person)
        _text_cell(ws, r, 7, ct.phone)
        _text_cell(ws, r, 8, ct.email)
        c = ws.cell(r, 9, ok_str)
        _apply(c, font=ok_font, fill=FILL_WHITE, border=BORDER_LIGHT,
               alignment=_al("center"))
        _row_height(ws, r, 22)
        _text_cell(ws, r, 10, "", align="left")
        r += 1


# ─────────────────────────────────────────────────────────────
# Sheet 8: 회신 추적
# ─────────────────────────────────────────────────────────────

def _build_sheet_reply_tracking(
    wb,
    ctx: ReportContext,
    replies: list[ConfirmationReplyInfo],
    all_decisions: list[PartyDecision],
) -> None:
    ws = wb.create_sheet("회신 추적")

    for c, w in [(1, 4), (2, 28), (3, 16), (4, 16), (5, 16), (6, 12), (7, 14)]:
        _set_col_width(ws, c, w)

    prep_date   = ctx.prep_date   or date.today()
    review_date = ctx.review_date or date.today()

    _write_doc_header(
        ws, ctx.company_name, "조회서 회신 추적",
        ctx.period_end, ctx.preparer, ctx.reviewer,
        prep_date, review_date, wp_no="회신추적",
        last_col=7,
    )

    r = 5
    _subheader_row(ws, r, "PDF 회신 처리 결과", col_end=7)
    r += 1

    # 요약 통계
    if replies:
        total_sent = len([d for d in all_decisions if d.final_sampled and not d.is_excluded])
        replied    = [x for x in replies if x.status != "미회신"]
        matched    = [x for x in replies if x.status == "matched"]
        mismatch   = [x for x in replies if x.status == "mismatch"]
        no_reply   = [x for x in replies if x.status == "미회신"]

        summary_data = [
            ("발송 건수",  total_sent, FONT_BODY),
            ("회신 수령",  len(replied), FONT_GREEN),
            ("일치",       len(matched),  FONT_GREEN),
            ("불일치",     len(mismatch), FONT_RED),
            ("미회신",     len(no_reply), FONT_AMBER),
        ]
        for col_idx, (lbl, val, f) in enumerate(summary_data, 1):
            c = ws.cell(r, col_idx, lbl)
            _apply(c, font=FONT_FADED, fill=FILL_BG_SUB, border=BORDER_LIGHT,
                   alignment=_al("center"))
        r += 1
        for col_idx, (lbl, val, f) in enumerate(summary_data, 1):
            c = ws.cell(r, col_idx, val)
            _apply(c, font=f, fill=FILL_BG_SUB, border=BORDER_LIGHT,
                   alignment=_al("center"), number_format=NUMFMT_INT)
        _row_height(ws, r, 28)
        r += 2

    # 상세 표
    for col_idx, h in enumerate(["No", "거래처명", "회신 금액", "장부가", "차이", "상태", "회신일자"], 1):
        _header_cell(ws, r, col_idx, h)
    r += 1

    final_map = {d.name: d.balance for d in all_decisions if d.final_sampled and not d.is_excluded}

    if not replies:
        c = ws.cell(r, 2, "Step 4 미완료 — 회신 데이터 없음")
        _apply(c, font=FONT_FADED, fill=FILL_WHITE, border=BORDER_LIGHT, alignment=_al("left"))
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        return

    for seq, rep in enumerate(replies, 1):
        status_label = {
            "matched": "일치", "mismatch": "불일치",
            "needs_review": "검토필요", "미회신": "미회신",
        }.get(rep.status, rep.status)

        status_font = (
            FONT_GREEN  if status_label == "일치"    else
            FONT_RED    if status_label == "불일치"   else
            FONT_AMBER  if status_label == "검토필요" else
            FONT_FADED
        )

        ledger_bal  = final_map.get(rep.party_name, 0.0)
        reply_amt   = rep.extracted_balance
        diff        = (reply_amt - ledger_bal) if reply_amt is not None else None

        _text_cell(ws, r, 1, seq, align="center")
        _text_cell(ws, r, 2, rep.party_name)
        c = ws.cell(r, 3, reply_amt)
        _apply(c, font=FONT_BODY, fill=FILL_WHITE, border=BORDER_LIGHT,
               alignment=_al("right"), number_format=NUMFMT_INT)
        _row_height(ws, r, 22)
        _num_cell(ws, r, 4, ledger_bal)
        c = ws.cell(r, 5, diff)
        diff_font = FONT_RED if diff and abs(diff) > 0 else FONT_BODY
        _apply(c, font=diff_font, fill=FILL_WHITE, border=BORDER_LIGHT,
               alignment=_al("right"), number_format=NUMFMT_INT)
        c = ws.cell(r, 6, status_label)
        _apply(c, font=status_font, fill=FILL_WHITE, border=BORDER_LIGHT,
               alignment=_al("center"))
        c = ws.cell(r, 7, rep.reply_date or "")
        _apply(c, font=FONT_BODY, fill=FILL_WHITE, border=BORDER_LIGHT,
               alignment=_al("center"))
        r += 1


# ─────────────────────────────────────────────────────────────
# Sheet 3: 조회서 (UploadGuide 양식 C1~C15 + 회신 컬럼 C16~C21)
# ─────────────────────────────────────────────────────────────

def _normalize_for_match(name: str) -> str:
    """거래처명 정규화 — 공백·법인 접미사·특수문자 제거, 소문자."""
    import re
    import unicodedata
    name = unicodedata.normalize("NFKC", name)
    # 법인 접미사 제거
    name = re.sub(
        r"유한회사|주식회사|\(주\)|㈜|Co\.,\s*Ltd\.?|Ltd\.?|Inc\.?|有限公司|株式会社",
        "", name, flags=re.IGNORECASE,
    )
    # 공백·특수문자 제거
    name = re.sub(r"[\s\(\)\[\]\.\,\-\_]+", "", name)
    return name.lower()


def _build_sheet_uploadguide_confirmation(
    wb,
    ctx: ReportContext,
    upload_guide_data: UploadGuideData | None,
    replies: list[ConfirmationReplyInfo],
    alt_procedures: list[AlternativeProcedureEntry],
) -> None:
    """조회서 시트 — UploadGuide Sheet1 행 그대로(C1~C15) + 회신 상태 컬럼(C16~C21).

    데이터 출처:
      C1~C15: UploadGuide send_targets의 계정과목별 원본 행
      C16: 회신상태 — PDF 회신 매칭 → "원본", 대체적절차 있음 → "대체적", 없음 → "미회신"
      C17: 회신금액, C18: 차이, C19: 일치여부, C20: 회신일자, C21: 비고
    """
    ws = wb.create_sheet("조회서")

    # 컬럼 너비 설정
    col_widths = {
        1: 8,   # C1: 채권채무구분
        2: 18,  # C2: 계정과목명
        3: 8,   # C3: 통화
        4: 15,  # C4: 조회금액
        5: 8,   # C5: 통화2
        6: 15,  # C6: 조회금액2
        7: 28,  # C7: 거래처명
        8: 8,   # C8: 국가
        9: 10,  # C9: 거래처 구분
        10: 14, # C10: 사업자번호
        11: 10, # C11: 대표자명
        12: 12, # C12: 담당자명
        13: 18, # C13: 담당자 전화
        14: 24, # C14: 담당자 이메일
        15: 12, # C15: 필수항목
        16: 10, # C16: 회신상태
        17: 14, # C17: 회신금액
        18: 14, # C18: 차이
        19: 10, # C19: 일치여부
        20: 12, # C20: 회신일자
        21: 20, # C21: 비고
    }
    for col, width in col_widths.items():
        _set_col_width(ws, col, width)

    prep_date   = ctx.prep_date   or date.today()
    review_date = ctx.review_date or date.today()

    _write_doc_header(
        ws, ctx.company_name, "조회서 — UploadGuide 발송 현황 및 회신 결과",
        ctx.period_end, ctx.preparer, ctx.reviewer,
        prep_date, review_date, wp_no="조회서",
        last_col=21,
    )

    r = 5
    _subheader_row(ws, r, "UploadGuide 발송 내역 + 회신 상태", col_end=21)
    r += 1

    # 헤더 행
    headers = [
        "채권채무구분", "계정과목명", "통화", "조회금액", "통화2", "조회금액2",
        "거래처명", "국가", "거래처구분", "사업자번호", "대표자명",
        "담당자명", "전화번호", "이메일", "필수항목",
        "회신상태", "회신금액", "차이", "일치여부", "회신일자", "비고",
    ]
    for i, h in enumerate(headers, 1):
        _header_cell(ws, r, i, h)
    r += 1

    if upload_guide_data is None or not upload_guide_data.send_targets:
        c = ws.cell(r, 1, "UploadGuide 미제공 — 조회서 데이터 없음")
        _apply(c, font=FONT_FADED, fill=FILL_WHITE, border=BORDER_LIGHT,
               alignment=_al("left"))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=21)
        return

    # 회신 매칭 인덱스 구축 (party_name_matched 기준)
    # key: (normalize(party), normalize(account)) → ConfirmationReplyInfo
    reply_index: dict[tuple[str, str], ConfirmationReplyInfo] = {}
    for rep in replies:
        pnorm = _normalize_for_match(rep.party_name)
        # ConfirmationReplyInfo 에는 계정과목 정보가 없으므로 party_name만으로 매칭
        # 동일 거래처 여러 건이면 마지막 우선 (matched > mismatch > needs_review > 미회신)
        status_priority = {"matched": 4, "mismatch": 3, "needs_review": 2, "미회신": 1}
        existing = reply_index.get((pnorm, ""))
        if existing is None or (
            status_priority.get(rep.status, 0) > status_priority.get(existing.status, 0)
        ):
            reply_index[(pnorm, "")] = rep

    # 대체적 절차 인덱스 구축
    alt_index: dict[str, AlternativeProcedureEntry] = {}
    for proc in alt_procedures:
        alt_index[_normalize_for_match(proc.party_name)] = proc

    # 각 행 작성
    status_label_map = {
        "matched": "원본", "mismatch": "원본", "needs_review": "원본", "미회신": "미회신",
    }
    match_label_map = {
        "matched": "일치", "mismatch": "불일치", "needs_review": "검토필요",
    }

    for ct in upload_guide_data.send_targets:
        # 계정과목별 행 (accounts 리스트)
        accounts = ct.accounts if ct.accounts else [("", "KRW", 0.0)]

        for acct_name, currency, amount in accounts:
            pnorm = _normalize_for_match(ct.name)
            rep = reply_index.get((pnorm, ""))

            # 회신 상태 결정
            if rep and rep.status in ("matched", "mismatch", "needs_review"):
                reply_status = "원본"
                reply_amount = rep.extracted_balance
                diff_val = (
                    (reply_amount - amount) if (reply_amount is not None and amount) else None
                )
                match_label = match_label_map.get(rep.status, "검토필요")
                reply_date_val = rep.reply_date or ""
                note_val = ""
                reply_font = (
                    FONT_GREEN  if rep.status == "matched"  else
                    FONT_RED    if rep.status == "mismatch" else
                    FONT_AMBER
                )
            elif pnorm in alt_index:
                proc = alt_index[pnorm]
                reply_status = "대체적"
                reply_amount = proc.covered_amount
                diff_val = None
                match_label = ""
                reply_date_val = ""
                note_val = proc.conclusion or ""
                reply_font = FONT_AMBER
            else:
                reply_status = "미회신"
                reply_amount = None
                diff_val = None
                match_label = ""
                reply_date_val = ""
                note_val = ""
                reply_font = FONT_FADED

            # UploadGuide 원본 컬럼 (C1~C15)
            _text_cell(ws, r, 1, "채권" if acct_name in _AR_ACCOUNTS else ("채무" if acct_name in _AP_ACCOUNTS else ""), align="center")
            _text_cell(ws, r, 2, acct_name)
            _text_cell(ws, r, 3, currency, align="center")
            if amount:
                _num_cell(ws, r, 4, amount)
            else:
                _text_cell(ws, r, 4, "")
            _text_cell(ws, r, 5, "", align="center")  # 통화2
            _text_cell(ws, r, 6, "")                  # 조회금액2
            _text_cell(ws, r, 7, ct.name)
            _text_cell(ws, r, 8, ct.country or "국내", align="center")
            _text_cell(ws, r, 9, "사업자", align="center")
            _text_cell(ws, r, 10, ct.business_no)
            _text_cell(ws, r, 11, ct.ceo_name)
            _text_cell(ws, r, 12, ct.contact_person)
            _text_cell(ws, r, 13, ct.phone)
            _text_cell(ws, r, 14, ct.email)
            # C15: 필수항목 기재여부
            has_required = bool(ct.email or ct.phone)
            _text_cell(ws, r, 15, "완료" if has_required else "미완료",
                       font=FONT_GREEN if has_required else FONT_RED, align="center")

            # 회신 컬럼 (C16~C21)
            # C16: 회신상태
            c16 = ws.cell(r, 16, reply_status)
            _apply(c16, font=reply_font, fill=FILL_WHITE, border=BORDER_LIGHT,
                   alignment=_al("center"))
            _row_height(ws, r, 22)

            # C17: 회신금액
            if reply_amount is not None:
                _num_cell(ws, r, 17, reply_amount)
            else:
                _text_cell(ws, r, 17, "")

            # C18: 차이
            if diff_val is not None:
                c18 = ws.cell(r, 18, diff_val)
                diff_font = FONT_RED if abs(diff_val) > 0 else FONT_BODY
                _apply(c18, font=diff_font, fill=FILL_WHITE, border=BORDER_LIGHT,
                       alignment=_al("right"), number_format=NUMFMT_INT)
            else:
                _text_cell(ws, r, 18, "")

            # C19: 일치여부
            if match_label:
                match_font = (
                    FONT_GREEN if match_label == "일치"    else
                    FONT_RED   if match_label == "불일치"  else
                    FONT_AMBER
                )
                c19 = ws.cell(r, 19, match_label)
                _apply(c19, font=match_font, fill=FILL_WHITE, border=BORDER_LIGHT,
                       alignment=_al("center"))
            else:
                _text_cell(ws, r, 19, "")

            # C20: 회신일자
            _text_cell(ws, r, 20, reply_date_val, align="center")

            # C21: 비고
            _text_cell(ws, r, 21, note_val)

            r += 1

    # 집계 요약
    r += 1
    _subheader_row(ws, r, "회신 현황 요약", col_end=21)
    r += 1
    total_rows = len(upload_guide_data.send_targets)
    replied = sum(
        1 for ct in upload_guide_data.send_targets
        if _normalize_for_match(ct.name) in {_normalize_for_match(rep.party_name)
                                               for rep in replies
                                               if rep.status in ("matched","mismatch","needs_review")}
    )
    alt_covered = sum(
        1 for ct in upload_guide_data.send_targets
        if _normalize_for_match(ct.name) in alt_index
        and _normalize_for_match(ct.name) not in {
            _normalize_for_match(rep.party_name)
            for rep in replies if rep.status in ("matched","mismatch","needs_review")
        }
    )
    no_reply = total_rows - replied - alt_covered

    for label, val, font in [
        ("발송 거래처 수", total_rows, FONT_BODY),
        ("원본 회신",      replied,    FONT_GREEN),
        ("대체적 절차",    alt_covered, FONT_AMBER),
        ("미회신",         no_reply,   FONT_RED if no_reply > 0 else FONT_GREEN),
    ]:
        _text_cell(ws, r, 1, label, font=FONT_BOLD)
        c = ws.cell(r, 2, val)
        _apply(c, font=font, fill=FILL_WHITE, border=BORDER_LIGHT,
               alignment=_al("right"), number_format=NUMFMT_INT)
        _row_height(ws, r, 22)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=21)
        r += 1


# ─────────────────────────────────────────────────────────────
# Sheet 10: 대체적 절차
# ─────────────────────────────────────────────────────────────

def _build_sheet_alt_procedures(
    wb,
    ctx: ReportContext,
    procedures: list[AlternativeProcedureEntry],
) -> None:
    ws = wb.create_sheet("대체적 절차")

    for c, w in [(1, 4), (2, 28), (3, 10), (4, 16), (5, 14),
                 (6, 22), (7, 14), (8, 10), (9, 10), (10, 22)]:
        _set_col_width(ws, c, w)

    prep_date   = ctx.prep_date   or date.today()
    review_date = ctx.review_date or date.today()

    _write_doc_header(
        ws, ctx.company_name, "대체적 절차 — 미회신·불일치 거래처",
        ctx.period_end, ctx.preparer, ctx.reviewer,
        prep_date, review_date, wp_no="대체",
        last_col=10,
    )

    r = 5
    _subheader_row(ws, r, "대체적 절차 수행 내역", col_end=10)
    r += 1

    headers = ["No", "거래처명", "사유", "장부가", "절차유형",
               "증빙명세", "커버금액", "커버리지%", "결론", "감사인 메모"]
    for i, h in enumerate(headers, 1):
        _header_cell(ws, r, i, h)
    r += 1

    if not procedures:
        c = ws.cell(r, 2, "대체적 절차 대상 없음 (Step 5 미완료 또는 전원 일치)")
        _apply(c, font=FONT_FADED, fill=FILL_WHITE, border=BORDER_LIGHT,
               alignment=_al("left"))
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
        r += 1
    else:
        for i, proc in enumerate(procedures, 1):
            conclusion_font = {
                "충분":  FONT_GREEN,
                "부분":  FONT_AMBER,
                "미해소": FONT_RED,
            }.get(proc.conclusion, FONT_BODY)

            _text_cell(ws, r, 1, i, align="center")
            _text_cell(ws, r, 2, proc.party_name)
            _text_cell(ws, r, 3, proc.reason, align="center")
            _num_cell(ws, r, 4, proc.ledger_balance)
            _text_cell(ws, r, 5, proc.procedure_type, align="center")
            _text_cell(ws, r, 6, "; ".join(proc.evidence_names) if proc.evidence_names else "")
            _num_cell(ws, r, 7, proc.covered_amount)
            _pct_cell(ws, r, 8, proc.coverage_ratio)
            c = ws.cell(r, 9, proc.conclusion)
            _apply(c, font=conclusion_font, fill=FILL_WHITE, border=BORDER_LIGHT,
                   alignment=_al("center"))
            _row_height(ws, r, 22)
            _text_cell(ws, r, 10, proc.auditor_notes or "")
            r += 1

    # 집계 요약
    r += 1
    _subheader_row(ws, r, "집계 요약", col_end=10)
    r += 1

    if procedures:
        sufficient = sum(1 for p in procedures if p.conclusion == "충분")
        partial    = sum(1 for p in procedures if p.conclusion == "부분")
        unresolved = sum(1 for p in procedures if p.conclusion == "미해소")
        total_covered = sum(p.covered_amount or 0 for p in procedures)
        total_ledger  = sum(p.ledger_balance or 0 for p in procedures)
        overall_ratio = total_covered / total_ledger if total_ledger > 0 else None

        for label, val, fmt, f in [
            ("충분 건수",   sufficient,    NUMFMT_INT, FONT_GREEN),
            ("부분 건수",   partial,       NUMFMT_INT, FONT_AMBER),
            ("미해소 건수", unresolved,    NUMFMT_INT, FONT_RED),
            ("장부가 합계", total_ledger,  NUMFMT_INT, FONT_BODY),
            ("커버금액 합계", total_covered, NUMFMT_INT, FONT_BODY),
        ]:
            _text_cell(ws, r, 1, label, font=FONT_BOLD)
            c = ws.cell(r, 2, val)
            _apply(c, font=f, fill=FILL_WHITE, border=BORDER_LIGHT,
                   alignment=_al("right"), number_format=fmt)
            _row_height(ws, r, 22)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
            r += 1

        if overall_ratio is not None:
            _text_cell(ws, r, 1, "전체 커버리지", font=FONT_BOLD)
            cov_font = (
                FONT_GREEN if overall_ratio >= 0.95 else
                FONT_AMBER if overall_ratio >= 0.50 else
                FONT_RED
            )
            _pct_cell(ws, r, 2, overall_ratio, font=cov_font)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
