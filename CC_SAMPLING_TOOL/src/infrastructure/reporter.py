"""
조서 Excel 출력 — C100-1/2/3 + AA100 시리즈 형식 재현

시트:
  1) C100 control sheet — 최종 샘플링 거래처 + 회신 추적
  2) C100-1 표본규모 결정
  3) C100-2 Key item 추출 (모집단 완전성 + Key item 표시)
  4) C100-3 표본 추출 (MUS)
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.domain.mus import MUSResult
from src.domain.population import CompletenessCheck, PartyDecision
from src.domain.sample_size import SampleSizeResult


# --- 공통 스타일 -------------------------------------------------------
_THIN = Side(style="thin", color="888888")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
KEY_ITEM_FILL = PatternFill("solid", fgColor="FFF2CC")
SAMPLED_FILL = PatternFill("solid", fgColor="C6EFCE")
EXCLUDED_FILL = PatternFill("solid", fgColor="F4CCCC")

HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="맑은 고딕", size=14, bold=True)
BODY_FONT = Font(name="맑은 고딕", size=10)


def _h(ws, row, col, text, fill=HEADER_FILL, font=HEADER_FONT, align="center"):
    c = ws.cell(row, col, text)
    c.font = font
    c.fill = fill
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = BORDER
    return c


def _b(ws, row, col, value, number_format=None, align="right"):
    c = ws.cell(row, col, value)
    c.font = BODY_FONT
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = BORDER
    if number_format:
        c.number_format = number_format
    return c


@dataclass
class ReportContext:
    company_name: str
    period_end: date
    kind: str                  # "receivable" | "payable"
    preparer: str = ""
    reviewer: str = ""
    prep_date: date | None = None
    review_date: date | None = None
    workpaper_no_prefix: str = "C100"   # 채권=C100, 채무=AA100


def build_report(
    out_path: str | Path,
    ctx: ReportContext,
    completeness: CompletenessCheck,
    size_result: SampleSizeResult,
    decisions: list[PartyDecision],
    mus_result: MUSResult,
    performance_materiality: float,
    population_amount: float,
    exclusion_notes: dict[str, str] | None = None,
) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _sheet_control(wb, ctx, decisions, mus_result)
    _sheet_size(wb, ctx, size_result, performance_materiality, population_amount, decisions)
    _sheet_key_item(wb, ctx, completeness, size_result, decisions, exclusion_notes or {})
    _sheet_mus(wb, ctx, size_result, mus_result, decisions)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# --- Sheet 1: Control --------------------------------------------------
def _sheet_control(wb, ctx, decisions, mus_result):
    title = "채권 조회서 control sheet" if ctx.kind == "receivable" else "채무 조회서 control sheet"
    ws = wb.create_sheet(f"{ctx.workpaper_no_prefix} control sheet")

    ws.cell(1, 1, ctx.company_name).font = TITLE_FONT
    ws.cell(2, 1, title).font = Font(bold=True, size=12)
    ws.cell(3, 1, f"기준일: {ctx.period_end}").font = BODY_FONT
    ws.cell(1, 6, f"작성자: {ctx.preparer}")
    ws.cell(2, 6, f"검토자: {ctx.reviewer}")
    ws.cell(1, 8, f"조서번호: {ctx.workpaper_no_prefix}")

    ws.cell(5, 1, "1. 조회방법: 적극적 조회").font = Font(bold=True)
    ws.cell(6, 1, "2. 최종 샘플링 거래처 (Key item + MUS 표본 + 특관자)").font = Font(bold=True)

    headers = ["No", "거래처명", "잔액", "구분", "특관자", "비고"]
    for i, h in enumerate(headers, start=2):
        _h(ws, 8, i, h)

    sampled = [d for d in decisions if d.final_sampled]
    sampled.sort(key=lambda d: -d.balance)
    for idx, d in enumerate(sampled, start=1):
        r = 8 + idx
        kind_label = []
        if d.is_key_item:
            kind_label.append("Key item")
        if d.is_representative:
            kind_label.append("Representative")
        if d.is_related_party and not (d.is_key_item or d.is_representative):
            kind_label.append("특관자")
        _b(ws, r, 2, idx, align="center")
        _b(ws, r, 3, d.name, align="left")
        _b(ws, r, 4, d.balance, number_format="#,##0")
        _b(ws, r, 5, " + ".join(kind_label), align="center")
        _b(ws, r, 6, "Y" if d.is_related_party else "", align="center")
        _b(ws, r, 7, "", align="left")
        if d.is_key_item:
            for col in range(2, 8):
                ws.cell(r, col).fill = KEY_ITEM_FILL
        elif d.is_representative:
            for col in range(2, 8):
                ws.cell(r, col).fill = SAMPLED_FILL

    total_row = 9 + len(sampled)
    _h(ws, total_row, 2, "합계", fill=SUBHEADER_FILL, font=Font(bold=True))
    _h(ws, total_row, 3, "", fill=SUBHEADER_FILL)
    _b(ws, total_row, 4, sum(d.balance for d in sampled), number_format="#,##0")

    _autosize(ws, [2, 4, 8, 18, 14, 10, 18])


# --- Sheet 2: 표본규모 ---------------------------------------------
def _sheet_size(wb, ctx, size, pm, pop, decisions):
    ws = wb.create_sheet(f"{ctx.workpaper_no_prefix}-1 표본규모 결정")
    ws.cell(1, 1, ctx.company_name).font = TITLE_FONT
    title = "매출채권 조회(1) - 표본규모 결정" if ctx.kind == "receivable" else "매입채무 조회(1) - 조회대상 결정"
    ws.cell(2, 1, title).font = Font(bold=True, size=12)
    ws.cell(3, 1, f"기준일: {ctx.period_end}").font = BODY_FONT
    ws.cell(1, 6, f"조서번호: {ctx.workpaper_no_prefix}-1")

    ws.cell(5, 1, "1. 감사목적").font = Font(bold=True)
    purpose = "보고기간말 현재 매출채권의 실재성 검토를 위한 적정한 표본 규모 결정" if ctx.kind == "receivable" \
        else "보고기간말 현재 매입채무의 완전성 검토를 위한 조회대상 결정"
    ws.cell(6, 2, purpose).font = BODY_FONT

    ws.cell(8, 1, "2. 조회대상 표본").font = Font(bold=True)
    headers = ["구분", "조서번호", "개수", "금액", "표본추출방법"]
    for i, h in enumerate(headers, start=3):
        _h(ws, 10, i, h)

    ki = [d for d in decisions if d.is_key_item]
    rep = [d for d in decisions if d.is_representative and not d.is_key_item]
    rp = [d for d in decisions if d.is_related_party and d.final_sampled and not (d.is_key_item or d.is_representative)]

    ki_amt = sum(d.balance for d in ki)
    rep_amt = sum(d.balance for d in rep)
    rp_amt = sum(d.balance for d in rp)

    _b(ws, 11, 2, "Key item", align="left")
    _b(ws, 11, 3, f"{ctx.workpaper_no_prefix}-2", align="center")
    _b(ws, 11, 4, len(ki), number_format="#,##0")
    _b(ws, 11, 5, ki_amt, number_format="#,##0")
    _b(ws, 11, 6, "전수", align="center")

    _b(ws, 12, 2, "Representative sample", align="left")
    _b(ws, 12, 3, f"{ctx.workpaper_no_prefix}-3", align="center")
    _b(ws, 12, 4, len(rep), number_format="#,##0")
    _b(ws, 12, 5, rep_amt, number_format="#,##0")
    _b(ws, 12, 6, "MUS 방법", align="center")

    if rp:
        _b(ws, 13, 2, "특관자 추가", align="left")
        _b(ws, 13, 4, len(rp), number_format="#,##0")
        _b(ws, 13, 5, rp_amt, number_format="#,##0")
        _b(ws, 13, 6, "전수", align="center")
        sum_row = 14
    else:
        sum_row = 13

    _b(ws, sum_row, 2, "합계", align="left")
    _b(ws, sum_row, 4, len(ki) + len(rep) + len(rp), number_format="#,##0")
    _b(ws, sum_row, 5, ki_amt + rep_amt + rp_amt, number_format="#,##0")

    _b(ws, sum_row + 1, 2, "모집단", align="left")
    _b(ws, sum_row + 1, 4, len(decisions), number_format="#,##0")
    _b(ws, sum_row + 1, 5, pop, number_format="#,##0")

    _b(ws, sum_row + 2, 2, "Coverage", align="left")
    cov_cnt = (len(ki) + len(rep) + len(rp)) / len(decisions) if decisions else 0
    cov_amt = (ki_amt + rep_amt + rp_amt) / pop if pop else 0
    _b(ws, sum_row + 2, 4, cov_cnt, number_format="0.0%")
    _b(ws, sum_row + 2, 5, cov_amt, number_format="0.0%")

    # 3. 표본규모 결정근거
    r = sum_row + 5
    ws.cell(r, 1, "3. 표본규모 결정근거").font = Font(bold=True)
    r += 1
    items = [
        ("모집단금액", pop, "#,##0"),
        ("수행중요성", pm, "#,##0"),
        ("Key item 비율", size.key_item_ratio, "0.0%"),
        ("Key item 기준금액", size.key_item_threshold, "#,##0"),
        ("Key item 금액 합계", ki_amt, "#,##0"),
        ("Key item 개수", len(ki), "#,##0"),
        ("잔여 모집단", size.remaining_population, "#,##0"),
        ("Base sample size", size.base_sample_size, "0.000"),
        ("Confidence factor", size.confidence_factor, "0.00"),
        ("Final sample size", size.final_sample_size, "#,##0"),
        ("표본간격", size.sample_interval, "#,##0"),
    ]
    for lbl, val, fmt in items:
        _b(ws, r, 2, lbl, align="left")
        _b(ws, r, 5, val, number_format=fmt)
        r += 1

    _autosize(ws, [2, 28, 14, 10, 22, 14, 18])


# --- Sheet 3: 모집단 완전성 + Key item ------------------------------
def _sheet_key_item(wb, ctx, comp, size, decisions, exclusion_notes):
    ws = wb.create_sheet(f"{ctx.workpaper_no_prefix}-2 Key item 추출")
    ws.cell(1, 1, ctx.company_name).font = TITLE_FONT
    title = "매출채권 조회(2) - Key item 추출" if ctx.kind == "receivable" else "매입채무 조회(2) - 조회대상 결정"
    ws.cell(2, 1, title).font = Font(bold=True, size=12)
    ws.cell(3, 1, f"기준일: {ctx.period_end}").font = BODY_FONT
    ws.cell(1, 6, f"조서번호: {ctx.workpaper_no_prefix}-2")

    ws.cell(5, 1, "1. 모집단 완전성 검토").font = Font(bold=True)
    headers = ["구분", "회사제시 명세서", "재무제표", "차이", "차이소명"]
    for i, h in enumerate(headers, start=2):
        _h(ws, 7, i, h)

    for idx, row in enumerate(comp.by_group, start=1):
        r = 7 + idx
        _b(ws, r, 2, row["group"], align="left")
        _b(ws, r, 3, row["ledger"], number_format="#,##0")
        _b(ws, r, 4, row["fs"], number_format="#,##0")
        _b(ws, r, 5, row["diff"], number_format="#,##0")
        _b(ws, r, 6, row["note"], align="left")

    total_r = 8 + len(comp.by_group)
    _h(ws, total_r, 2, "합계", fill=SUBHEADER_FILL, font=Font(bold=True))
    _b(ws, total_r, 3, comp.total_ledger, number_format="#,##0")
    _b(ws, total_r, 4, comp.total_fs, number_format="#,##0")
    _b(ws, total_r, 5, comp.total_diff, number_format="#,##0")

    # 발송제외
    r = total_r + 3
    ws.cell(r, 1, "2. 발송제외 검토").font = Font(bold=True)
    r += 1
    _h(ws, r, 2, "거래처명")
    _h(ws, r, 3, "잔액")
    _h(ws, r, 4, "제외사유")
    r += 1
    excluded = [d for d in decisions if d.is_excluded]
    if not excluded:
        _b(ws, r, 2, "(해당 없음)", align="left")
        r += 1
    else:
        for d in excluded:
            _b(ws, r, 2, d.name, align="left")
            _b(ws, r, 3, d.balance, number_format="#,##0")
            _b(ws, r, 4, d.exclusion_reason, align="left")
            for c in range(2, 5):
                ws.cell(r, c).fill = EXCLUDED_FILL
            r += 1

    # Key item 추출 기준
    r += 2
    ws.cell(r, 1, "3. Key item 추출 기준").font = Font(bold=True)
    r += 1
    _b(ws, r, 2, "Key item 비율 (PM 대비)", align="left")
    _b(ws, r, 5, size.key_item_ratio, number_format="0.0%")
    r += 1
    _b(ws, r, 2, "Key item 기준금액", align="left")
    _b(ws, r, 5, size.key_item_threshold, number_format="#,##0")

    # 거래처별 결과
    r += 3
    ws.cell(r, 1, "4. 거래처별 분류").font = Font(bold=True)
    r += 2
    headers2 = ["거래처명", "잔액", "발송제외", "특관자", "Key item", "Representative", "최종샘플링"]
    for i, h in enumerate(headers2, start=2):
        _h(ws, r, i, h)
    r += 1

    for d in decisions:
        _b(ws, r, 2, d.name, align="left")
        _b(ws, r, 3, d.balance, number_format="#,##0")
        _b(ws, r, 4, "Y" if d.is_excluded else "N", align="center")
        _b(ws, r, 5, "Y" if d.is_related_party else "N", align="center")
        _b(ws, r, 6, "Y" if d.is_key_item else "N", align="center")
        _b(ws, r, 7, "Y" if d.is_representative else "N", align="center")
        _b(ws, r, 8, "Y" if d.final_sampled else "N", align="center")
        if d.is_excluded:
            for c in range(2, 9):
                ws.cell(r, c).fill = EXCLUDED_FILL
        elif d.is_key_item:
            for c in range(2, 9):
                ws.cell(r, c).fill = KEY_ITEM_FILL
        elif d.final_sampled:
            for c in range(2, 9):
                ws.cell(r, c).fill = SAMPLED_FILL
        r += 1

    _autosize(ws, [2, 32, 18, 16, 16, 12, 12, 14, 14])


# --- Sheet 4: MUS 추출 ----------------------------------------------
def _sheet_mus(wb, ctx, size, mus, decisions):
    ws = wb.create_sheet(f"{ctx.workpaper_no_prefix}-3 표본 추출 MUS")
    ws.cell(1, 1, ctx.company_name).font = TITLE_FONT
    title = "매출채권 조회(3) - 표본 추출 (MUS방법)" if ctx.kind == "receivable" else "매입채무 조회(3) - 표본 추출 (MUS방법)"
    ws.cell(2, 1, title).font = Font(bold=True, size=12)
    ws.cell(3, 1, f"기준일: {ctx.period_end}").font = BODY_FONT
    ws.cell(1, 6, f"조서번호: {ctx.workpaper_no_prefix}-3")

    ws.cell(5, 1, "1. 표본추출방법: MUS (Monetary Unit Sampling)").font = Font(bold=True)

    ws.cell(7, 1, "2. 모수").font = Font(bold=True)
    _b(ws, 8, 2, "잔여 모집단", align="left")
    _b(ws, 8, 5, size.remaining_population, number_format="#,##0")
    _b(ws, 9, 2, "표본규모", align="left")
    _b(ws, 9, 5, size.final_sample_size, number_format="#,##0")
    _b(ws, 10, 2, "표본간격", align="left")
    _b(ws, 10, 5, size.sample_interval, number_format="#,##0")
    _b(ws, 11, 2, "임의출발점 r0", align="left")
    _b(ws, 11, 5, mus.random_start, number_format="#,##0")

    ws.cell(13, 1, "3. 추출 내역").font = Font(bold=True)
    headers = ["거래처명", "잔액", "누적금액", "# of selections", "표본간격", "Selection remainder", "Sample 추출?"]
    for i, h in enumerate(headers, start=2):
        _h(ws, 15, i, h)

    for idx, sel in enumerate(mus.selections, start=1):
        r = 15 + idx
        _b(ws, r, 2, sel.name, align="left")
        _b(ws, r, 3, sel.balance, number_format="#,##0")
        _b(ws, r, 4, sel.cumulative, number_format="#,##0")
        _b(ws, r, 5, sel.selections, align="center")
        _b(ws, r, 6, size.sample_interval, number_format="#,##0")
        _b(ws, r, 7, sel.remainder_after, number_format="#,##0")
        _b(ws, r, 8, "Y" if sel.hit else "N", align="center")
        if sel.hit:
            for c in range(2, 9):
                ws.cell(r, c).fill = SAMPLED_FILL

    _autosize(ws, [2, 32, 14, 14, 14, 14, 16, 12])


def _autosize(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
