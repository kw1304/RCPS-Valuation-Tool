"""
템플릿 기반 조서 출력 — 원본 7620 조서와 동일 서식

핵심:
  - templates/cc_template.xlsx 를 베이스로 사용
  - 셀 서식(border, font, fill, alignment) 보존
  - 계산값을 직접 채워넣음 (formula 의존 없음)
  - 거래처 수가 다르면 행 자동 확장/축소

시트 매핑:
  C100 조회서 control sheet (MUS)  — 최종 발송 거래처 목록
  C100-1 표본규모 결정 (MUS)        — PM·CF·표본규모 산출
  C100-2 Key item 추출 (MUS)        — 모집단 완전성 + Key item 표
  C100-3 표본 추출(MUS)             — MUS 추출 내역

AA100 시리즈는 대응 시트가 원본에 있으면 동일 처리.
"""
from __future__ import annotations

import copy
import shutil
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.domain.mus import MUSResult
from src.domain.population import CompletenessCheck, PartyDecision
from src.domain.sample_size import SampleSizeResult


ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_PATH = ROOT / "templates" / "cc_template.xlsx"

# 시트 이름 (원본 그대로)
SHEET_C100_CONTROL = "C100 조회서 control sheet (MUS)"
SHEET_C100_1 = "C100-1 표본규모 결정 (MUS)"
SHEET_C100_2 = "C100-2 Key item 추출 (MUS)"
SHEET_C100_3 = "C100-3 표본 추출(MUS)"

# 계정 그룹 → C100-2 컬럼 매핑
RECEIVABLE_GROUP_COL = {
    "외상매출금": 3, "받을어음": 4, "미수금": 5, "선급금": 6,
    "장기대여금": 7, "임차보증금": 8, "기타보증금": 9,
}
PAYABLE_GROUP_COL = {
    "외상매입금": 3, "지급어음(외담대외상매입금)": 4, "미지급금": 5,
    "선수금": 6, "임대보증금": 7,
}


@dataclass
class ReportContext:
    company_name: str
    period_end: date
    kind: str
    preparer: str = ""
    reviewer: str = ""
    prep_date: date | None = None
    review_date: date | None = None
    workpaper_no_prefix: str = "C100"


@dataclass
class AlternativeProcedureEntry:
    """대체적 절차 시트 기입용 데이터 클래스."""
    party_name: str
    reason: str           # "미회신" | "차이" | "회신거부" | "기타"
    ledger_balance: float | None
    procedure_type: str
    evidence_names: list[str]       # 증빙 파일명 목록
    covered_amount: float | None
    coverage_ratio: float | None    # 0.0 ~ 1.0
    conclusion: str
    auditor_notes: str | None = None


def build_template_report(
    out_path: str | Path,
    ctx: ReportContext,
    completeness: CompletenessCheck,
    size_result: SampleSizeResult,
    decisions: list[PartyDecision],
    mus_result: MUSResult,
    performance_materiality: float,
    population_amount: float,
    exclusion_rows: list[dict] | None = None,
    alternative_procedures: list[AlternativeProcedureEntry] | None = None,
) -> None:
    """원본 7620 양식 그대로 채워서 출력"""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(TEMPLATE_PATH, out_path)

    wb = openpyxl.load_workbook(out_path)

    # 헤더 정보 공통 갱신
    _write_header_all_sheets(wb, ctx)

    # C100-1
    _fill_c100_1(
        wb[SHEET_C100_1], ctx, size_result, performance_materiality,
        population_amount, decisions,
    )

    # C100-2
    _fill_c100_2(
        wb[SHEET_C100_2], ctx, completeness, size_result,
        performance_materiality, decisions, exclusion_rows or [],
    )

    # C100-3
    _fill_c100_3(
        wb[SHEET_C100_3], ctx, size_result, mus_result, population_amount,
    )

    # Control sheet
    _fill_control_sheet(wb[SHEET_C100_CONTROL], ctx, decisions)

    # 대체적 절차 시트 (있는 경우)
    if alternative_procedures:
        _fill_alternative_procedures(wb, ctx, alternative_procedures)

    wb.save(out_path)


def _fill_alternative_procedures(
    wb,
    ctx: ReportContext,
    procedures: list[AlternativeProcedureEntry],
) -> None:
    """대체적 절차 시트 자동 작성.

    원본 템플릿에 "대체적 절차" 시트가 있으면 데이터를 채우고,
    없으면 새 시트를 추가한다.
    """
    SHEET_NAME = "대체적 절차"

    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb.create_sheet(SHEET_NAME)
        # 헤더 작성
        headers = ["No", "거래처명", "사유", "장부가", "절차유형",
                   "증빙명세", "커버금액", "커버리지%", "결론", "감사인 메모"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(1, i, h)
            cell.font = Font(bold=True, size=9)

    # 데이터 시작행 탐색: 기존 내용 밑에 써야 할 수도 있으나
    # Week 5에서는 R2부터 덮어씀 (단순 접근)
    start_row = 2

    for i, proc in enumerate(procedures):
        r = start_row + i
        ws.cell(r, 1, i + 1)
        ws.cell(r, 2, proc.party_name)
        ws.cell(r, 3, proc.reason)
        ws.cell(r, 4, proc.ledger_balance)
        ws.cell(r, 5, proc.procedure_type)
        ws.cell(r, 6, "; ".join(proc.evidence_names) if proc.evidence_names else "")
        ws.cell(r, 7, proc.covered_amount)
        if proc.coverage_ratio is not None:
            ws.cell(r, 8, f"{proc.coverage_ratio * 100:.1f}%")
        ws.cell(r, 9, proc.conclusion)
        ws.cell(r, 10, proc.auditor_notes or "")

        # 결론별 배경색
        _CONCLUSION_FILL = {
            "충분":        PatternFill("solid", fgColor="D1FAE5"),
            "부분":        PatternFill("solid", fgColor="FEF9C3"),
            "미해소":      PatternFill("solid", fgColor="FEE2E2"),
            "needs_review": PatternFill("solid", fgColor="F1F5F9"),
        }
        fill = _CONCLUSION_FILL.get(proc.conclusion)
        if fill:
            ws.cell(r, 9).fill = fill

    # 헤더 행 회사명
    ws.cell(1, 1).value = ctx.company_name if ws.cell(1, 1).value is None else ws.cell(1, 1).value


# ─────────────────────────────────────────────────────────────
def _write_header_all_sheets(wb, ctx: ReportContext):
    """모든 시트 상단 회사명·작성자·검토자·일자 갱신"""
    prep_date = ctx.prep_date or date.today()
    review_date = ctx.review_date or date.today()
    for sname in (SHEET_C100_1, SHEET_C100_2, SHEET_C100_3, SHEET_C100_CONTROL):
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        ws.cell(1, 1, ctx.company_name)
        ws.cell(3, 1, ctx.period_end)
        # 작성자/검토자/일자 — 시트별로 위치 다름
        if sname == SHEET_C100_CONTROL:
            ws.cell(1, 12, ctx.preparer)
            ws.cell(2, 12, ctx.reviewer)
            ws.cell(1, 14, prep_date)
            ws.cell(2, 14, review_date)
        else:
            ws.cell(1, 5, ctx.preparer)
            ws.cell(2, 5, ctx.reviewer)
            ws.cell(1, 7, prep_date)
            ws.cell(2, 7, review_date)


# ─────────────────────────────────────────────────────────────
def _fill_c100_1(ws, ctx, size, pm, pop, decisions):
    """표본규모 결정 시트 — 핵심 셀 직접 갱신"""
    ki = [d for d in decisions if d.is_key_item]
    rep = [d for d in decisions if d.is_representative and not d.is_key_item]

    ki_amt = sum(d.balance for d in ki)
    rep_amt = sum(d.balance for d in rep)
    total_n = len(ki) + len(rep)
    total_amt = ki_amt + rep_amt

    # 2. 조회대상 표본 (R11~R15)
    ws.cell(11, 4, len(ki))
    ws.cell(11, 5, ki_amt)
    ws.cell(11, 7, ki_amt / pop if pop else 0)
    ws.cell(12, 4, len(rep))
    ws.cell(12, 5, rep_amt)
    ws.cell(13, 4, total_n)
    ws.cell(13, 5, total_amt)
    ws.cell(14, 4, len(decisions) - sum(1 for d in decisions if d.is_excluded))
    ws.cell(14, 5, pop)
    # Coverage
    ws.cell(15, 4, total_n / len(decisions) if decisions else 0)
    ws.cell(15, 5, total_amt / pop if pop else 0)

    # 3. 표본규모 결정근거
    ws.cell(23, 5, pop)              # 모집단금액
    ws.cell(24, 5, pm)               # 수행중요성
    ws.cell(25, 5, size.key_item_threshold)
    ws.cell(26, 5, ki_amt)
    ws.cell(27, 5, len(ki))
    ws.cell(28, 5, size.base_sample_size)
    ws.cell(29, 5, size.confidence_factor)
    # 30, 31 위험·통제 의존 (드롭다운) — 그대로 두거나 갱신
    ws.cell(32, 5, size.final_sample_size)


# ─────────────────────────────────────────────────────────────
def _fill_c100_2(ws, ctx, comp: CompletenessCheck, size, pm, decisions, exclusion_rows):
    """Key item 추출 시트
    - R12~R24: 모집단 완전성 (기말금액)
    - R29~R38: 발송제외 (선택)
    - R42~R44: PM·비율·기준금액
    - R47~: 거래처 매트릭스
    """
    group_col_map = (RECEIVABLE_GROUP_COL if ctx.kind == "receivable" else PAYABLE_GROUP_COL)

    # 1) 완전성 검토 표 — comp.by_group 기반
    #    원본 행: R13 매출채권 합계, R16~24 그룹별 (외상매출금, 받을어음, 미수금, ...)
    #    회사제시(C) · 재무제표(D) · 차이(E) · 비고(F)
    #    기존 데이터 클리어 (R13~R24)
    for r in range(13, 25):
        for c in (3, 4, 5, 6):
            ws.cell(r, c).value = None

    if comp:
        ws.cell(13, 3, comp.total_ledger)
        ws.cell(13, 4, comp.total_fs)
        ws.cell(13, 5, comp.total_ledger - comp.total_fs)
        # 그룹별 — R16부터
        r = 16
        for row in comp.by_group:
            if r > 24:
                break
            ws.cell(r, 2, row["group"])
            ws.cell(r, 3, row["ledger"])
            ws.cell(r, 4, row["fs"])
            ws.cell(r, 5, row["diff"])
            if row["note"]:
                ws.cell(r, 6, row["note"])
            r += 1
        # 합계 행
        ws.cell(24, 3, sum(r2["ledger"] for r2 in comp.by_group))
        ws.cell(24, 4, sum(r2["fs"] for r2 in comp.by_group))
        ws.cell(24, 5, sum(r2["diff"] for r2 in comp.by_group))

    # 2) 발송제외 표 R29~R38
    for r in range(30, 39):
        for c in (2, 3, 4, 5, 6, 7, 8, 9, 10):
            ws.cell(r, c).value = None
    r = 30
    excluded = [d for d in decisions if d.is_excluded]
    for d in excluded[:8]:
        ws.cell(r, 2, d.name)
        ws.cell(r, 4, d.balance)
        ws.cell(r, 6, d.exclusion_reason)
        r += 1

    # 3) PM·비율·기준금액
    ws.cell(42, 4, pm)
    ws.cell(43, 4, size.key_item_ratio)
    ws.cell(44, 4, size.key_item_threshold)

    # 4) 거래처 매트릭스 R47~
    _write_party_matrix_c100_2(ws, decisions, group_col_map, size.key_item_threshold)


def _write_party_matrix_c100_2(ws, decisions, group_col_map, threshold):
    """C100-2 거래처 매트릭스 — 발송제외 제외, 잔액 > 0"""
    parties = [d for d in decisions if not d.is_excluded and d.balance > 0]
    parties.sort(key=lambda d: d.name)

    start_row = 47
    template_rows = (47, 48)  # 짝/홀 스타일 참고용
    template_styles = _capture_row_styles(ws, template_rows[0], col_range=(2, 16))

    # 기존 데이터 영역 클리어 (R47~R103)
    for r in range(start_row, 104):
        for c in range(2, 16):
            cell = ws.cell(r, c)
            cell.value = None

    # 새 데이터 작성
    n = len(parties)
    needed_last_row = start_row + n - 1

    # 행 부족 시 행 추가
    if needed_last_row > 103:
        extra = needed_last_row - 103
        ws.insert_rows(104, amount=extra)
        # 추가 행에 스타일 복사
        for offset in range(extra):
            _apply_row_styles(ws, 104 + offset, template_styles, col_range=(2, 16))

    # 데이터 쓰기 — 계정 그룹별 컬럼 분해
    for i, d in enumerate(parties):
        r = start_row + i
        ws.cell(r, 2, d.name)
        # 계정 그룹별 잔액
        for group, col in group_col_map.items():
            amt = d.by_account.get(group, 0)
            if amt:
                ws.cell(r, col, amt)
        # 매출채권 계 / 매입채무 계 (J=10)
        ws.cell(r, 10, d.balance)
        ws.cell(r, 11, d.balance if d.is_key_item else 0)
        ws.cell(r, 12, "Y" if d.is_key_item else "N")
        ws.cell(r, 13, "Y" if d.is_representative else "N")
        ws.cell(r, 14, "Y" if d.is_related_party else "N")
        ws.cell(r, 15, "Y" if d.final_sampled else "N")

    # 합계 행 (마지막 행 + 1)
    last_data_row = start_row + n - 1
    sum_row = last_data_row + 1
    ws.cell(sum_row, 2, "합계")
    ws.cell(sum_row, 10, sum(d.balance for d in parties))
    ws.cell(sum_row, 11, sum(d.balance for d in parties if d.is_key_item))


# ─────────────────────────────────────────────────────────────
def _fill_c100_3(ws, ctx, size, mus, pop):
    """표본 추출 (MUS) 시트
    - D13~D16: 모집단·표본규모·표본간격·임의출발점
    - R22~: MUS 추출 내역
    """
    # 잔여 모집단 (Key item 제외)
    ws.cell(13, 4, size.remaining_population)
    ws.cell(14, 4, size.final_sample_size)
    ws.cell(15, 4, size.sample_interval)
    ws.cell(16, 4, mus.random_start)

    # MUS 추출 표 R22~
    start_row = 22
    template_styles = _capture_row_styles(ws, start_row, col_range=(2, 9))

    # 클리어
    for r in range(start_row, 73):
        for c in range(2, 9):
            ws.cell(r, c).value = None

    selections = mus.selections
    n = len(selections)
    needed_last = start_row + n - 1
    if needed_last > 72:
        extra = needed_last - 72
        ws.insert_rows(73, amount=extra)
        for offset in range(extra):
            _apply_row_styles(ws, 73 + offset, template_styles, col_range=(2, 9))

    cumulative = 0.0
    for i, s in enumerate(selections):
        r = start_row + i
        cumulative += s.balance
        ws.cell(r, 2, s.name)
        ws.cell(r, 3, s.balance)
        ws.cell(r, 4, s.cumulative)
        ws.cell(r, 5, s.selections)
        ws.cell(r, 6, size.sample_interval)
        ws.cell(r, 7, s.remainder_after)
        ws.cell(r, 8, "Y" if s.hit else "N")

    # 합계 (마지막 행 + 1)
    sum_row = start_row + n
    ws.cell(sum_row, 2, "합계")
    ws.cell(sum_row, 3, sum(s.balance for s in selections))
    ws.cell(sum_row, 5, sum(s.selections for s in selections))


# ─────────────────────────────────────────────────────────────
def _fill_control_sheet(ws, ctx, decisions):
    """Control sheet — 최종 발송 거래처
    R7~8: 헤더
    R9~: 데이터
    """
    final = sorted(
        [d for d in decisions if d.final_sampled],
        key=lambda d: -d.balance,
    )

    start_row = 9
    template_styles = _capture_row_styles(ws, start_row, col_range=(2, 18))

    # 클리어 (R9~R43)
    for r in range(start_row, 44):
        for c in range(2, 18):
            ws.cell(r, c).value = None

    n = len(final)
    needed_last = start_row + n - 1
    if needed_last > 43:
        extra = needed_last - 43
        ws.insert_rows(44, amount=extra)
        for offset in range(extra):
            _apply_row_styles(ws, 44 + offset, template_styles, col_range=(2, 18))

    # control sheet 컬럼 매핑 (R8 헤더 기준)
    # 채권:  4=외상매출금 5=받을어음 6=미수금 7=선급금 8=임차보증금 9=장기대여금 10=채권요약
    # 채무: 11=외상매입금 12=미지급금 13=임대보증금 14=채무요약
    AR_COL = {"외상매출금":4, "받을어음":5, "미수금":6, "선급금":7,
              "임차보증금":8, "장기대여금":9}
    AP_COL = {"외상매입금":11, "지급어음(외담대외상매입금)":11,
              "미지급금":12, "임대보증금":13}

    for i, d in enumerate(final):
        r = start_row + i
        ws.cell(r, 2, i + 1)
        ws.cell(r, 3, d.name)

        # 채권 측 분해
        ar_total = 0
        for g, amt in d.by_account.items():
            if g in AR_COL and amt:
                ws.cell(r, AR_COL[g], amt)
                ar_total += amt
        if ar_total:
            ws.cell(r, 10, ar_total)

        # 채무 측 분해
        ap_total = 0
        for g, amt in d.by_account.items():
            if g in AP_COL and amt:
                ws.cell(r, AP_COL[g], amt)
                ap_total += amt
        if ap_total:
            ws.cell(r, 14, ap_total)

        ws.cell(r, 15, ar_total + ap_total or d.balance)

    # 총합계 행 (마지막 행 + 1)
    sum_row = start_row + n
    ws.cell(sum_row, 3, "총합계")
    total = sum(d.balance for d in final)
    if ctx.kind == "receivable":
        ws.cell(sum_row, 4, total)
        ws.cell(sum_row, 10, total)
    else:
        ws.cell(sum_row, 11, total)
        ws.cell(sum_row, 14, total)
    ws.cell(sum_row, 15, total)


# ─────────────────────────────────────────────────────────────
# 스타일 복사 유틸
# ─────────────────────────────────────────────────────────────
def _capture_row_styles(ws: Worksheet, row: int, col_range: tuple[int, int]):
    """행 스타일 dict로 캡처"""
    styles = {}
    for c in range(col_range[0], col_range[1]):
        cell = ws.cell(row, c)
        styles[c] = {
            "font": copy.copy(cell.font),
            "fill": copy.copy(cell.fill),
            "border": copy.copy(cell.border),
            "alignment": copy.copy(cell.alignment),
            "number_format": cell.number_format,
        }
    return styles


def _apply_row_styles(ws: Worksheet, row: int, styles: dict, col_range: tuple[int, int]):
    for c in range(col_range[0], col_range[1]):
        s = styles.get(c)
        if not s:
            continue
        cell = ws.cell(row, c)
        cell.font = s["font"]
        cell.fill = s["fill"]
        cell.border = s["border"]
        cell.alignment = s["alignment"]
        cell.number_format = s["number_format"]
