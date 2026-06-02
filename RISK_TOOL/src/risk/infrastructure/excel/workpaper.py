from __future__ import annotations
import openpyxl
from openpyxl.styles import Font, PatternFill
from risk.application.assess_risk_uc import RiskResult

_FILL = {"red": PatternFill("solid", fgColor="FFC7CE"),
         "yellow": PatternFill("solid", fgColor="FFEB9C"),
         "green": PatternFill("solid", fgColor="C6EFCE"),
         "na": PatternFill("solid", fgColor="D9D9D9")}  # 회색 = 데이터없음/보류
_FOLLOWUP = {
    "ar_turnover": "매출채권 조회·기수령 검토·연령분석",
    "accrual_quality": "발생액 분석·수익인식 cutoff 검토",
    "debt_ratio": "차입약정 위반·만기구조·계속기업 평가",
    "interest_coverage": "계속기업 가정·차입금 상환능력 검토",
    "revenue_change": "수익인식 정책·이상거래 표본 검토",
}


def build_workpaper(res: RiskResult, path: str) -> str:
    wb = openpyxl.Workbook()
    # 표지
    ws = wb.active; ws.title = "표지"
    ws["A1"] = "감사전 위험평가 조서 (ISA 315)"; ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "대상회사"; ws["B3"] = res.company
    ws["A4"] = "종합위험등급"; ws["B4"] = res.grade.grade if res.grade else "-"
    if res.materiality:
        ws["A5"] = "수행중요성(PM)"; ws["B5"] = res.materiality.pm
        ws["A6"] = "중요성 benchmark"; ws["B6"] = res.materiality.benchmark
    if res.error:
        ws["A8"] = "오류"; ws["B8"] = res.error

    # 재무요약
    ws2 = wb.create_sheet("재무요약")
    ws2.append(["연도", "매출", "영업이익", "당기순이익", "자산", "부채", "자본", "영업CF"])
    for y in res.years:
        ws2.append([y.year, y.revenue, y.operating_income, y.net_income,
                    y.total_assets, y.total_liabilities, y.total_equity, y.operating_cf])

    # 위험평가매트릭스 (계정×주장은 신호 매핑 요약)
    ws3 = wb.create_sheet("위험평가매트릭스")
    ws3.append(["축", "지표", "신호", "값", "기준", "AI코멘트"])
    for s in res.signals:
        row = [s.axis, s.label, s.level, s.value, s.threshold, res.comments.get(s.code, "")]
        ws3.append(row)
        # na 안전 fallback: 미지정 level은 green이 아니라 na(회색)로 처리해
        # 감사인이 "데이터없음/보류"를 정상 green과 구별하도록 함.
        ws3.cell(ws3.max_row, 3).fill = _FILL.get(s.level, _FILL["na"])

    # 4축신호상세
    ws4 = wb.create_sheet("신호상세")
    ws4.append(["축", "code", "지표", "신호", "값", "기준", "비고"])
    for s in res.signals:
        ws4.append([s.axis, s.code, s.label, s.level, s.value, s.threshold, s.note])

    # 외부리스크
    ws5 = wb.create_sheet("외부리스크")
    ws5.append(["키워드", "제목", "날짜", "요약", "출처"])
    for h in res.news:
        ws5.append([getattr(h, "keyword", ""), getattr(h, "title", ""),
                    getattr(h, "date", ""), getattr(h, "summary", ""), getattr(h, "url", "")])
    if not res.news:
        ws5.append(["", "특이사항 없음", "", "", ""])
    # 축4 DART 공시이벤트 (구 fixture 호환: getattr 가드)
    disclosures = getattr(res, "disclosures", []) or []
    if disclosures:
        ws5.append(["", "", "", "", ""])  # 구분행
        for d in disclosures:
            ws5.append(["공시", d.get("report_nm", ""), d.get("rcept_dt", ""),
                        "", d.get("rcept_no", "")])

    # 후속절차 — yellow/red만 (green·na 제외)
    ws6 = wb.create_sheet("후속감사절차")
    ws6.append(["지표", "신호", "권고절차"])
    for s in res.signals:
        if s.level in ("yellow", "red"):
            ws6.append([s.label, s.level, _FOLLOWUP.get(s.code, "추가 검토 절차 설계")])

    wb.save(path)
    return path
