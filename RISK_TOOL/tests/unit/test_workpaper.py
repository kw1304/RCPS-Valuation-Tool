import openpyxl
from risk.application.assess_risk_uc import RiskResult
from risk.domain.financial import FinancialYear
from risk.domain.materiality import Materiality
from risk.domain.risk_grade import RiskGrade
from risk.domain.thresholds import Signal
from risk.infrastructure.excel.workpaper import build_workpaper


def test_workpaper_sheets(tmp_path):
    res = RiskResult("테스트", [FinancialYear(2025, revenue=1000)],
                     Materiality(5, 3.75, "revenue"),
                     signals=[], grade=RiskGrade("낮음", 0, 0))
    p = tmp_path / "wp.xlsx"
    build_workpaper(res, str(p))
    wb = openpyxl.load_workbook(p)
    assert "표지" in wb.sheetnames
    assert "위험평가매트릭스" in wb.sheetnames


def test_financial_cells_have_comma_format(tmp_path):
    res = RiskResult("콤마", [FinancialYear(2025, revenue=1128265886250,
                                          operating_income=23450000)],
                     Materiality(5000000, 3.75, "revenue"),
                     signals=[], grade=RiskGrade("낮음", 0, 0))
    p = tmp_path / "fmt.xlsx"
    build_workpaper(res, str(p))
    wb = openpyxl.load_workbook(p)
    ws = wb["재무요약"]
    assert ws.cell(row=2, column=2).number_format == "#,##0"  # 매출
    assert ws.cell(row=2, column=8).number_format == "#,##0"  # 영업CF
    # PM 셀(표지 B5)
    assert wb["표지"]["B5"].number_format == "#,##0"


def test_structured_events_rows_at_top(tmp_path):
    res = RiskResult("이벤트", [FinancialYear(2025, revenue=1000)],
                     Materiality(5, 3.75, "revenue"),
                     signals=[], grade=RiskGrade("낮음", 0, 0),
                     events=[{"type": "소송", "summary": "하도급 소송",
                              "impact": "우발부채", "date": "2025-03",
                              "source": "http://x"}])
    p = tmp_path / "ev.xlsx"
    build_workpaper(res, str(p))
    wb = openpyxl.load_workbook(p)
    ws = wb["외부리스크"]
    assert ws.cell(row=1, column=1).value == "유형"
    assert ws.cell(row=2, column=1).value == "소송"


def test_na_signal_gets_gray_fill(tmp_path):
    # na 신호(데이터 없음) → 빌드 성공 + 매트릭스 신호셀 회색(D9D9D9)
    na_sig = Signal("analytical", "revenue_change", "매출 증감률", "na",
                    None, "±10%황/±30%적", note="데이터 없음 — 신호 보류")
    res = RiskResult("NA테스트", [FinancialYear(2025, revenue=1000)],
                     Materiality(5, 3.75, "revenue"),
                     signals=[na_sig], grade=RiskGrade("낮음", 0, 0))
    p = tmp_path / "na.xlsx"
    build_workpaper(res, str(p))
    wb = openpyxl.load_workbook(p)
    ws = wb["위험평가매트릭스"]
    # 헤더(1행) + 신호(2행), 신호셀은 3번째 컬럼
    cell = ws.cell(row=2, column=3)
    assert cell.fill.fgColor.rgb in ("00D9D9D9", "FFD9D9D9", "D9D9D9")


def test_followup_excludes_green_and_na(tmp_path):
    # 후속감사절차: yellow/red만, green·na 제외
    sigs = [
        Signal("analytical", "revenue_change", "매출 증감률", "red", 40.0, "t"),
        Signal("fraud", "effective_tax", "유효세율", "yellow", 8.0, "t"),
        Signal("going_concern", "current_ratio", "유동비율", "green", 200.0, "t"),
        Signal("analytical", "inv_turnover", "재고회전율", "na", None, "t"),
    ]
    res = RiskResult("필터테스트", [FinancialYear(2025, revenue=1000)],
                     Materiality(5, 3.75, "revenue"),
                     signals=sigs, grade=RiskGrade("높음", 1, 1))
    p = tmp_path / "fu.xlsx"
    build_workpaper(res, str(p))
    wb = openpyxl.load_workbook(p)
    ws = wb["후속감사절차"]
    labels = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    labels = [x for x in labels if x]
    assert "매출 증감률" in labels      # red 포함
    assert "유효세율" in labels         # yellow 포함
    assert "유동비율" not in labels     # green 제외
    assert "재고회전율" not in labels   # na 제외
