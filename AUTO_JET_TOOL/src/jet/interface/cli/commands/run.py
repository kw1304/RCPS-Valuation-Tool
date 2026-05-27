"""jet run 커맨드 — 감사조서 Excel 생성 CLI.

사용 예:
    jet run \\
      --workpaper templates/workpapers/blank.yaml \\
      --schema configs/schema_mapping/auto_detect.yaml \\
      --input "INPUT/GL_2025.xlsx" \\
      --hr "INPUT/HR_2025.xlsx" \\
      --output "OUTPUT/7400_FY2025.xlsx"
"""

from __future__ import annotations

import logging
from datetime import date
from pathlib import Path
from typing import Annotated, Optional

import typer
import yaml

from jet.application.pipeline.normalizer import Normalizer
from jet.application.pipeline.schema_mapper import SchemaMapper
from jet.application.workpaper.template_loader import TemplateLoader
from jet.domain.rules.a01_integrity import A01DataIntegrity
from jet.domain.rules.a02_dr_cr_balance import A02DrCrBalance
from jet.domain.rules.a03_tb_rollforward import A03TBRollforward
from jet.domain.rules.b01_large_pl import B01LargePLItem
from jet.domain.rules.b02_unmatched_account import B02UnmatchedAccount
from jet.domain.rules.b03_new_account import B03NewlyCreatedAccount
from jet.domain.rules.b04_seldom_used import B04SeldomUsedAccount
from jet.domain.rules.b05_unusual_user import B05UnusualUser
from jet.domain.rules.b06_inappropriate_user import B06InappropriateUser
from jet.domain.rules.b07_backdated_entry import B07BackDatedEntry
from jet.domain.rules.b08_doc_type_account import B08DocTypeAccountCombo
from jet.domain.rules.b09_counter_account import B09CounterAccountAnalysis
from jet.domain.rules.base import RuleContext
from jet.infrastructure.io.sap_gl_loader import SapGlLoader
from jet.infrastructure.reporters.excel_reporter import ExcelReporter

logger = logging.getLogger(__name__)

# 룰 코드 → 클래스 매핑
_RULE_CLASSES = {
    "A01": A01DataIntegrity,
    "A02": A02DrCrBalance,
    "A03": A03TBRollforward,
    "B01": B01LargePLItem,
    "B02": B02UnmatchedAccount,
    "B03": B03NewlyCreatedAccount,
    "B04": B04SeldomUsedAccount,
    "B05": B05UnusualUser,
    "B06": B06InappropriateUser,
    "B07": B07BackDatedEntry,
    "B08": B08DocTypeAccountCombo,
    "B09": B09CounterAccountAnalysis,
}


def run(
    workpaper: Annotated[
        Path,
        typer.Option("--workpaper", help="조서 템플릿 YAML 경로"),
    ],
    schema: Annotated[
        Path,
        typer.Option("--schema", help="스키마 매핑 YAML 경로"),
    ],
    input_path: Annotated[
        Path,
        typer.Option("--input", help="GL 입력 파일 경로 (Excel)"),
    ],
    output: Annotated[
        Path,
        typer.Option("--output", help="출력 Excel 파일 경로"),
    ],
    hr_path: Annotated[
        Optional[Path],
        typer.Option("--hr", help="HR 마스터 파일 경로"),
    ] = None,
    coa_path: Annotated[
        Optional[Path],
        typer.Option("--coa", help="COA 마스터 파일 경로"),
    ] = None,
    tb_path: Annotated[
        Optional[Path],
        typer.Option("--tb", help="합계잔액시산표 파일 경로"),
    ] = None,
    doctype_path: Annotated[
        Optional[Path],
        typer.Option("--doctype", help="전표유형 파일 경로"),
    ] = None,
    company: Annotated[
        Optional[str],
        typer.Option("--company", help="회사명 (YAML 값 오버라이드)"),
    ] = None,
    period_end_override: Annotated[
        Optional[str],
        typer.Option("--period-end", help="결산일 (YYYY-MM-DD, YAML 값 오버라이드)"),
    ] = None,
    cache_dir: Annotated[
        Optional[Path],
        typer.Option("--cache-dir", help="Parquet 캐시 디렉토리 (기본: data/cache)"),
    ] = None,
    force_reload: Annotated[
        bool,
        typer.Option("--force-reload/--no-force-reload", help="캐시 무시하고 재로드"),
    ] = False,
    verbose: Annotated[
        bool,
        typer.Option("--verbose/--no-verbose", help="상세 로그 출력"),
    ] = False,
) -> None:
    """GL 데이터로 JET 감사조서 Excel을 생성한다.

    SAP 취합본을 적재하여 A01~A03·B01~B09를 검증하고
    결과를 조서 양식 Excel로 출력한다.
    """
    _configure_logging(verbose)

    # ── 1. 조서 템플릿 로드 ───────────────────────────────────────────────
    typer.echo(f"조서 템플릿 로드: {workpaper}")
    spec = TemplateLoader().load(workpaper)

    # CLI --company / --period-end 오버라이드
    if company or period_end_override:
        from dataclasses import replace as dc_replace
        spec = dc_replace(
            spec,
            company=company or spec.company,
            period_end=period_end_override or spec.period_end,
        )

    typer.echo(f"  회사: {spec.company} / 결산일: {spec.period_end} / 활성룰: {len(spec.enabled_scenarios)}종")

    # ── 2. 마스터 파일 경로 결정 (CLI > YAML 기본값) ──────────────────────
    master_files = spec.master_files
    _hr_path = hr_path or _resolve_path(master_files.get("hr"))
    _coa_path = coa_path or _resolve_path(master_files.get("coa"))
    _tb_path = tb_path or _resolve_path(master_files.get("tb"))
    _dt_path = doctype_path or _resolve_path(master_files.get("doctype"))

    # ── 3. 마스터 적재 ────────────────────────────────────────────────────
    coa_master = None
    doc_type_master = None
    hr_master = None
    tb_master = None
    user_df = None

    if _coa_path and _coa_path.exists():
        typer.echo(f"COA 마스터 적재: {_coa_path}")
        from jet.infrastructure.io.coa_loader import CoaLoader
        coa_master = CoaLoader().load(_coa_path)
        typer.echo(f"  COA 적재 완료: {len(coa_master):,}계정")
    else:
        typer.echo("COA 마스터 없음 - B01/B02/B03/B04 제한적 실행")

    if _dt_path and _dt_path.exists():
        typer.echo(f"전표유형 마스터 적재: {_dt_path}")
        from jet.infrastructure.io.doc_type_loader import DocTypeLoader
        doc_type_master = DocTypeLoader().load(_dt_path)
        typer.echo(f"  전표유형 적재 완료: {len(doc_type_master):,}종")

    if _hr_path and _hr_path.exists():
        typer.echo(f"HR 마스터 적재: {_hr_path}")
        from jet.infrastructure.io.hr_loader import HRLoader
        hr_master = HRLoader().load(_hr_path)
        typer.echo(
            f"  HR 적재 완료: 재직자 {len(hr_master.active_employees):,}명 / "
            f"퇴직자 {len(hr_master.retired_employees):,}명"
        )
    else:
        typer.echo("HR 마스터 없음 — B05 제한적 실행")

    tb_master_prior = None
    if _tb_path and _tb_path.exists():
        typer.echo(f"TB 마스터 적재: {_tb_path}")
        from jet.infrastructure.io.tb_loader import TbLoader
        # load_with_prior()로 당기·전기 TB를 동시에 적재.
        # B03 신규계정 fallback(COA created_date 없을 때 TB 비교)에 전기 TB 사용.
        # COA 없으면 계정코드 첫자리 fallback으로 부호 결정.
        tb_loader = TbLoader(coa_master=coa_master)
        tb_master, tb_master_prior = tb_loader.load_with_prior(_tb_path)
        prior_msg = f" / 전기 {len(tb_master_prior):,}계정" if tb_master_prior else " / 전기 없음"
        typer.echo(f"  TB 적재 완료: 당기 {len(tb_master):,}계정{prior_msg}")

    # USER 리스트 (USR02) 로드
    user_path = Path("INPUT/extracted/8. 사용자리스트 USR02.XLSX")
    if user_path.exists():
        try:
            import pandas as pd
            user_df = pd.read_excel(user_path, engine="openpyxl", dtype=str)
            typer.echo(f"  사용자 리스트 적재 완료: {len(user_df):,}건")
        except Exception as e:
            logger.warning("USER 리스트 적재 실패: %s", e)

    # ── 4. 스키마 매핑 로드 ───────────────────────────────────────────────
    typer.echo(f"스키마 매핑 로드: {schema}")
    with schema.open(encoding="utf-8") as fh:
        schema_dict = yaml.safe_load(fh)
    mapper = SchemaMapper(schema_dict)

    # ── 5. SAP GL 적재 ────────────────────────────────────────────────────
    resolved_cache = cache_dir or (Path.cwd() / "data" / "cache")
    typer.echo(f"SAP GL 적재 시작: {input_path}")
    loader = SapGlLoader(cache_dir=resolved_cache, force_reload=force_reload)
    raw_df = loader.load(input_path)
    typer.echo(f"  적재 완료: {len(raw_df):,}라인")

    # ── 6. 스키마 매핑 + 정규화 ──────────────────────────────────────────
    typer.echo("컬럼 매핑 + 정규화 중...")
    mapped_df = mapper.map(raw_df)
    normalizer = Normalizer(
        amount_unit=mapper.amount_unit,
        account_code_pad=mapper.account_code_pad,
    )
    entries, norm_report = normalizer.normalize(mapped_df)
    typer.echo(
        f"  정규화 완료: {norm_report.success_count:,}건 성공 / "
        f"{norm_report.quarantine_count}건 격리"
    )

    # 시스템 자동전표 통계 출력
    sys_entries = sum(1 for e in entries if e.is_system_generated)
    if sys_entries:
        typer.echo(f"  시스템 자동전표(SYSTEM-*): {sys_entries:,}건 (정상 처리됨)")

    # ── 7. 룰 실행 컨텍스트 ──────────────────────────────────────────────
    context = RuleContext(
        period_start=date(2025, 1, 1),
        period_end=date(2025, 12, 31),
        coa_master=coa_master,
        doc_type_master=doc_type_master,
        hr_master=hr_master,
        tb_master=tb_master,
        tb_master_prior=tb_master_prior,
    )
    results: dict = {}

    # ── 8. 룰 실행 ────────────────────────────────────────────────────────
    enabled_codes = {s.code for s in spec.enabled_scenarios}
    # B06는 비활성이어도 항상 실행 (waive/active 시트 생성용)
    # Q01은 룰 클래스가 없는 메타 시나리오이므로 실행 제외
    all_exec_codes = (enabled_codes | {"B06"}) - {"Q01"}

    for code in sorted(all_exec_codes):
        rule_cls = _RULE_CLASSES.get(code)
        if rule_cls is None:
            continue

        typer.echo(f"{code} 룰 실행 중...")
        rule = rule_cls()

        # 시나리오별 params 주입
        scenario = next((s for s in spec.scenarios if s.code == code), None)
        params = scenario.params if scenario and scenario.params else {}
        rule.configure(params)

        result = rule.apply(entries, context)
        results[code] = result

        _print_rule_result(code, result, typer)

    # ── 9. Excel 생성 ─────────────────────────────────────────────────────
    typer.echo(f"\nExcel 조서 생성 중: {output}")
    master_data = {
        "coa": coa_master,
        "hr": hr_master,
        "doc_types": doc_type_master,
        "tb": tb_master,
        "user_df": user_df,
    }
    reporter = ExcelReporter()
    final_path = reporter.write(
        spec=spec,
        results=results,
        output_path=output,
        master_data=master_data,
    )

    typer.echo(f"\n생성 완료: {final_path}")

    # 시트 목록 출력
    import openpyxl
    try:
        wb_check = openpyxl.load_workbook(str(final_path), read_only=True)
        typer.echo(f"시트 수: {len(wb_check.sheetnames)}개")
        typer.echo(f"시트 목록: {', '.join(wb_check.sheetnames)}")
        wb_check.close()
    except Exception:
        pass


def _resolve_path(raw: str | None) -> Path | None:
    """상대경로 문자열을 Path로 변환한다."""
    if not raw:
        return None
    p = Path(raw)
    return p


def _print_rule_result(code: str, result: "RuleResult", echo: typer.echo) -> None:  # type: ignore[type-arg]
    """룰 실행 결과를 CLI에 출력한다."""
    from jet.domain.entities.rule_result import RuleResult

    finding_cnt = result.finding_count

    extra_info = ""
    if code == "A01":
        stats = result.extra.get("integrity_stats")
        if stats:
            extra_info = f"형식오류 {stats.format_error_count}건 / 필수결측 {stats.missing_required_count}건"
    elif code == "A02":
        dr_stats = result.extra.get("dr_cr_stats")
        if dr_stats:
            extra_info = f"전표 {dr_stats.total_entries:,}건 / 불일치 {dr_stats.imbalanced_count}건"
    elif code == "A03":
        extra_info = f"검증 계정 {result.extra.get('accounts_checked', 0):,}개 / 불일치 {result.extra.get('mismatch_count', 0)}건"
    elif code == "B05":
        extra_info = (
            f"미등록 {result.extra.get('not_registered_count', 0):,}건 / "
            f"퇴직후 {result.extra.get('post_retirement_count', 0):,}건 / "
            f"시스템 {result.extra.get('system_account_count', 0):,}건 / "
            f"외부계약직 {result.extra.get('external_count', 0):,}건"
        )
    elif code == "B06":
        extra_info = "데이터 미입수 (Waived)"
        finding_cnt = 0
    elif code == "B09":
        sub_total = sum(
            len(sr.findings) for sr in result.extra.get("sub_results", [])
        )
        extra_info = f"희소조합 {finding_cnt:,}건 / 서브시나리오 {sub_total:,}건"
    else:
        if finding_cnt > 0:
            extra_info = f"적출 {finding_cnt:,}건"
        else:
            extra_info = "적출 없음"

    typer.echo(f"  {code} 완료: {extra_info}")


def _configure_logging(verbose: bool) -> None:
    """로깅 레벨을 설정한다."""
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
        datefmt="%H:%M:%S",
    )
