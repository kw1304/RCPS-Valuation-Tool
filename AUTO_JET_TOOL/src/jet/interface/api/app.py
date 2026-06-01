"""JET 자동화 툴 Flask 웹 API.

RCPS 평가툴(c:/Claude/rcps_valuation/api/app.py)과 동일한 패턴으로 구현한다.
Basic Auth, /healthz, 정적 파일 서빙, 백그라운드 실행 지원.

환경변수:
    APP_USER         : Basic Auth 사용자 (기본 admin)
    APP_PASSWORD     : Basic Auth 비밀번호 (비어 있으면 인증 비활성화 — 로컬용)
    JET_PORT         : 서버 포트 (기본 5050)
    JET_DATA_DIR     : 업로드/출력 임시 저장 위치 (기본 ./data/runtime)
    JET_MAX_UPLOAD_MB: 업로드 최대 크기 MB (기본 200)
"""

from __future__ import annotations

import logging
import os
import shutil
import sys
import threading
import traceback
import uuid
from datetime import date, datetime, timedelta
from functools import wraps
from pathlib import Path
from typing import Any

from flask import Flask, Response, jsonify, request, send_file, send_from_directory

# 경로 계산: app.py → api/ → interface/ → jet/ → src/ → PROJECT_ROOT
# parents[3] = src/, parents[4] = PROJECT_ROOT (c:/Claude/AUTO_JET_TOOL)
_PROJECT_ROOT = Path(__file__).resolve().parents[4]
_SRC_DIR = _PROJECT_ROOT / "src"
if str(_SRC_DIR) not in sys.path:
    sys.path.insert(0, str(_SRC_DIR))

logger = logging.getLogger(__name__)

# ── 프론트엔드 디렉토리 ─────────────────────────────────────────────────────
_FRONTEND_DIR = Path(__file__).parent / "frontend"

# ── Flask 앱 초기화 ─────────────────────────────────────────────────────────
app = Flask(__name__, static_folder=str(_FRONTEND_DIR))
app.config["MAX_CONTENT_LENGTH"] = int(
    os.environ.get("JET_MAX_UPLOAD_MB", "200")
) * 1024 * 1024

# ── 환경변수 ────────────────────────────────────────────────────────────────
_AUTH_USER = os.environ.get("APP_USER", "admin")
_AUTH_PASS = os.environ.get("APP_PASSWORD", "")

_DATA_DIR = Path(os.environ.get("JET_DATA_DIR", str(_PROJECT_ROOT / "data" / "runtime")))
_DATA_DIR.mkdir(parents=True, exist_ok=True)

# ── configs 디렉토리 ────────────────────────────────────────────────────────
_CONFIGS_DIR = _PROJECT_ROOT / "configs"
_SCHEMA_MAPPING_DIR = _CONFIGS_DIR / "schema_mapping"

# ── 실행 상태 저장소 (메모리) ───────────────────────────────────────────────
# run_id → {status, progress, result, error, created_at}
_runs: dict[str, dict[str, Any]] = {}
_runs_lock = threading.Lock()


# ── Basic Auth ───────────────────────────────────────────────────────────────
@app.before_request
def _require_auth() -> Response | None:
    """헬스체크를 제외한 모든 요청에 Basic Auth를 적용한다."""
    if request.path == "/healthz":
        return None
    if not _AUTH_PASS:
        return None  # 비밀번호 미설정 시 통과 (로컬)
    auth = request.authorization
    if auth and auth.username == _AUTH_USER and auth.password == _AUTH_PASS:
        return None
    return Response(
        "Authentication required.", 401,
        {"WWW-Authenticate": 'Basic realm="JET Auto Tool"'},
    )


# ── 헬스체크 ─────────────────────────────────────────────────────────────────
@app.route("/healthz")
def healthz():
    """UptimeRobot 등 외부 모니터링 keep-alive용."""
    # 등록된 룰 수 계산
    try:
        from jet.domain.rules.base import Rule
        _ensure_rules_loaded()
        rules_count = len(Rule.get_registry())
    except Exception:
        rules_count = 0
    return jsonify({"status": "ok", "version": "0.1.0", "rules_count": rules_count}), 200


# ── 정적 파일 (프론트엔드 SPA) ───────────────────────────────────────────────
@app.route("/")
def index():
    """프론트엔드 SPA index.html을 서빙한다."""
    return send_from_directory(str(_FRONTEND_DIR), "index.html")


@app.route("/static/<path:filename>")
def static_files(filename: str):
    """정적 자산 서빙."""
    static_dir = _FRONTEND_DIR / "static"
    if static_dir.exists():
        return send_from_directory(str(static_dir), filename)
    return jsonify({"error": "not found"}), 404


# ── 스키마 매핑 목록 ─────────────────────────────────────────────────────────
@app.route("/api/schema-mappings")
def list_schema_mappings():
    """configs/schema_mapping/ 에 있는 YAML 파일 목록을 반환한다."""
    result = []
    if not _SCHEMA_MAPPING_DIR.exists():
        logger.warning("스키마 매핑 디렉토리 없음: %s", _SCHEMA_MAPPING_DIR)
        return jsonify(result)

    for yaml_file in sorted(_SCHEMA_MAPPING_DIR.glob("*.yaml")):
        mapping_id = yaml_file.stem
        friendly_name = _mapping_id_to_name(mapping_id)
        description = ""
        # 첫 줄 주석에서 설명 추출
        try:
            with yaml_file.open(encoding="utf-8") as fh:
                first_line = fh.readline().strip()
            if first_line.startswith("#"):
                description = first_line.lstrip("# ").strip()
        except Exception:
            pass
        result.append({
            "id": mapping_id,
            "name": friendly_name,
            "description": description,
        })

    return jsonify(result)


def _mapping_id_to_name(mapping_id: str) -> str:
    """파일 ID를 사람이 읽기 좋은 이름으로 변환한다."""
    name_map = {
        "auto_detect":    "표준 자동 추론",
        "standard_korean": "한국 일반 ERP",
        "sap_standard":   "SAP 표준",
        "duzon":          "더존 ERP",
        "default":        "기본 SAP",
    }
    return name_map.get(mapping_id, mapping_id.replace("_", " ").title())



# ── 룰 목록 ─────────────────────────────────────────────────────────────────
# 룰 메타데이터 (코드 순으로 정렬)
_RULE_META: list[dict] = [
    {
        "code": "A01",
        "name": "Data Integrity Test",
        "category": "integrity",
        "description": "GL 데이터의 필수 필드 존재 여부, 날짜·금액·계정코드 형식 유효성 검증",
        "default_enabled": True,
        "params_schema": {},
    },
    {
        "code": "A02",
        "name": "Transaction DR/CR Test",
        "category": "integrity",
        "description": "전표번호별 차변/대변 합계 일치 여부 검증",
        "default_enabled": True,
        "params_schema": {},
    },
    {
        "code": "A03",
        "name": "Trial Balance Rollforward",
        "category": "integrity",
        "description": "합계잔액시산표(TB)와 GL 데이터 계정별 잔액 일치 검증",
        "default_enabled": True,
        "params_schema": {},
    },
    {
        "code": "B01",
        "name": "Large P/L Items Test",
        "category": "risk",
        "description": "매출액 대비 중요성 금액(PM) 초과 손익계정 분개 적출",
        "default_enabled": True,
        "params_schema": {
            "materiality_ratio": {"type": "number", "default": 0.005, "label": "중요성 비율"},
        },
    },
    {
        "code": "B02",
        "name": "Unmatched Accounts Test",
        "category": "risk",
        "description": "COA 마스터에 미등록된 계정코드 사용 분개 적출",
        "default_enabled": True,
        "params_schema": {},
    },
    {
        "code": "B03",
        "name": "Newly Created Accounts Test",
        "category": "risk",
        "description": "당기 중 새로 생성된 계정과목 관련 분개 적출",
        "default_enabled": True,
        "params_schema": {},
    },
    {
        "code": "B04",
        "name": "Seldom Used Accounts",
        "category": "risk",
        "description": "사용 빈도 극히 낮은 계정과목 분개 적출",
        "default_enabled": True,
        "params_schema": {
            "max_usage_count": {"type": "integer", "default": 5, "label": "최대 사용 횟수"},
        },
    },
    {
        "code": "B05",
        "name": "Unusual User Test",
        "category": "risk",
        "description": "HR 대비 비정상 분개 입력자(퇴직자·미등록자) 적출",
        "default_enabled": True,
        "params_schema": {},
    },
    {
        "code": "B06",
        "name": "Inappropriate User Test",
        "category": "risk",
        "description": "처리자/승인자 직무분리 위반 사용자 분개 적출 (데이터 미입수 시 Waived)",
        "default_enabled": False,
        "params_schema": {
            "waived": {"type": "boolean", "default": True, "label": "Waived 처리"},
            "waive_reason": {"type": "string", "default": "승인자 컬럼 미포함", "label": "Waive 사유"},
        },
    },
    {
        "code": "B07",
        "name": "Back Dated Entries Test",
        "category": "risk",
        "description": "전기일보다 늦게 입력된 소급/역행 분개 적출",
        "default_enabled": True,
        "params_schema": {
            "max_delay_days": {"type": "integer", "default": 30, "label": "허용 지연일수"},
        },
    },
    {
        "code": "B08",
        "name": "전표유형-계정분석",
        "category": "risk",
        "description": "전표유형별 사용계정 조합의 이상 유무 검토",
        "default_enabled": True,
        "params_schema": {
            "min_frequency": {"type": "integer", "default": 2, "label": "최소 빈도"},
        },
    },
    {
        "code": "B09",
        "name": "상대계정분석",
        "category": "risk",
        "description": "전표 내 상대계정 조합 추출하여 이상 조합 검토",
        "default_enabled": True,
        "params_schema": {
            "min_frequency": {"type": "integer", "default": 2, "label": "최소 빈도"},
        },
    },
    {
        "code": "B10",
        "name": "적요 부재·짧음 검사",
        "category": "risk",
        "description": "ISA 240 §A43 — 적요가 없거나 너무 짧거나 의미 없는 키워드만 포함된 분개 적출",
        "default_enabled": True,
        "params_schema": {
            "min_length": {"type": "integer", "default": 5, "label": "최소 적요 길이"},
        },
    },
    {
        "code": "B11",
        "name": "결산일 인접 분개",
        "category": "risk",
        "description": "ISA 240 §A43 — 회계연도 종료일 ±10일 이내 분개 및 결산 후 소급입력 분개 적출",
        "default_enabled": True,
        "params_schema": {
            "days_before": {"type": "integer", "default": 10, "label": "결산일 전 일수"},
            "days_after": {"type": "integer", "default": 10, "label": "결산일 후 일수"},
        },
    },
    {
        "code": "B12",
        "name": "결산조정 분개",
        "category": "risk",
        "description": "ISA 240 §A43 — 결산조정 전표유형(SA·SU 등) 분개 중 결산일 인접·이후 분개 적출",
        "default_enabled": True,
        "params_schema": {
            "days_around_period_end": {"type": "integer", "default": 10, "label": "결산일 전후 일수"},
        },
    },
]


def _ensure_rules_loaded() -> None:
    """룰 클래스들을 임포트하여 Rule._registry에 등록한다."""
    try:
        import jet.domain.rules.a01_integrity  # noqa: F401
        import jet.domain.rules.a02_dr_cr_balance  # noqa: F401
        import jet.domain.rules.a03_tb_rollforward  # noqa: F401
        import jet.domain.rules.b01_large_pl  # noqa: F401
        import jet.domain.rules.b02_unmatched_account  # noqa: F401
        import jet.domain.rules.b03_new_account  # noqa: F401
        import jet.domain.rules.b04_seldom_used  # noqa: F401
        import jet.domain.rules.b05_unusual_user  # noqa: F401
        import jet.domain.rules.b06_inappropriate_user  # noqa: F401
        import jet.domain.rules.b07_backdated_entry  # noqa: F401
        import jet.domain.rules.b08_doc_type_account  # noqa: F401
        import jet.domain.rules.b09_counter_account  # noqa: F401
        import jet.domain.rules.b10_no_description  # noqa: F401
        import jet.domain.rules.b11_period_end_proximity  # noqa: F401
        import jet.domain.rules.b12_top_side_adjustment  # noqa: F401
    except Exception as exc:
        logger.warning("룰 임포트 중 일부 실패: %s", exc)


@app.route("/api/rules")
def list_rules():
    """등록된 모든 룰 메타데이터를 반환한다."""
    return jsonify(_RULE_META)


# ── GL 미리보기 + 자동 매핑 추천 ─────────────────────────────────────────────
@app.route("/api/upload-gl-preview", methods=["POST"])
def upload_gl_preview():
    """GL 파일을 업로드받아 헤더 + 상위 10행 + 자동 매핑 추천을 반환한다.

    Request: multipart/form-data
        gl_file: GL 파일 (첫 번째 파일만 미리보기 대상)

    Response:
        headers: 원본 헤더 목록
        preview_rows: 상위 10행 (list of list)
        suggested_mapping: {표준필드: 매칭컬럼 or null}
        unmapped_required: 매칭 실패한 필수 필드 목록
        dr_cr_format: 자동 감지된 차대변 형식
    """
    from jet.application.pipeline.column_aliases import suggest_mapping, detect_dr_cr_format
    from jet.application.pipeline.schema_mapper import REQUIRED_FIELDS

    gl_files_raw = request.files.getlist("gl_file")
    if not gl_files_raw:
        return jsonify({"error": "GL 파일을 업로드해 주세요"}), 400

    gl_file = gl_files_raw[0]
    if not gl_file.filename:
        return jsonify({"error": "GL 파일을 선택해 주세요"}), 400

    # 임시 저장
    preview_dir = _DATA_DIR / "preview_tmp"
    preview_dir.mkdir(parents=True, exist_ok=True)
    tmp_path = preview_dir / (str(uuid.uuid4()) + Path(gl_file.filename).suffix.lower())
    gl_file.save(str(tmp_path))

    try:
        from jet.infrastructure.io.gl_loader_factory import load_gl_files
        raw_df = load_gl_files([tmp_path], cache_dir=None)
    except Exception as exc:
        tmp_path.unlink(missing_ok=True)
        return jsonify({"error": f"파일 읽기 실패: {exc}"}), 400
    finally:
        tmp_path.unlink(missing_ok=True)

    headers = list(raw_df.columns)

    # 상위 10행 직렬화
    preview_rows = []
    for _, row in raw_df.head(10).iterrows():
        serialized = []
        for v in row:
            if v is None:
                serialized.append(None)
            elif hasattr(v, "isoformat"):
                serialized.append(v.isoformat())
            else:
                serialized.append(str(v) if str(v) not in ("nan", "None", "") else None)
        preview_rows.append(serialized)

    # 자동 매핑 추천
    suggested = suggest_mapping(headers)
    dr_cr_format = detect_dr_cr_format(suggested)

    # 필수 미매칭 계산 (debit_amount/credit_amount는 signed_amount+indicator 콤비로도 OK)
    unmapped_required = []
    for req_field in REQUIRED_FIELDS:
        if suggested.get(req_field) is None:
            # debit/credit은 single_signed/single_indicator 형식이면 OK
            if req_field in ("debit_amount", "credit_amount"):
                if dr_cr_format != "split":
                    continue
            unmapped_required.append(req_field)

    return jsonify({
        "headers": headers,
        "preview_rows": preview_rows,
        "suggested_mapping": suggested,
        "unmapped_required": unmapped_required,
        "dr_cr_format": dr_cr_format,
        "total_rows": len(raw_df),
    })


# ── 새 실행 시작 (POST /api/runs) ───────────────────────────────────────────
@app.route("/api/runs", methods=["POST"])
def create_run():
    """GL 파일 업로드 후 백그라운드 JET 실행을 시작한다."""
    import json

    # GL 파일 필수 검증 (단일 또는 다중)
    _SUPPORTED_GL_EXTS = {".xlsx", ".xls", ".csv", ".txt", ".parquet", ".zip"}
    gl_files_raw = request.files.getlist("gl_file")
    if not gl_files_raw:
        return jsonify({"error": "GL 파일을 업로드해 주세요 (gl_file 필드 필수)"}), 400

    # 빈 파일명 제거
    gl_files_raw = [f for f in gl_files_raw if f.filename]
    if not gl_files_raw:
        return jsonify({"error": "GL 파일을 선택해 주세요"}), 400

    # 확장자 검증
    for gf in gl_files_raw:
        gl_ext_check = Path(gf.filename).suffix.lower()
        if gl_ext_check not in _SUPPORTED_GL_EXTS:
            return jsonify({"error": f"지원하지 않는 GL 파일 형식: {gl_ext_check}. 지원: {sorted(_SUPPORTED_GL_EXTS)}"}), 400

    # run_id 생성 및 디렉토리 준비
    run_id = str(uuid.uuid4())
    run_dir = _DATA_DIR / run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    # 파일 저장 (다중 지원)
    gl_paths: list[Path] = []
    for i, gf in enumerate(gl_files_raw):
        gl_ext = Path(gf.filename).suffix.lower()
        gl_path = run_dir / f"gl_{i:02d}{gl_ext}"
        gf.save(str(gl_path))
        gl_paths.append(gl_path)

    gl_path = gl_paths[0]  # 하위 호환용

    # 선택 파일들 저장
    optional_files: dict[str, Path | None] = {
        "hr": None,
        "coa": None,
        "tb": None,
        "doctype": None,
    }
    for key in optional_files:
        if key + "_file" in request.files:
            f = request.files[key + "_file"]
            if f.filename:
                ext = Path(f.filename).suffix.lower()
                saved = run_dir / f"{key}{ext}"
                f.save(str(saved))
                optional_files[key] = saved

    # workpaper_spec 파싱
    workpaper_spec_raw = request.form.get("workpaper_spec", "{}")
    try:
        workpaper_spec = json.loads(workpaper_spec_raw)
    except json.JSONDecodeError:
        workpaper_spec = {}

    # 기본값 채우기
    workpaper_spec.setdefault("company", "")
    workpaper_spec.setdefault("period_end", str(date.today().replace(month=12, day=31).replace(year=date.today().year - 1)))
    workpaper_spec.setdefault("preparer", "")
    workpaper_spec.setdefault("reviewer", "")
    workpaper_spec.setdefault("workpaper_code", "7400")
    workpaper_spec.setdefault("title", "")

    schema_mapping_id = request.form.get("schema_mapping_id", "auto_detect")

    # rule_overrides: {"A01": true, "B06": false, ...}
    rule_overrides_raw = request.form.get("rule_overrides", "{}")
    try:
        rule_overrides = json.loads(rule_overrides_raw)
    except json.JSONDecodeError:
        rule_overrides = {}

    # rule_params: {"B07": {"max_delay_days": 30}, ...}
    rule_params_raw = request.form.get("rule_params", "{}")
    try:
        rule_params = json.loads(rule_params_raw)
    except json.JSONDecodeError:
        rule_params = {}

    # 실행 상태 초기화
    with _runs_lock:
        _runs[run_id] = {
            "status": "running",
            "progress": {"stage": "loading", "pct": 0},
            "total_rows": 0,
            "normalized_rows": 0,
            "rules_done": [],
            "rules_total": 0,
            "result": None,
            "error": None,
            "created_at": datetime.utcnow().isoformat(),
            "output_path": None,
        }

    # column_mapping_overrides: 사용자가 UI에서 조정한 매핑
    # {"entry_no": "전표번호", "user_id": "사원코드", ...}
    column_mapping_raw = request.form.get("column_mapping", "{}")
    try:
        column_mapping_overrides = json.loads(column_mapping_raw)
    except json.JSONDecodeError:
        column_mapping_overrides = {}

    # 백그라운드 스레드로 실행
    t = threading.Thread(
        target=_run_jet_background,
        args=(
            run_id,
            run_dir,
            gl_paths,
            optional_files,
            schema_mapping_id,
            workpaper_spec,
            rule_overrides,
            rule_params,
            column_mapping_overrides,
        ),
        daemon=True,
    )
    t.start()

    return jsonify({"run_id": run_id}), 202


def _update_run(run_id: str, **kwargs: Any) -> None:
    """실행 상태를 스레드 안전하게 갱신한다."""
    with _runs_lock:
        if run_id in _runs:
            _runs[run_id].update(kwargs)


def _run_jet_background(
    run_id: str,
    run_dir: Path,
    gl_paths: list[Path],
    optional_files: dict[str, Path | None],
    schema_mapping_id: str,
    workpaper_spec: dict,
    rule_overrides: dict,
    rule_params: dict,
    column_mapping_overrides: dict | None = None,
) -> None:
    """백그라운드에서 JET 파이프라인을 실행한다.

    진행 단계: loading → normalizing → rule:A01 → ... → reporting → done
    """
    import yaml
    from jet.application.pipeline.normalizer import Normalizer
    from jet.application.pipeline.schema_mapper import SchemaMapper
    from jet.application.workpaper.template_loader import TemplateLoader
    from jet.application.workpaper.workpaper_spec import WorkpaperSpec, ScenarioSpec
    from jet.domain.rules.base import RuleContext
    from jet.infrastructure.reporters.excel_reporter import ExcelReporter

    start_time = datetime.utcnow()

    try:
        # ── 1단계: 룰 임포트 및 스키마 로드 ─────────────────────────────────
        _update_run(run_id, progress={"stage": "loading", "pct": 5})
        _ensure_rules_loaded()

        # 스키마 매핑 로드
        schema_yaml_path = _SCHEMA_MAPPING_DIR / f"{schema_mapping_id}.yaml"
        if not schema_yaml_path.exists():
            schema_yaml_path = _SCHEMA_MAPPING_DIR / "auto_detect.yaml"
        if not schema_yaml_path.exists():
            schema_yaml_path = _SCHEMA_MAPPING_DIR / "default.yaml"
        if not schema_yaml_path.exists():
            raise FileNotFoundError(f"스키마 매핑 파일을 찾을 수 없습니다: {schema_mapping_id}")

        with schema_yaml_path.open(encoding="utf-8") as fh:
            schema_dict = yaml.safe_load(fh)

        # 사용자 UI 매핑 오버라이드 적용
        if column_mapping_overrides:
            if "mappings" not in schema_dict:
                schema_dict["mappings"] = {}
            schema_dict["mappings"].update(column_mapping_overrides)

        # auto_detect 모드: 별칭 사전으로 매핑 자동 추론 (GL 로딩 전에는 헤더 불명)
        # → GL 로드 후 suggest_mapping()으로 채운다 (3단계에서 처리)
        is_auto_detect = schema_dict.get("auto_detect", False)

        mapper = SchemaMapper(schema_dict) if not is_auto_detect else None

        # ── 2단계: 마스터 파일 적재 ──────────────────────────────────────────
        _update_run(run_id, progress={"stage": "loading", "pct": 10})

        coa_master = None
        doc_type_master = None
        hr_master = None
        tb_master = None
        user_df = None

        if optional_files.get("coa") and optional_files["coa"].exists():
            from jet.infrastructure.io.coa_loader import CoaLoader
            coa_master = CoaLoader().load(optional_files["coa"])

        if optional_files.get("doctype") and optional_files["doctype"].exists():
            from jet.infrastructure.io.doc_type_loader import DocTypeLoader
            doc_type_master = DocTypeLoader().load(optional_files["doctype"])

        if optional_files.get("hr") and optional_files["hr"].exists():
            from jet.infrastructure.io.hr_loader import HRLoader
            hr_master = HRLoader().load(optional_files["hr"])

        if optional_files.get("tb") and optional_files["tb"].exists():
            from jet.infrastructure.io.tb_loader import TbLoader
            tb_master = TbLoader().load(optional_files["tb"])

        # ── 3단계: GL 데이터 적재 ────────────────────────────────────────────
        _update_run(run_id, progress={"stage": "loading", "pct": 20})

        from jet.infrastructure.io.gl_loader_factory import load_gl_files
        cache_dir = run_dir / "cache"
        raw_df = load_gl_files(gl_paths, cache_dir=cache_dir)

        # auto_detect 모드: 헤더를 보고 자동 매핑 추론
        if is_auto_detect or mapper is None:
            from jet.application.pipeline.column_aliases import suggest_mapping, detect_dr_cr_format
            headers = list(raw_df.columns)
            suggested = suggest_mapping(headers)
            # column_mapping_overrides 재적용 (사용자 조정 우선)
            if column_mapping_overrides:
                suggested.update(column_mapping_overrides)
            dr_cr_fmt = detect_dr_cr_format(suggested)
            # suggested에서 표준 14종만 추출 (signed_amount, dr_cr_indicator 제외)
            from jet.application.pipeline.schema_mapper import ALL_STANDARD_FIELDS
            final_mappings = {f: suggested.get(f) for f in ALL_STANDARD_FIELDS}
            schema_dict["mappings"] = final_mappings
            schema_dict["dr_cr_format"] = dr_cr_fmt
            schema_dict.pop("auto_detect", None)
            mapper = SchemaMapper(schema_dict)

        total_rows = len(raw_df)
        _update_run(run_id, total_rows=total_rows, progress={"stage": "loading", "pct": 30})

        # ── 4단계: 정규화 ─────────────────────────────────────────────────────
        _update_run(run_id, progress={"stage": "normalizing", "pct": 35})

        mapped_df = mapper.map(raw_df)
        normalizer = Normalizer(
            amount_unit=mapper.amount_unit,
            account_code_pad=mapper.account_code_pad,
        )
        entries, norm_report = normalizer.normalize(mapped_df)
        normalized_rows = norm_report.success_count
        _update_run(
            run_id,
            normalized_rows=normalized_rows,
            progress={"stage": "normalizing", "pct": 40},
        )

        # ── 5단계: 룰 실행 컨텍스트 조립 ────────────────────────────────────
        period_end_str = workpaper_spec.get("period_end", "2025-12-31")
        try:
            period_end = date.fromisoformat(period_end_str)
        except (ValueError, TypeError):
            period_end = date(2025, 12, 31)
        period_start = period_end.replace(month=1, day=1)

        context = RuleContext(
            period_start=period_start,
            period_end=period_end,
            coa_master=coa_master,
            doc_type_master=doc_type_master,
            hr_master=hr_master,
            tb_master=tb_master,
        )

        # ── 6단계: 실행할 룰 결정 ────────────────────────────────────────────
        from jet.interface.cli.commands.run import _RULE_CLASSES

        # 기본 활성 룰 계산
        enabled_by_default = {
            m["code"] for m in _RULE_META if m["default_enabled"]
        }
        # rule_overrides 적용
        enabled_codes: set[str] = set()
        for code in _RULE_CLASSES:
            if code in rule_overrides:
                if rule_overrides[code]:
                    enabled_codes.add(code)
            elif code in enabled_by_default:
                enabled_codes.add(code)

        # B06는 항상 실행 (waive 시트 생성용)
        all_exec_codes = enabled_codes | {"B06"}
        rules_total = len(all_exec_codes)
        _update_run(run_id, rules_total=rules_total)

        # ── 7단계: 룰별 실행 ─────────────────────────────────────────────────
        results: dict = {}
        rules_done: list[str] = []

        # 진행률 계산: 40% ~ 85% 구간을 룰 실행에 할당
        rule_pct_start = 40
        rule_pct_range = 45

        for idx, code in enumerate(sorted(all_exec_codes)):
            rule_cls = _RULE_CLASSES.get(code)
            if rule_cls is None:
                continue

            pct = rule_pct_start + int(rule_pct_range * idx / max(rules_total, 1))
            _update_run(run_id, progress={"stage": f"rule:{code}", "pct": pct})

            rule = rule_cls()
            params = dict(rule_params.get(code, {}))

            # B07 파라미터 기본값 주입
            if code == "B07" and "max_delay_days" not in params:
                params["max_delay_days"] = 30
            # B04 파라미터 기본값 주입
            if code == "B04" and "max_usage_count" not in params:
                params["max_usage_count"] = 5
            # B01 파라미터 기본값 주입
            if code == "B01":
                params.setdefault("materiality_ratio", 0.005)
            # B06 waive 처리
            if code == "B06":
                params.setdefault("waived", True)
                params.setdefault("waive_reason", "승인자 컬럼 미포함")
            # B08·B09는 v2.0에서 min_frequency 미사용 (B01 매출 임계치·계정명 키워드 사용)
            # B10 파라미터 기본값 주입
            if code == "B10":
                params.setdefault("min_length", 5)
            # B11 파라미터 기본값 주입
            if code == "B11":
                params.setdefault("days_before", 10)
                params.setdefault("days_after", 10)
            # B12 파라미터 기본값 주입
            if code == "B12":
                params.setdefault("days_around_period_end", 10)

            rule.configure(params)
            result = rule.apply(entries, context)
            results[code] = result
            rules_done.append(code)
            _update_run(run_id, rules_done=list(rules_done))

        # ── 8단계: Excel 조서 생성 ──────────────────────────────────────────
        _update_run(run_id, progress={"stage": "reporting", "pct": 87})

        # WorkpaperSpec 동적 생성
        spec = _build_workpaper_spec(
            workpaper_spec, rule_overrides, rule_params, enabled_codes
        )

        master_data = {
            "coa": coa_master,
            "hr": hr_master,
            "doc_types": doc_type_master,
            "tb": tb_master,
            "user_df": user_df,
            "entries": entries,  # Stats_AutoManual 시트용
        }

        output_path = run_dir / "output.xlsx"
        reporter = ExcelReporter()
        reporter.write(
            spec=spec,
            results=results,
            output_path=output_path,
            master_data=master_data,
        )

        # ── 9단계: 결과 집계 ─────────────────────────────────────────────────
        _update_run(run_id, progress={"stage": "reporting", "pct": 95})

        duration_sec = (datetime.utcnow() - start_time).total_seconds()
        total_findings = sum(r.finding_count for r in results.values())

        scenarios_summary = []
        for code in sorted(results.keys()):
            r = results[code]
            finding_cnt = r.finding_count
            if finding_cnt == 0:
                status = "pass"
            elif code == "B06":
                status = "waived"
                finding_cnt = 0
            else:
                status = "fail"
            rule_meta = next((m for m in _RULE_META if m["code"] == code), None)
            scenario_entry = {
                "code": code,
                "name": rule_meta["name"] if rule_meta else code,
                "findings": finding_cnt,
                "status": status,
            }
            # 룰별 경고 surface — pass 시 warn으로 승격
            warning_msg = None
            if r.extra.get("sales_identification_failed"):
                warning_msg = (
                    "매출 식별 실패 — 계정과목명 키워드 매칭 0건. "
                    "임계치 0원으로 검증 skip. absolute_threshold 주입 검토."
                )
            elif r.extra.get("coa_missing_waive"):
                warning_msg = (
                    r.extra.get("coa_missing_reason")
                    or "COA 마스터 미제공 — 미등록 계정 적출 불가, 면제 처리."
                )
            elif r.extra.get("skipped_reason") == "gl_entries_empty":
                warning_msg = "GL 분개 0건 — 입력 데이터 정상성 확인 필요."
            if warning_msg:
                scenario_entry["warning"] = warning_msg
                if status == "pass":
                    scenario_entry["status"] = "warn"
            scenarios_summary.append(scenario_entry)

        run_result = {
            "run_id": run_id,
            "status": "done",
            "summary": {
                "total_rows": total_rows,
                "normalized_rows": normalized_rows,
                "total_findings": total_findings,
                "duration_sec": round(duration_sec, 1),
                "output_size_mb": round(output_path.stat().st_size / 1024 / 1024, 2) if output_path.exists() else 0,
            },
            "scenarios": scenarios_summary,
            "output_path": str(output_path),
        }

        _update_run(
            run_id,
            status="done",
            progress={"stage": "done", "pct": 100},
            result=run_result,
            output_path=str(output_path),
        )

        # 결과 캐시에 RuleResult 저장 (미리보기용)
        with _runs_lock:
            _runs[run_id]["_rule_results"] = results

        logger.info("JET 실행 완료: run_id=%s, duration=%.1f초", run_id, duration_sec)

    except Exception as exc:
        tb_str = traceback.format_exc()
        logger.exception("JET 실행 실패: run_id=%s", run_id)
        _update_run(
            run_id,
            status="failed",
            progress={"stage": "failed", "pct": 0},
            error={"message": str(exc), "traceback": tb_str[-2000:]},
        )


def _build_workpaper_spec(
    spec_dict: dict,
    rule_overrides: dict,
    rule_params: dict,
    enabled_codes: set,
) -> "WorkpaperSpec":
    """웹 폼 입력으로부터 WorkpaperSpec 객체를 생성한다."""
    from jet.application.workpaper.workpaper_spec import ScenarioSpec, WorkpaperSpec

    try:
        period_end = date.fromisoformat(spec_dict.get("period_end", "2025-12-31"))
    except (ValueError, TypeError):
        period_end = date(2025, 12, 31)

    scenarios = []
    for meta in _RULE_META:
        code = meta["code"]
        is_enabled = code in enabled_codes
        params = dict(rule_params.get(code, {}))
        if code == "B06":
            params.setdefault("waived", True)
            params.setdefault("waive_reason", "승인자 컬럼 미포함")
        scenarios.append(
            ScenarioSpec(
                code=code,
                name=meta["name"],
                objective=meta["description"],
                rule=code,
                enabled=is_enabled,
                params=params,
            )
        )

    return WorkpaperSpec(
        company=spec_dict.get("company", ""),
        period_end=period_end,
        preparer=spec_dict.get("preparer", ""),
        reviewer=spec_dict.get("reviewer", ""),
        prepared_date=date.today(),
        reviewed_date=date.today(),
        workpaper_code=spec_dict.get("workpaper_code", "7400"),
        title=spec_dict.get("title", ""),
        scenarios=scenarios,
        master_files={},
    )


# ── 실행 상태 폴링 ────────────────────────────────────────────────────────────
@app.route("/api/runs/<run_id>/status")
def get_run_status(run_id: str):
    """실행 진행 상황을 반환한다."""
    with _runs_lock:
        run = _runs.get(run_id)
    if run is None:
        return jsonify({"error": "실행 ID를 찾을 수 없습니다"}), 404

    return jsonify({
        "run_id": run_id,
        "status": run["status"],
        "progress": run["progress"],
        "total_rows": run["total_rows"],
        "normalized_rows": run["normalized_rows"],
        "rules_done": run["rules_done"],
        "rules_total": run["rules_total"],
        "error": run.get("error"),
    })


# ── 실행 결과 ─────────────────────────────────────────────────────────────────
@app.route("/api/runs/<run_id>/result")
def get_run_result(run_id: str):
    """완료된 실행의 결과 요약을 반환한다."""
    with _runs_lock:
        run = _runs.get(run_id)
    if run is None:
        return jsonify({"error": "실행 ID를 찾을 수 없습니다"}), 404
    if run["status"] != "done":
        return jsonify({"error": f"아직 완료되지 않았습니다 (status: {run['status']})"}), 400

    return jsonify(run["result"])


# ── 시트 미리보기 ─────────────────────────────────────────────────────────────
@app.route("/api/runs/<run_id>/preview/<sheet>")
def preview_sheet(run_id: str, sheet: str):
    """출력 Excel의 특정 시트 상위 100행을 반환한다."""
    with _runs_lock:
        run = _runs.get(run_id)
    if run is None:
        return jsonify({"error": "실행 ID를 찾을 수 없습니다"}), 404
    if run["status"] != "done":
        return jsonify({"error": "아직 완료되지 않았습니다"}), 400

    output_path = run.get("output_path")
    if not output_path or not Path(output_path).exists():
        return jsonify({"error": "출력 파일을 찾을 수 없습니다"}), 404

    try:
        import openpyxl
        wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            # 부분 일치 시도
            matched = [s for s in wb.sheetnames if sheet in s or s in sheet]
            if matched:
                sheet = matched[0]
            else:
                return jsonify({"error": f"시트를 찾을 수 없습니다: {sheet}", "available": wb.sheetnames}), 404

        ws = wb[sheet]
        rows_raw = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows_raw:
            return jsonify({"sheet": sheet, "total_rows": 0, "columns": [], "rows": []})

        # 첫 행을 헤더로 사용
        headers = [str(c) if c is not None else "" for c in rows_raw[0]]
        data_rows = rows_raw[1:101]  # 상위 100행
        total_rows = max(0, len(rows_raw) - 1)

        # 직렬화 (date/datetime → str)
        def _serialize_cell(v: Any) -> Any:
            if v is None:
                return None
            if isinstance(v, (date, datetime)):
                return v.isoformat()
            return v

        serialized_rows = [
            [_serialize_cell(c) for c in row]
            for row in data_rows
        ]

        return jsonify({
            "sheet": sheet,
            "total_rows": total_rows,
            "columns": headers,
            "rows": serialized_rows,
        })

    except Exception as exc:
        logger.exception("시트 미리보기 실패: run_id=%s, sheet=%s", run_id, sheet)
        return jsonify({"error": str(exc)}), 500


# ── 시트 목록 ─────────────────────────────────────────────────────────────────
@app.route("/api/runs/<run_id>/sheets")
def list_sheets(run_id: str):
    """출력 Excel의 시트 목록을 반환한다."""
    with _runs_lock:
        run = _runs.get(run_id)
    if run is None:
        return jsonify({"error": "실행 ID를 찾을 수 없습니다"}), 404
    if run["status"] != "done":
        return jsonify({"error": "아직 완료되지 않았습니다"}), 400

    output_path = run.get("output_path")
    if not output_path or not Path(output_path).exists():
        return jsonify({"error": "출력 파일을 찾을 수 없습니다"}), 404

    try:
        import openpyxl
        wb = openpyxl.load_workbook(output_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return jsonify({"sheets": sheets})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


# ── 조서 Excel 다운로드 ───────────────────────────────────────────────────────
@app.route("/api/runs/<run_id>/download")
def download_result(run_id: str):
    """완료된 실행의 조서 Excel 파일을 다운로드한다."""
    with _runs_lock:
        run = _runs.get(run_id)
    if run is None:
        return jsonify({"error": "실행 ID를 찾을 수 없습니다"}), 404
    if run["status"] != "done":
        return jsonify({"error": "아직 완료되지 않았습니다"}), 400

    output_path = run.get("output_path")
    if not output_path or not Path(output_path).exists():
        return jsonify({"error": "출력 파일을 찾을 수 없습니다"}), 404

    # 다운로드 파일명: 회사명_조서번호_날짜.xlsx
    result = run.get("result", {})
    company = ""
    if result and isinstance(result.get("scenarios"), list):
        # run 결과에서 company 정보 추출 불필요 — workpaper_spec에서 가져옴
        pass

    download_name = f"JET_감사조서_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"

    return send_file(
        output_path,
        as_attachment=True,
        download_name=download_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── 실행 삭제 ─────────────────────────────────────────────────────────────────
@app.route("/api/runs/<run_id>", methods=["DELETE"])
def delete_run(run_id: str):
    """실행 데이터(업로드/출력 파일)를 삭제한다 (개인정보 보호)."""
    with _runs_lock:
        run = _runs.pop(run_id, None)

    if run is None:
        return jsonify({"error": "실행 ID를 찾을 수 없습니다"}), 404

    run_dir = _DATA_DIR / run_id
    try:
        if run_dir.exists():
            shutil.rmtree(str(run_dir))
    except Exception as exc:
        logger.warning("run_dir 삭제 실패: %s — %s", run_dir, exc)

    return jsonify({"message": f"실행 {run_id} 삭제 완료"})


# ── 오래된 실행 정리 (7일 초과) ─────────────────────────────────────────────
def _cleanup_old_runs() -> None:
    """7일 이상 지난 실행 디렉토리를 삭제한다 (Lazy 방식)."""
    cutoff = datetime.utcnow() - timedelta(days=7)
    try:
        if not _DATA_DIR.exists():
            return
        for run_dir in _DATA_DIR.iterdir():
            if not run_dir.is_dir():
                continue
            try:
                mtime = datetime.utcfromtimestamp(run_dir.stat().st_mtime)
                if mtime < cutoff:
                    shutil.rmtree(str(run_dir))
                    with _runs_lock:
                        _runs.pop(run_dir.name, None)
                    logger.info("오래된 실행 정리: %s", run_dir.name)
            except Exception:
                pass
    except Exception:
        pass


@app.before_request
def _lazy_cleanup() -> None:
    """요청 시마다 1% 확률로 오래된 실행을 정리한다."""
    import random
    if random.random() < 0.01:
        threading.Thread(target=_cleanup_old_runs, daemon=True).start()
