import openpyxl
import shutil
from pathlib import Path
from src.infrastructure.excel_writer.ac_filler import ACFiller, SHEET_CONFIG
from src.infrastructure.excel_writer.color_swap import apply_toss_palette

ROOT = Path(__file__).resolve().parents[2]
TEMPLATE = ROOT / "templates" / "4150_AC_template.xlsx"

def test_borders_and_merges_preserved(tmp_path):
    out = tmp_path / "out.xlsx"
    shutil.copy(TEMPLATE, out)
    src_wb = openpyxl.load_workbook(TEMPLATE)
    src_merges = {s: list(src_wb[s].merged_cells.ranges) for s in src_wb.sheetnames}
    src_wb.close()
    filler = ACFiller(out)
    filler.fill_section("AC1", [])
    apply_toss_palette(filler.wb)
    filler.save()
    out_wb = openpyxl.load_workbook(out)
    for s in out_wb.sheetnames:
        if s not in src_merges: continue
        out_merges = list(out_wb[s].merged_cells.ranges)
        assert len(out_merges) == len(src_merges[s]), f"merge 영역 변경됨: {s}"
    out_wb.close()
