import shutil
import openpyxl
from pathlib import Path
from src.infrastructure.excel_writer.color_swap import apply_toss_palette

ROOT = Path(__file__).resolve().parents[2]
TEMPLATE = ROOT / "templates" / "4150_AC_template.xlsx"

def test_swap_title_to_toss_blue(tmp_path):
    out = tmp_path / "out.xlsx"
    shutil.copy(TEMPLATE, out)
    wb = openpyxl.load_workbook(out)
    apply_toss_palette(wb)
    wb.save(out)
    wb2 = openpyxl.load_workbook(out)
    ac1 = [s for s in wb2.sheetnames if s.startswith("AC1.")][0]
    title_cell = wb2[ac1]["A2"]
    fg = title_cell.fill.fgColor.rgb if title_cell.fill and title_cell.fill.fgColor else None
    assert fg and "3182F6" in (fg or "").upper()
