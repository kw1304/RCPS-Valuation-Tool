import shutil
import openpyxl
from pathlib import Path
from decimal import Decimal
from datetime import date
from src.infrastructure.excel_writer.ac_filler import ACFiller
from src.domain.ac_models import FinancialAsset, Borrowing

ROOT = Path(__file__).resolve().parents[2]
TEMPLATE = ROOT / "templates" / "4150_AC_template.xlsx"

def test_fill_ac1_writes_to_row_11(tmp_path):
    out = tmp_path / "out.xlsx"
    shutil.copy(TEMPLATE, out)
    filler = ACFiller(out)
    recs = [
        FinancialAsset(bc_no="BC-1", bank="국민은행", asset_type="deposit",
                       product="보통예금-내맘대로통장", account_no="0936-01-01",
                       currency="KRW", balance=Decimal("10218")),
    ]
    filler.fill_section("AC1", recs)
    filler.save()
    wb = openpyxl.load_workbook(out, data_only=False)
    ws = [s for s in wb.sheetnames if s.startswith("AC1.") or s.startswith("AC1 ")][0]
    assert wb[ws]["C11"].value == "BC-1"
    assert wb[ws]["D11"].value == "국민은행"
    assert wb[ws]["H11"].value == 10218

def test_fill_ac2_writes_to_row_12(tmp_path):
    out = tmp_path / "out2.xlsx"
    shutil.copy(TEMPLATE, out)
    filler = ACFiller(out)
    recs = [
        Borrowing(bc_no="BC-2", bank="기업은행", contract_type="일반자금대출",
                  limit_amt=Decimal("14500000000"), limit_ccy="KRW",
                  balance=Decimal("14500000000"), balance_ccy="KRW",
                  contract_date=date(2025, 6, 10))
    ]
    filler.fill_section("AC2", recs)
    filler.save()
    wb = openpyxl.load_workbook(out, data_only=False)
    ws = [s for s in wb.sheetnames if s.startswith("AC2.") or s.startswith("AC2 ")][0]
    assert wb[ws]["C12"].value == "BC-2"
