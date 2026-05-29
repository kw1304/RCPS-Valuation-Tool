import json
import shutil
from datetime import datetime
from pathlib import Path
from sqlmodel import Session, select
import openpyxl
from openpyxl.cell.cell import MergedCell
from src.infrastructure.db.models import Project, Counterparty, ExtractedRecord, FileAsset
from src.infrastructure.excel_writer.ac_filler import ACFiller, SHEET_CONFIG
from src.infrastructure.excel_writer.color_swap import apply_toss_palette, mark_low_confidence
from src.infrastructure.cs_loader import ControlSheetLoader
from src.infrastructure.union_monthly import parse_collateral_or_guarantee, parse_union_monthly
from src.domain.ac_models import (
    FinancialAsset, Borrowing, Derivative, Guarantee,
    Collateral, BillCheck, Insurance, GeneralDeal,
)
from src.domain.party_normalize import PartyNormalizer

ROOT = Path(__file__).resolve().parents[2]
TEMPLATE = ROOT / "templates" / "4150_AC_template.xlsx"
OUTPUT_DIR = ROOT / "OUTPUT"

MODEL_BY_SECTION = {
    "AC1": FinancialAsset, "AC2": Borrowing, "AC3": Derivative,
    "AC4": Guarantee, "AC5": Collateral, "AC6": BillCheck,
    "AC7": Insurance, "AC8": GeneralDeal,
}


def export_4150(session: Session, project_id: int) -> Path:
    """Export 4150 AC workpaper from extracted records and counterparty data."""
    project = session.get(Project, project_id)
    if project is None:
        raise ValueError("project not found")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fy = project.fiscal_date[:4]
    out_path = OUTPUT_DIR / f"4150_AC_금융기관조회_{project.name}_FY{fy}_{ts}.xlsx"

    # Copy template
    shutil.copy(TEMPLATE, out_path)
    filler = ACFiller(out_path)

    # Stamp project info (회사명·기준일) — propagates via formulas to AC1~AC10
    _stamp_project_info(filler.wb, project.name, project.fiscal_date)

    # Clear V1 example data rows (예: V1엔 코스맥스비티아이 전기 데이터가 채워져 있음)
    _clear_data_rows(filler.wb)

    # Get counterparties + file assets
    cps = list(session.exec(
        select(Counterparty).where(Counterparty.project_id == project_id)
    ).all())
    files = {f.kind: f for f in session.exec(
        select(FileAsset).where(FileAsset.project_id == project_id)
    ).all()}
    norm = PartyNormalizer.load(ROOT / "configs")

    # Fill AC control sheet + AC0
    _fill_control_sheet(filler.wb, cps)
    _fill_ac0(filler.wb, cps, files, norm)

    # Fill AC1~AC8: ExtractedRecord
    for ac in MODEL_BY_SECTION:
        records_raw = session.exec(
            select(ExtractedRecord).where(
                ExtractedRecord.project_id == project_id,
                ExtractedRecord.ac_section == ac,
            )
        ).all()

        Model = MODEL_BY_SECTION[ac]
        models = [Model.model_validate_json(r.payload_json) for r in records_raw]

        filler.fill_section(ac, models)

        cfg = SHEET_CONFIG[ac]
        ws = filler.wb[cfg["sheet_name"]] if cfg["sheet_name"] in filler.wb.sheetnames else None
        if ws:
            for idx, raw in enumerate(records_raw):
                if raw.confidence == "low":
                    for col in cfg["cols"].keys():
                        mark_low_confidence(ws, cfg["start_row"] + idx, col)

    # Apply Toss palette
    apply_toss_palette(filler.wb)
    filler.save()

    return out_path


def _safe_write(sheet, cell_ref: str, value) -> None:
    """Merged cell 에 쓰기 시도 시 조용히 skip."""
    cell = sheet[cell_ref]
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def _fill_control_sheet(wb, cps: list[Counterparty]):
    """Fill AC control sheet with counterparty summary."""
    sheet = next((wb[s] for s in wb.sheetnames if "control sheet" in s.lower()), None)
    if sheet is None:
        return

    for i, cp in enumerate(cps):
        r = 6 + i
        _safe_write(sheet, f"B{r}", cp.bc_no)
        _safe_write(sheet, f"C{r}", cp.canonical_name)
        if cp.branch:
            _safe_write(sheet, f"D{r}", cp.branch)
        _safe_write(sheet, f"E{r}", cp.channel or "")
        _safe_write(sheet, f"F{r}", cp.address or "")
        _safe_write(sheet, f"J{r}", "회신" if cp.response_arrived else "미회신")


def _fill_ac0(wb, cps: list[Counterparty], files: dict, norm):
    """V1 AC0 시트의 5개 sub-section을 각각 fill.

    V1 AC0 구조:
      R11    Section 1 header: 전기 금융조회 대상 / 회사 list 포함? / 제외사유
      R12+   Section 1 data (전기 CS list)
      R47    Section 2 header: 은행연합회 월보 / 회사 list 포함? / 제외사유
      R48+   Section 2 data (월보 list)
      R71    Section 3 header: 구분(계정) / 계정별원장상 금융기관 / 회사 list 포함?
      R72+   Section 3 data (G/L 계정별)
      R91    Section 4 header: 담보·보증명세서 / 회사 list 포함?
      R92+   Section 4 data
      R102   Section 5 header: 구분 / 부서 / 인터넷상 주소 / G:일치여부
      R103+  Section 5 data (우편 조회처 주소)
    """
    sheet = next((wb[s] for s in wb.sheetnames if s.startswith("AC0.")), None)
    if sheet is None:
        return

    # CS 에 포함된 (canonical, branch) 집합 — 비교 기준
    cs_keys: set[tuple[str, str | None]] = set()
    cs_rows = []
    if "cs" in files:
        cs_rows = ControlSheetLoader(Path(files["cs"].stored_path)).load_bc_rows()
        for r in cs_rows:
            text = " ".join(filter(None, [r.get("name"), r.get("branch")]))
            np = norm.normalize(text or "")
            cs_keys.add((np.canonical, np.branch))

    def _in_cs(canon: str, branch: str | None) -> str:
        return "Y" if (canon, branch) in cs_keys else "N"

    # === Section 1: 전기 CS list ===
    if "prior_cs" in files:
        prior_rows = ControlSheetLoader(Path(files["prior_cs"].stored_path)).load_bc_rows()
        seen = set()
        row_idx = 12
        for pr in prior_rows:
            text = " ".join(filter(None, [pr.get("name"), pr.get("branch")]))
            np = norm.normalize(text or "")
            key = (np.canonical, np.branch)
            if key in seen:
                continue
            seen.add(key)
            display = np.canonical + (f" {np.branch}" if np.branch else "")
            _safe_write(sheet, f"C{row_idx}", display)
            _safe_write(sheet, f"D{row_idx}", _in_cs(np.canonical, np.branch))
            row_idx += 1
            if row_idx > 44:
                break

    # === Section 2: 은행연합회 월보 list ===
    if "union" in files:
        names = parse_union_monthly(Path(files["union"].stored_path))
        seen = set()
        row_idx = 48
        for n in names:
            np = norm.normalize(n)
            if not np.matched:
                continue
            key = (np.canonical, np.branch)
            if key in seen:
                continue
            seen.add(key)
            display = np.canonical + (f" {np.branch}" if np.branch else "")
            _safe_write(sheet, f"C{row_idx}", display)
            _safe_write(sheet, f"D{row_idx}", _in_cs(np.canonical, np.branch))
            row_idx += 1
            if row_idx > 67:
                break

    # === Section 3: G/L 계정별원장 검토 ===
    # G/L sampling 결과 (bs_balance 또는 pl_volume != 0 = G/L 발견)
    row_idx = 72
    seen = set()
    for cp in cps:
        if cp.bs_balance == 0 and cp.pl_volume == 0:
            continue  # CS-only counterparty, G/L에 흔적 없음
        key = (cp.canonical_name, cp.branch)
        if key in seen:
            continue
        seen.add(key)
        # 계정 구분: 잔액 있으면 B/S, 거래액 있으면 P/L
        if cp.bs_balance != 0:
            acc_label = "B/S 잔액"
        else:
            acc_label = "P/L 거래"
        display = cp.canonical_name + (f" {cp.branch}" if cp.branch else "")
        _safe_write(sheet, f"C{row_idx}", acc_label)
        _safe_write(sheet, f"D{row_idx}", display)
        _safe_write(sheet, f"E{row_idx}", _in_cs(cp.canonical_name, cp.branch))
        row_idx += 1
        if row_idx > 88:
            break

    # === Section 4: 담보·보증 명세서 ===
    listed_names = set()
    if "collateral" in files:
        for n in parse_collateral_or_guarantee(Path(files["collateral"].stored_path)):
            np = norm.normalize(n)
            if np.matched:
                listed_names.add((np.canonical, np.branch))
    if "guarantee" in files:
        for n in parse_collateral_or_guarantee(Path(files["guarantee"].stored_path)):
            np = norm.normalize(n)
            if np.matched:
                listed_names.add((np.canonical, np.branch))
    row_idx = 92
    for canon, branch in sorted(listed_names, key=lambda kb: (kb[0], kb[1] or "")):
        display = canon + (f" {branch}" if branch else "")
        _safe_write(sheet, f"C{row_idx}", display)
        _safe_write(sheet, f"D{row_idx}", _in_cs(canon, branch))
        row_idx += 1
        if row_idx > 100:
            break

    # === Section 5: 우편 조회처 주소 유효성 ===
    # CS의 우편 채널 row를 그대로
    row_idx = 103
    for r in cs_rows:
        if "우편" not in (r.get("channel") or ""):
            continue
        # 우리 counterparty 매칭
        text = " ".join(filter(None, [r.get("name"), r.get("branch")]))
        np = norm.normalize(text or "")
        cp = next((c for c in cps if (c.canonical_name, c.branch) == (np.canonical, np.branch)), None)
        addr_valid = "Y" if (cp and cp.address_valid == "ok") else "N"
        display = np.canonical + (f" {np.branch}" if np.branch else "")
        _safe_write(sheet, f"C{row_idx}", display)
        _safe_write(sheet, f"D{row_idx}", r.get("branch") or "")
        _safe_write(sheet, f"E{row_idx}", r.get("address") or "")
        _safe_write(sheet, f"G{row_idx}", addr_valid)
        row_idx += 1
        if row_idx > 110:
            break


def _stamp_project_info(wb, company_name: str, fiscal_date: str):
    """Stamp 회사명·기준일 onto each AC sheet.
    AC control sheet의 A1·A3에 박으면 AC1~AC10은 formula로 자동 참조."""
    # 회사명: 보통 'A1' 위치. AC1~AC10은 formula로 control sheet를 참조함.
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("AC1~AC8"):
            continue  # divider sheet
        ws = wb[sheet_name]
        # A1: 회사명 (control sheet에서만 직접 stamp; 나머지는 formula 유지)
        if "control sheet" in sheet_name.lower() or sheet_name.startswith("AC0."):
            _safe_write(ws, "A1", f"{company_name} 주식회사")
        # A3: 기준일
        _safe_write(ws, "A3", fiscal_date)


# AC0는 5개 sub-section의 data row 영역만 clear (header·절차텍스트·결론 보존)
_AC0_SUBSECTION_RANGES = [
    (12, 44),    # Section 1: 전기 CS list data
    (48, 67),    # Section 2: 월보 list data
    (72, 88),    # Section 3: G/L 계정별 data
    (92, 100),   # Section 4: 담보·보증 data
    (103, 110),  # Section 5: 우편 주소 data
]

# AC1~AC8 data row 영역 (start_row, footer 직전 행) — V1 예시 데이터 clear용
_DATA_REGION = {
    "AC1.": (11, 128),   # 금융자산 list
    "AC2.": (12, 45),    # 차입금 list
    "AC3.": (12, 14),    # 파생상품 list
    "AC4.": (13, 65),    # 지급보증
    "AC5.": (12, 60),    # 담보제공자산
    "AC6.": (13, 43),    # 어음·수표
    "AC7.": (12, 45),    # 보험
    "AC8.": (12, 21),    # 리스
}

# 각 시트별 clear 대상 column 범위 (헤더 column 보존)
_DATA_COLS = "CDEFGHIJKLMNOPQ"

# 보호 키워드 (footer·합계·소제목 보존)
_PROTECT_KEYWORDS = ("합계","소계","계 :","Total","total","해당없음","결론","감사인 의견","구분")


def _clear_data_rows(wb):
    """V1 template의 전년도 예시 데이터를 비움. 헤더·서식·병합·footer·결론 보존."""
    # AC0 — sub-section별 data row만 (header·절차·결론 텍스트 절대 건드리지 않음)
    ac0 = next((wb[s] for s in wb.sheetnames if s.startswith("AC0.")), None)
    if ac0 is not None:
        for start, end in _AC0_SUBSECTION_RANGES:
            for r in range(start, min(end, ac0.max_row) + 1):
                for col_letter in _DATA_COLS:
                    _try_clear(ac0, col_letter, r)
    # AC1~AC8 — 데이터 영역 전체 clear
    for prefix, (start, end) in _DATA_REGION.items():
        sheet_name = next((s for s in wb.sheetnames if s.startswith(prefix)), None)
        if sheet_name is None:
            continue
        ws = wb[sheet_name]
        for r in range(start, min(end, ws.max_row) + 1):
            for col_letter in _DATA_COLS:
                _try_clear(ws, col_letter, r)


def _try_clear(ws, col: str, row: int):
    try:
        cell = ws[f"{col}{row}"]
        if isinstance(cell, MergedCell):
            return
        val = cell.value
        if val is None:
            return
        if isinstance(val, str) and any(k in val for k in _PROTECT_KEYWORDS):
            return
        cell.value = None
    except Exception:
        pass
