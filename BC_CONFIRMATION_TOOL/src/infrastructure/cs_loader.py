from pathlib import Path
import openpyxl


class ControlSheetLoader:
    """4150 'AC 금융기관조회서 control sheet' 시트 파싱.

    표준 컬럼 매핑 (row 6+ data):
      B=BC번호 C=금융기관명 D=지점 E=회신구분 F=주소 H=담당자 I=전화 J=회신여부
    """

    SHEET_KEYWORD = "control sheet"
    DATA_START_ROW = 6
    COL_MAP = {
        "bc_no": "B", "name": "C", "branch": "D", "channel": "E",
        "address": "F", "contact": "H", "phone": "I", "response_status": "J",
    }

    def __init__(self, path: Path):
        self.path = path

    def load_bc_rows(self) -> list[dict]:
        wb = openpyxl.load_workbook(self.path, data_only=True)
        target_sheet = None
        for name in wb.sheetnames:
            if self.SHEET_KEYWORD in name.lower() or "control sheet" in name:
                target_sheet = wb[name]
                break
        if target_sheet is None:
            wb.close()
            return []
        ws = target_sheet
        rows = []
        for r in range(self.DATA_START_ROW, ws.max_row + 1):
            bc = ws[f"B{r}"].value
            if not bc or not str(bc).startswith("BC-"):
                continue
            rows.append({
                key: (ws[f"{col}{r}"].value or None)
                for key, col in self.COL_MAP.items()
            })
        wb.close()
        return rows
