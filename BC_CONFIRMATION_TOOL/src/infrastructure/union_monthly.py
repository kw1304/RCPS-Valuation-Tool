from pathlib import Path
import openpyxl
import re

FIN_TERMS = ["은행","증권","보험","캐피탈","카드","저축은행","금융","신협","수협","농협","산업","수출입"]

def parse_collateral_or_guarantee(path: Path) -> list[str]:
    """담보·연대보증 명세서에서 금융기관 이름 후보 추출.

    rule: 모든 시트·셀을 sweep, FIN_TERMS 키워드가 포함된 문자열을 추출.
    중복 제거.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    seen: set[str] = set()
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if not isinstance(v, str):
                    continue
                s = v.strip()
                if not s or len(s) > 80:
                    continue
                if any(t in s for t in FIN_TERMS):
                    seen.add(s)
    wb.close()
    return sorted(seen)

def parse_union_monthly(path: Path) -> list[str]:
    """은행연합회 월보 (Excel or PDF). MVP: Excel만 지원. PDF는 Phase 2."""
    if path.suffix.lower() not in {".xlsx",".xls"}:
        return []  # PDF는 추후
    return parse_collateral_or_guarantee(path)
