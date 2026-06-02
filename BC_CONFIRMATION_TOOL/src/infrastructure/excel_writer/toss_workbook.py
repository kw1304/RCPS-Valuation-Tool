"""Fresh Toss-style 4150 AC workbook generator.

V1 의 구성요소 (시트·섹션·컬럼·결론) 보존하되 from-scratch 생성하여
Toss 디자인 컨셉 (Pretendard, 깔끔한 헤더·테이블·여백) 적용.
"""
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Toss palette
TOSS_BLUE = "FF3182F6"
TOSS_BLUE_DARK = "FF1B64DA"
TOSS_BLUE_LIGHT = "FFEAF2FE"
COOL_GRAY_50 = "FFF9FAFB"
COOL_GRAY_100 = "FFF2F4F6"
COOL_GRAY_200 = "FFE5E8EB"
COOL_GRAY_900 = "FF191F28"
COOL_GRAY_600 = "FF4E5968"
SUCCESS_GREEN = "FFE0F8EF"
WARN_YELLOW = "FFFFF7E0"
DANGER_RED = "FFFFEAEC"

FONT_NAME = "맑은 고딕"  # Excel 한글 표준; 시스템 Pretendard 있으면 그것

# Reusable styles
_thin_border_color = COOL_GRAY_200
_thin_side = Side(border_style="thin", color=_thin_border_color)
BORDER_ALL = Border(top=_thin_side, bottom=_thin_side, left=_thin_side, right=_thin_side)
BORDER_BOTTOM = Border(bottom=_thin_side)


def _font(size=10, bold=False, color=COOL_GRAY_900):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)


def _fill(rgb):
    return PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")


def _align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _write_title_block(ws, company: str, sheet_title: str, fiscal_date: str, ref_code: str = ""):
    """시트 상단 title block — 회사명 / 시트 제목 / 기준일 + 메타 (작성자/검토자)."""
    # Row 1: 회사명 (Toss blue bar)
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = company
    c.font = _font(size=14, bold=True, color="FFFFFFFF")
    c.fill = _fill(TOSS_BLUE)
    c.alignment = _align(h="left", v="center")
    ws.row_dimensions[1].height = 30

    # Row 2: 시트 제목 (subtitle)
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = sheet_title
    c.font = _font(size=12, bold=True, color=COOL_GRAY_900)
    c.fill = _fill(TOSS_BLUE_LIGHT)
    c.alignment = _align(h="left", v="center")
    # ref code
    if ref_code:
        ws.merge_cells("G2:H2")
        c = ws["G2"]
        c.value = ref_code
        c.font = _font(size=11, bold=True, color=TOSS_BLUE_DARK)
        c.fill = _fill(TOSS_BLUE_LIGHT)
        c.alignment = _align(h="right", v="center")
    ws.row_dimensions[2].height = 24

    # Row 3: 기준일
    ws.merge_cells("A3:F3")
    c = ws["A3"]
    c.value = f"기준일: {fiscal_date}"
    c.font = _font(size=10, color=COOL_GRAY_600)
    c.alignment = _align(h="left", v="center")
    ws.row_dimensions[3].height = 20


def _write_procedure_block(ws, start_row: int, title: str, lines: list[str]) -> int:
    """감사목적·감사절차 텍스트 블록. Returns next available row."""
    r = start_row
    ws.cell(row=r, column=1).value = title
    ws.cell(row=r, column=1).font = _font(size=11, bold=True, color=TOSS_BLUE_DARK)
    ws.cell(row=r, column=1).alignment = _align(h="left", v="center")
    ws.row_dimensions[r].height = 22
    r += 1
    for line in lines:
        ws.cell(row=r, column=2).value = line
        ws.cell(row=r, column=2).font = _font(size=10, color=COOL_GRAY_600)
        ws.cell(row=r, column=2).alignment = _align(h="left", v="top", wrap=True)
        ws.row_dimensions[r].height = 20
        r += 1
    return r + 1  # blank row spacer


def _write_table_header(ws, row: int, headers: list[str], col_widths: list[int] | None = None):
    """테이블 헤더 row — Toss blue dark + white text + bold."""
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i)
        c.value = h
        c.font = _font(size=10, bold=True, color="FFFFFFFF")
        c.fill = _fill(TOSS_BLUE_DARK)
        c.alignment = _align(h="center", v="center")
        c.border = BORDER_ALL
    ws.row_dimensions[row].height = 28
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def _write_data_row(ws, row: int, values: list, alt: bool = False):
    """데이터 row — alternating bg + thin border + Pretendard."""
    bg = COOL_GRAY_50 if alt else "FFFFFFFF"
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i)
        if v is not None:
            c.value = v
        c.font = _font(size=10)
        c.fill = _fill(bg)
        is_num = isinstance(v, (int, float)) and not isinstance(v, bool)
        if is_num:
            # 금액 천단위 콤마 + 우측정렬(잘림 방지 위해 wrap 끔 — 숫자는 줄바꿈 불요).
            c.number_format = "#,##0"
            c.alignment = _align(h="right", v="center")
        else:
            c.alignment = _align(h="left", v="center", wrap=True)
        c.border = BORDER_ALL
    # 행 높이 고정 제거 — wrap 된 텍스트가 잘리지 않도록 Excel 자동 높이에 맡김.


def _write_status_cell(ws, row: int, col: int, status: str):
    """상태 셀 — Y/N/△ tag-style 색상."""
    c = ws.cell(row=row, column=col)
    c.value = status
    s = str(status).strip()
    if s == "Y" or s.startswith("Y"):
        c.fill = _fill(SUCCESS_GREEN)
        c.font = _font(size=10, bold=True, color="FF00C896")
    elif s == "N":
        c.fill = _fill(DANGER_RED)
        c.font = _font(size=10, bold=True, color="FFF04452")
    elif "△" in s or "검토" in s or "수기" in s:
        c.fill = _fill(WARN_YELLOW)
        c.font = _font(size=10, bold=True, color="FFF2A40C")
    else:
        c.font = _font(size=10)
    c.alignment = _align(h="center", v="center")
    c.border = BORDER_ALL


def _write_conclusion(ws, start_row: int, text: str) -> int:
    """결론 블록 — 강조 박스 스타일."""
    r = start_row + 1   # spacer
    ws.cell(row=r, column=1).value = "3. 결론"
    ws.cell(row=r, column=1).font = _font(size=11, bold=True, color=TOSS_BLUE_DARK)
    ws.row_dimensions[r].height = 22
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    c = ws.cell(row=r, column=2)
    c.value = text
    c.font = _font(size=10, color=COOL_GRAY_900)
    c.fill = _fill(COOL_GRAY_50)
    c.alignment = _align(h="left", v="center", wrap=True)
    c.border = BORDER_ALL
    ws.row_dimensions[r].height = 26
    return r + 1


# ─── Sheet builders ──────────────────────────────────────────────

def build_control_sheet(wb: Workbook, company: str, fiscal_date: str, cps: list):
    ws = wb.create_sheet("AC control sheet")
    _write_title_block(ws, company, "AC. 금융기관조회서 control sheet", fiscal_date, "CS")
    headers = ["조서번호", "금융기관", "지점", "조회방식", "주소", "담당자", "전화", "회신여부"]
    _write_table_header(ws, 5, headers, col_widths=[10, 18, 15, 12, 40, 12, 14, 10])
    for i, cp in enumerate(cps):
        row = 6 + i
        values = [
            cp.bc_no, cp.canonical_name, cp.branch or "",
            cp.channel or "", cp.address or "",
            "", "",  # 담당자·전화 (CS에 있으면 별도 fetch — 현재 N/A)
            "회신" if cp.response_arrived else "미회신",
        ]
        _write_data_row(ws, row, values, alt=(i % 2 == 1))
        _write_status_cell(ws, row, 8, values[7])
    # 회신여부 색
    return ws


def build_ac0(wb: Workbook, company: str, fiscal_date: str, sections: dict):
    """AC0 — 5섹션 + 결론.
    sections: dict with keys 'prior'/'union'/'gl'/'collateral'/'address',
              each value = list of dicts."""
    ws = wb.create_sheet("AC0. 금융조회 대상 완전성")
    _write_title_block(ws, company, "AC0. 금융조회 대상의 완전성 검토", fiscal_date, "AC0")
    # 감사목적
    r = _write_procedure_block(ws, 5,
        "1. 감사목적",
        ["금융조회 대상의 완전성 확인"],
    )
    # 감사절차 title
    ws.cell(row=r, column=1).value = "2. 감사절차"
    ws.cell(row=r, column=1).font = _font(size=11, bold=True, color=TOSS_BLUE_DARK)
    ws.row_dimensions[r].height = 22
    r += 2

    # Section 1: 전기 CS
    r = _write_section_title(ws, r, "1) 전기 금융조회 대상 list와 당기 회사 제시 list 비교")
    _write_table_header(ws, r, ["전기 금융조회 대상", "회사 제시 list 포함?", "제외사유 및 타당성"], col_widths=[35, 18, 35])
    r += 1
    for i, item in enumerate(sections.get("prior", [])):
        _write_data_row(ws, r, [item["name"], "", item.get("reason", "")], alt=(i % 2 == 1))
        _write_status_cell(ws, r, 2, item["status"])
        r += 1
    r += 1

    # Section 2: 월보
    r = _write_section_title(ws, r, "2) 은행연합회 월보 ↔ 회사 제시 list 비교")
    _write_table_header(ws, r, ["월보상 금융기관", "회사 제시 list 포함?", "제외사유 및 타당성"], col_widths=[35, 18, 35])
    r += 1
    for i, item in enumerate(sections.get("union", [])):
        _write_data_row(ws, r, [item["name"], "", item.get("reason", "")], alt=(i % 2 == 1))
        _write_status_cell(ws, r, 2, item["status"])
        r += 1
    r += 1

    # Section 3: G/L 계정별
    r = _write_section_title(ws, r, "3) G/L 계정별원장에서 발견한 금융기관 ↔ 회사 제시 list 비교")
    _write_table_header(ws, r, ["구분 (잔액·거래)", "금융기관", "회사 제시 list 포함?", "제외사유 및 타당성"], col_widths=[14, 28, 18, 35])
    r += 1
    for i, item in enumerate(sections.get("gl", [])):
        _write_data_row(ws, r, [item["label"], item["name"], "", item.get("reason", "")], alt=(i % 2 == 1))
        _write_status_cell(ws, r, 3, item["status"])
        r += 1
    r += 1

    # Section 4: 담보·보증 명세서
    r = _write_section_title(ws, r, "4) 담보·지급보증 명세서 list ↔ 회사 제시 list 비교")
    _write_table_header(ws, r, ["명세서상 금융기관", "회사 제시 list 포함?", "제외사유 및 타당성"], col_widths=[35, 18, 35])
    r += 1
    for i, item in enumerate(sections.get("collateral", [])):
        _write_data_row(ws, r, [item["name"], "", item.get("reason", "")], alt=(i % 2 == 1))
        _write_status_cell(ws, r, 2, item["status"])
        r += 1
    r += 1

    # Section 5: 우편 주소
    r = _write_section_title(ws, r, "5) 우편 발송 거래처 주소 유효성 검토")
    _write_table_header(ws, r, ["조서번호", "금융기관", "지점·부서", "주소", "유효성"], col_widths=[10, 20, 18, 50, 14])
    r += 1
    for i, item in enumerate(sections.get("address", [])):
        _write_data_row(ws, r, [item.get("bc_no", ""), item["name"], item.get("branch", ""), item.get("address", ""), ""], alt=(i % 2 == 1))
        _write_status_cell(ws, r, 5, item["status"])
        r += 1

    # 결론
    _write_conclusion(ws, r, "회사 제시 금융기관 list 완전성 검토 결과, 위 사항 외 특이사항 없음.")
    return ws


def _write_section_title(ws, row: int, title: str) -> int:
    """sub-section 제목 (① 회계감사 절차 단계)."""
    ws.cell(row=row, column=1).value = title
    ws.cell(row=row, column=1).font = _font(size=10, bold=True, color=COOL_GRAY_900)
    ws.cell(row=row, column=1).fill = _fill(COOL_GRAY_100)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.row_dimensions[row].height = 22
    return row + 1


_MONEY_HEADER_KW = ("금액", "잔액", "액", "평가", "balance")


def _write_total_row(ws, row: int, records: list, data_start_row: int, headers: list[str]):
    """합계행 — 금액성 컬럼별 =SUM(). 재무제표·주석 대사용.

    금액(금액·잔액·평가액 등) 헤더의 숫자 컬럼만 합산한다. 이자율·설정순위·수량·기준가
    등은 합계가 무의미하므로 제외."""
    n_cols = len(headers)
    numeric_cols = set()
    for rec in records:
        for ci, v in enumerate(rec):
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                numeric_cols.add(ci)
    sum_cols = {ci for ci in numeric_cols
                if ci < n_cols and any(k in headers[ci] for k in _MONEY_HEADER_KW)}
    for i in range(n_cols):
        c = ws.cell(row=row, column=i + 1)
        if i == 0:
            c.value = "합계"
        elif i in sum_cols and row > data_start_row:
            col = get_column_letter(i + 1)
            c.value = f"=SUM({col}{data_start_row}:{col}{row - 1})"
            c.number_format = "#,##0"   # 합계 천단위 콤마
        c.font = _font(size=10, bold=True, color=COOL_GRAY_900)
        c.fill = _fill(COOL_GRAY_100)
        c.alignment = _align(h="left" if i == 0 else "right", v="center")
        c.border = BORDER_ALL
    ws.row_dimensions[row].height = 24


def _disp_len(v) -> float:
    """셀 표시 길이 추정(한글·전각은 1.8배 가중). 수식 셀은 넉넉히 14."""
    if v is None:
        return 0.0
    if isinstance(v, str) and v.startswith("="):
        return 14.0
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        s = f"{v:,.0f}"
    else:
        s = str(v)
    return sum(1.8 if ord(ch) > 0x1100 else 1.0 for ch in s)


def _autofit_table(ws, header_row: int, last_row: int, n_cols: int,
                   min_w: float = 8.0, max_w: float = 46.0):
    """테이블 영역(헤더~합계행)만 보고 열폭을 내용에 맞춤 — 잘림 방지.

    제목·결론 등 병합 텍스트 행은 제외(테이블 행만 스캔)해 표가 왜곡되지 않게 한다."""
    for ci in range(1, n_cols + 1):
        w = 0.0
        for r in range(header_row, last_row + 1):
            w = max(w, _disp_len(ws.cell(row=r, column=ci).value))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(w + 2.0, min_w), max_w)


def _build_record_sheet(wb: Workbook, sheet_name: str, ref_code: str,
                       company: str, fiscal_date: str,
                       title: str, purpose: str,
                       headers: list[str], col_widths: list[int],
                       records: list[list], conclusion: str):
    """AC1~AC8 공통 builder."""
    ws = wb.create_sheet(sheet_name)
    _write_title_block(ws, company, title, fiscal_date, ref_code)
    r = _write_procedure_block(ws, 5, "1. 감사목적", [purpose])
    r = _write_procedure_block(ws, r, "2. 감사절차", [
        f"1) 금융조회서상 회사의 {ref_code.replace('AC','').strip('.')}건 거래내역을 요약 정리함",
        "2) 회사 장부와 대사하여 차이 여부 확인",
    ])
    hdr_row = r
    _write_table_header(ws, r, headers, col_widths)
    r += 1
    data_start = r
    for i, rec in enumerate(records):
        _write_data_row(ws, r, rec, alt=(i % 2 == 1))
        r += 1
    if records:
        _write_total_row(ws, r, records, data_start, headers)
        r += 1
    _autofit_table(ws, hdr_row, r - 1, len(headers))
    _write_conclusion(ws, r, conclusion)
    return ws


def build_ac1_assets(wb, company, fiscal_date, records, detail_records=None):
    """AC1 금융자산 — V1 구조: ① 은행 예금 + ② 증권사 자산 + ③ 유가증권 상세명세."""
    detail_records = detail_records or []
    ws = wb.create_sheet("AC1. 금융자산")
    _write_title_block(ws, company, "AC1. 금융조회서 요약 — 금융자산", fiscal_date, "AC1")
    r = _write_procedure_block(ws, 5, "1. 감사목적", ["회사의 금융자산 실재성 검토"])
    r = _write_procedure_block(ws, r, "2. 감사절차", [
        "1) 금융조회서상 회사의 금융자산을 요약 정리함",
        "2) 회사 장부와 대사하여 차이 여부 확인",
    ])
    bank_recs = [rec for rec in records if rec.get("category", "bank") == "bank"]
    sec_recs = [rec for rec in records if rec.get("category") == "securities"]

    # === ① 은행 예금·신탁 ===
    r = _write_section_title(ws, r, "① 은행 예금·신탁")
    bank_headers = ["조서번호", "금융기관명", "금융상품 종류", "계좌번호", "통화",
                    "금액", "이자율", "최종이자지급일", "만기일", "인출제한 등"]
    _write_table_header(ws, r, bank_headers,
                        col_widths=[10, 16, 28, 20, 8, 16, 10, 14, 14, 18])
    r += 1
    bank_start = r
    bank_rows = []
    for i, rec in enumerate(bank_recs):
        values = [
            rec.get("bc_no",""), rec.get("bank",""), rec.get("product",""),
            rec.get("account_no","") or "", rec.get("currency","KRW"),
            float(rec.get("balance") or 0),
            float(rec.get("interest_rate") or 0),
            str(rec.get("last_interest_date","") or ""),
            str(rec.get("maturity","") or ""),
            rec.get("withdrawal_limit","") or "",
        ]
        bank_rows.append(values)
        _write_data_row(ws, r, values, alt=(i % 2 == 1))
        r += 1
    if bank_rows:
        _write_total_row(ws, r, bank_rows, bank_start, bank_headers)
        r += 1
    r += 1

    # === ② 증권사 자산 (주식·신탁·펀드 등) ===
    r = _write_section_title(ws, r, "② 증권사 자산 (주식·신탁·펀드 등)")
    sec_headers = ["조서번호", "금융기관명", "금융상품 종류", "계좌번호", "통화",
                   "금액", "예수금", "신용설정 보증금", "미수금액", "담보제공·처분제한"]
    _write_table_header(ws, r, sec_headers,
                        col_widths=[10, 16, 24, 20, 8, 16, 12, 14, 12, 20])
    r += 1
    sec_start = r
    sec_rows = []
    for i, rec in enumerate(sec_recs):
        values = [
            rec.get("bc_no",""), rec.get("bank",""), rec.get("product",""),
            rec.get("account_no","") or "", rec.get("currency","KRW"),
            float(rec.get("balance") or 0),
            float(rec.get("deposit_money") or 0) if rec.get("deposit_money") else "",
            float(rec.get("margin_deposit") or 0) if rec.get("margin_deposit") else "",
            float(rec.get("receivable") or 0) if rec.get("receivable") else "",
            rec.get("collateral_restriction","") or "",
        ]
        sec_rows.append(values)
        _write_data_row(ws, r, values, alt=(i % 2 == 1))
        r += 1
    if sec_rows:
        _write_total_row(ws, r, sec_rows, sec_start, sec_headers)
        r += 1

    # === ③ 유가증권 상세명세 (종목별) ===
    if detail_records:
        r += 1
        r = _write_section_title(ws, r, "③ 유가증권 상세명세 (종목별)")
        det_headers = ["조서번호", "금융기관명", "계좌번호", "종목명", "수량",
                       "기준가", "평가액", "담보수량", "비고(질권·담보)"]
        _write_table_header(ws, r, det_headers,
                            col_widths=[10, 14, 20, 20, 14, 14, 18, 14, 20])
        r += 1
        det_start = r
        det_rows = []
        for i, rec in enumerate(detail_records):
            values = [
                rec.get("bc_no",""), rec.get("bank",""),
                rec.get("account_no","") or "",
                rec.get("ticker_name","") or "",
                float(rec.get("quantity") or 0) if rec.get("quantity") else "",
                float(rec.get("base_price") or 0) if rec.get("base_price") else "",
                float(rec.get("valuation") or 0) if rec.get("valuation") else "",
                float(rec.get("collateral_qty") or 0) if rec.get("collateral_qty") else "",
                rec.get("collateral_type","") or "",
            ]
            det_rows.append(values)
            _write_data_row(ws, r, values, alt=(i % 2 == 1))
            r += 1
        if det_rows:
            _write_total_row(ws, r, det_rows, det_start, det_headers)
            r += 1

    # === 결론 ===
    _write_conclusion(ws, r, "회사 장부상 금융자산을 금융조회서상 금액과 대사한 바, 적정함.")
    return ws


def build_ac2_borrowings(wb, company, fiscal_date, records):
    rows = [
        [r.get("bc_no",""), r.get("bank",""), r.get("contract_type",""),
         r.get("limit_ccy","KRW"), float(r.get("limit_amt",0)),
         r.get("balance_ccy","KRW"), float(r.get("balance",0)),
         str(r.get("contract_date","")), str(r.get("maturity","")), r.get("rate","")]
        for r in records
    ]
    return _build_record_sheet(wb, "AC2. 차입금", "AC2",
        company, fiscal_date,
        "AC2. 금융조회서 요약 — 차입금",
        "회사의 차입금 완전성 검토",
        ["조서번호", "금융기관", "대출종류", "한도통화", "한도금액", "잔액통화", "대출금액", "대출일", "만기일", "이자율"],
        [10, 16, 24, 8, 16, 8, 16, 12, 12, 10],
        rows,
        "회사 장부상 차입금을 금융조회서상 금액과 대사한 바, 적정함.",
    )


def build_ac3_derivatives(wb, company, fiscal_date, records):
    rows = [
        [r.get("bc_no",""), r.get("bank",""), r.get("instrument",""),
         str(r.get("contract_date","")), r.get("buy_ccy",""), float(r.get("buy_amt",0)),
         r.get("sell_ccy",""), float(r.get("sell_amt",0)), str(r.get("maturity",""))]
        for r in records
    ]
    return _build_record_sheet(wb, "AC3. 파생상품", "AC3",
        company, fiscal_date,
        "AC3. 금융조회서 요약 — 파생상품",
        "회사 파생상품계약의 회계처리 적정성 검토",
        ["조서번호", "금융기관", "계약종류", "계약일", "매입통화", "매입금액", "매도통화", "매도금액", "만기일"],
        [10, 16, 24, 12, 10, 16, 10, 16, 12],
        rows,
        "회사 파생상품계약 평가액은 회사 장부에 적절히 반영됨.",
    )


_AC4_DIR_LABEL = {"received": "제공받음", "provided": "제공"}


def build_ac4_guarantees(wb, company, fiscal_date, records):
    # 회사가 받은 지급보증(우발자산)과 제공한 연대보증(우발부채)은 회계성격이 반대 →
    # 구분 컬럼으로 명시(혼재 시 주석 오도 방지).
    rows = [
        [r.get("bc_no",""), r.get("bank",""),
         _AC4_DIR_LABEL.get(r.get("direction"), r.get("direction","")),
         r.get("guarantee_type",""),
         r.get("limit_ccy","KRW"), float(r.get("limit_amt",0)),
         r.get("balance_ccy","KRW"), float(r.get("balance",0)), str(r.get("maturity",""))]
        for r in records
    ]
    return _build_record_sheet(wb, "AC4. 지급보증", "AC4",
        company, fiscal_date,
        "AC4. 금융조회서 요약 — 지급보증",
        "회사의 지급보증 등 약정사항 주석의 적정성 검토",
        ["조서번호", "금융기관", "구분", "보증내용", "한도통화", "한도금액", "잔액통화", "실행금액", "만기일"],
        [10, 16, 10, 28, 8, 16, 8, 16, 12],
        rows,
        "회사의 지급보증 등 약정사항 주석은 적정하게 공시됨.",
    )


def build_ac5_collateral(wb, company, fiscal_date, records):
    rows = [
        [r.get("bc_no",""), r.get("bank",""), r.get("collateral_type",""),
         r.get("creditor",""), r.get("issuer",""),
         float(r.get("book_amount",0)), float(r.get("appraised_amount") or 0),
         float(r.get("senior_lien") or 0), r.get("priority","")]
        for r in records
    ]
    return _build_record_sheet(wb, "AC5. 담보제공자산", "AC5",
        company, fiscal_date,
        "AC5. 금융조회서 요약 — 담보제공자산",
        "회사의 담보제공자산 주석의 적정성 검토",
        ["조서번호", "금융기관", "담보종류", "담보권자", "소유자", "장부금액", "감정금액", "선순위설정금액", "설정순위"],
        [10, 16, 24, 18, 18, 16, 16, 16, 10],
        rows,
        "회사의 담보제공자산 주석은 적정하게 공시됨.",
    )


def build_ac6_bills(wb, company, fiscal_date, records):
    rows = [
        [r.get("bc_no",""), r.get("bank",""), r.get("kind",""), int(r.get("count",0)), float(r.get("balance",0))]
        for r in records
    ]
    return _build_record_sheet(wb, "AC6. 어음·수표", "AC6",
        company, fiscal_date,
        "AC6. 금융조회서 요약 — 어음·수표",
        "회사의 미회수 어음·수표 우발부채 검토",
        ["조서번호", "금융기관", "종류", "매수", "잔액"],
        [10, 16, 30, 10, 16],
        rows,
        "회사의 어음·수표 관련 부채 및 주석은 적정함.",
    )


def build_ac7_insurance(wb, company, fiscal_date, records):
    rows = [
        [r.get("bc_no",""), r.get("bank",""), r.get("product",""), r.get("policy_no",""),
         float(r.get("coverage_amount") or 0), float(r.get("premium") or 0),
         str(r.get("start_date","")), str(r.get("end_date",""))]
        for r in records
    ]
    return _build_record_sheet(wb, "AC7. 보험가입내역", "AC7",
        company, fiscal_date,
        "AC7. 금융조회서 요약 — 보험가입내역",
        "회사의 보험 관련 비용·선급/미지급비용·주석 공시 적정성 검토",
        ["조서번호", "금융기관", "보험종류", "증권번호", "부보금액", "연간보험료", "시작일", "종료일"],
        [10, 16, 24, 18, 16, 14, 12, 12],
        rows,
        "회사의 보험 관련 비용·주석은 적정함.",
    )


def build_ac8_lease(wb, company, fiscal_date, records):
    rows = [
        [r.get("bc_no",""), r.get("bank",""), r.get("asset_type",""), r.get("account_no",""),
         str(r.get("deal_date","")), r.get("deal_type",""), float(r.get("outstanding") or 0), r.get("period","")]
        for r in records
    ]
    return _build_record_sheet(wb, "AC8. 리스거래", "AC8",
        company, fiscal_date,
        "AC8. 금융조회서 요약 — 리스 거래",
        "회사의 리스 분류·관련 비용·주석 공시 적정성 검토",
        ["조서번호", "금융기관", "리스자산", "계약번호", "거래일", "거래종류", "미상환잔액", "리스기간"],
        [10, 16, 22, 18, 12, 14, 16, 16],
        rows,
        "회사의 리스 분류·관련 비용·주석은 적정함.",
    )


def build_workbook(company: str, fiscal_date: str,
                   cps: list, ac0_sections: dict,
                   ac1_recs: list, ac2_recs: list, ac3_recs: list, ac4_recs: list,
                   ac5_recs: list, ac6_recs: list, ac7_recs: list, ac8_recs: list,
                   ac1_detail_recs: list | None = None) -> Workbook:
    """Build complete Toss-style 4150 workbook."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    build_control_sheet(wb, company, fiscal_date, cps)
    build_ac0(wb, company, fiscal_date, ac0_sections)
    build_ac1_assets(wb, company, fiscal_date, ac1_recs, detail_records=ac1_detail_recs or [])
    build_ac2_borrowings(wb, company, fiscal_date, ac2_recs)
    build_ac3_derivatives(wb, company, fiscal_date, ac3_recs)
    build_ac4_guarantees(wb, company, fiscal_date, ac4_recs)
    build_ac5_collateral(wb, company, fiscal_date, ac5_recs)
    build_ac6_bills(wb, company, fiscal_date, ac6_recs)
    build_ac7_insurance(wb, company, fiscal_date, ac7_recs)
    build_ac8_lease(wb, company, fiscal_date, ac8_recs)
    return wb
