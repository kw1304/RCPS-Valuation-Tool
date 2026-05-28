from openpyxl.styles import PatternFill, Font
from openpyxl.workbook import Workbook

TOSS = {
    "primary":      "FF3182F6",
    "primary_dark": "FF1B64DA",
    "light_bg":     "FFF2F4F6",
    "warning":      "FFFFF7E0",
}

TITLE_ROWS = [1, 2, 3]
HEADER_ROWS_BY_PREFIX = {
    "AC0.": [11, 12],
    "AC1.": [10, 11],
    "AC2.": [10, 11],
    "AC3.": [10, 11],
    "AC4.": [11, 12],
    "AC5.": [11, 12],
    "AC6.": [11, 12],
    "AC7.": [10, 11],
    "AC8.": [10, 11],
    "AC ":  [5],          # control sheet header
}

def _fill_row(ws, row: int, hex_argb: str, font_white: bool = True):
    pf = PatternFill(start_color=hex_argb, end_color=hex_argb, fill_type="solid")
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=c)
        if cell.value is None and c > 1:
            continue
        cell.fill = pf
        if font_white:
            old = cell.font
            cell.font = Font(name=old.name, size=old.size, bold=old.bold,
                             italic=old.italic, color="FFFFFFFF")

def apply_toss_palette(wb: Workbook):
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # title rows
        for r in TITLE_ROWS:
            if ws.max_row >= r:
                _fill_row(ws, r, TOSS["primary"], font_white=True)
        # header rows by sheet prefix
        for prefix, rows in HEADER_ROWS_BY_PREFIX.items():
            if sheet_name.startswith(prefix):
                for r in rows:
                    if ws.max_row >= r:
                        _fill_row(ws, r, TOSS["primary_dark"], font_white=True)
                break

def mark_low_confidence(ws, row: int, col: str, comment: str = "OCR 신뢰도 낮음 - 검토 필요"):
    from openpyxl.comments import Comment
    cell = ws[f"{col}{row}"]
    cell.fill = PatternFill(start_color=TOSS["warning"], end_color=TOSS["warning"], fill_type="solid")
    cell.comment = Comment(comment, "BC tool")
