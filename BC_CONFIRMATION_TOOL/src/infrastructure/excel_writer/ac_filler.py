import openpyxl
from copy import copy
from pathlib import Path
from decimal import Decimal
from datetime import date
from openpyxl.cell.cell import MergedCell
from src.domain.ac_models import (
    FinancialAsset, Borrowing, Derivative, Guarantee,
    Collateral, BillCheck, Insurance, GeneralDeal,
)

SHEET_CONFIG = {
    "AC1": {"sheet_name": "AC1. 금융자산", "start_row": 11, "cols": {
        "C": "bc_no", "D": "bank", "E": "product", "F": "account_no",
        "G": "currency", "H": "balance", "I": "interest_rate", "J": "open_date",
    }},
    "AC2": {"sheet_name": "AC2. 차입금", "start_row": 12, "cols": {
        "C": "bc_no", "D": "bank", "E": "contract_type",
        "F": "limit_ccy", "G": "limit_amt", "H": "balance_ccy", "I": "balance",
        "J": "contract_date",
    }},
    "AC3": {"sheet_name": "AC3. 파생상품", "start_row": 12, "cols": {
        "C": "bc_no", "D": "instrument", "E": "contract_date",
        "F": "buy_ccy", "G": "buy_amt", "H": "sell_ccy", "I": "sell_amt",
    }},
    "AC4": {"sheet_name": "AC4. 지급보증", "start_row": 13, "cols": {
        "C": "bc_no", "D": "bank", "E": "guarantee_type",
        "F": "limit_ccy", "G": "limit_amt", "H": "balance_ccy", "I": "balance",
        "J": "maturity",
    }},
    "AC5": {"sheet_name": "AC5. 담보제공자산", "start_row": 12, "cols": {
        "C": "bc_no", "D": "bank", "E": "collateral_type",
        "F": "creditor", "G": "issuer", "H": "book_amount", "I": "appraised_amount",
        "J": "priority",
    }},
    "AC6": {"sheet_name": "AC6. 어음.수표", "start_row": 13, "cols": {
        "C": "bc_no", "D": "bank", "E": "kind", "G": "count",
    }},
    "AC7": {"sheet_name": "AC7. 보험가입내역", "start_row": 12, "cols": {
        "C": "bc_no", "D": "bank", "E": "product", "F": "policy_no",
        "G": "coverage_amount", "H": "premium", "I": "start_date", "J": "end_date",
    }},
    "AC8": {"sheet_name": "AC8. 리스거래", "start_row": 12, "cols": {
        "C": "bc_no", "D": "bank", "E": "asset_type", "F": "account_no",
        "G": "deal_date", "H": "deal_type", "I": "outstanding", "J": "period",
    }},
}

class ACFiller:
    def __init__(self, template_path: Path):
        self.path = template_path
        self.wb = openpyxl.load_workbook(template_path)

    # AC sheet별 최대 data row (결론 직전). fill_section은 이 한계 초과 시 skip.
    _MAX_DATA_ROW = {
        "AC1": 128, "AC2": 47, "AC3": 26, "AC4": 61,
        "AC5": 60, "AC6": 43, "AC7": 45, "AC8": 20,
    }

    def fill_section(self, ac: str, records: list):
        cfg = SHEET_CONFIG[ac]
        # exact match first, prefix fallback (sheet 이름이 미세하게 다를 수 있음)
        ws = None
        if cfg["sheet_name"] in self.wb.sheetnames:
            ws = self.wb[cfg["sheet_name"]]
        else:
            prefix = cfg["sheet_name"].split(".")[0] + "."
            for name in self.wb.sheetnames:
                if name.startswith(prefix):
                    ws = self.wb[name]; break
        if ws is None:
            return
        start = cfg["start_row"]
        max_row = self._MAX_DATA_ROW.get(ac, start + 999)
        for idx, rec in enumerate(records):
            row = start + idx
            if row > max_row:
                # 결론·합계 영역 침범 방지 — 초과 record는 skip (사용자 검토 column에 별도 표시 가능)
                break
            self._ensure_row_style(ws, start, row)
            for col, attr in cfg["cols"].items():
                val = self._extract(rec, attr)
                if val is not None:
                    cell = ws[f"{col}{row}"]
                    if not isinstance(cell, MergedCell):
                        cell.value = val

    def _extract(self, rec, attr: str):
        if hasattr(rec, attr):
            v = getattr(rec, attr)
            if isinstance(v, Decimal):
                return float(v)
            if isinstance(v, date):
                return v
            return v
        return None

    def _ensure_row_style(self, ws, source_row: int, target_row: int):
        if target_row == source_row:
            return
        for col_idx in range(1, ws.max_column + 1):
            src = ws.cell(row=source_row, column=col_idx)
            tgt = ws.cell(row=target_row, column=col_idx)
            if isinstance(src, MergedCell) or isinstance(tgt, MergedCell):
                continue
            if src.has_style:
                tgt.font          = copy(src.font)
                tgt.fill          = copy(src.fill)
                tgt.border        = copy(src.border)
                tgt.alignment     = copy(src.alignment)
                tgt.number_format = src.number_format
                tgt.protection    = copy(src.protection)

    def save(self, dest: Path | None = None):
        self.wb.save(dest or self.path)
