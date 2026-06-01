import logging
from pathlib import Path
from typing import Iterator
import openpyxl

logger = logging.getLogger("bc.gl")

# SAP 보조부원장에서 통화를 담는 컬럼 후보(레이아웃마다 표기 상이).
_CCY_COLS = ("통화", "통화키", "통화코드", "Currency", "CURR", "Crcy", "WAERS")
_KRW_LIKE = {"KRW", "WON", "원", "₩", ""}


class GLLoader:
    """대용량 G/L Excel을 stream으로 읽음. read_only=True로 메모리 절약.

    두 가지 포맷 지원:
    1. 표준 포맷: 컬럼 "계정 과목", "거래처", "금액"
    2. SAP 보조부원장 포맷: 컬럼 "계정", "계정명", "차변금액(현지통화)", "대변금액(현지통화)"
       - 금융계정에서 거래처는 계정명에 내재 (예: "기업은행 257-... 보통예금")
       - 출력 시 "계정 과목" → 계정명, "거래처" → 계정명, "금액" → 차변-대변
    """

    def __init__(self, path: Path):
        self.path = path

    def iter_rows(self, sheet: str | None = None) -> Iterator[dict]:
        wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            wb.close()
            return
        cols = [str(h).strip() if h is not None else "" for h in header]
        is_sap = "계정명" in cols and "계정 과목" not in cols
        # SAP 금액 컬럼: 회사통화(KRW 환산) > 현지통화. 현지통화만 있으면 비KRW 행이
        # 원화와 섞여 단위 오집계되므로, 회사통화 컬럼이 있으면 그쪽을 우선한다.
        debit_col = "차변금액" if "차변금액" in cols else "차변금액(현지통화)"
        credit_col = "대변금액" if "대변금액" in cols else "대변금액(현지통화)"
        ccy_col = next((c for c in _CCY_COLS if c in cols), None)
        local_only = debit_col.endswith("(현지통화)")
        nonkrw = 0
        for row in it:
            if not any(v not in (None, "") for v in row):
                continue
            d = {cols[i]: row[i] for i in range(min(len(cols), len(row)))}
            if is_sap:
                acc_nm = str(d.get("계정명") or "").strip()
                if not acc_nm:
                    continue
                debit = self._to_float(d.get(debit_col))
                credit = self._to_float(d.get(credit_col))
                amount = debit - credit
                ccy = (str(d.get(ccy_col)).strip().upper() if ccy_col and d.get(ccy_col) else "")
                # 현지통화만 있고 비KRW면 단위 혼합 위험 — 카운트해 경고(환산 불가시 가시화).
                if local_only and ccy and ccy not in _KRW_LIKE:
                    nonkrw += 1
                # 표준 포맷으로 정규화하여 yield (통화 표기 보존)
                yield {
                    "계정 과목": acc_nm,
                    "거래처": acc_nm,   # 금융계정에서 거래처=계정명 (은행명 포함)
                    "금액": amount,
                    "통화": ccy or None,
                    # 원본 필드도 보존
                    **{k: v for k, v in d.items()},
                }
            else:
                if not d.get("계정 과목"):
                    continue
                yield d
        if nonkrw:
            logger.warning(
                "GL 현지통화 컬럼만 존재 + 비KRW 행 %d건 — 원화환산 없이 금액 합산됨(단위혼합 주의). "
                "회사통화(KRW) 금액 컬럼이 있는 원장 사용 권장.", nonkrw)
        wb.close()

    @staticmethod
    def _to_float(v) -> float:
        if v is None or v == "":
            return 0.0
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0
