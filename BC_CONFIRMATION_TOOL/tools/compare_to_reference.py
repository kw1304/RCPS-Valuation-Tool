"""참고조서(정답) ↔ 툴 파싱결과 비교 하니스 (지속적 매핑 개선용).

목적
----
사람이 완성한 참고조서(4150 AC 금융기관 조회 …xlsx)의 AC 시트 값과,
툴이 회신본 PDF 들을 파싱해 만든 레코드를 **AC별 금액 집합**으로 비교한다.
컬럼 미세정합이 아니라 "정답의 금액이 툴 출력에 들어왔는가 / 툴이 헛것을
만들었는가"를 빠르게 본다 — 매핑 개선의 회귀 가드.

비교 대상 금액(AC별)
  AC1 : 금액(balance)
  AC2 : 한도금액(limit_amt) + 대출금액(balance)  ← POSITIONAL 두 컬럼 모두
  AC4 : 한도금액 + 실행금액(balance)
  AC5 : 장부가(book_amount)  [+ 감정금액(appraised_amount) 있으면]
  AC7 : 부보금액(coverage_amount)

사용
----
  python tools/compare_to_reference.py \
      "INPUT/4150_AC 금융기관 조회_코스맥스비티아이_FY2025_V1.xlsx" \
      "INPUT/온라인" [우편디렉터리]

출력
----
  AC별 MISSED(정답엔 있는데 툴엔 없음) / EXTRA(툴엔 있는데 정답엔 없음) 목록 +
  per-AC 요약(#정답행, #matched, #missed, #extra).

견고성
------
  - 금액은 정수 원화(소수 절사, 콤마 제거)로 정규화해 비교(멀티셋).
  - 0 은 noise 가 많아 비교에서 제외(매칭/미스 판정 대상 아님).
  - 참고시트 헤더 행을 자동 탐색해 금액 컬럼 인덱스를 잡는다(레이아웃 변동 견고).
"""
from __future__ import annotations

import sys
import glob
import re
from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402

from src.infrastructure.pdf.extractor import extract_rows  # noqa: E402
from src.infrastructure.pdf.ocr import ocr_pdf  # noqa: E402
from src.infrastructure.pdf.form_fingerprint import identify_form  # noqa: E402
from src.infrastructure.pdf.section_splitter import split_sections  # noqa: E402
from src.infrastructure.pdf.filename_parser import parse_filename  # noqa: E402
from src.application.parse_response_uc import route_or_classify, _dispatch  # noqa: E402
from src.infrastructure.pdf.row_parsers.fallback import fallback_parse  # noqa: E402
from src.domain.record_dedup import dedup_key  # noqa: E402


# ── 금액 정규화 ────────────────────────────────────────────────────────────
def _to_int_amount(v) -> int | None:
    """셀/문자열을 정수 원화 금액으로. 비금액·0·None → None.

    '126,598,004.00(126,645,985.00)' 같은 합성 셀은 첫 금액만 취한다."""
    if v is None:
        return None
    if isinstance(v, (int, float, Decimal)):
        try:
            iv = int(Decimal(str(v)))
        except (InvalidOperation, ValueError):
            return None
        return iv if iv != 0 else None
    s = str(v).strip()
    if not s:
        return None
    # 괄호 안 보조금액 제거(평가액/직전잔액 등)
    s = re.split(r"[(（]", s, maxsplit=1)[0].strip()
    s = s.replace(",", "")
    m = re.match(r"-?\d+(?:\.\d+)?$", s)
    if not m:
        return None
    try:
        iv = int(Decimal(s))
    except (InvalidOperation, ValueError):
        return None
    return iv if iv != 0 else None


# ── 참고조서 읽기 ──────────────────────────────────────────────────────────
# 시트 prefix → (라벨, [비교할 헤더명들])
_REF_SHEETS = {
    "AC1": ("금융자산", ["금액"]),
    "AC2": ("차입금", ["한도금액", "대출금액"]),
    # ① 회사가 제공받은(한도액/실행금액) + ② 회사가 타인에게 제공한 연대보증.
    # 이 시트는 통화/금액 2단 stacked 구조라 단일 헤더행 모델로 금액컬럼을 못 잡는다.
    # → AC4 는 헤더 무관하게 모든 데이터 셀에서 100만↑ 원화 금액을 수집한다(아래 특례).
    "AC4": ("지급보증", []),
    "AC5": ("담보제공자산", ["장부가", "장부금액", "감정금액", "설정금액", "담보금액"]),
    "AC7": ("보험가입내역", ["부보금액"]),
}


def _find_sheet(wb, prefix: str):
    for name in wb.sheetnames:
        # 'AC2. 차입금' / 'AC2.차입금' 등 — prefix 가 'AC2' 면 'AC2.' 로 시작
        head = name.replace(" ", "")
        if head.startswith(prefix + ".") or head.startswith(prefix + "."):
            return wb[name]
    return None


def _header_amount_cols(ws, header_names: list[str]) -> tuple[int, list[int]]:
    """헤더 행 번호와, header_names 중 하나와 일치하는 금액 컬럼 인덱스 목록을 찾는다.

    헤더가 2줄에 걸쳐(병합) 있을 수 있어 앞 30행을 스캔, 가장 많은 매칭을 가진 행 채택.
    반환: (data_start_row, [col_idx, ...]). 못 찾으면 (0, [])."""
    best_row, best_cols = 0, []
    wanted = [h.replace(" ", "") for h in header_names]
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        cols = []
        for ci, cell in enumerate(row):
            if cell is None:
                continue
            txt = str(cell).replace(" ", "")
            if any(w and w in txt for w in wanted):
                cols.append(ci)
        if len(cols) > len(best_cols):
            best_row, best_cols = ri, cols
    return best_row, best_cols


# AC4 등 stacked 시트는 헤더 매핑 대신 전체 금액 스캔. 100만 미만(증권번호 index·
# 순위 등)은 잡지 않는다.
_MIN_REF_AMOUNT = 1_000_000


def _scan_amounts_min(ws, minv: int) -> Counter:
    """시트 전체 데이터 셀에서 minv 이상 원화 금액을 멀티셋으로 수집."""
    amounts: Counter = Counter()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            iv = _to_int_amount(cell)
            if iv is not None and iv >= minv:
                amounts[iv] += 1
    return amounts


def read_reference_amounts(xlsx: Path) -> dict[str, Counter]:
    """참고조서 → {AC: Counter(정수금액)}. 0/비금액 제외."""
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    out: dict[str, Counter] = {}
    for ac, (label, headers) in _REF_SHEETS.items():
        ws = _find_sheet(wb, ac)
        if ws is None:
            out[ac] = Counter()
            continue
        # AC4 특례: 통화/금액 2단 stacked + ①② 2개 표 → 헤더 컬럼 매핑이 불안정.
        # 시트 전체 데이터 셀에서 100만↑ 원화 금액을 수집(지급보증 시트는 금액 전용).
        if ac == "AC4":
            out[ac] = _scan_amounts_min(ws, _MIN_REF_AMOUNT)
            continue
        hdr_row, cols = _header_amount_cols(ws, headers)
        amounts: Counter = Counter()
        if cols:
            for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
                for ci in cols:
                    if ci < len(row):
                        iv = _to_int_amount(row[ci])
                        if iv is not None:
                            amounts[iv] += 1
        out[ac] = amounts
    return out


# ── 툴 파싱 ────────────────────────────────────────────────────────────────
# 레코드(payload dict) → 비교할 금액 필드(AC별)
_TOOL_FIELDS = {
    "AC1": ["balance"],
    "AC2": ["limit_amt", "balance"],
    "AC4": ["limit_amt", "balance"],
    "AC5": ["book_amount", "appraised_amount"],
    "AC7": ["coverage_amount"],
}


# 비교에 포함하는 AC (참고조서가 가진 시트). AC8 은 참고조서에 없어 제외.
_COMPARE_ACS = ["AC1", "AC2", "AC4", "AC5", "AC7"]


def _add_fallback_amounts(out: dict[str, Counter], text: str, bc_no: str, bank: str) -> None:
    """우편/unknown → fallback_parse. 행별 최대 금액(보조금액 아닌 대표값)을 AC별 수집."""
    for rec in fallback_parse(text, bc_no=bc_no, bank=bank):
        ac = rec["ac_section"]
        if ac not in _COMPARE_ACS:
            continue
        amts = [_to_int_amount(a) for a in rec["payload"].get("amounts", [])]
        amts = [a for a in amts if a is not None]
        if not amts:
            continue
        out[ac][max(amts)] += 1


def parse_pdf_dir(pdf_dir: Path, fallback_only: bool = False) -> dict[str, Counter]:
    """회신본 PDF 디렉터리 → {AC: Counter(정수금액)} (툴 파싱 결과).

    fallback_only=True 면 정형 파서를 쓰지 않고 fallback_parse 만 적용한다
    (우편/OCR 디렉터리용). False 면 정형 family 는 섹션 파서로, unknown 은
    fallback 으로 파싱한다(온라인 디렉터리에 섞인 비정형 회신 포함).

    이중계상 방지: 한 디렉터리에 개별 회신본 + 합본 스캔이 함께 있으면 같은
    holding 이 두 번 파싱된다. 정형 레코드는 dedup_key 로 디렉터리 전체에서
    한 번만 집계한다(합본 스캔은 bc_no·bank 가 비어 키에서 제외됨).
    또한 AC1_DETAIL(유가증권 종목 상세명세)은 참고조서 AC1(요약)과 다른 영역이므로
    AC1 비교 버킷에 섞지 않는다(별도/제외)."""
    out: dict[str, Counter] = {ac: Counter() for ac in _COMPARE_ACS}
    seen: set = set()  # dedup: 디렉터리 전체 정형 레코드 식별자
    pdfs = sorted(glob.glob(str(pdf_dir / "*.pdf")))
    for p in pdfs:
        path = Path(p)
        meta = parse_filename(path.name)
        bc_no = meta.get("bc_no") or ""
        bank = meta.get("bank_raw") or ""
        try:
            text = extract_rows(path)
        except Exception as e:
            print(f"  [WARN] extract 실패 {path.name}: {e}", file=sys.stderr)
            continue
        # 스캔 PDF(디지털 텍스트 거의 없음) → OCR (production parse_response_uc 와 동일 기준).
        # 우편 보험사 회신서(흥국화재·예별/MG손해보험)는 이미지 스캔이라 OCR 없이는
        # 부보금액을 전혀 못 잡는다.
        if len(text.strip()) < 80:
            try:
                text = ocr_pdf(path).get("text", "") or text
            except Exception as e:
                print(f"  [WARN] OCR 실패 {path.name}: {e}", file=sys.stderr)
        family = identify_form(text)
        if fallback_only or family == "unknown":
            _add_fallback_amounts(out, text, bc_no, bank)
            continue
        blocks = split_sections(text)
        for section_no, block in blocks.items():
            route = route_or_classify(family, section_no, block)
            if not route:
                continue
            ac = route["ac"]
            # AC1_DETAIL(종목 상세명세)은 AC1(요약) 비교 대상이 아니다 — 제외.
            if ac == "AC1_DETAIL":
                continue
            store_ac = ac
            if store_ac not in _TOOL_FIELDS:
                continue
            parse_block = route.get("block", block)
            try:
                recs = _dispatch(ac, parse_block, bc_no, bank, route)
            except Exception:
                continue
            for rec in recs:
                d = rec.model_dump() if hasattr(rec, "model_dump") else dict(rec)
                k = dedup_key(store_ac, d)
                if k is not None:
                    if k in seen:
                        continue  # 합본 스캔 등 중복 holding → 한 번만 집계
                    seen.add(k)
                for fld in _TOOL_FIELDS[store_ac]:
                    iv = _to_int_amount(d.get(fld))
                    if iv is not None:
                        out[store_ac][iv] += 1
    return out


def _merge_counters(a: dict[str, Counter], b: dict[str, Counter]) -> dict[str, Counter]:
    out: dict[str, Counter] = {ac: Counter() for ac in _COMPARE_ACS}
    for ac in _COMPARE_ACS:
        out[ac] = (a.get(ac, Counter()) + b.get(ac, Counter()))
    return out


# ── 비교 ───────────────────────────────────────────────────────────────────
def _fmt(n: int) -> str:
    return f"{n:,}"


def _amount_sum(c: Counter) -> int:
    """멀티셋 금액 총합 = Σ(금액 × 건수)."""
    return sum(amt * cnt for amt, cnt in c.items())


def _pct_err(tool_sum: int, ref_sum: int) -> float | None:
    """abs(tool−ref)/ref ×100. ref_sum==0 이면 None(정의 불가)."""
    if ref_sum == 0:
        return None
    return abs(tool_sum - ref_sum) / ref_sum * 100.0


def compare(ref: dict[str, Counter], tool: dict[str, Counter]) -> None:
    acs = _COMPARE_ACS
    print("=" * 72)
    print("참고조서(정답) ↔ 툴 파싱 금액집합 비교  (0·비금액 제외)")
    print("=" * 72)
    summary = []
    for ac in acs:
        r = ref.get(ac, Counter())
        t = tool.get(ac, Counter())
        # 멀티셋 차집합
        missed = r - t   # 정답엔 있는데 툴엔 부족
        extra = t - r    # 툴엔 있는데 정답엔 없음
        matched_multiset = (r & t)
        n_ref = sum(r.values())
        n_matched = sum(matched_multiset.values())
        n_missed = sum(missed.values())
        n_extra = sum(extra.values())
        ref_sum = _amount_sum(r)
        tool_sum = _amount_sum(t)
        pct = _pct_err(tool_sum, ref_sum)
        summary.append((ac, n_ref, n_matched, n_missed, n_extra,
                        ref_sum, tool_sum, pct))

        pct_str = f"{pct:.1f}%" if pct is not None else "n/a"
        print(f"\n── {ac} ── (정답 {n_ref}건, matched {n_matched}, "
              f"missed {n_missed}, extra {n_extra} | "
              f"오차 {pct_str})")
        print(f"  금액합  정답 {_fmt(ref_sum)} | 툴 {_fmt(tool_sum)}")
        if missed:
            print("  MISSED (정답엔 있는데 툴엔 없음):")
            for amt, c in sorted(missed.items(), reverse=True):
                tag = f" x{c}" if c > 1 else ""
                print(f"    - {_fmt(amt)}{tag}")
        if extra:
            print("  EXTRA (툴엔 있는데 정답엔 없음):")
            for amt, c in sorted(extra.items(), reverse=True):
                tag = f" x{c}" if c > 1 else ""
                print(f"    + {_fmt(amt)}{tag}")
        if not missed and not extra:
            print("  (완전 일치)")

    print("\n" + "=" * 72)
    print("요약  AC | #정답 | matched | missed | extra |   오차%  | 정답금액합 → 툴금액합")
    print("-" * 72)
    tot_ref = tot_tool = 0
    for ac, n_ref, n_matched, n_missed, n_extra, ref_sum, tool_sum, pct in summary:
        tot_ref += ref_sum
        tot_tool += tool_sum
        pct_str = f"{pct:6.1f}%" if pct is not None else "   n/a "
        print(f"  {ac:5} | {n_ref:5} | {n_matched:7} | {n_missed:6} | {n_extra:5} | "
              f"{pct_str} | {_fmt(ref_sum)} → {_fmt(tool_sum)}")
    print("=" * 72)
    overall = _pct_err(tot_tool, tot_ref)
    overall_str = f"{overall:.1f}%" if overall is not None else "n/a"
    print(f"전체 금액합  정답 {_fmt(tot_ref)} | 툴 {_fmt(tot_tool)}")
    print(f"OVERALL AMOUNT ERROR: {overall_str}")


def main(argv: list[str]) -> int:
    if len(argv) < 3:
        print(__doc__)
        print("사용: python tools/compare_to_reference.py "
              "<참고조서.xlsx> <온라인디렉터리> [우편디렉터리]")
        return 2
    xlsx = Path(argv[1])
    online_dir = Path(argv[2])
    postal_dir = Path(argv[3]) if len(argv) >= 4 else None
    if not xlsx.exists():
        print(f"참고조서 없음: {xlsx}", file=sys.stderr)
        return 2
    if not online_dir.is_dir():
        print(f"온라인 디렉터리 없음: {online_dir}", file=sys.stderr)
        return 2
    print(f"참고조서: {xlsx}")
    print(f"온라인  : {online_dir}")
    ref = read_reference_amounts(xlsx)
    tool = parse_pdf_dir(online_dir)
    if postal_dir is not None:
        if postal_dir.is_dir():
            print(f"우편    : {postal_dir} (fallback/OCR)")
            postal = parse_pdf_dir(postal_dir, fallback_only=True)
            tool = _merge_counters(tool, postal)
        else:
            print(f"  [WARN] 우편 디렉터리 없음, 건너뜀: {postal_dir}", file=sys.stderr)
    compare(ref, tool)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
